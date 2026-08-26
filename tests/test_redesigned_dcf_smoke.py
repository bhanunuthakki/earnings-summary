"""Smoke test for the redesigned-DCF builder (execution/build_redesigned_dcf.py).

Generates a minimal FMP fixture for a fake ticker, runs the builder as a
subprocess (the way the driver invokes it), and asserts the workbook has the
ten expected sheets, the headline value cell, the Dashboard dropdowns, and no
column-A label that accidentally became a formula (the leading-'=' bug).
"""

from __future__ import annotations

import json
import os
import sqlite3
import subprocess
import sys
from pathlib import Path
from typing import cast

import openpyxl
import pytest

from dcf import redesign

BUILDER = Path(__file__).resolve().parents[1] / "execution" / "build_redesigned_dcf.py"
SHEETS = [
    "Cover",
    "Dashboard",
    "Assumptions",
    "Color Code",
    "WACC",
    "Model",
    "Financials",
    "Consensus",
    "Valuation",
    "Sensitivity",
    "Monte Carlo",
]


def _write_primary_bridge_facts(
    repo: Path, ticker: str, latest: dict[str, object], *, currency: str
) -> None:
    conn = sqlite3.connect(repo / "data" / "portfolio.db")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY, ticker TEXT NOT NULL, source_type TEXT NOT NULL,
            fetched_at TEXT NOT NULL, source_url TEXT, source_quality_tier TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS financial_facts (
            id INTEGER PRIMARY KEY, ticker TEXT NOT NULL, period_end TEXT NOT NULL,
            fiscal_period_type TEXT NOT NULL, line_item TEXT NOT NULL, value NUMERIC NOT NULL,
            currency TEXT, unit TEXT NOT NULL, source_doc_id INTEGER NOT NULL, locator TEXT
        );
        CREATE VIEW IF NOT EXISTS v_financial_facts_resolved_current AS
            SELECT * FROM financial_facts;
        """
    )
    conn.execute(
        "INSERT OR REPLACE INTO documents VALUES (1, ?, 'sec_xbrl', "
        "'2026-01-15T00:00:00+00:00', 'https://www.sec.gov/example', 'sec_official')",
        (ticker,),
    )
    conn.executemany(
        "INSERT OR REPLACE INTO financial_facts "
        "(id,ticker,period_end,fiscal_period_type,line_item,value,currency,unit,source_doc_id,locator) "
        "VALUES (?,?,?,?,?,?,?,?,1,NULL)",
        [
            (
                1,
                ticker,
                latest["date"],
                latest["period"],
                "cash_and_short_term_investments",
                latest["cashAndShortTermInvestments"],
                currency,
                "actual",
            ),
            (
                2,
                ticker,
                latest["date"],
                latest["period"],
                "total_debt",
                latest["totalDebt"],
                currency,
                "actual",
            ),
            (
                3,
                ticker,
                latest["date"],
                latest["period"],
                "finance_lease_liability",
                latest.get("financeLeaseLiability", 0),
                currency,
                "actual",
            ),
        ],
    )
    conn.commit()
    conn.close()


def _write_fixture(repo: Path, ticker: str) -> None:
    fmp = repo / "data" / "historical" / "fmp"
    fmp.mkdir(parents=True, exist_ok=True)
    inc, bal, cf = [], [], []
    rev = 250.0
    for year in (2022, 2023, 2024, 2025):
        for q in ("Q1", "Q2", "Q3", "Q4"):
            rev *= 1.03
            inc.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "reportedCurrency": "USD",
                    "date": f"{year}-03-31",
                    "revenue": rev * 1e6,
                    "costOfRevenue": rev * 0.50 * 1e6,
                    "grossProfit": rev * 0.50 * 1e6,
                    "researchAndDevelopmentExpenses": rev * 0.12 * 1e6,
                    "sellingGeneralAndAdministrativeExpenses": rev * 0.15 * 1e6,
                    "operatingExpenses": rev * 0.40 * 1e6,
                    "operatingIncome": rev * 0.12 * 1e6,
                    "netIncome": rev * 0.09 * 1e6,
                    "weightedAverageShsOutDil": 100 * 1e6,
                }
            )
            bal.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "reportedCurrency": "USD",
                    "date": f"{year}-03-31",
                    "cashAndShortTermInvestments": rev * 0.30 * 1e6,
                    "totalCurrentAssets": rev * 0.60 * 1e6,
                    "propertyPlantEquipmentNet": rev * 0.50 * 1e6,
                    "totalAssets": rev * 1.50 * 1e6,
                    "totalCurrentLiabilities": rev * 0.30 * 1e6,
                    "longTermDebt": rev * 0.20 * 1e6,
                    "totalDebt": rev * 0.20 * 1e6,
                    "financeLeaseLiability": 0,
                    "totalStockholdersEquity": rev * 0.80 * 1e6,
                }
            )
            cf.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "depreciationAndAmortization": rev * 0.08 * 1e6,
                    "stockBasedCompensation": rev * 0.05 * 1e6,
                    "changeInWorkingCapital": -rev * 0.01 * 1e6,
                    "operatingCashFlow": rev * 0.15 * 1e6,
                    "capitalExpenditure": -rev * 0.10 * 1e6,
                    "freeCashFlow": rev * 0.05 * 1e6,
                }
            )
    (fmp / f"{ticker}_income_statement_quarterly.json").write_text(
        json.dumps(inc), encoding="utf-8"
    )
    (fmp / f"{ticker}_balance_sheet_quarterly.json").write_text(json.dumps(bal), encoding="utf-8")
    (fmp / f"{ticker}_cash_flow_quarterly.json").write_text(json.dumps(cf), encoding="utf-8")
    (fmp / f"{ticker}_profile.json").write_text(
        json.dumps(
            [
                {
                    "companyName": "Test Co",
                    "sector": "Tech",
                    "beta": 1.2,
                    "price": 50.0,
                    "currency": "USD",
                }
            ]
        ),
        encoding="utf-8",
    )
    est = [
        {
            "date": f"{y}-12-31",
            "revenueAvg": 1100 * (1.10 ** (y - 2026)) * 1e6,
            "netIncomeAvg": 120 * (1.10 ** (y - 2026)) * 1e6,
            "ebitdaAvg": 200 * 1e6,
            "ebitAvg": 150 * 1e6,
            "sgaExpenseAvg": 160 * 1e6,
            "epsAvg": 1.2 * (1.10 ** (y - 2026)),
        }
        for y in range(2026, 2031)
    ]
    (fmp / f"{ticker}_analyst_estimates_annual.json").write_text(json.dumps(est), encoding="utf-8")
    _write_primary_bridge_facts(
        repo,
        ticker,
        cast("dict[str, object]", bal[-1]),
        currency="USD",
    )


def test_builder_produces_valid_nine_sheet_workbook(tmp_path: Path) -> None:
    repo = tmp_path / "repo"
    _write_fixture(repo, "TESTCO")
    dest = tmp_path / "TESTCO.xlsx"
    env = dict(os.environ, DCF_TICKER="TESTCO", DCF_REPO_ROOT=str(repo), DCF_DEST=str(dest))
    proc = subprocess.run(
        [sys.executable, str(BUILDER)],
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    assert proc.returncode == 0, proc.stderr
    assert proc.stdout.startswith("RESULT\tTESTCO"), proc.stdout
    assert dest.exists()

    wb = openpyxl.load_workbook(str(dest))
    assert wb.sheetnames == SHEETS

    # the headline value cell exists on the Valuation walk
    val = wb["Valuation"]
    assert any(
        str(val.cell(r, 1).value).strip() == "VALUE / SHARE" for r in range(1, val.max_row + 1)
    )

    # the two terminal dropdowns survived onto the Dashboard
    dv = {str(d.sqref) for d in wb["Dashboard"].data_validations.dataValidation}
    assert "B43" in dv and "B44" in dv

    # no column-A label accidentally became a formula (the leading-'=' bug) on the
    # text-label sheets (the Monte Carlo sheet legitimately has formula bin labels)
    for name in ("Cover", "Dashboard", "Valuation", "Model"):
        ws = wb[name]
        for row in ws.iter_rows(min_col=1, max_col=1):
            v = row[0].value
            assert not (isinstance(v, str) and v.startswith("=")), (
                f"label formula at {name}!{row[0].coordinate}: {v}"
            )


def test_builder_uses_exact_primary_debt_when_normalized_aggregate_is_missing(
    tmp_path: Path,
) -> None:
    repo = tmp_path / "repo"
    _write_fixture(repo, "TESTCO")
    balance_path = repo / "data/historical/fmp/TESTCO_balance_sheet_quarterly.json"
    balance = json.loads(balance_path.read_text(encoding="utf-8"))
    expected_debt = balance[-1]["totalDebt"]
    for row in balance:
        cash = row.pop("cashAndShortTermInvestments")
        row["cashAndCashEquivalents"] = cash - 7_000_000
        row["shortTermInvestments"] = 7_000_000
        row.pop("totalDebt")
        row["longTermDebt"] = expected_debt + 17_000_000
        row["shortTermDebt"] = 9_000_000
    balance_path.write_text(json.dumps(balance), encoding="utf-8")

    proc = _run_builder(repo, "TESTCO", tmp_path / "TESTCO.xlsx")

    assert proc.returncode == 0, proc.stderr
    country_events: list[dict[str, object]] = []
    for line in proc.stderr.splitlines():
        if not line.startswith("{"):
            continue
        payload_raw: object = json.loads(line)
        if not isinstance(payload_raw, dict):
            continue
        payload = cast("dict[str, object]", payload_raw)
        if payload.get("event") == "dcf_country_risk_context":
            country_events.append(payload)
    assert len(country_events) == 1
    assert country_events[0]["authority"] == "systematic_default_zero"
    assert country_events[0]["source_record"] is None
    inputs = redesign.read_inputs(tmp_path / "TESTCO.xlsx")
    assert inputs is not None
    latest = balance[-1]
    assert inputs.cash_m == pytest.approx(
        (latest["cashAndCashEquivalents"] + latest["shortTermInvestments"]) / 1e6
    )
    assert inputs.total_debt_m == pytest.approx(expected_debt / 1e6)
    assert inputs.total_debt_m != pytest.approx(
        (latest["longTermDebt"] + latest["shortTermDebt"]) / 1e6
    )


def test_builder_fails_loudly_on_partial_debt_or_missing_shares(tmp_path: Path) -> None:
    repo = tmp_path / "partial-debt"
    _write_fixture(repo, "TESTCO")
    balance_path = repo / "data/historical/fmp/TESTCO_balance_sheet_quarterly.json"
    balance = json.loads(balance_path.read_text(encoding="utf-8"))
    for row in balance:
        row.pop("totalDebt")
    balance_path.write_text(json.dumps(balance), encoding="utf-8")
    conn = sqlite3.connect(repo / "data" / "portfolio.db")
    conn.execute("DELETE FROM financial_facts WHERE line_item='total_debt'")
    conn.commit()
    conn.close()

    partial = _run_builder(repo, "TESTCO", tmp_path / "partial.xlsx")

    assert partial.returncode != 0
    assert '"event": "dcf_equity_bridge_unavailable"' in partial.stderr
    assert '"verified_interest_bearing_debt_only"' in partial.stderr

    shares_repo = tmp_path / "missing-shares"
    _write_fixture(shares_repo, "TESTCO")
    income_path = shares_repo / "data/historical/fmp/TESTCO_income_statement_quarterly.json"
    income = json.loads(income_path.read_text(encoding="utf-8"))
    for row in income:
        row.pop("weightedAverageShsOutDil")
    income_path.write_text(json.dumps(income), encoding="utf-8")

    missing_shares = _run_builder(shares_repo, "TESTCO", tmp_path / "shares.xlsx")

    assert missing_shares.returncode != 0
    assert '"positive_diluted_shares"' in missing_shares.stderr


def test_builder_preserves_existing_workbook_when_primary_cash_is_missing(tmp_path: Path) -> None:
    repo = tmp_path / "missing-primary-cash"
    _write_fixture(repo, "TESTCO")
    conn = sqlite3.connect(repo / "data" / "portfolio.db")
    conn.execute("DELETE FROM financial_facts WHERE line_item='cash_and_short_term_investments'")
    conn.commit()
    conn.close()
    destination = tmp_path / "existing.xlsx"
    sentinel = b"existing-workbook-must-survive"
    destination.write_bytes(sentinel)

    result = _run_builder(repo, "TESTCO", destination)

    assert result.returncode != 0
    assert '"event": "dcf_equity_bridge_unavailable"' in result.stderr
    assert '"cash_and_short_term_investments"' in result.stderr
    assert not result.stdout.startswith("RESULT\t")
    assert destination.read_bytes() == sentinel


# --------------------------------------------------------------------------- #
# graceful degradation: short histories (clean SKIP) + base-year segment gaps
# (single-seg fallback) instead of crashing on out-of-range/zero-denominator data
# --------------------------------------------------------------------------- #
def _statement_rows(
    year: int, q: str, rev: float
) -> tuple[dict[str, object], dict[str, object], dict[str, object]]:
    """One quarter of income/balance/cash-flow FMP rows (raw dollars)."""
    inc: dict[str, object] = {
        "fiscalYear": year,
        "period": q,
        "reportedCurrency": "USD",
        "date": f"{year}-03-31",
        "revenue": rev * 1e6,
        "costOfRevenue": rev * 0.50 * 1e6,
        "grossProfit": rev * 0.50 * 1e6,
        "researchAndDevelopmentExpenses": rev * 0.12 * 1e6,
        "sellingGeneralAndAdministrativeExpenses": rev * 0.15 * 1e6,
        "operatingExpenses": rev * 0.40 * 1e6,
        "operatingIncome": rev * 0.12 * 1e6,
        "netIncome": rev * 0.09 * 1e6,
        "weightedAverageShsOutDil": 100 * 1e6,
    }
    bal: dict[str, object] = {
        "fiscalYear": year,
        "period": q,
        "reportedCurrency": "USD",
        "date": f"{year}-03-31",
        "cashAndShortTermInvestments": rev * 0.30 * 1e6,
        "totalCurrentAssets": rev * 0.60 * 1e6,
        "propertyPlantEquipmentNet": rev * 0.50 * 1e6,
        "totalAssets": rev * 1.50 * 1e6,
        "totalCurrentLiabilities": rev * 0.30 * 1e6,
        "longTermDebt": rev * 0.20 * 1e6,
        "totalDebt": rev * 0.20 * 1e6,
        "financeLeaseLiability": 0,
        "totalStockholdersEquity": rev * 0.80 * 1e6,
    }
    cf: dict[str, object] = {
        "fiscalYear": year,
        "period": q,
        "depreciationAndAmortization": rev * 0.08 * 1e6,
        "stockBasedCompensation": rev * 0.05 * 1e6,
        "changeInWorkingCapital": -rev * 0.01 * 1e6,
        "operatingCashFlow": rev * 0.15 * 1e6,
        "capitalExpenditure": -rev * 0.10 * 1e6,
        "freeCashFlow": rev * 0.05 * 1e6,
    }
    return inc, bal, cf


def _write_quarters(
    repo: Path,
    ticker: str,
    quarters: list[tuple[int, str]],
    *,
    seg_quarters: tuple[tuple[int, str], ...] = (),
) -> None:
    """Write income/balance/cash-flow (+ optional product segments + estimates) FMP
    fixtures for an explicit list of (year, quarter) pairs. `seg_quarters` get a
    two-segment product split — pass a subset that omits the base FY to mimic FMP's
    gappy segment disclosure (the LITE shape)."""
    fmp = repo / "data" / "historical" / "fmp"
    fmp.mkdir(parents=True, exist_ok=True)
    inc: list[dict[str, object]] = []
    bal: list[dict[str, object]] = []
    cf: list[dict[str, object]] = []
    pseg: list[dict[str, object]] = []
    seg_set = set(seg_quarters)
    rev = 250.0
    for year, q in quarters:
        rev *= 1.03
        i, b, c = _statement_rows(year, q, rev)
        inc.append(i)
        bal.append(b)
        cf.append(c)
        if (year, q) in seg_set:
            pseg.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "data": {"Components": rev * 0.6 * 1e6, "Systems": rev * 0.4 * 1e6},
                }
            )
    (fmp / f"{ticker}_income_statement_quarterly.json").write_text(
        json.dumps(inc), encoding="utf-8"
    )
    (fmp / f"{ticker}_balance_sheet_quarterly.json").write_text(json.dumps(bal), encoding="utf-8")
    (fmp / f"{ticker}_cash_flow_quarterly.json").write_text(json.dumps(cf), encoding="utf-8")
    if pseg:
        (fmp / f"{ticker}_product_segments_quarterly.json").write_text(
            json.dumps(pseg), encoding="utf-8"
        )
    (fmp / f"{ticker}_profile.json").write_text(
        json.dumps([{"companyName": "Test Co", "beta": 1.2, "price": 50.0, "currency": "USD"}]),
        encoding="utf-8",
    )
    est = [
        {
            "date": f"{y}-12-31",
            "revenueAvg": 1100 * (1.10 ** (y - 2026)) * 1e6,
            "netIncomeAvg": 120 * (1.10 ** (y - 2026)) * 1e6,
            "ebitdaAvg": 200 * 1e6,
            "ebitAvg": 150 * 1e6,
            "sgaExpenseAvg": 160 * 1e6,
            "epsAvg": 1.2 * (1.10 ** (y - 2026)),
        }
        for y in range(2026, 2031)
    ]
    (fmp / f"{ticker}_analyst_estimates_annual.json").write_text(json.dumps(est), encoding="utf-8")
    _write_primary_bridge_facts(repo, ticker, bal[-1], currency="USD")


def _run_builder(repo: Path, ticker: str, dest: Path) -> subprocess.CompletedProcess[str]:
    env = dict(os.environ, DCF_TICKER=ticker, DCF_REPO_ROOT=str(repo), DCF_DEST=str(dest))
    return subprocess.run(
        [sys.executable, str(BUILDER)],
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def test_builder_skips_when_no_quarterly_history(tmp_path: Path) -> None:
    """A name that just IPO'd: FMP returns a profile but no quarterly statements
    (FRVO). The builder must SKIP cleanly, not IndexError on an empty quarter list."""
    repo = tmp_path / "repo"
    fmp = repo / "data" / "historical" / "fmp"
    fmp.mkdir(parents=True, exist_ok=True)
    (fmp / "IPOCO_profile.json").write_text(
        json.dumps([{"companyName": "IPO Co", "price": 20.0, "currency": "USD"}]), encoding="utf-8"
    )
    dest = tmp_path / "IPOCO.xlsx"
    proc = _run_builder(repo, "IPOCO", dest)
    assert proc.returncode == 0, proc.stderr
    assert proc.stdout.startswith("SKIP\tIPOCO"), proc.stdout
    assert "no quarterly FMP history" in proc.stdout
    assert not dest.exists()


def test_builder_skips_when_no_complete_fiscal_year(tmp_path: Path) -> None:
    """A couple of quarters in, but no full four-quarter fiscal year yet — SKIP
    rather than indexing into an empty full-FY list."""
    repo = tmp_path / "repo"
    _write_quarters(repo, "YOUNGCO", [(2025, "Q1"), (2025, "Q2")])
    dest = tmp_path / "YOUNGCO.xlsx"
    proc = _run_builder(repo, "YOUNGCO", dest)
    assert proc.returncode == 0, proc.stderr
    assert proc.stdout.startswith("SKIP\tYOUNGCO"), proc.stdout
    assert "no complete fiscal year yet" in proc.stdout
    assert not dest.exists()


def test_builder_falls_back_to_single_seg_on_base_year_segment_gap(tmp_path: Path) -> None:
    """Segments reported for old years + the newest partial-year quarters, but with a
    gap over the base fiscal year (the LITE shape): the per-segment base revenue is
    zero. The builder must fall back to a single total-company revenue line and still
    produce a real, non-zero valuation instead of dividing by zero (line 444)."""
    repo = tmp_path / "repo"
    quarters = [(y, q) for y in (2022, 2023, 2024, 2025) for q in ("Q1", "Q2", "Q3", "Q4")]
    quarters += [(2026, "Q1"), (2026, "Q2")]
    # segments only for 2022/2023 and the newest 2026 quarters — NOTHING for the base
    # fiscal year (2025), exactly the disclosure gap that zeroed out LITE.
    seg_quarters = tuple(
        [(y, q) for y in (2022, 2023) for q in ("Q1", "Q2", "Q3", "Q4")]
        + [(2026, "Q1"), (2026, "Q2")]
    )
    _write_quarters(repo, "GAPCO", quarters, seg_quarters=seg_quarters)
    dest = tmp_path / "GAPCO.xlsx"
    proc = _run_builder(repo, "GAPCO", dest)
    assert proc.returncode == 0, proc.stderr
    fields = proc.stdout.splitlines()[0].split("\t")
    assert fields[0] == "RESULT" and fields[1] == "GAPCO", proc.stdout
    assert fields[5] == "single", (
        f"expected single-seg fallback, got seg={fields[5]}: {proc.stdout}"
    )
    assert float(fields[2]) > 0, f"degenerate (zero) valuation: {proc.stdout}"
    assert dest.exists()
    assert openpyxl.load_workbook(str(dest)).sheetnames == SHEETS


def test_builder_handles_semiannual_filer(tmp_path: Path) -> None:
    """A semi-annual filer (the BHP shape) reports only H1/H2 as Q2/Q4 each fiscal
    year — the two halves sum to the fiscal year exactly as four quarters do. The
    builder must DETECT the 2-period cadence and build (not SKIP on 'no complete
    fiscal year'), aggregating Q2+Q4 into each FY actual."""
    repo = tmp_path / "repo"
    half_years = [(y, q) for y in (2022, 2023, 2024, 2025) for q in ("Q2", "Q4")]
    _write_quarters(repo, "SEMICO", half_years)
    dest = tmp_path / "SEMICO.xlsx"
    proc = _run_builder(repo, "SEMICO", dest)
    assert proc.returncode == 0, proc.stderr
    fields = proc.stdout.splitlines()[0].split("\t")
    assert fields[0] == "RESULT" and fields[1] == "SEMICO", proc.stdout
    assert float(fields[2]) > 0, f"degenerate valuation: {proc.stdout}"
    assert dest.exists()
    wb = openpyxl.load_workbook(str(dest))
    assert wb.sheetnames == SHEETS
    # the Financials history is half-yearly: Q2/Q4 columns only, never Q1/Q3
    fs = wb["Financials"]
    quarters = {
        str(fs.cell(1, c).value).split()[0]
        for c in range(2, fs.max_column + 1)
        if isinstance(fs.cell(1, c).value, str) and str(fs.cell(1, c).value).startswith("Q")
    }
    assert quarters == {"Q2", "Q4"}, quarters
    # the Model carries FY actual columns built off the half-year sums
    md = wb["Model"]
    fy_headers = [md.cell(1, c).value for c in range(2, md.max_column + 1)]
    assert "FY2025A" in fy_headers and "FY2026E" in fy_headers, fy_headers
