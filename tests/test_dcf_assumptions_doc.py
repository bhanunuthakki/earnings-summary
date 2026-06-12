"""Tests for the DCF assumption-provenance layer (``src/dcf/assumptions_doc.py``):

* the immutable ``opus_baseline`` block — seeded once, never re-seeded, and
  untouched by the workbook→JSON sync that rewrites the ``redesign`` block;
* the ``assumption_overrides`` ledger — records the Opus value + first-seen
  date when a workbook input diverges from the baseline, clears on revert;
* per-input classification (Opus / user-edited / builder default), including
  the beta display-rounding tolerance and the capex unit caveat;
* the Assumptions sheet + Dashboard cell comments, written against an open
  workbook (live current-value formulas, static source columns).

Pure-Python: workbooks are fabricated in-memory (the builder/refresher
integration lives in ``test_dcf_redesign.py``).
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from dcf import assumptions_doc as doc  # noqa: E402
from dcf import redesign  # noqa: E402

_TODAY = date(2026, 6, 12)

_BLOCK: dict[str, object] = {
    "dcf_applicable": True,
    "business_model": "operating",
    "segments": {
        "Total company": {"near_term_growth": 0.10, "terminal_growth": 0.03},
    },
    "near_term_op_margin": 0.20,
    "terminal_op_margin": 0.25,
    "tax_rate": 0.24,
    "capex_pct_revenue_2026": 0.06,
    "terminal_capex_da": 1.05,
    "terminal_method": "Exit multiple",
    "exit_basis": "EV/EBITDA",
    "exit_multiple": 12.0,
    "terminal_growth_g": 0.03,
    "beta": 1.2435,
    "risk_free_rate": 0.043,
    "equity_risk_premium": 0.045,
    "cost_of_debt": 0.045,
    "narrative": "REDESIGN NARRATIVE",
    "reasoning": "KEY JUDGMENTS",
}

_INP = redesign.RedesignInputs(
    segments=("Total company",),
    base_revenue_by_segment={"Total company": 1000.0},
    near_growth_by_segment={"Total company": 0.10},
    terminal_growth_by_segment={"Total company": 0.03},
    near_op_margin=0.20,
    terminal_op_margin=0.25,
    tax_rate=0.24,
    capex_2026_m=60.0,
    terminal_capex_da=1.05,
    da_ratio=0.05,
    consensus_years=5,
    wacc=0.09,
    beta=1.244,  # the builder writes round(beta, 3) into B40
    risk_free_rate=0.043,
    equity_risk_premium=0.045,
    cost_of_debt=0.045,
    terminal_method="Exit multiple",
    terminal_basis="EV/EBITDA",
    exit_multiple=12.0,
    terminal_growth_g=0.03,
    current_price=50.0,
    cash_m=100.0,
    total_debt_m=200.0,
    diluted_shares_m=100.0,
    fx_to_usd=1.0,
)


def _write_json(path: Path, data: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def _dashboard_workbook(segments: tuple[str, ...] = ("Total company",)) -> openpyxl.Workbook:
    """A minimal workbook carrying just a Dashboard with segment labels —
    enough for row construction, sheet placement, and cell comments."""
    wb = openpyxl.Workbook()
    first = wb.active
    assert first is not None
    first.title = redesign.DASHBOARD_SHEET
    for i, s in enumerate(segments):
        first.cell(row=redesign.SEG_ROW0 + i, column=1, value=s)
    return wb


# --------------------------------------------------------------------------- #
# opus_baseline
# --------------------------------------------------------------------------- #
def test_baseline_seeds_once_and_survives_sync_mutation(tmp_path: Path) -> None:
    p = tmp_path / "data" / "dcf_assumptions" / "TESTCO.json"
    _write_json(p, {"redesign": dict(_BLOCK)})

    b1 = doc.ensure_opus_baseline(p, today=_TODAY)
    assert b1 is not None
    assert b1.seeded is True
    assert b1.set_by == "seeded_from_redesign_block"
    assert b1.as_of == "2026-06-12"
    assert b1.scalar("exit_multiple") == 12.0
    assert b1.scalar("capex_pct_revenue_2026") == 0.06
    # prose/flags are not part of the baseline
    assert "narrative" not in b1.values and "dcf_applicable" not in b1.values

    # Simulate the workbook→JSON sync rewriting the redesign block: the
    # baseline must NOT re-seed or follow the mutation.
    data = json.loads(p.read_text(encoding="utf-8"))
    data["redesign"]["exit_multiple"] = 8.0
    _write_json(p, data)
    b2 = doc.ensure_opus_baseline(p, today=date(2027, 1, 1))
    assert b2 is not None
    assert b2.scalar("exit_multiple") == 12.0
    assert b2.as_of == "2026-06-12"  # original seed date, not today


def test_baseline_absent_and_corrupt_cases(tmp_path: Path) -> None:
    assert doc.ensure_opus_baseline(tmp_path / "nope" / "X.json") is None
    p = tmp_path / "NOBLOCK.json"
    _write_json(p, {"ticker": "NOBLOCK"})
    assert doc.ensure_opus_baseline(p) is None
    bad = tmp_path / "BAD.json"
    bad.write_text("{not json", encoding="utf-8")
    with pytest.raises(doc.ProvenanceError):
        doc.ensure_opus_baseline(bad)


def test_baseline_from_opus_pass_is_authoritative() -> None:
    payload = doc.baseline_from_opus_pass(dict(_BLOCK), today=_TODAY)
    assert payload["set_by"] == "opus-4.8"
    assert payload["seeded"] is False
    assert payload["as_of"] == "2026-06-12"
    values = payload["values"]
    assert isinstance(values, dict)
    assert values["exit_multiple"] == 12.0


# --------------------------------------------------------------------------- #
# override ledger
# --------------------------------------------------------------------------- #
def test_ledger_records_divergence_then_clears_on_revert(tmp_path: Path) -> None:
    import dataclasses

    p = tmp_path / "TESTCO.json"
    _write_json(p, {"redesign": dict(_BLOCK)})
    baseline = doc.ensure_opus_baseline(p, today=_TODAY)

    edited = dataclasses.replace(_INP, exit_multiple=8.0)
    ledger = doc.update_override_ledger(p, edited, baseline, today=_TODAY)
    assert ledger["exit_multiple"] == {"opus_value": 12.0, "overridden_on": "2026-06-12"}
    on_disk = json.loads(p.read_text(encoding="utf-8"))["assumption_overrides"]
    assert on_disk["exit_multiple"]["opus_value"] == 12.0

    # A later refresh with the same edit keeps the ORIGINAL first-seen date.
    ledger = doc.update_override_ledger(p, edited, baseline, today=date(2027, 1, 1))
    assert ledger["exit_multiple"]["overridden_on"] == "2026-06-12"

    # Reverting to the Opus value clears the entry.
    ledger = doc.update_override_ledger(p, _INP, baseline, today=date(2027, 1, 1))
    assert "exit_multiple" not in ledger
    assert "exit_multiple" not in json.loads(p.read_text(encoding="utf-8"))["assumption_overrides"]


def test_ledger_beta_display_rounding_is_not_an_override(tmp_path: Path) -> None:
    """Baseline beta 1.2435 reads back as the builder's round-3 1.244 — that
    must classify as Opus, not user-edited."""
    p = tmp_path / "TESTCO.json"
    _write_json(p, {"redesign": dict(_BLOCK)})
    baseline = doc.ensure_opus_baseline(p, today=_TODAY)
    ledger = doc.update_override_ledger(p, _INP, baseline, today=_TODAY)
    assert "beta" not in ledger


# --------------------------------------------------------------------------- #
# classification + sheet + comments
# --------------------------------------------------------------------------- #
def test_rows_classify_opus_user_and_default(tmp_path: Path) -> None:
    import dataclasses

    p = tmp_path / "TESTCO.json"
    block = dict(_BLOCK)
    del block["beta"]  # no Opus WACC drivers for this name -> builder default
    _write_json(p, {"redesign": block})
    baseline = doc.ensure_opus_baseline(p, today=_TODAY)

    edited = dataclasses.replace(_INP, near_op_margin=0.30)
    ledger = doc.update_override_ledger(p, edited, baseline, today=_TODAY)

    wb = _dashboard_workbook()
    rows = {r.label: r for r in doc.build_assumption_rows(wb, edited, baseline, ledger)}

    assert rows["Exit multiple"].source == doc.SOURCE_OPUS
    assert rows["Exit multiple"].baseline_display == "12.0x"

    margin = rows["Near-term operating margin"]
    assert margin.source == doc.SOURCE_USER
    assert "overridden from Opus 20.0% on 2026-06-12" in margin.note

    assert rows["Beta (levered)"].source == doc.SOURCE_DEFAULT
    assert rows["Total company — near-term growth"].source == doc.SOURCE_OPUS

    capex = rows["2026 capex ($M)"]
    assert capex.source == doc.SOURCE_OPUS
    assert "6% of FY1E revenue" in capex.note and "not tracked" in capex.note

    # Scenario Δs at the documented seeds read as seed defaults.
    assert rows["Bull Δ — exit multiple"].source == "seed default"


def test_write_provenance_sheet_comments_and_ledger(tmp_path: Path) -> None:
    import dataclasses

    wb_path = tmp_path / "TESTCO.xlsx"
    _dashboard_workbook().save(str(wb_path))
    p = tmp_path / "data" / "dcf_assumptions" / "TESTCO.json"
    _write_json(p, {"redesign": dict(_BLOCK)})

    edited = dataclasses.replace(_INP, exit_multiple=8.0)
    counts = doc.write_provenance(
        wb_path, edited, p, ticker="TESTCO", update_ledger=True, today=_TODAY
    )
    assert counts[doc.SOURCE_USER] >= 1 and counts[doc.SOURCE_OPUS] >= 1

    wb = openpyxl.load_workbook(str(wb_path))
    try:
        # Placed directly after the Dashboard; read-only documentation sheet.
        names = wb.sheetnames
        assert doc.ASSUMPTIONS_SHEET in names
        assert names.index(doc.ASSUMPTIONS_SHEET) == names.index(redesign.DASHBOARD_SHEET) + 1

        ws = wb[doc.ASSUMPTIONS_SHEET]
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert "REDESIGN NARRATIVE" in text
        assert "KEY JUDGMENTS" in text
        assert "overridden from Opus 12.0x on 2026-06-12" in text
        # Current values are live links, not stale literals.
        assert "=Dashboard!B45" in text

        comment = wb[redesign.DASHBOARD_SHEET]["B45"].comment
        assert comment is not None
        assert "overridden from Opus 12.0x on 2026-06-12" in comment.text
        opus_comment = wb[redesign.DASHBOARD_SHEET]["B29"].comment
        assert opus_comment is not None and "Opus" in opus_comment.text
    finally:
        wb.close()

    # The ledger landed on disk; the baseline stayed at the Opus value.
    data = json.loads(p.read_text(encoding="utf-8"))
    assert data["assumption_overrides"]["exit_multiple"]["opus_value"] == 12.0
    assert data["opus_baseline"]["values"]["exit_multiple"] == 12.0


def test_write_provenance_without_assumptions_file(tmp_path: Path) -> None:
    """No JSON at all: every input is a builder default; the sheet still
    renders (with the no-Opus explainer) and nothing is created on disk."""
    wb_path = tmp_path / "TESTCO.xlsx"
    _dashboard_workbook().save(str(wb_path))
    p = tmp_path / "data" / "dcf_assumptions" / "TESTCO.json"

    counts = doc.write_provenance(
        wb_path, _INP, p, ticker="TESTCO", update_ledger=True, today=_TODAY
    )
    assert doc.SOURCE_USER not in counts
    assert counts[doc.SOURCE_DEFAULT] > 10
    assert not p.exists()

    wb = openpyxl.load_workbook(str(wb_path))
    try:
        ws = wb[doc.ASSUMPTIONS_SHEET]
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert "no Opus assumptions file" in text
    finally:
        wb.close()


def test_sheet_is_replaced_in_place_on_rewrite(tmp_path: Path) -> None:
    wb_path = tmp_path / "TESTCO.xlsx"
    _dashboard_workbook().save(str(wb_path))
    p = tmp_path / "TESTCO.json"
    _write_json(p, {"redesign": dict(_BLOCK)})
    doc.write_provenance(wb_path, _INP, p, ticker="TESTCO", update_ledger=False, today=_TODAY)
    doc.write_provenance(wb_path, _INP, p, ticker="TESTCO", update_ledger=False, today=_TODAY)
    wb = openpyxl.load_workbook(str(wb_path))
    try:
        assert wb.sheetnames.count(doc.ASSUMPTIONS_SHEET) == 1
    finally:
        wb.close()
