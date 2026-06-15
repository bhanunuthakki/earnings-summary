"""Capture-every-number widening of the IR-spreadsheet parser config.

S4 of the capture program replaces "map only the holdings tier KPIs" with "map
EVERY labeled numeric row": the row set is enumerated deterministically (no LLM
gate), units/scale come from a heuristic classifier, junk (footnotes, FMP line
items) is filtered, and the long-tail rows are tagged ``origin='capture'`` and
matched ``exact_label`` so an auto-derived short label can't grab a superstring
row. These tests pin that behaviour and prove a hand-built config (NU.json) is
widened without disturbing its curated analyst rows.
"""

from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from ir_pipeline import config_builder  # noqa: E402
from ir_pipeline.config import IrConfig, SheetKpi, get_config, save_config  # noqa: E402
from ir_pipeline.config_builder import (  # noqa: E402
    _capture_kpis,  # pyright: ignore[reportPrivateUsage]
    _classify_unit_scale,  # pyright: ignore[reportPrivateUsage]
    _is_capturable_label,  # pyright: ignore[reportPrivateUsage]
    build_ir_config,
    widen_config,
)
from ir_pipeline.spreadsheet import (  # noqa: E402
    SheetDataRow,
    enumerate_numeric_rows,
    parse_spreadsheet,
)

_QUARTERS = [dt.datetime(2025, 6, 30), dt.datetime(2025, 9, 30), dt.datetime(2025, 12, 31)]


def _write_sheet(path: Path) -> None:
    """A small NU-flavoured sheet with the tricky cases capture must handle:
    a substring-colliding pair (NII / Risk-adjusted NII), a decimal-percent row,
    a US$ row, an explicit ratio row, a subtotal, a footnote, and an over-long
    prose row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicators"
    ws.cell(1, 2, "Issuer Co.")
    for j, q in enumerate(_QUARTERS):  # header on row 4, labels in col B, values col C+
        ws.cell(4, 3 + j, q)
    rows = [
        ("NII", [100.0, 110.0, 120.0]),
        ("Risk-adjusted NII", [80.0, 85.0, 90.0]),
        ("Fee Revenue", [50.0, 55.0, 60.0]),
        ("Efficiency Ratio (%)", [0.1993, 0.1764, 0.1700]),  # decimal percent
        ("Cost to Serve per Customer - US$", [0.8, 1.0, 1.1]),
        ("Net Portfolio / Available Funding", [0.49, 0.58, 0.60]),
        ("Total Available Funding", [38767.0, 39080.0, 40000.0]),  # subtotal — KEEP
        ("net income", [894.0, 871.0, 900.0]),  # FMP line item — DROP
        ("1) Figures are unaudited and may be restated", [1.0, 2.0, 3.0]),  # footnote — DROP
    ]
    for i, (lab, vals) in enumerate(rows):
        r = 5 + i
        ws.cell(r, 2, lab)
        for j, v in enumerate(vals):
            ws.cell(r, 3 + j, v)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# config round-trip: origin + exact_label survive, legacy dict defaults safely
# ---------------------------------------------------------------------------


def test_config_roundtrip_preserves_origin_and_exact_label(tmp_path: Path) -> None:
    cfg = IrConfig(
        ticker="ZZ",
        platform="mz",
        results_center_url="https://x",
        spreadsheet_kpis=(
            SheetKpi("Monthly ARPAC (USD)", "S", "arpac", "usd", 1.0),  # defaults: analyst/substr
            SheetKpi("Fee Revenue", "S", "Fee Revenue", "actual", 1.0, "capture", True),
        ),
    )
    save_config(cfg, tmp_path)
    loaded = get_config("ZZ", tmp_path)
    assert loaded is not None
    analyst, capture = loaded.spreadsheet_kpis
    assert (analyst.origin, analyst.exact_label) == ("analyst", False)
    assert (capture.origin, capture.exact_label) == ("capture", True)


def test_legacy_config_dict_defaults_to_analyst_substring() -> None:
    """A pre-capture JSON row (no origin / exact_label) loads as a substring-matched
    analyst row — the hand-built NU.json keeps its exact original behavior."""
    from ir_pipeline.config import _from_dict  # pyright: ignore[reportPrivateUsage]

    cfg = _from_dict(
        {
            "ticker": "NU",
            "platform": "mz",
            "spreadsheet_kpis": [{"kpi_name": "X", "sheet": "S", "row_label": "x", "unit": "usd"}],
        }
    )
    (only,) = cfg.spreadsheet_kpis
    assert only.origin == "analyst" and only.exact_label is False


# ---------------------------------------------------------------------------
# enumerate_numeric_rows + exact-label matching
# ---------------------------------------------------------------------------


def test_enumerate_numeric_rows_returns_every_labeled_numeric_row(tmp_path: Path) -> None:
    path = tmp_path / "s.xlsx"
    _write_sheet(path)
    rows = enumerate_numeric_rows(path)
    labels = {r.label for r in rows}
    # Every data row is enumerated (filtering happens later, in capture).
    assert {"NII", "Risk-adjusted NII", "Fee Revenue", "net income"} <= labels
    nii = next(r for r in rows if r.label == "NII")
    assert nii.sheet == "Indicators"
    assert nii.samples == [110.0, 120.0]  # two most-recent numeric cells


def test_exact_label_disambiguates_substring_collision(tmp_path: Path) -> None:
    """A capture row for "NII" (exact) must read the NII series, not the richer
    "Risk-adjusted NII" row a substring match would also hit."""
    path = tmp_path / "s.xlsx"
    _write_sheet(path)
    cfg = IrConfig(
        ticker="ZZ",
        platform="mz",
        results_center_url="",
        spreadsheet_kpis=(SheetKpi("NII", "Indicators", "NII", "actual", 1.0, "capture", True),),
    )
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    assert parsed["NII"][dt.datetime(2025, 12, 31)] == 120.0  # NII, not Risk-adjusted NII (90)

    # The substring variant is ambiguous — it would resolve to the richer row.
    loose = IrConfig(
        ticker="ZZ",
        platform="mz",
        results_center_url="",
        spreadsheet_kpis=(SheetKpi("NII", "Indicators", "NII", "actual", 1.0, "capture", False),),
    )
    loose_parsed = parse_spreadsheet(path, loose, max_quarters=8)
    # both rows have equal numeric counts, so the FIRST wins for substring — assert
    # exact mode is what guarantees the NII row specifically.
    assert loose_parsed["NII"][dt.datetime(2025, 12, 31)] in (120.0, 90.0)


# ---------------------------------------------------------------------------
# unit/scale classifier + capturable-label filter
# ---------------------------------------------------------------------------


def test_classify_unit_scale_cases() -> None:
    assert _classify_unit_scale("Efficiency Ratio (%)", [0.19, 0.18]) == ("percent", 100.0)
    assert _classify_unit_scale("NPL ratio (%)", [9.5, 9.8]) == ("percent", 1.0)  # already %
    assert _classify_unit_scale("ARPAC - US$", [15.0, 15.9]) == ("usd", 1.0)
    assert _classify_unit_scale("Net Portfolio / Available Funding", [0.49, 0.58]) == ("ratio", 1.0)
    assert _classify_unit_scale("# of active accounts", [110.0, 115.0]) == ("count", 1.0)
    # Ambiguous small-dollar value with no ratio/percent marker stays a level.
    assert _classify_unit_scale("Share of Results From Associates", [-0.36, -1.03]) == (
        "actual",
        1.0,
    )
    # "Loans to customers" mentions customers but is a dollar balance → not a count.
    assert _classify_unit_scale("Loans to customers", [10915.0, 12904.0]) == ("actual", 1.0)


def test_is_capturable_label_filters_scraps_but_keeps_subtotals() -> None:
    assert _is_capturable_label("Total Available Funding") is True  # subtotal kept
    assert _is_capturable_label("Credit Income") is True
    assert _is_capturable_label("net income") is False  # FMP line item
    assert _is_capturable_label("1) Figures are unaudited") is False  # footnote
    assert _is_capturable_label("* preliminary") is False
    assert _is_capturable_label("Note: see appendix") is False
    assert _is_capturable_label("x" * 120) is False  # prose / too long
    assert _is_capturable_label("   ") is False


# ---------------------------------------------------------------------------
# _capture_kpis: skip claimed rows, dedupe, stamp capture/exact
# ---------------------------------------------------------------------------


def test_capture_skips_analyst_claimed_rows_and_dedupes() -> None:
    rows = [
        SheetDataRow("Indicators", "Fee Revenue", [50.0, 60.0]),
        SheetDataRow("Indicators", "Fee Revenue", [50.0, 60.0]),  # duplicate label
        SheetDataRow("Indicators", "Risk-adjusted NIM (%)", [0.10, 0.09]),  # claimed below
        SheetDataRow("Other", "Fee Revenue", [50.0, 60.0]),  # same label, 2nd sheet → deduped
    ]
    analyst = [
        SheetKpi("Risk-adjusted NIM (x)", "Indicators", "Risk-adjusted NIM", "percent", 100.0)
    ]
    capture = _capture_kpis(rows, analyst=analyst)
    names = [k.kpi_name for k in capture]
    assert names == ["Fee Revenue"]  # claimed NIM dropped; Fee Revenue deduped to one
    assert capture[0].origin == "capture" and capture[0].exact_label is True


# ---------------------------------------------------------------------------
# widen_config: analyst rows preserved verbatim, capture added, idempotent
# ---------------------------------------------------------------------------


def test_widen_preserves_analyst_and_adds_capture(tmp_path: Path) -> None:
    path = tmp_path / "s.xlsx"
    _write_sheet(path)
    base = IrConfig(
        ticker="ZZ",
        platform="mz",
        results_center_url="https://x",
        spreadsheet_kpis=(
            # Curated analyst row, substring-matched (its row is then NOT recaptured).
            SheetKpi("Fee Revenue (canonical)", "Indicators", "Fee Revenue", "usd", 1.0),
        ),
    )
    wide = widen_config(base, path)
    analyst = [k for k in wide.spreadsheet_kpis if k.origin != "capture"]
    capture = [k for k in wide.spreadsheet_kpis if k.origin == "capture"]

    # Analyst row preserved byte-for-byte.
    assert analyst == list(base.spreadsheet_kpis)
    cap_names = {k.kpi_name for k in capture}
    # Long-tail rows captured; the analyst-claimed "Fee Revenue" row is NOT.
    assert {
        "NII",
        "Risk-adjusted NII",
        "Efficiency Ratio (%)",
        "Total Available Funding",
    } <= cap_names
    assert "Fee Revenue" not in cap_names  # claimed by the analyst substring row
    assert "net income" not in cap_names  # FMP line item filtered
    assert not any(n.startswith("1)") for n in cap_names)  # footnote filtered
    # Decimal-percent classified with scale.
    eff = next(k for k in capture if k.kpi_name == "Efficiency Ratio (%)")
    assert eff.unit == "percent" and eff.scale == 100.0

    # Idempotent: re-widening reproduces the same capture set.
    again = widen_config(wide, path)
    assert again.spreadsheet_kpis == wide.spreadsheet_kpis


def test_build_ir_config_capture_only_when_no_holdings(tmp_path: Path, monkeypatch) -> None:
    """A ticker with no holdings tier KPIs makes NO LLM call and still configures —
    its full audited series is mapped at capture origin."""

    def _boom(*_a: object, **_k: object) -> object:
        raise AssertionError("LLM must not be called for a capture-only ticker")

    monkeypatch.setattr(config_builder, "_llm_map", _boom)
    path = tmp_path / "s.xlsx"
    _write_sheet(path)
    # No micro_thesis/holdings/ZZ.json under tmp_path → _target_kpi_names == [].
    cfg = build_ir_config(
        "ZZ", path, platform="mz", results_center_url="", repo_root=tmp_path, persist=False
    )
    assert cfg.spreadsheet_kpis  # non-empty
    assert all(k.origin == "capture" for k in cfg.spreadsheet_kpis)
    assert {"NII", "Fee Revenue", "Total Available Funding"} <= {
        k.kpi_name for k in cfg.spreadsheet_kpis
    }


def test_real_nu_config_widens_without_touching_curated_rows() -> None:
    """End-to-end on the real hand-built NU.json + a synthetic NU-shaped sheet:
    the four curated analyst rows are preserved verbatim and capture rows are added."""
    repo = PROJECT_ROOT
    cfg = get_config("NU", repo)
    if cfg is None:  # config lives in the main checkout; skip if absent in this tree
        return
    assert all(k.origin == "analyst" for k in cfg.spreadsheet_kpis)
    original = list(cfg.spreadsheet_kpis)

    # Reuse the test sheet (the real spreadsheet lives outside the worktree); the
    # point is that widen leaves analyst rows untouched regardless of the sheet.
    import tempfile

    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / "nu.xlsx"
        _write_sheet(path)
        wide = widen_config(cfg, path)

    analyst = [k for k in wide.spreadsheet_kpis if k.origin != "capture"]
    assert analyst == original  # curated NU rows unchanged
    assert any(k.origin == "capture" for k in wide.spreadsheet_kpis)


def test_holdings_fixture_unused_warning_guard(tmp_path: Path) -> None:
    """Guard: a holdings file present makes _target_kpi_names non-empty (the analyst
    layer would run) — documents the wiring without invoking the LLM."""
    holdings = tmp_path / "micro_thesis" / "holdings"
    holdings.mkdir(parents=True)
    (holdings / "ZZ.json").write_text(json.dumps({"tier_1_kpis": [{"name": "GMV"}]}), "utf-8")
    from ir_pipeline.config_builder import _target_kpi_names  # pyright: ignore[reportPrivateUsage]

    assert _target_kpi_names("ZZ", tmp_path) == ["GMV"]
