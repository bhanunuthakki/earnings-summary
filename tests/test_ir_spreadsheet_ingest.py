"""End-to-end ingest of a capture-widened IR-spreadsheet config into kpi_facts.

Proves the S4 capture path on a real (synthetic) spreadsheet through the real
parse → ingest pipeline:

* capture rows land at IR_DOC tier with ``definition_origin='capture'``;
* a curated analyst series is NOT disturbed (regression) — its definition stays
  ``analyst`` and the claimed row is never re-minted as a capture duplicate;
* a captured label canonicalizes onto an existing series instead of fragmenting;
* an audited spreadsheet value SUPERSEDES the lower-tier LLM value for the same
  metric (tier precedence — the interaction with S3's broad LLM capture).
"""

from __future__ import annotations

import datetime as dt
import sqlite3
import sys
from collections.abc import Callable
from pathlib import Path
from typing import ParamSpec, TypeVar, cast

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import ir_pipeline.ingest as ingest_module  # noqa: E402
from ir_pipeline.config import IrConfig, SheetKpi  # noqa: E402
from ir_pipeline.config_builder import build_ir_config, widen_config  # noqa: E402
from ir_pipeline.ingest import ingest_spreadsheet_kpis  # noqa: E402
from ir_pipeline.spreadsheet import parse_spreadsheet  # noqa: E402
from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

_Q = [dt.datetime(2025, 6, 30), dt.datetime(2025, 9, 30), dt.datetime(2025, 12, 31)]

_P = ParamSpec("_P")
_R = TypeVar("_R")


def _capture_kwargs(
    function: Callable[_P, _R],
    captured: dict[str, object],
) -> Callable[_P, _R]:
    def wrapped(*args: _P.args, **kwargs: _P.kwargs) -> _R:
        captured.update(kwargs)
        return function(*args, **kwargs)

    return wrapped


_SCHEMA = """
CREATE TABLE kpi_definitions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticker TEXT NOT NULL, name TEXT NOT NULL, unit TEXT NOT NULL,
    primary_source TEXT NOT NULL, fallback_source TEXT, ir_url TEXT,
    threshold_tier TEXT, threshold_low FLOAT, threshold_high FLOAT, notes TEXT,
    definition_origin TEXT NOT NULL DEFAULT 'analyst',
    reporting_cadence TEXT NOT NULL DEFAULT 'quarterly',
    UNIQUE(ticker, name)
);
CREATE TABLE kpi_facts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticker TEXT NOT NULL, period_end TIMESTAMP NOT NULL, fiscal_period_type TEXT NOT NULL,
    kpi_definition_id INTEGER NOT NULL, value NUMERIC(24,6) NOT NULL, unit TEXT NOT NULL,
    source_doc_id INTEGER NOT NULL, confidence FLOAT NOT NULL DEFAULT 1.0,
    extracted_by TEXT, supersedes_id INTEGER, locator TEXT, source_excerpt TEXT
);
CREATE UNIQUE INDEX uq_kpi_facts_provenance
    ON kpi_facts (ticker, period_end, fiscal_period_type, kpi_definition_id, source_doc_id);
CREATE TABLE validation_issues (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id TEXT NOT NULL, source_doc_id INTEGER, ticker TEXT,
    severity TEXT NOT NULL, rule TEXT NOT NULL, raw_value TEXT, expected TEXT,
    raised_at TIMESTAMP NOT NULL, resolved_at TIMESTAMP
);
CREATE TABLE documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticker TEXT, source_type TEXT, doc_type TEXT, period_end TIMESTAMP,
    file_path TEXT, sha256 TEXT, fetched_at TIMESTAMP, fetch_status TEXT,
    raw_bytes_size INTEGER, source_quality_tier TEXT
);
CREATE TABLE ingestion_runs (
    run_id TEXT PRIMARY KEY, started_at TIMESTAMP, ended_at TIMESTAMP,
    directive TEXT, ticker_scope TEXT, status TEXT, error_summary TEXT
);
"""


def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    c.executescript(_SCHEMA)
    c.commit()
    return c


def _write_sheet(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Indicators"
    for j, q in enumerate(_Q):  # header row 4, labels col B, values col C+
        ws.cell(4, 3 + j, q)
    rows = [
        ("Fee Revenue", [50.0, 55.0, 60.0]),
        ("NII", [100.0, 110.0, 120.0]),
        ("Net interest margin (%)", [0.100, 0.095, 0.094]),  # decimal percent
        ("Total Available Funding", [38000.0, 39000.0, 40000.0]),
    ]
    for i, (lab, vals) in enumerate(rows):
        r = 5 + i
        ws.cell(r, 2, lab)
        for j, v in enumerate(vals):
            ws.cell(r, 3 + j, v)
    wb.save(str(path))


def _seed_doc(conn: sqlite3.Connection, *, ticker: str, period: dt.datetime) -> int:
    cur = conn.execute(
        "INSERT INTO documents (ticker, source_type, doc_type, period_end, fetched_at, "
        "fetch_status) VALUES (?, 'ir_doc', 'press_release', ?, ?, 'ok')",
        (ticker, period, dt.datetime(2026, 1, 1)),
    )
    conn.commit()
    return int(cur.lastrowid or 0)


def _defs(conn: sqlite3.Connection, ticker: str) -> dict[str, sqlite3.Row]:
    return {
        str(r["name"]): r
        for r in conn.execute(
            "SELECT * FROM kpi_definitions WHERE ticker = ?", (ticker.upper(),)
        ).fetchall()
    }


def _facts(conn: sqlite3.Connection, def_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT * FROM kpi_facts WHERE kpi_definition_id = ? ORDER BY period_end", (def_id,)
    ).fetchall()


def _series(conn: sqlite3.Connection, def_id: int) -> dict[str, float]:
    """A definition's facts as ``{period_end_str: value}``.

    This connection registers no datetime converter (mirroring the prod
    connection), so ``period_end`` reads back as the adapted string — key on
    ``str(period)`` to look one up, exactly as the schema stores it.
    """
    return {str(f["period_end"]): float(f["value"]) for f in _facts(conn, def_id)}


_Q4 = str(_Q[2])  # the 2025-12-31 period as stored (adapted string)


# ---------------------------------------------------------------------------


def test_capture_only_ticker_ingests_at_ir_doc_with_capture_origin(tmp_path: Path) -> None:
    conn = _conn()
    path = tmp_path / "co.xlsx"
    _write_sheet(path)
    cfg = build_ir_config(
        "CO", path, platform="mz", results_center_url="", repo_root=tmp_path, persist=False
    )
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    inserted, doc_id = ingest_spreadsheet_kpis(conn, "CO", cfg, parsed, path)

    assert inserted > 0
    # The spreadsheet document is registered at IR_DOC tier.
    doc = conn.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    assert doc["source_type"] == "ir_doc" and doc["doc_type"] == "ir_historical_spreadsheet"

    defs = _defs(conn, "CO")
    assert {"Fee Revenue", "NII", "Net interest margin (%)", "Total Available Funding"} <= set(defs)
    # Every minted definition is capture origin (no curated watchlist for CO).
    assert all(d["definition_origin"] == "capture" for d in defs.values())
    # Facts carry the deterministic IR-spreadsheet extractor tag at the IR doc.
    nii_facts = _facts(conn, int(defs["NII"]["id"]))
    assert len(nii_facts) == len(_Q)
    assert all(
        f["extracted_by"] == "ir_spreadsheet" and f["source_doc_id"] == doc_id for f in nii_facts
    )
    # Decimal percent scaled (0.094 -> 9.4%).
    nim = _series(conn, int(defs["Net interest margin (%)"]["id"]))
    assert round(nim[_Q4], 1) == 9.4
    assert [
        row[0]
        for row in conn.execute(
            "SELECT status FROM ingestion_runs WHERE directive='ingest_ir_spreadsheet'"
        ).fetchall()
    ] == ["ok"]


def test_ingest_fingerprints_source_config_and_parsed_values(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    conn = _conn()
    path = tmp_path / "co.xlsx"
    _write_sheet(path)
    cfg = build_ir_config(
        "CO", path, platform="mz", results_center_url="", repo_root=tmp_path, persist=False
    )
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    captured: dict[str, object] = {}
    real_start = ingest_module.start_run
    monkeypatch.setattr(ingest_module, "start_run", _capture_kwargs(real_start, captured))
    ingest_spreadsheet_kpis(conn, "CO", cfg, parsed, path)

    inputs_obj = captured["invocation_inputs"]
    assert isinstance(inputs_obj, dict)
    inputs = cast(dict[str, object], inputs_obj)
    source_obj = inputs["source"]
    assert isinstance(source_obj, dict)
    source = cast(dict[str, object], source_obj)
    assert isinstance(source["sha256"], str)
    assert isinstance(inputs["config_sha256"], str)
    assert isinstance(inputs["parsed_sha256"], str)
    assert captured["deduplicate_completed"] is True


def test_curated_analyst_series_not_disturbed_by_widening(tmp_path: Path) -> None:
    """Regression: a hand-built analyst row keeps its definition + origin, and the
    row it claims is never re-minted as a capture duplicate."""
    conn = _conn()
    # Seed the curated analyst definition + a prior fact for one period.
    conn.execute(
        "INSERT INTO kpi_definitions (ticker, name, unit, primary_source, definition_origin) "
        "VALUES ('CO', 'Net Fee Revenue (USD)', 'actual', 'ir_doc', 'analyst')"
    )
    analyst_id = int(conn.execute("SELECT id FROM kpi_definitions").fetchone()["id"])
    prior_doc = _seed_doc(conn, ticker="CO", period=_Q[0])
    conn.execute(
        "INSERT INTO kpi_facts (ticker, period_end, fiscal_period_type, kpi_definition_id, "
        "value, unit, source_doc_id, extracted_by) VALUES "
        "('CO', ?, 'Q2', ?, '50', 'actual', ?, 'llm:haiku')",
        (_Q[0], analyst_id, prior_doc),
    )
    conn.commit()

    path = tmp_path / "co.xlsx"
    _write_sheet(path)
    # A curated analyst config that maps the "Fee Revenue" row to the canonical name.
    base = IrConfig(
        ticker="CO",
        platform="mz",
        results_center_url="",
        spreadsheet_kpis=(
            SheetKpi("Net Fee Revenue (USD)", "Indicators", "Fee Revenue", "usd", 1.0),
        ),
    )
    cfg = widen_config(base, path)
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    ingest_spreadsheet_kpis(conn, "CO", cfg, parsed, path)

    defs = _defs(conn, "CO")
    # The curated definition still exists, still analyst origin.
    assert defs["Net Fee Revenue (USD)"]["id"] == analyst_id
    assert defs["Net Fee Revenue (USD)"]["definition_origin"] == "analyst"
    # No capture duplicate for the claimed "Fee Revenue" row.
    assert "Fee Revenue" not in defs
    # The analyst series received the spreadsheet's values (IR_DOC).
    assert _series(conn, analyst_id)[_Q4] == 60.0
    # Capture rows for the OTHER labels still landed at capture origin.
    assert defs["NII"]["definition_origin"] == "capture"


def test_capture_label_reuses_existing_definition(tmp_path: Path) -> None:
    """A captured raw label canonicalizes onto an existing same-metric series
    (unit-only difference) rather than minting a fragmented duplicate."""
    conn = _conn()
    conn.execute(
        "INSERT INTO kpi_definitions (ticker, name, unit, primary_source, definition_origin) "
        "VALUES ('CO', 'NII (USD)', 'actual', 'ir_doc', 'capture')"
    )
    existing_id = int(conn.execute("SELECT id FROM kpi_definitions").fetchone()["id"])
    conn.commit()

    path = tmp_path / "co.xlsx"
    _write_sheet(path)
    cfg = build_ir_config(
        "CO", path, platform="mz", results_center_url="", repo_root=tmp_path, persist=False
    )
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    ingest_spreadsheet_kpis(conn, "CO", cfg, parsed, path)

    defs = _defs(conn, "CO")
    # "NII" canonicalized onto "NII (USD)" — no bare "NII" duplicate minted.
    assert "NII" not in defs
    assert _series(conn, existing_id)[_Q4] == 120.0  # the existing series received the values


def test_ir_doc_value_supersedes_llm_incumbent_for_same_metric(tmp_path: Path) -> None:
    """Tier precedence: an LLM-extracted value for a metric the spreadsheet also
    carries is superseded by the audited IR_DOC value (the S3↔S4 interaction)."""
    conn = _conn()
    # An LLM capture (S3) already minted "NII" and stored a value for 2025-12-31.
    conn.execute(
        "INSERT INTO kpi_definitions (ticker, name, unit, primary_source, definition_origin) "
        "VALUES ('CO', 'NII', 'actual', 'ir_doc', 'capture')"
    )
    nii_id = int(conn.execute("SELECT id FROM kpi_definitions").fetchone()["id"])
    llm_doc = _seed_doc(conn, ticker="CO", period=_Q[2])
    conn.execute(
        "INSERT INTO kpi_facts (ticker, period_end, fiscal_period_type, kpi_definition_id, "
        "value, unit, source_doc_id, extracted_by) VALUES "
        "('CO', ?, 'Q4', ?, '999', 'actual', ?, 'llm:claude-haiku')",
        (_Q[2], nii_id, llm_doc),
    )
    conn.commit()

    path = tmp_path / "co.xlsx"
    _write_sheet(path)
    cfg = build_ir_config(
        "CO", path, platform="mz", results_center_url="", repo_root=tmp_path, persist=False
    )
    parsed = parse_spreadsheet(path, cfg, max_quarters=8)
    ingest_spreadsheet_kpis(conn, "CO", cfg, parsed, path)

    q4 = [f for f in _facts(conn, nii_id) if str(f["period_end"]) == _Q4]
    # History is append-only: the audited row points to the lower-tier incumbent.
    assert len(q4) == 2
    assert q4[0]["extracted_by"] == "llm:claude-haiku"
    assert q4[1]["extracted_by"] == "ir_spreadsheet"
    assert q4[1]["supersedes_id"] == q4[0]["id"]
    assert float(q4[1]["value"]) == 120.0


def test_post_cutover_ingest_dual_writes_evidence_and_preserves_history(
    tmp_path: Path,
    migrated_db: Callable[..., Path],
) -> None:
    repo_root = tmp_path / "repo"
    db_path = migrated_db(repo_root / "data" / "portfolio.db")
    spreadsheet_dir = repo_root / "data" / "ir_spreadsheets" / "CO"
    spreadsheet_dir.mkdir(parents=True)
    first_path = spreadsheet_dir / "co-first.xlsx"
    second_path = spreadsheet_dir / "co-second.xlsx"
    _write_sheet(first_path)
    _write_sheet(second_path)
    workbook = openpyxl.load_workbook(second_path)
    workbook["Indicators"].cell(6, 5, 121.0)
    workbook.save(second_path)
    cfg = IrConfig(
        ticker="CO",
        platform="mz",
        results_center_url="https://example.invalid/results",
        spreadsheet_kpis=(SheetKpi("NII", "Indicators", "NII", "actual"),),
    )
    period = dt.datetime(2025, 12, 31)

    conn = connect_sqlite(
        db_path,
        role=SQLiteConnectionRole.WRITER,
        schema_preflight=True,
    )
    conn.row_factory = sqlite3.Row
    try:
        first_inserted, first_doc = ingest_spreadsheet_kpis(
            conn,
            "CO",
            cfg,
            {"NII": {period: 120.0}},
            first_path,
            repo_root=repo_root,
        )
        conn.execute(
            "UPDATE documents SET fetched_at = '2025-01-01 00:00:00' WHERE id = ?",
            (first_doc,),
        )
        second_inserted, second_doc = ingest_spreadsheet_kpis(
            conn,
            "CO",
            cfg,
            {"NII": {period: 121.0}},
            second_path,
            repo_root=repo_root,
        )
        facts = conn.execute(
            "SELECT id, value, supersedes_id, source_doc_id FROM kpi_facts "
            "WHERE ticker = 'CO' ORDER BY id"
        ).fetchall()
        evidence_count = conn.execute(
            "SELECT COUNT(*) FROM evidence_document_versions WHERE legacy_document_id IN (?, ?)",
            (first_doc, second_doc),
        ).fetchone()[0]
        revision_count = conn.execute(
            "SELECT COUNT(*) FROM fact_observation_revisions "
            "WHERE fact_table = 'kpi_facts' AND fact_row_id IN (?, ?)",
            (facts[0]["id"], facts[1]["id"]),
        ).fetchone()[0]
    finally:
        conn.close()

    assert (first_inserted, second_inserted) == (1, 1)
    assert evidence_count == 2
    assert revision_count == 2
    assert len(facts) == 2
    assert facts[1]["supersedes_id"] == facts[0]["id"]
    assert (facts[0]["source_doc_id"], facts[1]["source_doc_id"]) == (
        first_doc,
        second_doc,
    )
