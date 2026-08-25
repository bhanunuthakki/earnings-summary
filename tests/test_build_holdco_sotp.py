# pyright: reportPrivateUsage=false
"""Unit tests for the BN holdco SOTP (execution/build_holdco_sotp.py) — the
2026-06 calibration contract:

  * the value-of-record matches the hand-computed bucket sums;
  * the carry bucket is disclosure-shaped (legacy margin, not an employee-pool
    deduction; future term takes BN-NET dollars);
  * scenarios follow the S6 bear/base/bull convention and order;
  * marks load from data/dcf_assumptions/<T>.json["sotp"]["marks"], yellow-cell
    edits in a v2 workbook override them (price never captured), and effective
    values sync back to the JSON without touching the justification notes;
  * the in-sheet formula chain evaluates (via the `formulas` lib) to exactly
    the Python mirror.
"""

from __future__ import annotations

import json
import sqlite3
import sys
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path
from typing import cast

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "execution"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import build_holdco_sotp as h  # noqa: E402


@pytest.fixture
def tmp_repo(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Point the module at an empty repo root + workbook dest."""
    monkeypatch.setattr(h, "REPO", tmp_path)
    monkeypatch.setattr(h, "DEST", tmp_path / "dcf" / "BN.xlsx")
    return tmp_path


def _write_marks(repo: Path, marks: dict[str, object]) -> Path:
    path = repo / "data" / "dcf_assumptions" / "BN.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"ticker": "BN", "sotp": {"marks": marks}}), encoding="utf-8")
    return path


# --------------------------------------------------------------------------- #
# mirror math
# --------------------------------------------------------------------------- #
def test_value_matches_hand_computed_buckets() -> None:
    s = h.Sotp()
    am = s.bam_fre * s.fre_mult * s.bn_own
    carry = (
        s.carry_accrued_gross * s.carry_legacy_margin * (1 - s.carry_haircut)
        + s.carry_future_net_annual * s.carry_future_mult
    ) * (1 - s.carry_tax)
    bws = s.bws_de * s.bws_mult
    ic = s.ic_listed + s.ic_private * (1 - s.ic_re_haircut)
    corp = s.corp_recourse_debt + s.corp_preferred + s.corp_overhead_pv
    eq, vps = h.value(s)
    assert abs(eq - (am + carry + bws + ic - corp)) < 1e-9
    assert abs(vps - eq * 1000.0 / s.shares_m) < 1e-9


def test_carry_is_disclosure_shaped() -> None:
    """Legacy margin applies to the accrued stack only; the future term takes
    BN-net dollars straight (the 33% royalty is already inside the net figure)."""
    s = h.Sotp(
        carry_accrued_gross=10.0,
        carry_legacy_margin=0.65,
        carry_haircut=0.30,
        carry_future_net_annual=2.0,
        carry_future_mult=5.0,
        carry_tax=0.0,
    )
    assert abs(h._carry(s) - (10.0 * 0.65 * 0.70 + 2.0 * 5.0)) < 1e-9


def test_scenarios_bear_below_base_below_bull() -> None:
    s = h.Sotp()
    base = h.value(s)[1]
    bull = h._scn(s, carry_haircut=0.0, ic_re_haircut=0.0)
    bear = h._scn(s, carry_zero=True, ic_private=0.0)
    assert bear < base < bull
    # bear really zeroes the contested buckets: AM + BWS + listed - corp only
    floor = (
        s.bam_fre * s.fre_mult * s.bn_own
        + s.bws_de * s.bws_mult
        + s.ic_listed
        - (s.corp_recourse_debt + s.corp_preferred + s.corp_overhead_pv)
    )
    assert abs(bear - floor * 1000.0 / s.shares_m) < 1e-9


def test_reverse_valuation_is_the_existing_carry_private_re_residual() -> None:
    s = h.Sotp()
    eq, vps = h.value(s)

    snapshot = h.reverse_valuation(s, eq, vps)

    assert snapshot is not None
    levers = cast("list[dict[str, object]]", snapshot["levers"])
    lever = levers[0]
    expected = s.price * s.shares_m / 1000.0 - (h._am(s) + h._bws(s) + s.ic_listed - h._corp(s))
    assert lever["id"] == "implied_carry_private_re_value"
    assert lever["method"] == "bridge_residual"
    assert lever["implied_value"] == pytest.approx(expected)


# --------------------------------------------------------------------------- #
# JSON marks + capture-inject + sync-back
# --------------------------------------------------------------------------- #
def test_json_marks_override_defaults(tmp_repo: Path) -> None:
    _write_marks(
        tmp_repo,
        {
            "bws_mult": 15.0,  # bare number form
            "ic_listed": {"value": 20.0, "note": "test note"},  # rich form
            "unknown_field": {"value": 1.0},  # ignored, not an error
        },
    )
    s, notes = h._load("BN")
    assert s.bws_mult == 15.0
    assert s.ic_listed == 20.0
    assert notes["ic_listed"] == "test note"
    assert not hasattr(s, "unknown_field")


def test_workbook_edit_survives_rebuild(tmp_repo: Path) -> None:
    """The capture-inject loop: edit a yellow cell, reload — the edit wins over
    the JSON mark; a rebuild then writes the edited value back into the sheet."""
    _write_marks(tmp_repo, {"bws_mult": 12.0})
    s, _ = h._load("BN")
    assert s.bws_mult == 12.0
    h.build(s, h.DEST)
    wb = openpyxl.load_workbook(h.DEST)
    wb["Dashboard"].cell(row=h._SOTP_ROW["bws_mult"], column=2, value=14.5)
    wb.save(h.DEST)
    s2, _ = h._load("BN")
    assert s2.bws_mult == 14.5  # workbook edit beats the JSON's 12.0


def test_price_is_never_captured(tmp_repo: Path) -> None:
    s = h.Sotp()
    h.build(s, h.DEST)
    wb = openpyxl.load_workbook(h.DEST)
    wb["Dashboard"].cell(row=h._SOTP_ROW["price"], column=2, value=999.0)
    wb.save(h.DEST)
    s2, _ = h._load("BN")
    assert s2.price == h.Sotp().price  # no FMP profile in tmp → default, not 999


def test_v1_workbook_without_marker_is_not_captured(tmp_repo: Path) -> None:
    """Pre-calibration workbooks must not re-poison the calibrated marks."""
    s = h.Sotp()
    h.build(s, h.DEST)
    wb = openpyxl.load_workbook(h.DEST)
    wb["Dashboard"]["D1"] = None  # strip the v2 marker (≈ a legacy workbook)
    wb["Dashboard"].cell(row=h._SOTP_ROW["bam_fre"], column=2, value=4.0)
    wb.save(h.DEST)
    assert h._capture_bn_inputs(h.DEST) == {}
    s2, _ = h._load("BN")
    assert s2.bam_fre == h.Sotp().bam_fre


def test_sync_back_updates_values_and_preserves_notes(tmp_repo: Path) -> None:
    path = _write_marks(tmp_repo, {"bws_mult": {"value": 12.0, "note": "keep me"}})
    s = h.Sotp(bws_mult=14.0)
    assert h._sync_sotp_json("BN", s)
    data = json.loads(path.read_text(encoding="utf-8"))
    marks = data["sotp"]["marks"]
    assert marks["bws_mult"] == {"value": 14.0, "note": "keep me"}
    assert marks["ic_listed"]["value"] == s.ic_listed  # created for unlisted fields
    assert "price" not in marks  # live, never synced
    assert data["sotp"]["last_synced"]


def test_bn_owner_json_is_not_mutated_when_dcf_persistence_does_not_succeed(
    tmp_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_marks(tmp_repo, {"bws_mult": {"value": 12.0, "note": "keep me"}})
    before = path.read_text(encoding="utf-8")
    s = h.Sotp(bws_mult=14.0)
    monkeypatch.setattr(h, "persist_dcf_run", lambda *_args, **_kwargs: False)

    persisted, sync_result = h._persist_then_sync_bn(s, *h.value(s), {"marks": {}})

    assert persisted is False
    assert sync_result.status.startswith("not_attempted")
    assert path.read_text(encoding="utf-8") == before


def test_holdco_persistence_degrades_truthfully_on_legacy_schema_without_sync_columns(
    tmp_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    db_path = tmp_repo / "data" / "portfolio.db"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.touch()
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE dcf_runs (id INTEGER PRIMARY KEY, ticker TEXT)")
    captured: list[object] = []

    @contextmanager
    def fake_connect(*_args: object, **_kwargs: object) -> Iterator[sqlite3.Connection]:
        yield conn

    monkeypatch.setattr(h, "connect_sqlite", fake_connect)
    assert h.persist_mod is not None
    monkeypatch.setattr(h.persist_mod, "upsert", lambda _conn, row: captured.append(row))

    s = h.Sotp()
    eq, vps = h.value(s)
    persisted = h.persist_dcf_run(
        eq,
        vps,
        s.price,
        s.ke,
        {"model": "holdco_sotp", "marks": {}},
        h.SyncResult("synced", datetime(2026, 8, 25, tzinfo=UTC).replace(tzinfo=None)),
    )

    assert persisted is True
    assert captured
    row = captured[0]
    assert getattr(row, "assumptions_sync_status") is None
    assert getattr(row, "assumptions_synced_at") is None
    assert getattr(row, "provenance") is None


# --------------------------------------------------------------------------- #
# workbook artifacts
# --------------------------------------------------------------------------- #
def test_scenarios_sheet_carries_bear_base_bull_statics(tmp_repo: Path) -> None:
    s = h.Sotp()
    h.build(s, h.DEST, scenarios=(30.0, 48.0, 55.0))
    ws = openpyxl.load_workbook(h.DEST)["Scenarios"]
    assert [ws.cell(row=r, column=2).value for r in (4, 5, 6)] == [30.0, 48.0, 55.0]
    assert "Bear" in str(ws["A4"].value)
    assert "Bull" in str(ws["A6"].value)


def test_dashboard_notes_render_in_column_c(tmp_repo: Path) -> None:
    s = h.Sotp()
    h.build(s, h.DEST, notes={"bws_mult": "why thirteen"})
    ws = openpyxl.load_workbook(h.DEST)["Dashboard"]
    assert ws.cell(row=h._SOTP_ROW["bws_mult"], column=3).value == "why thirteen"


def test_in_sheet_formulas_match_python_mirror(tmp_repo: Path) -> None:
    """The bank-model precedent: evaluate the workbook with the `formulas` lib
    and require the in-sheet value/share to equal the value-of-record."""
    formulas = pytest.importorskip("formulas")
    s = h.Sotp()
    h.build(s, h.DEST)
    xl = formulas.ExcelModel().loads(str(h.DEST)).finish()
    sol = xl.calculate()
    keys = [k for k in sol if k.upper().endswith("'[BN.XLSX]SOTP'!B9")]
    assert keys, "SOTP!B9 (value/share) not found in the evaluated model"
    sheet_vps = float(sol[keys[0]].value[0, 0])
    assert abs(sheet_vps - h.value(s)[1]) < 1e-6
