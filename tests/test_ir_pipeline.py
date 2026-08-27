"""IR-spreadsheet parser: layout auto-detection, label match, unit scaling.

Builds a synthetic "Managerial indicators" sheet mirroring Nu Holdings' Historical
Data file (offset header row, labels in column B, quarters across columns, and a
decimal-stored percent row) and asserts the real NU config extracts the right
canonical KPI series with `scale` applied.
"""

from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from ir_pipeline import config_builder  # noqa: E402
from ir_pipeline.config import (  # noqa: E402
    IrConfig,
    SheetKpi,
    get_config,
    save_config,
)
from ir_pipeline.discover import (  # noqa: E402
    discover_documents,
    discover_history_hybrid,
)
from ir_pipeline.discover._docmeta import CandidateDoc, classify  # noqa: E402
from ir_pipeline.spreadsheet import (  # noqa: E402
    _header_row,
    _parse_period,
    parse_spreadsheet,
)

_QUARTERS = [
    dt.datetime(2025, 3, 31),
    dt.datetime(2025, 6, 30),
    dt.datetime(2025, 9, 30),
    dt.datetime(2025, 12, 31),
]


def _write_nu_like_sheet(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Managerial indicators"
    ws.cell(1, 2, "Nu Holdings Ltd.")
    # Header row offset down (mirrors the real file's row 9); labels in col B (2),
    # quarter dates from col C (3). One date is a dd/mm/yyyy string like the real file.
    for j, q in enumerate(_QUARTERS):
        cell = q.strftime("%d/%m/%Y") if j == 3 else q
        ws.cell(9, 3 + j, cell)
    ws.cell(11, 2, "Average Revenue per Active Customer (ARPAC) - US$")
    for j, v in enumerate([11.6, 12.5, 13.8, 15.0]):
        ws.cell(11, 3 + j, v)
    # Decimal-stored percent: 0.0928 -> 9.28% after scale=100.
    ws.cell(13, 2, "Risk-adjusted NIM (%)")
    for j, v in enumerate([0.0928, 0.099, 0.108, 0.105]):
        ws.cell(13, 3 + j, v)
    wb.save(str(path))


def test_parse_period_handles_datetime_and_ddmmyyyy() -> None:
    assert _parse_period(dt.datetime(2026, 3, 31)) == dt.datetime(2026, 3, 31)
    assert _parse_period("31/03/2026") == dt.datetime(2026, 3, 31)
    assert _parse_period("2026-03-31 00:00:00") == dt.datetime(2026, 3, 31)
    assert _parse_period("not a date") is None


def test_q4cdn_config_has_no_precise_adapter() -> None:
    config = IrConfig(
        ticker="ZZZ",
        platform="q4cdn",
        results_center_url="https://example.test/results",
    )
    with pytest.raises(ValueError, match=r"precise|generic|direct"):
        discover_documents(config)


def test_q4cdn_config_uses_generic_history_fallback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = IrConfig(
        ticker="ZZZ",
        platform="q4cdn",
        results_center_url="https://example.test/results",
    )
    calls: list[tuple[str, int, int]] = []

    def fake_history(*, ir_url: str, max_quarters: int, timeout_ms: int) -> list[CandidateDoc]:
        calls.append((ir_url, max_quarters, timeout_ms))
        return []

    monkeypatch.setattr("ir_pipeline.discover.generic.discover_document_history", fake_history)
    assert (
        discover_history_hybrid(
            ir_url=config.results_center_url,
            config=config,
            max_quarters=4,
            timeout_ms=1234,
        )
        == []
    )
    assert calls == [(config.results_center_url, 4, 1234)]


def test_header_row_detection_finds_offset_date_row(tmp_path: Path) -> None:
    path = tmp_path / "nu.xlsx"
    _write_nu_like_sheet(path)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    rows = [tuple(r) for r in wb["Managerial indicators"].iter_rows(values_only=True)]
    idx, col_map = _header_row(rows)
    assert idx == 8  # 0-based -> Excel row 9
    assert len(col_map) == 4


def test_parse_extracts_canonical_kpis_with_scale(tmp_path: Path) -> None:
    path = tmp_path / "nu.xlsx"
    _write_nu_like_sheet(path)
    parsed = parse_spreadsheet(path, get_config("NU"), max_quarters=8)

    arpac = parsed["Monthly ARPAC (USD)"]
    assert arpac[dt.datetime(2025, 12, 31)] == 15.0
    assert arpac[dt.datetime(2025, 3, 31)] == 11.6

    nim = parsed["Risk-adjusted NIM (NIM minus cost of risk)"]
    # Stored as decimals; scale=100 turns 0.105 into 10.5%.
    assert round(nim[dt.datetime(2025, 12, 31)], 2) == 10.5
    assert round(nim[dt.datetime(2025, 3, 31)], 2) == 9.28

    # NPL rows live on a 'NPLs' sheet absent from this fixture — skipped, not an error.
    assert "NPL 15-90d" not in parsed


def test_max_quarters_truncates_to_recent(tmp_path: Path) -> None:
    path = tmp_path / "nu.xlsx"
    _write_nu_like_sheet(path)
    parsed = parse_spreadsheet(path, get_config("NU"), max_quarters=2)
    arpac = parsed["Monthly ARPAC (USD)"]
    assert set(arpac) == {dt.datetime(2025, 9, 30), dt.datetime(2025, 12, 31)}


# ---------------------------------------------------------------------------
# Discovery: doc classification + platform dispatch (no live browser)
# ---------------------------------------------------------------------------


def test_classify_identifies_doc_types_by_filename() -> None:
    assert classify("Nu Holdings Historical Data 1Q26.xlsx") == "spreadsheet"
    assert classify("1T26 Results Presentation.pdf") == "deck"
    assert classify("1Q26 Earnings Press Release.pdf") == "press_release"
    assert classify("Transcript 1Q26.pdf") == "transcript"
    assert classify("Some Unrelated Governance Doc.pdf") is None


def test_discover_dispatch_rejects_unknown_platform() -> None:
    cfg = IrConfig(ticker="ZZZ", platform="nope", results_center_url="https://x")
    with pytest.raises(ValueError, match="discovery adapter"):
        discover_documents(cfg)


# ---------------------------------------------------------------------------
# Config persistence + LLM-generated parser config
# ---------------------------------------------------------------------------


def test_config_save_load_roundtrip(tmp_path: Path) -> None:
    cfg = IrConfig(
        ticker="ZZ",
        platform="mz",
        results_center_url="https://ir.example",
        spreadsheet_kpis=(
            SheetKpi("Monthly ARPAC (USD)", "Sheet1", "arpac", "usd", 1.0),
            SheetKpi("Risk-adjusted NIM", "Sheet1", "nim", "percent", 100.0),
        ),
    )
    path = save_config(cfg, tmp_path)
    assert path.exists()
    loaded = get_config("ZZ", tmp_path)
    assert loaded is not None
    assert loaded.platform == "mz"
    assert loaded.results_center_url == "https://ir.example"
    assert len(loaded.spreadsheet_kpis) == 2
    nim = loaded.spreadsheet_kpis[1]
    assert nim.kpi_name == "Risk-adjusted NIM" and nim.scale == 100.0


def test_build_ir_config_maps_rows_via_llm(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """The builder feeds the sheet structure + canonical KPIs to the LLM and
    assembles a persisted config from the mapping (LLM mocked)."""
    sheet = tmp_path / "hist.xlsx"
    _write_nu_like_sheet(sheet)
    holdings_dir = tmp_path / "micro_thesis" / "holdings"
    holdings_dir.mkdir(parents=True)
    (holdings_dir / "ZZ.json").write_text(
        json.dumps(
            {
                "tier_1_kpis": [{"name": "Monthly ARPAC (USD)"}],
                "chart_priorities": ["Risk-adjusted NIM (NIM minus cost of risk)"],
            }
        ),
        encoding="utf-8",
    )
    fake: dict[str, object] = {
        "Monthly ARPAC (USD)": {
            "sheet": "Managerial indicators",
            "row_label": "Average Revenue",
            "unit": "usd",
            "scale": 1,
        },
        "Risk-adjusted NIM (NIM minus cost of risk)": {
            "sheet": "Managerial indicators",
            "row_label": "Risk-adjusted NIM",
            "unit": "percent",
            "scale": 100,
        },
    }

    def fake_structured_call(*_args: object, **_kwargs: object) -> dict[str, object]:
        return fake

    monkeypatch.setattr("llm.structured.call_llm_structured", fake_structured_call)
    cfg = config_builder.build_ir_config(
        "ZZ", sheet, platform="mz", results_center_url="https://x", repo_root=tmp_path
    )
    by_name = {s.kpi_name: s for s in cfg.spreadsheet_kpis}
    assert "Monthly ARPAC (USD)" in by_name
    assert by_name["Risk-adjusted NIM (NIM minus cost of risk)"].scale == 100.0
    assert (tmp_path / "micro_thesis" / "ir_config" / "ZZ.json").exists()
