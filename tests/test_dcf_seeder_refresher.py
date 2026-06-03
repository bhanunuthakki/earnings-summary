"""Tests for the DCF workbook lifecycle: seeder (build from scratch),
refresher (preserve INPUTS, recompute everything else), and the contract
that the resulting workbook is parseable by `workbook_reader`.

The big property under test: the user can edit Forecast.INPUTS cells in
Excel between refreshes, and those edits propagate through to the
`dcf_runs` row that feeds the briefs. We exercise this end-to-end by
simulating a user edit (Y1 growth bumped from auto-derived to 25%) and
checking that the refreshed Valuation reflects the new assumption.
"""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import cast

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from dcf import forecast, refresher, seeder
from dcf.workbook_reader import read_valuation


# A miniature FMP-shaped fixture. 8 quarters so we can compute TTM + prior TTM
# (for Y1 growth derivation). The TTM quarters (first 4) carry the full set of
# line items the P&L -> FCF bridge needs (gross/R&D/SG&A, D&A/SBC, working
# capital).
#
# TTM sums (for assertion math):
#   revenue            = 4_100M
#   grossProfit        = 2_050M  -> gross margin = 50%
#   R&D                =   615M  -> 15% of revenue
#   SG&A               =   615M  -> 15% of revenue
#   incomeBeforeTax    =   950M ; incomeTaxExpense = 237M -> tax 24.95%
#   D&A                =   205M  -> 5% of revenue
#   SBC                =   328M  -> 8% of revenue
#   capitalExpenditure = -190M   -> capex/D&A = 190/205 = 0.93x
# Balance (latest): receivables 337M -> DSO 30; payables 225M -> DPO 40 (on
#   COGS 2_050M); deferred revenue 410M -> 10% of revenue.
def _income_q(
    rev: float, op: float, pretax: float, tax: float, **extra: object
) -> dict[str, object]:
    return {
        "revenue": rev,
        "costOfRevenue": rev * 0.5,
        "grossProfit": rev * 0.5,
        "researchAndDevelopmentExpenses": rev * 0.15,
        "sellingGeneralAndAdministrativeExpenses": rev * 0.15,
        "operatingIncome": op,
        "incomeBeforeTax": pretax,
        "incomeTaxExpense": tax,
        "weightedAverageShsOutDil": 100_000_000,
        **extra,
    }


_INCOME_RECORDS: list[dict[str, object]] = [
    {
        "fiscalYear": "2025",
        "period": "Q4",
        "date": "2025-12-31",
        **_income_q(1_100_000_000, 220_000_000, 280_000_000, 70_000_000, netIncome=165_000_000),
    },
    {
        "fiscalYear": 2025,
        "period": "Q3",
        "date": "2025-09-30",
        **_income_q(1_050_000_000, 200_000_000, 250_000_000, 62_000_000),
    },
    {
        "fiscalYear": 2025,
        "period": "Q2",
        "date": "2025-06-30",
        **_income_q(1_000_000_000, 180_000_000, 220_000_000, 55_000_000),
    },
    {
        "fiscalYear": 2025,
        "period": "Q1",
        "date": "2025-03-31",
        **_income_q(950_000_000, 160_000_000, 200_000_000, 50_000_000),
    },
    # Prior TTM — 2024 at ~10% lower run-rate. Q4 carries a share count so the
    # net-share-change driver derives (100M now vs 98M a year ago -> +2.04%).
    {
        "fiscalYear": 2024,
        "period": "Q4",
        "date": "2024-12-31",
        "revenue": 1_000_000_000,
        "weightedAverageShsOutDil": 98_000_000,
    },
    {"fiscalYear": 2024, "period": "Q3", "date": "2024-09-30", "revenue": 950_000_000},
    {"fiscalYear": 2024, "period": "Q2", "date": "2024-06-30", "revenue": 900_000_000},
    {"fiscalYear": 2024, "period": "Q1", "date": "2024-03-31", "revenue": 850_000_000},
]
_BALANCE_RECORDS: list[dict[str, object]] = [
    {
        "fiscalYear": 2025,
        "period": "Q4",
        "date": "2025-12-31",
        "totalAssets": 5_000_000_000,
        "totalStockholdersEquity": 3_000_000_000,
        "longTermDebt": 800_000_000,
        "netReceivables": 337_000_000,
        "accountPayables": 225_000_000,
        "deferredRevenue": 410_000_000,
    },
]


def _cashflow_q(
    capex: float, da: float, sbc: float, fcf: float, **extra: object
) -> dict[str, object]:
    return {
        "capitalExpenditure": capex,
        "depreciationAndAmortization": da,
        "stockBasedCompensation": sbc,
        "freeCashFlow": fcf,
        **extra,
    }


_CASHFLOW_RECORDS: list[dict[str, object]] = [
    {
        "fiscalYear": 2025,
        "period": "Q4",
        "date": "2025-12-31",
        **_cashflow_q(
            -50_000_000,
            55_000_000,
            88_000_000,
            200_000_000,
            operatingCashFlow=250_000_000,
            netIncome=165_000_000,
        ),
    },
    {
        "fiscalYear": 2025,
        "period": "Q3",
        "date": "2025-09-30",
        **_cashflow_q(-55_000_000, 52_500_000, 84_000_000, 180_000_000),
    },
    {
        "fiscalYear": 2025,
        "period": "Q2",
        "date": "2025-06-30",
        **_cashflow_q(-45_000_000, 50_000_000, 80_000_000, 170_000_000),
    },
    {
        "fiscalYear": 2025,
        "period": "Q1",
        "date": "2025-03-31",
        **_cashflow_q(-40_000_000, 47_500_000, 76_000_000, 150_000_000),
    },
]


def _set_assumption(ws: object, label: str, value: float, fy: int = 0) -> None:
    """Edit a per-year assumption-grid cell (col A label -> forecast column).

    Mirrors a user editing a yellow cell in Excel. `fy` is the 0-based forecast
    year (FY1 == 0). Sets every forecast column when `fy` is None-equivalent via
    -1, so a single edit can flatten a whole row.
    """
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)
    for r in range(1, 30):
        if ws.cell(row=r, column=1).value == label:
            if fy < 0:
                for col in range(3, 3 + 10):
                    if ws.cell(row=r, column=col).value is not None:
                        ws.cell(row=r, column=col, value=value)
            else:
                ws.cell(row=r, column=3 + fy, value=value)
            return
    raise AssertionError(f"assumption row {label!r} not found")


@pytest.fixture
def fmp_dir(tmp_path: Path) -> Path:
    """Drop the three statement JSONs into a temp dir for the test ticker."""
    d = tmp_path / "fmp"
    d.mkdir()
    (d / "TEST_income_statement_quarterly.json").write_text(json.dumps(_INCOME_RECORDS))
    (d / "TEST_balance_sheet_quarterly.json").write_text(json.dumps(_BALANCE_RECORDS))
    (d / "TEST_cash_flow_quarterly.json").write_text(json.dumps(_CASHFLOW_RECORDS))
    return d


def _facts_db_fy(
    tmp_path: Path,
    ticker: str,
    line_items_by_year: dict[str, dict[int, float]],
    *,
    tier: str = "s1_provisional",
) -> Path:
    """Build a portfolio.db with documents + annual (FY) financial_facts rows.

    Mirrors the shape the S-1 extractor writes: all facts point at one
    `s1_provisional`-tier document (the lowest precedence). `line_items_by_year`
    maps a canonical line_item to {fiscal_year: value}; each value lands on a
    Dec-31 FY period. Exercises the no-FMP fallback the loader reads from.
    """
    p = tmp_path / "portfolio.db"
    conn = sqlite3.connect(str(p))
    try:
        conn.executescript(
            """
            CREATE TABLE documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_quality_tier VARCHAR,
                fetched_at DATETIME
            );
            CREATE TABLE financial_facts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticker VARCHAR NOT NULL,
                period_end DATETIME NOT NULL,
                fiscal_period_type VARCHAR NOT NULL,
                line_item VARCHAR NOT NULL,
                value NUMERIC(24,6) NOT NULL,
                currency VARCHAR(3),
                unit VARCHAR NOT NULL,
                source_doc_id INTEGER NOT NULL,
                confidence FLOAT NOT NULL DEFAULT 1.0
            );
            """
        )
        conn.execute(
            "INSERT INTO documents(id, source_quality_tier, fetched_at) VALUES (1, ?, ?)",
            (tier, "2026-05-01 00:00:00"),
        )
        for line_item, by_year in line_items_by_year.items():
            for year, value in by_year.items():
                conn.execute(
                    "INSERT INTO financial_facts(ticker, period_end, fiscal_period_type, "
                    "line_item, value, unit, source_doc_id) VALUES (?,?,?,?,?,?,1)",
                    (ticker, f"{year}-12-31 00:00:00", "FY", line_item, float(value), "actual"),
                )
        conn.commit()
    finally:
        conn.close()
    return p


# Two fiscal years of S-1-shaped annuals (actual dollars; EPS per-share; shares
# as a raw count) — modeled on FRVO's S-1 figures.
_S1_FACTS: dict[str, dict[int, float]] = {
    "revenue": {2024: 199_000.0, 2025: 138_000.0},
    "operating_income": {2024: -41_838_000.0, 2025: -48_806_000.0},
    "net_income": {2024: -41_110_000.0, 2025: -57_788_000.0},
    "eps_diluted": {2024: -3.31, 2025: -5.66},
    "weighted_avg_shares_diluted": {2024: 12_438_000.0, 2025: 12_462_000.0},
    "total_assets": {2024: 531_299_000.0, 2025: 1_365_168_000.0},
    "long_term_debt": {2024: 39_019_000.0, 2025: 172_837_000.0},
    "capital_expenditure": {2024: -178_693_000.0, 2025: -465_659_000.0},
    "free_cash_flow": {2024: -233_441_000.0, 2025: -497_416_000.0},
}


# ---------------------------------------------------------------------------
# Forecast module — derivation + projection math
# ---------------------------------------------------------------------------


def _flat_inputs(**over: object) -> forecast.ForecastInputs:
    """A ForecastInputs with flat per-year driver series; override any field."""
    h_raw = over.pop("forecast_years", 5)
    h = int(h_raw) if isinstance(h_raw, (int, float)) else 5

    def num(key: str, default: float) -> float:
        v = over.pop(key, None)
        return float(v) if isinstance(v, (int, float)) else default

    def series(key: str, default: float) -> list[float]:
        v = over.pop(key, None)
        if isinstance(v, list):
            return [float(x) for x in cast("list[float]", v)]
        return [float(v) if isinstance(v, (int, float)) else default] * h

    return forecast.ForecastInputs(
        base_revenue_M=num("base_revenue_M", 1000.0),
        diluted_shares_M=num("diluted_shares_M", 100.0),
        terminal_multiple=num("terminal_multiple", 15.0),
        forecast_years=h,
        revenue_growth_pct=series("revenue_growth_pct", 0.10),
        gross_margin_pct=series("gross_margin_pct", 0.60),
        rnd_pct=series("rnd_pct", 0.10),
        sga_pct=series("sga_pct", 0.15),
        sbc_pct=series("sbc_pct", 0.05),
        da_pct=series("da_pct", 0.04),
        capex_to_da=series("capex_to_da", 1.0),
        dso_days=series("dso_days", 30.0),
        dpo_days=series("dpo_days", 30.0),
        deferred_rev_pct=series("deferred_rev_pct", 0.0),
        tax_rate_pct=series("tax_rate_pct", 0.25),
        net_share_change_pct=series("net_share_change_pct", 0.02),
    )


def test_derive_initial_inputs_from_history() -> None:
    inputs = forecast.derive_initial_inputs(
        _INCOME_RECORDS,
        _CASHFLOW_RECORDS,
        _BALANCE_RECORDS,
        terminal_growth_pct=0.025,
        forecast_years=5,
    )
    assert inputs.base_revenue_M == pytest.approx(4100.0, rel=1e-3)  # TTM revenue
    assert inputs.diluted_shares_M == pytest.approx(100.0)
    assert inputs.terminal_multiple == 15.0
    assert inputs.forecast_years == 5
    # Y1 growth = 4100/3700 - 1 = 10.81%, decaying to terminal 2.5% by FY5.
    assert inputs.revenue_growth_pct[0] == pytest.approx(0.1081, abs=0.001)
    assert inputs.revenue_growth_pct[-1] == pytest.approx(0.025, abs=1e-6)
    # Ratios default to their TTM value, held flat across the horizon.
    assert inputs.gross_margin_pct == pytest.approx([0.50] * 5)  # 2050/4100
    assert inputs.rnd_pct == pytest.approx([0.15] * 5)  # 615/4100
    assert inputs.sga_pct == pytest.approx([0.15] * 5)  # 615/4100
    assert inputs.sbc_pct[0] == pytest.approx(0.08, abs=0.002)  # 328/4100
    assert inputs.da_pct[0] == pytest.approx(0.05, abs=0.002)  # 205/4100
    assert inputs.capex_to_da[0] == pytest.approx(0.93, abs=0.02)  # 190/205
    assert inputs.dso_days[0] == pytest.approx(30.0, abs=0.6)
    assert inputs.dpo_days[0] == pytest.approx(40.0, abs=0.6)
    assert inputs.deferred_rev_pct[0] == pytest.approx(0.10, abs=0.005)
    assert inputs.tax_rate_pct[0] == pytest.approx(0.2495, abs=0.001)  # 237/950
    assert inputs.net_share_change_pct[0] == pytest.approx(0.0204, abs=0.002)  # 100/98 - 1


def test_compute_projections_bridge() -> None:
    inputs = _flat_inputs(
        base_revenue_M=1000.0,
        revenue_growth_pct=0.10,
        gross_margin_pct=0.60,
        rnd_pct=0.10,
        sga_pct=0.15,
        sbc_pct=0.05,
        da_pct=0.04,
        capex_to_da=1.0,
        tax_rate_pct=0.25,
        net_share_change_pct=0.02,
    )
    proj = forecast.compute_projections(inputs, base_year=2026)
    assert proj.years == [2027, 2028, 2029, 2030, 2031]
    # FY1: revenue 1000*1.10 = 1100; gross 660; op income 660-110-165 = 385.
    assert proj.revenue_M[0] == pytest.approx(1100.0)
    assert proj.gross_profit_M[0] == pytest.approx(660.0)
    assert proj.operating_income_M[0] == pytest.approx(385.0)
    assert proj.operating_margin_pct[0] == pytest.approx(0.35, abs=1e-6)
    # SBC is added back in CFO but re-charged to reach Valuation FCF:
    assert proj.sbc_M[0] == pytest.approx(55.0)  # 5% of 1100
    assert proj.valuation_fcf_M[0] == pytest.approx(proj.fcff_M[0] - proj.sbc_M[0])
    assert proj.cfo_M[0] == pytest.approx(
        proj.nopat_M[0] + proj.da_M[0] + proj.sbc_M[0] - proj.delta_nwc_M[0]
    )
    assert proj.fcff_M[0] == pytest.approx(proj.cfo_M[0] - proj.capex_M[0])
    # Shares dilute at +2%/yr: FY1 = 102, compounding.
    assert proj.shares_M[0] == pytest.approx(102.0)
    assert proj.shares_M[-1] == pytest.approx(100.0 * 1.02**5, rel=1e-6)


def test_heavy_capex_year_depresses_fcf_then_recovers() -> None:
    """A heavy-capex early year (AMZN-style AI buildout) should depress FCF, and
    a per-year Capex/D&A that normalises down should let it recover — the
    per-year grid gives the user that control."""
    inputs = _flat_inputs(
        base_revenue_M=600_000.0,
        revenue_growth_pct=0.10,
        gross_margin_pct=0.45,
        rnd_pct=0.05,
        sga_pct=0.10,
        da_pct=0.10,
        sbc_pct=0.0,
        capex_to_da=[3.0, 2.5, 2.0, 1.5, 1.0],  # elevated then normalising
        tax_rate_pct=0.25,
        net_share_change_pct=0.0,
    )
    proj = forecast.compute_projections(inputs, base_year=2026)
    # Heavy early capex (3x D&A) suppresses FY1 FCF well below the normalised FY5.
    assert proj.valuation_fcf_M[0] < proj.valuation_fcf_M[-1]
    assert proj.capex_M[0] > proj.capex_M[-1]


# ---------------------------------------------------------------------------
# Seeder
# ---------------------------------------------------------------------------


def test_seeder_builds_three_sheets_from_scratch(tmp_path: Path, fmp_dir: Path) -> None:
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    assert wb.sheetnames == [
        seeder.HISTORICALS_SHEET,
        seeder.FORECAST_SHEET,
        seeder.VALUATION_SHEET,
    ]


def test_seeder_derives_forecast_inputs_from_history(tmp_path: Path, fmp_dir: Path) -> None:
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    inputs = forecast.read_inputs_from_sheet(wb[seeder.FORECAST_SHEET])
    # Same values as the standalone derive_initial_inputs test — confirms the
    # seeder writes the grid and read_inputs_from_sheet round-trips it.
    assert inputs.base_revenue_M == pytest.approx(4100.0, rel=1e-3)
    assert inputs.revenue_growth_pct[0] == pytest.approx(0.1081, abs=0.005)
    assert inputs.gross_margin_pct[0] == pytest.approx(0.50, abs=0.01)
    assert inputs.rnd_pct[0] == pytest.approx(0.15, abs=0.01)
    assert inputs.sbc_pct[0] == pytest.approx(0.08, abs=0.005)
    assert inputs.da_pct[0] == pytest.approx(0.05, abs=0.005)
    assert inputs.tax_rate_pct[0] == pytest.approx(0.2495, abs=0.001)


def test_seeded_workbook_parses_via_workbook_reader(tmp_path: Path, fmp_dir: Path) -> None:
    """The whole point: workbook_reader must read what the seeder writes,
    so the PV calc + dcf_runs persist + briefs chain stays unbroken."""
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    snap = read_valuation(out, valuation_year=2026)
    # Base year (2026) is the actual; 2027-2031 are forecasts.
    assert snap.latest_actual_year == 2026
    assert snap.forecast_years == [2027, 2028, 2029, 2030, 2031]
    # All forecast years carry numeric FCFs.
    assert all(isinstance(snap.fcf_by_year[y], float) for y in snap.forecast_years)
    # Diluted shares baseline is the user input.
    assert snap.shares_by_year[2026] == pytest.approx(100.0)


def test_seeder_refuses_to_overwrite_without_force(tmp_path: Path, fmp_dir: Path) -> None:
    out = tmp_path / "TEST.xlsx"
    out.write_text("existing")
    with pytest.raises(seeder.SeederError, match="refusing to overwrite"):
        seeder.seed_workbook("TEST", fmp_dir, out)


def test_seeder_force_overwrites(tmp_path: Path, fmp_dir: Path) -> None:
    out = tmp_path / "TEST.xlsx"
    out.write_text("existing")
    seeder.seed_workbook("TEST", fmp_dir, out, force=True)
    assert out.read_bytes()[:2] == b"PK"


def test_seeder_missing_fmp_raises(tmp_path: Path) -> None:
    empty_fmp = tmp_path / "fmp"
    empty_fmp.mkdir()
    with pytest.raises(seeder.SeederError, match="no income statement records"):
        seeder.seed_workbook("NOPE", empty_fmp, tmp_path / "NOPE.xlsx")


# ---------------------------------------------------------------------------
# Seeder — financial_facts fallback (recently-IPO'd / S-1-anchored tickers)
# ---------------------------------------------------------------------------


def test_seeder_falls_back_to_financial_facts_when_fmp_absent(
    tmp_path: Path,
) -> None:
    """No FMP quarterly files → seed Historicals from annual financial_facts.

    One column per fiscal year; values scaled to $M (EPS unscaled, shares in
    millions). This is the S-1-anchored path the whole change exists for."""
    db = _facts_db_fy(tmp_path, "S1CO", _S1_FACTS)
    empty_fmp = tmp_path / "fmp"
    empty_fmp.mkdir()
    out = tmp_path / "S1CO.xlsx"

    seeder.seed_workbook("S1CO", empty_fmp, out, base_year=2026, db_path=db)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    assert wb.sheetnames == [
        seeder.HISTORICALS_SHEET,
        seeder.FORECAST_SHEET,
        seeder.VALUATION_SHEET,
    ]
    hist = wb[seeder.HISTORICALS_SHEET]
    headers = [hist.cell(row=1, column=c).value for c in range(2, hist.max_column + 1)]
    assert headers == ["FY2024", "FY2025"]  # one column per fiscal year, oldest first

    rev_row = _find_label_row(hist, "Revenue")
    assert hist.cell(row=rev_row, column=2).value == pytest.approx(0.199)  # 199k → $M
    assert hist.cell(row=rev_row, column=3).value == pytest.approx(0.138)
    fcf_row = _find_label_row(hist, "Free Cash Flow")
    assert hist.cell(row=fcf_row, column=3).value == pytest.approx(-497.416)
    eps_row = _find_label_row(hist, "Diluted EPS")
    assert hist.cell(row=eps_row, column=3).value == pytest.approx(-5.66)  # per-share, unscaled
    shares_row = _find_label_row(hist, "Diluted Shares (M)")
    assert hist.cell(row=shares_row, column=3).value == pytest.approx(12.462)  # count → M


def test_seeder_fallback_derives_forecast_inputs_from_annuals(
    tmp_path: Path,
) -> None:
    """Forecast INPUTS come off the annuals: base revenue + shares from the
    latest FY, Y1 growth from the YoY of the two fiscal years (via the TTM-window
    padding)."""
    db = _facts_db_fy(tmp_path, "S1CO", _S1_FACTS)
    empty_fmp = tmp_path / "fmp"
    empty_fmp.mkdir()
    out = tmp_path / "S1CO.xlsx"

    seeder.seed_workbook("S1CO", empty_fmp, out, base_year=2026, db_path=db)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    inputs = forecast.read_inputs_from_sheet(wb[seeder.FORECAST_SHEET])
    assert inputs.base_revenue_M == pytest.approx(0.14, abs=0.005)  # latest FY 138k
    assert inputs.diluted_shares_M == pytest.approx(12.46, abs=0.01)  # latest FY
    # Y1 growth = YoY of the two fiscal years: 138k / 199k - 1.
    assert inputs.revenue_growth_pct[0] == pytest.approx(138_000.0 / 199_000.0 - 1, abs=0.001)
    # The S-1 facts carry no gross/R&D/SBC lines, so those default to the
    # steady-state shapes rather than crashing.
    assert inputs.gross_margin_pct[0] == pytest.approx(0.50)
    assert inputs.tax_rate_pct[0] == pytest.approx(0.25)


def test_seeder_prefers_fmp_when_both_fmp_and_facts_present(tmp_path: Path, fmp_dir: Path) -> None:
    """FMP quarterly files are the primary source — facts are only a fallback.
    When both exist, the seeder uses FMP (quarterly columns, not FY)."""
    db = _facts_db_fy(tmp_path, "TEST", {"revenue": {2024: 1.0, 2025: 2.0}})
    out = tmp_path / "TEST.xlsx"

    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026, db_path=db)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    hist = wb[seeder.HISTORICALS_SHEET]
    headers = [hist.cell(row=1, column=c).value for c in range(2, hist.max_column + 1)]
    assert "Q4 2025" in headers  # FMP quarterly path
    assert "FY2025" not in headers


def test_seeder_raises_when_db_has_no_rows_for_ticker(tmp_path: Path) -> None:
    """db_path given but no FY facts for THIS ticker → still raises (no silent
    empty workbook)."""
    db = _facts_db_fy(tmp_path, "OTHER", {"revenue": {2025: 1.0}})
    empty_fmp = tmp_path / "fmp"
    empty_fmp.mkdir()
    with pytest.raises(seeder.SeederError, match="no income statement records"):
        seeder.seed_workbook("MISSING", empty_fmp, tmp_path / "MISSING.xlsx", db_path=db)


def test_refresh_uses_financial_facts_fallback(tmp_path: Path) -> None:
    """The refresh path also falls back to facts: re-running on an existing
    S-1-seeded workbook (no FMP) re-writes Historicals and preserves INPUTS."""
    db = _facts_db_fy(tmp_path, "S1CO", _S1_FACTS)
    empty_fmp = tmp_path / "fmp"
    empty_fmp.mkdir()
    out = tmp_path / "S1CO.xlsx"
    seeder.seed_workbook("S1CO", empty_fmp, out, base_year=2026, db_path=db)

    # Simulate a user edit (FY1 revenue growth), then refresh via the facts fallback.
    wb = openpyxl.load_workbook(str(out))
    _set_assumption(wb[seeder.FORECAST_SHEET], "Revenue Growth %", 0.5, fy=0)
    wb.save(str(out))

    result = refresher.refresh_historicals(
        out, empty_fmp, ticker="S1CO", base_year=2026, db_path=db
    )
    assert result.historicals_cells_written > 0
    assert result.forecast_inputs.revenue_growth_pct[0] == pytest.approx(0.5)  # edit preserved

    wb2 = openpyxl.load_workbook(str(out), data_only=True)
    hist = wb2[seeder.HISTORICALS_SHEET]
    headers = [hist.cell(row=1, column=c).value for c in range(2, hist.max_column + 1)]
    assert headers == ["FY2024", "FY2025"]


# ---------------------------------------------------------------------------
# Refresher — INPUTS preservation contract
# ---------------------------------------------------------------------------


def test_refresh_preserves_user_edited_inputs(tmp_path: Path, fmp_dir: Path) -> None:
    """The core user-edit contract: open the workbook, change Forecast INPUT
    cells, save. Re-running the refresher must read those edited values back and
    use them (not re-derive from FMP). We edit FY1 revenue growth, gross margin
    (all years), and the tax rate so the contract covers scalars + the grid."""
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    wb = openpyxl.load_workbook(str(out))
    fws = wb[seeder.FORECAST_SHEET]
    _set_assumption(fws, "Revenue Growth %", 0.25, fy=0)  # FY1 growth -> 25%
    _set_assumption(fws, "Gross Margin %", 0.70, fy=-1)  # all years -> 70%
    _set_assumption(fws, "Tax Rate %", 0.20, fy=-1)
    wb.save(str(out))

    result = refresher.refresh_historicals(out, fmp_dir, ticker="TEST", base_year=2026)

    # The edited values MUST round-trip through the refresher.
    assert result.forecast_inputs.revenue_growth_pct[0] == pytest.approx(0.25)
    assert result.forecast_inputs.gross_margin_pct[0] == pytest.approx(0.70)
    assert result.forecast_inputs.tax_rate_pct[0] == pytest.approx(0.20)
    # And the projection MUST reflect the growth edit: FY1 revenue = 4100 * 1.25 = 5125.
    assert result.projections.revenue_M[0] == pytest.approx(5125.0, abs=1.0)
    # FY1 gross profit reflects the edited 70% gross margin: 5125 * 0.70 = 3587.5.
    assert result.projections.gross_profit_M[0] == pytest.approx(3587.5, abs=2.0)


def test_refresh_updates_valuation_fcf(tmp_path: Path, fmp_dir: Path) -> None:
    """After a user-input edit, workbook_reader must see the new FCFs in
    Valuation — that's what feeds the PV calc and dcf_runs."""
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    fcf_before = read_valuation(out, valuation_year=2026).fcf_by_year[2027]

    # Bump gross margin from the auto-derived 50% to 80% — gross profit (and so
    # operating income and Valuation FCF) jumps; the 2027 FCF must rise.
    wb = openpyxl.load_workbook(str(out))
    _set_assumption(wb[seeder.FORECAST_SHEET], "Gross Margin %", 0.80, fy=-1)
    wb.save(str(out))

    refresher.refresh_historicals(out, fmp_dir, ticker="TEST", base_year=2026)
    fcf_after = read_valuation(out, valuation_year=2026).fcf_by_year[2027]

    assert fcf_after > fcf_before * 1.2


def test_refresh_recomputes_historicals_from_latest_fmp(tmp_path: Path, fmp_dir: Path) -> None:
    """Historicals reflects current FMP, even if FMP changed between seed
    and refresh. Forecast INPUTS are NOT re-derived (preserve user edits)."""
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    # Mutate the FMP file's revenue for the newest quarter.
    income_path = fmp_dir / "TEST_income_statement_quarterly.json"
    income = json.loads(income_path.read_text())
    income[0]["revenue"] = 9_999_000_000
    income_path.write_text(json.dumps(income))

    refresher.refresh_historicals(out, fmp_dir, ticker="TEST", base_year=2026)

    wb = openpyxl.load_workbook(str(out), data_only=True)
    hist = wb[seeder.HISTORICALS_SHEET]
    revenue_row = _find_label_row(hist, "Revenue")
    # Find the Q4 2025 column — it's the rightmost quarter in our test data.
    q4_col = None
    for c in range(2, hist.max_column + 1):
        if hist.cell(row=1, column=c).value == "Q4 2025":
            q4_col = c
            break
    assert q4_col is not None
    assert hist.cell(row=revenue_row, column=q4_col).value == pytest.approx(9999.0)

    # INPUTS should NOT have changed — Base Revenue still reflects seed-time
    # TTM, not the post-refresh FMP value.
    inputs = forecast.read_inputs_from_sheet(wb[seeder.FORECAST_SHEET])
    assert inputs.base_revenue_M == pytest.approx(4100.0, rel=1e-3)


def test_refresh_is_idempotent_when_inputs_unchanged(tmp_path: Path, fmp_dir: Path) -> None:
    out = tmp_path / "TEST.xlsx"
    seeder.seed_workbook("TEST", fmp_dir, out, base_year=2026)

    refresher.refresh_historicals(out, fmp_dir, ticker="TEST", base_year=2026)
    inputs_first = _read_inputs(out)
    fcf_first = read_valuation(out, valuation_year=2026).fcf_by_year[2027]

    refresher.refresh_historicals(out, fmp_dir, ticker="TEST", base_year=2026)
    inputs_second = _read_inputs(out)
    fcf_second = read_valuation(out, valuation_year=2026).fcf_by_year[2027]

    assert inputs_first == inputs_second
    assert fcf_first == fcf_second


def test_refresher_rejects_workbook_missing_forecast_sheet(tmp_path: Path, fmp_dir: Path) -> None:
    dest = tmp_path / "TEST.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet.title = "Valuation"
    wb.save(str(dest))
    with pytest.raises(refresher.RefresherError, match="missing required sheets"):
        refresher.refresh_historicals(dest, fmp_dir, ticker="TEST")


def test_refresher_missing_workbook_raises(tmp_path: Path) -> None:
    with pytest.raises(refresher.RefresherError, match="workbook not found"):
        refresher.refresh_historicals(tmp_path / "missing.xlsx", tmp_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _find_label_row(ws: Worksheet, label: str) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    raise AssertionError(f"label {label!r} not in column A")


def _read_inputs(path: Path) -> forecast.ForecastInputs:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    inputs = forecast.read_inputs_from_sheet(wb[seeder.FORECAST_SHEET])
    wb.close()
    return inputs
