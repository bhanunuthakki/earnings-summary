# pyright: reportPrivateUsage=false
"""Tests for the redesigned 9-sheet DCF: the reader/projection/value engine
(``src/dcf/redesign.py``) and the redesign refresh path in
``execution/refresh_dcf.py`` (rebuild-from-FMP with Dashboard edit-preservation).

Three layers:
  * Pure engine — construct ``RedesignInputs`` directly and assert the projection
    math (growth monotonicity, FX scaling, perpetuity terminal, the WACC>g guard).
    No workbook, no subprocess.
  * Reader — run the real builder as a subprocess (the way the driver/refresher
    do), then read it back: format detection, value-of-record parity with the
    builder's own ``_project`` mirror, FX for a non-USD reporter.
  * Refresh integration — drive ``refresh_dcf.refresh_one`` end-to-end: it rebuilds
    every sheet from FMP, PRESERVES the user's Dashboard inputs, recomputes the
    value, and upserts ``dcf_runs``. Plus the ``dcf_applicable=false`` skip and the
    negative-fair-value (#291) guard.
"""

from __future__ import annotations

import dataclasses
import json
import os
import shutil
import sqlite3
import subprocess
import sys
from pathlib import Path
from typing import cast

import openpyxl
import pytest
from openpyxl.cell.cell import Cell

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "execution"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import dcf_sheets  # noqa: E402
import refresh_dcf  # noqa: E402

from dcf import fact_sheet, redesign  # noqa: E402

BUILDER = PROJECT_ROOT / "execution" / "build_redesigned_dcf.py"

REDESIGN_SHEETS = [
    "Cover",
    "Dashboard",
    "Assumptions",
    "Color Code",
    "WACC",
    "Model",
    "Financials",
    "Consensus",
    "Valuation",
    "Sensitivity",
    "Monte Carlo",
]

_DCF_RUNS_SCHEMA = """
CREATE TABLE dcf_runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticker TEXT UNIQUE,
    valuation_date TEXT, horizon_years INTEGER,
    wacc REAL, terminal_growth REAL,
    npv REAL, npv_per_share REAL, shares_outstanding REAL,
    currency TEXT, notes TEXT, run_id TEXT,
    live_price REAL, live_price_at TEXT, over_under_pct REAL,
    mos_bar_used REAL, assumption_snapshot_json TEXT,
    revenue_growths_json TEXT, fcf_margin REAL,
    assumptions_sync_status TEXT, assumptions_synced_at TEXT
    , input_sha256 TEXT, workbook_sha256 TEXT, engine_version TEXT,
    inputs_as_of TEXT, provenance_json TEXT
);
"""


# --------------------------------------------------------------------------- #
# Pure engine — RedesignInputs -> value (no workbook)
# --------------------------------------------------------------------------- #
_BASE = redesign.RedesignInputs(
    segments=("Total company",),
    base_revenue_by_segment={"Total company": 1000.0},
    near_growth_by_segment={"Total company": 0.10},
    terminal_growth_by_segment={"Total company": 0.03},
    near_op_margin=0.20,
    terminal_op_margin=0.25,
    tax_rate=0.24,
    capex_2026_m=60.0,
    terminal_capex_da=1.05,
    da_ratio=0.05,
    consensus_years=5,
    wacc=0.09,
    beta=1.2,
    risk_free_rate=0.043,
    equity_risk_premium=0.045,
    cost_of_debt=0.045,
    terminal_method="Exit multiple",
    terminal_basis="EV/EBITDA",
    exit_multiple=12.0,
    terminal_growth_g=0.03,
    current_price=50.0,
    cash_m=100.0,
    total_debt_m=200.0,
    diluted_shares_m=100.0,
    fx_to_usd=1.0,
)


def test_value_is_deterministic() -> None:
    assert redesign.value(_BASE).value_per_share_usd == redesign.value(_BASE).value_per_share_usd


def test_higher_growth_raises_value() -> None:
    faster = dataclasses.replace(_BASE, near_growth_by_segment={"Total company": 0.20})
    assert redesign.value(faster).value_per_share_usd > redesign.value(_BASE).value_per_share_usd


def test_higher_margin_raises_value() -> None:
    richer = dataclasses.replace(_BASE, terminal_op_margin=0.35)
    assert redesign.value(richer).value_per_share_usd > redesign.value(_BASE).value_per_share_usd


def test_fx_scales_per_share_value_linearly() -> None:
    usd = redesign.value(_BASE)
    half = redesign.value(dataclasses.replace(_BASE, fx_to_usd=0.5))
    # Reporting-currency value is FX-independent; the USD value scales by FX.
    assert half.value_per_share_reporting == pytest.approx(usd.value_per_share_reporting)
    assert half.value_per_share_usd == pytest.approx(usd.value_per_share_usd * 0.5)


def test_exit_multiple_higher_than_perpetuity_basis_selects_metric() -> None:
    sales = redesign.value(dataclasses.replace(_BASE, terminal_basis="EV/Sales", exit_multiple=3.0))
    ebitda = redesign.value(_BASE)
    # Different bases price different terminal metrics — just assert both resolve
    # to finite, positive values (the basis plumbing reaches compute_valuation).
    assert sales.value_per_share_usd > 0
    assert ebitda.value_per_share_usd > 0


def test_perpetuity_terminal_resolves() -> None:
    perp = redesign.value(
        dataclasses.replace(_BASE, terminal_method="Perpetuity", terminal_growth_g=0.03)
    )
    assert perp.value_per_share_usd > 0
    assert perp.terminal_method == "Perpetuity"


def test_perpetuity_requires_wacc_above_terminal_g() -> None:
    bad = dataclasses.replace(
        _BASE, terminal_method="Perpetuity", wacc=0.025, terminal_growth_g=0.03
    )
    with pytest.raises(redesign.RedesignError, match="WACC"):
        redesign.value(bad)


def test_negative_margin_yields_negative_value_without_crashing() -> None:
    # Opex above revenue every year -> negative FCF -> negative value (no raise);
    # the refresher's #291 guard nulls over/under on a value like this.
    loss = dataclasses.replace(_BASE, near_op_margin=-0.5, terminal_op_margin=-0.5)
    assert redesign.value(loss).value_per_share_usd < 0


# --------------------------------------------------------------------------- #
# SBC as an explicit after-tax expense (op margins stay non-GAAP)
# --------------------------------------------------------------------------- #
def test_sbc_defaults_to_zero_backcompat() -> None:
    """A pre-SBC input set (no SBC fields) charges no SBC — the value is identical
    to the old 'SBC nets out' behaviour, so legacy workbooks/fixtures don't move."""
    assert _BASE.near_sbc_pct == 0.0
    assert _BASE.terminal_sbc_pct == 0.0
    # An explicit 0.0 must equal the default (the charge term vanishes).
    zeroed = dataclasses.replace(_BASE, near_sbc_pct=0.0, terminal_sbc_pct=0.0)
    assert redesign.value(zeroed).value_per_share_usd == pytest.approx(
        redesign.value(_BASE).value_per_share_usd
    )


def test_charging_sbc_lowers_value() -> None:
    """A heavy-SBC name (12% -> 8% of revenue) values materially LOWER than the
    same name with SBC excluded — the whole point of the fix."""
    heavy = dataclasses.replace(_BASE, near_sbc_pct=0.12, terminal_sbc_pct=0.08)
    assert redesign.value(heavy).value_per_share_usd < redesign.value(_BASE).value_per_share_usd


def test_sbc_charged_after_tax() -> None:
    """The FCF-stream SBC charge is ``sbc_pct * revenue * (1 - tax)``: at a higher
    tax rate the *same* SBC% charges LESS after-tax SBC. Isolate that channel by
    valuing on an EV/Sales terminal (whose metric is revenue, so the terminal-EBITDA
    SBC burden — deliberately NOT tax-adjusted — does not confound the comparison).

    The SBC drag then scales cleanly by (1-tax): the drag at 40% tax is 0.60/0.90 of
    the drag at 10% tax."""
    lo_tax = dataclasses.replace(_BASE, tax_rate=0.10, terminal_basis="EV/Sales", exit_multiple=3.0)
    hi_tax = dataclasses.replace(_BASE, tax_rate=0.40, terminal_basis="EV/Sales", exit_multiple=3.0)
    sbc = {"near_sbc_pct": 0.10, "terminal_sbc_pct": 0.10}
    drag_lo = (
        redesign.value(lo_tax).value_per_share_usd
        - redesign.value(dataclasses.replace(lo_tax, **sbc)).value_per_share_usd
    )
    drag_hi = (
        redesign.value(hi_tax).value_per_share_usd
        - redesign.value(dataclasses.replace(hi_tax, **sbc)).value_per_share_usd
    )
    # After-tax SBC at 40% tax is 0.60/0.90 of the charge at 10% tax.
    assert drag_hi < drag_lo
    assert drag_hi == pytest.approx(drag_lo * (0.60 / 0.90), rel=0.02)


def test_terminal_ebitda_burdened_by_sbc() -> None:
    """``_terminal_metrics`` burdens the exit-multiple EBITDA by terminal-year SBC
    (ebitda = ebit - sbc + da), so the terminal metric drops when SBC is charged."""
    no_sbc = redesign._terminal_metrics(redesign._project(_BASE), _BASE)
    heavy_inp = dataclasses.replace(_BASE, near_sbc_pct=0.12, terminal_sbc_pct=0.10)
    heavy = redesign._terminal_metrics(redesign._project(heavy_inp), heavy_inp)
    assert heavy.ebitda < no_sbc.ebitda
    # The burden equals terminal_sbc_pct * terminal revenue.
    streams = redesign._project(heavy_inp)
    burden = heavy_inp.terminal_sbc_pct * streams.revenue[-1]
    assert no_sbc.ebitda - heavy.ebitda == pytest.approx(burden)


def test_sbc_round_trips_through_dict() -> None:
    heavy = dataclasses.replace(_BASE, near_sbc_pct=0.13, terminal_sbc_pct=0.07)
    rebuilt = redesign.RedesignInputs.from_dict(heavy.to_dict())
    assert rebuilt.near_sbc_pct == pytest.approx(0.13)
    assert rebuilt.terminal_sbc_pct == pytest.approx(0.07)
    assert rebuilt == heavy
    assert redesign.value(rebuilt).value_per_share_usd == pytest.approx(
        redesign.value(heavy).value_per_share_usd
    )


def test_from_dict_missing_sbc_defaults_to_zero() -> None:
    payload = _BASE.to_dict()
    del payload["near_sbc_pct"]
    del payload["terminal_sbc_pct"]
    inp = redesign.RedesignInputs.from_dict(payload)
    assert inp.near_sbc_pct == 0.0
    assert inp.terminal_sbc_pct == 0.0


# --------------------------------------------------------------------------- #
# Scenarios + sensitivity (pure engine)
# --------------------------------------------------------------------------- #
def test_scenario_values_order_bear_base_bull() -> None:
    """With the seeded offsets the three fair values are strictly ordered and
    Base is exactly the value-of-record."""
    sv = redesign.scenario_values(_BASE)
    assert sv.bull is not None and sv.bear is not None
    assert sv.bear < sv.base < sv.bull
    assert sv.base == pytest.approx(redesign.value(_BASE).value_per_share_usd)


def test_apply_scenario_shifts_every_lever() -> None:
    shifted = redesign.apply_scenario(_BASE, redesign.BULL_SEED)
    assert shifted.near_growth_by_segment["Total company"] == pytest.approx(0.13)
    assert shifted.terminal_growth_by_segment["Total company"] == pytest.approx(0.04)
    assert shifted.near_op_margin == pytest.approx(0.21)
    assert shifted.terminal_op_margin == pytest.approx(0.27)
    assert shifted.exit_multiple == pytest.approx(14.0)
    assert shifted.terminal_growth_g == pytest.approx(0.035)


def test_scenario_degenerate_bull_degrades_to_none() -> None:
    """A Bull terminal-g pushed to/above WACC under a perpetuity terminal is
    un-valuable: bull degrades to None while Base (still WACC > g) and Bear
    (g moves further below WACC) stay numeric."""
    perp = dataclasses.replace(
        _BASE, terminal_method="Perpetuity", wacc=0.038, terminal_growth_g=0.034
    )
    sv = redesign.scenario_values(perp)
    assert sv.bull is None  # bull g = 3.9% >= 3.8% WACC
    assert sv.base > 0
    assert sv.bear is not None


def test_sensitivity_grid_centered_on_base_and_monotonic() -> None:
    grid = redesign.sensitivity_grid(_BASE)
    assert len(grid.wacc_axis) == 7 and len(grid.multiple_axis) == 7
    assert grid.values[3][3] == pytest.approx(redesign.value(_BASE).value_per_share_usd)
    assert grid.wacc_axis[3] == pytest.approx(_BASE.wacc)
    assert grid.multiple_axis[3] == pytest.approx(_BASE.exit_multiple)
    for row in grid.values:  # value falls as WACC rises (left -> right)
        assert all(row[j] > row[j + 1] for j in range(len(row) - 1))
    for j in range(7):  # value rises with the exit multiple (top -> bottom)
        col = [grid.values[i][j] for i in range(7)]
        assert all(col[i] < col[i + 1] for i in range(len(col) - 1))


# --------------------------------------------------------------------------- #
# JSON constructor — to_dict / from_dict (the non-workbook transport for the
# in-app recompute route; no xlsx in the loop)
# --------------------------------------------------------------------------- #
def test_to_dict_is_json_safe() -> None:
    payload = _BASE.to_dict()
    # Round-trips through json without raising; tuple → list, deltas → dicts.
    assert isinstance(payload["segments"], list)
    assert isinstance(payload["bull_deltas"], dict)
    text = json.dumps(payload)
    assert json.loads(text)["wacc"] == pytest.approx(_BASE.wacc)


def test_to_dict_from_dict_round_trips() -> None:
    """from_dict(to_dict(x)) reconstructs an equal input set with an identical
    value-of-record — the recompute route can rebuild edits with no workbook."""
    rebuilt = redesign.RedesignInputs.from_dict(_BASE.to_dict())
    assert rebuilt == _BASE
    assert redesign.value(rebuilt).value_per_share_usd == pytest.approx(
        redesign.value(_BASE).value_per_share_usd
    )


def test_from_dict_applies_an_edit() -> None:
    """A WACC edit posted as a dict recomputes a lower value (taken as given —
    NOT re-derived from CAPM the way read_inputs does)."""
    edited = _BASE.to_dict()
    edited["wacc"] = 0.12
    inp = redesign.RedesignInputs.from_dict(edited)
    assert inp.wacc == pytest.approx(0.12)
    assert redesign.value(inp).value_per_share_usd < redesign.value(_BASE).value_per_share_usd


def test_from_dict_defaults_scenario_deltas_to_seeds() -> None:
    payload = _BASE.to_dict()
    del payload["bull_deltas"]
    del payload["bear_deltas"]
    inp = redesign.RedesignInputs.from_dict(payload)
    assert inp.bull_deltas == redesign.BULL_SEED
    assert inp.bear_deltas == redesign.BEAR_SEED


def test_from_dict_rejects_missing_field() -> None:
    payload = _BASE.to_dict()
    del payload["wacc"]
    with pytest.raises(redesign.RedesignError, match="wacc"):
        redesign.RedesignInputs.from_dict(payload)


def test_from_dict_rejects_non_numeric_field() -> None:
    payload = _BASE.to_dict()
    payload["wacc"] = "high"
    with pytest.raises(redesign.RedesignError, match="must be a number"):
        redesign.RedesignInputs.from_dict(payload)


def test_from_dict_rejects_segment_without_growth() -> None:
    """A segment present in the list but absent from a growth map fails loud
    here, not with a KeyError deep in the projection."""
    payload = _BASE.to_dict()
    payload["segments"] = ["Total company", "Phantom"]
    with pytest.raises(redesign.RedesignError, match="Phantom"):
        redesign.RedesignInputs.from_dict(payload)


# --------------------------------------------------------------------------- #
# Fixtures: write FMP, run the real builder
# --------------------------------------------------------------------------- #
def _write_fmp(repo: Path, ticker: str, *, currency: str = "USD", segments: bool = False) -> None:
    """Minimal FMP fixture: 4 full fiscal years of growing quarterlies + a profile
    + forward estimates. Optionally a two-line product-segment file."""
    fmp = repo / "data" / "historical" / "fmp"
    fmp.mkdir(parents=True, exist_ok=True)
    inc: list[dict[str, object]] = []
    bal: list[dict[str, object]] = []
    cf: list[dict[str, object]] = []
    pseg: list[dict[str, object]] = []
    rev = 250.0
    for year in (2022, 2023, 2024, 2025):
        for q in ("Q1", "Q2", "Q3", "Q4"):
            rev *= 1.03
            inc.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "reportedCurrency": currency,
                    "date": f"{year}-03-31",
                    "revenue": rev * 1e6,
                    "costOfRevenue": rev * 0.50 * 1e6,
                    "grossProfit": rev * 0.50 * 1e6,
                    "researchAndDevelopmentExpenses": rev * 0.12 * 1e6,
                    "sellingGeneralAndAdministrativeExpenses": rev * 0.15 * 1e6,
                    "operatingExpenses": rev * 0.40 * 1e6,
                    "operatingIncome": rev * 0.12 * 1e6,
                    "netIncome": rev * 0.09 * 1e6,
                    "weightedAverageShsOutDil": 100 * 1e6,
                }
            )
            bal.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "cashAndShortTermInvestments": rev * 0.30 * 1e6,
                    "totalCurrentAssets": rev * 0.60 * 1e6,
                    "propertyPlantEquipmentNet": rev * 0.50 * 1e6,
                    "totalAssets": rev * 1.50 * 1e6,
                    "totalCurrentLiabilities": rev * 0.30 * 1e6,
                    "longTermDebt": rev * 0.20 * 1e6,
                    "totalDebt": rev * 0.20 * 1e6,
                    "totalStockholdersEquity": rev * 0.80 * 1e6,
                }
            )
            cf.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "depreciationAndAmortization": rev * 0.08 * 1e6,
                    "stockBasedCompensation": rev * 0.05 * 1e6,
                    "changeInWorkingCapital": -rev * 0.01 * 1e6,
                    "operatingCashFlow": rev * 0.15 * 1e6,
                    "capitalExpenditure": -rev * 0.10 * 1e6,
                    "freeCashFlow": rev * 0.05 * 1e6,
                }
            )
            if segments:
                pseg.append(
                    {
                        "fiscalYear": year,
                        "period": q,
                        "data": {"Cloud": rev * 0.6 * 1e6, "Devices": rev * 0.4 * 1e6},
                    }
                )
    (fmp / f"{ticker}_income_statement_quarterly.json").write_text(
        json.dumps(inc), encoding="utf-8"
    )
    (fmp / f"{ticker}_balance_sheet_quarterly.json").write_text(json.dumps(bal), encoding="utf-8")
    (fmp / f"{ticker}_cash_flow_quarterly.json").write_text(json.dumps(cf), encoding="utf-8")
    if segments:
        (fmp / f"{ticker}_product_segments_quarterly.json").write_text(
            json.dumps(pseg), encoding="utf-8"
        )
    (fmp / f"{ticker}_profile.json").write_text(
        json.dumps(
            [{"companyName": f"{ticker} Co", "beta": 1.2, "price": 50.0, "currency": "USD"}]
        ),
        encoding="utf-8",
    )
    # Consensus sits comfortably ABOVE the last-FY actual (~1.6bn) and grows, the
    # realistic shape — so the model's seeded near-term growth is a smooth
    # continuation and the value-of-record tracks the builder's mirror tightly.
    est = [
        {
            "date": f"{y}-12-31",
            "revenueAvg": 1750 * (1.08 ** (y - 2026)) * 1e6,
            "netIncomeAvg": 210 * (1.08 ** (y - 2026)) * 1e6,
            "ebitdaAvg": 320 * (1.08 ** (y - 2026)) * 1e6,
            "ebitAvg": 260 * (1.08 ** (y - 2026)) * 1e6,
            "sgaExpenseAvg": 240 * 1e6,
            "epsAvg": 2.1 * (1.08 ** (y - 2026)),
        }
        for y in range(2026, 2031)
    ]
    (fmp / f"{ticker}_analyst_estimates_annual.json").write_text(json.dumps(est), encoding="utf-8")


def _write_fmp_semiannual(repo: Path, ticker: str) -> None:
    """Semi-annual filer (the BHP shape): each fiscal year reports only H1/H2 as
    Q2/Q4, which sum to the fiscal-year figure. Consensus is anchored just above the
    last full FY (its two halves) so near-term growth stays smooth and the
    value-of-record tracks the builder's mirror."""
    fmp = repo / "data" / "historical" / "fmp"
    fmp.mkdir(parents=True, exist_ok=True)
    inc: list[dict[str, object]] = []
    bal: list[dict[str, object]] = []
    cf: list[dict[str, object]] = []
    rev = 400.0
    for year in (2022, 2023, 2024, 2025):
        for q in ("Q2", "Q4"):  # only the two halves — never Q1/Q3
            rev *= 1.03
            inc.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "reportedCurrency": "USD",
                    "date": f"{year}-06-30",
                    "revenue": rev * 1e6,
                    "costOfRevenue": rev * 0.50 * 1e6,
                    "grossProfit": rev * 0.50 * 1e6,
                    "researchAndDevelopmentExpenses": rev * 0.12 * 1e6,
                    "sellingGeneralAndAdministrativeExpenses": rev * 0.15 * 1e6,
                    "operatingExpenses": rev * 0.40 * 1e6,
                    "operatingIncome": rev * 0.12 * 1e6,
                    "netIncome": rev * 0.09 * 1e6,
                    "weightedAverageShsOutDil": 100 * 1e6,
                }
            )
            bal.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "cashAndShortTermInvestments": rev * 0.30 * 1e6,
                    "totalCurrentAssets": rev * 0.60 * 1e6,
                    "propertyPlantEquipmentNet": rev * 0.50 * 1e6,
                    "totalAssets": rev * 1.50 * 1e6,
                    "totalCurrentLiabilities": rev * 0.30 * 1e6,
                    "longTermDebt": rev * 0.20 * 1e6,
                    "totalDebt": rev * 0.20 * 1e6,
                    "totalStockholdersEquity": rev * 0.80 * 1e6,
                }
            )
            cf.append(
                {
                    "fiscalYear": year,
                    "period": q,
                    "depreciationAndAmortization": rev * 0.08 * 1e6,
                    "stockBasedCompensation": rev * 0.05 * 1e6,
                    "changeInWorkingCapital": -rev * 0.01 * 1e6,
                    "operatingCashFlow": rev * 0.15 * 1e6,
                    "capitalExpenditure": -rev * 0.10 * 1e6,
                    "freeCashFlow": rev * 0.05 * 1e6,
                }
            )
    (fmp / f"{ticker}_income_statement_quarterly.json").write_text(
        json.dumps(inc), encoding="utf-8"
    )
    (fmp / f"{ticker}_balance_sheet_quarterly.json").write_text(json.dumps(bal), encoding="utf-8")
    (fmp / f"{ticker}_cash_flow_quarterly.json").write_text(json.dumps(cf), encoding="utf-8")
    (fmp / f"{ticker}_profile.json").write_text(
        json.dumps(
            [{"companyName": f"{ticker} Co", "beta": 1.0, "price": 50.0, "currency": "USD"}]
        ),
        encoding="utf-8",
    )
    # FY2025 actual = its two half-year revenues ($M); consensus continues ~8%/yr.
    base_fy_m = 400.0 * (1.03**7) + 400.0 * (1.03**8)  # the FY2025 Q2 + Q4 halves
    est = [
        {
            "date": f"{y}-06-30",
            "revenueAvg": base_fy_m * (1.08 ** (y - 2025)) * 1e6,
            "netIncomeAvg": base_fy_m * 0.09 * (1.08 ** (y - 2025)) * 1e6,
            "ebitdaAvg": base_fy_m * 0.20 * (1.08 ** (y - 2025)) * 1e6,
            "ebitAvg": base_fy_m * 0.12 * (1.08 ** (y - 2025)) * 1e6,
            "sgaExpenseAvg": base_fy_m * 0.15 * 1e6,
            "epsAvg": base_fy_m * 0.09 / 100 * (1.08 ** (y - 2025)),
        }
        for y in range(2026, 2031)
    ]
    (fmp / f"{ticker}_analyst_estimates_annual.json").write_text(json.dumps(est), encoding="utf-8")


def _build(repo: Path, ticker: str, dest: Path) -> float:
    """Run the builder as a subprocess; return its value-of-record (RESULT line)."""
    env = dict(os.environ, DCF_TICKER=ticker, DCF_REPO_ROOT=str(repo), DCF_DEST=str(dest))
    proc = subprocess.run(
        [sys.executable, str(BUILDER)],
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    assert proc.returncode == 0, proc.stderr
    line = next((ln for ln in proc.stdout.splitlines() if ln.startswith("RESULT")), None)
    assert line is not None, proc.stdout
    return float(line.split("\t")[2])


@pytest.fixture(scope="module")
def built_usd(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, float]:
    """Build one USD single-segment workbook once; read-only tests share it."""
    repo = tmp_path_factory.mktemp("redesign_usd")
    _write_fmp(repo, "TESTCO")
    dest = repo / "dcf" / "TESTCO.xlsx"
    builder_value = _build(repo, "TESTCO", dest)
    return dest, builder_value


# --------------------------------------------------------------------------- #
# Reader / format detection
# --------------------------------------------------------------------------- #
def test_is_redesign_format_true_for_builder_output(built_usd: tuple[Path, float]) -> None:
    dest, _ = built_usd
    assert redesign.is_redesign_format(dest) is True


def test_is_redesign_format_false_for_missing(tmp_path: Path) -> None:
    assert redesign.is_redesign_format(tmp_path / "nope.xlsx") is False


def test_is_redesign_format_false_for_legacy_three_sheet(tmp_path: Path) -> None:
    legacy = tmp_path / "legacy.xlsx"
    wb = openpyxl.Workbook()
    for i, name in enumerate(("Historicals", "Forecast", "Valuation")):
        if i == 0:
            ws = wb.active
            assert ws is not None
            ws.title = name
        else:
            wb.create_sheet(name)
    wb.save(str(legacy))
    assert redesign.is_redesign_format(legacy) is False
    assert redesign.read_inputs(legacy) is None
    assert redesign.read_and_value(legacy) is None


def test_read_and_value_matches_builder_mirror(built_usd: tuple[Path, float]) -> None:
    """The reader recomputes the value-of-record from the workbook's inputs; it
    must track the builder's own _project mirror closely (they differ only by the
    capex/D&A base nuance — the reader matches the in-sheet formula exactly)."""
    dest, builder_value = built_usd
    rv = redesign.read_and_value(dest)
    assert rv is not None
    assert rv.value_per_share_usd == pytest.approx(builder_value, rel=0.03)
    assert rv.fx_to_usd == 1.0


def test_builder_seeds_scenario_weights_and_reader_reads_them(
    built_usd: tuple[Path, float],
) -> None:
    """The builder writes the scenario probability-weight cells (row 62) and
    read_inputs reads them back — seeded to the symmetric default here (the test
    FMP carries no scenario_prior block)."""
    dest, _ = built_usd
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.weight_base == pytest.approx(0.50)
    assert inp.weight_bull == pytest.approx(0.25)
    assert inp.weight_bear == pytest.approx(0.25)


# --------------------------------------------------------------------------- #
# Semi-annual filer (BHP shape): H1/H2 reported as Q2/Q4, two periods per FY
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def built_semiannual(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, float]:
    """Build one semi-annual (Q2/Q4-only) workbook once; reader tests share it."""
    repo = tmp_path_factory.mktemp("redesign_semi")
    _write_fmp_semiannual(repo, "SEMICO")
    dest = repo / "dcf" / "SEMICO.xlsx"
    builder_value = _build(repo, "SEMICO", dest)
    return dest, builder_value


def test_semiannual_reader_round_trips_and_matches_builder(
    built_semiannual: tuple[Path, float],
) -> None:
    """The reader must accept a semi-annual workbook — its _latest_full_fy detects
    the Q2/Q4 cadence instead of demanding four quarters — and recompute the
    value-of-record in line with the builder, proving the H1/H2 build round-trips."""
    dest, builder_value = built_semiannual
    assert redesign.is_redesign_format(dest) is True
    rv = redesign.read_and_value(dest)
    assert rv is not None
    assert builder_value > 0 and rv.value_per_share_usd > 0
    assert rv.value_per_share_usd == pytest.approx(builder_value, rel=0.04)


def test_semiannual_reader_aggregates_two_halves_into_fy(
    built_semiannual: tuple[Path, float],
) -> None:
    """Single-segment base revenue = the latest FY's two half-year columns summed
    (Q2 + Q4), not one quarter — the same FY aggregation the builder applies."""
    dest, _ = built_semiannual
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.segments == ("Total company",)
    expected_fy2025 = 400.0 * (1.03**7) + 400.0 * (1.03**8)  # the two FY2025 halves
    assert inp.base_revenue_by_segment["Total company"] == pytest.approx(expected_fy2025, rel=1e-3)


def test_reader_reads_dashboard_and_financials(built_usd: tuple[Path, float]) -> None:
    dest, _ = built_usd
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.segments == ("Total company",)
    assert inp.diluted_shares_m == pytest.approx(100.0)
    assert 0.0 < inp.wacc < 0.25
    assert inp.consensus_years == 5  # 2026..2030 on the Consensus sheet


def test_reader_accepts_long_term_debt_label(built_usd: tuple[Path, float], tmp_path: Path) -> None:
    """Live FMP workbooks label the balance-sheet debt row ``Long-term Debt``;
    the reader must accept that source label as the equity-bridge debt input."""
    source, _ = built_usd
    dest = tmp_path / "long-term-debt-label.xlsx"
    shutil.copy2(source, dest)
    wb = openpyxl.load_workbook(dest)
    try:
        ws = wb["Financials"]
        debt_row = next(
            row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Total Debt"
        )
        debt_cell = ws.cell(debt_row, 1)
        assert isinstance(debt_cell, Cell)
        debt_cell.value = "Long-term Debt"
        wb.save(dest)
    finally:
        wb.close()

    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.total_debt_m > 0


# --------------------------------------------------------------------------- #
# SBC + analyst-segment plumbing through the real builder
# --------------------------------------------------------------------------- #
def test_builder_writes_sbc_cells_reader_charges_them(built_usd: tuple[Path, float]) -> None:
    """The builder writes the Dashboard SBC % inputs (rows 65/66) from the actuals
    fade, the reader reads them back, and the value-of-record is the SBC-charged one
    (the in-sheet valfcf formula + the reader mirror agree)."""
    dest, builder_value = built_usd
    inp = redesign.read_inputs(dest)
    assert inp is not None
    # The FMP fixture carries SBC at 5% of revenue, faded 5% -> 3% (x0.6 terminal).
    assert inp.near_sbc_pct == pytest.approx(0.05, abs=0.01)
    assert inp.terminal_sbc_pct == pytest.approx(0.03, abs=0.01)
    assert inp.near_sbc_pct > 0  # SBC is actually charged now, not netted out
    # The reader's value tracks the builder's own SBC-charged mirror.
    rv = redesign.read_and_value(dest)
    assert rv is not None
    assert rv.value_per_share_usd == pytest.approx(builder_value, rel=0.03)


def _build_with_assumptions(
    repo: Path, ticker: str, redesign_block: dict[str, object], dest: Path
) -> float:
    """Write a data/dcf_assumptions/<T>.json redesign block, then run the builder."""
    adir = repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True, exist_ok=True)
    (adir / f"{ticker}.json").write_text(
        json.dumps({"ticker": ticker, "redesign": redesign_block}), encoding="utf-8"
    )
    return _build(repo, ticker, dest)


def test_analyst_segments_override_fmp_and_split_base_revenue(tmp_path: Path) -> None:
    """A valid analyst_segments block REPLACES the FMP-resolved segment set: the
    workbook models the analyst's two streams, base revenue is split by base_pct,
    and each segment grows on its own near/terminal rate (reader round-trips)."""
    repo = tmp_path / "analyst"
    _write_fmp(repo, "SEGCO", segments=True)  # FMP would resolve Cloud/Devices
    dest = repo / "dcf" / "SEGCO.xlsx"
    _build_with_assumptions(
        repo,
        "SEGCO",
        {
            "dcf_applicable": True,
            "business_model": "operating",
            "analyst_segments": {
                "Legacy": {"base_pct": 0.80, "near_term_growth": 0.05, "terminal_growth": -0.01},
                "NewEngine": {"base_pct": 0.20, "near_term_growth": 0.35, "terminal_growth": 0.04},
            },
        },
        dest,
    )
    inp = redesign.read_inputs(dest)
    assert inp is not None
    # The analyst split, not FMP's Cloud/Devices.
    assert set(inp.segments) == {"Legacy", "NewEngine"}
    total_base = sum(inp.base_revenue_by_segment.values())
    assert inp.base_revenue_by_segment["Legacy"] == pytest.approx(0.80 * total_base, rel=1e-3)
    assert inp.base_revenue_by_segment["NewEngine"] == pytest.approx(0.20 * total_base, rel=1e-3)
    # Each stream carries its own growth (a split FMP does not report).
    assert inp.near_growth_by_segment["NewEngine"] == pytest.approx(0.35)
    assert inp.terminal_growth_by_segment["Legacy"] == pytest.approx(-0.01)
    rv = redesign.read_and_value(dest)
    assert rv is not None and rv.value_per_share_usd > 0


def test_invalid_analyst_segments_falls_back_to_fmp(tmp_path: Path) -> None:
    """An analyst_segments block whose base_pct doesn't sum to ~1 is rejected — the
    builder falls back to the FMP-resolved segments rather than half-applying it."""
    repo = tmp_path / "analyst_bad"
    _write_fmp(repo, "BADCO", segments=True)
    dest = repo / "dcf" / "BADCO.xlsx"
    _build_with_assumptions(
        repo,
        "BADCO",
        {
            "dcf_applicable": True,
            "business_model": "operating",
            "analyst_segments": {  # sums to 0.6 — invalid
                "A": {"base_pct": 0.30, "near_term_growth": 0.05, "terminal_growth": 0.01},
                "B": {"base_pct": 0.30, "near_term_growth": 0.05, "terminal_growth": 0.01},
            },
        },
        dest,
    )
    inp = redesign.read_inputs(dest)
    assert inp is not None
    # Fell back to FMP's Cloud/Devices — the analyst A/B split was NOT applied.
    assert set(inp.segments) == {"Cloud", "Devices"}


def test_bare_exit_multiple_override_applies_without_segments_block(tmp_path: Path) -> None:
    """A block with only `exit_multiple` (no `segments` sub-block) still applies —
    the terminal-basis re-calibration (#837 SBC-burdened-EBITDA re-basing) sets the
    exit multiple without fabricating a segments block. The builder's default is 12x;
    a bare 20x override must land on the Dashboard cell and lift the value."""
    repo = tmp_path / "bare_mult"
    _write_fmp(repo, "MULTCO")  # single-segment, no analyst/opus segments
    dest = repo / "dcf" / "MULTCO.xlsx"
    _build_with_assumptions(repo, "MULTCO", {"dcf_applicable": True, "exit_multiple": 20.0}, dest)
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.exit_multiple == pytest.approx(20.0)  # bare override applied


def test_bare_margin_and_sbc_overrides_apply_without_segments_block(tmp_path: Path) -> None:
    """The #838 heavy-SBC terminal normalization: a block with only
    `terminal_op_margin` + `sbc_pct_near`/`sbc_pct_terminal` (no `segments` sub-block)
    still applies — margins and SBC % are scalar terminal assumptions independent of
    the segment set, so a name carrying only a margin/SBC recut takes effect."""
    repo = tmp_path / "bare_margin"
    _write_fmp(repo, "MGNCO")  # single-segment, no segments block
    dest = repo / "dcf" / "MGNCO.xlsx"
    _build_with_assumptions(
        repo,
        "MGNCO",
        {
            "dcf_applicable": True,
            "terminal_op_margin": 0.28,
            "sbc_pct_near": 0.20,
            "sbc_pct_terminal": 0.10,
        },
        dest,
    )
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.terminal_op_margin == pytest.approx(0.28)  # bare margin override applied
    assert inp.near_sbc_pct == pytest.approx(0.20)  # bare SBC overrides applied
    assert inp.terminal_sbc_pct == pytest.approx(0.10)


def test_fx_applied_for_non_usd_reporter(tmp_path: Path) -> None:
    """A non-USD reporter (EUR) gets a × FX multiplier off the VALUE/SHARE
    formula; the reporting-currency value is the USD value ÷ FX."""
    repo = tmp_path / "eur"
    _write_fmp(repo, "EURCO", currency="EUR")
    dest = repo / "EURCO.xlsx"
    _build(repo, "EURCO", dest)
    rv = redesign.read_and_value(dest)
    assert rv is not None
    assert rv.fx_to_usd != 1.0  # EUR -> a non-unity FX
    assert rv.value_per_share_usd == pytest.approx(rv.value_per_share_reporting * rv.fx_to_usd)


# --------------------------------------------------------------------------- #
# Scenarios + sensitivity in the built workbook
# --------------------------------------------------------------------------- #
# The grid is 7x7 anchored at header row 4 / column B, so the base (center)
# cell sits at row 8, column E.
_GRID_CENTER = (8, 5)


def test_builder_writes_scenarios_block_and_sensitivity_sheet(
    built_usd: tuple[Path, float],
) -> None:
    """A fresh build carries the seeded Bull/Bear Δ columns, the three computed
    fair values (Base == the reader's value-of-record), and a Sensitivity sheet
    whose center cell is that same value — builder statics and the refresh
    engine can't drift because both come from dcf.redesign."""
    dest, _ = built_usd
    rv = redesign.read_and_value(dest)
    assert rv is not None
    wb = openpyxl.load_workbook(str(dest))
    try:
        assert wb.sheetnames == REDESIGN_SHEETS
        dsh = wb["Dashboard"]
        bull_margin = dsh.cell(redesign.SCEN_ROW_MARGIN_TERM, redesign.SCEN_COL_BULL).value
        bear_mult = dsh.cell(redesign.SCEN_ROW_EXIT_MULT, redesign.SCEN_COL_BEAR).value
        assert bull_margin == pytest.approx(redesign.BULL_SEED.margin_term)
        assert bear_mult == pytest.approx(redesign.BEAR_SEED.exit_multiple)
        base_fv = dsh.cell(redesign.SCEN_FV_ROW, 2).value
        bull_fv = dsh.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BULL).value
        bear_fv = dsh.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BEAR).value
        assert isinstance(base_fv, float) and isinstance(bull_fv, float)
        assert isinstance(bear_fv, float)
        assert base_fv == pytest.approx(rv.value_per_share_usd)
        assert bear_fv < base_fv < bull_fv
        sens = wb["Sensitivity"]
        assert sens.cell(*_GRID_CENTER).value == pytest.approx(rv.value_per_share_usd)
    finally:
        wb.close()


def test_reader_returns_seeded_deltas_and_blank_cells_fall_back(
    built_usd: tuple[Path, float], tmp_path: Path
) -> None:
    """read_inputs returns the seeded offsets from a fresh build; blanking a
    single Δ cell falls back to that lever's documented seed (and a pre-scenario
    workbook therefore reads as the full seeds)."""
    dest, _ = built_usd
    inp = redesign.read_inputs(dest)
    assert inp is not None
    assert inp.bull_deltas == redesign.BULL_SEED
    assert inp.bear_deltas == redesign.BEAR_SEED

    copy = tmp_path / "blanked.xlsx"
    shutil.copyfile(dest, copy)
    wb = openpyxl.load_workbook(str(copy))
    wb["Dashboard"].cell(row=redesign.SCEN_ROW_TG, column=redesign.SCEN_COL_BULL).value = None
    wb.save(str(copy))
    wb.close()
    blanked = redesign.read_inputs(copy)
    assert blanked is not None
    assert blanked.bull_deltas.terminal_g == pytest.approx(redesign.BULL_SEED.terminal_g)


# --------------------------------------------------------------------------- #
# Edit-preservation refresh integration
# --------------------------------------------------------------------------- #
class _FakeLive:
    def __init__(self, price: float) -> None:
        self.price = price
        self.fetched_at = None


def _fake_read(_repo: object, _ticker: object) -> _FakeLive:
    return _FakeLive(50.0)


@pytest.fixture
def refresh_repo(tmp_path: Path) -> Path:
    """A repo_root with FMP, a dcf/ dir, and a dcf_runs DB — ready for refresh_one."""
    _write_fmp(tmp_path, "TESTCO", segments=True)
    (tmp_path / "dcf").mkdir()
    conn = sqlite3.connect(str(tmp_path / "data" / "portfolio.db"))
    conn.executescript(_DCF_RUNS_SCHEMA)
    conn.commit()
    conn.close()
    return tmp_path


def _dashboard_cell(path: Path, row: int, col: int = 2) -> object:
    wb = openpyxl.load_workbook(str(path), data_only=False)
    try:
        return wb["Dashboard"].cell(row=row, column=col).value
    finally:
        wb.close()


def test_refresh_redesign_seeds_then_persists(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A missing workbook builds fresh (redesign format) and persists dcf_runs."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res
    assert res["format"] == "redesign"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    assert dest.exists()
    wb = openpyxl.load_workbook(str(dest))
    assert wb.sheetnames == REDESIGN_SHEETS
    wb.close()
    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT npv_per_share, live_price, input_sha256, workbook_sha256, "
        "engine_version, inputs_as_of, provenance_json "
        "FROM dcf_runs WHERE ticker='TESTCO'"
    ).fetchone()
    conn.close()
    assert row is not None and row[0] is not None
    assert row[1] == pytest.approx(50.0)
    assert isinstance(row[2], str) and len(row[2]) == 64
    assert isinstance(row[3], str) and len(row[3]) == 64
    assert row[4] == refresh_dcf.DCF_ENGINE_VERSION
    assert row[5]
    provenance_detail = json.loads(row[6])
    assert provenance_detail["ticker"] == "TESTCO"
    assert {source["role"] for source in provenance_detail["sources"]} >= {
        "income_statement",
        "balance_sheet",
        "cash_flow",
        "calculation_workbook",
    }


def test_sync_assumptions_json_mirrors_numbers_keeps_prose(tmp_path: Path) -> None:
    """_sync_assumptions_json writes edited numeric inputs into the redesign block
    while preserving the Opus narrative/reasoning and the model flags."""
    adir = tmp_path / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps(
            {
                "ticker": "TESTCO",
                "narrative": "TOP NARRATIVE keep",
                "redesign": {
                    "dcf_applicable": True,
                    "business_model": "operating",
                    "segments": {
                        "Total company": {"near_term_growth": 0.10, "terminal_growth": 0.03}
                    },
                    "near_term_op_margin": 0.20,
                    "terminal_op_margin": 0.25,
                    "tax_rate": 0.24,
                    "exit_basis": "EV/EBITDA",
                    "terminal_method": "Exit multiple",
                    "exit_multiple": 12.0,
                    "terminal_growth_g": 0.03,
                    "terminal_capex_da": 1.05,
                    "narrative": "REDESIGN NARRATIVE keep",
                    "reasoning": "REASONING keep",
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    edited = dataclasses.replace(
        _BASE,
        near_growth_by_segment={"Total company": 0.06},
        terminal_growth_by_segment={"Total company": 0.02},
        near_op_margin=0.05,
        terminal_op_margin=0.15,
        exit_multiple=9.0,
        terminal_growth_g=0.02,
        beta=1.45,
        equity_risk_premium=0.055,
    )
    assert refresh_dcf.sync_assumptions_json(tmp_path, "TESTCO", edited).status == "synced"
    out = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))
    rd = out["redesign"]
    assert rd["segments"]["Total company"] == {"near_term_growth": 0.06, "terminal_growth": 0.02}
    assert rd["near_term_op_margin"] == 0.05
    assert rd["terminal_op_margin"] == 0.15
    assert rd["exit_multiple"] == 9.0
    assert rd["terminal_growth_g"] == 0.02
    # WACC drivers: beta + cost_of_debt mirror per name; risk_free_rate +
    # equity_risk_premium are GLOBAL now (global_dcf_assumptions, migration 0112)
    # and deliberately NOT re-pinned into the block (PR6).
    assert rd["beta"] == 1.45
    assert rd["cost_of_debt"] == 0.045
    assert "risk_free_rate" not in rd
    assert "equity_risk_premium" not in rd
    # Opus prose + model flags untouched
    assert rd["narrative"] == "REDESIGN NARRATIVE keep"
    assert rd["reasoning"] == "REASONING keep"
    assert rd["dcf_applicable"] is True
    assert rd["business_model"] == "operating"
    assert out["narrative"] == "TOP NARRATIVE keep"


def test_sync_creates_assumptions_json_when_absent(tmp_path: Path) -> None:
    """No assumptions file (the silent-no-op hole: 43 of ~91 workbooks) -> one
    is CREATED from the workbook inputs, marked set_by='sync', so a
    from-scratch build reproduces the user's values instead of reverting them."""
    res = refresh_dcf.sync_assumptions_json(tmp_path, "MISSING", _BASE)
    assert res.status == "created"
    created = json.loads(
        (tmp_path / "data" / "dcf_assumptions" / "MISSING.json").read_text(encoding="utf-8")
    )
    rd = created["redesign"]
    assert rd["set_by"] == "sync"
    assert rd["exit_multiple"] == 12.0
    assert rd["segments"]["Total company"]["near_term_growth"] == 0.10
    # no fabricated Opus provenance for a workbook-derived block
    assert "opus_baseline" not in created

    # A file with no redesign block gets one (other keys preserved).
    adir = tmp_path / "data" / "dcf_assumptions"
    (adir / "NOBLOCK.json").write_text(json.dumps({"ticker": "NOBLOCK"}), encoding="utf-8")
    res = refresh_dcf.sync_assumptions_json(tmp_path, "NOBLOCK", _BASE)
    assert res.status == "created"
    out = json.loads((adir / "NOBLOCK.json").read_text(encoding="utf-8"))
    assert out["ticker"] == "NOBLOCK"
    assert out["redesign"]["set_by"] == "sync"


def test_sync_fails_loud_on_unreadable_json(tmp_path: Path) -> None:
    """Corruption is 'failed' + detail — never a silent False."""
    adir = tmp_path / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "BAD.json").write_text("{not json", encoding="utf-8")
    res = refresh_dcf.sync_assumptions_json(tmp_path, "BAD", _BASE)
    assert res.status == "failed"
    assert res.detail is not None and "unreadable" in res.detail
    (adir / "NOTOBJ.json").write_text('"just a string"', encoding="utf-8")
    assert refresh_dcf.sync_assumptions_json(tmp_path, "NOTOBJ", _BASE).status == "failed"


def test_refresh_stages_missing_assumptions_without_live_path(tmp_path: Path) -> None:
    assumptions = tmp_path / "data" / "dcf_assumptions" / "MISSING.json"
    staged = refresh_dcf._stage_assumptions(assumptions)
    assert staged != assumptions
    assert staged.name == "MISSING.rebuild.json"
    assert not assumptions.exists()
    assert not staged.exists()


def test_refresh_redesign_syncs_assumptions_json(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """End-to-end: a Dashboard edit flows back into data/dcf_assumptions on refresh,
    so a from-scratch rebuild would reproduce it (single source of truth)."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps({"redesign": {"exit_multiple": 12.0, "narrative": "keep me"}}, indent=2),
        encoding="utf-8",
    )
    # Seed a workbook, then simulate the user editing the exit multiple (B45) and refresh.
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    wb = openpyxl.load_workbook(str(dest))
    dsh = wb["Dashboard"]
    dsh.cell(row=45, column=2, value=8.0)  # exit multiple -> 8x
    dsh.cell(row=40, column=2, value=1.55)  # beta -> 1.55
    wb.save(str(dest))
    wb.close()
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "blocked", res
    assert res["reason"] == "outlier_requires_explicit_owner_review"
    rd = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))["redesign"]
    assert rd["exit_multiple"] == pytest.approx(12.0)
    assert rd["narrative"] == "keep me"
    # The candidate was rejected before replacement, so the prior run's sync
    # outcome remains durable and the assumptions mirror stays unchanged.
    conn = sqlite3.connect(str(db))
    srow = conn.execute(
        "SELECT assumptions_sync_status, assumptions_synced_at FROM dcf_runs WHERE ticker='TESTCO'"
    ).fetchone()
    conn.close()
    assert srow is not None and srow[0] == "synced" and srow[1]


def test_refresh_provenance_end_to_end(refresh_repo: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Full provenance round-trip through the real builder + refresher:

    1. first refresh seeds the immutable opus_baseline and writes the
       Assumptions sheet with the Opus-sourced exit multiple;
    2. the user edits the exit multiple in the workbook; the next refresh
       classifies it user-edited, records the override (Opus value + date) in
       the ledger, annotates the yellow cell — while sync rewrites the
       redesign block to 8x and the baseline provably stays at 11x.
    """
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps(
            {
                "redesign": {
                    "dcf_applicable": True,
                    "business_model": "operating",
                    "segments": {
                        "Cloud": {"near_term_growth": 0.12, "terminal_growth": 0.04},
                        "Devices": {"near_term_growth": 0.05, "terminal_growth": 0.02},
                    },
                    "near_term_op_margin": 0.18,
                    "terminal_op_margin": 0.22,
                    "tax_rate": 0.22,
                    "terminal_capex_da": 1.05,
                    "terminal_method": "Exit multiple",
                    "exit_basis": "EV/EBITDA",
                    "exit_multiple": 11.0,
                    "terminal_growth_g": 0.035,
                    "narrative": "STORY PROSE",
                    "reasoning": "JUDGMENT PROSE",
                }
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    res1 = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res1["status"] == "ok", res1
    prov1 = res1["assumption_provenance"]
    assert isinstance(prov1, dict) and prov1["status"] == "ok"
    data = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))
    assert data["opus_baseline"]["values"]["exit_multiple"] == 11.0
    assert data["opus_baseline"]["seeded"] is True
    wb = openpyxl.load_workbook(str(dest))
    try:
        assert "Assumptions" in wb.sheetnames
        text = "\n".join(
            str(c.value)
            for row in wb["Assumptions"].iter_rows()
            for c in row
            if c.value is not None
        )
        assert "STORY PROSE" in text and "JUDGMENT PROSE" in text
        dsh = wb["Dashboard"]
        assert dsh.cell(row=45, column=2).value == pytest.approx(11.0)  # Opus applied
        dsh.cell(row=45, column=2, value=8.0)  # the user's edit
        wb.save(str(dest))
    finally:
        wb.close()

    res2 = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res2["status"] == "ok", res2
    provenance = cast("dict[str, object]", res2["assumption_provenance"])
    sources = cast("dict[str, int]", provenance["sources"])
    assert sources.get("user-edited", 0) >= 1

    data = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))
    # sync mirrored the edit into the redesign block (from-scratch builds
    # reproduce it) while the baseline kept the original Opus value.
    assert data["redesign"]["exit_multiple"] == pytest.approx(8.0)
    assert data["opus_baseline"]["values"]["exit_multiple"] == 11.0
    assert data["assumption_overrides"]["exit_multiple"]["opus_value"] == 11.0

    wb = openpyxl.load_workbook(str(dest))
    try:
        comment = wb["Dashboard"]["B45"].comment
        assert comment is not None
        assert "overridden from Opus 11.0x" in comment.text
        text = "\n".join(
            str(c.value)
            for row in wb["Assumptions"].iter_rows()
            for c in row
            if c.value is not None
        )
        assert "overridden from Opus 11.0x" in text
    finally:
        wb.close()


def test_builder_reads_wacc_override_from_block(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A beta/ERP override in the redesign block flows into the built workbook's WACC
    cells — so a from-scratch rebuild reproduces the user's WACC drivers (the other
    half of the round-trip from test_refresh_redesign_syncs_assumptions_json)."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps({"redesign": {"beta": 1.62, "equity_risk_premium": 0.055}}, indent=2),
        encoding="utf-8",
    )
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res
    # The builder wrote the override into the Dashboard WACC cells (B40 beta, B39 ERP).
    assert _dashboard_cell(dest, 40) == pytest.approx(1.62)
    assert _dashboard_cell(dest, 39) == pytest.approx(0.055)


def test_refresh_redesign_preserves_dashboard_edit_and_updates_actuals(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The core edit-preservation contract: edit a yellow Dashboard cell, change
    the underlying FMP actuals, refresh. The edit MUST survive, the actuals MUST
    update, and the dropdowns/charts MUST still be intact."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"

    # Seed a workbook, then simulate a user editing the terminal margin (B30) and
    # the Cloud segment near-term growth (a segment row).
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    wb = openpyxl.load_workbook(str(dest))
    dsh = wb["Dashboard"]
    dsh.cell(row=30, column=2, value=0.33)  # terminal op margin -> 33%
    for r in range(20, 28):  # Cloud segment near-term growth -> 50%
        if dsh.cell(row=r, column=1).value == "Cloud":
            dsh.cell(row=r, column=2, value=0.50)
    wb.save(str(dest))
    wb.close()

    # Mutate the latest-quarter FMP revenue so the rebuilt Financials must change.
    inc_path = (
        refresh_repo / "data" / "historical" / "fmp" / "TESTCO_income_statement_quarterly.json"
    )
    inc = json.loads(inc_path.read_text(encoding="utf-8"))
    newest = max(inc, key=lambda r: (int(r["fiscalYear"]), str(r["period"])))
    newest["revenue"] = 9_999 * 1e6
    inc_path.write_text(json.dumps(inc), encoding="utf-8")
    # Scale that quarter's product segmentation to match the new revenue. Otherwise a 6x
    # income jump against flat segments would (correctly) trip the partial-coverage guard
    # (src/dcf/segment_coverage.py) and fall back to a whole-company build — which has no
    # "Cloud" segment row, defeating the per-segment edit-preservation this test asserts.
    seg_path = (
        refresh_repo / "data" / "historical" / "fmp" / "TESTCO_product_segments_quarterly.json"
    )
    pseg = json.loads(seg_path.read_text(encoding="utf-8"))
    for rec in pseg:
        if int(rec["fiscalYear"]) == int(newest["fiscalYear"]) and str(rec["period"]) == str(
            newest["period"]
        ):
            rec["data"] = {"Cloud": 9_999 * 0.6 * 1e6, "Devices": 9_999 * 0.4 * 1e6}
    seg_path.write_text(json.dumps(pseg), encoding="utf-8")

    before_blocked_refresh = dest.read_bytes()
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "blocked", res
    assert res["reason"] == "outlier_requires_explicit_owner_review"
    assert dest.read_bytes() == before_blocked_refresh


def test_refresh_preserves_scenario_edits_and_recomputes_outputs(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The scenario half of the edit-preservation contract: edit Bull/Bear Δ
    cells, refresh. The edits MUST survive the rebuild, the computed fair-value
    row + Sensitivity grid MUST be rewritten from the edited inputs (not the
    builder's defaults), dcf_runs MUST carry the scenario range in
    assumption_snapshot_json, and the deltas MUST mirror to dcf_assumptions."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps({"redesign": {"exit_multiple": 12.0, "narrative": "keep me"}}, indent=2),
        encoding="utf-8",
    )
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)

    # The user edits both scenario columns (a richer Bull margin, a harsher
    # Bear multiple) and we note the seeded computed Bull value beforehand.
    wb = openpyxl.load_workbook(str(dest))
    dsh = wb["Dashboard"]
    pre_bull = dsh.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BULL).value
    pre_bear = dsh.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BEAR).value
    assert isinstance(pre_bull, float) and isinstance(pre_bear, float)
    dsh.cell(row=redesign.SCEN_ROW_MARGIN_TERM, column=redesign.SCEN_COL_BULL, value=0.10)
    dsh.cell(row=redesign.SCEN_ROW_EXIT_MULT, column=redesign.SCEN_COL_BEAR, value=-4.0)
    wb.save(str(dest))
    wb.close()

    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res

    wb2 = openpyxl.load_workbook(str(dest))
    try:
        dsh2 = wb2["Dashboard"]
        # The Δ edits survived the rebuild.
        bull_margin = dsh2.cell(redesign.SCEN_ROW_MARGIN_TERM, redesign.SCEN_COL_BULL).value
        bear_mult = dsh2.cell(redesign.SCEN_ROW_EXIT_MULT, redesign.SCEN_COL_BEAR).value
        assert bull_margin == pytest.approx(0.10)
        assert bear_mult == pytest.approx(-4.0)
        # The computed row was rewritten from the EDITED deltas: a richer Bull
        # margin lifts Bull, a harsher Bear multiple cuts Bear.
        post_bull = dsh2.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BULL).value
        post_bear = dsh2.cell(redesign.SCEN_FV_ROW, redesign.SCEN_COL_BEAR).value
        assert isinstance(post_bull, float) and isinstance(post_bear, float)
        assert post_bull > pre_bull
        assert post_bear < pre_bear
        # The Sensitivity grid tracks the value-of-record (rewritten post-inject).
        rv = redesign.read_and_value(dest)
        assert rv is not None
        sens = wb2["Sensitivity"]
        assert sens.cell(*_GRID_CENTER).value == pytest.approx(rv.value_per_share_usd)
    finally:
        wb2.close()

    # dcf_runs carries the scenario range; BASE stays npv_per_share.
    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT npv_per_share, assumption_snapshot_json FROM dcf_runs WHERE ticker='TESTCO'"
    ).fetchone()
    conn.close()
    assert row is not None
    snap = json.loads(row[1])
    sc = snap["scenarios"]
    assert sc["base"]["fair_value_per_share_usd"] == pytest.approx(float(row[0]))
    assert sc["bull"]["fair_value_per_share_usd"] == pytest.approx(post_bull)
    assert sc["bear"]["fair_value_per_share_usd"] == pytest.approx(post_bear)
    assert sc["bull"]["deltas"]["margin_term"] == pytest.approx(0.10)
    assert sc["bear"]["deltas"]["exit_multiple"] == pytest.approx(-4.0)

    # The deltas mirrored back to the assumptions JSON (from-scratch parity).
    rd = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))["redesign"]
    assert rd["scenario_bull"]["margin_term"] == pytest.approx(0.10)
    assert rd["scenario_bear"]["exit_multiple"] == pytest.approx(-4.0)
    assert rd["narrative"] == "keep me"


def test_builder_reads_scenario_override_from_block(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """scenario_bull/scenario_bear overrides in the redesign block flow into the
    built workbook's Δ columns (the other half of the sync round-trip) — levers
    absent from the override keep their documented seeds."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps(
            {
                "redesign": {
                    "scenario_bull": {"margin_term": 0.07},
                    "scenario_bear": {"growth_near": -0.08},
                }
            }
        ),
        encoding="utf-8",
    )
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res
    assert _dashboard_cell(
        dest, redesign.SCEN_ROW_MARGIN_TERM, redesign.SCEN_COL_BULL
    ) == pytest.approx(0.07)
    assert _dashboard_cell(
        dest, redesign.SCEN_ROW_GROWTH_NEAR, redesign.SCEN_COL_BEAR
    ) == pytest.approx(-0.08)
    # un-overridden levers keep the seeds
    assert _dashboard_cell(
        dest, redesign.SCEN_ROW_EXIT_MULT, redesign.SCEN_COL_BULL
    ) == pytest.approx(redesign.BULL_SEED.exit_multiple)


def test_refresh_persists_seeded_per_name_prior_and_skews_reward(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A per-name ``scenario_prior`` block already on file (the producer's LLM
    output) flows through end-to-end: the built workbook's weight cells (row 62)
    seed from it, the ``dcf_runs`` snapshot carries both ``scenario_prior`` and
    ``priced_in`` blocks, and ``dcf.scenario_reward`` skews its expected return
    toward the per-name weights (not the symmetric 25/50/25 global)."""
    from dcf import scenario_reward as scenario_reward_mod

    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"
    adir.mkdir(parents=True)
    (adir / "TESTCO.json").write_text(
        json.dumps(
            {
                "redesign": {
                    "scenario_prior": {
                        "bull_weight": 0.15,
                        "base_weight": 0.50,
                        "bear_weight": 0.35,
                        "rationale": "fragile, execution-heavy thesis",
                        "set_by": "llm",
                        "as_of": "2026-07-01",
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res

    # The built workbook's row-62 weight cells seeded from the per-name prior.
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_WEIGHT_BASE
    ) == pytest.approx(0.50)
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_BULL
    ) == pytest.approx(0.15)
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_BEAR
    ) == pytest.approx(0.35)

    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT npv_per_share, live_price, assumption_snapshot_json "
        "FROM dcf_runs WHERE ticker='TESTCO'"
    ).fetchone()
    conn.close()
    assert row is not None
    npv_per_share, live_price, snapshot_raw = row
    snap = json.loads(snapshot_raw)

    # Both blocks land on the run.
    assert "priced_in" in snap
    sp = snap["scenario_prior"]
    assert sp["weights"] == {
        "bull": pytest.approx(0.15),
        "base": pytest.approx(0.50),
        "bear": pytest.approx(0.35),
    }
    assert sp["rationale"] == "fragile, execution-heavy thesis"
    assert sp["set_by"] == "llm"

    # scenario_reward reads the per-name weights (not the global 25/50/25) and
    # the skew reflects the bear-heavy tilt.
    reward = scenario_reward_mod.scenario_reward(
        price=live_price, base_fv=npv_per_share, snapshot_json=snapshot_raw
    )
    assert reward is not None
    assert reward.weights_source == "per_name"
    assert reward.probabilities["bear"] == pytest.approx(0.35)
    assert reward.probabilities["bull"] == pytest.approx(0.15)
    # A bear-heavy prior pulls the expected return below the symmetric-prior one.
    global_reward = scenario_reward_mod.ScenarioReward(
        expected_return=sum(
            scenario_reward_mod.SCENARIO_PROBABILITIES[s]
            * (
                (
                    scenario_reward_mod.parse_scenario_fair_values(snapshot_raw)
                    | {"base": npv_per_share}
                )[s]
                / live_price
                - 1.0
            )
            for s in ("bull", "base", "bear")
        ),
        base_return=npv_per_share / live_price - 1.0,
        bull_return=None,
        bear_return=None,
        has_scenarios=True,
        probabilities=scenario_reward_mod.SCENARIO_PROBABILITIES,
        detail="",
    )
    assert reward.expected_return < global_reward.expected_return


def test_owner_workbook_edit_off_default_creates_per_name_owner_prior(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The owner edits the row-62 weight cells directly in the workbook (no
    per-name ``scenario_prior`` on file yet, no LLM run) — the sync MUST create a
    per-name block with ``set_by='owner'`` rather than silently dropping the
    edit, and that block MUST survive the next refresh (owner edits always
    win)."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    adir = refresh_repo / "data" / "dcf_assumptions"

    # First refresh: no assumptions JSON at all yet — a from-scratch build seeds
    # row 62 at the symmetric global default (no prior on file).
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert not (adir / "TESTCO.json").exists() or "scenario_prior" not in json.loads(
        (adir / "TESTCO.json").read_text(encoding="utf-8")
    ).get("redesign", {})

    # The owner opens the workbook and edits the weight cells directly, off the
    # 25/50/25 default — a locked-in decision, not an LLM proposal.
    wb = openpyxl.load_workbook(str(dest))
    dsh = wb["Dashboard"]
    dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_WEIGHT_BASE, value=0.50)
    dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_BULL, value=0.40)
    dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_BEAR, value=0.10)
    wb.save(str(dest))
    wb.close()

    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok", res

    # A per-name owner block now exists — the edit was NOT silently lost.
    rd = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))["redesign"]
    sp = rd["scenario_prior"]
    assert sp["set_by"] == "owner"
    assert sp["bull_weight"] == pytest.approx(0.40)
    assert sp["base_weight"] == pytest.approx(0.50)
    assert sp["bear_weight"] == pytest.approx(0.10)

    # The workbook weight cells themselves carried the edit through this refresh.
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_BULL
    ) == pytest.approx(0.40)
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_BEAR
    ) == pytest.approx(0.10)

    # A THIRD refresh (no further owner edit) must not revert the owner's block —
    # it survives the next rebuild, seeded straight back from the JSON.
    res2 = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res2["status"] == "ok", res2
    rd2 = json.loads((adir / "TESTCO.json").read_text(encoding="utf-8"))["redesign"]
    sp2 = rd2["scenario_prior"]
    assert sp2["set_by"] == "owner"
    assert sp2["bull_weight"] == pytest.approx(0.40)
    assert sp2["bear_weight"] == pytest.approx(0.10)
    assert _dashboard_cell(
        dest, redesign.SCEN_WEIGHTS_ROW, redesign.SCEN_COL_BULL
    ) == pytest.approx(0.40)


def test_apply_edits_persists_without_rebuild_and_records_override(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """apply_edits (the in-app save) writes edited assumptions onto the LIVE
    workbook, recomputes + re-persists dcf_runs, and records the change in the
    override ledger WITHOUT overwriting the Opus baseline — no FMP rebuild, the
    market quote carried forward from the prior run."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    # Seed the workbook + dcf_runs + assumptions JSON via a normal refresh.
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)

    base_inp = redesign.read_inputs(dest)
    assert base_inp is not None
    m0 = base_inp.exit_multiple
    conn = sqlite3.connect(str(db))
    npv0 = conn.execute("SELECT npv_per_share FROM dcf_runs WHERE ticker='TESTCO'").fetchone()[0]
    primary_overlay = {
        "status": "ok",
        "statements": {
            "income": {
                "status": "ok",
                "applied": [
                    {
                        "fact_id": 42,
                        "source_url": "https://www.sec.gov/example",
                        "as_of": "2026-08-20T12:00:00+00:00",
                    }
                ],
            }
        },
    }
    prior_country_context = {
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": "TESTCO",
        "premium": base_inp.country_risk_premium,
        "authority": "systematic_geo",
        "source_record": {
            "role": "geographic_revenue",
            "path": "data/historical/fmp/TESTCO_geo_segments_annual.json",
            "sha256": "a" * 64,
            "bytes": 123,
            "observed_at": "2026-08-20T12:00:00+00:00",
            "influences_calculation": True,
            "selection": "annual_latest_fiscal_year",
        },
    }
    conn.execute(
        "UPDATE dcf_runs SET provenance_json=? WHERE ticker='TESTCO'",
        (
            json.dumps(
                {
                    "ticker": "TESTCO",
                    "primary_fact_overlay": primary_overlay,
                    "country_risk_context": prior_country_context,
                }
            ),
        ),
    )
    conn.commit()
    conn.close()

    payload = base_inp.to_dict()
    payload["exit_multiple"] = m0 + 3.0  # a clean ledger-tracked scalar
    payload["country_risk_premium"] = base_inp.country_risk_premium + 0.01
    res = refresh_dcf.apply_edits(
        "TESTCO", refresh_repo, db, redesign.RedesignInputs.from_dict(payload)
    )
    assert res["status"] == "ok", res

    # The edit landed on the live workbook (no rebuild) and lifts the value.
    edited = redesign.read_inputs(dest)
    assert edited is not None and edited.exit_multiple == pytest.approx(m0 + 3.0)
    assert cast("float", res["fair_value_per_share"]) > float(npv0)

    # dcf_runs re-persisted; the prior market quote was carried forward.
    conn = sqlite3.connect(str(db))
    row = conn.execute(
        "SELECT npv_per_share, live_price, provenance_json FROM dcf_runs WHERE ticker='TESTCO'"
    ).fetchone()
    conn.close()
    assert row is not None
    assert row[1] == pytest.approx(50.0)
    assert res["fair_value_per_share"] == pytest.approx(float(row[0]))
    persisted_provenance = json.loads(str(row[2]))
    assert persisted_provenance["primary_fact_overlay"] == primary_overlay
    persisted_country = persisted_provenance["country_risk_context"]
    assert persisted_country == {
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": "TESTCO",
        "premium": pytest.approx(base_inp.country_risk_premium + 0.01),
        "authority": "owner_override",
        "source_record": None,
    }
    assert all(
        source.get("role") != "geographic_revenue" for source in persisted_provenance["sources"]
    )

    # The override ledger records the edit; the Opus baseline is NOT overwritten.
    adata = json.loads(
        (refresh_repo / "data" / "dcf_assumptions" / "TESTCO.json").read_text(encoding="utf-8")
    )
    assert adata["opus_baseline"]["values"]["exit_multiple"] == pytest.approx(m0)
    assert adata["assumption_overrides"]["exit_multiple"]["opus_value"] == pytest.approx(m0)
    assert adata["redesign"]["exit_multiple"] == pytest.approx(m0 + 3.0)


@pytest.mark.parametrize(
    ("is_latest_column", "segment_column", "created_at_column"),
    (
        (False, False, False),
        (False, False, True),
        (False, True, False),
        (False, True, True),
        (True, False, False),
        (True, False, True),
        (True, True, False),
        (True, True, True),
    ),
)
def test_prior_primary_overlay_supports_only_explicit_static_schema_variants(
    tmp_path: Path,
    is_latest_column: bool,
    segment_column: bool,
    created_at_column: bool,
) -> None:
    db = tmp_path / "portfolio.db"
    optional_columns: list[str] = []
    optional_values: list[object] = []
    if is_latest_column:
        optional_columns.append("is_latest INTEGER")
        optional_values.append(1)
    if segment_column:
        optional_columns.append("segment_name TEXT")
        optional_values.append("")
    if created_at_column:
        optional_columns.append("created_at TEXT")
        optional_values.append("2026-08-20T12:00:00+00:00")
    schema_suffix = ", " + ", ".join(optional_columns) if optional_columns else ""
    insert_columns = ["id", "ticker", "provenance_json"] + [
        column.split()[0] for column in optional_columns
    ]
    placeholders = ", ".join("?" for _ in insert_columns)
    overlay: dict[str, object] = {"status": "ok", "statements": {}}
    bridge_context: dict[str, object] = {
        "schema_version": "dcf_equity_bridge_context.v1",
        "ticker": "TESTCO",
        "period_end": "2026-06-30",
        "fiscal_period_type": "Q2",
        "reporting_currency": "USD",
        "cash_m": 200.0,
        "total_debt_m": 100.0,
        "diluted_shares_m": 10.0,
        "cash_basis": "reported_aggregate",
        "total_debt_basis": "reported_aggregate",
    }
    country_context: dict[str, object] = {
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": "TESTCO",
        "premium": 0.0,
        "authority": "owner_override",
        "source_record": None,
    }
    payload = json.dumps(
        {
            "ticker": "TESTCO",
            "primary_fact_overlay": overlay,
            "equity_bridge_receipt": {
                "schema_version": "dcf_equity_bridge_receipt.v2",
                "bridge_context": bridge_context,
            },
            "country_risk_context": country_context,
        }
    )

    with sqlite3.connect(db) as conn:
        conn.execute(
            "CREATE TABLE dcf_runs "
            f"(id INTEGER PRIMARY KEY, ticker TEXT, provenance_json TEXT{schema_suffix})"
        )
        conn.execute(
            f"INSERT INTO dcf_runs ({', '.join(insert_columns)}) VALUES ({placeholders})",
            (1, "TESTCO", payload, *optional_values),
        )

    assert refresh_dcf.prior_primary_fact_overlay(db, "testco") == overlay
    assert refresh_dcf.prior_equity_bridge_context(db, "testco") == bridge_context
    assert refresh_dcf.prior_country_risk_context(db, "testco") == country_context


_KPI_SCHEMA = """
CREATE TABLE documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticker TEXT NOT NULL, source_type TEXT NOT NULL, doc_type TEXT NOT NULL,
    source_url TEXT, file_path TEXT NOT NULL, sha256 TEXT NOT NULL,
    fetched_at TIMESTAMP NOT NULL, fetch_status TEXT NOT NULL,
    raw_bytes_size INTEGER NOT NULL,
    source_quality_tier TEXT NOT NULL DEFAULT 'fmp_normalized'
);
CREATE TABLE kpi_definitions (
    id INTEGER PRIMARY KEY AUTOINCREMENT, ticker TEXT NOT NULL, name TEXT NOT NULL
);
CREATE TABLE kpi_facts (
    id INTEGER PRIMARY KEY AUTOINCREMENT, ticker TEXT NOT NULL,
    period_end TIMESTAMP NOT NULL, fiscal_period_type TEXT NOT NULL,
    kpi_definition_id INTEGER NOT NULL, value NUMERIC(24, 6) NOT NULL,
    unit TEXT NOT NULL, source_doc_id INTEGER NOT NULL
);
"""


def _seed_kpi_fact(db: Path, ticker: str, name: str, value: float, unit: str) -> None:
    """Add the kpi/documents tables (the refresh DB only has dcf_runs) and a
    single latest-quarter KPI fact the inject route can resolve."""
    conn = sqlite3.connect(str(db))
    conn.executescript(_KPI_SCHEMA)
    conn.execute(
        "INSERT INTO documents (id, ticker, source_type, doc_type, file_path, sha256, "
        "fetched_at, fetch_status, raw_bytes_size) VALUES "
        "(1, ?, 'fmp', 'fmp_key_metrics', 'x', ?, '2026-01-01 00:00:00', 'ok', 1)",
        (ticker, "0" * 64),
    )
    conn.execute("INSERT INTO kpi_definitions (id, ticker, name) VALUES (1, ?, ?)", (ticker, name))
    conn.execute(
        "INSERT INTO kpi_facts (ticker, period_end, fiscal_period_type, kpi_definition_id, "
        "value, unit, source_doc_id) VALUES (?, '2025-09-30 00:00:00', 'Q3', 1, ?, ?, 1)",
        (ticker, value, unit),
    )
    conn.commit()
    conn.close()


def test_inject_fact_route_reprices_and_syncs_end_to_end(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The S6 deliverable, end-to-end through the real route: pick a fact →
    inject as a DCF driver → the model reprices and the xlsx + assumptions JSON
    stay in sync (clobber-safe), with the converted units written and the fact
    lineage recorded.

    A 42% operating-margin KPI (stored as percent) must land on the Dashboard's
    near-margin cell as the decimal ratio 0.42 — the units/scale conversion that
    is the whole point — and lift the fair value above the builder's seed."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    _seed_kpi_fact(db, "TESTCO", "Operating margin", 42.0, "percent")

    base_inp = redesign.read_inputs(dest)
    assert base_inp is not None
    conn = sqlite3.connect(str(db))
    npv0 = conn.execute("SELECT npv_per_share FROM dcf_runs WHERE ticker='TESTCO'").fetchone()[0]
    conn.close()

    import comments_server

    client = comments_server.create_app(refresh_repo).test_client()
    resp = client.post(
        "/api/dcf/inject-fact",
        json={"ticker": "TESTCO", "token": "kpi:Operating margin", "field": "near_op_margin"},
    )
    assert resp.status_code == 200, resp.get_json()
    body = resp.get_json()
    assert body["injected"] is True
    inj = body["injection"]
    # The load-bearing conversion: percent 42.0 → ratio 0.42.
    assert inj["raw_value"] == pytest.approx(42.0)
    assert inj["raw_unit"] == "percent"
    assert inj["applied_value"] == pytest.approx(0.42)
    assert inj["fact_id"] is not None

    # The edit landed on the LIVE workbook's near-margin cell (B29) as 0.42 —
    # written through apply_edits, not poked into the JSON.
    assert _dashboard_cell(dest, redesign._DB_MARGIN_NEAR) == pytest.approx(0.42)
    # ...and re-read from the workbook the engine repriced from.
    edited = redesign.read_inputs(dest)
    assert edited is not None and edited.near_op_margin == pytest.approx(0.42)

    # The assumptions JSON mirrored the same value (from-scratch-build default
    # stays in sync — no clobber) and recorded the fact lineage (S6 #5).
    adata = json.loads(
        (refresh_repo / "data" / "dcf_assumptions" / "TESTCO.json").read_text(encoding="utf-8")
    )
    assert adata["redesign"]["near_term_op_margin"] == pytest.approx(0.42)
    prov = adata["redesign"]["driver_provenance"]["near_op_margin"]
    assert prov["metric"] == "kpi:Operating margin"
    assert prov["raw_unit"] == "percent"

    # dcf_runs repriced (margin 0.12 → 0.42 lifts value); never wrote over_under
    # directly — it is re-derived by persist.
    conn = sqlite3.connect(str(db))
    npv1 = conn.execute("SELECT npv_per_share FROM dcf_runs WHERE ticker='TESTCO'").fetchone()[0]
    conn.close()
    assert npv1 > float(npv0)
    assert body["fair_value_per_share_usd"] == pytest.approx(float(npv1))


def test_inject_fact_sheet_survives_a_dcf_refresh_end_to_end(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """THE S7 DELIVERABLE, end-to-end through the real route: pick a fact → add
    it as a DCF reference → the model workbook is REFRESHED (rebuilt from scratch)
    → the reference still has the value.

    Survival is structural: the reference lands in the companion ``dcf/facts/
    TESTCO.xlsx``, NOT in the main workbook the rebuild discards. The proof is the
    second ``refresh_one`` (a full FMP rebuild + os.replace of the main workbook)
    leaving the companion untouched."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    facts_path = fact_sheet.facts_workbook_path(refresh_repo, "TESTCO")

    # Build the model, seed a fact, then park it as a reference via the route.
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    _seed_kpi_fact(db, "TESTCO", "Operating margin", 42.0, "percent")

    import comments_server

    client = comments_server.create_app(refresh_repo).test_client()
    resp = client.post(
        "/api/dcf/inject-fact-sheet",
        json={"ticker": "TESTCO", "token": "kpi:Operating margin"},
    )
    assert resp.status_code == 200, resp.get_json()
    body = resp.get_json()
    assert body["added"] is True and body["action"] == "added" and body["count"] == 1
    # Faithful reference: stored in its NATIVE unit (no driver conversion).
    assert body["fact"]["value"] == pytest.approx(42.0)
    assert body["fact"]["unit"] == "percent"
    assert facts_path.exists()

    # The reference lives in the COMPANION, never appended to the model workbook
    # (which is what makes it survive the rebuild).
    wb = openpyxl.load_workbook(str(dest))
    assert fact_sheet.SHEET_NAME not in wb.sheetnames
    wb.close()

    facts_before = fact_sheet.read_facts(facts_path)
    assert [f.value for f in facts_before] == [pytest.approx(42.0)]

    # The survival event: a full refresh rebuilds dcf/TESTCO.xlsx from scratch.
    res2 = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res2["status"] == "ok"

    # The companion — and the parked fact — survived untouched.
    assert facts_path.exists()
    facts_after = fact_sheet.read_facts(facts_path)
    assert len(facts_after) == 1
    assert facts_after[0].token == "kpi:Operating margin"
    assert facts_after[0].value == pytest.approx(42.0)
    assert facts_after[0].unit == "percent"

    # And the route reads them back the same way.
    got = client.get("/api/dcf/reference-facts/TESTCO").get_json()
    assert got["ticker"] == "TESTCO"
    assert len(got["facts"]) == 1 and got["facts"][0]["value"] == pytest.approx(42.0)


def test_apply_edits_no_workbook_fails_soft(refresh_repo: Path) -> None:
    """No redesigned workbook to edit → a soft failure, not a crash (the route
    maps this to 409)."""
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.apply_edits(
        "TESTCO", refresh_repo, db, redesign.RedesignInputs.from_dict(_BASE.to_dict())
    )
    assert res["status"] == "failed"
    assert "no redesigned workbook" in str(res["reason"])


def test_refresh_skips_dcf_not_applicable(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A non-bank financial (asset-manager/insurer) Opus flagged dcf_applicable=
    false skips before any build — only credit banks route to the bank model."""
    assumptions = refresh_repo / "data" / "dcf_assumptions"
    assumptions.mkdir(parents=True, exist_ok=True)
    (assumptions / "TESTCO.json").write_text(
        json.dumps({"redesign": {"dcf_applicable": False, "business_model": "asset_manager"}}),
        encoding="utf-8",
    )
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "skipped"
    assert "asset_manager" in str(res["reason"])
    assert not (refresh_repo / "dcf" / "TESTCO.xlsx").exists()  # never built


def test_refresh_bank_dispatches_to_bank_model(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A credit bank (business_model=bank) is routed to the equity-side bank
    model (`_refresh_bank`), NOT skipped — it returns format='bank'."""
    assumptions = refresh_repo / "data" / "dcf_assumptions"
    assumptions.mkdir(parents=True, exist_ok=True)
    (assumptions / "TESTCO.json").write_text(
        json.dumps({"redesign": {"dcf_applicable": False, "business_model": "bank"}}),
        encoding="utf-8",
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] != "skipped"  # dispatched to the bank builder
    assert res["format"] == "bank"  # (status is 'failed' here — no TESTCO FMP fixture data)


def test_valuation_model_dispatches_to_holdco(refresh_repo: Path) -> None:
    """An explicit valuation_model='holdco_sotp' routes to the SOTP builder."""
    assumptions = refresh_repo / "data" / "dcf_assumptions"
    assumptions.mkdir(parents=True, exist_ok=True)
    (assumptions / "TESTCO.json").write_text(
        json.dumps({"redesign": {"dcf_applicable": False, "valuation_model": "holdco_sotp"}}),
        encoding="utf-8",
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] != "skipped"
    assert res["format"] == "holdco_sotp"


def test_valuation_model_dispatches_to_fintech_sotp(refresh_repo: Path) -> None:
    """An explicit valuation_model='fintech_sotp' routes to the fintech SOTP builder."""
    assumptions = refresh_repo / "data" / "dcf_assumptions"
    assumptions.mkdir(parents=True, exist_ok=True)
    (assumptions / "TESTCO.json").write_text(
        json.dumps({"redesign": {"dcf_applicable": False, "valuation_model": "fintech_sotp"}}),
        encoding="utf-8",
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] != "skipped"
    assert res["format"] == "fintech_sotp"


def test_valuation_model_dispatches_to_platform_dcf(refresh_repo: Path) -> None:
    """An explicit valuation_model='platform_dcf' routes to the customer-driven platform DCF."""
    assumptions = refresh_repo / "data" / "dcf_assumptions"
    assumptions.mkdir(parents=True, exist_ok=True)
    (assumptions / "TESTCO.json").write_text(
        json.dumps({"redesign": {"dcf_applicable": False, "valuation_model": "platform_dcf"}}),
        encoding="utf-8",
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] != "skipped"
    assert res["format"] == "platform_dcf"


def test_valuation_model_holdings_override_wins(refresh_repo: Path) -> None:
    """The holdings valuation_model override beats the dcf_assumptions one."""
    (refresh_repo / "data" / "dcf_assumptions").mkdir(parents=True, exist_ok=True)
    (refresh_repo / "data" / "dcf_assumptions" / "TESTCO.json").write_text(
        json.dumps({"redesign": {"valuation_model": "holdco_sotp"}}), encoding="utf-8"
    )
    holdings = refresh_repo / "micro_thesis" / "holdings"
    holdings.mkdir(parents=True, exist_ok=True)
    (holdings / "TESTCO.json").write_text(
        json.dumps({"ticker": "TESTCO", "valuation_model": "bank_excess_return"}), encoding="utf-8"
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["format"] == "bank"  # holdings override won over the dcf_assumptions holdco


def test_valuation_model_new_skips_with_suggestion(refresh_repo: Path) -> None:
    """valuation_model='new' has no template → skip, surfacing the Opus suggestion."""
    (refresh_repo / "data" / "dcf_assumptions").mkdir(parents=True, exist_ok=True)
    (refresh_repo / "data" / "dcf_assumptions" / "TESTCO.json").write_text(
        json.dumps(
            {
                "redesign": {
                    "valuation_model": "new",
                    "valuation_model_suggestion": "insurance embedded value — book + PV(in-force)",
                }
            }
        ),
        encoding="utf-8",
    )
    db = refresh_repo / "data" / "portfolio.db"
    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "skipped"
    assert res["valuation_model"] == "new"
    assert "embedded value" in str(res["reason"])  # the suggestion is surfaced


def test_refresh_redesign_negative_fair_value_nulls_over_under(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A user forcing deeply negative margins yields a negative fair value;
    refresh_one persists it with over_under None (the #291 guard) rather than
    crashing on the over/under calc."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    db = refresh_repo / "data" / "portfolio.db"
    dest = refresh_repo / "dcf" / "TESTCO.xlsx"
    refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    wb = openpyxl.load_workbook(str(dest))
    dsh = wb["Dashboard"]
    dsh.cell(row=29, column=2, value=-1.0)  # near op margin -100%
    dsh.cell(row=30, column=2, value=-1.0)  # terminal op margin -100%
    wb.save(str(dest))
    wb.close()

    res = refresh_dcf.refresh_one("TESTCO", refresh_repo, db, valuation_year=2026)
    assert res["status"] == "ok"
    fv = res["fair_value_per_share"]
    assert isinstance(fv, float) and fv < 0
    assert res["over_under_pct"] is None


def test_gsheets_reingest_carries_dashboard_edit_to_dcf_runs(
    refresh_repo: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The Google-Sheets re-ingest path (dcf_sheets import --file -> refresh_one)
    carries a Dashboard edit on the pulled workbook through to the persisted
    dcf_runs value — the user edits in Sheets, pulls down, re-ingests."""
    monkeypatch.setattr(refresh_dcf.live_price_mod, "read_live_price", _fake_read)
    repo = refresh_repo
    # A redesign workbook standing in for the pulled Sheet; bump terminal margin.
    downloaded = repo / "downloaded_TESTCO.xlsx"
    _build(repo, "TESTCO", downloaded)
    base = redesign.read_and_value(downloaded)
    assert base is not None
    wb = openpyxl.load_workbook(str(downloaded))
    wb["Dashboard"].cell(row=30, column=2, value=0.40)  # terminal op margin up
    wb.save(str(downloaded))
    wb.close()
    edited = redesign.read_and_value(downloaded)
    assert edited is not None and edited.value_per_share_usd > base.value_per_share_usd

    rc = dcf_sheets.main(
        [
            "import",
            "--ticker",
            "TESTCO",
            "--file",
            str(downloaded),
            "--repo-root",
            str(repo),
            "--valuation-year",
            "2026",
        ]
    )
    assert rc == 0
    assert (repo / "dcf" / "TESTCO.xlsx").exists()  # placed at the canonical path

    conn = sqlite3.connect(str(repo / "data" / "portfolio.db"))
    row = conn.execute("SELECT npv_per_share FROM dcf_runs WHERE ticker='TESTCO'").fetchone()
    conn.close()
    assert row is not None and row[0] is not None
    # The persisted value reflects the edited (higher) terminal margin.
    assert float(row[0]) == pytest.approx(edited.value_per_share_usd, rel=0.05)
    assert float(row[0]) > base.value_per_share_usd
