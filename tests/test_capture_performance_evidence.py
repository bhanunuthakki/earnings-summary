from __future__ import annotations

import json
import shlex
import shutil
import subprocess
import sys
from pathlib import Path

import pytest

import quality.performance as performance
from execution import benchmark_performance_workload as workload
from execution.capture_performance_evidence import main as capture_cli
from quality.performance import (
    COHORT_REGISTRY,
    CausalRunEnvelope,
    FrozenPerformanceCohort,
    capture_performance_evidence,
)


def _envelope_command() -> str:
    payload = {
        "sql_statements": 20,
        "rows": 20,
        "elapsed_seconds": 0.02,
        "peak_rss_bytes": 1024,
        "alembic_revision": "head",
        "query_plan_sha256": "a" * 64,
        "connection_role": "request_scoped_read",
        "stage": "route-render",
        "revision": "current",
    }
    return shlex.join([sys.executable, "-c", f"import json;print(json.dumps({payload!r}))"])


def _source_analysis_fixture(tmp_path: Path) -> tuple[Path, str, str]:
    """Build a tiny real Git repository for fast immutable-source tests."""
    root = tmp_path / "source-analysis-repo"
    (root / "src" / "quality").mkdir(parents=True)
    (root / "execution").mkdir()
    for relative in (
        "src/quality/__init__.py",
        "src/quality/duplicates.py",
        "execution/analyze_code_duplicates.py",
        "execution/sqlite_bootstrap.py",
    ):
        source = Path.cwd() / relative
        destination = root / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source, destination)
    sample = root / "src" / "sample.py"
    sample.write_text(
        "def fixture_function(value: int) -> int:\n    return value + 1\n", encoding="utf-8"
    )
    subprocess.run(["git", "-C", str(root), "init", "--quiet"], check=True)
    subprocess.run(
        ["git", "-C", str(root), "config", "user.email", "audit@example.invalid"], check=True
    )
    subprocess.run(["git", "-C", str(root), "config", "user.name", "Audit Fixture"], check=True)
    subprocess.run(["git", "-C", str(root), "add", "."], check=True)
    subprocess.run(
        ["git", "-C", str(root), "commit", "--quiet", "-m", "fixture baseline"], check=True
    )
    baseline = subprocess.check_output(
        ["git", "-C", str(root), "rev-parse", "HEAD"], text=True
    ).strip()
    sample.write_text(
        "def fixture_function(value: int) -> int:\n    return value + 2\n", encoding="utf-8"
    )
    subprocess.run(["git", "-C", str(root), "add", str(sample)], check=True)
    subprocess.run(
        ["git", "-C", str(root), "commit", "--quiet", "-m", "fixture current"], check=True
    )
    current = subprocess.check_output(
        ["git", "-C", str(root), "rev-parse", "HEAD"], text=True
    ).strip()
    return root, baseline, current


def test_route_cohort_is_frozen_and_has_causal_contract() -> None:
    cohort = FrozenPerformanceCohort(
        cohort="route_cold_warm",
        declared_command=_envelope_command(),
        route_count=20,
        route_names=tuple(f"route-{i}" for i in range(20)),
    )
    receipt = capture_performance_evidence(
        Path.cwd(),
        cohort,
        provenance="mac_guidance",
        baseline_revision="base",
        current_revision="current",
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.baseline.timing.count == 7
    assert receipt.causal_evidence.query_plan_sha256 == "a" * 64
    assert receipt.paired_identity is False
    assert FrozenPerformanceCohort.model_config.get("frozen") is True


def test_missing_cohort_companions_hold() -> None:
    cohort = FrozenPerformanceCohort(cohort="migrations", declared_command="")
    receipt = capture_performance_evidence(
        Path.cwd(),
        cohort,
        provenance="mac_guidance",
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.baseline.adaptive_verdict == "hold"
    assert any("alembic_revision" in reason for reason in receipt.baseline.hold_reasons)


def test_registry_is_canonical_and_invalid_envelope_holds() -> None:
    assert set(COHORT_REGISTRY) == {
        "integrity",
        "migrations",
        "route_cold_warm",
        "dcf",
        "source_analysis",
        "ci",
    }
    cohort = COHORT_REGISTRY["dcf"]
    bad = capture_performance_evidence(
        Path.cwd(),
        cohort,
        samples=7,
        provenance="mac_guidance",
        baseline_revision="base",
        current_revision="current",
    )
    assert bad.baseline.status == "HOLD"
    assert "paired execution provenance" in " ".join(bad.baseline.hold_reasons)


def test_forged_sidecar_and_revision_labels_can_never_pass() -> None:
    """Self-reported metrics are not proof of paired revision execution."""
    receipt = capture_performance_evidence(
        Path.cwd(),
        FrozenPerformanceCohort(
            cohort="ci",
            declared_command=_envelope_command(),
        ),
        provenance="approved_windows_production_shaped",
        baseline_revision="base",
        current_revision="current",
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.paired_identity is False
    assert any("revision-aware harness" in reason for reason in receipt.baseline.hold_reasons)


def test_cli_has_no_caller_selected_workload_or_sidecar() -> None:
    """The CLI must select only the frozen registry workload."""
    with pytest.raises(SystemExit):
        capture_cli(
            [
                "--cohort",
                "ci",
                "--output",
                ".tmp/quality/forged.json",
                "--provenance",
                "mac_guidance",
                "--command",
                "python -c 'print(\"fake\")'",
            ]
        )


def test_source_analysis_runs_real_paired_revisions(tmp_path: Path) -> None:
    root, baseline, current = _source_analysis_fixture(tmp_path)
    receipt = capture_performance_evidence(
        root,
        COHORT_REGISTRY["source_analysis"],
        samples=7,
        provenance="mac_guidance",
        baseline_revision=baseline,
        current_revision=current,
        timeout_seconds=30,
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.paired_identity is True
    assert len(receipt.causal_runs) == 16
    assert {run.revision for run in receipt.causal_runs} == {baseline, current}
    assert all(run.rows > 0 for run in receipt.causal_runs)
    assert all(run.peak_rss_bytes == 0 for run in receipt.causal_runs)
    assert all(run.rss_semantics == "unavailable" for run in receipt.causal_runs)
    assert all(run.stage == "source-analysis" for run in receipt.causal_runs)
    assert all(sample.label == "cold" for sample in receipt.baseline.timing_samples)
    assert receipt.baseline.source_analysis is not None
    assert receipt.baseline.source_analysis.cache_disposition == "no-cache"
    assert receipt.baseline.source_analysis.cache_hits == 0
    assert receipt.baseline.source_analysis.cache_misses == sum(
        run.cache_misses or 0 for run in receipt.causal_runs
    )
    assert receipt.baseline.source_analysis.parsed_once is True
    assert receipt.baseline.source_analysis.warmup_count == 2
    assert len(receipt.baseline.source_analysis.baseline_warmup_seconds) == 1
    assert len(receipt.baseline.source_analysis.current_warmup_seconds) == 1
    assert receipt.baseline.source_analysis.paired_delta_bootstrap_ci_95 is not None
    assert sum(run.phase == "warmup" for run in receipt.causal_runs) == 2
    assert all(run.cache_state == "no-cache" for run in receipt.causal_runs)
    assert len(receipt.baseline.hold_reasons) == len(set(receipt.baseline.hold_reasons))


def test_source_analysis_uses_collector_pinned_scanner_for_malicious_revision(
    tmp_path: Path,
) -> None:
    root, baseline, current = _source_analysis_fixture(tmp_path)
    scanner = root / "execution" / "analyze_code_duplicates.py"
    scanner.write_text(
        "import json, sys\n"
        "from pathlib import Path\n"
        "out = Path(sys.argv[sys.argv.index('--out') + 1])\n"
        "out.parent.mkdir(parents=True, exist_ok=True)\n"
        "out.write_text(json.dumps({'files_scanned': 999, 'scoped_revision': "
        "sys.argv[sys.argv.index('--revision') + 1], 'commit_hash': 'forged', "
        "'scanner_hash': 'forged'}))\n",
        encoding="utf-8",
    )
    subprocess.run(["git", "-C", str(root), "add", str(scanner)], check=True)
    subprocess.run(
        ["git", "-C", str(root), "commit", "--quiet", "-m", "malicious scanner"], check=True
    )
    current = subprocess.check_output(
        ["git", "-C", str(root), "rev-parse", "HEAD"], text=True
    ).strip()
    receipt = getattr(performance, "_paired_source_analysis")(
        root,
        COHORT_REGISTRY["source_analysis"],
        samples=7,
        provenance="mac_guidance",
        baseline_revision=baseline,
        current_revision=current,
        timeout_seconds=30,
        config_paths=None,
    )
    assert receipt.paired_identity is True
    assert all(run.rows != 999 for run in receipt.causal_runs)
    assert receipt.baseline.source_analysis is not None
    assert receipt.baseline.source_analysis.trusted_scanner_sha256
    assert receipt.baseline.source_analysis.trusted_scanner_wrapper_sha256


def test_public_source_analysis_rejects_noncanonical_command_without_execution(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    marker = tmp_path / "executed"
    command = shlex.join(
        [sys.executable, "-c", f"from pathlib import Path; Path({str(marker)!r}).touch()"]
    )
    cohort = FrozenPerformanceCohort(cohort="source_analysis", declared_command=command)

    def fail_if_launched(*args: object, **kwargs: object) -> object:
        raise AssertionError("noncanonical source workload must not execute")

    monkeypatch.setattr(performance.subprocess, "run", fail_if_launched)
    receipt = capture_performance_evidence(
        tmp_path,
        cohort,
        provenance="mac_guidance",
        baseline_revision="base",
        current_revision="current",
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.baseline.timing.count == 0
    assert not marker.exists()
    assert any("frozen registry" in reason for reason in receipt.baseline.hold_reasons)


def test_canonical_ci_returns_typed_hold_without_subprocess(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    def fail_if_launched(*args: object, **kwargs: object) -> object:
        raise AssertionError("CI protocol template must not execute as a workload")

    monkeypatch.setattr(performance.subprocess, "run", fail_if_launched)
    receipt = capture_performance_evidence(
        tmp_path,
        COHORT_REGISTRY["ci"],
        provenance="mac_guidance",
        baseline_revision="base",
        current_revision="current",
    )
    assert receipt.baseline.status == "HOLD"
    assert receipt.baseline.timing.count == 0
    assert any(
        "collect_paired_ci_performance.py" in reason for reason in receipt.baseline.hold_reasons
    )


def test_real_workload_cli_emits_one_envelope_and_no_checkout_database(
    capsys: pytest.CaptureFixture[str],
) -> None:
    checkout_db = Path.cwd() / "data" / "portfolio.db"
    existed = checkout_db.exists()
    assert workload.main(["--workload", "integrity", "--repo-root", str(Path.cwd())]) == 0
    captured = capsys.readouterr()
    assert len(captured.out.splitlines()) == 1
    envelope = CausalRunEnvelope.model_validate_json(captured.out)
    assert envelope.stage == "integrity"
    assert envelope.rss_semantics == "process_high_water"
    assert envelope.rows >= 0
    assert envelope.sql_statements > 0
    assert envelope.alembic_invocations == 1
    assert envelope.migration_elapsed_seconds is not None
    assert envelope.migration_elapsed_seconds > 0
    assert envelope.schema_object_count is not None
    assert envelope.schema_object_count > 0
    assert envelope.connection_role == "read"
    stderr_events = [json.loads(line) for line in captured.err.splitlines()]
    assert [event["event"] for event in stderr_events] == [
        "bha115_workload_started",
        "bha115_workload_finished",
    ]
    assert checkout_db.exists() is existed


def test_workload_cli_rejects_forged_sidecar_and_only_accepts_typed_args() -> None:
    with pytest.raises(SystemExit):
        workload.main(["--workload", "integrity", "--sql-statements", "99"])


def test_dcf_registry_runs_disposable_staged_workload() -> None:
    from execution import benchmark_dcf_workload

    result = subprocess.run(
        [
            sys.executable,
            str(Path(benchmark_dcf_workload.__file__)),
            "--repo-root",
            str(Path.cwd()),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == 0
    envelope = CausalRunEnvelope.model_validate_json(result.stdout)
    assert envelope.stage == "dcf"
    assert envelope.rss_semantics == "process_plus_children_high_water_upper_bound"
    assert envelope.artifact_sheet_names == (
        "Cover",
        "Dashboard",
        "Assumptions",
        "Color Code",
        "WACC",
        "Model",
        "Financials",
        "Consensus",
        "Valuation",
        "Sensitivity",
        "Monte Carlo",
    )
    assert envelope.rows == 11
    assert envelope.artifact_sha256 and envelope.formula_sha256 and envelope.receipt_sha256
    assert envelope.artifact_parity_sha256
    assert envelope.semantic_parity is True
    assert envelope.artifact_byte_parity is not None
    assert envelope.alembic_invocations == 1
    assert envelope.migration_elapsed_seconds is not None
    assert envelope.migration_elapsed_seconds > 0
    assert envelope.schema_object_count is not None
    assert envelope.schema_object_count > 0
    events = [json.loads(line) for line in result.stderr.splitlines()]
    assert [event["event"] for event in events] == [
        "dcf_workload_started",
        "dcf_stage_finished",
        "dcf_stage_finished",
        "dcf_stage_finished",
        "dcf_stage_finished",
        "dcf_workload_finished",
    ]
    builder_events = [event for event in events if event.get("stage") == "builder"]
    assert builder_events and builder_events[0]["peak_rss_bytes"] > 0
    parity_events = [event for event in events if event.get("stage") == "parity-build"]
    assert (
        parity_events and parity_events[0]["peak_rss_bytes"] >= builder_events[0]["peak_rss_bytes"]
    )
    assert all(
        event["rss_semantics"] == "process_plus_children_high_water_upper_bound"
        for event in events
        if "stage" in event
    )


def test_dcf_semantic_hash_includes_defined_names(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.workbook.defined_name import DefinedName

    from execution.benchmark_dcf_workload import _semantic_hash

    workbook_path = tmp_path / "defined-name.xlsx"
    workbook = Workbook()
    workbook.defined_names.add(DefinedName("ForecastInput", attr_text="'Sheet'!$A$1"))
    workbook.save(workbook_path)
    before = _semantic_hash(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook.defined_names["ForecastInput"].attr_text = "'Sheet'!$A$2"
    workbook.save(workbook_path)
    after = _semantic_hash(workbook_path)
    assert before[0] != after[0]
    assert before[1] == after[1]
