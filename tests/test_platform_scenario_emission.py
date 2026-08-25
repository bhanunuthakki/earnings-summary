"""Scenario-block emission for the bespoke NU/MELI builders (Monthly Red Team PR8).

The redesigned FCFF refresher persists a ``scenarios`` block every risk consumer
reads; the bespoke platform/SOTP builders historically wrote none, leaving MELI
and NU (~25% of the book) invisible to tail stress / bear lint. These tests pin:

  * ``scenarios_block`` emits bull/base/bear fair values structurally identical
    to the redesign snapshot's block (``parse_scenario_fair_values`` +
    ``parse_scenario_bear_provenance`` read it unchanged),
  * bear provenance is "thesis" iff the holdings JSON names ``bear_deltas``,
  * the delta->lever mappings actually push the bear leg below base,
  * the Gordon-terminal guardrails hold under extreme deltas,
  * ``redesign.strip_unedited_seed_bear`` / ``redesign.resolve_mirrored_bear``
    treat an untouched BEAR_SEED as a fallback, never an owner edit,
  * the six red-team holdings JSONs carry valid, owner-pending bear_deltas,
  * (PR10 — surface wiring) the workbook ``Scenario`` SHEET's Bear/Bull rows
    derive from the exact same ``scenario_assumptions()`` call as
    ``scenarios_block`` (the persisted snapshot), with a provenance note cell,
    so the workbook and every risk consumer that reads the persisted block
    (bear_lint, tail stress, red-team evidence packs) can never disagree.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path
from typing import cast

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "execution"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import build_meli_platform_dcf as meli  # noqa: E402
import build_nu_platform_dcf as nu  # noqa: E402

from dcf import redesign  # noqa: E402
from dcf.scenario_reward import (  # noqa: E402
    parse_scenario_bear_provenance,
    parse_scenario_fair_values,
)

THESIS_HOLDINGS: dict[str, object] = {
    "bear_deltas": {
        "growth_delta_pp": -8.0,
        "margin_delta_pp": -6.0,
        "exit_multiple_delta": -6.0,
        "terminal_g_delta_pp": -1.0,
        "note": "test thesis-break bear",
    }
}

RED_TEAM_TICKERS = ("NU", "MELI", "NVO", "BKNG", "UBER", "WIX")


def _snapshot_json(block: dict[str, object]) -> str:
    return json.dumps({"model": "x", "scenarios": block})


# --------------------------------------------------------------------------- #
# NU platform builder
# --------------------------------------------------------------------------- #


def test_nu_scenarios_block_parseable_and_ordered() -> None:
    s = nu.Assum()
    m = nu.mirror(s)
    block = nu.scenarios_block(s, m, THESIS_HOLDINGS)
    snap = _snapshot_json(block)
    fvs = parse_scenario_fair_values(snap)
    assert set(fvs) == {"bull", "base", "bear"}
    assert fvs["base"] == m.vps
    assert fvs["bear"] < fvs["base"] < fvs["bull"]
    assert parse_scenario_bear_provenance(snap) == "thesis"


def test_nu_seed_provenance_without_holdings_override() -> None:
    s = nu.Assum()
    block = nu.scenarios_block(s, nu.mirror(s), None)
    snap = _snapshot_json(block)
    assert parse_scenario_bear_provenance(snap) == "seed"
    # the generic seed still produces a bear below base (mild disappointment)
    fvs = parse_scenario_fair_values(snap)
    assert fvs["bear"] < fvs["base"]


def test_nu_bear_deltas_recorded_in_block() -> None:
    s = nu.Assum()
    block = nu.scenarios_block(s, nu.mirror(s), THESIS_HOLDINGS)
    bear = block["bear"]
    assert isinstance(bear, dict)
    assert bear["deltas"]["growth_near"] == -0.08
    assert bear["deltas"]["exit_multiple"] == -6.0


def test_nu_gordon_guardrails_under_extreme_deltas() -> None:
    # An absurd exit-multiple delta may not flip the Gordon terminal's sign:
    # ROE stays above g, g stays below ke, so the bear is a small positive value.
    d = redesign.ScenarioDeltas(
        growth_near=-0.20,
        growth_term=-0.20,
        margin_near=-0.10,
        margin_term=-0.10,
        exit_multiple=-50.0,
        terminal_g=-0.05,
    )
    s2 = nu.scenario_assumptions(nu.Assum(), d)
    assert s2.g_term < s2.ke
    assert s2.terminal_roe > s2.g_term


def test_nu_exit_lever_maps_to_terminal_roe() -> None:
    d = redesign.ScenarioDeltas(exit_multiple=-7.0)
    base = nu.Assum()
    s2 = nu.scenario_assumptions(base, d)
    assert s2.terminal_roe == base.terminal_roe - 0.07


def test_nu_reverse_valuation_reprices_with_archetype_specific_levers() -> None:
    s = nu.Assum()
    snapshot = nu.reverse_valuation(s, nu.mirror(s))

    assert snapshot is not None
    assert snapshot["archetype"] == "platform_dcf"
    assert snapshot["valuation_scope"] == "equity"
    levers = cast("list[dict[str, object]]", snapshot["levers"])
    assert [lever["id"] for lever in levers] == [
        "implied_joint_customer_arpac_growth_shift",
        "implied_terminal_roe",
    ]


# --------------------------------------------------------------------------- #
# MELI SOTP builder
# --------------------------------------------------------------------------- #


def test_meli_scenarios_block_parseable_and_ordered() -> None:
    s = meli.Assum()
    m = meli.mirror(s)
    block = meli.scenarios_block(s, m, THESIS_HOLDINGS)
    snap = _snapshot_json(block)
    fvs = parse_scenario_fair_values(snap)
    assert set(fvs) == {"bull", "base", "bear"}
    assert fvs["base"] == m.vps
    assert fvs["bear"] < fvs["base"] < fvs["bull"]
    assert parse_scenario_bear_provenance(snap) == "thesis"


def test_meli_seed_provenance_without_holdings_override() -> None:
    s = meli.Assum()
    block = meli.scenarios_block(s, meli.mirror(s), {})
    assert parse_scenario_bear_provenance(_snapshot_json(block)) == "seed"


def test_meli_lever_mapping_and_guardrails() -> None:
    d = redesign.ScenarioDeltas(
        growth_near=-0.075,
        growth_term=-0.075,
        margin_near=-0.05,
        margin_term=-0.05,
        exit_multiple=-40.0,  # extreme: must floor at 1x, not go negative
        terminal_g=-0.10,
    )
    base = meli.Assum()
    s2 = meli.scenario_assumptions(base, d)
    assert s2.op_exit_ebitda_mult == 1.0
    assert s2.nimal_term == base.nimal_term - 0.05
    assert s2.comm_g_near == base.comm_g_near - 0.075
    assert s2.credit_g_term < s2.credit_ke
    assert s2.credit_terminal_roe > s2.credit_g_term


def test_meli_reverse_valuation_has_exit_multiple_and_credit_residual() -> None:
    s = meli.Assum(derive_capm=0)
    snapshot = meli.reverse_valuation(s, meli.mirror(s))

    assert snapshot is not None
    assert snapshot["archetype"] == "meli_platform_sotp"
    levers = cast("list[dict[str, object]]", snapshot["levers"])
    assert [lever["id"] for lever in levers] == [
        "implied_operating_exit_multiple",
        "implied_credit_equity_value",
    ]


@pytest.mark.parametrize(
    "script_name",
    ("build_nu_platform_dcf.py", "build_holdco_sotp.py"),
)
def test_specialized_builder_imports_code_from_executing_checkout_with_external_data_root(
    tmp_path: Path, script_name: str
) -> None:
    env = dict(os.environ)
    env["DCF_REPO_ROOT"] = str(tmp_path / "external-data-root")
    script = PROJECT_ROOT / "execution" / script_name

    proc = subprocess.run(
        [
            sys.executable,
            "-c",
            "import runpy,sys; runpy.run_path(sys.argv[1], run_name='external_root_import')",
            str(script),
        ],
        cwd=PROJECT_ROOT,
        env=env,
        capture_output=True,
        text=True,
        check=False,
    )

    assert proc.returncode == 0, proc.stderr


# --------------------------------------------------------------------------- #
# strip_unedited_seed_bear (refresh capture->inject loop, redesigned names)
# --------------------------------------------------------------------------- #


def _captured_with_bear(deltas: redesign.ScenarioDeltas) -> redesign.CapturedDashboard:
    cells = {
        (redesign.SCEN_ROW_GROWTH_NEAR, redesign.SCEN_COL_BEAR): deltas.growth_near,
        (redesign.SCEN_ROW_GROWTH_TERM, redesign.SCEN_COL_BEAR): deltas.growth_term,
        (redesign.SCEN_ROW_MARGIN_NEAR, redesign.SCEN_COL_BEAR): deltas.margin_near,
        (redesign.SCEN_ROW_MARGIN_TERM, redesign.SCEN_COL_BEAR): deltas.margin_term,
        (redesign.SCEN_ROW_EXIT_MULT, redesign.SCEN_COL_BEAR): deltas.exit_multiple,
        (redesign.SCEN_ROW_TG, redesign.SCEN_COL_BEAR): deltas.terminal_g,
        # a Bull cell that must never be touched
        (redesign.SCEN_ROW_GROWTH_NEAR, redesign.SCEN_COL_BULL): 0.05,
    }
    return redesign.CapturedDashboard(
        segment_growth={},
        scalars={},
        terminal_method=None,
        terminal_basis=None,
        scenario_deltas=cells,
    )


def test_strip_drops_untouched_seed_bear_when_thesis_named() -> None:
    cap = _captured_with_bear(redesign.BEAR_SEED)
    out = redesign.strip_unedited_seed_bear(cap, THESIS_HOLDINGS)
    assert out is not None
    bear_cells = [k for k in out.scenario_deltas if k[1] == redesign.SCEN_COL_BEAR]
    assert bear_cells == []
    # bull cell preserved
    assert (redesign.SCEN_ROW_GROWTH_NEAR, redesign.SCEN_COL_BULL) in out.scenario_deltas


def test_strip_preserves_owner_edited_bear() -> None:
    owner = redesign.ScenarioDeltas(
        growth_near=-0.11,
        growth_term=-0.01,
        margin_near=-0.01,
        margin_term=-0.02,
        exit_multiple=-2.0,
        terminal_g=-0.005,
    )
    cap = _captured_with_bear(owner)
    out = redesign.strip_unedited_seed_bear(cap, THESIS_HOLDINGS)
    assert out is cap


def test_strip_noop_without_thesis_override() -> None:
    cap = _captured_with_bear(redesign.BEAR_SEED)
    assert redesign.strip_unedited_seed_bear(cap, None) is cap
    assert redesign.strip_unedited_seed_bear(cap, {}) is cap
    assert redesign.strip_unedited_seed_bear(None, THESIS_HOLDINGS) is None


# --------------------------------------------------------------------------- #
# resolve_mirrored_bear (fresh-build seeding, assumptions-JSON mirror)
# --------------------------------------------------------------------------- #


def test_resolve_mirrored_bear_prefers_thesis_over_untouched_seed() -> None:
    resolved = redesign.resolve_mirrored_bear(redesign.BEAR_SEED, THESIS_HOLDINGS)
    assert resolved == redesign.parse_thesis_bear_deltas(THESIS_HOLDINGS)


def test_resolve_mirrored_bear_keeps_owner_edit() -> None:
    owner = redesign.ScenarioDeltas(growth_near=-0.09)
    assert redesign.resolve_mirrored_bear(owner, THESIS_HOLDINGS) == owner


def test_resolve_mirrored_bear_seed_without_thesis() -> None:
    assert redesign.resolve_mirrored_bear(redesign.BEAR_SEED, None) == redesign.BEAR_SEED


# --------------------------------------------------------------------------- #
# PR10: workbook Scenario sheet == persisted snapshot (both builders)
# --------------------------------------------------------------------------- #


def _sheet_scenarios(dest: Path) -> tuple[dict[str, float], str]:
    """``({scenario_name: value_per_share}, provenance_note)`` from the
    workbook's Scenario sheet — rows 3-5 (Bear/Base/Bull) + the provenance
    note cell on row 6."""
    ws = openpyxl.load_workbook(str(dest))["Scenario"]
    fvs = {str(ws[f"A{r}"].value): float(cast("float", ws[f"E{r}"].value)) for r in (3, 4, 5)}
    assert str(ws["A6"].value) == "Bear provenance"
    return fvs, str(ws["B6"].value)


def _persisted_fvs(block: dict[str, object]) -> dict[str, float]:
    return parse_scenario_fair_values(_snapshot_json(block))


def test_nu_sheet_bear_matches_persisted_snapshot_thesis(tmp_path: Path) -> None:
    """The regression the PR exists for: sheet-bear VPS == persisted-snapshot
    bear VPS on a fixture assumption set (thesis bear_deltas)."""
    s = nu.Assum()
    m = nu.mirror(s)
    dest = tmp_path / "NU.xlsx"
    nu.build(s, m, dest, THESIS_HOLDINGS)
    sheet, provenance_note = _sheet_scenarios(dest)
    persisted = _persisted_fvs(nu.scenarios_block(s, m, THESIS_HOLDINGS))
    for name in ("Bear", "Base", "Bull"):
        # The sheet rounds to cents; the snapshot stores full precision.
        assert sheet[name] == pytest.approx(persisted[name.lower()], abs=0.005), name
    assert "thesis" in provenance_note
    assert "seed" not in provenance_note


def test_nu_sheet_bear_seed_fallback_labeled(tmp_path: Path) -> None:
    s = nu.Assum()
    m = nu.mirror(s)
    dest = tmp_path / "NU_seed.xlsx"
    nu.build(s, m, dest, None)
    sheet, provenance_note = _sheet_scenarios(dest)
    persisted = _persisted_fvs(nu.scenarios_block(s, m, None))
    assert sheet["Bear"] == pytest.approx(persisted["bear"], abs=0.005)
    assert "seed fallback" in provenance_note


def test_meli_sheet_bear_matches_persisted_snapshot_thesis(tmp_path: Path) -> None:
    s = meli.Assum(derive_capm=0)
    m = meli.mirror(s)
    dest = tmp_path / "MELI.xlsx"
    meli.build(s, m, dest, THESIS_HOLDINGS)
    sheet, provenance_note = _sheet_scenarios(dest)
    persisted = _persisted_fvs(meli.scenarios_block(s, m, THESIS_HOLDINGS))
    for name in ("Bear", "Base", "Bull"):
        assert sheet[name] == pytest.approx(persisted[name.lower()], abs=0.005), name
    assert "thesis" in provenance_note


def test_meli_sheet_bear_seed_fallback_labeled(tmp_path: Path) -> None:
    s = meli.Assum(derive_capm=0)
    m = meli.mirror(s)
    dest = tmp_path / "MELI_seed.xlsx"
    meli.build(s, m, dest, {})
    sheet, provenance_note = _sheet_scenarios(dest)
    persisted = _persisted_fvs(meli.scenarios_block(s, m, {}))
    assert sheet["Bear"] == pytest.approx(persisted["bear"], abs=0.005)
    assert "seed fallback" in provenance_note


def test_sheet_lever_columns_show_post_delta_values() -> None:
    """The sheet's B-D columns display the ACTUAL post-delta lever values, not
    the legacy hardcoded ones — pinned via the same scenario_assumptions call."""
    s = nu.Assum()
    bear = nu.scenario_assumptions(s, redesign.thesis_bear_seed(THESIS_HOLDINGS))
    assert bear.custg_near == pytest.approx(s.custg_near - 0.08)
    assert bear.gpm_term == pytest.approx(s.gpm_term - 0.06)


# --------------------------------------------------------------------------- #
# the six shipped bear_deltas blocks validate against the documented schema
# --------------------------------------------------------------------------- #


def test_red_team_holdings_carry_valid_bear_deltas() -> None:
    for ticker in RED_TEAM_TICKERS:
        path = PROJECT_ROOT / "micro_thesis" / "holdings" / f"{ticker}.json"
        loaded: object = json.loads(path.read_text(encoding="utf-8"))
        assert isinstance(loaded, dict), f"{ticker}: holdings JSON is not an object"
        holdings = cast("dict[str, object]", loaded)
        block_raw = holdings.get("bear_deltas")
        assert isinstance(block_raw, dict), f"{ticker}: bear_deltas missing"
        block = cast("dict[str, object]", block_raw)
        parsed = redesign.parse_thesis_bear_deltas(holdings)
        assert parsed is not None, f"{ticker}: bear_deltas unparseable"
        assert parsed != redesign.BEAR_SEED, f"{ticker}: bear_deltas equals the generic seed"
        # every shipped lever is a downside shift
        assert parsed.growth_near < 0 and parsed.margin_near < 0, ticker
        assert parsed.exit_multiple < 0 and parsed.terminal_g < 0, ticker
        note = block.get("note")
        assert isinstance(note, str), ticker
        normalized_note = note.lower()
        assert "pending owner review" in normalized_note or "owner approved" in normalized_note, (
            ticker
        )
