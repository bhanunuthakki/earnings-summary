"""Tests for the DCF reference-sheet companion workbook (capture-every-number S7).

``dcf.fact_sheet`` writes a picked fact into ``dcf/facts/<T>.xlsx`` — a companion
the refresh NEVER rebuilds, so an injected reference survives every model refresh
(the end-to-end survival proof lives in ``test_dcf_redesign``; here we cover the
pure read/write/upsert primitives in isolation).
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from dcf import fact_sheet  # noqa: E402


def _fact(token: str, value: float, **kw: object) -> fact_sheet.ReferenceFact:
    base: dict[str, object] = {
        "token": token,
        "label": token.split(":")[-1],
        "value": value,
        "unit": "percent",
        "period_end": "2025-09-30",
        "source": "fmp_key_metrics",
        "fact_id": 7,
        "captured_on": "2026-06-15",
    }
    base.update(kw)
    return fact_sheet.ReferenceFact(**base)  # type: ignore[arg-type]


def test_path_is_under_facts_subdir() -> None:
    """The companion lives in a SUBDIRECTORY so the non-recursive dcf/*.xlsx
    globs that enumerate ticker workbooks never mistake it for a ticker."""
    p = fact_sheet.facts_workbook_path(Path("/repo"), "nu")
    assert p == Path("/repo") / "dcf" / "facts" / "NU.xlsx"
    assert p.parent.name == "facts"


def test_upsert_creates_workbook_sheet_and_row(tmp_path: Path) -> None:
    path = fact_sheet.facts_workbook_path(tmp_path, "TESTCO")
    assert not path.exists()
    outcome = fact_sheet.upsert_fact(path, _fact("kpi:Operating margin", 42.0))
    assert outcome == {"action": "added", "row": 2, "count": 1}
    assert path.exists()

    wb = openpyxl.load_workbook(str(path))
    assert fact_sheet.SHEET_NAME in wb.sheetnames
    ws = wb[fact_sheet.SHEET_NAME]
    assert ws.cell(row=1, column=1).value == "Metric"  # header present
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=2, column=2).value == "kpi:Operating margin"
    assert ws.cell(row=2, column=3).value == 42.0
    wb.close()


def test_round_trip_read(tmp_path: Path) -> None:
    path = fact_sheet.facts_workbook_path(tmp_path, "TESTCO")
    fact_sheet.upsert_fact(path, _fact("fin:revenue", 1234.5, unit="millions", fact_id=None))
    facts = fact_sheet.read_facts(path)
    assert len(facts) == 1
    f = facts[0]
    assert f.token == "fin:revenue"
    assert f.value == 1234.5
    assert f.unit == "millions"
    assert f.period_end == "2025-09-30"
    assert f.fact_id is None


def test_distinct_tokens_append_as_separate_rows(tmp_path: Path) -> None:
    path = fact_sheet.facts_workbook_path(tmp_path, "TESTCO")
    fact_sheet.upsert_fact(path, _fact("kpi:Operating margin", 42.0))
    out2 = fact_sheet.upsert_fact(path, _fact("fin:revenue", 1000.0, unit="millions"))
    assert out2 == {"action": "added", "row": 3, "count": 2}
    tokens = [f.token for f in fact_sheet.read_facts(path)]
    assert tokens == ["kpi:Operating margin", "fin:revenue"]


def test_same_token_upserts_in_place(tmp_path: Path) -> None:
    """Re-injecting the same metric UPDATES its row (newer value/period), never
    stacking duplicates."""
    path = fact_sheet.facts_workbook_path(tmp_path, "TESTCO")
    fact_sheet.upsert_fact(path, _fact("kpi:ROE", 18.0, period_end="2025-06-30"))
    out2 = fact_sheet.upsert_fact(path, _fact("kpi:ROE", 21.5, period_end="2025-09-30"))
    assert out2 == {"action": "updated", "row": 2, "count": 1}
    facts = fact_sheet.read_facts(path)
    assert len(facts) == 1
    assert facts[0].value == 21.5
    assert facts[0].period_end == "2025-09-30"


def test_read_missing_file_is_empty(tmp_path: Path) -> None:
    assert fact_sheet.read_facts(tmp_path / "dcf" / "facts" / "NOPE.xlsx") == []


def test_read_workbook_without_sheet_is_empty(tmp_path: Path) -> None:
    other = tmp_path / "other.xlsx"
    openpyxl.Workbook().save(str(other))
    assert fact_sheet.read_facts(other) == []
