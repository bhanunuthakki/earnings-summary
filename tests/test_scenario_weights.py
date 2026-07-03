"""PR D — per-name scenario weights in the workbook + assumptions JSON.

Covers the plumbing that carries the LLM/owner Bull/Base/Bear weights from the
JSON through the workbook to the dcf_runs snapshot:
  * redesign.normalize_scenario_weights + from_dict/to_dict round-trip;
  * capture_dashboard/inject_dashboard preserve the weight cells (owner edit wins);
  * refresh_dcf._mirror_scenario_prior_weights (JSON mirror + owner-provenance flip);
  * refresh_dcf._scenario_prior_snapshot (the snapshot block);
  * set_scenario_priors.write_scenario_prior_to_json (the producer's JSON writer).
"""

from __future__ import annotations

import dataclasses
import json
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "execution"))

import refresh_dcf  # noqa: E402
import set_scenario_priors as ssp  # noqa: E402

from dcf import redesign  # noqa: E402
from dcf.scenario_prior import ScenarioPrior  # noqa: E402

_BASE = redesign.RedesignInputs(
    segments=("Cloud",),
    base_revenue_by_segment={"Cloud": 1000.0},
    near_growth_by_segment={"Cloud": 0.10},
    terminal_growth_by_segment={"Cloud": 0.03},
    near_op_margin=0.20,
    terminal_op_margin=0.25,
    tax_rate=0.24,
    capex_2026_m=60.0,
    terminal_capex_da=1.05,
    da_ratio=0.05,
    consensus_years=5,
    wacc=0.09,
    beta=1.2,
    risk_free_rate=0.043,
    equity_risk_premium=0.045,
    cost_of_debt=0.045,
    terminal_method="Exit multiple",
    terminal_basis="EV/EBITDA",
    exit_multiple=12.0,
    terminal_growth_g=0.03,
    current_price=50.0,
    cash_m=100.0,
    total_debt_m=200.0,
    diluted_shares_m=100.0,
    fx_to_usd=1.0,
)


# --------------------------------------------------------------------------- #
# 1. normalize + from_dict/to_dict round-trip
# --------------------------------------------------------------------------- #
def test_normalize_scenario_weights() -> None:
    assert redesign.normalize_scenario_weights(0.3, 0.5, 0.2) == (0.3, 0.5, 0.2)
    # Renormalizes a non-summing edit.
    b, m, r = redesign.normalize_scenario_weights(30, 50, 20)
    assert (round(b, 4), round(m, 4), round(r, 4)) == (0.3, 0.5, 0.2)
    # Degenerate (negative / zero-sum) → symmetric default.
    assert redesign.normalize_scenario_weights(-1, 0.5, 0.5) == (0.25, 0.5, 0.25)
    assert redesign.normalize_scenario_weights(0, 0, 0) == (0.25, 0.5, 0.25)


def test_inputs_default_weights_are_global() -> None:
    assert (_BASE.weight_bull, _BASE.weight_base, _BASE.weight_bear) == (0.25, 0.50, 0.25)


def test_from_dict_to_dict_round_trips_weights() -> None:
    edited = dataclasses.replace(_BASE, weight_bull=0.15, weight_base=0.50, weight_bear=0.35)
    d = edited.to_dict()
    assert d["weight_bear"] == 0.35
    back = redesign.RedesignInputs.from_dict(d)
    assert (back.weight_bull, back.weight_base, back.weight_bear) == (0.15, 0.50, 0.35)


def test_from_dict_normalizes_and_defaults_weights() -> None:
    d = _BASE.to_dict()
    d["weight_bull"], d["weight_base"], d["weight_bear"] = 15.0, 50.0, 35.0  # percent form
    back = redesign.RedesignInputs.from_dict(d)
    assert round(back.weight_bear, 4) == 0.35
    # Missing weights → global default.
    d2 = _BASE.to_dict()
    for k in ("weight_bull", "weight_base", "weight_bear"):
        d2.pop(k, None)
    back2 = redesign.RedesignInputs.from_dict(d2)
    assert (back2.weight_bull, back2.weight_base, back2.weight_bear) == (0.25, 0.50, 0.25)


# --------------------------------------------------------------------------- #
# 2. capture/inject the weight cells (owner edit wins)
# --------------------------------------------------------------------------- #
def _minimal_redesign_wb(
    path: Path, *, base: float | None, bull: float | None, bear: float | None
) -> None:
    """A workbook with just the redesign marker sheets + the weight cells — enough
    for is_redesign_format + capture/inject (no financials, no read_inputs)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = redesign.DASHBOARD_SHEET
    for name in (redesign.MODEL_SHEET, redesign.FINANCIALS_SHEET, redesign.CONSENSUS_SHEET):
        wb.create_sheet(name)
    dsh = wb[redesign.DASHBOARD_SHEET]
    if base is not None:
        dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_WEIGHT_BASE, value=base)
    if bull is not None:
        dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_BULL, value=bull)
    if bear is not None:
        dsh.cell(row=redesign.SCEN_WEIGHTS_ROW, column=redesign.SCEN_COL_BEAR, value=bear)
    wb.save(str(path))


def test_capture_reads_weight_cells(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _minimal_redesign_wb(wb, base=0.5, bull=0.3, bear=0.2)
    cap = redesign.capture_dashboard(wb)
    assert cap is not None
    assert cap.scenario_weights == {"bull": 0.3, "base": 0.5, "bear": 0.2}


def test_capture_absent_weights_stays_empty(tmp_path: Path) -> None:
    wb = tmp_path / "wb.xlsx"
    _minimal_redesign_wb(wb, base=None, bull=None, bear=None)
    cap = redesign.capture_dashboard(wb)
    assert cap is not None
    assert cap.scenario_weights == {}


def test_inject_restores_captured_weights(tmp_path: Path) -> None:
    src = tmp_path / "src.xlsx"
    _minimal_redesign_wb(src, base=0.45, bull=0.40, bear=0.15)
    cap = redesign.capture_dashboard(src)
    assert cap is not None

    # A freshly-"rebuilt" workbook seeded with the global default...
    rebuilt = tmp_path / "rebuilt.xlsx"
    _minimal_redesign_wb(rebuilt, base=0.5, bull=0.25, bear=0.25)
    redesign.inject_dashboard(rebuilt, cap, current_price=None)
    # ...gets the owner's captured weights back.
    recap = redesign.capture_dashboard(rebuilt)
    assert recap is not None
    assert recap.scenario_weights == {"bull": 0.40, "base": 0.45, "bear": 0.15}


def test_capture_from_inputs_carries_weights() -> None:
    inp = dataclasses.replace(_BASE, weight_bull=0.2, weight_base=0.5, weight_bear=0.3)
    cap = redesign.capture_from_inputs(inp)
    assert cap.scenario_weights == {"bull": 0.2, "base": 0.5, "bear": 0.3}


# --------------------------------------------------------------------------- #
# 3. refresh_dcf: JSON mirror + owner-provenance flip
# --------------------------------------------------------------------------- #
def test_mirror_skips_global_with_no_block() -> None:
    rd: dict[str, object] = {}
    refresh_dcf._mirror_scenario_prior_weights(rd, _BASE)  # pyright: ignore[reportPrivateUsage]
    assert "scenario_prior" not in rd  # global default + no block → keep JSON clean


def test_mirror_writes_owner_weights_off_default() -> None:
    rd: dict[str, object] = {}
    inp = dataclasses.replace(_BASE, weight_bull=0.15, weight_base=0.50, weight_bear=0.35)
    refresh_dcf._mirror_scenario_prior_weights(rd, inp)  # pyright: ignore[reportPrivateUsage]
    sp = rd["scenario_prior"]
    assert isinstance(sp, dict)
    assert sp["bear_weight"] == 0.35 and sp["set_by"] == "owner"


def test_mirror_preserves_llm_rationale_and_flips_provenance_on_edit() -> None:
    # An LLM block on file; the workbook weights now differ → provenance flips to
    # owner but the rationale is preserved.
    rd: dict[str, object] = {
        "scenario_prior": {
            "bull_weight": 0.25,
            "base_weight": 0.50,
            "bear_weight": 0.25,
            "rationale": "fragile thesis",
            "set_by": "llm",
        }
    }
    inp = dataclasses.replace(_BASE, weight_bull=0.15, weight_base=0.50, weight_bear=0.35)
    refresh_dcf._mirror_scenario_prior_weights(rd, inp)  # pyright: ignore[reportPrivateUsage]
    sp = rd["scenario_prior"]
    assert isinstance(sp, dict)
    assert sp["set_by"] == "owner"  # owner edited away from the LLM's weights
    assert sp["rationale"] == "fragile thesis"  # narrative preserved
    assert sp["bear_weight"] == 0.35


def test_mirror_keeps_llm_provenance_when_unchanged() -> None:
    rd: dict[str, object] = {
        "scenario_prior": {
            "bull_weight": 0.15,
            "base_weight": 0.50,
            "bear_weight": 0.35,
            "rationale": "fragile",
            "set_by": "llm",
        }
    }
    inp = dataclasses.replace(_BASE, weight_bull=0.15, weight_base=0.50, weight_bear=0.35)
    refresh_dcf._mirror_scenario_prior_weights(rd, inp)  # pyright: ignore[reportPrivateUsage]
    sp = rd["scenario_prior"]
    assert isinstance(sp, dict)
    assert sp["set_by"] == "llm"  # weights match → still the LLM's call


def test_scenario_prior_snapshot_block() -> None:
    inp = dataclasses.replace(_BASE, weight_bull=0.15, weight_base=0.50, weight_bear=0.35)
    meta: dict[str, object] = {"rationale": "high multiple", "set_by": "llm", "as_of": "2026-07-02"}
    block = refresh_dcf._scenario_prior_snapshot(inp, meta)  # pyright: ignore[reportPrivateUsage]
    assert block["weights"] == {"bull": 0.15, "base": 0.50, "bear": 0.35}
    assert block["rationale"] == "high multiple"
    assert block["set_by"] == "llm"
    # No meta + global weights → global provenance, empty rationale.
    block2 = refresh_dcf._scenario_prior_snapshot(_BASE, None)  # pyright: ignore[reportPrivateUsage]
    assert block2["set_by"] == "global"
    assert block2["rationale"] == ""


# --------------------------------------------------------------------------- #
# 4. producer JSON writer (owner wins / global skip / no block)
# --------------------------------------------------------------------------- #
def _llm_prior() -> ScenarioPrior:
    return ScenarioPrior(0.15, 0.50, 0.35, "fragile thesis", "llm", "2026-07-02")


def test_writer_writes_into_existing_redesign_block(tmp_path: Path) -> None:
    path = tmp_path / "NU.json"
    path.write_text(json.dumps({"redesign": {"segments": {}}}), encoding="utf-8")
    status = ssp.write_scenario_prior_to_json(path, _llm_prior())
    assert status == "written"
    data = json.loads(path.read_text(encoding="utf-8"))
    sp = data["redesign"]["scenario_prior"]
    assert sp["bear_weight"] == 0.35 and sp["rationale"] == "fragile thesis"
    assert data["redesign"]["segments"] == {}  # untouched


def test_writer_respects_owner(tmp_path: Path) -> None:
    path = tmp_path / "NU.json"
    path.write_text(
        json.dumps({"redesign": {"scenario_prior": {"set_by": "owner", "bear_weight": 0.4}}}),
        encoding="utf-8",
    )
    assert ssp.write_scenario_prior_to_json(path, _llm_prior()) == "skipped_owner"
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["redesign"]["scenario_prior"]["bear_weight"] == 0.4  # owner's kept


def test_writer_skips_global_and_missing_block(tmp_path: Path) -> None:
    path = tmp_path / "NU.json"
    path.write_text(json.dumps({"redesign": {}}), encoding="utf-8")
    g = ScenarioPrior(0.25, 0.50, 0.25, "sym", "global", "2026-07-02")
    assert ssp.write_scenario_prior_to_json(path, g) == "skipped_global"
    # No redesign block → never fabricated.
    path.write_text(json.dumps({"other": 1}), encoding="utf-8")
    assert ssp.write_scenario_prior_to_json(path, _llm_prior()) == "skipped_no_redesign_block"
    # No file at all.
    assert ssp.write_scenario_prior_to_json(tmp_path / "missing.json", _llm_prior()) == (
        "skipped_no_redesign_block"
    )


def test_writer_overwrite_owner_flag(tmp_path: Path) -> None:
    path = tmp_path / "NU.json"
    path.write_text(
        json.dumps({"redesign": {"scenario_prior": {"set_by": "owner", "bear_weight": 0.4}}}),
        encoding="utf-8",
    )
    status = ssp.write_scenario_prior_to_json(path, _llm_prior(), respect_owner=False)
    assert status == "written"
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["redesign"]["scenario_prior"]["set_by"] == "llm"
