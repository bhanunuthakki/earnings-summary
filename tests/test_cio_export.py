"""Tests for the Personal-CIO xlsx export (src/dashboard/cio_export.py).

The CIO content (alerts → queued actions → thesis ledger) is surfaced as HTML
but previously had no spreadsheet export. The exporter reads the substrate via
the stores and writes a three-sheet workbook; it must produce a valid file even
when the substrate is empty.
"""

from __future__ import annotations

import json
from collections.abc import Callable
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

import dashboard.cio_export as cio_export
from alerts import QueuedActionRow, compute_signature_sha, fire_alert, queue_action
from dashboard.cio_export import export_cio_workbook
from user_state.ledger import append_entry

_PRIOR_HEAD = "0059_kpi_facts_restatement"
_FIRED_AT = datetime(2026, 6, 1, 12, 0)  # naive-UTC, per the repo convention


@pytest.fixture
def db_path(tmp_path: Path, migrated_db: Callable[..., Path]) -> Path:
    return migrated_db(tmp_path / "portfolio.db", stamp=_PRIOR_HEAD)


def test_export_empty_substrate_is_header_only(tmp_path: Path, db_path: Path) -> None:
    out = export_cio_workbook(tmp_path / "cio.xlsx", db_path=db_path)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Alerts", "Queued Actions", "Thesis Ledger"]
    # Each sheet has only its header row when the substrate is empty.
    assert wb["Alerts"].max_row == 1
    assert wb["Queued Actions"].max_row == 1
    assert wb["Thesis Ledger"].max_row == 1
    assert [c.value for c in wb["Alerts"][1]][:3] == ["id", "fired_at", "ticker"]


def test_export_populated_substrate(tmp_path: Path, db_path: Path) -> None:
    alert = fire_alert(
        ticker="NU",
        trigger_kind="earnings_tone",
        fired_at=_FIRED_AT,
        evidence_json=json.dumps({"summary": "NIM contracted 100bps QoQ"}),
        signature_sha=compute_signature_sha("earnings_tone", "NU", {"k": "v"}),
        db_path=db_path,
    )
    queue_action(
        alert_id=alert.id,
        action_kind="thesis_update",
        payload={"note": "watch NIM"},
        db_path=db_path,
    )
    append_entry(
        ticker="NU",
        entry_kind="thesis_update",
        body="Machine summary: NIM under pressure",
        source_alert_id=alert.id,
        db_path=db_path,
    )
    append_entry(
        ticker="NU",
        entry_kind="thesis_update",
        body="Owner thesis: NIM under pressure",
        db_path=db_path,
    )

    wb = load_workbook(export_cio_workbook(tmp_path / "cio.xlsx", db_path=db_path))

    alerts_row = [c.value for c in wb["Alerts"][2]]
    assert "NU" in alerts_row
    assert "earnings_tone" in alerts_row
    assert any("NIM contracted" in str(c) for c in alerts_row)  # summary flattened in

    actions_row = [c.value for c in wb["Queued Actions"][2]]
    assert "thesis_update" in actions_row
    assert any("watch NIM" in str(c) for c in actions_row)  # payload serialized in

    ledger_row = [c.value for c in wb["Thesis Ledger"][2]]
    assert any("Owner thesis: NIM under pressure" in str(c) for c in ledger_row)
    assert wb["Thesis Ledger"].max_row == 2
    assert not any(
        "Machine summary: NIM under pressure" in str(cell.value)
        for row in wb["Thesis Ledger"].iter_rows()
        for cell in row
    )


def test_export_uses_single_batched_actions_lookup(
    tmp_path: Path, db_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    first = fire_alert(
        ticker="AAA",
        trigger_kind="kpi_inflection",
        fired_at=_FIRED_AT,
        evidence_json=json.dumps({"summary": "first summary"}),
        signature_sha=compute_signature_sha("kpi_inflection", "AAA", {"k": "first"}),
        db_path=db_path,
    )
    second = fire_alert(
        ticker="BBB",
        trigger_kind="earnings_tone",
        fired_at=_FIRED_AT + timedelta(days=1),
        evidence_json=json.dumps({"summary": "second summary"}),
        signature_sha=compute_signature_sha("earnings_tone", "BBB", {"k": "second"}),
        db_path=db_path,
    )
    queue_action(
        alert_id=first.id,
        action_kind="thesis_update",
        payload={"seq": "a1"},
        db_path=db_path,
    )
    queue_action(
        alert_id=first.id,
        action_kind="bear_append",
        payload={"seq": "a2"},
        db_path=db_path,
    )
    queue_action(
        alert_id=second.id,
        action_kind="thesis_update",
        payload={"seq": "b1"},
        db_path=db_path,
    )

    batch_calls: list[list[int]] = []
    orig_batch = cio_export.list_queued_actions_for_alerts

    def counting_batch(
        alert_ids: list[int], db_path: Path | str | None = None
    ) -> dict[int, list[QueuedActionRow]]:
        batch_calls.append(list(alert_ids))
        return orig_batch(alert_ids, db_path=db_path)

    assert not hasattr(cio_export, "list_queued_actions_for_alert")
    monkeypatch.setattr(cio_export, "list_queued_actions_for_alerts", counting_batch)

    wb = load_workbook(export_cio_workbook(tmp_path / "cio.xlsx", db_path=db_path))

    assert batch_calls == [[second.id, first.id]]

    # Alerts newest-first; Queued Actions grouped in that alert order,
    # oldest-first within each alert: b1, a1, a2.
    assert wb["Alerts"].max_row == 3
    rows = list(wb["Queued Actions"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    assert str(rows[0][2]) == "BBB"
    assert "b1" in str(rows[0][7])
    assert str(rows[1][2]) == "AAA"
    assert "a1" in str(rows[1][7])
    assert str(rows[2][2]) == "AAA"
    assert "a2" in str(rows[2][7])
    assert rows[0][1] == second.id
    assert rows[1][1] == first.id
    assert rows[2][1] == first.id


def test_export_empty_alerts_does_not_query_actions(
    tmp_path: Path, db_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    batch_calls: list[list[int]] = []
    orig_batch = cio_export.list_queued_actions_for_alerts

    def counting_batch(
        alert_ids: list[int], db_path: Path | str | None = None
    ) -> dict[int, list[QueuedActionRow]]:
        batch_calls.append(list(alert_ids))
        return orig_batch(alert_ids, db_path=db_path)

    assert not hasattr(cio_export, "list_queued_actions_for_alert")
    monkeypatch.setattr(cio_export, "list_queued_actions_for_alerts", counting_batch)

    out = export_cio_workbook(tmp_path / "cio.xlsx", db_path=db_path)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Alerts", "Queued Actions", "Thesis Ledger"]
    assert wb["Alerts"].max_row == 1
    assert wb["Queued Actions"].max_row == 1
    assert wb["Thesis Ledger"].max_row == 1
    assert batch_calls == []
