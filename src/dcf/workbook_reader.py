"""Extract the FCF stream and diluted-shares snapshot from a DCF workbook.

The canonical workbook follows the META/GOOG/AMZN reference pattern in
`examples/dcf/`: a sheet named `Valuation` with year headers in row 1
(columns 3+), a single `FCF` row of data (label "FCF" — distinct from the
"FCF " section header), and a `#diluted shares outstanding` row. The reader
is intentionally tolerant about row positions — it scans column-A labels —
so workbooks evolved beyond the seed template still parse.

Reader returns a typed payload; the caller (valuation.py) handles math
and dcf_runs persistence.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

VALUATION_SHEET = "Valuation"
FCF_LABELS = (
    "fcf",
)  # exact, lowercased, stripped — section header "FCF " (with trailing space) is excluded
SHARES_LABELS = ("#diluted shares outstanding", "diluted shares outstanding", "diluted shares")
MAX_SCAN_ROW = 60
MAX_SCAN_COL = 25


class WorkbookReadError(Exception):
    """The workbook is missing required structure (sheet, row, or data)."""


@dataclass(frozen=True)
class WorkbookSnapshot:
    """What `read_valuation` extracts from a workbook for the PV calculator.

    - `fcf_by_year`: full year → FCF map (millions of the workbook's reporting
      currency). Both historical and forecast years included.
    - `shares_by_year`: full year → diluted shares outstanding map (millions).
    - `forecast_years`: ordered list of years to use as the forecast horizon —
      everything strictly after `latest_actual_year`.
    - `latest_actual_year`: the last fiscal year that already reported (used
      to anchor the "current shares" baseline for per-share fair value).
    """

    workbook_path: str
    fcf_by_year: dict[int, float]
    shares_by_year: dict[int, float]
    forecast_years: list[int]
    latest_actual_year: int


def read_valuation(workbook_path: Path, valuation_year: int) -> WorkbookSnapshot:
    """Open `workbook_path` and extract FCF + shares data needed for PV calc.

    `valuation_year` is the cutoff: years strictly greater than it are
    treated as forecast; years ≤ it are treated as actuals. The latest
    actual year provides the shares-outstanding baseline.
    """
    if not workbook_path.exists():
        raise WorkbookReadError(f"workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    if VALUATION_SHEET not in wb.sheetnames:
        raise WorkbookReadError(
            f"no '{VALUATION_SHEET}' sheet in {workbook_path.name} (sheets: {wb.sheetnames})"
        )
    ws = wb[VALUATION_SHEET]

    year_columns = _scan_year_columns(ws)
    if not year_columns:
        raise WorkbookReadError("no year headers found in Valuation row 1 columns 1–25")

    fcf_row = _find_data_row(ws, FCF_LABELS)
    if fcf_row is None:
        raise WorkbookReadError("no 'FCF' data row found in column A of Valuation sheet")

    shares_row = _find_data_row(ws, SHARES_LABELS)
    if shares_row is None:
        raise WorkbookReadError("no diluted-shares row found in column A of Valuation sheet")

    fcf_by_year = _extract_year_series(ws, fcf_row, year_columns)
    shares_by_year = _extract_year_series(ws, shares_row, year_columns)

    actuals = [y for y in fcf_by_year if y <= valuation_year]
    forecasts = sorted(y for y in fcf_by_year if y > valuation_year)
    if not actuals:
        raise WorkbookReadError(
            f"no actual FCF years ≤ {valuation_year} found in {workbook_path.name}"
        )
    if not forecasts:
        raise WorkbookReadError(
            f"no forecast FCF years > {valuation_year} found in {workbook_path.name}"
        )

    return WorkbookSnapshot(
        workbook_path=str(workbook_path),
        fcf_by_year=fcf_by_year,
        shares_by_year=shares_by_year,
        forecast_years=forecasts,
        latest_actual_year=max(actuals),
    )


def _scan_year_columns(ws: Worksheet) -> list[tuple[int, int]]:
    """Return [(column_index, year)] from row 1, only plausible fiscal years."""
    found: list[tuple[int, int]] = []
    for col in range(1, MAX_SCAN_COL + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, (int, float)) and 2010 <= int(v) <= 2050:
            found.append((col, int(v)))
    return found


def _find_data_row(ws: Worksheet, accepted_labels: tuple[str, ...]) -> int | None:
    """Scan column A for the first row whose stripped+lowered label exactly
    matches one of `accepted_labels` AND has at least one numeric value
    in the year columns. The exact-match keeps us from picking up section
    headers like "FCF " (with trailing space) or "Free cash flow margin".
    """
    for row in range(1, MAX_SCAN_ROW + 1):
        raw = ws.cell(row=row, column=1).value
        if not isinstance(raw, str):
            continue
        label = raw.strip().lower()
        if label not in accepted_labels:
            continue
        # Confirm the row has at least one numeric value in scan range.
        for col in range(2, MAX_SCAN_COL + 1):
            cell_val = ws.cell(row=row, column=col).value
            if isinstance(cell_val, (int, float)):
                return row
    return None


def _extract_year_series(
    ws: Worksheet, row: int, year_columns: list[tuple[int, int]]
) -> dict[int, float]:
    """Build year → value dict, skipping non-numeric cells."""
    out: dict[int, float] = {}
    for col, year in year_columns:
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            out[year] = float(v)
    return out
