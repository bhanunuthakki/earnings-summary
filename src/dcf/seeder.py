"""Build a fresh DCF workbook for a ticker from FMP data.

The seeder produces `dcf/<TICKER>.xlsx` with three sheets:

  Historicals — program-owned table of 30 line items x 20 quarters from
                FMP. Refreshed each cycle by `refresher.refresh_historicals`.

  Forecast    — INPUTS zone (user-editable, auto-derived from this ticker's
                TTM history at seed time) + PROJECTED zone (program-owned,
                rewritten each refresh from current INPUTS). See
                `dcf.forecast` for the math.

  Valuation   — year-header row + FCF row + diluted-shares row. The current
                year column carries the ticker's TTM FCF; forecast columns
                carry projected FCFs from the Forecast sheet. This is what
                `workbook_reader.read_valuation` parses on refresh, what the
                PV calc consumes, and what feeds `dcf_runs` (and therefore
                the briefs).

Source data: standardized FMP statements
(`{TICKER}_income_statement_quarterly.json`,
`{TICKER}_balance_sheet_quarterly.json`, `{TICKER}_cash_flow_quarterly.json`)
— stable camelCase fields across all tickers, unlike the per-company XBRL
`as_reported_*` files.

Fallback for recently-IPO'd tickers: when the FMP quarterly files are absent
(FMP returns empty until the first 10-Q), the seeder falls back to annual (FY)
figures from the `financial_facts` table — e.g. the audited S-1 statements
written by `compute.s1_financials`. The fallback yields one Historicals column
per fiscal year (necessarily sparse), and is only taken when `db_path` is
provided. See `_load_quarterly_records_from_facts`.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from dcf import forecast as forecast_mod

HISTORICALS_SHEET = "Historicals"
FORECAST_SHEET = "Forecast"
VALUATION_SHEET = "Valuation"
DEFAULT_QUARTERS = 20
MILLIONS = 1_000_000.0
_FCF_LABEL = "FCF"
_SHARES_LABEL = "#diluted shares outstanding"
# Period types loaded from financial_facts in the no-FMP fallback. Annual only:
# the fallback's motivating case (S-1-anchored tickers) has FY data only, and
# mixing FY with Q4 would collide on a shared period_end (both end Dec 31).
_ANNUAL_PERIOD_TYPES: tuple[str, ...] = ("FY",)


class SeederError(Exception):
    """Workbook seeding cannot proceed (missing inputs, refusing to overwrite)."""


@dataclass(frozen=True)
class _Row:
    """One Historicals line: (label, source-statement, FMP field, scale).

    `fact_line_item` is the canonical `financial_facts` line_item this row maps
    to (the snake_case name the FMP/S-1 extractors write); None when no
    canonical fact exists. Used only by the financial_facts fallback — the FMP
    path reads `field` off the raw statement JSON.
    """

    label: str
    source: str  # 'income' | 'balance' | 'cashflow'
    field: str
    scale: float
    fact_line_item: str | None = None


# Historicals row layout — fixed order so the refresher's column writes land
# on the same rows the seeder wrote.
_INCOME_ROWS: tuple[_Row, ...] = (
    _Row("Revenue", "income", "revenue", MILLIONS, "revenue"),
    _Row("Cost of Revenue", "income", "costOfRevenue", MILLIONS, "cost_of_revenue"),
    _Row("Gross Profit", "income", "grossProfit", MILLIONS, "gross_profit"),
    _Row(
        "R&D Expense",
        "income",
        "researchAndDevelopmentExpenses",
        MILLIONS,
        "research_and_development",
    ),
    _Row("SG&A Expense", "income", "sellingGeneralAndAdministrativeExpenses", MILLIONS, "sga"),
    _Row("Operating Expenses", "income", "operatingExpenses", MILLIONS, "operating_expenses"),
    _Row("Operating Income", "income", "operatingIncome", MILLIONS, "operating_income"),
    _Row("Net Income", "income", "netIncome", MILLIONS, "net_income"),
    _Row("Diluted EPS", "income", "epsDiluted", 1.0, "eps_diluted"),
    _Row(
        "Diluted Shares (M)",
        "income",
        "weightedAverageShsOutDil",
        MILLIONS,
        "weighted_avg_shares_diluted",
    ),
)
_BALANCE_ROWS: tuple[_Row, ...] = (
    _Row(
        "Cash & ST Investments",
        "balance",
        "cashAndShortTermInvestments",
        MILLIONS,
        "cash_and_short_term_investments",
    ),
    _Row("Receivables", "balance", "netReceivables", MILLIONS, "net_receivables"),
    _Row("Inventory", "balance", "inventory", MILLIONS, "inventory"),
    _Row("Total Current Assets", "balance", "totalCurrentAssets", MILLIONS, "total_current_assets"),
    _Row(
        "PP&E (net)",
        "balance",
        "propertyPlantEquipmentNet",
        MILLIONS,
        "property_plant_equipment_net",
    ),
    _Row("Goodwill", "balance", "goodwill", MILLIONS, "goodwill"),
    _Row("Total Assets", "balance", "totalAssets", MILLIONS, "total_assets"),
    _Row(
        "Total Current Liabilities",
        "balance",
        "totalCurrentLiabilities",
        MILLIONS,
        "total_current_liabilities",
    ),
    _Row("Long-term Debt", "balance", "longTermDebt", MILLIONS, "long_term_debt"),
    _Row(
        "Total Equity", "balance", "totalStockholdersEquity", MILLIONS, "total_stockholders_equity"
    ),
)
_CASHFLOW_ROWS: tuple[_Row, ...] = (
    _Row("Net Income (CF)", "cashflow", "netIncome", MILLIONS, "net_income_cf"),
    _Row(
        "D&A",
        "cashflow",
        "depreciationAndAmortization",
        MILLIONS,
        "depreciation_and_amortization_cf",
    ),
    _Row(
        "Stock-based Compensation",
        "cashflow",
        "stockBasedCompensation",
        MILLIONS,
        "stock_based_compensation",
    ),
    _Row(
        "Change in Working Capital",
        "cashflow",
        "changeInWorkingCapital",
        MILLIONS,
        "change_in_working_capital",
    ),
    _Row("Operating Cash Flow", "cashflow", "operatingCashFlow", MILLIONS, "operating_cash_flow"),
    _Row("Capex", "cashflow", "capitalExpenditure", MILLIONS, "capital_expenditure"),
    _Row("Free Cash Flow", "cashflow", "freeCashFlow", MILLIONS, "free_cash_flow"),
    # FMP cash-flow JSON carries `netCommonStockIssuance`, but no extractor maps
    # it to a canonical financial_facts line_item — blank in the fallback.
    _Row("Net Stock Issuance", "cashflow", "netCommonStockIssuance", MILLIONS, None),
    _Row("Dividends Paid", "cashflow", "commonDividendsPaid", MILLIONS, "common_dividends_paid"),
    _Row("Net Change in Cash", "cashflow", "netChangeInCash", MILLIONS, "net_change_in_cash"),
)


def seed_workbook(
    ticker: str,
    fmp_quarterly_dir: Path,
    output_path: Path,
    *,
    force: bool = False,
    quarters: int = DEFAULT_QUARTERS,
    forecast_years: int = forecast_mod.DEFAULT_FORECAST_YEARS,
    terminal_growth_pct: float = forecast_mod.DEFAULT_TERMINAL_GROWTH,
    base_year: int | None = None,
    db_path: Path | None = None,
) -> None:
    """Build a fresh DCF workbook for `ticker` at `output_path`.

    Raises `SeederError` if `output_path` exists and `force=False`, or if
    neither the FMP quarterly files nor — when `db_path` is given — the FY
    rows in `financial_facts` yield any records. Builds the workbook from
    scratch (no template copy) — every sheet is program-defined.

    `base_year` defaults to the current calendar year. Forecast columns
    start at `base_year + 1`. The current-year column on the Valuation
    sheet carries TTM actuals.

    `db_path` enables the financial_facts fallback for recently-IPO'd tickers
    with no FMP quarterly files (see module docstring).
    """
    ticker = ticker.upper()
    if output_path.exists() and not force:
        raise SeederError(
            f"refusing to overwrite existing workbook: {output_path} (pass force=True)"
        )

    quarters_data = _load_quarterly_records(ticker, fmp_quarterly_dir, quarters, db_path=db_path)
    if _fmp_quarterly_available(ticker, fmp_quarterly_dir):
        income_records = _load_records(ticker, fmp_quarterly_dir, "income_statement")
        cashflow_records = _load_records(ticker, fmp_quarterly_dir, "cash_flow")
        balance_records = _load_records(ticker, fmp_quarterly_dir, "balance_sheet")
    else:
        # financial_facts fallback: the per-fiscal-year Historicals records ARE
        # the forecast deriver's source. Pad each annual record into a
        # 4-quarter TTM window so derive_initial_inputs reads one fiscal year
        # per window (see _pad_annual_records).
        income_records = _pad_annual_records([q.income for q in reversed(quarters_data)])
        cashflow_records = _pad_annual_records([q.cashflow for q in reversed(quarters_data)])
        balance_records = _pad_annual_records([q.balance for q in reversed(quarters_data)])

    inputs = forecast_mod.derive_initial_inputs(
        income_records,
        cashflow_records,
        balance_records,
        terminal_growth_pct=terminal_growth_pct,
        forecast_years=forecast_years,
    )
    resolved_base_year = base_year if base_year is not None else date.today().year
    projections = forecast_mod.compute_projections(inputs, resolved_base_year)
    ttm_fcf_M = _ttm_fcf_millions(cashflow_records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    default = wb.active
    assert default is not None
    wb.remove(default)

    hist_ws = wb.create_sheet(HISTORICALS_SHEET)
    _write_historicals(hist_ws, quarters_data)

    forecast_ws = wb.create_sheet(FORECAST_SHEET)
    forecast_mod.write_inputs_section(forecast_ws, inputs)
    forecast_mod.write_projections_section(forecast_ws, projections)

    val_ws = wb.create_sheet(VALUATION_SHEET)
    write_valuation_sheet(val_ws, inputs, projections, resolved_base_year, ttm_fcf_M)

    wb.save(str(output_path))


def write_historicals_sheet(
    workbook_path: Path,
    fmp_quarterly_dir: Path,
    ticker: str,
    *,
    quarters: int = DEFAULT_QUARTERS,
    db_path: Path | None = None,
) -> int:
    """Open `workbook_path`, rewrite ONLY its Historicals sheet, save.

    Returns the count of numeric cells written. Kept as a thin public
    helper so the refresher can reuse the layout without going through the
    full seed path. `db_path` enables the financial_facts fallback (see
    `_load_quarterly_records`).
    """
    quarters_data = _load_quarterly_records(
        ticker.upper(), fmp_quarterly_dir, quarters, db_path=db_path
    )
    wb = openpyxl.load_workbook(str(workbook_path))
    if HISTORICALS_SHEET in wb.sheetnames:
        del wb[HISTORICALS_SHEET]
    ws = wb.create_sheet(HISTORICALS_SHEET)
    cells = _write_historicals(ws, quarters_data)
    # Move Historicals to position 0 if it just got appended at the end.
    sheets = wb.sheetnames
    idx = sheets.index(HISTORICALS_SHEET)
    wb.move_sheet(HISTORICALS_SHEET, offset=-idx)
    wb.save(str(workbook_path))
    return cells


def write_valuation_sheet(
    ws: Worksheet,
    inputs: forecast_mod.ForecastInputs,
    projections: forecast_mod.ForecastProjections,
    base_year: int,
    ttm_fcf_M: float | None,
) -> None:
    """Write the Valuation sheet in the layout `workbook_reader` expects.

    Column A holds row labels; row 1 holds year headers starting at
    `base_year` (the current-year actual) and continuing through the
    forecast horizon. Two data rows:

      FCF — base_year column = TTM FCF (actual, from FMP); forecast columns =
            `projections.valuation_fcf_M` (the SBC-charged FCF the bridge
            produces — see dcf.forecast).

      #diluted shares outstanding — base_year column = current diluted shares;
            forecast columns = `projections.shares_M` (the evolving, diluting
            path). The workbook_reader uses the base year's (current) shares
            value as the per-share denominator.
    """
    # Wipe whatever was there (refresher reuses this on every cycle).
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    bold = Font(bold=True)
    note = Font(italic=True, color="666666")

    ws.cell(row=1, column=1, value=None)
    ws.cell(row=1, column=2, value=base_year).font = bold
    for i, year in enumerate(projections.years):
        ws.cell(row=1, column=3 + i, value=year).font = bold

    ws.cell(row=2, column=1, value=_FCF_LABEL).font = bold
    if ttm_fcf_M is not None:
        c = ws.cell(row=2, column=2, value=ttm_fcf_M)
        c.number_format = "#,##0"
    for i, fcf in enumerate(projections.valuation_fcf_M):
        c = ws.cell(row=2, column=3 + i, value=fcf)
        c.number_format = "#,##0"

    ws.cell(row=3, column=1, value=_SHARES_LABEL).font = bold
    base_shares = ws.cell(row=3, column=2, value=inputs.diluted_shares_M)
    base_shares.number_format = "#,##0"
    for i, sh in enumerate(projections.shares_M):
        c = ws.cell(row=3, column=3 + i, value=sh)
        c.number_format = "#,##0"

    ws.cell(
        row=5,
        column=1,
        value="Rewritten on every refresh from Forecast sheet inputs. Edit Forecast, not here.",
    ).font = note

    ws.column_dimensions["A"].width = 32
    from openpyxl.utils import get_column_letter

    for i in range(1 + len(projections.years)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12


def _ttm_fcf_millions(cashflow_records: list[dict[str, object]]) -> float | None:
    total = 0.0
    seen = 0
    for r in cashflow_records[:4]:
        v = r.get("freeCashFlow")
        if isinstance(v, (int, float)):
            total += float(v)
            seen += 1
    if seen == 0:
        return None
    return total / MILLIONS


def _load_records(ticker: str, fmp_quarterly_dir: Path, statement: str) -> list[dict[str, object]]:
    return _read_fmp(fmp_quarterly_dir / f"{ticker.upper()}_{statement}_quarterly.json")


# ---------------------------------------------------------------------------
# Historicals helpers — unchanged from the previous seeder, factored to write
# into a passed-in worksheet so the refresher can reuse the row layout.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _QuarterRecord:
    """One quarter's data, merged across the 3 statement files."""

    period_label: str
    income: dict[str, object]
    balance: dict[str, object]
    cashflow: dict[str, object]


def _load_quarterly_records(
    ticker: str,
    fmp_quarterly_dir: Path,
    quarters: int,
    *,
    db_path: Path | None = None,
) -> list[_QuarterRecord]:
    """Build N periods of merged income+balance+cashflow records, oldest first.

    Primary source: FMP standardized quarterly statements. Fallback (only when
    FMP yields nothing AND `db_path` is set): annual (FY) figures from the
    `financial_facts` table — for recently-IPO'd tickers FMP hasn't backfilled
    yet. Raises `SeederError` if neither source yields records.
    """
    fmp_quarters = _load_quarterly_records_fmp(ticker, fmp_quarterly_dir, quarters)
    if fmp_quarters:
        return fmp_quarters
    if db_path is not None:
        facts_quarters = _load_quarterly_records_from_facts(ticker, db_path, quarters)
        if facts_quarters:
            return facts_quarters
    raise SeederError(
        f"no income statement records for {ticker} "
        f"(no FMP quarterly files, no FY financial_facts fallback)"
    )


def _fmp_quarterly_available(ticker: str, fmp_quarterly_dir: Path) -> bool:
    """True iff the FMP quarterly income file yields ≥1 indexable quarter.

    Mirrors the source-selection in `_load_quarterly_records`: when False, the
    seeder is on the financial_facts fallback path.
    """
    income_raw = _read_fmp(fmp_quarterly_dir / f"{ticker}_income_statement_quarterly.json")
    return bool(_index_by_period(income_raw))


def _load_quarterly_records_fmp(
    ticker: str, fmp_quarterly_dir: Path, quarters: int
) -> list[_QuarterRecord]:
    """Build merged quarter records from FMP statement JSONs. `[]` if none."""
    income_raw = _read_fmp(fmp_quarterly_dir / f"{ticker}_income_statement_quarterly.json")
    balance_raw = _read_fmp(fmp_quarterly_dir / f"{ticker}_balance_sheet_quarterly.json")
    cashflow_raw = _read_fmp(fmp_quarterly_dir / f"{ticker}_cash_flow_quarterly.json")

    income_by_key = _index_by_period(income_raw)
    balance_by_key = _index_by_period(balance_raw)
    cashflow_by_key = _index_by_period(cashflow_raw)

    if not income_by_key:
        return []

    sorted_keys = sorted(income_by_key.keys(), reverse=True)[:quarters]
    sorted_keys.reverse()  # oldest -> newest left-to-right

    return [
        _QuarterRecord(
            period_label=f"{period} {year}",
            income=income_by_key.get((year, period), {}),
            balance=balance_by_key.get((year, period), {}),
            cashflow=cashflow_by_key.get((year, period), {}),
        )
        for (year, period) in sorted_keys
    ]


def _load_quarterly_records_from_facts(
    ticker: str, db_path: Path, quarters: int
) -> list[_QuarterRecord]:
    """Build annual Historicals records from `financial_facts`. `[]` if none.

    Loads each Historicals row that has a canonical `fact_line_item` as an
    annual (FY) series — tier-aware via the timeseries loader, so a later
    SEC/FMP-sourced fact supersedes an S-1-provisional one for the same period.
    Values land under the row's FMP `field` key so the existing
    `_extract_value` / `_write_historicals` path (and the forecast deriver)
    consume them unchanged. One `_QuarterRecord` per fiscal year, oldest first,
    capped at `quarters`.
    """
    from timeseries.loaders import load_financial_series

    # period_end -> {'income'|'balance'|'cashflow' -> {FMP field: value}}
    by_period: dict[datetime, dict[str, dict[str, object]]] = {}
    for spec in (*_INCOME_ROWS, *_BALANCE_ROWS, *_CASHFLOW_ROWS):
        if spec.fact_line_item is None:
            continue
        series = load_financial_series(
            ticker,
            spec.fact_line_item,
            db_path=db_path,
            period_types=_ANNUAL_PERIOD_TYPES,
        )
        for obs in series:
            slot = by_period.setdefault(
                obs.period_end, {"income": {}, "balance": {}, "cashflow": {}}
            )
            slot[spec.source][spec.field] = obs.value

    if not by_period:
        return []

    sorted_periods = sorted(by_period)[-quarters:]  # oldest -> newest, capped
    return [
        _QuarterRecord(
            period_label=f"FY{pe.year}",
            income=by_period[pe]["income"],
            balance=by_period[pe]["balance"],
            cashflow=by_period[pe]["cashflow"],
        )
        for pe in sorted_periods
    ]


def _pad_annual_records(
    annual_newest_first: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Space annual records into 4-quarter TTM windows for the forecast deriver.

    `derive_initial_inputs` sums records[:4] as the current TTM and records[4:8]
    as the prior-year TTM. Placing each fiscal year at a window boundary —
    followed by 3 empty dicts `_ttm_sum` skips — makes each window read exactly
    one fiscal year, so an annual-only ticker derives a correct base revenue and
    YoY growth instead of summing several years together.
    """
    out: list[dict[str, object]] = []
    for rec in annual_newest_first:
        out.append(rec)
        out.extend(({}, {}, {}))
    return out


def _read_fmp(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    raw_obj: object = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw_obj, list):
        return []
    raw = cast("list[object]", raw_obj)
    return [cast("dict[str, object]", r) for r in raw if isinstance(r, dict)]


def _index_by_period(
    records: list[dict[str, object]],
) -> dict[tuple[int, str], dict[str, object]]:
    """`fiscalYear` ships as int OR str across FMP files; coerce defensively."""
    out: dict[tuple[int, str], dict[str, object]] = {}
    for r in records:
        year_raw = r.get("fiscalYear")
        period = r.get("period")
        if not (isinstance(period, str) and period.startswith("Q")):
            continue
        try:
            year = int(year_raw) if isinstance(year_raw, (int, str)) else None
        except ValueError:
            year = None
        if year is not None:
            out[(year, period)] = r
    return out


def _write_historicals(ws: Worksheet, quarters: list[_QuarterRecord]) -> int:
    bold = Font(bold=True)
    n_cols = len(quarters)

    ws.cell(row=1, column=1, value="Line Item").font = bold
    for i, q in enumerate(quarters):
        ws.cell(row=1, column=2 + i, value=q.period_label).font = bold

    row = 2
    data_cells = 0
    for section, rows in (
        ("INCOME STATEMENT", _INCOME_ROWS),
        ("BALANCE SHEET", _BALANCE_ROWS),
        ("CASH FLOW", _CASHFLOW_ROWS),
    ):
        ws.cell(row=row, column=1, value=section).font = bold
        row += 1
        for spec in rows:
            ws.cell(row=row, column=1, value=spec.label)
            for i, q in enumerate(quarters):
                val = _extract_value(q, spec)
                if val is not None:
                    ws.cell(row=row, column=2 + i, value=val)
                    data_cells += 1
            row += 1
        row += 1  # spacer between sections

    ws.cell(row=row, column=1, value="Units: USD millions except per-share").font = Font(
        italic=True
    )

    _autosize_historicals_columns(ws, n_cols)
    return data_cells


def _extract_value(q: _QuarterRecord, spec: _Row) -> float | None:
    if spec.source == "income":
        src = q.income
    elif spec.source == "balance":
        src = q.balance
    elif spec.source == "cashflow":
        src = q.cashflow
    else:
        return None
    raw = src.get(spec.field)
    if not isinstance(raw, (int, float)):
        return None
    return float(raw) / spec.scale


def _autosize_historicals_columns(ws: Worksheet, n_cols: int) -> None:
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = 32
    for i in range(n_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12
