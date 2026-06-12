"""Assumption provenance for the redesigned DCF workbook.

The yellow Dashboard cells are bare literals: nothing in the workbook says
which values came from the Opus assumption pass (``data/dcf_assumptions/
<T>.json["redesign"]``), which the user has since overridden, and which are
builder defaults — and the Opus ``narrative``/``reasoning`` prose never
surfaces next to the numbers it justifies. This module closes that gap with
three pieces, all regenerated on every build/refresh (the same survive-the-
rebuild contract as the Sensitivity sheet):

* an **opus_baseline** block in the assumptions JSON — an immutable snapshot
  of the Opus-set values. ``sync_assumptions_json`` deliberately mirrors
  workbook edits back into the ``redesign`` block (single source of truth for
  from-scratch builds), which destroys the original Opus values — so
  provenance needs its own copy that sync never touches. Written fresh by
  each Opus pass; seeded once from the current ``redesign`` block for names
  whose pass predates provenance tracking (flagged ``seeded`` — those values
  may already include user edits, the best record that still exists).
* an **assumption_overrides** ledger in the same JSON — per input, the Opus
  value and the date the refresher first saw the workbook diverge from it
  (cleared when the value returns to baseline). This is what lets the sheet
  say "overridden from Opus 11.5% on 2026-06-12" rather than just "edited".
* the read-only **Assumptions sheet** + native comments on the yellow cells —
  one row per Dashboard input: live current value (cross-sheet formula, so it
  never goes stale between refreshes), source (Opus / user-edited / builder
  default), the Opus value, and a note; headed by the Opus narrative and
  reasoning prose. Comments carry the same per-cell verdict into
  hover-in-place, and survive the Sheets round-trip as notes.

Capture/inject never touches this sheet (it is not a marker sheet and holds
no inputs); the builder and the refresher both rewrite it after the input
cells settle, so it reflects the injected values, not the builder defaults.
"""

from __future__ import annotations

import json
import math
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from dcf import redesign as redesign_mod

ASSUMPTIONS_SHEET = "Assumptions"

# Source labels — the sheet's Source column and the cell-comment headline.
SOURCE_OPUS = "Opus"
SOURCE_USER = "user-edited"
SOURCE_DEFAULT = "builder default"
SOURCE_SYNCED = "workbook (synced)"  # block created by sync, no Opus pass behind it

# The numeric redesign-block fields snapshotted into opus_baseline, with the
# scalar RedesignInputs attribute each one maps to (segments are handled
# separately). capex_pct_revenue_2026 is snapshotted but never compared — the
# JSON stores a fraction of FY1E revenue while the Dashboard stores $M, so
# edit-detection for that one cell is out of reach without the revenue base.
_BASELINE_SCALAR_FIELDS: tuple[tuple[str, str], ...] = (
    ("near_term_op_margin", "near_op_margin"),
    ("terminal_op_margin", "terminal_op_margin"),
    ("tax_rate", "tax_rate"),
    ("terminal_capex_da", "terminal_capex_da"),
    ("terminal_method", "terminal_method"),
    ("exit_basis", "terminal_basis"),
    ("exit_multiple", "exit_multiple"),
    ("terminal_growth_g", "terminal_growth_g"),
    ("beta", "beta"),
    ("risk_free_rate", "risk_free_rate"),
    ("equity_risk_premium", "equity_risk_premium"),
    ("cost_of_debt", "cost_of_debt"),
)
_BASELINE_PASSTHROUGH_FIELDS: tuple[str, ...] = ("capex_pct_revenue_2026",)

# The builder writes beta rounded to 3dp (B40) — a baseline of 1.2435 reads
# back as 1.244 without anyone touching it. Everything else round-trips exact.
_FIELD_ABS_TOL: dict[str, float] = {"beta": 5e-4}
_DEFAULT_REL_TOL = 1e-6
_DEFAULT_ABS_TOL = 1e-9


class ProvenanceError(Exception):
    """The assumptions JSON exists but cannot be read/updated — surfaced, not
    swallowed (a silently-skipped provenance write would quietly mislabel
    user edits as Opus values on the next sheet)."""


@dataclass(frozen=True)
class OpusBaseline:
    """The immutable Opus-set values + where they came from."""

    as_of: str
    set_by: str
    seeded: bool
    values: dict[str, object]  # flat scalars + "segments" sub-dict

    def scalar(self, field: str) -> object | None:
        return self.values.get(field)


@dataclass(frozen=True)
class AssumptionRow:
    """One Assumptions-sheet row / one yellow-cell comment."""

    label: str
    cell: str  # Dashboard address, e.g. "B29"
    fmt: str | None  # number format for the live current-value formula
    source: str
    baseline_display: str  # the Opus value, formatted ("—" when none)
    note: str


# --------------------------------------------------------------------------- #
# JSON helpers
# --------------------------------------------------------------------------- #
def _load_assumptions(path: Path) -> dict[str, object] | None:
    """The assumptions JSON as a dict, ``None`` when the file doesn't exist.

    A file that exists but is unreadable raises ``ProvenanceError`` — that is
    corruption, not absence.
    """
    if not path.exists():
        return None
    try:
        raw: object = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        raise ProvenanceError(f"unreadable assumptions JSON at {path}: {e}") from e
    if not isinstance(raw, dict):
        raise ProvenanceError(f"assumptions JSON at {path} is not an object")
    return cast("dict[str, object]", raw)


def _write_assumptions(path: Path, data: dict[str, object]) -> None:
    try:
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except OSError as e:
        raise ProvenanceError(f"cannot write assumptions JSON at {path}: {e}") from e


def snapshot_baseline_values(redesign_block: dict[str, object]) -> dict[str, object]:
    """The numeric/string assumption fields of a redesign block, copied for
    ``opus_baseline.values`` (prose, flags and bookkeeping keys excluded)."""
    out: dict[str, object] = {}
    segs = redesign_block.get("segments")
    if isinstance(segs, dict):
        out["segments"] = cast("object", json.loads(json.dumps(segs)))  # deep copy
    for field, _attr in _BASELINE_SCALAR_FIELDS:
        if field in redesign_block:
            out[field] = redesign_block[field]
    for field in _BASELINE_PASSTHROUGH_FIELDS:
        if field in redesign_block:
            out[field] = redesign_block[field]
    return out


def ensure_opus_baseline(
    assumptions_path: Path, *, today: date | None = None
) -> OpusBaseline | None:
    """Load the ``opus_baseline`` block, seeding it once when absent.

    Seeding copies the CURRENT ``redesign`` block values — for names whose
    Opus pass predates provenance tracking this is the best remaining record
    (post-sync it may already contain user edits; ``seeded`` flags that).
    Returns ``None`` when there is no assumptions file or no redesign block
    (nothing to base provenance on — every input is then a builder default).
    """
    data = _load_assumptions(assumptions_path)
    if data is None:
        return None
    existing_raw = data.get("opus_baseline")
    if isinstance(existing_raw, dict):
        existing = cast("dict[str, object]", existing_raw)
        values = existing.get("values")
        return OpusBaseline(
            as_of=str(existing.get("as_of", "")),
            set_by=str(existing.get("set_by", "")),
            seeded=bool(existing.get("seeded", False)),
            values=cast("dict[str, object]", values) if isinstance(values, dict) else {},
        )
    block = data.get("redesign")
    if not isinstance(block, dict):
        return None
    baseline = OpusBaseline(
        as_of=(today or date.today()).isoformat(),
        set_by="seeded_from_redesign_block",
        seeded=True,
        values=snapshot_baseline_values(cast("dict[str, object]", block)),
    )
    data["opus_baseline"] = {
        "as_of": baseline.as_of,
        "set_by": baseline.set_by,
        "seeded": True,
        "values": baseline.values,
    }
    _write_assumptions(assumptions_path, data)
    return baseline


def baseline_from_opus_pass(
    redesign_block: dict[str, object], *, today: date | None = None
) -> dict[str, object]:
    """The ``opus_baseline`` payload a fresh Opus pass writes alongside its
    redesign block — authoritative (not seeded), so a re-run of the pass
    resets provenance to the new Opus values."""
    return {
        "as_of": (today or date.today()).isoformat(),
        "set_by": "opus-4.8",
        "seeded": False,
        "values": snapshot_baseline_values(redesign_block),
    }


# --------------------------------------------------------------------------- #
# Field comparison + classification
# --------------------------------------------------------------------------- #
def _values_match(field: str, current: object, baseline: object) -> bool:
    if isinstance(current, str) or isinstance(baseline, str):
        return isinstance(current, str) and isinstance(baseline, str) and current == baseline
    if isinstance(current, (int, float)) and isinstance(baseline, (int, float)):
        return math.isclose(
            float(current),
            float(baseline),
            rel_tol=_DEFAULT_REL_TOL,
            abs_tol=_FIELD_ABS_TOL.get(field, _DEFAULT_ABS_TOL),
        )
    return False


def _parse_ledger(raw: object) -> dict[str, dict[str, object]]:
    """The assumption_overrides ledger off a loaded JSON value (tolerant)."""
    if not isinstance(raw, dict):
        return {}
    out: dict[str, dict[str, object]] = {}
    for k, v in cast("dict[str, object]", raw).items():
        if isinstance(v, dict):
            out[str(k)] = dict(cast("dict[str, object]", v))
    return out


def _seg_field_id(segment: str, key: str) -> str:
    """Ledger key for a per-segment growth input ('::' never appears in FMP
    segment names; scalar fields use the bare redesign-block key)."""
    prefix = "seg_near" if key == "near_term_growth" else "seg_term"
    return f"{prefix}::{segment}"


def _current_fields(inp: redesign_mod.RedesignInputs) -> dict[str, object]:
    """Every comparable input as ``field_id -> current value`` (the ledger's
    and classifier's shared view of the workbook)."""
    out: dict[str, object] = {}
    for s in inp.segments:
        out[_seg_field_id(s, "near_term_growth")] = inp.near_growth_by_segment[s]
        out[_seg_field_id(s, "terminal_growth")] = inp.terminal_growth_by_segment[s]
    for field, attr in _BASELINE_SCALAR_FIELDS:
        out[field] = getattr(inp, attr)
    return out


def _baseline_fields(baseline: OpusBaseline) -> dict[str, object]:
    out: dict[str, object] = {}
    segs = baseline.values.get("segments")
    if isinstance(segs, dict):
        for s, block_raw in cast("dict[str, object]", segs).items():
            if not isinstance(block_raw, dict):
                continue
            block = cast("dict[str, object]", block_raw)
            for key in ("near_term_growth", "terminal_growth"):
                v = block.get(key)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    out[_seg_field_id(str(s), key)] = float(v)
    for field, _attr in _BASELINE_SCALAR_FIELDS:
        if field in baseline.values:
            out[field] = baseline.values[field]
    return out


def update_override_ledger(
    assumptions_path: Path,
    inp: redesign_mod.RedesignInputs,
    baseline: OpusBaseline | None,
    *,
    today: date | None = None,
) -> dict[str, dict[str, object]]:
    """Reconcile ``assumption_overrides`` with the workbook's current inputs.

    A field that newly diverges from the Opus baseline gets an entry recording
    the Opus value and today's date (the refresh that first SAW the edit — the
    closest observable to when the user made it); a field back at baseline
    drops its entry. Returns the up-to-date ledger; no file write when nothing
    changed. With no baseline there is nothing to diverge from — the existing
    ledger (normally empty) is returned untouched.
    """
    data = _load_assumptions(assumptions_path)
    if data is None:
        return {}
    ledger = _parse_ledger(data.get("assumption_overrides"))
    if baseline is None:
        return ledger

    base_fields = _baseline_fields(baseline)
    changed = False
    for field_id, current in _current_fields(inp).items():
        base = base_fields.get(field_id)
        if base is None:
            continue  # Opus never set this field — builder default, not an override
        if _values_match(field_id, current, base):
            if field_id in ledger:
                del ledger[field_id]
                changed = True
        elif field_id not in ledger:
            ledger[field_id] = {
                "opus_value": base,
                "overridden_on": (today or date.today()).isoformat(),
            }
            changed = True
    if changed:
        data["assumption_overrides"] = ledger
        _write_assumptions(assumptions_path, data)
    return ledger


# --------------------------------------------------------------------------- #
# Row construction
# --------------------------------------------------------------------------- #
_PCT = "0.0%"
_MULT = '0.0"x"'
_USD = "#,##0"
_NUM3 = "0.000"
_SIGNED_PCT = "+0.0%;-0.0%"
_SIGNED_MULT = '+0.0"x";-0.0"x"'


def _fmt_value(value: object, fmt: str | None) -> str:
    """Human text for a baseline value (sheet column + comment prose)."""
    if value is None:
        return "—"
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        v = float(value)
        if fmt in (_PCT, _SIGNED_PCT):
            return f"{v * 100:.1f}%"
        if fmt in (_MULT, _SIGNED_MULT):
            return f"{v:.1f}x"
        if fmt == _NUM3:
            return f"{v:.3f}"
        return f"{v:,.0f}"
    return str(value)


def _classify(
    field_id: str,
    current: object,
    baseline: OpusBaseline | None,
    ledger: dict[str, dict[str, object]],
    *,
    block_origin: str | None = None,
) -> tuple[str, object | None, str]:
    """(source, baseline_value, note) for one input.

    ``block_origin`` distinguishes a redesign block written by a real Opus
    pass from one created by the workbook sync for a name that never had a
    pass (``set_by: "sync"``) — those values are the user's own, not Opus's.
    """
    base = _baseline_fields(baseline).get(field_id) if baseline is not None else None
    if base is None:
        if block_origin == "sync":
            return SOURCE_SYNCED, None, "no Opus pass — value synced from this workbook"
        return SOURCE_DEFAULT, None, ""
    if _values_match(field_id, current, base):
        return SOURCE_OPUS, base, ""
    entry = ledger.get(field_id, {})
    opus_v = entry.get("opus_value", base)
    on = entry.get("overridden_on")
    note = f"overridden from Opus {_fmt_value(opus_v, _field_fmt(field_id))}"
    if isinstance(on, str) and on:
        note += f" on {on}"
    return SOURCE_USER, base, note


def _field_fmt(field_id: str) -> str | None:
    if field_id.startswith(("seg_near::", "seg_term::")):
        return _PCT
    return {
        "near_term_op_margin": _PCT,
        "terminal_op_margin": _PCT,
        "tax_rate": _PCT,
        "terminal_capex_da": _MULT,
        "terminal_method": None,
        "exit_basis": None,
        "exit_multiple": _MULT,
        "terminal_growth_g": _PCT,
        "beta": _NUM3,
        "risk_free_rate": _PCT,
        "equity_risk_premium": _PCT,
        "cost_of_debt": _PCT,
    }.get(field_id)


_SCALAR_ROW_DEFS: tuple[tuple[str, str, int], ...] = (
    # (field_id, label, dashboard row) — col B unless noted
    ("near_term_op_margin", "Near-term operating margin", 29),
    ("terminal_op_margin", "Terminal operating margin", 30),
    ("tax_rate", "Tax rate", 31),
    ("terminal_capex_da", "Terminal capex / D&A", 35),
    ("risk_free_rate", "Risk-free rate (10Y)", 38),
    ("equity_risk_premium", "Equity risk premium", 39),
    ("beta", "Beta (levered)", 40),
    ("cost_of_debt", "Pre-tax cost of debt", 41),
    ("terminal_method", "Terminal method", 43),
    ("exit_basis", "Exit basis", 44),
    ("exit_multiple", "Exit multiple", 45),
    ("terminal_growth_g", "Terminal growth (g)", 46),
)

_SCEN_LEVER_LABELS: tuple[tuple[int, str, str, str], ...] = (
    # (dashboard row, lever label, deltas attr, number format)
    (redesign_mod.SCEN_ROW_GROWTH_NEAR, "segment growth, near-term", "growth_near", _SIGNED_PCT),
    (redesign_mod.SCEN_ROW_GROWTH_TERM, "segment growth, terminal", "growth_term", _SIGNED_PCT),
    (redesign_mod.SCEN_ROW_MARGIN_NEAR, "operating margin, near-term", "margin_near", _SIGNED_PCT),
    (redesign_mod.SCEN_ROW_MARGIN_TERM, "operating margin, terminal", "margin_term", _SIGNED_PCT),
    (redesign_mod.SCEN_ROW_EXIT_MULT, "exit multiple", "exit_multiple", _SIGNED_MULT),
    (redesign_mod.SCEN_ROW_TG, "terminal growth (g)", "terminal_g", _SIGNED_PCT),
)


def _segment_rows_on_sheet(wb: Workbook, segments: tuple[str, ...]) -> dict[str, int]:
    """Dashboard row per segment name, read off column A (exact, survives a
    skipped/blank row where positional indexing would drift)."""
    dsh = wb[redesign_mod.DASHBOARD_SHEET]
    out: dict[str, int] = {}
    wanted = set(segments)
    for row in range(redesign_mod.SEG_ROW0, redesign_mod.SEG_ROW_MAX + 1):
        v = dsh.cell(row=row, column=1).value
        name = v.strip() if isinstance(v, str) else None
        if name and name in wanted and name not in out:
            out[name] = row
    return out


def build_assumption_rows(
    wb: Workbook,
    inp: redesign_mod.RedesignInputs,
    baseline: OpusBaseline | None,
    ledger: dict[str, dict[str, object]],
    *,
    block_origin: str | None = None,
) -> list[AssumptionRow]:
    """One row per Dashboard input, classified against the Opus baseline."""
    rows: list[AssumptionRow] = []
    seg_rows = _segment_rows_on_sheet(wb, inp.segments)

    for s in inp.segments:
        sheet_row = seg_rows.get(s)
        if sheet_row is None:
            continue
        for key, col, what in (
            ("near_term_growth", "B", "near-term growth"),
            ("terminal_growth", "C", "terminal growth"),
        ):
            fid = _seg_field_id(s, key)
            current = (
                inp.near_growth_by_segment[s]
                if key == "near_term_growth"
                else inp.terminal_growth_by_segment[s]
            )
            source, base, note = _classify(
                fid, current, baseline, ledger, block_origin=block_origin
            )
            rows.append(
                AssumptionRow(
                    label=f"{s} — {what}",
                    cell=f"{col}{sheet_row}",
                    fmt=_PCT,
                    source=source,
                    baseline_display=_fmt_value(base, _PCT),
                    note=note,
                )
            )

    for fid, label, dash_row in _SCALAR_ROW_DEFS:
        current = _current_fields(inp)[fid]
        source, base, note = _classify(fid, current, baseline, ledger, block_origin=block_origin)
        fmt = _field_fmt(fid)
        rows.append(
            AssumptionRow(
                label=label,
                cell=f"B{dash_row}",
                fmt=fmt,
                source=source,
                baseline_display=_fmt_value(base, fmt),
                note=note,
            )
        )

    # 2026 capex ($M): the JSON stores a fraction of FY1E revenue, the cell
    # stores $M — not numerically comparable, so no edit-detection; surface
    # where the number came from instead of guessing.
    capex_pct = baseline.scalar("capex_pct_revenue_2026") if baseline is not None else None
    if isinstance(capex_pct, (int, float)) and not isinstance(capex_pct, bool):
        capex_source = SOURCE_OPUS
        capex_note = (
            f"set from Opus {float(capex_pct) * 100:.0f}% of FY1E revenue; "
            "edits not tracked (unit differs from the JSON)"
        )
    else:
        capex_source = SOURCE_DEFAULT if block_origin != "sync" else SOURCE_SYNCED
        capex_note = "last-year capex +10% unless set to guidance; edits not tracked"
    rows.append(
        AssumptionRow(
            label="2026 capex ($M)",
            cell="B34",
            fmt=_USD,
            source=capex_source,
            baseline_display="—",
            note=capex_note,
        )
    )

    # Scenario Δs: seeds are code constants (BULL_SEED/BEAR_SEED), not Opus
    # values — source is "seed default" vs "user-edited" by comparison.
    for dash_row, lever, attr, fmt in _SCEN_LEVER_LABELS:
        for col, scen, deltas, seed in (
            ("C", "Bull", inp.bull_deltas, redesign_mod.BULL_SEED),
            ("D", "Bear", inp.bear_deltas, redesign_mod.BEAR_SEED),
        ):
            cur = getattr(deltas, attr)
            seed_v = getattr(seed, attr)
            edited = not _values_match(f"scen_{attr}", cur, seed_v)
            rows.append(
                AssumptionRow(
                    label=f"{scen} Δ — {lever}",
                    cell=f"{col}{dash_row}",
                    fmt=fmt,
                    source=SOURCE_USER if edited else "seed default",
                    baseline_display=_fmt_value(seed_v, fmt),
                    note=f"documented seed {_fmt_value(seed_v, fmt)}" if edited else "",
                )
            )
    return rows


# --------------------------------------------------------------------------- #
# Sheet + comment writers
# --------------------------------------------------------------------------- #
_TITLE_FONT = Font(bold=True, size=15)
_SUB_FONT = Font(italic=True, color="666666")
_HDR_FONT = Font(bold=True)
_BAND_FONT = Font(bold=True, color="FFFFFF")
_BAND_FILL = PatternFill("solid", fgColor="1F4E78")
_LINK_FONT = Font(color="008000")
_USER_FONT = Font(bold=True, color="9C5700")
_WRAP = Alignment(wrap_text=True, vertical="top")

_COLS = ("Input", "Cell", "Current", "Source", "Opus value", "Note")
_COL_WIDTHS = (38, 7, 12, 16, 12, 52)
_COMMENT_AUTHOR = "DCF provenance"


def _band(ws: Worksheet, row: int, text: str) -> None:
    for c in range(1, len(_COLS) + 1):
        ws.cell(row, c).fill = _BAND_FILL
    ws.cell(row, 1, "  " + text).font = _BAND_FONT


def _prose_block(ws: Worksheet, row: int, text: str, *, n_rows: int) -> int:
    """Write wrapped prose merged across the table width; returns next row."""
    end = row + n_rows - 1
    ws.merge_cells(start_row=row, start_column=1, end_row=end, end_column=len(_COLS))
    cell = ws.cell(row, 1, text)
    cell.alignment = _WRAP
    return end + 2


def write_assumptions_sheet(
    wb: Workbook,
    rows: list[AssumptionRow],
    *,
    ticker: str,
    narrative: str,
    reasoning: str,
    baseline: OpusBaseline | None,
) -> None:
    """(Re)write the read-only Assumptions sheet directly after the Dashboard.

    Current values are live cross-sheet formulas (they track edits made after
    this write); Source / Opus value / Note are static, refreshed by every
    build/refresh — the subtitle says so.
    """
    if ASSUMPTIONS_SHEET in wb.sheetnames:
        idx = wb.sheetnames.index(ASSUMPTIONS_SHEET)
        wb.remove(wb[ASSUMPTIONS_SHEET])
    else:
        names = wb.sheetnames
        idx = (
            names.index(redesign_mod.DASHBOARD_SHEET) + 1
            if redesign_mod.DASHBOARD_SHEET in names
            else len(names)
        )
    ws = wb.create_sheet(ASSUMPTIONS_SHEET, idx)
    ws.sheet_view.showGridLines = False
    for j, w in enumerate(_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(1 + j)].width = w

    ws.cell(1, 1, f"{ticker} — Assumptions & Reasoning").font = _TITLE_FONT
    ws.cell(
        2,
        1,
        "Provenance of every Dashboard input. Read-only — regenerated by each "
        "build/refresh. Current is a live link to the Dashboard; Source / Opus "
        "value / Note are as of the last refresh_dcf run.",
    ).font = _SUB_FONT

    row = 4
    if baseline is not None:
        provenance = (
            f"Opus pass {baseline.as_of}"
            if not baseline.seeded
            else (
                f"seeded {baseline.as_of} from the synced assumptions block "
                "(pass predates provenance tracking — may already include your edits)"
            )
        )
        ws.cell(row, 1, "Assumptions baseline").font = _HDR_FONT
        ws.cell(row, 2, provenance)
        row += 2
    else:
        ws.cell(row, 1, "Assumptions baseline").font = _HDR_FONT
        ws.cell(row, 2, "no Opus assumptions file — every input is a builder default")
        row += 2

    if narrative:
        _band(ws, row, "THE STORY (Opus narrative)")
        row = _prose_block(ws, row + 1, narrative, n_rows=4)
    if reasoning:
        _band(ws, row, "KEY JUDGMENTS (Opus reasoning)")
        row = _prose_block(ws, row + 1, reasoning, n_rows=4)

    _band(ws, row, "DASHBOARD INPUTS — value · source · reasoning")
    row += 1
    for j, h in enumerate(_COLS):
        ws.cell(row, 1 + j, h).font = _HDR_FONT
    row += 1
    for r in rows:
        ws.cell(row, 1, r.label)
        ws.cell(row, 2, r.cell)
        cur = ws.cell(row, 3, f"=Dashboard!{r.cell}")
        cur.font = _LINK_FONT
        if r.fmt:
            cur.number_format = r.fmt
        src = ws.cell(row, 4, r.source)
        if r.source == SOURCE_USER:
            src.font = _USER_FONT
        ws.cell(row, 5, r.baseline_display)
        note = ws.cell(row, 6, r.note)
        note.alignment = _WRAP
        row += 1


def annotate_dashboard_comments(wb: Workbook, rows: list[AssumptionRow]) -> None:
    """Per-cell provenance as native comments on the yellow Dashboard inputs
    (Sheets shows them as notes after an export). Replaced on every write."""
    dsh = wb[redesign_mod.DASHBOARD_SHEET]
    for r in rows:
        if r.source == SOURCE_USER and r.note:
            text = f"User-edited — {r.note}"
        elif r.source == SOURCE_OPUS:
            text = f"Opus: {r.baseline_display}"
            if r.note:
                text += f"\n{r.note}"
        else:
            text = r.source if not r.note else f"{r.source} — {r.note}"
        dsh[r.cell].comment = Comment(text, _COMMENT_AUTHOR)


def write_provenance_into(
    wb: Workbook,
    inp: redesign_mod.RedesignInputs,
    assumptions_path: Path,
    *,
    ticker: str,
    update_ledger: bool,
    today: date | None = None,
) -> dict[str, int]:
    """Provenance onto an already-open workbook (the builder's path — it saves
    once at the end): ensure the baseline, reconcile the override ledger
    (refresher only — a from-scratch build reads the same JSON its inputs came
    from, so divergence there is already recorded), then rewrite the sheet +
    cell comments. Does NOT save the workbook.

    Returns per-source row counts (the refresher surfaces them in its result).
    Raises ``ProvenanceError`` when the assumptions JSON exists but cannot be
    read or updated — never silently skips.
    """
    baseline = ensure_opus_baseline(assumptions_path, today=today)
    data = _load_assumptions(assumptions_path)
    if update_ledger:
        ledger = update_override_ledger(assumptions_path, inp, baseline, today=today)
    else:
        ledger = _parse_ledger(data.get("assumption_overrides") if data is not None else None)

    block = data.get("redesign") if data is not None else None
    block_d: dict[str, object] = cast("dict[str, object]", block) if isinstance(block, dict) else {}
    origin = block_d.get("set_by")
    narrative = block_d.get("narrative")
    if not isinstance(narrative, str) or not narrative:
        top = data.get("narrative") if data is not None else None
        narrative = top if isinstance(top, str) else ""
    reasoning = block_d.get("reasoning")
    reasoning = reasoning if isinstance(reasoning, str) else ""

    rows = build_assumption_rows(
        wb, inp, baseline, ledger, block_origin=origin if isinstance(origin, str) else None
    )
    write_assumptions_sheet(
        wb, rows, ticker=ticker, narrative=narrative, reasoning=reasoning, baseline=baseline
    )
    annotate_dashboard_comments(wb, rows)

    counts: dict[str, int] = {}
    for r in rows:
        counts[r.source] = counts.get(r.source, 0) + 1
    return counts


def write_provenance(
    workbook_path: Path,
    inp: redesign_mod.RedesignInputs,
    assumptions_path: Path,
    *,
    ticker: str,
    update_ledger: bool,
    today: date | None = None,
) -> dict[str, int]:
    """``write_provenance_into`` against a workbook on disk (the refresher's
    path: capture → rebuild → inject → recompute → THIS → swap into place)."""
    wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    try:
        counts = write_provenance_into(
            wb,
            inp,
            assumptions_path,
            ticker=ticker,
            update_ledger=update_ledger,
            today=today,
        )
        wb.save(str(workbook_path))
    finally:
        wb.close()
    return counts
