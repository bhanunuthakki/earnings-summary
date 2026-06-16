"""Inject a picked DIY fact as a DCF *reference sheet* (capture-every-number S7).

The non-driver counterpart of :mod:`dcf.fact_drivers` (S6). Where a driver flows a
fact's value into the FCFF model and reprices, a *reference* fact lands as a
self-contained note attached to the ticker's DCF — value + unit + period +
source/provenance — so a number the owner found while slicing data is parked
alongside the valuation for context, without touching any assumption cell.

WHY A SEPARATE COMPANION WORKBOOK (the S7 decision, with rationale)
------------------------------------------------------------------
The genuinely hard part of S7 is *survival*. The canonical model workbook
``dcf/<T>.xlsx`` is rebuilt FROM SCRATCH on every refresh:
``execution/refresh_dcf._refresh_redesign`` runs the builder into a fresh
``openpyxl.Workbook()`` (``execution/build_redesigned_dcf.py``), then
``redesign.inject_dashboard`` re-applies ONLY the Dashboard yellow cells before
``os.replace`` swaps the rebuild over the live file. Any sheet appended to the
main workbook is therefore DESTROYED on the next refresh — there is no
preservation seam for non-Dashboard sheets.

Two ways to make an injected reference survive were considered:

* **(a) capture/reinject an appended sheet inside the rebuild** — add a new
  capture+reinject pass for a whole non-Dashboard sheet, threaded through
  ``_refresh_redesign`` next to the Dashboard one. This is net-new behavior in
  the most fragile part of the system; a future builder change or sheet-count
  assumption could silently wipe it, and openpyxl cell/style copying is fiddly.

* **(b) a separate companion workbook the refresh NEVER rebuilds** — CHOSEN.
  Survival is *structural*, not behavioral: the refresh path neither reads nor
  writes this file, so there is nothing to get subtly wrong. It also keeps the
  reference reference-only (literal values + provenance, no formula links the
  offline engine couldn't evaluate anyway — formula-feeding is S6's job).

The companion lives in a SUBDIRECTORY — ``dcf/facts/<T>.xlsx`` — not a sibling
``dcf/<T>_facts.xlsx``. ``dcf/*.xlsx`` is globbed by
``build_all_redesigned_dcf.default_tickers`` and ``pipeline.dcf_coverage_panel``,
both of which would mis-read a sibling as a ticker named ``<T>_FACTS``. Those
globs are non-recursive (``glob("*.xlsx")``), so a subdirectory file is invisible
to them — the same reason ``dcf/redesign/`` already nests sample builds.

This module is pure (no ``execution/`` dependency, unit-tested in isolation); the
server route (``POST /api/dcf/inject-fact-sheet``) resolves the fact's latest
value via :func:`dcf.fact_drivers.resolve_fact_value` (so company-doc
``fact_overrides`` win — S2, exactly as the driver path does) and calls
:func:`upsert_fact` here.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

# The companion workbook's single sheet. Plain-prose name (not a redesign marker
# sheet) so it is never mistaken for the model itself.
SHEET_NAME = "Reference Facts"

# Column layout (1-indexed) + their header labels. ``token`` (col 2) is the
# upsert key — re-injecting the same metric updates its row in place rather than
# stacking duplicates.
_COL_LABEL = 1
_COL_TOKEN = 2
_COL_VALUE = 3
_COL_UNIT = 4
_COL_PERIOD = 5
_COL_SOURCE = 6
_COL_FACT_ID = 7
_COL_CAPTURED = 8
_HEADERS: tuple[tuple[int, str], ...] = (
    (_COL_LABEL, "Metric"),
    (_COL_TOKEN, "Token"),
    (_COL_VALUE, "Value"),
    (_COL_UNIT, "Unit"),
    (_COL_PERIOD, "Period end"),
    (_COL_SOURCE, "Source"),
    (_COL_FACT_ID, "Fact ID"),
    (_COL_CAPTURED, "Captured"),
)
_COL_WIDTHS: dict[int, int] = {1: 34, 2: 30, 3: 16, 4: 12, 5: 14, 6: 24, 7: 10, 8: 14}
_FIRST_DATA_ROW = 2


@dataclass(frozen=True, slots=True)
class ReferenceFact:
    """One picked fact parked on a ticker's DCF reference sheet — a faithful,
    self-contained snapshot (the value is stored in its native unit, NOT
    converted; reference, not a driver). ``token`` is the metric token that
    produced it (the upsert key + provenance back-link to the DIY picker)."""

    token: str
    label: str
    value: float
    unit: str | None
    period_end: str  # 'YYYY-MM-DD'
    source: str
    fact_id: int | None
    captured_on: str  # 'YYYY-MM-DD'


def facts_workbook_path(repo_root: Path, ticker: str) -> Path:
    """The companion reference workbook for a ticker: ``dcf/facts/<T>.xlsx``.

    Nested under ``dcf/facts/`` so it is invisible to the non-recursive
    ``dcf/*.xlsx`` globs that enumerate ticker workbooks (see module docstring).
    """
    return repo_root / "dcf" / "facts" / f"{ticker.upper()}.xlsx"


def _text(ws: Worksheet, row: int, col: int) -> str | None:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    return str(v).strip() or None


def _ensure_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """Return the (created-if-absent) reference sheet with a styled header row."""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    # A brand-new workbook ships with one default sheet; reuse it as ours.
    ws = wb.active
    if ws is not None and ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
        ws.title = SHEET_NAME
    else:
        ws = wb.create_sheet(SHEET_NAME)
    for col, label in _HEADERS:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    return ws


def _find_row(ws: Worksheet, token: str) -> int | None:
    """The row carrying ``token`` (col 2), or None — the upsert lookup."""
    for row in range(_FIRST_DATA_ROW, ws.max_row + 1):
        if _text(ws, row, _COL_TOKEN) == token:
            return row
    return None


def _write_row(ws: Worksheet, row: int, fact: ReferenceFact) -> None:
    ws.cell(row=row, column=_COL_LABEL, value=fact.label)
    ws.cell(row=row, column=_COL_TOKEN, value=fact.token)
    ws.cell(row=row, column=_COL_VALUE, value=fact.value)
    ws.cell(row=row, column=_COL_UNIT, value=fact.unit)
    ws.cell(row=row, column=_COL_PERIOD, value=fact.period_end)
    ws.cell(row=row, column=_COL_SOURCE, value=fact.source)
    ws.cell(row=row, column=_COL_FACT_ID, value=fact.fact_id)
    ws.cell(row=row, column=_COL_CAPTURED, value=fact.captured_on)


def upsert_fact(path: Path, fact: ReferenceFact) -> dict[str, object]:
    """Write ``fact`` into the companion workbook at ``path``, creating the file
    + sheet if needed and updating the row in place when ``fact.token`` already
    exists (never stacking duplicates).

    Returns ``{"action": "added"|"updated", "row": <int>, "count": <int>}``.
    Self-contained literal values only — no formula links to the model (the
    offline engine can't evaluate them; that is the driver path's job).
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        try:
            wb = openpyxl.load_workbook(str(path))
        except (OSError, KeyError, ValueError, InvalidFileException):
            # A corrupt/foreign file is replaced rather than blocking the inject.
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()
    try:
        ws = _ensure_sheet(wb)
        existing = _find_row(ws, fact.token)
        if existing is not None:
            action, row = "updated", existing
        else:
            action, row = "added", _next_data_row(ws)
        _write_row(ws, row, fact)
        count = _data_row_count(ws)
        wb.save(str(path))
    finally:
        wb.close()
    return {"action": action, "row": row, "count": count}


def _next_data_row(ws: Worksheet) -> int:
    """The first empty data row (the append target)."""
    row = _FIRST_DATA_ROW
    while _text(ws, row, _COL_TOKEN) is not None:
        row += 1
    return row


def _data_row_count(ws: Worksheet) -> int:
    return sum(
        1
        for row in range(_FIRST_DATA_ROW, ws.max_row + 1)
        if _text(ws, row, _COL_TOKEN) is not None
    )


def read_facts(path: Path) -> list[ReferenceFact]:
    """Read every reference fact back from the companion workbook (ascending row
    order). Returns ``[]`` for a missing/corrupt file or one without the sheet."""
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except (OSError, KeyError, ValueError, InvalidFileException):
        return []
    try:
        if SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[SHEET_NAME]
        out: list[ReferenceFact] = []
        for row in range(_FIRST_DATA_ROW, ws.max_row + 1):
            token = _text(ws, row, _COL_TOKEN)
            if token is None:
                continue
            out.append(
                ReferenceFact(
                    token=token,
                    label=_text(ws, row, _COL_LABEL) or token,
                    value=_cell_float(ws, row, _COL_VALUE),
                    unit=_text(ws, row, _COL_UNIT),
                    period_end=_text(ws, row, _COL_PERIOD) or "",
                    source=_text(ws, row, _COL_SOURCE) or "",
                    fact_id=_cell_int(ws, row, _COL_FACT_ID),
                    captured_on=_text(ws, row, _COL_CAPTURED) or "",
                )
            )
        return out
    finally:
        wb.close()


def _cell_float(ws: Worksheet, row: int, col: int) -> float:
    v = ws.cell(row=row, column=col).value
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else 0.0


def _cell_int(ws: Worksheet, row: int, col: int) -> int | None:
    v = ws.cell(row=row, column=col).value
    return int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None
