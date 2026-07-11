"""Read a redesigned 9-sheet DCF workbook and recompute its value-of-record.

The redesigned workbook (``execution/build_redesigned_dcf.py``) is fully
formula-driven: openpyxl cannot evaluate its formulas, so this module reads the
*inputs* instead — the yellow Dashboard assumption cells (the user-owned layer)
plus the blue Financials actuals — and recomputes fair value in Python,
mirroring the in-sheet formulas exactly and bridging through
``dcf.valuation.compute_valuation``.

The projection here is the live counterpart of the builder's ``_project``
mirror: same segment-growth fade (convex near→terminal, see
``GROWTH_FADE_CURVATURE``), same margin ramp, same FCFF bridge, same
EV-multiple / perpetuity terminal, same ``× FX`` conversion for non-USD
reporters. The difference is that this reads
*arbitrary* (possibly user-edited) Dashboard inputs, so the persisted value
tracks whatever the user sees when they open the workbook.

Also provides the edit-preservation primitives the refresher uses:
``capture_dashboard`` snapshots the user-owned Dashboard inputs before a rebuild;
``inject_dashboard`` writes them back into a freshly-rebuilt workbook (refreshing
current price from the live quote).

Scenarios + sensitivity (S6): the Dashboard carries a SCENARIOS block whose
Bull/Bear columns are DELTAS vs the Base yellow cells (``ScenarioDeltas``;
seeded with the documented ``BULL_SEED``/``BEAR_SEED`` offsets, user-editable,
preserved across rebuilds). ``scenario_values`` computes all three fair values;
``sensitivity_grid`` computes the WACC × exit-multiple grid; both are written as
STATIC cells (``write_computed_outputs``) since the in-sheet formula chain only
evaluates the Base inputs.
"""

from __future__ import annotations

import dataclasses
import re
from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from dcf import valuation as val_mod

# --------------------------------------------------------------------------- #
# Structural constants — must track the builder's layout (the DB map + _project)
# --------------------------------------------------------------------------- #
DASHBOARD_SHEET = "Dashboard"
FINANCIALS_SHEET = "Financials"
CONSENSUS_SHEET = "Consensus"
VALUATION_SHEET = "Valuation"
MODEL_SHEET = "Model"
# Sheets that uniquely identify the redesigned format (vs the legacy
# Historicals/Forecast/Valuation seeder workbook).
_REDESIGN_MARKER_SHEETS = (DASHBOARD_SHEET, MODEL_SHEET, FINANCIALS_SHEET, CONSENSUS_SHEET)

N_FC = 10  # forecast years (matches the builder)
NWC_PCT = 0.005  # working-capital draw on incremental revenue (builder literal)

# Growth fade shape. Each segment's growth decays from its near-term rate to its
# terminal rate as ``g_t = g_T + (g_1 - g_T) * ((N-1-j)/(N-1))**CURVATURE``.
# CURVATURE = 1.0 is a straight-line glide; > 1.0 is CONVEX — deceleration is
# front-loaded (fast early, flattening late), the empirical shape of how
# hyper-growth fades (the law of large numbers bites hardest while the base is
# small). This constant is the DEFAULT; the builder calibrates a per-name value
# (``dcf.fade_calibration``) so each name's revenue path best-fits its consensus
# and writes it to the Dashboard, which this engine reads back — so the curvature
# is a real driver (``RedesignInputs.growth_fade_curvature``), defaulting here
# only when no consensus / no override pins it. The builder writes the identical
# closed form into the Model sheet, so the in-cell value and this mirror never drift.
GROWTH_FADE_CURVATURE = 2.0

SEG_ROW0 = 20  # first Dashboard segment row
SEG_ROW_MAX = 27  # segments occupy rows 20..27 (the PROFITABILITY band sits at 28)

# Dashboard column-B input cells at fixed addresses (the builder's DB map).
_DB_MARGIN_NEAR = 29
_DB_MARGIN_TERM = 30
_DB_TAX = 31
_DB_CAPEX26 = 34
_DB_TERM_CAPEX_DA = 35
_DB_RF = 38
_DB_ERP = 39
_DB_BETA = 40
_DB_KD = 41
_DB_METHOD = 43
_DB_BASIS = 44
_DB_MULT = 45
_DB_TG = 46
# Country risk premium (revenue-weighted Damodaran CRP) lives at row 47 — a free
# row between Terminal growth (46) and Current price (48), so no other Dashboard
# address shifts. Optional: a pre-CRP workbook with no row 47 reads back as 0.0.
_DB_CRP = 47
_DB_PRICE = 48
# Growth-fade curvature (per-name, calibrated to consensus by the builder) at the
# free row 49. Optional: a pre-curvature workbook reads back as the default 2.0.
_DB_CURV = 49

# SCENARIOS block (Dashboard rows 50+): Bull/Bear expressed as DELTAS vs the Base
# yellow cells, one lever per row. Column C = Bull Δ, column D = Bear Δ (both
# yellow inputs); row 58 carries the three Python-computed fair values (static —
# rewritten by every rebuild/refresh, like the Sensitivity grid).
SCEN_ROW_GROWTH_NEAR = 52  # Δ on every segment's near-term growth (pts)
SCEN_ROW_GROWTH_TERM = 53  # Δ on every segment's terminal growth (pts)
SCEN_ROW_MARGIN_NEAR = 54  # Δ on near-term operating margin (pts)
SCEN_ROW_MARGIN_TERM = 55  # Δ on terminal operating margin (pts)
SCEN_ROW_EXIT_MULT = 56  # Δ on the exit multiple (turns)
SCEN_ROW_TG = 57  # Δ on terminal growth g (pts)
SCEN_FV_ROW = 58  # computed fair value / share: B=Base, C=Bull, D=Bear
SCEN_COL_BULL = 3  # column C
SCEN_COL_BEAR = 4  # column D
_SCEN_DELTA_ROWS: tuple[int, ...] = (
    SCEN_ROW_GROWTH_NEAR,
    SCEN_ROW_GROWTH_TERM,
    SCEN_ROW_MARGIN_NEAR,
    SCEN_ROW_MARGIN_TERM,
    SCEN_ROW_EXIT_MULT,
    SCEN_ROW_TG,
)

# Scenario PROBABILITY WEIGHTS (Dashboard row 62, below the SCENARIOS block's
# fair-value/upside rows 58-59). Three yellow cells the owner can edit — Base(B) /
# Bull(C) / Bear(D), matching the row-51 column headers — seeded from the LLM
# scenario_prior (or the symmetric default) and preserved across refresh like the
# other yellow inputs (owner edit wins). ``dcf.scenario_reward`` consumes them
# (PR E) with the global 25/50/25 as the fallback. A pre-weights workbook has no
# row 62, so the weights read back as the symmetric default.
SCEN_WEIGHTS_LABEL_ROW = 61
SCEN_WEIGHTS_ROW = 62
SCEN_COL_WEIGHT_BASE = 2  # column B (matches the "Base" header column)
DEFAULT_SCENARIO_WEIGHTS: dict[str, float] = {"bull": 0.25, "base": 0.50, "bear": 0.25}

# STOCK-BASED COMPENSATION band (Dashboard rows 64+). Two yellow owner inputs — the
# near-term and terminal SBC % of revenue — sit BELOW the SCENARIOS block so no
# existing Dashboard address (segments 20-27, scalars 29-49, SCENARIOS 50-62)
# shifts. SBC is charged as an explicit after-tax expense in ``_project`` (the
# operating margins stay NON-GAAP / SBC-excluded, comparable across names) and the
# terminal EBITDA is burdened by SBC before the exit multiple applies. Optional: a
# pre-SBC workbook has no rows 65/66, so both read back as 0.0 — i.e. exactly the
# old "SBC never charged" behaviour, keeping legacy workbooks/fixtures unchanged.
_DB_SBC_BAND = 64
_DB_SBC_NEAR = 65
_DB_SBC_TERM = 66

SENSITIVITY_SHEET = "Sensitivity"

# The numeric scalar inputs (Dashboard col B) preserved across a refresh, keyed by
# row. Current price (row 48) is deliberately excluded — it is a market datum the
# refresher always refreshes from the live quote, not a user assumption.
_PRESERVED_SCALAR_ROWS: tuple[int, ...] = (
    _DB_MARGIN_NEAR,
    _DB_MARGIN_TERM,
    _DB_TAX,
    _DB_CAPEX26,
    _DB_TERM_CAPEX_DA,
    _DB_RF,
    _DB_ERP,
    _DB_BETA,
    _DB_KD,
    _DB_CRP,
    _DB_MULT,
    _DB_TG,
    _DB_SBC_NEAR,
    _DB_SBC_TERM,
)

PERPETUITY = "Perpetuity"
EXIT_MULTIPLE = "Exit multiple"


class RedesignError(Exception):
    """The workbook is not a readable redesigned-format DCF, or carries a
    degenerate assumption the engine cannot value."""


# --------------------------------------------------------------------------- #
# Dataclasses
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ScenarioDeltas:
    """A Bull/Bear scenario as offsets applied on top of the Base inputs.

    Growth deltas apply uniformly to EVERY segment's near-term/terminal growth
    (the same uniform-shift mechanic the builder uses to reconcile segment
    momentum to consensus); margin/multiple/g deltas shift the corresponding
    Dashboard scalar. Expressing scenarios as deltas keeps them meaningful when
    the user later edits a Base cell — Bull stays "Base + 2pts margin" rather
    than a stale absolute.
    """

    growth_near: float = 0.0
    growth_term: float = 0.0
    margin_near: float = 0.0
    margin_term: float = 0.0
    exit_multiple: float = 0.0
    terminal_g: float = 0.0

    @classmethod
    def from_dict(cls, data: Mapping[str, object]) -> ScenarioDeltas:
        """Reconstruct a ScenarioDeltas from a JSON-decoded dict (the inverse of
        ``dataclasses.asdict``). A missing lever defaults to 0.0 (no shift);
        seed fallback for an absent *whole* block is the caller's concern
        (``RedesignInputs.from_dict``). Raises ``RedesignError`` on a non-numeric
        lever."""

        def lever(key: str) -> float:
            v = data.get(key, 0.0)
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise RedesignError(f"scenario delta {key!r} must be a number")
            return float(v)

        return cls(
            growth_near=lever("growth_near"),
            growth_term=lever("growth_term"),
            margin_near=lever("margin_near"),
            margin_term=lever("margin_term"),
            exit_multiple=lever("exit_multiple"),
            terminal_g=lever("terminal_g"),
        )


# The documented seed offsets a fresh build writes into the Bull/Bear columns —
# a deliberately moderate spread the user then edits per name.
BULL_SEED = ScenarioDeltas(
    growth_near=0.03,
    growth_term=0.01,
    margin_near=0.01,
    margin_term=0.02,
    exit_multiple=2.0,
    terminal_g=0.005,
)
BEAR_SEED = ScenarioDeltas(
    growth_near=-0.03,
    growth_term=-0.01,
    margin_near=-0.01,
    margin_term=-0.02,
    exit_multiple=-2.0,
    terminal_g=-0.005,
)


# --------------------------------------------------------------------------- #
# Thesis-calibrated bear override (Monthly Red Team Phase 1, guard 3).
#
# BEAR_SEED is a generic "mild disappointment" offset — the same -3pt growth /
# -1pt margin / -2x exit shift for every name, regardless of what the analyst's
# own thesis says would actually break it. ``micro_thesis/holdings/<T>.json`` may
# carry an optional ``bear_deltas`` block (documented in
# ``directives/holdings_json_schema.md``) that overrides the SEED default (never
# an owner's workbook edit — that machinery, ``scenario_bull``/``scenario_bear``
# in ``data/dcf_assumptions/<T>.json`` mirrored by
# ``refresh_dcf._apply_inputs_to_block``, is untouched and still wins whenever it
# is on file). The holdings schema is intentionally simpler than the workbook's
# 6-lever ``ScenarioDeltas`` — one growth/margin/exit-multiple/terminal-g pp-or-
# turns delta each, applied uniformly to the near+term legs — so an analyst names
# the thesis-break magnitude without hand-deriving six numbers.
# --------------------------------------------------------------------------- #


def parse_thesis_bear_deltas(holdings: Mapping[str, object] | None) -> ScenarioDeltas | None:
    """The holdings JSON's thesis-calibrated ``bear_deltas`` override, or
    ``None`` when absent/malformed (a name with no override, or a block with no
    recognized numeric lever, is exactly ``None`` — the caller falls back to
    ``BEAR_SEED``). ``holdings`` is the already-loaded ``micro_thesis/holdings/
    <T>.json`` dict — this function does no I/O.

    Recognized keys (all optional, in percentage points except the exit
    multiple, which is in turns): ``growth_delta_pp`` (shifts BOTH near- and
    terminal-segment growth), ``margin_delta_pp`` (shifts both near- and
    terminal operating margin), ``exit_multiple_delta`` (shifts the exit
    multiple), ``terminal_g_delta_pp`` (shifts terminal growth g). A lever the
    block doesn't set falls back to ``BEAR_SEED``'s value for that lever, so a
    thesis only needs to name the specific break, not re-derive the whole range.
    """
    if not isinstance(holdings, Mapping):
        return None
    raw_obj = holdings.get("bear_deltas")
    if not isinstance(raw_obj, Mapping):
        return None
    raw = cast("Mapping[str, object]", raw_obj)

    def _num(key: str) -> float | None:
        v = raw.get(key)
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return None
        return float(v)

    growth_pp = _num("growth_delta_pp")
    margin_pp = _num("margin_delta_pp")
    exit_delta = _num("exit_multiple_delta")
    term_g_pp = _num("terminal_g_delta_pp")
    if growth_pp is None and margin_pp is None and exit_delta is None and term_g_pp is None:
        return None  # block present but no recognized lever — not a real override

    growth = growth_pp / 100.0 if growth_pp is not None else BEAR_SEED.growth_near
    margin = margin_pp / 100.0 if margin_pp is not None else BEAR_SEED.margin_near
    exit_multiple = exit_delta if exit_delta is not None else BEAR_SEED.exit_multiple
    terminal_g = term_g_pp / 100.0 if term_g_pp is not None else BEAR_SEED.terminal_g
    return ScenarioDeltas(
        growth_near=growth,
        growth_term=growth,
        margin_near=margin,
        margin_term=margin,
        exit_multiple=exit_multiple,
        terminal_g=terminal_g,
    )


def thesis_bear_seed(holdings: Mapping[str, object] | None) -> ScenarioDeltas:
    """``BEAR_SEED``, overridden by the holdings JSON's thesis-calibrated
    ``bear_deltas`` when present — the default a FRESH scenario build (no owner
    workbook edit on file yet) seeds the Bear column from."""
    return parse_thesis_bear_deltas(holdings) or BEAR_SEED


def classify_bear_provenance(
    bear_deltas: ScenarioDeltas, holdings: Mapping[str, object] | None
) -> str:
    """``"seed"`` | ``"thesis"`` | ``"owner"`` for the bear deltas actually in use
    (a run's ``inp.bear_deltas``, read back from the workbook), by comparing
    against ``BEAR_SEED`` and the holdings JSON's thesis override (if any).

    Pure — no I/O; the caller has already loaded ``holdings``. ``"owner"`` is
    the catch-all: a workbook Dashboard edit away from both the generic seed and
    the thesis override is, by construction, a hand edit."""
    if bear_deltas == BEAR_SEED:
        return "seed"
    thesis = parse_thesis_bear_deltas(holdings)
    if thesis is not None and bear_deltas == thesis:
        return "thesis"
    return "owner"


@dataclass(frozen=True)
class RedesignInputs:
    """Everything ``value()`` needs, read from a redesigned workbook.

    Flows/stocks are in the workbook's reporting currency, millions; shares are
    millions; ``fx_to_usd`` converts the per-share value to USD (1.0 for USD
    reporters). ``consensus_years`` is the margin-ramp window (the builder's
    ``ncons``).
    """

    segments: tuple[str, ...]
    base_revenue_by_segment: dict[str, float]
    near_growth_by_segment: dict[str, float]
    terminal_growth_by_segment: dict[str, float]
    near_op_margin: float
    terminal_op_margin: float
    tax_rate: float
    capex_2026_m: float
    terminal_capex_da: float
    da_ratio: float
    consensus_years: int
    wacc: float
    beta: float
    risk_free_rate: float
    equity_risk_premium: float
    cost_of_debt: float
    terminal_method: str
    terminal_basis: str
    exit_multiple: float
    terminal_growth_g: float
    current_price: float
    cash_m: float
    total_debt_m: float
    diluted_shares_m: float
    fx_to_usd: float
    # Operation-weighted Damodaran country risk premium, ADDED to the CAPM cost
    # of equity (ke = rf + beta*erp + crp). 0.0 for a US/mature name. Defaulted
    # so a pre-CRP workbook (or a test fixture) reads back as 0.0 — i.e. exactly
    # the old behaviour. The discount rate of record is ``wacc`` (already
    # CRP-inclusive once the builder/reader fold it in), so ``value()`` reads
    # ``wacc`` directly; this field is carried for the workbook round-trip and
    # for ``read_inputs``' CAPM re-derivation.
    country_risk_premium: float = 0.0
    # Per-name growth-fade curvature, calibrated to consensus by the builder
    # (``dcf.fade_calibration``). Defaults to the module constant so a
    # pre-curvature workbook / test fixture reproduces the old global behaviour.
    growth_fade_curvature: float = GROWTH_FADE_CURVATURE
    # Stock-based compensation charged as an explicit % of revenue (Dashboard rows
    # 65/66, yellow — owner-editable). Non-GAAP operating margins EXCLUDE SBC (they
    # stay comparable across names), so ``_project`` charges an after-tax SBC
    # expense — ``sbc_pct * revenue * (1 - tax)`` — and ``_terminal_metrics``
    # burdens the exit-multiple EBITDA by SBC. The near-term rate applies in the
    # first forecast year and fades linearly to the terminal rate over the window,
    # mirroring the builder's ``sbc_pct`` fade. Defaults to 0.0 so a pre-SBC
    # workbook / fixture reproduces the old "SBC never charged" behaviour exactly.
    near_sbc_pct: float = 0.0
    terminal_sbc_pct: float = 0.0
    # Bull/Bear scenario offsets (Dashboard SCENARIOS block). Frozen instances
    # are immutable, so the shared seed defaults are safe; a pre-scenario
    # workbook reads back as the seeds, giving every name a meaningful range.
    bull_deltas: ScenarioDeltas = BULL_SEED
    bear_deltas: ScenarioDeltas = BEAR_SEED
    # Scenario PROBABILITY weights (Dashboard row 62, yellow — owner-editable,
    # seeded from the LLM scenario_prior). Default to the symmetric global prior so
    # a pre-weights workbook reads back as today's 25/50/25 behaviour. Normalized to
    # sum 1 at read; ``dcf.scenario_reward`` consumes them (PR E).
    weight_bull: float = 0.25
    weight_base: float = 0.50
    weight_bear: float = 0.25

    def to_dict(self) -> dict[str, object]:
        """A JSON-safe dict of every input — the inverse of :meth:`from_dict`.

        The transport for the in-app recompute route (``/api/dcf/recompute``):
        segments as a list and the Bull/Bear ``ScenarioDeltas`` as plain dicts,
        so the whole assumption set round-trips through ``json.dumps`` and back
        with NO xlsx in the loop. The card edits a few fields and POSTs the
        result; the server rebuilds via :meth:`from_dict` and runs the pure
        engine.
        """
        payload: dict[str, object] = dataclasses.asdict(self)
        payload["segments"] = list(self.segments)
        return payload

    @classmethod
    def from_dict(cls, data: Mapping[str, object]) -> RedesignInputs:
        """Build a :class:`RedesignInputs` from a JSON-decoded dict — the
        non-workbook constructor (the inverse of :meth:`to_dict`).

        Where :func:`read_inputs` reads the yellow Dashboard cells and *derives*
        WACC from the CAPM inputs, this takes every value as given: the in-app
        card edits WACC (and the terminal/margin/growth levers) directly, then
        POSTs the edited set so the pure engine recomputes with no Sheets/Excel
        round-trip. Raises :class:`RedesignError` with a field-named message on a
        missing, non-numeric, or structurally invalid input, mirroring the
        ``read_inputs`` validation contract.
        """

        def req_float(key: str) -> float:
            if key not in data:
                raise RedesignError(f"missing DCF input: {key!r}")
            v = data[key]
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise RedesignError(f"DCF input {key!r} must be a number")
            return float(v)

        def seg_map(key: str) -> dict[str, float]:
            raw = data.get(key)
            if not isinstance(raw, Mapping):
                raise RedesignError(f"DCF input {key!r} must be a segment->number object")
            out: dict[str, float] = {}
            for seg, v in cast("Mapping[str, object]", raw).items():
                if isinstance(v, bool) or not isinstance(v, (int, float)):
                    raise RedesignError(f"DCF input {key!r}[{seg!r}] must be a number")
                out[str(seg)] = float(v)
            return out

        raw_segs = data.get("segments")
        if not isinstance(raw_segs, (list, tuple)) or not raw_segs:
            raise RedesignError("DCF input 'segments' must be a non-empty list")
        segments = tuple(str(s) for s in cast("tuple[object, ...]", raw_segs))

        base_rev = seg_map("base_revenue_by_segment")
        near = seg_map("near_growth_by_segment")
        term = seg_map("terminal_growth_by_segment")
        for seg in segments:
            for label, m in (
                ("base_revenue_by_segment", base_rev),
                ("near_growth_by_segment", near),
                ("terminal_growth_by_segment", term),
            ):
                if seg not in m:
                    raise RedesignError(f"DCF input {label!r} missing segment {seg!r}")

        cy_raw = data.get("consensus_years")
        if isinstance(cy_raw, bool) or not isinstance(cy_raw, (int, float)):
            raise RedesignError("DCF input 'consensus_years' must be an integer")
        consensus_years = int(cy_raw)

        method = data.get("terminal_method")
        if not isinstance(method, str) or not method.strip():
            raise RedesignError("DCF input 'terminal_method' must be a non-empty string")
        basis = data.get("terminal_basis")
        if not isinstance(basis, str) or not basis.strip():
            raise RedesignError("DCF input 'terminal_basis' must be a non-empty string")

        crp_raw = data.get("country_risk_premium", 0.0)
        if isinstance(crp_raw, bool) or not isinstance(crp_raw, (int, float)):
            raise RedesignError("DCF input 'country_risk_premium' must be a number")
        country_risk_premium = float(crp_raw)

        curv_raw = data.get("growth_fade_curvature", GROWTH_FADE_CURVATURE)
        if isinstance(curv_raw, bool) or not isinstance(curv_raw, (int, float)):
            raise RedesignError("DCF input 'growth_fade_curvature' must be a number")
        growth_fade_curvature = float(curv_raw)

        def opt_sbc(key: str) -> float:
            v = data.get(key, 0.0)
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise RedesignError(f"DCF input {key!r} must be a number")
            return float(v)

        near_sbc_pct = opt_sbc("near_sbc_pct")
        terminal_sbc_pct = opt_sbc("terminal_sbc_pct")

        bull_raw = data.get("bull_deltas")
        bull = (
            ScenarioDeltas.from_dict(cast("Mapping[str, object]", bull_raw))
            if isinstance(bull_raw, Mapping)
            else BULL_SEED
        )
        bear_raw = data.get("bear_deltas")
        bear = (
            ScenarioDeltas.from_dict(cast("Mapping[str, object]", bear_raw))
            if isinstance(bear_raw, Mapping)
            else BEAR_SEED
        )

        def opt_weight(key: str, default: float) -> float:
            v = data.get(key, default)
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                return default
            return float(v)

        w_bull, w_base, w_bear = normalize_scenario_weights(
            opt_weight("weight_bull", DEFAULT_SCENARIO_WEIGHTS["bull"]),
            opt_weight("weight_base", DEFAULT_SCENARIO_WEIGHTS["base"]),
            opt_weight("weight_bear", DEFAULT_SCENARIO_WEIGHTS["bear"]),
        )

        return cls(
            segments=segments,
            base_revenue_by_segment=base_rev,
            near_growth_by_segment=near,
            terminal_growth_by_segment=term,
            near_op_margin=req_float("near_op_margin"),
            terminal_op_margin=req_float("terminal_op_margin"),
            tax_rate=req_float("tax_rate"),
            capex_2026_m=req_float("capex_2026_m"),
            terminal_capex_da=req_float("terminal_capex_da"),
            da_ratio=req_float("da_ratio"),
            consensus_years=consensus_years,
            wacc=req_float("wacc"),
            beta=req_float("beta"),
            risk_free_rate=req_float("risk_free_rate"),
            equity_risk_premium=req_float("equity_risk_premium"),
            cost_of_debt=req_float("cost_of_debt"),
            terminal_method=method.strip(),
            terminal_basis=basis.strip(),
            exit_multiple=req_float("exit_multiple"),
            terminal_growth_g=req_float("terminal_growth_g"),
            current_price=req_float("current_price"),
            cash_m=req_float("cash_m"),
            total_debt_m=req_float("total_debt_m"),
            diluted_shares_m=req_float("diluted_shares_m"),
            fx_to_usd=req_float("fx_to_usd"),
            country_risk_premium=country_risk_premium,
            growth_fade_curvature=growth_fade_curvature,
            near_sbc_pct=near_sbc_pct,
            terminal_sbc_pct=terminal_sbc_pct,
            bull_deltas=bull,
            bear_deltas=bear,
            weight_bull=w_bull,
            weight_base=w_base,
            weight_bear=w_bear,
        )


def normalize_scenario_weights(bull: float, base: float, bear: float) -> tuple[float, float, float]:
    """Normalize three raw scenario weights to a sum-1 simplex, or fall back to the
    symmetric default when they are non-positive/negative (a degenerate cell edit).

    Deliberately does NOT enforce the LLM's non-degenerate bounds (base>=0.30 etc.)
    — those gate the *model's* proposals (dcf.scenario_prior); the owner editing the
    workbook cells is trusted, so any positive simplex the owner types is honored."""
    if bull < 0 or base < 0 or bear < 0:
        return (
            DEFAULT_SCENARIO_WEIGHTS["bull"],
            DEFAULT_SCENARIO_WEIGHTS["base"],
            DEFAULT_SCENARIO_WEIGHTS["bear"],
        )
    s = bull + base + bear
    if s <= 0:
        return (
            DEFAULT_SCENARIO_WEIGHTS["bull"],
            DEFAULT_SCENARIO_WEIGHTS["base"],
            DEFAULT_SCENARIO_WEIGHTS["bear"],
        )
    return (bull / s, base / s, bear / s)


@dataclass(frozen=True)
class RedesignValuation:
    """The recomputed value-of-record and the headline intermediates."""

    value_per_share_usd: float
    value_per_share_reporting: float
    operating_value_usd_m: float  # enterprise value, USD millions
    equity_value_usd_m: float
    fcff_stream_m: list[float]  # valuation FCF, reporting ccy
    forecast_revenue_m: list[float]
    wacc: float
    terminal_method: str
    terminal_basis: str
    exit_multiple: float
    fx_to_usd: float
    diluted_shares_m: float
    cash_m: float
    total_debt_m: float
    current_price: float


@dataclass(frozen=True)
class CapturedDashboard:
    """The user-owned Dashboard inputs, snapshotted for edit-preservation.

    ``segment_growth`` is keyed by segment *name* (not row) so it survives FMP
    adding/removing/reordering a segment between refreshes; ``scalars`` is keyed
    by Dashboard row. ``scenario_deltas`` is keyed by ``(row, column)`` over the
    SCENARIOS block's Bull/Bear delta cells; a pre-scenario workbook captures an
    empty dict, so the rebuilt workbook keeps its seeded defaults. Current price
    is intentionally absent (refreshed live).
    """

    segment_growth: dict[str, tuple[float, float]]
    scalars: dict[int, float]
    terminal_method: str | None
    terminal_basis: str | None
    scenario_deltas: dict[tuple[int, int], float] = dataclasses.field(
        default_factory=dict[tuple[int, int], float]
    )
    # Scenario probability weights keyed "bull"/"base"/"bear" (Dashboard row 62 —
    # owner-editable yellow cells). Empty on a pre-weights workbook, so the rebuilt
    # workbook keeps its freshly-seeded weights. Owner edits win: captured before a
    # rebuild and re-injected after, exactly like the other yellow inputs.
    scenario_weights: dict[str, float] = dataclasses.field(default_factory=dict[str, float])


# --------------------------------------------------------------------------- #
# Small typed cell helpers (openpyxl values are untyped unions)
# --------------------------------------------------------------------------- #
def _num(ws: Worksheet, row: int, col: int) -> float | None:
    v = ws.cell(row=row, column=col).value
    if isinstance(v, bool):  # bool is an int subclass — never a numeric input
        return None
    return float(v) if isinstance(v, (int, float)) else None


def _text(ws: Worksheet, row: int, col: int) -> str | None:
    v = ws.cell(row=row, column=col).value
    return v.strip() if isinstance(v, str) else None


# --------------------------------------------------------------------------- #
# Format detection
# --------------------------------------------------------------------------- #
def is_redesign_format(workbook_path: Path) -> bool:
    """True iff the workbook carries the redesigned-format marker sheets.

    Cheap and tolerant: a missing/corrupt file is simply "not redesign".
    """
    if not workbook_path.exists():
        return False
    try:
        wb = openpyxl.load_workbook(str(workbook_path), read_only=True)
    except (OSError, KeyError, ValueError, InvalidFileException):
        return False
    try:
        names = set(wb.sheetnames)
    finally:
        wb.close()
    return all(s in names for s in _REDESIGN_MARKER_SHEETS)


# --------------------------------------------------------------------------- #
# Financials sheet reader (blue actuals)
# --------------------------------------------------------------------------- #
_QUARTER_RE = re.compile(r"Q([1-4])\s+(\d{4})")


def _quarter_columns(fs: Worksheet) -> dict[int, tuple[int, int]]:
    """Map Financials column index -> (fiscal_year, quarter) from the row-1 headers."""
    out: dict[int, tuple[int, int]] = {}
    for col in range(2, fs.max_column + 1):
        label = fs.cell(row=1, column=col).value
        if not isinstance(label, str):
            continue
        m = _QUARTER_RE.match(label.strip())
        if m:
            out[col] = (int(m.group(2)), int(m.group(1)))
    return out


def _detect_fy_quarters(quarters_by_year: dict[int, set[int]]) -> set[int]:
    """The quarter numbers that make up ONE fiscal year for this issuer.

    Generalises "all four quarters" to any consistent cadence — a semi-annual
    filer (e.g. BHP) shows only ``{2, 4}``, whose two half-year columns sum to the
    fiscal year. The cadence is the largest quarter-set recurring across >=2
    fiscal years (so the current partial year never defines it); falls back to
    ``{1, 2, 3, 4}`` when history is too short to establish one.
    """
    counts: dict[frozenset[int], int] = defaultdict(int)
    for qs in quarters_by_year.values():
        if qs:
            counts[frozenset(qs)] += 1
    recurring = [qs for qs, n in counts.items() if n >= 2]
    return set(max(recurring, key=len)) if recurring else {1, 2, 3, 4}


def _latest_full_fy(qcols: dict[int, tuple[int, int]]) -> int:
    """The most recent fiscal year carrying this issuer's full period set (all four
    quarters for a quarterly filer; both halves for a semi-annual one)."""
    quarters_by_year: dict[int, set[int]] = defaultdict(set)
    for _col, (year, q) in qcols.items():
        quarters_by_year[year].add(q)
    cadence = _detect_fy_quarters(quarters_by_year)
    full = [y for y, qs in quarters_by_year.items() if cadence <= qs]
    if not full:
        raise RedesignError("Financials sheet has no complete fiscal year")
    return max(full)


def _find_row(fs: Worksheet, label: str) -> int | None:
    """First column-A row whose stripped label equals ``label`` exactly."""
    for row in range(1, fs.max_row + 1):
        a = fs.cell(row=row, column=1).value
        if isinstance(a, str) and a.strip() == label:
            return row
    return None


def _fy_sum(fs: Worksheet, row: int, fy_cols: list[int]) -> float:
    """Sum a fiscal year's period columns for a row (four quarters, or two halves
    for a semi-annual filer), skipping blanks."""
    total = 0.0
    for col in fy_cols:
        v = _num(fs, row, col)
        if v is not None:
            total += v
    return total


def _latest_quarter_value(fs: Worksheet, row: int, last_col: int) -> float:
    """The latest-quarter value for a row; scan left if the last column is blank."""
    for col in range(last_col, 1, -1):
        v = _num(fs, row, col)
        if v is not None:
            return v
    return 0.0


# --------------------------------------------------------------------------- #
# Reading the full input set
# --------------------------------------------------------------------------- #
def _read_segments(dsh: Worksheet) -> tuple[list[str], dict[str, float], dict[str, float]]:
    """Read the Dashboard segment rows: names + near/terminal growth (yellow)."""
    names: list[str] = []
    near: dict[str, float] = {}
    term: dict[str, float] = {}
    for row in range(SEG_ROW0, SEG_ROW_MAX + 1):
        name = _text(dsh, row, 1)
        g_near = _num(dsh, row, 2)
        if not name or g_near is None:
            continue
        g_term = _num(dsh, row, 3)
        names.append(name)
        near[name] = g_near
        term[name] = g_term if g_term is not None else g_near
    if not names:
        raise RedesignError("Dashboard has no segment rows (expected at row 20+)")
    return names, near, term


def _consensus_window(wb: Workbook) -> int:
    """Number of consensus fiscal-year columns on the Consensus sheet (the
    builder's ``ncons``), clamped to [2, N_FC]."""
    cs = wb[CONSENSUS_SHEET]
    count = 0
    for col in range(2, N_FC + 2):
        v = cs.cell(row=2, column=col).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            count += 1
    return max(2, min(N_FC, count))


_FX_RE = re.compile(r"\*\s*([0-9]*\.?[0-9]+)\s*$")


def _read_fx(vs: Worksheet) -> float:
    """Parse the ``× FX`` multiplier off the Valuation 'VALUE / SHARE' formula.

    The builder writes ``=B<eq>/B<sh>*<fx>`` (e.g. ``*1.0`` for USD, ``*0.145``
    for a DKK reporter). The workbook is the single source of truth for FX, so we
    read it back rather than re-deriving a currency table here.
    """
    for row in range(1, vs.max_row + 1):
        a = vs.cell(row=row, column=1).value
        if isinstance(a, str) and a.strip() == "VALUE / SHARE":
            formula = vs.cell(row=row, column=2).value
            if isinstance(formula, str):
                m = _FX_RE.search(formula)
                if m:
                    return float(m.group(1))
    return 1.0


def read_inputs(workbook_path: Path) -> RedesignInputs | None:
    """Read every input ``value()`` needs from a redesigned workbook.

    Returns ``None`` if the workbook is not redesigned-format (so callers can
    fall back to the legacy reader during migration). Raises ``RedesignError``
    if the workbook *is* redesigned-format but structurally unreadable.
    """
    if not is_redesign_format(workbook_path):
        return None
    wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    try:
        dsh = wb[DASHBOARD_SHEET]
        fs = wb[FINANCIALS_SHEET]
        vs = wb[VALUATION_SHEET]

        segments, near, term = _read_segments(dsh)

        near_margin = _num(dsh, _DB_MARGIN_NEAR, 2)
        term_margin = _num(dsh, _DB_MARGIN_TERM, 2)
        tax = _num(dsh, _DB_TAX, 2)
        capex26 = _num(dsh, _DB_CAPEX26, 2)
        term_capex_da = _num(dsh, _DB_TERM_CAPEX_DA, 2)
        rf = _num(dsh, _DB_RF, 2)
        erp = _num(dsh, _DB_ERP, 2)
        beta = _num(dsh, _DB_BETA, 2)
        kd = _num(dsh, _DB_KD, 2)
        # Country risk premium is optional: a pre-CRP workbook has no row 47, so
        # a blank reads back as 0.0 (the old US-only behaviour) rather than failing.
        crp = _num(dsh, _DB_CRP, 2) or 0.0
        # Fade curvature is optional too: a pre-curvature workbook has no row 49,
        # so a blank reads back as the default (the old global behaviour).
        curv = _num(dsh, _DB_CURV, 2) or GROWTH_FADE_CURVATURE
        # SBC % inputs are optional: a pre-SBC workbook has no rows 65/66, so both
        # read back as 0.0 (the old "SBC never charged" behaviour).
        sbc_near = _num(dsh, _DB_SBC_NEAR, 2) or 0.0
        sbc_term = _num(dsh, _DB_SBC_TERM, 2) or 0.0
        exit_mult = _num(dsh, _DB_MULT, 2)
        tg = _num(dsh, _DB_TG, 2)
        price = _num(dsh, _DB_PRICE, 2)
        method = _text(dsh, _DB_METHOD, 2) or EXIT_MULTIPLE
        basis = _text(dsh, _DB_BASIS, 2) or "EV/EBITDA"
        bull = _read_scenario_deltas(dsh, SCEN_COL_BULL, BULL_SEED)
        bear = _read_scenario_deltas(dsh, SCEN_COL_BEAR, BEAR_SEED)
        w_bull, w_base, w_bear = _read_weights(dsh)

        required = {
            "near margin": near_margin,
            "terminal margin": term_margin,
            "tax": tax,
            "2026 capex": capex26,
            "terminal capex/D&A": term_capex_da,
            "risk-free rate": rf,
            "equity risk premium": erp,
            "beta": beta,
            "cost of debt": kd,
            "exit multiple": exit_mult,
            "terminal g": tg,
            "current price": price,
        }
        missing = [k for k, v in required.items() if v is None]
        if missing:
            raise RedesignError(f"Dashboard missing inputs: {', '.join(missing)}")

        qcols = _quarter_columns(fs)
        if not qcols:
            raise RedesignError("Financials sheet has no quarter columns")
        ly = _latest_full_fy(qcols)
        fy_cols = [c for c, (y, _q) in qcols.items() if y == ly]
        last_col = max(qcols)

        rev_row = _find_row(fs, "Revenue")
        da_row = _find_row(fs, "D&A")
        if rev_row is None or da_row is None:
            raise RedesignError("Financials sheet missing Revenue or D&A row")
        rev_ly = _fy_sum(fs, rev_row, fy_cols)
        if rev_ly <= 0:
            raise RedesignError(f"non-positive last-FY revenue ({rev_ly}) on Financials")
        da_ratio = _fy_sum(fs, da_row, fy_cols) / rev_ly

        base_rev: dict[str, float] = {}
        for s in segments:
            seg_row = _find_row(fs, s)
            # A single-segment model ("Total company") has no segment row — its
            # base revenue is total revenue, exactly as the builder's pseg_fy does.
            base_rev[s] = _fy_sum(fs, seg_row, fy_cols) if seg_row is not None else rev_ly

        cash_row = _find_row(fs, "Cash & ST Investments")
        debt_row = _find_row(fs, "Long-term Debt")
        shares_row = _find_row(fs, "Diluted Shares (M)")
        if cash_row is None or debt_row is None or shares_row is None:
            raise RedesignError("Financials sheet missing cash / debt / shares row")
        cash = _latest_quarter_value(fs, cash_row, last_col)
        debt = _latest_quarter_value(fs, debt_row, last_col)
        shares = _latest_quarter_value(fs, shares_row, last_col)
        if shares <= 0:
            raise RedesignError(f"non-positive diluted shares ({shares}) on Financials")

        ncons = _consensus_window(wb)
        fx = _read_fx(vs)
    finally:
        wb.close()

    # WACC mirrors the in-sheet WACC tab exactly: CAPM cost of equity (incl. the
    # revenue-weighted country risk premium), market-value weights from the
    # Dashboard price + Financials shares/debt.
    assert (
        rf is not None
        and erp is not None
        and beta is not None
        and kd is not None
        and tax is not None
        and price is not None
    )
    ke = rf + beta * erp + crp
    after_tax_kd = kd * (1.0 - tax)
    market_cap = price * shares
    equity_weight = market_cap / (market_cap + debt) if (market_cap + debt) > 0 else 1.0
    wacc = equity_weight * ke + (1.0 - equity_weight) * after_tax_kd

    assert (
        near_margin is not None
        and term_margin is not None
        and capex26 is not None
        and term_capex_da is not None
        and exit_mult is not None
        and tg is not None
    )
    return RedesignInputs(
        segments=tuple(segments),
        base_revenue_by_segment=base_rev,
        near_growth_by_segment=near,
        terminal_growth_by_segment=term,
        near_op_margin=near_margin,
        terminal_op_margin=term_margin,
        tax_rate=tax,
        capex_2026_m=capex26,
        terminal_capex_da=term_capex_da,
        da_ratio=da_ratio,
        consensus_years=ncons,
        wacc=wacc,
        beta=beta,
        risk_free_rate=rf,
        equity_risk_premium=erp,
        cost_of_debt=kd,
        terminal_method=method,
        terminal_basis=basis,
        exit_multiple=exit_mult,
        terminal_growth_g=tg,
        current_price=price,
        cash_m=cash,
        total_debt_m=debt,
        diluted_shares_m=shares,
        fx_to_usd=fx,
        country_risk_premium=crp,
        growth_fade_curvature=curv,
        near_sbc_pct=sbc_near,
        terminal_sbc_pct=sbc_term,
        bull_deltas=bull,
        bear_deltas=bear,
        weight_bull=w_bull,
        weight_base=w_base,
        weight_bear=w_bear,
    )


def _read_weights(dsh: Worksheet) -> tuple[float, float, float]:
    """Read the three scenario probability-weight cells (Base=B, Bull=C, Bear=D at
    ``SCEN_WEIGHTS_ROW``), normalized to sum 1. A pre-weights workbook (no row 62)
    reads back as the symmetric default — exactly the old global behaviour."""
    base = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_WEIGHT_BASE)
    bull = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_BULL)
    bear = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_BEAR)
    if base is None or bull is None or bear is None:
        return (
            DEFAULT_SCENARIO_WEIGHTS["bull"],
            DEFAULT_SCENARIO_WEIGHTS["base"],
            DEFAULT_SCENARIO_WEIGHTS["bear"],
        )
    return normalize_scenario_weights(bull, base, bear)


def _read_scenario_deltas(dsh: Worksheet, col: int, seed: ScenarioDeltas) -> ScenarioDeltas:
    """Read one scenario column (Bull or Bear) off the Dashboard SCENARIOS block.

    A blank cell falls back to that lever's seed — so a pre-scenario workbook
    (no block at all) reads as the full seed offsets, and clearing a single cell
    restores its documented default on the next refresh rather than silently
    zeroing the lever.
    """

    def cell(row: int, default: float) -> float:
        v = _num(dsh, row, col)
        return v if v is not None else default

    return ScenarioDeltas(
        growth_near=cell(SCEN_ROW_GROWTH_NEAR, seed.growth_near),
        growth_term=cell(SCEN_ROW_GROWTH_TERM, seed.growth_term),
        margin_near=cell(SCEN_ROW_MARGIN_NEAR, seed.margin_near),
        margin_term=cell(SCEN_ROW_MARGIN_TERM, seed.margin_term),
        exit_multiple=cell(SCEN_ROW_EXIT_MULT, seed.exit_multiple),
        terminal_g=cell(SCEN_ROW_TG, seed.terminal_g),
    )


# --------------------------------------------------------------------------- #
# Projection — the live mirror of the builder's _project / in-sheet formulas
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class _ProjectedStreams:
    revenue: list[float]
    ebit: list[float]
    da: list[float]
    valuation_fcf: list[float]


def _project(inp: RedesignInputs) -> _ProjectedStreams:
    """Project N_FC years of revenue / EBIT / D&A / valuation-FCF.

    Mirrors the in-sheet formulas: per-segment growth fades near→terminal over
    the 10y window on a CONVEX curve (front-loaded deceleration, ``g_t = g_T +
    (g_1 - g_T) * ((N-1-j)/(N-1))**GROWTH_FADE_CURVATURE``); operating margin
    ramps near→terminal by the end of the consensus window; D&A = ratio ×
    revenue; capex = D&A × (capex/D&A), where the 2026 ratio = capex / first-year
    D&A and fades to the terminal ratio; ΔNWC = 0.5% of incremental revenue;
    valuation FCF = NOPAT + D&A − capex − ΔNWC − after-tax SBC.

    SBC is charged as an explicit after-tax expense: ``sbc[j]*(1-tax)`` where
    ``sbc[j] = sbc_pct[j] * revenue[j]`` and the rate fades near→terminal linearly.
    The operating margin here is NON-GAAP (SBC excluded), so NOPAT over-taxes by
    the tax shield SBC would have provided; charging ``SBC*(1-tax)`` restores the
    GAAP-equivalent free cash flow (NOPAT_gaap + D&A − capex − ΔNWC, where
    NOPAT_gaap = (EBIT − SBC)*(1−tax) = NOPAT − SBC*(1−tax)). A pre-SBC workbook
    reads ``sbc_pct = 0`` and the term vanishes, reproducing the old behaviour.
    """
    ramp_steps = inp.consensus_years - 1
    fade_span = N_FC - 1

    def seg_growth(segment: str, j: int) -> float:
        g1 = inp.near_growth_by_segment[segment]
        gt = inp.terminal_growth_by_segment[segment]
        frac = ((fade_span - j) / fade_span) ** inp.growth_fade_curvature
        return gt + (g1 - gt) * frac

    def op_margin(j: int) -> float:
        spread = inp.terminal_op_margin - inp.near_op_margin
        return inp.near_op_margin + spread * min(1.0, j / ramp_steps)

    def sbc_pct(j: int) -> float:
        # Linear near→terminal fade over the window, matching the builder's
        # ``fade(sbc_near, sbc_term)`` and the in-sheet per-year SBC% row.
        return inp.near_sbc_pct + (inp.terminal_sbc_pct - inp.near_sbc_pct) * j / fade_span

    revenue: list[float] = []
    ebit: list[float] = []
    da: list[float] = []
    seg = dict(inp.base_revenue_by_segment)
    for j in range(N_FC):
        for s in inp.segments:
            seg[s] *= 1.0 + seg_growth(s, j)
        rev = sum(seg.values())
        revenue.append(rev)
        ebit.append(rev * op_margin(j))
        da.append(rev * inp.da_ratio)

    # 2026 capex/D&A ratio uses the *model's* first-year D&A (matches the in-sheet
    # formula), fading linearly to the terminal ratio over the window.
    da_2026 = da[0] if da[0] else 1.0
    cda0 = inp.capex_2026_m / da_2026

    valuation_fcf: list[float] = []
    prev_rev = sum(inp.base_revenue_by_segment.values())
    for j in range(N_FC):
        capex_da = cda0 + (inp.terminal_capex_da - cda0) * j / (N_FC - 1)
        capex = da[j] * capex_da
        nopat = ebit[j] * (1.0 - inp.tax_rate)
        delta_nwc = (revenue[j] - prev_rev) * NWC_PCT
        sbc_after_tax = sbc_pct(j) * revenue[j] * (1.0 - inp.tax_rate)
        valuation_fcf.append(nopat + da[j] - capex - delta_nwc - sbc_after_tax)
        prev_rev = revenue[j]

    return _ProjectedStreams(revenue, ebit, da, valuation_fcf)


def _terminal_metrics(streams: _ProjectedStreams, inp: RedesignInputs) -> val_mod.TerminalMetrics:
    """Terminal-year line items the exit multiple can apply to (reporting ccy).

    EBITDA is BURDENED by terminal-year SBC (``ebitda = ebit − sbc + da``) so the
    exit multiple applies to real, SBC-charged EBITDA — consistent with charging
    SBC as an operating expense in the FCF stream. For a heavy-SBC software name
    this lowers the terminal value; for a pre-SBC workbook (``sbc_pct = 0``) the
    burden is zero and this is the old ``ebit + da`` metric.
    """
    tax_rate = inp.tax_rate
    sbc_terminal = inp.terminal_sbc_pct * streams.revenue[-1]
    return val_mod.TerminalMetrics(
        revenue=streams.revenue[-1],
        ebit=streams.ebit[-1],
        ebitda=streams.ebit[-1] - sbc_terminal + streams.da[-1],
        fcf=streams.valuation_fcf[-1],
        net_income=streams.ebit[-1] * (1.0 - tax_rate),
    )


def value(inp: RedesignInputs) -> RedesignValuation:
    """Recompute the value-of-record from a redesigned workbook's inputs.

    Routes through ``compute_valuation`` (the Damodaran engine): the default
    exit-multiple terminal directly, and a perpetuity terminal as an equivalent
    EV/FCF multiple of (1+g)/(WACC−g). Multiplies the per-share result by FX to
    convert non-USD reporters to USD. Raises ``RedesignError`` only for a
    genuinely un-valuable assumption (perpetuity with WACC ≤ g).
    """
    streams = _project(inp)
    years = list(range(N_FC))  # discount exponents are positional, labels unused
    terminal = _terminal_metrics(streams, inp)

    if inp.terminal_method == PERPETUITY:
        if inp.wacc <= inp.terminal_growth_g:
            raise RedesignError(
                f"perpetuity terminal requires WACC ({inp.wacc:.4f}) > "
                f"terminal g ({inp.terminal_growth_g:.4f})"
            )
        basis = "EV/FCF"
        multiple = (1.0 + inp.terminal_growth_g) / (inp.wacc - inp.terminal_growth_g)
    else:
        basis = inp.terminal_basis
        multiple = inp.exit_multiple

    dv = val_mod.compute_valuation(
        streams.valuation_fcf,
        years,
        inp.wacc,
        basis=basis,
        terminal_multiple=multiple,
        terminal=terminal,
        cash_and_nonop=inp.cash_m,
        total_debt=inp.total_debt_m,
        diluted_shares_M=inp.diluted_shares_m,
    )

    return RedesignValuation(
        value_per_share_usd=dv.value_per_share * inp.fx_to_usd,
        value_per_share_reporting=dv.value_per_share,
        operating_value_usd_m=dv.operating_value * inp.fx_to_usd,
        equity_value_usd_m=dv.equity_value * inp.fx_to_usd,
        fcff_stream_m=list(streams.valuation_fcf),
        forecast_revenue_m=list(streams.revenue),
        wacc=inp.wacc,
        terminal_method=inp.terminal_method,
        terminal_basis=inp.terminal_basis,
        exit_multiple=inp.exit_multiple,
        fx_to_usd=inp.fx_to_usd,
        diluted_shares_m=inp.diluted_shares_m,
        cash_m=inp.cash_m,
        total_debt_m=inp.total_debt_m,
        current_price=inp.current_price,
    )


def read_and_value(workbook_path: Path) -> RedesignValuation | None:
    """Read a redesigned workbook and recompute its value-of-record.

    Returns ``None`` if the workbook is not redesigned-format.
    """
    inp = read_inputs(workbook_path)
    if inp is None:
        return None
    return value(inp)


# --------------------------------------------------------------------------- #
# Scenarios — Bull/Bear fair values from Base inputs + deltas
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ScenarioValues:
    """The three per-share fair values (USD). ``bull``/``bear`` are ``None``
    when that scenario's shifted inputs are un-valuable (e.g. a Bull terminal g
    pushed to/above WACC under a perpetuity terminal) — the Base contract is
    unchanged and still raises for a degenerate Base."""

    base: float
    bull: float | None
    bear: float | None


def apply_scenario(inp: RedesignInputs, deltas: ScenarioDeltas) -> RedesignInputs:
    """Base inputs shifted by one scenario's offsets.

    Growth deltas shift every segment uniformly; margin/multiple/g deltas shift
    the scalar. Deliberately un-clamped: an offset that drives the model out of
    its domain (negative exit multiple, perpetuity g ≥ WACC) should surface as a
    degenerate/None scenario value, not be silently repaired.
    """
    return dataclasses.replace(
        inp,
        near_growth_by_segment={
            s: g + deltas.growth_near for s, g in inp.near_growth_by_segment.items()
        },
        terminal_growth_by_segment={
            s: g + deltas.growth_term for s, g in inp.terminal_growth_by_segment.items()
        },
        near_op_margin=inp.near_op_margin + deltas.margin_near,
        terminal_op_margin=inp.terminal_op_margin + deltas.margin_term,
        exit_multiple=inp.exit_multiple + deltas.exit_multiple,
        terminal_growth_g=inp.terminal_growth_g + deltas.terminal_g,
    )


def scenario_values(inp: RedesignInputs) -> ScenarioValues:
    """Base/Bull/Bear per-share fair values (USD) from one input set."""

    def shifted(deltas: ScenarioDeltas) -> float | None:
        try:
            return value(apply_scenario(inp, deltas)).value_per_share_usd
        except RedesignError:
            return None

    return ScenarioValues(
        base=value(inp).value_per_share_usd,
        bull=shifted(inp.bull_deltas),
        bear=shifted(inp.bear_deltas),
    )


# --------------------------------------------------------------------------- #
# Sensitivity — WACC x exit-multiple grid
# --------------------------------------------------------------------------- #
_GRID_WACC_STEPS: tuple[float, ...] = (-0.015, -0.01, -0.005, 0.0, 0.005, 0.01, 0.015)
_GRID_MULT_STEPS: tuple[float, ...] = (-3.0, -2.0, -1.0, 0.0, 1.0, 2.0, 3.0)


@dataclass(frozen=True)
class SensitivityGrid:
    """A static WACC × exit-multiple grid of per-share fair values (USD).

    ``values[i][j]`` is the fair value at ``multiple_axis[i]`` × ``wacc_axis[j]``.
    Always computed on the exit-multiple terminal (the selected basis) — for a
    perpetuity-method name the grid is the exit-multiple cross-check, labeled as
    such on the sheet.
    """

    wacc_axis: tuple[float, ...]
    multiple_axis: tuple[float, ...]
    values: tuple[tuple[float, ...], ...]
    base_wacc: float
    base_multiple: float
    basis: str
    current_price: float


def sensitivity_grid(inp: RedesignInputs) -> SensitivityGrid:
    """Fair value / share over WACC ± 1.5% (50bp steps) × exit multiple ± 3 turns.

    The center cell is exactly ``value()`` of the Base inputs under an
    exit-multiple terminal — i.e. the value-of-record for exit-multiple names.
    """
    wacc_axis = tuple(inp.wacc + d for d in _GRID_WACC_STEPS)
    mult_axis = tuple(inp.exit_multiple + d for d in _GRID_MULT_STEPS)
    rows: list[tuple[float, ...]] = []
    for mult in mult_axis:
        row: list[float] = []
        for wacc in wacc_axis:
            shifted = dataclasses.replace(
                inp, wacc=wacc, exit_multiple=mult, terminal_method=EXIT_MULTIPLE
            )
            row.append(value(shifted).value_per_share_usd)
        rows.append(tuple(row))
    return SensitivityGrid(
        wacc_axis=wacc_axis,
        multiple_axis=mult_axis,
        values=tuple(rows),
        base_wacc=inp.wacc,
        base_multiple=inp.exit_multiple,
        basis=inp.terminal_basis,
        # The Dashboard price is already USD (the in-sheet upside compares the
        # FX-converted value/share against it directly) — no FX re-scaling here.
        current_price=inp.current_price,
    )


# --------------------------------------------------------------------------- #
# Static computed outputs — the Sensitivity sheet + Dashboard scenario values
# --------------------------------------------------------------------------- #
_PXS_FMT = '"$"#,##0.00'
_PCT_FMT = "0.0%"
_MULT_FMT = '0.0"x"'
_GRID_ROW0 = 4  # header row of the grid (axis labels); data rows follow
_GRID_COL0 = 2  # column B = first WACC column
_HDR_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=15)
_SUB_FONT = Font(italic=True, color="666666")
_BASE_FILL = PatternFill("solid", fgColor="DDEBF7")  # the builder's ACTH band
_ABOVE_PRICE_FONT = Font(color="008000")  # green = grid value above current price


def write_sensitivity_sheet(wb: Workbook, grid: SensitivityGrid) -> None:
    """(Re)write the static Sensitivity sheet on an open workbook.

    The grid is Python-computed (openpyxl can't evaluate formulas offline) and
    rewritten on every rebuild/refresh, so styling is applied directly: the base
    cell is shaded, values above the current price render green.
    """
    if SENSITIVITY_SHEET in wb.sheetnames:
        idx = wb.sheetnames.index(SENSITIVITY_SHEET)
        wb.remove(wb[SENSITIVITY_SHEET])
        ws = wb.create_sheet(SENSITIVITY_SHEET, idx)
    else:
        # Fresh insert: directly before Monte Carlo when present, else appended.
        names = wb.sheetnames
        idx = names.index("Monte Carlo") if "Monte Carlo" in names else len(names)
        ws = wb.create_sheet(SENSITIVITY_SHEET, idx)

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    for j in range(len(grid.wacc_axis)):
        ws.column_dimensions[get_column_letter(_GRID_COL0 + j)].width = 11

    ws.cell(1, 1, "Sensitivity — fair value / share (USD)").font = _TITLE_FONT
    ws.cell(
        2,
        1,
        f"WACC x exit multiple ({grid.basis}). Static grid recomputed by every "
        "rebuild/refresh — edit the Dashboard, re-run refresh_dcf to update. "
        "Shaded = base case; green = above current price.",
    ).font = _SUB_FONT

    corner = ws.cell(_GRID_ROW0, 1, f"{grid.basis} x  \\  WACC")
    corner.font = _HDR_FONT
    for j, w in enumerate(grid.wacc_axis):
        c = ws.cell(_GRID_ROW0, _GRID_COL0 + j, w)
        c.number_format = _PCT_FMT
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="right")
    for i, mult in enumerate(grid.multiple_axis):
        mc = ws.cell(_GRID_ROW0 + 1 + i, 1, mult)
        mc.number_format = _MULT_FMT
        mc.font = _HDR_FONT
        for j, val in enumerate(grid.values[i]):
            cell = ws.cell(_GRID_ROW0 + 1 + i, _GRID_COL0 + j, val)
            cell.number_format = _PXS_FMT
            if val > grid.current_price:
                cell.font = _ABOVE_PRICE_FONT
            if grid.multiple_axis[i] == grid.base_multiple and (
                grid.wacc_axis[j] == grid.base_wacc
            ):
                cell.fill = _BASE_FILL

    price_row = _GRID_ROW0 + len(grid.multiple_axis) + 2
    ws.cell(price_row, 1, "Current price").font = _HDR_FONT
    pc = ws.cell(price_row, 2, grid.current_price)
    pc.number_format = _PXS_FMT
    base_row = price_row + 1
    ws.cell(base_row, 1, "Base WACC / multiple").font = _HDR_FONT
    bw = ws.cell(base_row, 2, grid.base_wacc)
    bw.number_format = _PCT_FMT
    bm = ws.cell(base_row, 3, grid.base_multiple)
    bm.number_format = _MULT_FMT


def write_scenario_fair_values(wb: Workbook, sv: ScenarioValues) -> None:
    """Write the three computed fair values into Dashboard row ``SCEN_FV_ROW``.

    Static literals (not formulas): the in-sheet formula chain only evaluates
    the Base inputs, so Bull/Bear are Python-computed and refreshed on every
    rebuild/refresh — same contract as the Sensitivity grid. An un-valuable
    scenario writes "n/a" so the cell never carries a stale number.
    """
    dsh = wb[DASHBOARD_SHEET]
    for col, val in ((2, sv.base), (SCEN_COL_BULL, sv.bull), (SCEN_COL_BEAR, sv.bear)):
        cell = dsh.cell(row=SCEN_FV_ROW, column=col, value=val if val is not None else "n/a")
        if val is not None:
            cell.number_format = _PXS_FMT
        cell.font = _HDR_FONT


def write_computed_outputs(workbook_path: Path, inp: RedesignInputs) -> ScenarioValues:
    """Recompute + rewrite the static Python-computed cells from ``inp``.

    Called by the refresher AFTER ``inject_dashboard`` so the scenario fair
    values (Dashboard row 58) and the Sensitivity grid reflect the user's
    preserved edits, not the builder's from-scratch defaults. Returns the
    scenario values so the caller can persist them without re-deriving.
    """
    sv = scenario_values(inp)
    grid = sensitivity_grid(inp)
    wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    try:
        write_scenario_fair_values(wb, sv)
        write_sensitivity_sheet(wb, grid)
        wb.save(str(workbook_path))
    finally:
        wb.close()
    return sv


# --------------------------------------------------------------------------- #
# Edit-preservation — capture the user-owned Dashboard, re-inject after rebuild
# --------------------------------------------------------------------------- #
# Dashboard row -> the RedesignInputs scalar attribute it carries. The inverse of
# the read in ``read_inputs`` — used by ``capture_from_inputs`` to round-trip an
# edited input set back onto the yellow cells via ``inject_dashboard``. WACC is
# absent on purpose: it has no input cell (the WACC tab derives it from the CAPM
# drivers below), so an in-app WACC change persists only through beta/rf/ERP/Kd.
_SCALAR_ROW_ATTR: tuple[tuple[int, str], ...] = (
    (_DB_MARGIN_NEAR, "near_op_margin"),
    (_DB_MARGIN_TERM, "terminal_op_margin"),
    (_DB_TAX, "tax_rate"),
    (_DB_CAPEX26, "capex_2026_m"),
    (_DB_TERM_CAPEX_DA, "terminal_capex_da"),
    (_DB_RF, "risk_free_rate"),
    (_DB_ERP, "equity_risk_premium"),
    (_DB_BETA, "beta"),
    (_DB_KD, "cost_of_debt"),
    (_DB_CRP, "country_risk_premium"),
    (_DB_MULT, "exit_multiple"),
    (_DB_TG, "terminal_growth_g"),
    (_DB_SBC_NEAR, "near_sbc_pct"),
    (_DB_SBC_TERM, "terminal_sbc_pct"),
)


def capture_from_inputs(inp: RedesignInputs) -> CapturedDashboard:
    """A :class:`CapturedDashboard` built from an in-memory ``RedesignInputs``.

    The bridge from the JSON recompute path back to the workbook: an in-app save
    feeds this to :func:`inject_dashboard` to write the edited yellow cells onto
    the live ``dcf/<T>.xlsx`` (so the edits survive a from-scratch rebuild and
    Push-to-Sheets publishes them). Only the editable input cells are written —
    segment near/terminal growth, the scalar levers (margins, tax, capex, the
    CAPM drivers, exit multiple, terminal g), and the terminal method/basis;
    ``scenario_deltas`` is left empty so the workbook keeps its own Bull/Bear
    offset cells. Current price is a market datum, not captured here.
    """
    segment_growth = {
        s: (inp.near_growth_by_segment[s], inp.terminal_growth_by_segment[s]) for s in inp.segments
    }
    scalars = {row: float(getattr(inp, attr)) for row, attr in _SCALAR_ROW_ATTR}
    return CapturedDashboard(
        segment_growth=segment_growth,
        scalars=scalars,
        terminal_method=inp.terminal_method,
        terminal_basis=inp.terminal_basis,
        scenario_deltas={},
        scenario_weights={
            "bull": inp.weight_bull,
            "base": inp.weight_base,
            "bear": inp.weight_bear,
        },
    )


def capture_dashboard(workbook_path: Path) -> CapturedDashboard | None:
    """Snapshot the user-owned Dashboard inputs (yellow cells) for preservation.

    Returns ``None`` if the workbook is missing or not redesigned-format (a fresh
    build has nothing to preserve). Current price is intentionally not captured —
    the refresher always refreshes it from the live quote.
    """
    if not is_redesign_format(workbook_path):
        return None
    wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    try:
        dsh = wb[DASHBOARD_SHEET]
        segment_growth: dict[str, tuple[float, float]] = {}
        for row in range(SEG_ROW0, SEG_ROW_MAX + 1):
            name = _text(dsh, row, 1)
            g_near = _num(dsh, row, 2)
            if not name or g_near is None:
                continue
            g_term = _num(dsh, row, 3)
            segment_growth[name] = (g_near, g_term if g_term is not None else g_near)
        scalars: dict[int, float] = {}
        for r in _PRESERVED_SCALAR_ROWS:
            v = _num(dsh, r, 2)
            if v is not None:
                scalars[r] = v
        method = _text(dsh, _DB_METHOD, 2)
        basis = _text(dsh, _DB_BASIS, 2)
        scenario_deltas: dict[tuple[int, int], float] = {}
        for r in _SCEN_DELTA_ROWS:
            for c in (SCEN_COL_BULL, SCEN_COL_BEAR):
                v = _num(dsh, r, c)
                if v is not None:
                    scenario_deltas[(r, c)] = v
        scenario_weights: dict[str, float] = {}
        w_base = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_WEIGHT_BASE)
        w_bull = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_BULL)
        w_bear = _num(dsh, SCEN_WEIGHTS_ROW, SCEN_COL_BEAR)
        if w_base is not None and w_bull is not None and w_bear is not None:
            scenario_weights = {"bull": w_bull, "base": w_base, "bear": w_bear}
    finally:
        wb.close()
    return CapturedDashboard(
        segment_growth=segment_growth,
        scalars=scalars,
        terminal_method=method,
        terminal_basis=basis,
        scenario_deltas=scenario_deltas,
        scenario_weights=scenario_weights,
    )


def inject_dashboard(
    workbook_path: Path,
    captured: CapturedDashboard | None,
    *,
    current_price: float | None,
) -> None:
    """Write preserved Dashboard inputs into a freshly-rebuilt workbook in place.

    Segment growth is re-applied by segment *name* (a renamed/added/removed
    segment keeps the fresh default); scalars by row; current price (always) from
    the live quote when supplied. A ``None`` capture still refreshes price.
    """
    wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    try:
        dsh = wb[DASHBOARD_SHEET]
        if captured is not None:
            for row in range(SEG_ROW0, SEG_ROW_MAX + 1):
                name = _text(dsh, row, 1)
                if name and name in captured.segment_growth:
                    g_near, g_term = captured.segment_growth[name]
                    dsh.cell(row=row, column=2, value=g_near)
                    dsh.cell(row=row, column=3, value=g_term)
            for r, v in captured.scalars.items():
                dsh.cell(row=r, column=2, value=v)
            if captured.terminal_method is not None:
                dsh.cell(row=_DB_METHOD, column=2, value=captured.terminal_method)
            if captured.terminal_basis is not None:
                dsh.cell(row=_DB_BASIS, column=2, value=captured.terminal_basis)
            for (r, c), v in captured.scenario_deltas.items():
                dsh.cell(row=r, column=c, value=v)
            if captured.scenario_weights:
                sw = captured.scenario_weights
                if "base" in sw:
                    dsh.cell(row=SCEN_WEIGHTS_ROW, column=SCEN_COL_WEIGHT_BASE, value=sw["base"])
                if "bull" in sw:
                    dsh.cell(row=SCEN_WEIGHTS_ROW, column=SCEN_COL_BULL, value=sw["bull"])
                if "bear" in sw:
                    dsh.cell(row=SCEN_WEIGHTS_ROW, column=SCEN_COL_BEAR, value=sw["bear"])
        if current_price is not None:
            dsh.cell(row=_DB_PRICE, column=2, value=current_price)
        wb.save(str(workbook_path))
    finally:
        wb.close()
