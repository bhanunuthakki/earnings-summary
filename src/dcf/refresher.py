"""Refresh an existing DCF workbook: Historicals + projections + Valuation.

The seeder is responsible for the workbook's existence and initial layout
(three sheets: Historicals, Forecast, Valuation). The refresher is what
runs each cycle as new quarters land:

  * Historicals — rewritten from current FMP data. Always.
  * Forecast.INPUTS — preserved verbatim. The user's edits live here; the
    refresher never overwrites them. (Re-seed with `force=True` to reset.)
  * Forecast.PROJECTED — recomputed from the current INPUTS via Python and
    rewritten.
  * Valuation — rewritten from inputs + projections so `workbook_reader`
    sees the freshest fair-value math.

The user's iteration loop:

  1. Open `dcf/<TICKER>.xlsx` in Excel.
  2. Edit a cell in Forecast INPUTS (e.g., Y1 growth, FCF margin). Save.
  3. Re-run `python execution/refresh_dcf.py --ticker <TICKER>`.
  4. The refresher reads their edited inputs, recomputes projections +
     Valuation, persists the new fair value to `dcf_runs`.

The refresher refuses workbooks missing any of the three required sheets —
that means the seeder hasn't run, or the workbook was created by an older
toolchain. Either way: re-seed.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from dcf import forecast as forecast_mod
from dcf.seeder import (
    FORECAST_SHEET,
    HISTORICALS_SHEET,
    VALUATION_SHEET,
    SeederError,
    write_historicals_sheet,
    write_valuation_sheet,
)


class RefresherError(Exception):
    """Refresh cannot proceed (workbook missing, structurally invalid, etc.)."""


@dataclass(frozen=True)
class RefreshResult:
    workbook_path: str
    historicals_cells_written: int
    forecast_inputs: forecast_mod.ForecastInputs
    projections: forecast_mod.ForecastProjections
    base_year: int
    sheets: list[str]


def refresh_historicals(
    workbook_path: Path,
    fmp_quarterly_dir: Path,
    ticker: str | None = None,
    *,
    base_year: int | None = None,
    db_path: Path | None = None,
) -> RefreshResult:
    """Refresh Historicals + recompute Forecast.PROJECTED + Valuation.

    `ticker` defaults to the workbook's filename stem. `base_year` defaults
    to the current calendar year — forecast years run from `base_year + 1`
    through `base_year + inputs.forecast_years`.

    `db_path` enables the financial_facts fallback when refreshing Historicals
    for a ticker with no FMP quarterly files (see `seeder.write_historicals_sheet`).

    The function name is kept for backward compatibility with PR #142 — it
    now does more than just Historicals.
    """
    if not workbook_path.exists():
        raise RefresherError(f"workbook not found: {workbook_path}")

    resolved_ticker = (ticker or workbook_path.stem).upper()
    resolved_base_year = base_year if base_year is not None else date.today().year

    # Sheet guard — both Forecast and Valuation are required; Historicals will
    # be regenerated below regardless of its current state.
    wb_check = openpyxl.load_workbook(str(workbook_path), read_only=True)
    sheets_before = list(wb_check.sheetnames)
    wb_check.close()
    missing_sheets = {FORECAST_SHEET, VALUATION_SHEET} - set(sheets_before)
    if missing_sheets:
        raise RefresherError(
            f"{workbook_path.name} missing required sheets: {sorted(missing_sheets)} "
            f"— re-seed with force=True"
        )

    # Step 1 — Historicals. Reuses the seeder's helper so the row layout
    # stays single-sourced.
    try:
        cells = write_historicals_sheet(
            workbook_path, fmp_quarterly_dir, resolved_ticker, db_path=db_path
        )
    except SeederError as e:
        raise RefresherError(str(e)) from e

    # Step 2 — load forecast inputs (user-edited or seed-defaults), recompute
    # projections, rewrite Valuation. This open-modify-save is separate from
    # Historicals so each pass writes a clean workbook (openpyxl's load_workbook
    # + save loses the same data twice if you do it back-to-back without an
    # intervening read).
    wb = openpyxl.load_workbook(str(workbook_path))
    forecast_ws = wb[FORECAST_SHEET]
    try:
        inputs = forecast_mod.read_inputs_from_sheet(forecast_ws)
    except forecast_mod.ForecastError as e:
        raise RefresherError(f"Forecast inputs unreadable: {e}") from e

    projections = forecast_mod.compute_projections(inputs, resolved_base_year)
    forecast_mod.write_projections_section(forecast_ws, projections)

    val_ws = wb[VALUATION_SHEET]
    ttm_fcf_M = _ttm_fcf_from_historicals(wb[HISTORICALS_SHEET]) if HISTORICALS_SHEET in wb.sheetnames else None
    write_valuation_sheet(val_ws, inputs, projections, resolved_base_year, ttm_fcf_M)

    wb.save(str(workbook_path))

    # Confirm the sheet list survived the round-trip. A drift here means our
    # write code corrupted the workbook structure.
    wb_after = openpyxl.load_workbook(str(workbook_path), read_only=True)
    sheets_after = list(wb_after.sheetnames)
    wb_after.close()

    return RefreshResult(
        workbook_path=str(workbook_path),
        historicals_cells_written=cells,
        forecast_inputs=inputs,
        projections=projections,
        base_year=resolved_base_year,
        sheets=sheets_after,
    )


def _ttm_fcf_from_historicals(ws: Worksheet) -> float | None:
    """Sum the FCF row's last 4 quarter columns in Historicals.

    The seeder lays out the Historicals sheet with one quarter per column,
    newest on the right. We scan column A for the 'Free Cash Flow' label and
    sum the rightmost 4 numeric values from that row.
    """
    target_row: int | None = None
    for row in range(1, 60):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.strip().lower() == "free cash flow":
            target_row = row
            break
    if target_row is None:
        return None

    values: list[float] = []
    for col in range(2, 30):
        v = ws.cell(row=target_row, column=col).value
        if isinstance(v, (int, float)):
            values.append(float(v))
    if not values:
        return None
    return sum(values[-4:])
