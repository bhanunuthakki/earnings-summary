"""Forecast assumption derivation + projection math for DCF workbooks.

The workbook's Forecast sheet has two zones:

  INPUTS    — user-editable. A per-year assumption GRID (one column per forecast
              year) seeded from the ticker's own history, plus a handful of
              scalars. The refresher does NOT overwrite these on subsequent runs
              — re-seed with `force=True` to reset from latest FMP.

  PROJECTED — program-owned line-item bridge, rewritten each refresh from the
              current INPUTS via Python (not Excel formulas, so openpyxl reads
              stay deterministic).

The model is a full P&L -> FCF bridge (replacing the prior 7-scalar
op-margin/capex-intensity model), built to mirror the line-item analyst models
in `examples/dcf/` (ASML, FIG):

    revenue_t        = revenue_{t-1} x (1 + growth_t)           # growth decays Y1->terminal
    gross_profit_t   = revenue_t x gross_margin_t
    operating_inc_t  = gross_profit_t - rnd_t - sga_t           # R&D, SG&A each % of revenue
    nopat_t          = operating_inc_t x (1 - tax_t)
    cfo_t            = nopat_t + d&a_t + sbc_t - delta_nwc_t     # SBC added back (non-cash)
    fcff_t           = cfo_t - capex_t                          # capex = (capex/d&a)_t x d&a_t
    valuation_fcf_t  = fcff_t - sbc_t                           # SBC re-charged as a real cost

Working capital is days-driven: receivables (DSO), payables (DPO), and deferred
revenue (% of revenue) are each projected, and their period-over-period change
is `delta_nwc`. Diluted shares evolve by a net-share-change driver (SBC dilution
net of buybacks) so the projected-share row shows the dilution path.

SBC treatment: it is added back in CFO (it is genuinely non-cash) but subtracted
again to reach the valuation FCF, so stock comp is charged as the real economic
cost it is. That keeps the per-share denominator at *current* shares (no
double-count against the evolving-share row) — `workbook_reader` reads the
`valuation_fcf` series into the Valuation sheet's FCF row unchanged.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_FORECAST_YEARS = 5
DEFAULT_TERMINAL_GROWTH = 0.025

_TTM_QUARTERS = 4
_MILLIONS = 1_000_000.0
_MAX_HORIZON = 10
_DAYS = 365.0

# Per-driver clamps + fallbacks (applied when history is missing or degenerate).
_TAX_CLAMP = (0.10, 0.35)
_GROSS_MARGIN_CLAMP = (0.05, 0.95)
_DEFAULTS: dict[str, float] = {
    "gross_margin": 0.50,
    "rnd_pct": 0.10,
    "sga_pct": 0.20,
    "sbc_pct": 0.03,
    "da_pct": 0.04,
    "capex_to_da": 1.20,
    "dso_days": 45.0,
    "dpo_days": 30.0,
    "deferred_rev_pct": 0.0,
    "tax_rate": 0.25,
    "net_share_change": 0.0,
}

# --------------------------------------------------------------------------- #
# Forecast-sheet layout. Single source of truth for every read + write path.
# Column A = label; column B = base/TTM (reference); columns C.. = forecast years.
# --------------------------------------------------------------------------- #
_BASE_COL = 2  # column B — base/TTM reference
_FORECAST_COL0 = 3  # column C — first forecast year

# Scalars (column A label -> row, value in column B, yellow/editable).
_SCALAR_ROWS: dict[str, tuple[int, str]] = {
    "base_revenue": (3, "Base Revenue (TTM, $M)"),
    "diluted_shares": (4, "Diluted Shares (M)"),
    "terminal_multiple": (5, "Terminal FCF Multiple"),
    "forecast_years": (6, "Forecast Years"),
}

# Per-year assumption rows (yellow/editable across the forecast columns).
_ASSUMPTION_HEADER_ROW = 8
_ASSUMPTION_ROWS: dict[str, tuple[int, str, str]] = {
    # key: (row, label, number_format)
    "revenue_growth": (9, "Revenue Growth %", "0.00%"),
    "gross_margin": (10, "Gross Margin %", "0.00%"),
    "rnd_pct": (11, "R&D % of Revenue", "0.00%"),
    "sga_pct": (12, "SG&A % of Revenue", "0.00%"),
    "sbc_pct": (13, "SBC % of Revenue", "0.00%"),
    "da_pct": (14, "D&A % of Revenue", "0.00%"),
    "capex_to_da": (15, "Capex / D&A (x)", "0.00"),
    "dso_days": (16, "DSO (days)", "0"),
    "dpo_days": (17, "DPO (days)", "0"),
    "deferred_rev_pct": (18, "Deferred Rev % of Revenue", "0.00%"),
    "tax_rate": (19, "Tax Rate %", "0.00%"),
    "net_share_change": (20, "Net Share Change %/yr", "0.00%"),
}

# Projected bridge rows (program-owned, recomputed each refresh).
_PROJECTED_HEADER_ROW = 22
_PROJECTED_YEAR_ROW = 23
_PROJECTED_ROWS: list[tuple[str, str, str]] = [
    # (projections-field, label, number_format) — order = display order.
    ("revenue_M", "Revenue ($M)", "#,##0"),
    ("gross_profit_M", "Gross Profit ($M)", "#,##0"),
    ("rnd_M", "R&D ($M)", "#,##0"),
    ("sga_M", "SG&A ($M)", "#,##0"),
    ("operating_income_M", "Operating Income ($M)", "#,##0"),
    ("operating_margin_pct", "Operating Margin %", "0.00%"),
    ("nopat_M", "NOPAT ($M)", "#,##0"),
    ("da_M", "+ D&A ($M)", "#,##0"),
    ("sbc_M", "+ SBC ($M, non-cash)", "#,##0"),
    ("delta_nwc_M", "- Chg in NWC ($M)", "#,##0"),
    ("cfo_M", "= CFO ($M)", "#,##0"),
    ("capex_M", "- Capex ($M)", "#,##0"),
    ("fcff_M", "= FCFF ($M)", "#,##0"),
    ("valuation_fcf_M", "= Valuation FCF ($M, SBC-charged)", "#,##0"),
    ("fcf_margin_pct", "Valuation FCF Margin %", "0.00%"),
    ("shares_M", "Diluted Shares (projected, M)", "#,##0"),
]

_INPUT_FILL = PatternFill(start_color="FFF7DC", end_color="FFF7DC", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_NOTE_FONT = Font(italic=True, color="666666")
_MAX_SCAN_ROW = 60


@dataclass(frozen=True)
class ForecastInputs:
    """User-editable assumptions. Scalars + per-year driver series.

    Every `*_pct`/ratio list holds one value per forecast year (decimals;
    0.10 == 10%). `dso_days`/`dpo_days` are in days. List length == horizon.
    """

    base_revenue_M: float
    diluted_shares_M: float
    terminal_multiple: float
    forecast_years: int
    revenue_growth_pct: list[float]
    gross_margin_pct: list[float]
    rnd_pct: list[float]
    sga_pct: list[float]
    sbc_pct: list[float]
    da_pct: list[float]
    capex_to_da: list[float]
    dso_days: list[float]
    dpo_days: list[float]
    deferred_rev_pct: list[float]
    tax_rate_pct: list[float]
    net_share_change_pct: list[float]


@dataclass(frozen=True)
class ForecastProjections:
    """Full year-by-year P&L -> FCF bridge computed from `ForecastInputs`.

    `valuation_fcf_M` is the series the Valuation sheet's FCF row consumes;
    `shares_M` is the projected diluted-share path. All `_M` series are USD
    millions; `years` is the forecast year labels.
    """

    years: list[int]
    revenue_M: list[float]
    gross_profit_M: list[float]
    rnd_M: list[float]
    sga_M: list[float]
    operating_income_M: list[float]
    operating_margin_pct: list[float]
    nopat_M: list[float]
    da_M: list[float]
    sbc_M: list[float]
    delta_nwc_M: list[float]
    cfo_M: list[float]
    capex_M: list[float]
    fcff_M: list[float]
    valuation_fcf_M: list[float]
    fcf_margin_pct: list[float]
    shares_M: list[float]


class ForecastError(Exception):
    """Forecast inputs are missing or malformed in the workbook."""


# --------------------------------------------------------------------------- #
# Derivation — starter per-year assumptions from the ticker's history
# --------------------------------------------------------------------------- #
def derive_initial_inputs(
    income_records: list[dict[str, object]],
    cashflow_records: list[dict[str, object]],
    balance_records: list[dict[str, object]] | None = None,
    *,
    terminal_growth_pct: float = DEFAULT_TERMINAL_GROWTH,
    forecast_years: int = DEFAULT_FORECAST_YEARS,
) -> ForecastInputs:
    """Derive starter per-year assumptions from the ticker's TTM history.

    Records must be newest-first (FMP's native order). Each ratio defaults to
    its TTM value held flat across the horizon; revenue growth decays linearly
    from the TTM YoY rate to `terminal_growth_pct`. Missing inputs fall back to
    the `_DEFAULTS` steady-state shapes so a sparse/recently-IPO'd ticker still
    yields a viable model.
    """
    balance_records = balance_records or []
    horizon = max(1, min(forecast_years, _MAX_HORIZON))

    rev_ttm = _ttm_sum(income_records[:_TTM_QUARTERS], "revenue")
    base_revenue = (rev_ttm / _MILLIONS) if rev_ttm else 0.0
    rev_safe = rev_ttm if (rev_ttm and rev_ttm > 0) else None

    prior_rev = _ttm_sum(income_records[_TTM_QUARTERS : 2 * _TTM_QUARTERS], "revenue")
    y1_growth = (
        (rev_safe / prior_rev - 1.0)
        if (rev_safe is not None and prior_rev and prior_rev > 0)
        else terminal_growth_pct
    )

    gross_ttm = _ttm_sum(income_records[:_TTM_QUARTERS], "grossProfit")
    if gross_ttm is None:
        cogs_ttm = _ttm_sum(income_records[:_TTM_QUARTERS], "costOfRevenue")
        gross_ttm = (rev_ttm - cogs_ttm) if (rev_ttm is not None and cogs_ttm is not None) else None
    gross_margin = _ratio(gross_ttm, rev_safe, _DEFAULTS["gross_margin"])
    gross_margin = _clamp(gross_margin, *_GROSS_MARGIN_CLAMP)
    cogs_for_dpo = rev_safe * (1.0 - gross_margin) if rev_safe is not None else None

    rnd_pct = _ratio(
        _ttm_sum(income_records[:_TTM_QUARTERS], "researchAndDevelopmentExpenses"),
        rev_safe,
        _DEFAULTS["rnd_pct"],
    )
    sga_pct = _ratio(
        _ttm_sum(income_records[:_TTM_QUARTERS], "sellingGeneralAndAdministrativeExpenses"),
        rev_safe,
        _DEFAULTS["sga_pct"],
    )
    sbc_pct = _ratio(
        _ttm_sum(cashflow_records[:_TTM_QUARTERS], "stockBasedCompensation"),
        rev_safe,
        _DEFAULTS["sbc_pct"],
    )

    da_ttm = _ttm_sum(cashflow_records[:_TTM_QUARTERS], "depreciationAndAmortization")
    da_pct = _ratio(da_ttm, rev_safe, _DEFAULTS["da_pct"])
    capex_ttm = _ttm_sum(cashflow_records[:_TTM_QUARTERS], "capitalExpenditure")
    capex_to_da = (
        (abs(capex_ttm) / da_ttm)
        if (capex_ttm is not None and da_ttm and da_ttm > 0)
        else _DEFAULTS["capex_to_da"]
    )

    dso = _days(_latest_value(balance_records, "netReceivables"), rev_safe, _DEFAULTS["dso_days"])
    dpo = _days(
        _latest_value(balance_records, "accountPayables"), cogs_for_dpo, _DEFAULTS["dpo_days"]
    )
    deferred_pct = _ratio(
        _latest_value(balance_records, "deferredRevenue"), rev_safe, _DEFAULTS["deferred_rev_pct"]
    )

    pretax = _ttm_sum(income_records[:_TTM_QUARTERS], "incomeBeforeTax")
    tax_exp = _ttm_sum(income_records[:_TTM_QUARTERS], "incomeTaxExpense")
    tax_rate = (
        (tax_exp / pretax)
        if (pretax and pretax > 0 and tax_exp is not None)
        else _DEFAULTS["tax_rate"]
    )
    tax_rate = _clamp(tax_rate, *_TAX_CLAMP)

    shares_now = _latest_value(income_records, "weightedAverageShsOutDil")
    shares_year_ago = _quarter_value(income_records, "weightedAverageShsOutDil", _TTM_QUARTERS)
    net_share_change = (
        _clamp(shares_now / shares_year_ago - 1.0, -0.15, 0.15)
        if (shares_now and shares_year_ago and shares_year_ago > 0)
        else _DEFAULTS["net_share_change"]
    )
    diluted_shares = (shares_now / _MILLIONS) if shares_now else 1000.0

    growth_series = _linear_decay(y1_growth, terminal_growth_pct, horizon)

    def flat(v: float) -> list[float]:
        return [round(v, 4)] * horizon

    return ForecastInputs(
        base_revenue_M=round(base_revenue, 2),
        diluted_shares_M=round(diluted_shares, 2),
        terminal_multiple=15.0,
        forecast_years=horizon,
        revenue_growth_pct=[round(g, 4) for g in growth_series],
        gross_margin_pct=flat(gross_margin),
        rnd_pct=flat(rnd_pct),
        sga_pct=flat(sga_pct),
        sbc_pct=flat(sbc_pct),
        da_pct=flat(da_pct),
        capex_to_da=[round(capex_to_da, 2)] * horizon,
        dso_days=[round(dso, 1)] * horizon,
        dpo_days=[round(dpo, 1)] * horizon,
        deferred_rev_pct=flat(deferred_pct),
        tax_rate_pct=flat(tax_rate),
        net_share_change_pct=flat(net_share_change),
    )


# --------------------------------------------------------------------------- #
# Projection — the P&L -> FCF bridge
# --------------------------------------------------------------------------- #
def compute_projections(inputs: ForecastInputs, base_year: int) -> ForecastProjections:
    """Project the full line-item bridge over the horizon from `base_year + 1`.

    Working capital is days-driven; `delta_nwc` for year 1 is measured against
    the base-year NWC implied by the same DSO/DPO/deferred-rev assumptions on
    the base revenue, so the first forecast year isn't penalised by a phantom
    jump. Horizon is clamped to [1, 10].
    """
    horizon = max(1, min(inputs.forecast_years, _MAX_HORIZON))
    years = list(range(base_year + 1, base_year + 1 + horizon))

    def at(series: list[float], i: int, fallback: float) -> float:
        return series[i] if i < len(series) else (series[-1] if series else fallback)

    base_rev = inputs.base_revenue_M
    # Base-year NWC anchor (year-1 deltas measure against this).
    prev_nwc = _nwc(
        base_rev,
        at(inputs.gross_margin_pct, 0, _DEFAULTS["gross_margin"]),
        at(inputs.dso_days, 0, _DEFAULTS["dso_days"]),
        at(inputs.dpo_days, 0, _DEFAULTS["dpo_days"]),
        at(inputs.deferred_rev_pct, 0, _DEFAULTS["deferred_rev_pct"]),
    )

    rev = base_rev
    shares = inputs.diluted_shares_M
    cols: dict[str, list[float]] = {f: [] for f, _, _ in _PROJECTED_ROWS}
    for i in range(horizon):
        gm = at(inputs.gross_margin_pct, i, _DEFAULTS["gross_margin"])
        rev = rev * (1.0 + at(inputs.revenue_growth_pct, i, 0.0))
        gross = rev * gm
        rnd = rev * at(inputs.rnd_pct, i, _DEFAULTS["rnd_pct"])
        sga = rev * at(inputs.sga_pct, i, _DEFAULTS["sga_pct"])
        op_inc = gross - rnd - sga
        nopat = op_inc * (1.0 - at(inputs.tax_rate_pct, i, _DEFAULTS["tax_rate"]))
        da = rev * at(inputs.da_pct, i, _DEFAULTS["da_pct"])
        sbc = rev * at(inputs.sbc_pct, i, _DEFAULTS["sbc_pct"])
        capex = at(inputs.capex_to_da, i, _DEFAULTS["capex_to_da"]) * da
        nwc = _nwc(
            rev,
            gm,
            at(inputs.dso_days, i, _DEFAULTS["dso_days"]),
            at(inputs.dpo_days, i, _DEFAULTS["dpo_days"]),
            at(inputs.deferred_rev_pct, i, _DEFAULTS["deferred_rev_pct"]),
        )
        delta_nwc = nwc - prev_nwc
        prev_nwc = nwc
        cfo = nopat + da + sbc - delta_nwc
        fcff = cfo - capex
        valuation_fcf = fcff - sbc  # re-charge SBC as a real cost
        shares = shares * (1.0 + at(inputs.net_share_change_pct, i, 0.0))

        cols["revenue_M"].append(rev)
        cols["gross_profit_M"].append(gross)
        cols["rnd_M"].append(rnd)
        cols["sga_M"].append(sga)
        cols["operating_income_M"].append(op_inc)
        cols["operating_margin_pct"].append(op_inc / rev if rev else 0.0)
        cols["nopat_M"].append(nopat)
        cols["da_M"].append(da)
        cols["sbc_M"].append(sbc)
        cols["delta_nwc_M"].append(delta_nwc)
        cols["cfo_M"].append(cfo)
        cols["capex_M"].append(capex)
        cols["fcff_M"].append(fcff)
        cols["valuation_fcf_M"].append(valuation_fcf)
        cols["fcf_margin_pct"].append(valuation_fcf / rev if rev else 0.0)
        cols["shares_M"].append(shares)

    return ForecastProjections(years=years, **cols)


def _nwc(revenue: float, gross_margin: float, dso: float, dpo: float, deferred_pct: float) -> float:
    """Net working capital (a use of cash when it rises). Receivables tie up
    cash; payables and deferred revenue are sources, so they net down."""
    cogs = revenue * (1.0 - gross_margin)
    receivables = dso * revenue / _DAYS
    payables = dpo * cogs / _DAYS
    deferred = deferred_pct * revenue
    return receivables - payables - deferred


# --------------------------------------------------------------------------- #
# Sheet writers
# --------------------------------------------------------------------------- #
def write_inputs_section(ws: Worksheet, inputs: ForecastInputs) -> None:
    """Write the INPUTS zone: scalars + the per-year assumption grid. Yellow-
    fills every editable cell. The refresher must NOT call this (it would clobber
    user edits) — it only rewrites the PROJECTED zone."""
    horizon = inputs.forecast_years
    ws.cell(row=1, column=1, value="FORECAST INPUTS - edit the yellow cells").font = _HEADER_FONT
    ws.cell(
        row=2,
        column=1,
        value="(yellow = user-editable; one column per forecast year; refresh preserves your edits)",
    ).font = _NOTE_FONT

    scalars: dict[str, float | int] = {
        "base_revenue": inputs.base_revenue_M,
        "diluted_shares": inputs.diluted_shares_M,
        "terminal_multiple": inputs.terminal_multiple,
        "forecast_years": inputs.forecast_years,
    }
    for key, (row, label) in _SCALAR_ROWS.items():
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=_BASE_COL, value=scalars[key])
        c.fill = _INPUT_FILL
        c.number_format = (
            "#,##0.00" if key in ("base_revenue", "diluted_shares", "terminal_multiple") else "0"
        )

    # Assumption-grid header: forecast-year numbers across the columns.
    ws.cell(
        row=_ASSUMPTION_HEADER_ROW, column=1, value="ASSUMPTIONS (per forecast year)"
    ).font = _HEADER_FONT
    ws.cell(row=_ASSUMPTION_HEADER_ROW, column=_BASE_COL, value="TTM").font = _HEADER_FONT
    for j in range(horizon):
        ws.cell(
            row=_ASSUMPTION_HEADER_ROW, column=_FORECAST_COL0 + j, value=f"FY{j + 1}"
        ).font = _HEADER_FONT

    for key, (row, label, fmt) in _ASSUMPTION_ROWS.items():
        ws.cell(row=row, column=1, value=label)
        series = _series_of(inputs, key)
        for j in range(horizon):
            c = ws.cell(
                row=row,
                column=_FORECAST_COL0 + j,
                value=round(series[j], 4) if j < len(series) else None,
            )
            c.fill = _INPUT_FILL
            c.number_format = fmt

    _autosize(ws, horizon)


def write_projections_section(ws: Worksheet, projections: ForecastProjections) -> None:
    """Write/overwrite the PROJECTED bridge. Called by seeder (first write) and
    refresher (rewrite from current inputs). Clears the prior bridge first so a
    shorter horizon leaves no stale columns."""
    _clear_projections_section(ws)
    ws.cell(
        row=_PROJECTED_HEADER_ROW,
        column=1,
        value="PROJECTED - program-owned, recomputed each refresh from the inputs above",
    ).font = _HEADER_FONT

    ws.cell(row=_PROJECTED_YEAR_ROW, column=1, value="Year").font = _HEADER_FONT
    for i, year in enumerate(projections.years):
        ws.cell(row=_PROJECTED_YEAR_ROW, column=_FORECAST_COL0 + i, value=year).font = _HEADER_FONT

    for r, (field, label, fmt) in enumerate(_PROJECTED_ROWS, start=_PROJECTED_YEAR_ROW + 1):
        ws.cell(row=r, column=1, value=label)
        series: list[float] = getattr(projections, field)
        for i, v in enumerate(series):
            c = ws.cell(row=r, column=_FORECAST_COL0 + i, value=v)
            c.number_format = fmt


# --------------------------------------------------------------------------- #
# Sheet reader
# --------------------------------------------------------------------------- #
def read_inputs_from_sheet(ws: Worksheet) -> ForecastInputs:
    """Read the INPUTS zone back (scalars + the per-year grid), tolerant of the
    user having edited cells. Raises `ForecastError` on missing/invalid scalars."""
    horizon_raw = _scalar(ws, "forecast_years")
    if horizon_raw is None:
        raise ForecastError("Forecast sheet missing 'Forecast Years'")
    horizon = int(horizon_raw)
    if horizon < 1 or horizon > _MAX_HORIZON:
        raise ForecastError(f"forecast_years out of range [1, {_MAX_HORIZON}]: got {horizon}")

    base_rev = _scalar(ws, "base_revenue")
    shares = _scalar(ws, "diluted_shares")
    terminal = _scalar(ws, "terminal_multiple")
    if base_rev is None or shares is None:
        raise ForecastError("Forecast sheet missing 'Base Revenue' or 'Diluted Shares'")

    def grid(key: str) -> list[float]:
        row = _ASSUMPTION_ROWS[key][0]
        out: list[float] = []
        for j in range(horizon):
            v = ws.cell(row=row, column=_FORECAST_COL0 + j).value
            out.append(float(v) if isinstance(v, (int, float)) else _DEFAULTS.get(key, 0.0))
        return out

    return ForecastInputs(
        base_revenue_M=base_rev,
        diluted_shares_M=shares,
        terminal_multiple=terminal if terminal is not None else 15.0,
        forecast_years=horizon,
        revenue_growth_pct=grid("revenue_growth"),
        gross_margin_pct=grid("gross_margin"),
        rnd_pct=grid("rnd_pct"),
        sga_pct=grid("sga_pct"),
        sbc_pct=grid("sbc_pct"),
        da_pct=grid("da_pct"),
        capex_to_da=grid("capex_to_da"),
        dso_days=grid("dso_days"),
        dpo_days=grid("dpo_days"),
        deferred_rev_pct=grid("deferred_rev_pct"),
        tax_rate_pct=grid("tax_rate"),
        net_share_change_pct=grid("net_share_change"),
    )


# --------------------------------------------------------------------------- #
# Helpers (private)
# --------------------------------------------------------------------------- #
def _series_of(inputs: ForecastInputs, key: str) -> list[float]:
    return {
        "revenue_growth": inputs.revenue_growth_pct,
        "gross_margin": inputs.gross_margin_pct,
        "rnd_pct": inputs.rnd_pct,
        "sga_pct": inputs.sga_pct,
        "sbc_pct": inputs.sbc_pct,
        "da_pct": inputs.da_pct,
        "capex_to_da": inputs.capex_to_da,
        "dso_days": inputs.dso_days,
        "dpo_days": inputs.dpo_days,
        "deferred_rev_pct": inputs.deferred_rev_pct,
        "tax_rate": inputs.tax_rate_pct,
        "net_share_change": inputs.net_share_change_pct,
    }[key]


def _scalar(ws: Worksheet, key: str) -> float | None:
    row = _SCALAR_ROWS[key][0]
    v = ws.cell(row=row, column=_BASE_COL).value
    return float(v) if isinstance(v, (int, float)) else None


def _ttm_sum(records: list[dict[str, object]], field: str) -> float | None:
    total = 0.0
    seen = 0
    for r in records:
        v = r.get(field)
        if isinstance(v, (int, float)):
            total += float(v)
            seen += 1
    return total if seen > 0 else None


def _latest_value(records: list[dict[str, object]], field: str) -> float | None:
    for r in records:
        v = r.get(field)
        if isinstance(v, (int, float)):
            return float(v)
    return None


def _quarter_value(records: list[dict[str, object]], field: str, index: int) -> float | None:
    if index < len(records):
        v = records[index].get(field)
        if isinstance(v, (int, float)):
            return float(v)
    return None


def _ratio(numerator: float | None, denominator: float | None, fallback: float) -> float:
    if numerator is None or not denominator or denominator <= 0:
        return fallback
    return numerator / denominator


def _days(balance: float | None, flow_ttm: float | None, fallback: float) -> float:
    if balance is None or not flow_ttm or flow_ttm <= 0:
        return fallback
    return abs(balance) / flow_ttm * _DAYS


def _clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(v, hi))


def _linear_decay(start: float, end: float, n: int) -> list[float]:
    if n <= 0:
        return []
    if n == 1:
        return [start]
    step = (end - start) / (n - 1)
    return [start + i * step for i in range(n)]


def _autosize(ws: Worksheet, horizon: int) -> None:
    ws.column_dimensions["A"].width = 30
    for j in range(horizon + 1):
        ws.column_dimensions[get_column_letter(_BASE_COL + j)].width = 12


def _clear_projections_section(ws: Worksheet) -> None:
    last_row = _PROJECTED_YEAR_ROW + len(_PROJECTED_ROWS) + 1
    for row in range(_PROJECTED_HEADER_ROW, last_row + 1):
        for col in range(1, _FORECAST_COL0 + _MAX_HORIZON + 1):
            ws.cell(row=row, column=col, value=None)
