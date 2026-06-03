"""LLM-driven normalization of DCF forecast assumptions.

The seeder's `derive_initial_inputs` produces a FLAT projection — every ratio
held at its TTM value, growth fading linearly. That's a poor starting point: it
leaves over-earning names permanently over-earning and unprofitable names
permanently unprofitable, so the auto-drafts look terrible. This module asks
Opus to reason bottom-up about where each driver should NORMALIZE over the
horizon, grounded in valuation best practice, and rewrites the Forecast sheet's
per-year assumption grid.

Principles encoded in the prompt (Damodaran-grounded):
  * Revenue growth FADES gradually from the recent run-rate to a terminal
    ~GDP rate (2-3%); no sudden jumps (law of large numbers).
  * Operating margin NORMALIZES toward a defensible MATURE target for the
    business model — expanding for sub-scale names (operating leverage),
    compressing for over-earners. Never flat-extrapolate a peak or trough.
  * R&D / S&M&A and SBC decline as a % of revenue as the company scales.
  * Capex/D&A converges toward ~1.0-1.2x at maturity (steady-state reinvestment).
  * The terminal year MUST be FCF-positive — a DCF whose terminal FCF is
    negative is broken.
  * Each assumption ties to the company's actual trajectory + competitive
    position and the thesis, not a mechanical extrapolation.

The raw Opus response (narrative + per-driver rationale) is cached at
`data/dcf_assumptions/<TICKER>.json` so runs are reviewable and re-runnable
without re-calling the model.
"""

from __future__ import annotations

import dataclasses
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import cast

import openpyxl

from dcf import forecast as forecast_mod

# (driver key in the LLM JSON, the ForecastInputs list-attribute it overrides).
# Order = the order shown to the model. Days vs decimal vs multiple is documented
# in the prompt; parsing is unit-agnostic (it just reads the numbers).
_DRIVER_TO_ATTR: dict[str, str] = {
    "revenue_growth": "revenue_growth_pct",
    "gross_margin": "gross_margin_pct",
    "rnd_pct": "rnd_pct",
    "sga_pct": "sga_pct",
    "sbc_pct": "sbc_pct",
    "da_pct": "da_pct",
    "capex_to_da": "capex_to_da",
    "dso_days": "dso_days",
    "dpo_days": "dpo_days",
    "deferred_rev_pct": "deferred_rev_pct",
    "tax_rate": "tax_rate_pct",
    "net_share_change": "net_share_change_pct",
}

_PRINCIPLES = """\
- Revenue growth FADES gradually from the recent run-rate toward a terminal ~GDP \
rate (2-3%) by the final year. Large, mature franchises fade faster (law of large \
numbers); durable compounders fade slower. No sudden jumps.
- Operating margin NORMALIZES toward a defensible MATURE level for THIS business \
model and competitive position. Sub-scale / unprofitable names should show margin \
EXPANSION as they scale (operating leverage drops R&D/S&M/G&A as a % of revenue); \
names earning above their sustainable level should compress toward it. Never \
flat-extrapolate a current peak or trough.
- R&D, S&M/G&A, and SBC generally DECLINE as a % of revenue as a company scales.
- Reinvestment: Capex/D&A converges toward ~1.0-1.2x at maturity (capex ~ D&A in \
steady state). A heavy build-out phase can run higher early, then normalize.
- Working capital (DSO/DPO/deferred-revenue %) is roughly structural — keep it near \
the business's efficient level unless the thesis implies a shift.
- Tax rate trends toward a normalized statutory rate if today's is distorted.
- The TERMINAL YEAR MUST BE FCF-POSITIVE. If flat-current assumptions imply \
permanent losses, model the realistic path to profitability instead.
- Tie every number to the company's trajectory, competitive moat, and the thesis."""


class LlmAssumptionError(Exception):
    """The LLM response could not be parsed into a usable assumption grid."""


@dataclass(frozen=True)
class LlmAssumptions:
    """Parsed Opus output: the per-driver per-year overrides + the rationale."""

    ticker: str
    horizon: int
    narrative: str
    drivers: dict[str, list[float]]  # driver key -> per-year values
    notes: dict[str, str]  # driver key -> 1-line rationale


# --------------------------------------------------------------------------- #
# Context gathering
# --------------------------------------------------------------------------- #
def gather_context(ticker: str, workbook_path: Path, repo_root: Path) -> dict[str, object]:
    """Assemble what Opus reasons over: the naive baseline grid, a recent
    historicals snapshot, and the investment thesis."""
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    baseline = forecast_mod.read_inputs_from_sheet(wb[forecast_mod_sheet_name()])
    historicals = _historicals_summary(wb)
    thesis, name = _load_thesis(repo_root, ticker)
    return {
        "ticker": ticker.upper(),
        "name": name,
        "thesis": thesis,
        "horizon": baseline.forecast_years,
        "historicals": historicals,
        "baseline": _baseline_grid(baseline),
    }


def forecast_mod_sheet_name() -> str:
    # The Forecast sheet name lives in the seeder; avoid a hard import cycle.
    return "Forecast"


def _baseline_grid(inputs: forecast_mod.ForecastInputs) -> dict[str, list[float]]:
    return {
        key: [round(v, 4) for v in getattr(inputs, attr)] for key, attr in _DRIVER_TO_ATTR.items()
    }


def _historicals_summary(wb: openpyxl.Workbook) -> str:
    """Compact recent-trend block (last 8 quarter columns) for the key lines."""
    if "Historicals" not in wb.sheetnames:
        return "(no Historicals sheet)"
    ws = wb["Historicals"]
    labels = [str(ws.cell(row=1, column=c).value) for c in range(2, min(ws.max_column, 30) + 1)]
    wanted = ("Revenue", "Gross Profit", "Operating Income", "Free Cash Flow", "Diluted Shares (M)")
    lines: list[str] = []
    for target in wanted:
        row = _find_row(ws, target)
        if row is None:
            continue
        vals: list[str] = []
        for c in range(2, min(ws.max_column, 30) + 1):
            v = ws.cell(row=row, column=c).value
            vals.append(f"{v:,.0f}" if isinstance(v, (int, float)) else "")
        recent = [f"{lab}:{val}" for lab, val in zip(labels[-8:], vals[-8:], strict=False) if val]
        if recent:
            lines.append(f"{target}: " + ", ".join(recent))
    return "\n".join(lines) if lines else "(historicals unavailable)"


def _find_row(ws: object, label: str) -> int | None:
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)
    for r in range(1, 60):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None


def _load_thesis(repo_root: Path, ticker: str) -> tuple[str, str]:
    path = repo_root / "micro_thesis" / "holdings" / f"{ticker.upper()}.json"
    if not path.exists():
        return ("(no thesis on file)", ticker.upper())
    try:
        raw: object = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return ("(thesis unreadable)", ticker.upper())
    data = cast("dict[str, object]", raw) if isinstance(raw, dict) else {}
    thesis = data.get("thesis")
    name = data.get("name")
    return (
        thesis if isinstance(thesis, str) and thesis else "(no thesis on file)",
        name if isinstance(name, str) and name else ticker.upper(),
    )


# --------------------------------------------------------------------------- #
# Prompt + parse
# --------------------------------------------------------------------------- #
def build_prompt(context: dict[str, object]) -> str:
    horizon = int(cast("int", context.get("horizon", 5)))
    fy_list = ", ".join(f"FY{i + 1}" for i in range(horizon))
    baseline_json = json.dumps(context.get("baseline", {}), indent=2)
    driver_lines = "\n".join(
        f'    "{k}": {{"values": [<{horizon} numbers>], "note": "<why>"}},' for k in _DRIVER_TO_ATTR
    )
    return f"""\
You are a senior equity analyst building a DCF for {context.get("name")} ({context.get("ticker")}).

INVESTMENT THESIS (the narrative your numbers must reflect):
{context.get("thesis")}

RECENT HISTORICALS ($M unless noted; oldest->newest):
{context.get("historicals")}

NAIVE BASELINE ASSUMPTIONS — the current method holds every ratio FLAT at its TTM
value with growth fading linearly. This is the poor starting point you are
improving (decimals for %/ratios; dso/dpo in days; capex_to_da is a multiple):
{baseline_json}

The forecast horizon is {horizon} years ({fy_list}).

VALUATION BEST PRACTICE:
{_PRINCIPLES}

TASK: reason bottom-up about where each driver should NORMALIZE over the horizon
for THIS specific business — is it over- or under-earning today? where do margins
settle at maturity? how fast does growth fade? — then output the improved grid.

Output ONLY a JSON object (no prose outside it), exactly this shape:
{{
  "narrative": "<2-4 sentences: the forecast story>",
  "drivers": {{
{driver_lines}
  }}
}}
Units: %-fields and ratios are DECIMALS (0.25 == 25%); dso_days/dpo_days are days;
capex_to_da is a multiple (1.0 == capex equals D&A). Each "values" array has
EXACTLY {horizon} numbers, FY1 first."""


def parse_response(text: str, ticker: str, horizon: int) -> LlmAssumptions:
    """Extract the JSON object and validate each driver's per-year values.

    Drivers that are missing or malformed are simply omitted (the caller keeps
    the baseline for those), so a partial response still improves what it can.
    """
    obj = _extract_json(text)
    if obj is None:
        raise LlmAssumptionError("no JSON object found in the LLM response")
    narrative = obj.get("narrative")
    drivers_raw = obj.get("drivers")
    if not isinstance(drivers_raw, dict):
        raise LlmAssumptionError("response has no 'drivers' object")
    drivers: dict[str, list[float]] = {}
    notes: dict[str, str] = {}
    for key in _DRIVER_TO_ATTR:
        entry = cast("dict[str, object]", drivers_raw).get(key)
        if not isinstance(entry, dict):
            continue
        vals_raw = cast("dict[str, object]", entry).get("values")
        if not isinstance(vals_raw, list):
            continue
        nums = [float(v) for v in cast("list[object]", vals_raw) if isinstance(v, (int, float))]
        if len(nums) < horizon:
            continue  # incomplete row — keep baseline for it
        drivers[key] = nums[:horizon]
        note = cast("dict[str, object]", entry).get("note")
        notes[key] = note if isinstance(note, str) else ""
    if not drivers:
        raise LlmAssumptionError("no usable driver rows parsed from the response")
    return LlmAssumptions(
        ticker=ticker.upper(),
        horizon=horizon,
        narrative=narrative if isinstance(narrative, str) else "",
        drivers=drivers,
        notes=notes,
    )


def _extract_json(text: str) -> dict[str, object] | None:
    # Tolerate ```json fences and leading/trailing prose: grab the outermost {...}.
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    blob = fenced.group(1) if fenced else None
    if blob is None:
        start = text.find("{")
        end = text.rfind("}")
        blob = text[start : end + 1] if start != -1 and end > start else None
    if blob is None:
        return None
    try:
        parsed: object = json.loads(blob)
    except json.JSONDecodeError:
        return None
    return cast("dict[str, object]", parsed) if isinstance(parsed, dict) else None


# --------------------------------------------------------------------------- #
# Generate (LLM) + apply (workbook)
# --------------------------------------------------------------------------- #
def cache_path(repo_root: Path, ticker: str) -> Path:
    return repo_root / "data" / "dcf_assumptions" / f"{ticker.upper()}.json"


def generate_assumptions(
    ticker: str,
    workbook_path: Path,
    repo_root: Path,
    *,
    force: bool = False,
    run_id: str | None = None,
) -> LlmAssumptions:
    """Call Opus (cached on disk) and return the parsed per-year overrides."""
    cache = cache_path(repo_root, ticker)
    horizon = int(cast("int", gather_context(ticker, workbook_path, repo_root)["horizon"]))
    if cache.exists() and not force:
        return _from_cache(cache, ticker, horizon)

    context = gather_context(ticker, workbook_path, repo_root)
    horizon = int(cast("int", context["horizon"]))
    from llm.cli import call_llm

    response = call_llm(
        build_prompt(context),
        purpose="dcf_assumptions",
        ticker=ticker.upper(),
        run_id=run_id,
    )
    assumptions = parse_response(response, ticker, horizon)
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_text(
        json.dumps(
            {
                "ticker": assumptions.ticker,
                "narrative": assumptions.narrative,
                "drivers": {
                    k: {"values": v, "note": assumptions.notes.get(k, "")}
                    for k, v in assumptions.drivers.items()
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    return assumptions


def _from_cache(cache: Path, ticker: str, horizon: int) -> LlmAssumptions:
    raw: object = json.loads(cache.read_text(encoding="utf-8"))
    data = cast("dict[str, object]", raw) if isinstance(raw, dict) else {}
    drivers_raw = data.get("drivers")
    drivers: dict[str, list[float]] = {}
    notes: dict[str, str] = {}
    if isinstance(drivers_raw, dict):
        for key in _DRIVER_TO_ATTR:
            entry = cast("dict[str, object]", drivers_raw).get(key)
            if not isinstance(entry, dict):
                continue
            vals = cast("dict[str, object]", entry).get("values")
            if isinstance(vals, list):
                nums = [float(v) for v in cast("list[object]", vals) if isinstance(v, (int, float))]
                if len(nums) >= horizon:
                    drivers[key] = nums[:horizon]
                    note = cast("dict[str, object]", entry).get("note")
                    notes[key] = note if isinstance(note, str) else ""
    narrative = data.get("narrative")
    return LlmAssumptions(
        ticker=ticker.upper(),
        horizon=horizon,
        narrative=narrative if isinstance(narrative, str) else "",
        drivers=drivers,
        notes=notes,
    )


def apply_to_workbook(workbook_path: Path, assumptions: LlmAssumptions) -> int:
    """Overwrite the Forecast INPUTS grid with the LLM's per-year drivers,
    preserving the scalars (base revenue, shares, terminal multiple, horizon).
    Returns the number of driver rows replaced. Recompute via the refresher
    afterwards to push the new inputs into the Valuation sheet + dcf_runs."""
    wb = openpyxl.load_workbook(str(workbook_path))
    fws = wb[forecast_mod_sheet_name()]
    baseline = forecast_mod.read_inputs_from_sheet(fws)
    overrides: dict[str, list[float]] = {}
    for key, attr in _DRIVER_TO_ATTR.items():
        vals = assumptions.drivers.get(key)
        if vals and len(vals) >= baseline.forecast_years:
            overrides[attr] = vals[: baseline.forecast_years]
    if not overrides:
        return 0
    improved = dataclasses.replace(baseline, **overrides)
    forecast_mod.write_inputs_section(fws, improved)
    wb.save(str(workbook_path))
    return len(overrides)
