"""Export the Personal-CIO substrate to an .xlsx workbook.

The CIO content — alerts → their queued actions → the accepted thesis-ledger
entries — was surfaced as HTML (digest / feed / evidence drawer) but had no
spreadsheet export, unlike the per-ticker financial report
(``report/renderers/workbook.py``). This is the cross-ticker CIO equivalent:
one workbook, three sheets, read from the substrate via the alerts / ledger
stores.

Read-only and resilient: it degrades to a workbook with header-only sheets when
the substrate is empty, so "export my CIO" always produces a valid file.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Font

from alerts import AlertRow, list_alerts, list_queued_actions_for_alert
from identity import DEFAULT_USER_ID
from user_state.ledger import list_recent_entries


def _write_header(ws: object, headers: list[str], widths: list[int]) -> None:
    ws.append(headers)  # type: ignore[attr-defined]
    for cell in ws[1]:  # type: ignore[index]
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"  # type: ignore[attr-defined]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width  # type: ignore[attr-defined]


def _stamp(dt: datetime | None) -> str:
    # AlertRow/ledger datetimes are naive-UTC; render minute precision.
    return dt.isoformat(sep=" ", timespec="minutes") if dt is not None else ""


def _alert_summary(alert: AlertRow) -> str:
    """The alert's at-a-glance line — the evidence_json ``summary`` (or ``memo``)."""
    try:
        evidence = json.loads(alert.evidence_json)
    except (ValueError, TypeError):
        return ""
    if isinstance(evidence, dict):
        data = cast("dict[str, object]", evidence)
        return str(data.get("summary") or data.get("memo") or "")
    return ""


def _payload_str(payload: object) -> str:
    try:
        return json.dumps(payload, default=str, sort_keys=True)[:600]
    except (TypeError, ValueError):
        return str(payload)[:600]


def export_cio_workbook(
    output_path: Path,
    *,
    user_id: str = DEFAULT_USER_ID,
    db_path: Path | None = None,
    alert_limit: int = 500,
    ledger_limit: int = 500,
) -> Path:
    """Write the Personal-CIO workbook (Alerts / Queued Actions / Thesis Ledger)
    to ``output_path`` and return it. Sheets are header-only when empty."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet  # type: ignore[arg-type]

    alerts = list_alerts(user_id=user_id, limit=alert_limit, db_path=db_path)

    ws_alerts = wb.create_sheet("Alerts")
    _write_header(
        ws_alerts,
        ["id", "fired_at", "ticker", "trigger_kind", "status", "summary", "signature"],
        [6, 18, 10, 16, 11, 80, 20],
    )
    for alert in alerts:
        ws_alerts.append(
            [
                alert.id,
                _stamp(alert.fired_at),
                alert.ticker,
                alert.trigger_kind,
                alert.status,
                _alert_summary(alert),
                alert.signature_sha[:16],
            ]
        )

    ws_actions = wb.create_sheet("Queued Actions")
    _write_header(
        ws_actions,
        [
            "id",
            "alert_id",
            "ticker",
            "action_kind",
            "status",
            "created_at",
            "applied_at",
            "payload",
        ],
        [6, 9, 10, 22, 11, 18, 18, 80],
    )
    for alert in alerts:
        for action in list_queued_actions_for_alert(alert.id, db_path):
            ws_actions.append(
                [
                    action.id,
                    action.alert_id,
                    alert.ticker,
                    action.action_kind,
                    action.status,
                    _stamp(action.created_at),
                    _stamp(action.applied_at),
                    _payload_str(action.payload),
                ]
            )

    ws_ledger = wb.create_sheet("Thesis Ledger")
    _write_header(
        ws_ledger,
        ["id", "accepted_at", "ticker", "entry_kind", "source_alert_id", "body"],
        [6, 18, 10, 22, 16, 100],
    )
    for entry in list_recent_entries(user_id=user_id, limit=ledger_limit, db_path=db_path):
        ws_ledger.append(
            [
                entry.id,
                _stamp(entry.accepted_at),
                entry.ticker,
                entry.entry_kind,
                entry.source_alert_id if entry.source_alert_id is not None else "",
                entry.body,
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
