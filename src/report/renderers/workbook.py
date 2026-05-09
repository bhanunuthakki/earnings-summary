"""ReportSpec → DCF + supporting workbook (.xlsx).

GSheets-compatibility constraints:
  - No conditional formatting macros, no sparklines, no array formulas
  - Standard formulas only (SUM, NPV, IF, INDEX/MATCH, simple arithmetic)
  - Freeze panes ok; named ranges ok; basic number formats ok

Tabs:
  README                — generation metadata + missing-data report
  DCF_Quarterly         — historical revenue strip + annual forecast → NPV (formulas)
  DCF_Segments          — per-segment annual block, summed at top
  Quarterly_Financials  — last 12 quarters wide-form (IS / BS / CF)
  Annual_Financials     — last 10 fiscal years wide-form (placeholder until annual query wired)
  Segments_Quarterly    — 12 quarters × segments × {revenue, OI}
  KPIs                  — tier-1 KPI history with break conditions
  Provenance            — source-doc audit trail
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report.models import (
    QuarterlyLineItem,
    ReportSpec,
    SectionStatus,
    SegmentSeries,
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
SUBHEADER_FONT = Font(bold=True)


def render(spec: ReportSpec, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    _readme(wb, spec)
    _dcf_quarterly(wb, spec)
    _dcf_segments(wb, spec)
    _quarterly_financials(wb, spec)
    _annual_financials(wb, spec)
    _segments_quarterly(wb, spec)
    _kpis(wb, spec)
    _provenance(wb, spec)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _header_row(ws: object, row: int, values: list[str]) -> None:
    sheet = ws  # type: ignore[assignment]
    for col, val in enumerate(values, 1):
        cell = sheet.cell(row=row, column=col, value=val)  # type: ignore[attr-defined]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _subheader_row(ws: object, row: int, values: list[str]) -> None:
    sheet = ws  # type: ignore[assignment]
    for col, val in enumerate(values, 1):
        cell = sheet.cell(row=row, column=col, value=val)  # type: ignore[attr-defined]
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT


def _autosize(ws: object, max_col: int, width: int = 18) -> None:
    sheet = ws  # type: ignore[assignment]
    for c in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(c)].width = width  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------


def _readme(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("README", 0)
    ws["A1"] = f"{spec.ticker} — research workbook"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Generation date"
    ws["B3"] = spec.generation_date.isoformat()
    ws["A4"] = "Run ID"
    ws["B4"] = spec.run_id or "—"
    ws["A5"] = "Repo root"
    ws["B5"] = spec.repo_root

    ws["A7"] = "Section status"
    ws["A7"].font = Font(bold=True)
    statuses = [
        ("§1 Snapshot", spec.snapshot.status, _detail(spec.snapshot.missing)),
        ("§2 Thesis", spec.thesis.status, _detail(spec.thesis.missing)),
        ("§3 Financials (12Q)", spec.financials.status, _detail(spec.financials.missing)),
        ("§4 Segments (12Q)", spec.segments.status, _detail(spec.segments.missing)),
        ("§5 Earnings analysis", spec.earnings.status, _detail(spec.earnings.missing)),
        ("§6 IR documents", spec.ir_docs.status, _detail(spec.ir_docs.missing)),
        ("§7 Bear case", spec.bear_case.status, _detail(spec.bear_case.missing)),
        ("§8 Provenance", spec.provenance.status, _detail(spec.provenance.missing)),
    ]
    _header_row(ws, 8, ["Section", "Status", "Detail / fix command"])
    for i, (label, status, detail) in enumerate(statuses, start=9):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=status.value)
        ws.cell(row=i, column=3, value=detail)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 80


def _detail(missing: object) -> str:
    if missing is None:
        return ""
    fix = getattr(missing, "fix_command", "")
    detail = getattr(missing, "detail", "") or ""
    return f"{fix}  ·  {detail}".strip(" ·")


def _dcf_quarterly(wb: Workbook, spec: ReportSpec) -> None:
    """Historical quarterly revenue strip + annual-forecast DCF block.

    Cell layout (live formulas so user can tune assumptions in-sheet):
      Row 1: header
      Row 3: WACC, terminal-g (assumption inputs)
      Row 6+: 16Q historical revenue (oldest→newest) for context
      Row 25+: 10Y annual forecast (revenue, FCF, discount factor, PV)
      Row 40+: NPV summary
    """
    ws = wb.create_sheet("DCF_Quarterly")
    ws["A1"] = f"{spec.ticker} — DCF (quarterly history → annual forecast)"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "WACC"
    ws["B3"] = spec.snapshot.valuation.wacc or 0.09
    ws["B3"].number_format = "0.00%"
    ws["A4"] = "Terminal growth"
    ws["B4"] = spec.snapshot.valuation.terminal_growth or 0.025
    ws["B4"].number_format = "0.00%"
    ws["A5"] = "Periods / year"
    ws["B5"] = 1  # annual forecast horizon

    revenue = _find_line_item(spec, "Revenue")
    fcf = _find_line_item(spec, "Free cash flow")

    _subheader_row(ws, 7, ["Historical quarter", "Revenue (M)", "FCF (M)"])
    if revenue and fcf:
        for i, q in enumerate(revenue.quarters):
            ws.cell(row=8 + i, column=1, value=q)
            ws.cell(row=8 + i, column=2, value=revenue.values[i] if i < len(revenue.values) else None)
            ws.cell(row=8 + i, column=3, value=fcf.values[i] if i < len(fcf.values) else None)
    else:
        ws["A8"] = "(no quarterly revenue / FCF — section §3 missing)"

    forecast_start_row = 25
    ws.cell(row=forecast_start_row - 1, column=1, value="Annual forecast (10Y)").font = Font(bold=True)
    _subheader_row(
        ws,
        forecast_start_row,
        ["Year offset", "Revenue growth", "Revenue", "FCF margin", "FCF", "Discount factor", "PV of FCF"],
    )
    base_revenue = _ttm_from_series(revenue.values) if revenue else None
    if base_revenue is not None:
        ws.cell(row=forecast_start_row + 1, column=3, value=base_revenue)
        ws.cell(row=forecast_start_row + 1, column=1, value=0)
        ws.cell(row=forecast_start_row + 1, column=2, value=0).number_format = "0.0%"
        ws.cell(row=forecast_start_row + 1, column=4, value=0.20).number_format = "0.0%"
        ws.cell(row=forecast_start_row + 1, column=5, value=f"=C{forecast_start_row + 1}*D{forecast_start_row + 1}")
        ws.cell(row=forecast_start_row + 1, column=6, value=1.0)
        ws.cell(row=forecast_start_row + 1, column=7, value=f"=E{forecast_start_row + 1}*F{forecast_start_row + 1}")
        for i in range(1, 11):
            r = forecast_start_row + 1 + i
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=0.10).number_format = "0.0%"
            ws.cell(row=r, column=3, value=f"=C{r - 1}*(1+B{r})")
            ws.cell(row=r, column=4, value=f"=D{r - 1}").number_format = "0.0%"
            ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
            ws.cell(row=r, column=6, value=f"=1/((1+$B$3)^A{r})")
            ws.cell(row=r, column=7, value=f"=E{r}*F{r}")

        npv_row = forecast_start_row + 13
        ws.cell(row=npv_row, column=1, value="Sum PV of FCFs").font = Font(bold=True)
        ws.cell(row=npv_row, column=7, value=f"=SUM(G{forecast_start_row + 1}:G{forecast_start_row + 11})")
        ws.cell(row=npv_row + 1, column=1, value="Terminal value (Gordon)").font = Font(bold=True)
        ws.cell(
            row=npv_row + 1,
            column=7,
            value=f"=E{forecast_start_row + 11}*(1+$B$4)/($B$3-$B$4)/((1+$B$3)^10)",
        )
        ws.cell(row=npv_row + 2, column=1, value="NPV (enterprise)").font = Font(bold=True)
        ws.cell(row=npv_row + 2, column=7, value=f"=G{npv_row}+G{npv_row + 1}")
    else:
        ws.cell(row=forecast_start_row + 1, column=1, value="(base revenue unavailable — populate §3 first)")

    ws.freeze_panes = "B2"
    _autosize(ws, max_col=8)


def _ttm_from_series(values: list[float | None]) -> float | None:
    if not values or len(values) < 4:
        return None
    last_4 = values[-4:]
    if any(v is None for v in last_4):
        return None
    return sum(v for v in last_4 if v is not None)


def _find_line_item(spec: ReportSpec, name: str) -> QuarterlyLineItem | None:
    for li in spec.financials.line_items:
        if li.line_item == name:
            return li
    return None


def _dcf_segments(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("DCF_Segments")
    ws["A1"] = f"{spec.ticker} — segment DCF (per-segment annual)"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "WACC"
    ws["B3"] = spec.snapshot.valuation.wacc or 0.09
    ws["B3"].number_format = "0.00%"
    ws["A4"] = "Terminal growth"
    ws["B4"] = spec.snapshot.valuation.terminal_growth or 0.025
    ws["B4"].number_format = "0.00%"

    if spec.segments.status != SectionStatus.OK or not spec.segments.revenue_by_product:
        ws["A6"] = "(segment data missing — populate §4 first)"
        return

    _subheader_row(ws, 6, ["Segment", "TTM revenue (M)", "Assumed annual growth", "Assumed FCF margin", "NPV (M, Y0-Y10 + TV)"])
    for i, seg in enumerate(spec.segments.revenue_by_product):
        r = 7 + i
        ttm = _ttm_from_series(seg.values)
        ws.cell(row=r, column=1, value=seg.segment_name)
        ws.cell(row=r, column=2, value=ttm)
        ws.cell(row=r, column=3, value=0.10).number_format = "0.0%"
        ws.cell(row=r, column=4, value=0.20).number_format = "0.0%"
        ws.cell(
            row=r,
            column=5,
            value=(
                f"=IFERROR(B{r}*D{r}*((1-((1+C{r})/(1+$B$3))^10)/($B$3-C{r})) "
                f"+ B{r}*D{r}*((1+C{r})^10)*(1+$B$4)/($B$3-$B$4)/((1+$B$3)^10), \"-\")"
            ),
        )
    total_row = 7 + len(spec.segments.revenue_by_product) + 1
    ws.cell(row=total_row, column=1, value="Total NPV").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"=SUM(E7:E{total_row - 1})")

    ws.freeze_panes = "B2"
    _autosize(ws, max_col=6)


def _quarterly_financials(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("Quarterly_Financials")
    ws["A1"] = f"{spec.ticker} — quarterly financials (last 12)"
    ws["A1"].font = Font(bold=True, size=12)

    if spec.financials.status != SectionStatus.OK or not spec.financials.line_items:
        ws["A3"] = "(missing — populate §3 first)"
        return

    headers = ["Line item", "Unit", *spec.financials.quarter_labels, "QoQ", "YoY", "1Y CAGR (TTM)", "3Y CAGR (TTM)"]
    _header_row(ws, 3, headers)
    for i, li in enumerate(spec.financials.line_items):
        r = 4 + i
        ws.cell(row=r, column=1, value=li.line_item)
        ws.cell(row=r, column=2, value=li.unit)
        for j, v in enumerate(li.values):
            ws.cell(row=r, column=3 + j, value=v)
        n_quarters = len(li.values)
        col = 3 + n_quarters
        for v in (li.growth.qoq, li.growth.yoy, li.growth.cagr_1y_ttm, li.growth.cagr_3y_ttm):
            ws.cell(row=r, column=col, value=v).number_format = "0.0%"
            col += 1

    ws.freeze_panes = "C4"
    _autosize(ws, max_col=len(headers))


def _annual_financials(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("Annual_Financials")
    ws["A1"] = f"{spec.ticker} — annual financials (last {len(spec.financials.annual_years) or 10} FY)"
    ws["A1"].font = Font(bold=True, size=12)

    if not spec.financials.annual_line_items:
        ws["A3"] = "(no annual rows in metrics view — run extract_facts on annual statement docs)"
        return

    headers = ["Line item", "Unit", *[str(y) for y in spec.financials.annual_years]]
    _header_row(ws, 3, headers)
    for i, li in enumerate(spec.financials.annual_line_items):
        r = 4 + i
        ws.cell(row=r, column=1, value=li.line_item)
        ws.cell(row=r, column=2, value=li.unit)
        for j, v in enumerate(li.values):
            ws.cell(row=r, column=3 + j, value=v)
    ws.freeze_panes = "C4"
    _autosize(ws, max_col=len(headers))


def _segments_quarterly(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("Segments_Quarterly")
    ws["A1"] = f"{spec.ticker} — segments (last 12 quarters)"
    ws["A1"].font = Font(bold=True, size=12)

    if spec.segments.status != SectionStatus.OK:
        ws["A3"] = "(missing — populate §4 first)"
        return

    row = 3
    for label, group in (
        ("Revenue by product", spec.segments.revenue_by_product),
        ("Revenue by geography", spec.segments.revenue_by_geography),
        ("Operating income", spec.segments.operating_income),
    ):
        if not group:
            continue
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        row += 1
        headers = ["Segment", "Unit", *spec.segments.quarter_labels, "QoQ", "YoY", "1Y CAGR", "3Y CAGR"]
        _header_row(ws, row, headers)
        row += 1
        for s in group:
            row = _write_segment_row(ws, row, s)
        row += 2
    ws.freeze_panes = "C4"
    _autosize(ws, max_col=20)


def _write_segment_row(ws: object, row: int, s: SegmentSeries) -> int:
    sheet = ws  # type: ignore[assignment]
    sheet.cell(row=row, column=1, value=s.segment_name)  # type: ignore[attr-defined]
    sheet.cell(row=row, column=2, value=s.unit)  # type: ignore[attr-defined]
    for j, v in enumerate(s.values):
        sheet.cell(row=row, column=3 + j, value=v)  # type: ignore[attr-defined]
    col = 3 + len(s.values)
    for v in (s.growth.qoq, s.growth.yoy, s.growth.cagr_1y_ttm, s.growth.cagr_3y_ttm):
        cell = sheet.cell(row=row, column=col, value=v)  # type: ignore[attr-defined]
        cell.number_format = "0.0%"
        col += 1
    return row + 1


def _kpis(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("KPIs")
    ws["A1"] = f"{spec.ticker} — tier-1 / tier-2 / tier-3 KPI ledger"
    ws["A1"].font = Font(bold=True, size=12)

    if not spec.thesis.kpi_ledger:
        ws["A3"] = "(missing — holdings JSON not loaded or empty)"
        return

    _header_row(ws, 3, ["Tier", "KPI", "Unit", "Source", "Break", "Status"])
    for i, k in enumerate(spec.thesis.kpi_ledger):
        r = 4 + i
        ws.cell(row=r, column=1, value=k.tier)
        ws.cell(row=r, column=2, value=k.name)
        ws.cell(row=r, column=3, value=k.unit or "")
        ws.cell(row=r, column=4, value=k.source_hint or "")
        ws.cell(row=r, column=5, value=k.break_condition or "")
        ws.cell(row=r, column=6, value=k.current_status)
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60


def _provenance(wb: Workbook, spec: ReportSpec) -> None:
    ws = wb.create_sheet("Provenance")
    ws["A1"] = f"{spec.ticker} — source-doc audit"
    ws["A1"].font = Font(bold=True, size=12)

    if not spec.provenance.source_docs:
        ws["A3"] = "(no documents recorded for this ticker)"
        return

    _header_row(ws, 3, ["doc_type", "period_end", "file_path", "sha256", "fetched_at"])
    for i, d in enumerate(spec.provenance.source_docs):
        r = 4 + i
        ws.cell(row=r, column=1, value=d.doc_type)
        ws.cell(row=r, column=2, value=d.period_end or "")
        ws.cell(row=r, column=3, value=d.file_path)
        ws.cell(row=r, column=4, value=d.sha256 or "")
        ws.cell(row=r, column=5, value=d.fetched_at or "")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
