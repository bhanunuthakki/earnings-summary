"""Deterministic classifier for manually-uploaded IR documents.

Given a file dropped in `ir_documents/` at the root level, decide its
`(ticker, doc_type, period_end)` from filename heuristics + first-page content
fingerprinting. This is the manual-upload counterpart to the URL-manifest based
`fetch_ir_documents.py` flow.

No LLM calls. Every classification is rule-driven and explainable: the
`CategorizationResult` carries the textual evidence used. If neither filename
nor content gives a confident answer, a `CategorizationFailure` is returned and
the file is quarantined — never silently dropped.
"""

from __future__ import annotations

import hashlib
import logging
import re
from collections.abc import Sequence
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from models.documents import DocType
from models.ir_uploads import (
    CategorizationFailure,
    CategorizationResult,
    Confidence,
)

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Issuer / ticker registry
# ---------------------------------------------------------------------------


# (ticker, fiscal_calendar_id, ordered list of issuer-name substrings that
# appear in cover-page text). Order matters: more specific names first to
# avoid false-positive collisions (e.g. "Mercado Libre" before "MELI").
class FiscalCalendar(str):
    """Calendar ID — see `_period_end_for()` for mapping."""


_CAL_CALENDAR = "calendar"  # FY-end Dec 31 — Q1=Mar-31, Q2=Jun-30, Q3=Sep-30, Q4=Dec-31
_CAL_VEEV = "veeva"  # FY-end Jan 31 (Q1 ends Apr-30, Q4 ends Jan-31). RBRK shares this.
_CAL_APR30 = "apr30"  # generic FY-end Apr 30 — Q1=Jul-31, Q4=Apr-30. (NOT Rubrik; see below.)
_CAL_NVO = "nvo"  # Calendar, but H1=Q2, 9M=Q3, FY=Q4
_CAL_VISA = "visa"  # FY-end Sep 30 — Q1=Dec-31, Q2=Mar-31, Q3=Jun-30, Q4=Sep-30
_CAL_ORACLE = "oracle"  # FY-end May 31 — Q1=Aug-31, Q2=Nov-30, Q3=Feb-28, Q4=May-31

ISSUER_REGISTRY: list[tuple[str, str, tuple[str, ...]]] = [
    ("MELI", _CAL_CALENDAR, ("MercadoLibre", "Mercado Libre", "Mercado  Libre")),
    ("NU", _CAL_CALENDAR, ("Nu Holdings", "Nu's Investor", "Nubank")),
    # Rubrik has a January-31 fiscal year-end (FY26 Q1 ended Apr-30-2025) — the
    # SAME calendar as Veeva. It is NOT an April-30 issuer; the prior `_CAL_RUBRIK`
    # mapping shifted every RBRK period one quarter forward. See the transcript-
    # ingest path (`src/compute/transcript_ingest.py::_FYE_OFFSETS`) which also
    # treats RBRK as Jan-FYE.
    ("RBRK", _CAL_VEEV, ("Rubrik",)),
    ("NOW", _CAL_CALENDAR, ("ServiceNow",)),
    ("WIX", _CAL_CALENDAR, ("Wix.com", "Wix Ltd", "Wix's", "Wix ", "WIX ")),
    (
        "NVO",
        _CAL_NVO,
        (
            "Novo Nordisk",
            "Amounts in DKK million",
            "Amounts are in DKK",
            "Amounts  in  DKK",
        ),
    ),
    ("GOOG", _CAL_CALENDAR, ("Alphabet Inc", "Alphabet's")),
    (
        "META",
        _CAL_CALENDAR,
        (
            "Meta Platforms",
            "Meta Reports",
            "Meta Earnings",
            "investor.atmeta.com",
            "atmeta.com",
        ),
    ),
    ("AMZN", _CAL_CALENDAR, ("Amazon.com", "AMAZON.COM")),
    ("VEEV", _CAL_VEEV, ("Veeva Systems", "Veeva ")),
    # BN = Brookfield Corporation ONLY. "Brookfield Asset Management" is the
    # separate listed entity BAM (its own tracked ticker, synced into the store
    # registry with that alias) — claiming it here routed BAM's own filings to BN
    # in the content-fingerprint path, since curated entries win over the store.
    # (Pre-Dec-2022, BN was *named* "Brookfield Asset Management Inc."; those
    # historical covers would now route to BAM, but BN's IR docs in scope are all
    # post-rename and the financials are CIK-keyed via FMP/SEC, so this is safe.)
    ("BN", _CAL_CALENDAR, ("Brookfield Corporation",)),
    ("LLY", _CAL_CALENDAR, ("Eli Lilly and Company", "Eli Lilly", "Lilly ")),
    (
        "ABNB",
        _CAL_CALENDAR,
        (
            "Airbnb, Inc.",
            "Airbnb Inc.",
            "Airbnb (ABNB)",
            "Airbnb's",
        ),
    ),
    (
        "BKNG",
        _CAL_CALENDAR,
        (
            "Booking Holdings, Inc.",
            "Booking Holdings Inc.",
            "Booking Holdings",
            "Booking.com Holdings",
        ),
    ),
    (
        "SOFI",
        _CAL_CALENDAR,
        (
            "SoFi Technologies",
            "Social Finance, Inc.",
            "SoFi Reports",
            "SoFi's",
        ),
    ),
    (
        "DLO",
        _CAL_CALENDAR,
        (
            "DLocal Limited",
            "(NASDAQ:DLO)",
            "dLocal",
            "DLocal",
        ),
    ),
    # Evaluation list (added for the headless IR-document auto-fetch path). The
    # auto-fetch flow attributes the ticker from the crawl (a trusted hint), so
    # these issuer-name substrings matter mainly for *manual* uploads of these
    # names; the fiscal calendars are what the period math needs.
    ("V", _CAL_VISA, ("Visa Inc.", "VISA INC", "Visa Inc")),
    ("ORCL", _CAL_ORACLE, ("Oracle Corporation", "ORACLE CORPORATION", "Oracle Corp")),
    ("TEM", _CAL_CALENDAR, ("Tempus AI, Inc.", "Tempus AI", "Tempus's")),
    ("CRWV", _CAL_CALENDAR, ("CoreWeave, Inc.", "CoreWeave Inc", "CoreWeave")),
    ("NBIS", _CAL_CALENDAR, ("Nebius Group", "Nebius Group N.V.", "Nebius")),
]


# The registry the classifier actually consults. Defaults to the curated
# ISSUER_REGISTRY above; the auto-fetch categorizer overrides it per-run with
# `curated + data/issuer_registry.json` (see issuer_registry.effective_entries)
# so eval/portfolio tickers stay in sync with tracked_companies without a code
# edit. None resets to the curated list.
_RUNTIME_REGISTRY: list[tuple[str, str, tuple[str, ...]]] | None = None


def set_runtime_registry(entries: list[tuple[str, str, tuple[str, ...]]] | None) -> None:
    """Install the effective registry for this process (None resets to curated)."""
    global _RUNTIME_REGISTRY
    _RUNTIME_REGISTRY = entries


def _effective_registry() -> list[tuple[str, str, tuple[str, ...]]]:
    return _RUNTIME_REGISTRY if _RUNTIME_REGISTRY is not None else ISSUER_REGISTRY


# ---------------------------------------------------------------------------
# Doc-type detection
# ---------------------------------------------------------------------------

# Each rule: (DocType, regex, evidence-label). First rule that matches wins.
# Patterns are deliberately precise — they look for cover-page artifact labels,
# not random prose. Anchored to compiled regex for speed; case-insensitive.
_DOC_TYPE_RULES: list[tuple[DocType, re.Pattern[str], str]] = [
    # Order matters. Specific cover-page / opening-paragraph signals first;
    # broad phrases that often appear in disclaimers (e.g. "Conference Call",
    # "Annual Report") sit at the end so they only fire when nothing more
    # decisive matched.
    # ---- SEC filing covers: anchored to the first ~500 chars so disclaimer
    # mentions ("filing on Form 10-K for the period end…") in IR presentations
    # don't false-positive as 10-K/10-Q filings. Real SEC filings put the form
    # designator on the cover page within the first few hundred chars. ----
    (
        DocType.SEC_10Q,
        re.compile(r"\A[\s\S]{0,500}\bFORM\s+10-?Q\b", re.IGNORECASE),
        "sec_form_10q",
    ),
    (
        DocType.SEC_10K,
        re.compile(r"\A[\s\S]{0,500}\bFORM\s+10-?K\b", re.IGNORECASE),
        "sec_form_10k",
    ),
    (
        DocType.SEC_20F,
        re.compile(r"\A[\s\S]{0,500}\bFORM\s+20-?F\b", re.IGNORECASE),
        "sec_form_20f",
    ),
    (DocType.SEC_8K, re.compile(r"\A[\s\S]{0,500}\bFORM\s+8-?K\b", re.IGNORECASE), "sec_form_8k"),
    (DocType.SEC_6K, re.compile(r"\A[\s\S]{0,500}\bFORM\s+6-?K\b", re.IGNORECASE), "sec_form_6k"),
    # ---- Transcripts: only the strongest cover-page signals up-front. ----
    (
        DocType.IR_TRANSCRIPT,
        re.compile(r"\A\s*Operator\s*[:.]", re.IGNORECASE),
        "transcript_operator_opener",
    ),
    (
        DocType.IR_TRANSCRIPT,
        re.compile(r"\b(Earnings\s+Call\s+Script|Prepared\s+Remarks)\b", re.IGNORECASE),
        "transcript_title",
    ),
    (
        DocType.IR_TRANSCRIPT,
        # FactSet CallStreet-published transcripts open with "Corrected Transcript"
        # in the cover-page header. Specific enough to be unambiguous.
        re.compile(r"\bCorrected\s+Transcript\b|FactSet\s+CallStreet", re.IGNORECASE),
        "transcript_factset",
    ),
    # ---- Shareholder letters first: their bodies contain press-release-
    # style "today reported financial results" language verbatim, so the
    # cover-page title has to win over those broader signals. ----
    (
        DocType.IR_INVESTOR_UPDATE,
        re.compile(
            r"\b(Letter\s+to\s+(?:Our\s+)?Shareholders?|To\s+Our\s+Shareholders?|Shareholder\s+Letter)\b",
            re.IGNORECASE,
        ),
        "shareholder_letter_label",
    ),
    # ---- Non-quarterly events: investor days, AGMs, capital markets days,
    # broker conferences. Period_end on these rows is the event date, not a
    # fiscal-quarter end (handled in `classify_ir_file`). ----
    (
        DocType.IR_EVENT,
        re.compile(r"\bInvestor\s+Day\b", re.IGNORECASE),
        "investor_day_label",
    ),
    (
        DocType.IR_EVENT,
        re.compile(
            r"\b(?:Capital\s+Markets\s+Day|JPMorgan\s+Healthcare\s+Conference|J\.?P\.?\s*Morgan\s+Healthcare\s+Conference|Healthcare\s+Conference|Investor\s+Conference)\b",
            re.IGNORECASE,
        ),
        "conference_label",
    ),
    # ---- Presentations: cover-page titles. ----
    (
        DocType.IR_PRESENTATION,
        re.compile(r"\b(Investor|Earnings|Results)\s*Presentation\b", re.IGNORECASE),
        "presentation_label",
    ),
    (
        DocType.IR_PRESENTATION,
        re.compile(r"\bEarnings\s*Slides?\b", re.IGNORECASE),
        "earnings_slides_label",
    ),
    (
        DocType.IR_PRESENTATION,
        re.compile(r"\bCompany\s+Overview\b", re.IGNORECASE),
        "company_overview_label",
    ),
    # ---- Press releases: opening-paragraph / dateline signals. ----
    (
        DocType.IR_PRESS_RELEASE,
        re.compile(
            r"\btoday\s+(?:reported|announced|reports|announces)\s+(?:its\s+)?(?:financial\s+)?results?\b",
            re.IGNORECASE,
        ),
        "today_reported",
    ),
    (
        DocType.IR_PRESS_RELEASE,
        re.compile(
            r"\bReports?\s+(First|Second|Third|Fourth)\s+Quarter\b",
            re.IGNORECASE,
        ),
        "reports_quarter_phrase",
    ),
    (
        DocType.IR_PRESS_RELEASE,
        # dLocal / NU title pattern: "Q1'25 Earnings Release" as the document
        # title. Anchored to the first ~100 chars so transcripts that
        # casually reference "If you have not seen the Earnings Release..."
        # don't false-positive as a press release.
        re.compile(r"\A[\s\S]{0,100}\bEarnings\s+Release\b", re.IGNORECASE),
        "earnings_release_title",
    ),
    (
        DocType.IR_PRESS_RELEASE,
        # Press-release dateline: "CITY[, STATE-OR-COUNTRY][;,] Month DD, YYYY".
        # Accept all-caps ("MONTEVIDEO, Uruguay; October 29, 2025") *and* mixed
        # case ("Montevideo, Uruguay, May 14, 2025") since dLocal/NU IR releases
        # use the latter.
        re.compile(
            r"^[A-Z][A-Za-z\s]{2,40}(?:,\s*[A-Za-z\.]+)?[;,\s\-]+"
            r"(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}",
            re.MULTILINE,
        ),
        "press_release_dateline",
    ),
    # ---- Loose transcript fallback: only fires when none of the above
    # presentation/press-release/transcript-title rules matched. Disclaimers
    # in presentations and press releases routinely mention "earnings
    # conference call", which is why this is last. ----
    (
        DocType.IR_TRANSCRIPT,
        re.compile(r"\b(Earnings\s+Conference\s+Call|Conference\s+Call)\b", re.IGNORECASE),
        "transcript_phrase",
    ),
    # ---- Annual report tightened to require year context, so it doesn't
    # misfire on disclaimer phrases like "annual report on Form 20-F". ----
    (
        DocType.IR_INVESTOR_UPDATE,
        re.compile(
            r"\bAnnual\s+Report(?:\s+(?:for\s+|fiscal\s+year\s+))?\s+(?:20)?\d{2}\b",
            re.IGNORECASE,
        ),
        "annual_report_label",
    ),
]


# Rules to apply against `re.sub(r"\s+", "", text).lower()`. Maintain the same
# DocType outcomes as `_DOC_TYPE_RULES`, just glue the keywords together so we
# match pypdf's no-space output ("EarningsCall", "PreparedRemarks", etc).
_DOC_TYPE_SQUASHED_RULES: list[tuple[DocType, re.Pattern[str], str]] = [
    # Same priority ordering as `_DOC_TYPE_RULES`. Patterns assume
    # `re.sub(r"\s+", "", text).lower()` — i.e. pypdf's no-space output.
    # SEC FORM is intentionally NOT squashed: the boundary-anchored content
    # rule above is enough; relying on flat-text matching invites
    # collisions like "filingonform10-k" in disclaimers.
    (DocType.IR_TRANSCRIPT, re.compile(r"earningscallscript"), "transcript_title"),
    (DocType.IR_TRANSCRIPT, re.compile(r"preparedremarks"), "transcript_title"),
    (DocType.IR_TRANSCRIPT, re.compile(r"correctedtranscript"), "transcript_factset"),
    (DocType.IR_TRANSCRIPT, re.compile(r"factsetcallstreet"), "transcript_factset"),
    (DocType.IR_EVENT, re.compile(r"investorday"), "investor_day_label"),
    (DocType.IR_EVENT, re.compile(r"capitalmarketsday"), "conference_label"),
    (DocType.IR_EVENT, re.compile(r"healthcareconference"), "conference_label"),
    (DocType.IR_EVENT, re.compile(r"investorconference"), "conference_label"),
    (DocType.IR_INVESTOR_UPDATE, re.compile(r"lettertoshareholders"), "shareholder_letter_label"),
    (DocType.IR_INVESTOR_UPDATE, re.compile(r"toourshareholders"), "shareholder_letter_label"),
    (DocType.IR_INVESTOR_UPDATE, re.compile(r"shareholderletter"), "shareholder_letter_label"),
    (DocType.IR_PRESENTATION, re.compile(r"investorpresentation"), "presentation_label"),
    (DocType.IR_PRESENTATION, re.compile(r"earningspresentation"), "presentation_label"),
    (DocType.IR_PRESENTATION, re.compile(r"resultspresentation"), "presentation_label"),
    (DocType.IR_PRESENTATION, re.compile(r"companyoverview"), "company_overview_label"),
    (
        DocType.IR_PRESS_RELEASE,
        re.compile(r"todayreported(?:its)?(?:financial)?results"),
        "today_reported",
    ),
    (
        DocType.IR_PRESS_RELEASE,
        re.compile(r"reports(?:first|second|third|fourth)quarter"),
        "reports_quarter_phrase",
    ),
    (DocType.IR_TRANSCRIPT, re.compile(r"earningsconferencecall"), "transcript_phrase"),
    (DocType.IR_TRANSCRIPT, re.compile(r"earningscall"), "transcript_phrase"),
    (DocType.IR_TRANSCRIPT, re.compile(r"conferencecall"), "transcript_phrase"),
    (DocType.IR_INVESTOR_UPDATE, re.compile(r"annualreport20\d{2}"), "annual_report_label"),
]


# ---------------------------------------------------------------------------
# Period extraction
# ---------------------------------------------------------------------------

_RX_QUARTER_APOSTROPHE = re.compile(r"Q(?P<q>[1-4])['‘’ʼ]\s*(?P<yy>\d{2})\b")
_RX_QUARTER_SPACE = re.compile(r"Q(?P<q>[1-4])\s*(?P<y>20\d{2})\b")
_RX_QUARTER_FULL = re.compile(
    r"\b(?P<word>First|Second|Third|Fourth)\s+Quarter(?:\s+Fiscal)?\s+(?:20)?(?P<y>\d{2,4})\b",
    re.IGNORECASE,
)
_RX_QUARTER_FY = re.compile(
    r"\b(?P<word>First|Second|Third|Fourth)\s+Quarter\s+Fiscal\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
_RX_FY_QY = re.compile(r"\bQ(?P<q>[1-4])\s+FY\s*(?P<y>\d{2,4})\b", re.IGNORECASE)
_RX_NVO_PERIOD = re.compile(
    r"\b(?P<period>First\s+three\s+months|First\s+six\s+months|First\s+nine\s+months|Full\s+year)\s+(?:of\s+)?(?P<y>20\d{2})\b",
    re.IGNORECASE,
)

_WORD_TO_Q = {"first": 1, "second": 2, "third": 3, "fourth": 4}

# "Quarter Ended March 31, 2025" / "September 30 Quarter End 2025" — common in
# MELI's IR-published transcripts and conference-call decks.
_RX_QUARTER_ENDED_DATE = re.compile(
    r"\bQuarter\s+(?:E|e)nded?\s+"
    r"(?P<m>January|February|March|April|May|June|July|August|September|October|November|December)\s+"
    r"(?P<d>\d{1,2}),?\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
_RX_DATE_QUARTER_ENDED = re.compile(
    r"\b(?P<m>January|February|March|April|May|June|July|August|September|October|November|December)\s+"
    r"(?P<d>\d{1,2})\s+Quarter\s+End(?:ed)?\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
# "Fourth Quarter & Full Year 2025" / "Fourth Quarter and Full Year 2025"
_RX_QUARTER_FULL_YEAR = re.compile(
    r"\b(?P<word>First|Second|Third|Fourth)\s+Quarter\s+(?:&|and)\s+Full\s+Year\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
# 10-Q cover page: "For the quarterly period ended June 30, 2025"
_RX_SEC_QUARTERLY_PERIOD_ENDED = re.compile(
    r"\bFor\s+the\s+quarterly\s+period\s+ended[:\s]+"
    r"(?P<m>January|February|March|April|May|June|July|August|September|October|November|December)\s+"
    r"(?P<d>\d{1,2}),?\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
# 10-K cover page: "For the fiscal year ended: December 31, 2025"
_RX_SEC_FISCAL_YEAR_ENDED = re.compile(
    r"\bFor\s+the\s+fiscal\s+year\s+ended[:\s]+"
    r"(?P<m>January|February|March|April|May|June|July|August|September|October|November|December)\s+"
    r"(?P<d>\d{1,2}),?\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)
# Generic dateline used as event-date for IR_EVENT docs (Investor Day, JPM,
# Capital Markets Day decks). Match the FIRST month/day/year date in the
# fingerprint — these decks always lead with the event date on slide 1.
_RX_EVENT_DATE = re.compile(
    r"\b(?P<m>January|February|March|April|May|June|July|August|September|October|November|December)\s+"
    r"(?P<d>\d{1,2}),?\s+(?P<y>20\d{2})\b",
    re.IGNORECASE,
)

_MONTH_TO_Q: dict[str, int] = {
    "january": 1,
    "february": 1,
    "march": 1,
    "april": 2,
    "may": 2,
    "june": 2,
    "july": 3,
    "august": 3,
    "september": 3,
    "october": 4,
    "november": 4,
    "december": 4,
}


def _yy_to_yyyy(yy: int) -> int:
    """Two-digit year → four-digit; window: 2000-2099."""
    if yy < 100:
        return 2000 + yy
    return yy


def _period_end_for(ticker_calendar: str, year: int, q: int) -> date:
    """Map (calendar, year, q) → period-end ISO date.

    See `directives/fetch_ir_documents.md` for the per-issuer fiscal calendar.
    """
    if ticker_calendar == _CAL_CALENDAR:
        return [date(year, 3, 31), date(year, 6, 30), date(year, 9, 30), date(year, 12, 31)][q - 1]
    if ticker_calendar == _CAL_VEEV:
        # FY-N Q1 ends Apr 30 (N-1), Q4 ends Jan 31 (N).
        return [
            date(year - 1, 4, 30),
            date(year - 1, 7, 31),
            date(year - 1, 10, 31),
            date(year, 1, 31),
        ][q - 1]
    if ticker_calendar == _CAL_APR30:
        # Generic April-30 FYE: FY-N Q1 ends Jul 31 (N-1), Q4 ends Apr 30 (N).
        # No registered ticker uses this (Rubrik is Jan-FYE → _CAL_VEEV); it is
        # the calendar derived for a genuine April-30 issuer via fiscal_year_end.
        return [
            date(year - 1, 7, 31),
            date(year - 1, 10, 31),
            date(year, 1, 31),
            date(year, 4, 30),
        ][q - 1]
    if ticker_calendar == _CAL_NVO:
        return [date(year, 3, 31), date(year, 6, 30), date(year, 9, 30), date(year, 12, 31)][q - 1]
    if ticker_calendar == _CAL_VISA:
        # FY-N Q1 ends Dec 31 (N-1), Q4 ends Sep 30 (N).
        return [
            date(year - 1, 12, 31),
            date(year, 3, 31),
            date(year, 6, 30),
            date(year, 9, 30),
        ][q - 1]
    if ticker_calendar == _CAL_ORACLE:
        # FY-N Q1 ends Aug 31 (N-1), Q4 ends May 31 (N). Q3 uses Feb-28 as the
        # conventional last-day key (the existing calendars all use fixed dates).
        return [
            date(year - 1, 8, 31),
            date(year - 1, 11, 30),
            date(year, 2, 28),
            date(year, 5, 31),
        ][q - 1]
    if ticker_calendar.startswith("fye_"):
        # Generic non-December FYE derived from tracked_companies.fiscal_year_end.
        # 'fye_MM' = the fiscal year ENDS in month MM (year N); Q4 ends that
        # month-end and earlier quarters step back 3 months each. Curated entries
        # above override this for 52/53-week names whose ends aren't month-ends.
        try:
            fy_month = int(ticker_calendar[4:6])
        except ValueError:
            raise ValueError(f"Unknown calendar id: {ticker_calendar!r}") from None
        end_month, end_year = fy_month - 3 * (4 - q), year
        while end_month <= 0:
            end_month, end_year = end_month + 12, end_year - 1
        if end_month == 12:
            return date(end_year, 12, 31)
        return date(end_year, end_month + 1, 1) - timedelta(days=1)
    raise ValueError(f"Unknown calendar id: {ticker_calendar!r}")


def _calendar_for(ticker: str) -> str:
    for tk, cal, _ in _effective_registry():
        if tk == ticker:
            return cal
    raise ValueError(f"No calendar registered for ticker {ticker!r}")


# Map a `tracked_companies.fiscal_year_end` ("MM-DD") to a fiscal-calendar id.
_FYE_TO_CALENDAR: dict[str, str] = {
    "12-31": _CAL_CALENDAR,
    "01-31": _CAL_VEEV,
    "04-30": _CAL_APR30,
    "09-30": _CAL_VISA,
    "05-31": _CAL_ORACLE,
}


def calendar_id_from_fye(mmdd: str | None) -> str:
    """Derive a fiscal-calendar id from a ``fiscal_year_end`` ("MM-DD") string.

    Used by the headless auto-fetch path to attribute periods for tickers not in
    ``ISSUER_REGISTRY`` (the crawl supplies the ticker; the FYE supplies the
    calendar). Unknown / missing FYE falls back to the Dec-31 calendar. NVO's
    H1/9M/FY labeling can't be inferred from the FYE alone (it shares calendar's
    Dec-31), so NVO stays registry-only. An exact curated-table match wins (e.g.
    01-31 → veeva); any other non-December month falls back to the generic
    ``fye_MM`` calendar; December / missing / unparseable uses the Dec-31 calendar.
    """
    key = (mmdd or "").strip()
    if key in _FYE_TO_CALENDAR:
        return _FYE_TO_CALENDAR[key]
    try:
        month = int(key.split("-")[0])
    except (ValueError, IndexError):
        return _CAL_CALENDAR
    return f"fye_{month:02d}" if 1 <= month <= 11 else _CAL_CALENDAR


# ---------------------------------------------------------------------------
# Fingerprinting
# ---------------------------------------------------------------------------

_PDF_FINGERPRINT_PAGES = 2
_PDF_FINGERPRINT_CHARS = 6000


def _fingerprint_pdf_pymupdf(path: Path) -> str | None:
    """First ~2 pages via PyMuPDF (``fitz``), or None if it is unavailable/fails.

    PyMuPDF loads only the pages asked for and is markedly faster and far more
    robust than pypdf on large, graphics-heavy investor decks — a 200+ page
    Investor Day deck stalled pypdf for minutes (its xref-recovery rescan is the
    usual culprit) where fitz reads the first two pages in tens of milliseconds.
    Optional dep (bundled with the ``ir`` extra); a missing module or any parse
    error degrades to None so the caller falls back to pypdf — never a crash.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return None
    doc = None
    try:
        doc = fitz.open(str(path))
        n = min(_PDF_FINGERPRINT_PAGES, doc.page_count)
        chunks = [doc.load_page(i).get_text() or "" for i in range(n)]
    except Exception:  # PyMuPDF's error tree is wide; degrade to the pypdf path
        return None
    finally:
        if doc is not None:
            doc.close()
    return "\n".join(chunks)[:_PDF_FINGERPRINT_CHARS]


def _fingerprint_pdf_pypdf(path: Path) -> str:
    """First ~2 pages via pypdf (the always-available fallback extractor)."""
    reader = PdfReader(str(path))
    chunks: list[str] = []
    for i in range(min(_PDF_FINGERPRINT_PAGES, len(reader.pages))):
        chunks.append(reader.pages[i].extract_text() or "")
    return "\n".join(chunks)[:_PDF_FINGERPRINT_CHARS]


def fingerprint_pdf(path: Path) -> str:
    """Return the first ~2 pages of text from a PDF, capped at 6000 chars.

    Prefers PyMuPDF (fast + robust on heavy decks); falls back to pypdf when
    PyMuPDF is unavailable or errors, so behavior degrades rather than breaking.
    Both honor the same 2-page / 6000-char cap, so the downstream classifier sees
    an equivalent fingerprint regardless of which extractor produced it.
    """
    via_fitz = _fingerprint_pdf_pymupdf(path)
    if via_fitz is not None:
        return via_fitz
    return _fingerprint_pdf_pypdf(path)


def fingerprint_xlsx(path: Path) -> str:
    """Stringify the first sheet's first ~12 rows for substring matching.

    `wb.close()` is required on Windows: read_only mode keeps the underlying
    zipfile handle open until explicitly closed, which blocks any subsequent
    move/unlink of the source file.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        parts: list[str] = [f"sheets={wb.sheetnames}"]
        for sn in wb.sheetnames[:2]:
            ws = wb[sn]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 12:
                    break
                parts.append(" | ".join("" if v is None else str(v)[:80] for v in row))
        return "\n".join(parts)[:_PDF_FINGERPRINT_CHARS]
    finally:
        wb.close()


def fingerprint(path: Path) -> str:
    sfx = path.suffix.lower()
    if sfx == ".pdf":
        return fingerprint_pdf(path)
    if sfx == ".xlsx":
        return fingerprint_xlsx(path)
    raise ValueError(f"Unsupported extension {sfx!r} for {path.name}")


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Classification primitives
# ---------------------------------------------------------------------------


def _detect_ticker(text: str) -> tuple[str | None, list[str]]:
    """Substring-match against the issuer registry. Returns (ticker, evidence)."""
    for ticker, _cal, names in _effective_registry():
        for name in names:
            if name in text:
                return ticker, [f"issuer_name:{name!r}"]
    return None, []


_RX_WIX_CDN_PREFIX = re.compile(r"^(?:2e8ef1|4f4a31)_[0-9a-f]{32}\.pdf$", re.IGNORECASE)
_RX_NU_RESULTS_DECK = re.compile(r"^[1-4]Q\d{2} Results Presentation\.pdf$", re.IGNORECASE)
_RX_NU_TRANSCRIPT_FILE = re.compile(r"^Transcript [1-4]Q\d{2}\.pdf$", re.IGNORECASE)

# Filename-based ticker rules. First match wins. Patterns are anchored where
# possible (^ ... ) so a stray substring elsewhere in the name doesn't decide.
# Each rule is (regex, ticker, evidence-label). Compiled once at import.
_FILENAME_TICKER_RULES: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"^RBRK[-_]", re.IGNORECASE), "RBRK", "filename_prefix:RBRK"),
    (re.compile(r"^(?:ServiceNow|ER-Q)", re.IGNORECASE), "NOW", "filename_prefix:ServiceNow"),
    (re.compile(r"^Novo[-_]Nordisk", re.IGNORECASE), "NVO", "filename_prefix:novo-nordisk"),
    (_RX_WIX_CDN_PREFIX, "WIX", "filename_wix_cdn"),
    (_RX_NU_RESULTS_DECK, "NU", "filename_pattern:nu_default_naming"),
    (_RX_NU_TRANSCRIPT_FILE, "NU", "filename_pattern:nu_default_naming"),
    # Airbnb: own filenames + FactSet "CORRECTED-TRANSCRIPT_-Airbnb-Inc-ABNB-US-..."
    (
        re.compile(r"^(?:Airbnb[-_]|CORRECTED-TRANSCRIPT_-Airbnb-Inc-ABNB)", re.IGNORECASE),
        "ABNB",
        "filename_prefix:Airbnb",
    ),
    # Booking Holdings: BKNG-*, Q1-2026-BKNG-*, Q3-25-BKNG-*, Booking-Holdings-Inc-BKNG-*,
    # CORRECTED-TRANSCRIPT_-Booking-Holdings-Inc-BKNG-*
    (
        re.compile(
            r"^(?:BKNG[-_]|Q[1-4]-(?:\d{2}|20\d{2})-BKNG[-_]|Booking-Holdings-Inc-BKNG|CORRECTED-TRANSCRIPT_-Booking-Holdings-Inc-BKNG)",
            re.IGNORECASE,
        ),
        "BKNG",
        "filename_prefix:BKNG",
    ),
    # SoFi: SOFI-*, SoFi-*, SoFi_*
    (re.compile(r"^SoFi[-_]|^SOFI[-_]", re.IGNORECASE), "SOFI", "filename_prefix:SoFi"),
    # dLocal: dlocal_*, dLocal_*, https___investor.dlocal.com_*, plus the user's
    # specific bare-quarter naming (4q24_earnings_results_vf.pdf,
    # transcript_webcast_presentation_q4_2024.pdf — both verified dLocal-issued).
    (
        re.compile(
            r"^(?:dlocal[-_]|https___investor\.dlocal\.com_|[1-4]q\d{2}_earnings_|transcript_webcast_presentation_q[1-4]_20\d{2}\.pdf$)",
            re.IGNORECASE,
        ),
        "DLO",
        "filename_prefix:dlocal",
    ),
    # Veeva: VEEV-*, Veeva-*, 2025-Veeva-Investor-Day, Veeva-JPM-*, etc.
    (
        re.compile(r"^(?:VEEV[-_]|Veeva[-_]|\d{4}[-_]Veeva[-_])", re.IGNORECASE),
        "VEEV",
        "filename_prefix:Veeva",
    ),
    # Alphabet: Alphabet-*, 2025q1-alphabet-*
    (
        re.compile(r"^(?:Alphabet[-_]|\d{4}q[1-4]-alphabet[-_])", re.IGNORECASE),
        "GOOG",
        "filename_prefix:Alphabet",
    ),
]


def _detect_ticker_from_filename(name: str) -> tuple[str | None, list[str]]:
    """Filename prefix/pattern → ticker. Hex/UUID names without a known prefix
    fall through to content detection."""
    for rx, ticker, label in _FILENAME_TICKER_RULES:
        if rx.search(name):
            return ticker, [f"{label}:{name[:32]!r}"]
    return None, []


_RX_FILENAME_ANNUAL_REPORT = re.compile(r"annual[-_\s]?report", re.IGNORECASE)

# Filename-based doc-type fallbacks. Used when content rules return None
# (e.g. image-only PDFs, decks where pypdf extracts no useful text). First
# match wins. Order: most specific first, broad keywords last.
_FILENAME_DOC_TYPE_RULES: list[tuple[DocType, re.Pattern[str], str]] = [
    (
        DocType.IR_TRANSCRIPT,
        re.compile(
            r"earnings[-_](?:call|conference[-_]call)[-_]transcript|earnings[-_]transcript|^transcript[-_]|transcript[-_]webcast",
            re.IGNORECASE,
        ),
        "filename_transcript",
    ),
    (
        DocType.IR_INVESTOR_UPDATE,
        re.compile(r"shareholder[-_]letter", re.IGNORECASE),
        "filename_shareholder_letter",
    ),
    (
        DocType.IR_EVENT,
        re.compile(r"investor[-_]day", re.IGNORECASE),
        "filename_investor_day",
    ),
    (
        DocType.IR_EVENT,
        re.compile(
            r"(?:healthcare[-_]conference|jpm[-_].*?conference|capital[-_]markets[-_]day|investor[-_]conference)",
            re.IGNORECASE,
        ),
        "filename_conference",
    ),
    (
        DocType.IR_PRESENTATION,
        re.compile(
            r"(?:quarterly[-_]earnings[-_]presentation|earnings[-_]presentation|investor[-_]presentation|earnings[-_]slides?)",
            re.IGNORECASE,
        ),
        "filename_presentation",
    ),
    (
        DocType.IR_PRESS_RELEASE,
        re.compile(
            r"(?:earnings[-_]press[-_]release|earnings[-_]release|earnings[-_]results)",
            re.IGNORECASE,
        ),
        "filename_press_release",
    ),
    (
        DocType.IR_SUPPLEMENT,
        re.compile(
            r"(?:financial[-_]workbook|financial[-_]supplement|fact[-_]sheet)",
            re.IGNORECASE,
        ),
        "filename_supplement",
    ),
]


def _detect_doc_type(
    text: str,
    ext: str,
    filename: str = "",
) -> tuple[DocType | None, list[str]]:
    """Apply doc-type rules. xlsx files always classify as IR_SUPPLEMENT.

    Strategy: scan **all** rules and pick the one whose pattern appears
    earliest in the text. Rule order in `_DOC_TYPE_RULES` is the tiebreaker
    when two rules match at the same position. This handles MELI press
    releases that include the "To our Shareholders" letter as a later section
    of the same PDF — the press-release dateline appears first, so press
    release wins; a standalone shareholder letter has "To our Shareholders"
    at position 0, so it wins.

    pypdf often emits glyphs without spaces ("EarningsCall", "PreparedRemarks").
    The squashed rule set runs only when no full-text rule matched at all.

    The filename is consulted as a final fallback so unambiguous cases like
    `novo-nordisk-annual-report-2025.pdf` don't depend on regex hits in the
    body of the PDF (where "Annual Report" may not appear verbatim).
    """
    if ext == ".xlsx":
        return DocType.IR_SUPPLEMENT, ["xlsx_extension"]

    earliest_pos: int | None = None
    earliest_idx: int | None = None
    earliest_match: tuple[DocType, str, str] | None = None
    for idx, (doc_type, rx, label) in enumerate(_DOC_TYPE_RULES):
        m = rx.search(text)
        if m is None:
            continue
        pos = m.start()
        if (
            earliest_pos is None
            or pos < earliest_pos
            or (pos == earliest_pos and idx < earliest_idx)  # type: ignore[operator]
        ):
            earliest_pos = pos
            earliest_idx = idx
            earliest_match = (doc_type, label, m.group(0))
    if earliest_match is not None:
        doc_type, label, hit = earliest_match
        return doc_type, [f"{label}:{hit!r}@{earliest_pos}"]

    flat = re.sub(r"\s+", "", text).lower()
    for doc_type, rx, label in _DOC_TYPE_SQUASHED_RULES:
        m = rx.search(flat)
        if m:
            return doc_type, [f"{label}_squashed:{m.group(0)!r}"]
    if filename and _RX_FILENAME_ANNUAL_REPORT.search(filename):
        return DocType.IR_INVESTOR_UPDATE, ["filename_annual_report"]
    # Filename-pattern fallbacks fire only when no content rule matched.
    # Critical for image-only PDFs (e.g. Wix CDN cover decks) and decks where
    # pypdf returns near-empty text.
    if filename:
        for doc_type, rx, label in _FILENAME_DOC_TYPE_RULES:
            m = rx.search(filename)
            if m:
                return doc_type, [f"{label}:{m.group(0)!r}"]
    return None, []


def _fiscal_year_quarter_for_date(cal: str, period_end: date) -> tuple[int, int] | None:
    """Reverse of `_period_end_for`: the issuer's (fiscal_year, quarter) whose
    quarter-end equals `period_end` under `cal`.

    Returns None when `period_end` is not a recognized quarter-end for that
    calendar (e.g. a 52/53-week filer's off-month-end date), so the caller can
    fall back to a coarser heuristic.
    """
    q = _quarter_for_period_end(cal, period_end)
    if q is None:
        return None
    for fy in (period_end.year - 1, period_end.year, period_end.year + 1):
        try:
            if _period_end_for(cal, fy, q) == period_end:
                return fy, q
        except ValueError:
            continue
    return None


def _sec_cover_year_quarter(cal: str, m: re.Match[str]) -> tuple[int, int]:
    """(fiscal_year, quarter) for a SEC cover's EXPLICIT period-end date.

    The matched "...ended <Month DD, YYYY>" date IS the period end. Resolve it to
    the issuer's fiscal (year, quarter) under `cal` so a non-December FYE issuer
    (RBRK/VEEV Jan-31, V Sep-30, ORCL May-31) is not mis-shifted — the bug that
    arose from reading the month as a CALENDAR quarter and the literal year as a
    FISCAL year and re-mapping through `_period_end_for()`. Falls back to the
    calendar-quarter-of-month for an off-cycle date no calendar maps to a
    quarter-end (preserves prior behavior for that edge case).
    """
    month_num = _MONTH_TO_NUM[m.group("m").lower()]
    year = int(m.group("y"))
    try:
        explicit = date(year, month_num, int(m.group("d")))
    except ValueError:
        return year, _MONTH_TO_Q[m.group("m").lower()]
    yq = _fiscal_year_quarter_for_date(cal, explicit)
    if yq is not None:
        return yq
    return year, _MONTH_TO_Q[m.group("m").lower()]


def _detect_period(
    text: str,
    ticker_calendar: str,
) -> tuple[tuple[int, int] | None, list[str]]:
    """Find (year, quarter) in text. NVO H1/9M/FY mapped per the directive.

    Returns the *fiscal-year-and-quarter* pair the issuer would label this
    period with — `_period_end_for()` then converts to a calendar period_end.
    """
    # SEC 10-Q / 10-K cover pages always carry an explicit period-ended line.
    # Pre-empt the looser quarter regexes since 10-Qs frequently mention "first
    # quarter" prose later in the body; the cover phrase is the source of truth.
    m = _RX_SEC_QUARTERLY_PERIOD_ENDED.search(text)
    if m:
        return _sec_cover_year_quarter(ticker_calendar, m), [
            f"sec_quarterly_period_ended:{m.group(0)!r}"
        ]
    m = _RX_SEC_FISCAL_YEAR_ENDED.search(text)
    if m:
        return _sec_cover_year_quarter(ticker_calendar, m), [
            f"sec_fiscal_year_ended:{m.group(0)!r}"
        ]

    if ticker_calendar == _CAL_NVO:
        m = _RX_NVO_PERIOD.search(text)
        if m:
            phrase = m.group("period").lower().replace("\xa0", " ")
            year = int(m.group("y"))
            if "three months" in phrase:
                return (year, 1), [f"nvo_period:'three months {year}'"]
            if "six months" in phrase:
                return (year, 2), [f"nvo_period:'six months {year}'"]
            if "nine months" in phrase:
                return (year, 3), [f"nvo_period:'nine months {year}'"]
            if "full year" in phrase:
                return (year, 4), [f"nvo_period:'full year {year}'"]

    m = _RX_QUARTER_FULL_YEAR.search(text)
    if m:
        q = _WORD_TO_Q[m.group("word").lower()]
        y = int(m.group("y"))
        return (y, q), [f"quarter_full_year:{m.group(0)!r}"]

    m = _RX_QUARTER_ENDED_DATE.search(text)
    if m:
        q = _MONTH_TO_Q[m.group("m").lower()]
        y = int(m.group("y"))
        return (y, q), [f"quarter_ended_date:{m.group(0)!r}"]

    m = _RX_DATE_QUARTER_ENDED.search(text)
    if m:
        q = _MONTH_TO_Q[m.group("m").lower()]
        y = int(m.group("y"))
        return (y, q), [f"date_quarter_ended:{m.group(0)!r}"]

    m = _RX_FY_QY.search(text)
    if m:
        q = int(m.group("q"))
        y = _yy_to_yyyy(int(m.group("y")))
        return (y, q), [f"fy_qy:{m.group(0)!r}"]

    m = _RX_QUARTER_FY.search(text)
    if m:
        q = _WORD_TO_Q[m.group("word").lower()]
        y = _yy_to_yyyy(int(m.group("y")))
        return (y, q), [f"fiscal_word:{m.group(0)!r}"]

    m = _RX_QUARTER_FULL.search(text)
    if m:
        q = _WORD_TO_Q[m.group("word").lower()]
        y = _yy_to_yyyy(int(m.group("y")))
        return (y, q), [f"quarter_word:{m.group(0)!r}"]

    m = _RX_QUARTER_APOSTROPHE.search(text)
    if m:
        q = int(m.group("q"))
        y = _yy_to_yyyy(int(m.group("yy")))
        return (y, q), [f"quarter_apostrophe:{m.group(0)!r}"]

    m = _RX_QUARTER_SPACE.search(text)
    if m:
        q = int(m.group("q"))
        y = int(m.group("y"))
        return (y, q), [f"quarter_space:{m.group(0)!r}"]

    flat = re.sub(r"\s+", "", text).lower()
    m = re.search(r"q(?P<q>[1-4])(?P<y>20\d{2})", flat)
    if m:
        return (int(m.group("y")), int(m.group("q"))), [f"quarter_squashed:{m.group(0)!r}"]
    m = re.search(r"(?P<word>first|second|third|fourth)quarter(?:fiscal)?(?P<y>20\d{2})", flat)
    if m:
        q = _WORD_TO_Q[m.group("word")]
        return (int(m.group("y")), q), [f"quarter_word_squashed:{m.group(0)!r}"]

    return None, []


_MONTH_TO_NUM: dict[str, int] = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def _detect_event_period_end(text: str, filename: str) -> tuple[date | None, list[str]]:
    """For IR_EVENT decks, period_end is the event date (per DocType.IR_EVENT).

    Strategy: take the first explicit "Month DD, YYYY" date in the cover-page
    text. Investor day / conference decks lead with this on slide 1. If no
    date is found, fall back to a filename year hint and use Dec 31 as a
    placeholder so the file still gets categorized — the user can rename for
    a more specific date.
    """
    m = _RX_EVENT_DATE.search(text)
    if m:
        month_num = _MONTH_TO_NUM[m.group("m").lower()]
        d = int(m.group("d"))
        y = int(m.group("y"))
        try:
            return date(y, month_num, d), [f"event_date:{m.group(0)!r}"]
        except ValueError:
            pass
    # Filename year-only fallback: "2025-Veeva-Investor-Day.pdf" → 2025-12-31.
    rx_year = re.compile(r"\b(?P<y>20\d{2})\b")
    m2 = rx_year.search(filename)
    if m2:
        return date(int(m2.group("y")), 12, 31), [f"event_year_filename:{m2.group(0)!r}"]
    return None, []


def _filename_period_hint(name: str) -> tuple[tuple[int, int] | None, list[str]]:
    """Best-effort period hint from filename — used to disambiguate when
    content gives a less-specific match (e.g. cover slide doesn't say which Q).
    """
    # SoFi-style "Q4-FY-2025" naming: must be tried before the looser rx2/rx3
    # since they'd otherwise capture only the trailing "20\d{2}" segment.
    rx_fy = re.compile(r"\bq(?P<q>[1-4])[-_]fy[-_](?P<y>20\d{2})\b", re.IGNORECASE)
    m_fy = rx_fy.search(name)
    if m_fy:
        return (int(m_fy.group("y")), int(m_fy.group("q"))), [f"filename_period:{m_fy.group(0)!r}"]

    rx = re.compile(r"(?P<a>[1-4])Q(?P<b>\d{2})", re.IGNORECASE)
    m = rx.search(name)
    if m:
        q = int(m.group("a"))
        y = _yy_to_yyyy(int(m.group("b")))
        return (y, q), [f"filename_period:{m.group(0)!r}"]
    rx2 = re.compile(r"\bq(?P<q>[1-4])[-_+\s]+(?P<y>20\d{2})\b", re.IGNORECASE)
    m2 = rx2.search(name)
    if m2:
        return (int(m2.group("y")), int(m2.group("q"))), [f"filename_period:{m2.group(0)!r}"]
    rx3 = re.compile(r"\bq(?P<q>[1-4])[-_+\s]+(?P<yy>\d{2})\b", re.IGNORECASE)
    m3 = rx3.search(name)
    if m3:
        y = _yy_to_yyyy(int(m3.group("yy")))
        return (y, int(m3.group("q"))), [f"filename_period:{m3.group(0)!r}"]
    # YYYY-Q[1-4] / YYYYqN style (e.g. "2025-Q1", "2025q1"). Order: after rx
    # (which catches "1Q25") so the year-first form doesn't shadow it.
    rx_year_q = re.compile(r"\b(?P<y>20\d{2})[-_]?q(?P<q>[1-4])\b", re.IGNORECASE)
    m_year_q = rx_year_q.search(name)
    if m_year_q:
        return (int(m_year_q.group("y")), int(m_year_q.group("q"))), [
            f"filename_period:{m_year_q.group(0)!r}"
        ]
    rx4 = re.compile(r"annual[-_\s]report[-_\s](?P<y>20\d{2})", re.IGNORECASE)
    m4 = rx4.search(name)
    if m4:
        return (int(m4.group("y")), 4), [f"filename_annual:{m4.group(0)!r}"]
    return None, []


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def classify_ir_file(
    path: Path,
    *,
    ticker_hint: str | None = None,
    calendar_override: str | None = None,
) -> CategorizationResult | CategorizationFailure:
    """Classify a single IR-uploads file. Pure function — only reads the file.

    Returns a `CategorizationResult` when ticker, doc_type, and period are all
    determined. Otherwise a `CategorizationFailure` carrying the partial
    evidence so the user can repair (rename, delete, or extend the registry).

    `ticker_hint` is consulted only when the filename and content fingerprint
    fail to identify a ticker. Callers pass the parent-folder ticker when a
    file is dropped under `ir_documents/<TICKER>/...` — the path is then strong
    evidence for the issuer.

    `calendar_override` (the headless auto-fetch path): when set, the
    `ticker_hint` is trusted as authoritative — the file was downloaded from that
    issuer's own IR site into its folder — so the content/filename conflict gating
    is skipped, and this fiscal-calendar id attributes the period when the ticker
    has no `ISSUER_REGISTRY` calendar (a registered ticker still uses its registry
    calendar). The manual-upload path leaves it None and keeps strict behavior.
    """
    ext = path.suffix.lower()
    if ext not in {".pdf", ".xlsx"}:
        return CategorizationFailure(
            reason=f"unsupported_extension:{ext!r}",
            text_sample="",
        )

    try:
        raw_text = fingerprint(path)
    except Exception as e:  # pypdf / openpyxl can throw a wide tree
        return CategorizationFailure(
            reason=f"fingerprint_error:{type(e).__name__}:{e}",
            text_sample="",
        )

    # Many issuer PDFs land with multi-space inter-word formatting after pypdf
    # extraction (e.g. "Nu  Holdings"). Normalize whitespace once so substring
    # rules don't have to multiply-match every variant.
    text = re.sub(r"[ \t]+", " ", raw_text)

    file_ticker, file_ev = _detect_ticker_from_filename(path.name)
    content_ticker, content_ev = _detect_ticker(text)
    ticker_evidence = file_ev + content_ev

    if calendar_override is not None and ticker_hint is not None:
        # Auto-fetch: the file was downloaded from this ticker's own IR site into
        # its folder, so the path hint is authoritative — skip the filename/content
        # conflict gating that the strict manual-upload path enforces.
        ticker = ticker_hint
        ticker_evidence = [*ticker_evidence, "trusted_path_hint"]
    else:
        if file_ticker and content_ticker and file_ticker != content_ticker:
            return CategorizationFailure(
                reason=f"ticker_conflict:filename={file_ticker} content={content_ticker}",
                text_sample=text[:600],
                ticker_guess=file_ticker,
            )
        ticker = file_ticker or content_ticker
        if ticker is None and ticker_hint is not None:
            if any(t == ticker_hint for t, *_ in _effective_registry()):
                ticker = ticker_hint
                ticker_evidence = ticker_evidence + ["path_hint"]
        if ticker is None:
            return CategorizationFailure(
                reason="ticker_unidentified",
                text_sample=text[:600],
            )

    doc_type, doc_ev = _detect_doc_type(text, ext, path.name)
    if doc_type is None:
        return CategorizationFailure(
            reason="doc_type_unidentified",
            text_sample=text[:600],
            ticker_guess=ticker,
        )

    # Registered ticker → its registry calendar (most authoritative). Otherwise
    # fall back to the auto-fetch calendar_override; with neither, fail cleanly.
    try:
        cal = _calendar_for(ticker)
    except ValueError:
        if calendar_override is None:
            return CategorizationFailure(
                reason=f"no_calendar_for:{ticker}",
                text_sample=text[:600],
                ticker_guess=ticker,
                doc_type_guess=doc_type,
            )
        cal = calendar_override

    # IR_EVENT (investor day, broker conference, capital markets day) is keyed
    # by event date, not fiscal quarter. Pull the first explicit date out of
    # the cover-page text; fall back to a year-only filename hint.
    if doc_type == DocType.IR_EVENT:
        event_period_end, event_evidence = _detect_event_period_end(text, path.name)
        if event_period_end is None:
            return CategorizationFailure(
                reason="event_date_unidentified",
                text_sample=text[:600],
                ticker_guess=ticker,
                doc_type_guess=doc_type,
            )
        confidence = Confidence.HIGH if file_ev else Confidence.MEDIUM
        return CategorizationResult(
            ticker=ticker,
            doc_type=doc_type,
            period_end=event_period_end,
            period_label=event_period_end.isoformat(),
            confidence=confidence,
            ticker_evidence=ticker_evidence,
            doc_type_evidence=doc_ev,
            period_evidence=event_evidence,
        )

    fname_yq, fname_period_ev = _filename_period_hint(path.name)
    content_yq, content_period_ev = _detect_period(text, cal)

    # Filename takes precedence (it disambiguates near-identical cover slides).
    yq = fname_yq or content_yq
    period_evidence = fname_period_ev + content_period_ev
    if yq is None:
        return CategorizationFailure(
            reason="period_unidentified",
            text_sample=text[:600],
            ticker_guess=ticker,
            doc_type_guess=doc_type,
        )

    year, q = yq
    period_end = _period_end_for(cal, year, q)
    period_label = _period_label(cal, year, q)

    confidence = (
        Confidence.HIGH
        if (file_ev and (fname_yq or content_yq)) or (content_ev and content_yq)
        else Confidence.MEDIUM
    )

    return CategorizationResult(
        ticker=ticker,
        doc_type=doc_type,
        period_end=period_end,
        period_label=period_label,
        confidence=confidence,
        ticker_evidence=ticker_evidence,
        doc_type_evidence=doc_ev,
        period_evidence=period_evidence,
    )


def _period_label(cal: str, year: int, q: int) -> str:
    """Render a human-readable label, e.g. 'Q3 2024' or 'FY26 Q1'."""
    if cal in (_CAL_VEEV, _CAL_APR30, _CAL_VISA, _CAL_ORACLE):
        return f"FY{year % 100:02d} Q{q}"
    if cal == _CAL_NVO:
        return ["Q1", "H1", "9M", "FY"][q - 1] + f" {year}"
    return f"Q{q} {year}"


def canonical_path(
    root: Path,
    result: CategorizationResult,
    sha256: str,
    suffix: str,
) -> Path:
    """ir_documents/{TICKER}/{period_end_iso}/{doc_type}__{sha8}.{ext}.

    The sha8 prefix means rerunning with a modified file produces a new path,
    not an overwrite — preserving the supersession invariant from §2 of the
    data-provenance directive.
    """
    short = sha256[:8]
    return (
        root
        / result.ticker
        / result.period_end.isoformat()
        / f"{result.doc_type.value}__{short}{suffix.lower()}"
    )


_CANONICAL_PATH_RX = re.compile(
    r"^(?P<ticker>[A-Z][A-Z0-9.]*)/"
    r"(?P<period>\d{4}-\d{2}-\d{2})/"
    r"(?P<doc_type>ir_[a-z_]+)__(?P<sha8>[0-9a-f]{8})"
    r"\.(?P<ext>pdf|xlsx)$"
)


def parse_canonical_path(path: Path, root: Path) -> CategorizationResult | None:
    """Extract (ticker, doc_type, period_end) from a path already at canonical position.

    Returns None if the path doesn't match the `<TICKER>/<period_end>/<doc_type>__<sha8>.<ext>`
    layout produced by `canonical_path()`. Used by the reindexer fast-path so we
    can register pre-placed files without re-classifying them.
    """
    try:
        rel = path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return None
    m = _CANONICAL_PATH_RX.match(rel)
    if m is None:
        return None
    try:
        doc_type = DocType(m["doc_type"])
    except ValueError:
        return None
    period_end = date.fromisoformat(m["period"])
    try:
        cal = _calendar_for(m["ticker"])
    except ValueError:
        cal = _CAL_CALENDAR
    quarter = _quarter_for_period_end(cal, period_end)
    period_label = (
        _period_label(cal, period_end.year, quarter) if quarter else period_end.isoformat()
    )
    return CategorizationResult(
        ticker=m["ticker"],
        doc_type=doc_type,
        period_end=period_end,
        period_label=period_label,
        confidence=Confidence.HIGH,
        ticker_evidence=["canonical_path"],
        doc_type_evidence=["canonical_path"],
        period_evidence=["canonical_path"],
    )


def _quarter_for_period_end(cal: str, period_end: date) -> int | None:
    """Reverse of `_period_end_for`: derive Q-number from period_end given calendar."""
    if cal == _CAL_VEEV:
        return {4: 1, 7: 2, 10: 3, 1: 4}.get(period_end.month)
    if cal == _CAL_APR30:
        return {7: 1, 10: 2, 1: 3, 4: 4}.get(period_end.month)
    if cal == _CAL_VISA:
        return {12: 1, 3: 2, 6: 3, 9: 4}.get(period_end.month)
    if cal == _CAL_ORACLE:
        return {8: 1, 11: 2, 2: 3, 5: 4}.get(period_end.month)
    return {3: 1, 6: 2, 9: 3, 12: 4}.get(period_end.month)


def iter_uncategorized_files(root: Path) -> list[Path]:
    """Files anywhere under `ir_documents/` that may need processing.

    - Root-level files: arrived via manual drop, need full classification.
    - Files under `<TICKER>/...`: already in a ticker folder, may be at the
      canonical `<TICKER>/<period_end>/<doc_type>__<sha8>.<ext>` position
      (fast-path: parse-and-register), or somewhere else under the ticker dir
      (use the path-based ticker as a hint to the classifier).

    System folders prefixed with `_` (e.g. `_unsorted/`, `_events/`) and any
    hidden dotfiles are skipped.
    """
    out: list[Path] = []
    for p in sorted(root.rglob("*")):
        if not p.is_file() or p.name.startswith("."):
            continue
        if p.suffix.lower() not in {".pdf", ".xlsx"}:
            continue
        try:
            rel_parts = p.resolve().relative_to(root.resolve()).parts
        except ValueError:
            continue
        if rel_parts and (rel_parts[0].startswith("_") or rel_parts[0].startswith(".")):
            continue
        out.append(p)
    return out


def ticker_hint_from_path(path: Path, root: Path) -> str | None:
    """If `path` is under `<root>/<TICKER>/...`, return TICKER. Else None.

    Used by the categorizer when filename heuristics fail: if the user dropped
    a file inside a ticker subfolder, the path itself is strong evidence of
    which issuer it belongs to.
    """
    try:
        rel_parts = path.resolve().relative_to(root.resolve()).parts
    except ValueError:
        return None
    if len(rel_parts) < 2:
        return None
    head = rel_parts[0]
    if not re.match(r"^[A-Z][A-Z0-9.]*$", head):
        return None
    return head


__all__: Sequence[str] = (
    "ISSUER_REGISTRY",
    "CategorizationFailure",
    "CategorizationResult",
    "Confidence",
    "calendar_id_from_fye",
    "canonical_path",
    "classify_ir_file",
    "parse_canonical_path",
    "ticker_hint_from_path",
    "fingerprint",
    "fingerprint_pdf",
    "fingerprint_xlsx",
    "iter_uncategorized_files",
    "sha256_of",
)
