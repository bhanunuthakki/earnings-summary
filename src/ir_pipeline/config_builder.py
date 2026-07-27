"""Generate a ticker's IR-spreadsheet parser config — every row, not just the watchlist.

An issuer's "historical data" spreadsheet is its own audited numeric source, so
the capture-every-number program (`directives/capture_every_number_program.md`)
wants EVERY labeled data row in it — not only the handful of curated tier KPIs.
A config therefore carries two layers of ``SheetKpi`` (see ``config.SheetKpi``):

* **analyst** — the ticker's holdings ``tier_1_kpis`` / ``chart_priorities``,
  mapped to their canonical ``kpi_definitions.name`` by an LLM (it's a judgment
  task: which row is "Monthly ARPAC", and is a percent stored as a decimal). This
  is the original, high-value, hand-quality layer; it is unchanged.
* **capture** — every OTHER labeled numeric row, mapped DETERMINISTICALLY: the row
  set comes from ``spreadsheet.enumerate_numeric_rows`` (no LLM gate drops rows),
  unit/scale from a heuristic classifier, and the label is canonicalized to a
  ``kpi_definitions.name`` at INGEST time via ``canonical_metric_name``. Subtotals
  are kept (a subtotal is a reported number); footnotes / section captions /
  standard FMP line items are filtered so they don't become pickable junk.

``build_ir_config`` emits both layers for a brand-new ticker; ``widen_config``
adds the capture layer to an existing (possibly hand-built) config while keeping
its analyst rows byte-identical — the path that lifts the NU-only ceiling without
disturbing NU's curated series.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Literal, cast

import openpyxl
from pydantic import BaseModel, ConfigDict, Field, RootModel, TypeAdapter

from ir_pipeline.config import IrConfig, SheetKpi, save_config
from ir_pipeline.spreadsheet import SheetDataRow, _header_row, enumerate_numeric_rows

_PROJECT_ROOT = Path(__file__).resolve().parents[2]


class _SheetKpiMapping(BaseModel):
    model_config = ConfigDict(extra="forbid")

    sheet: str = Field(min_length=1)
    row_label: str = Field(min_length=1)
    unit: Literal["percent", "actual", "usd", "count", "ratio"]
    scale: Literal[1, 100]


class _SheetKpiMapResponse(RootModel[dict[str, _SheetKpiMapping]]):
    pass


# Standard financials line items — these flow from FMP/SEC at higher tiers, so
# they must NOT be sourced from the IR spreadsheet (that would create stray,
# wrong-unit duplicate KPI rows). Matches financials._LINE_ITEM_SPECS.
_FMP_LINE_ITEMS = frozenset(
    {
        "revenue",
        "gross profit",
        "operating income",
        "net income",
        "eps (diluted)",
        "operating cash flow",
        "free cash flow",
        "capex",
    }
)


def _target_kpi_names(ticker: str, repo_root: Path) -> list[str]:
    """Canonical KPI names to map to — the union of the ticker's holdings
    `tier_1_kpis` and `chart_priorities`, minus standard FMP line items.

    `chart_priorities` carries the granular series the report charts (e.g. NU's
    "NPL 15-90d" / "NPL 90d+"), which the spreadsheet stores as distinct rows;
    `tier_1_kpis` sometimes only has a combined form ("NPL 15d+ total") with no
    single row. Standard financials line items (Revenue, Net income, …) are
    excluded — they come from FMP, not IR spreadsheets. The LLM omits any
    remaining target the spreadsheet doesn't carry.
    """
    path = repo_root / "micro_thesis" / "holdings" / f"{ticker.upper()}.json"
    if not path.exists():
        return []
    holdings = json.loads(path.read_text(encoding="utf-8"))
    out: list[str] = []
    seen: set[str] = set()

    def _add(name: str) -> None:
        name = name.strip()
        if name and name not in seen and name.lower() not in _FMP_LINE_ITEMS:
            seen.add(name)
            out.append(name)

    for k in holdings.get("tier_1_kpis") or []:
        if isinstance(k, dict):
            _add(str(k.get("name", "")))
    for name in holdings.get("chart_priorities") or []:
        if isinstance(name, str):
            _add(name)
    return out


def dump_sheet_structure(xlsx_path: Path, label_col: int = 1, max_rows_per_sheet: int = 60) -> str:
    """Compact text view of each data sheet: labels + two recent sample values.

    Only sheets with a detectable quarter-date header are included; the samples
    let the model infer unit/scale (a 0.095 cell is a decimal percent → scale 100).
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    blocks: list[str] = []
    for sheet in wb.sheetnames:
        rows = cast(
            "list[tuple[object, ...]]",
            [tuple(r) for r in wb[sheet].iter_rows(values_only=True)],
        )
        _, col_map = _header_row(rows)
        if not col_map:
            continue
        recent = sorted(col_map)[-2:]
        recent_labels = [col_map[c].date().isoformat() for c in recent]
        lines = [f"## Sheet: {sheet!r}  (quarters across columns; latest = {recent_labels})"]
        for r in rows:
            lab = r[label_col] if label_col < len(r) else None
            if not (isinstance(lab, str) and lab.strip()):
                continue
            samples = [r[c] for c in recent if c < len(r)]
            if any(isinstance(s, (int, float)) and not isinstance(s, bool) for s in samples):
                vals = [round(s, 4) if isinstance(s, float) else s for s in samples]
                lines.append(f"  - {lab.strip()[:80]}  e.g. {vals}")
            if len(lines) > max_rows_per_sheet:
                lines.append("  - … (truncated)")
                break
        if len(lines) > 1:
            blocks.append("\n".join(lines))
    wb.close()
    return "\n\n".join(blocks)


def _build_prompt(ticker: str, structure: str, kpi_names: list[str]) -> str:
    kpi_block = "\n".join(f"- {n}" for n in kpi_names)
    return f"""You are configuring a parser for {ticker}'s investor-relations "historical data" \
spreadsheet. Below is its structure: each data sheet, with its rows (label + two \
recent sample values; quarters run across columns).

For EACH canonical KPI below, find the single spreadsheet row holding its quarterly \
time series:
{kpi_block}

Return ONLY a JSON object keyed by the EXACT canonical KPI name:
  {{"<kpi name>": {{"sheet": "<exact sheet title>", "row_label": "<unique substring of the row label>", "unit": "percent|actual|usd|count|ratio", "scale": 1 or 100}}}}

Rules:
- row_label: a short, case-insensitive SUBSTRING that uniquely identifies the row.
- scale: 100 ONLY if the row stores a percent as a decimal (sample like 0.095 = 9.5%); otherwise 1.
- unit: "percent" for margins / ratios / NPLs, "usd" or "actual" for dollar figures, "count" for customer counts.
- OMIT any KPI that is not present in the spreadsheet (e.g. ROE or CET1 are often not tabulated). Do NOT guess or invent a row.

Spreadsheet structure:
\"\"\"
{structure}
\"\"\"

Return ONLY the JSON object — no prose, no markdown fence."""


def _llm_map(ticker: str, structure: str, kpi_names: list[str]) -> dict[str, dict[str, object]]:
    """Single LLM call → {kpi_name: {sheet, row_label, unit, scale}}."""
    from llm.structured import call_llm_structured  # lazy: heavy import chain

    # One-time per-ticker onboarding step; mapping a full sheet structure takes
    # the model ~1-3 min, so allow generous headroom over the default timeout.
    parsed = call_llm_structured(
        _build_prompt(ticker, structure, kpi_names),
        purpose="ir_sheet_kpi_map",
        timeout_seconds=300,
        expect="object",
        schema=TypeAdapter(_SheetKpiMapResponse),
    )
    validated = _SheetKpiMapResponse.model_validate(parsed)
    return {
        name: spec.model_dump(mode="python")
        for name, spec in validated.root.items()
        if name in kpi_names
    }


def _analyst_kpis(ticker: str, xlsx_path: Path, repo_root: Path, label_col: int) -> list[SheetKpi]:
    """The curated layer: holdings tier KPIs mapped to canonical names via LLM.

    Unchanged from the original builder — preserves the high-quality hand-mapping
    of watchlist KPIs (which row is "Monthly ARPAC", decimal-vs-percent). Returns
    ``[]`` when the ticker has no holdings tier KPIs (a brand-new long-tail-only
    ticker) — no LLM call is made in that case.
    """
    kpi_names = _target_kpi_names(ticker, repo_root)
    if not kpi_names:
        return []
    structure = dump_sheet_structure(xlsx_path, label_col=label_col)
    mapping = _llm_map(ticker, structure, kpi_names) if structure else {}

    kpis: list[SheetKpi] = []
    for name in kpi_names:  # preserve holdings order; only keep mapped KPIs
        spec = mapping.get(name)
        if not isinstance(spec, dict) or not spec.get("sheet") or not spec.get("row_label"):
            continue
        raw_scale = spec.get("scale", 1.0)
        scale = (
            float(cast("float", raw_scale))
            if isinstance(raw_scale, (int, float)) and not isinstance(raw_scale, bool)
            else 1.0
        )
        kpis.append(
            SheetKpi(
                kpi_name=name,
                sheet=str(spec["sheet"]),
                row_label=str(spec["row_label"]),
                unit=str(spec.get("unit", "actual")),
                scale=scale,
                origin="analyst",
            )
        )
    return kpis


# ---------------------------------------------------------------------------
# Capture layer — map EVERY labeled numeric row (deterministic, no LLM gate)
# ---------------------------------------------------------------------------

# A label that is a footnote / source caption / numbered note, not a metric. A
# subtotal ("Total Available Funding") is deliberately NOT matched — it is still a
# reported number worth capturing — only true scraps are excluded.
_FOOTNOTE_RX = re.compile(
    r"^\s*(?:\(?\d+\s*[\).]|[*\u2020\u2021\u00a7\u2022\u00b7\u2013\u2014-]"
    r"|note\b|notes\b|source\b|footnote\b|disclaimer\b)",
    re.IGNORECASE,
)
# A capture label longer than this is almost certainly prose / a footnote, not a
# metric name — drop it rather than mint a definition keyed on a sentence.
_MAX_CAPTURE_LABEL = 90
# Currency markers that pin a row to a dollar/real amount (→ "usd").
_CURRENCY_MARKERS = ("US$", "USD", "R$", "$", "€", "£", "¥")
# Phrases that explicitly signal a COUNT — deliberately specific so an ambiguous
# mention ("Loans to customers" is a dollar balance, not a count) defaults to a
# level. A bare "customers"/"accounts" is NOT enough.
_COUNT_PHRASES = ("# of", "no. of", "number of", "active customers", "active accounts", "headcount")
# Explicit RATIO signals — only an "x / y" division or the literal word "ratio"
# converts a sub-unit decimal to a ratio; otherwise a small number is just a small
# level ("Share of Results From Associates" of -0.36 is $M, not a fraction).
_RATIO_MARKERS = (" / ", "ratio")


def _is_capturable_label(label: str) -> bool:
    """True iff ``label`` is a real metric row worth capturing (vs scrap)."""
    s = label.strip()
    if not s or len(s) > _MAX_CAPTURE_LABEL:
        return False
    if s.lower() in _FMP_LINE_ITEMS:
        return False
    return not _FOOTNOTE_RX.match(s)


def _classify_unit_scale(label: str, samples: list[float]) -> tuple[str, float]:
    """Heuristic (unit, scale) for a capture row from its label + sample values.

    Conservative and explainable rather than clever — the canonicalizer's
    unit-family gate is the real guard against a bad merge, so an imperfect unit
    here is recoverable. ``scale=100`` only for a percent stored as a decimal
    (every recent sample |v| < 1.5 on a ``%`` row, e.g. 0.0948 → 9.48%).
    """
    low = label.lower()
    decimalish = bool(samples) and all(abs(s) < 1.5 for s in samples)
    if "%" in label:
        return "percent", (100.0 if decimalish else 1.0)
    if any(tok in label for tok in _CURRENCY_MARKERS):
        return "usd", 1.0
    if "bps" in low:
        return "bps", 1.0
    if any(p in low for p in _COUNT_PHRASES):
        return "count", 1.0
    # Only an explicit ratio signal turns a sub-unit decimal into a ratio; an
    # unmarked number — large or small — is a level ("actual"), the safe default
    # that never crosses unit families with a same-stem dollar metric.
    if decimalish and any(m in low for m in _RATIO_MARKERS):
        return "ratio", 1.0
    return "actual", 1.0


def _claims(analyst: list[SheetKpi]) -> list[tuple[str, str]]:
    """`(sheet, lower-cased row_label)` each analyst row already covers — capture
    skips any sheet row whose label contains one of these so a curated KPI's row
    is never re-mapped as a duplicate capture entry."""
    return [(k.sheet, k.row_label.strip().lower()) for k in analyst]


def _is_claimed(row: SheetDataRow, claims: list[tuple[str, str]]) -> bool:
    lab = row.label.lower()
    return any(sheet == row.sheet and substr in lab for sheet, substr in claims)


def _capture_kpis(rows: list[SheetDataRow], *, analyst: list[SheetKpi]) -> list[SheetKpi]:
    """Build a capture ``SheetKpi`` for every uncovered, capturable numeric row.

    Deduped by case-insensitive label across the whole workbook (the first sheet a
    label appears on wins — the same metric repeated on a second sheet would only
    re-ingest identical values), and matched ``exact_label`` so an auto-derived
    short label can't grab a superstring row.
    """
    claims = _claims(analyst)
    out: list[SheetKpi] = []
    seen: set[str] = set()
    for row in rows:
        if not _is_capturable_label(row.label):
            continue
        if _is_claimed(row, claims):
            continue
        key = row.label.strip().lower()
        if key in seen:
            continue
        seen.add(key)
        unit, scale = _classify_unit_scale(row.label, row.samples)
        out.append(
            SheetKpi(
                kpi_name=row.label.strip(),  # raw label; canonicalized at ingest
                sheet=row.sheet,
                row_label=row.label.strip(),
                unit=unit,
                scale=scale,
                origin="capture",
                exact_label=True,
            )
        )
    return out


def build_ir_config(
    ticker: str,
    xlsx_path: Path,
    *,
    platform: str,
    results_center_url: str,
    repo_root: Path | None = None,
    label_col: int = 1,
    persist: bool = True,
) -> IrConfig:
    """Build `ticker`'s parser config: curated analyst layer + capture-all layer.

    The analyst layer LLM-maps the ticker's holdings tier KPIs (empty for a ticker
    with no holdings); the capture layer maps every OTHER labeled numeric row
    deterministically. The result is non-empty whenever the spreadsheet carries any
    capturable row — so a brand-new ticker with no watchlist still configures and
    ingests its full audited series.
    """
    root = repo_root or _PROJECT_ROOT
    analyst = _analyst_kpis(ticker, xlsx_path, root, label_col)
    rows = enumerate_numeric_rows(xlsx_path, label_col=label_col)
    capture = _capture_kpis(rows, analyst=analyst)

    cfg = IrConfig(
        ticker=ticker.upper(),
        platform=platform,
        results_center_url=results_center_url,
        spreadsheet_kpis=tuple(analyst) + tuple(capture),
        label_col=label_col,
    )
    if persist:
        save_config(cfg, root)
    return cfg


def widen_config(cfg: IrConfig, xlsx_path: Path) -> IrConfig:
    """Return `cfg` with its capture layer (re)derived from `xlsx_path`.

    The analyst rows (anything not ``origin='capture'``) are preserved VERBATIM —
    this is what keeps a hand-built config (NU.json) from regressing: its four
    curated series stay on their exact definitions while every other audited row is
    added at capture origin. Deterministic and idempotent: re-widening the same
    sheet reproduces the same capture set, and a new quarter that adds rows simply
    grows it. The caller persists the result.
    """
    analyst = [k for k in cfg.spreadsheet_kpis if k.origin != "capture"]
    rows = enumerate_numeric_rows(xlsx_path, label_col=cfg.label_col)
    capture = _capture_kpis(rows, analyst=analyst)
    return IrConfig(
        ticker=cfg.ticker,
        platform=cfg.platform,
        results_center_url=cfg.results_center_url,
        spreadsheet_kpis=tuple(analyst) + tuple(capture),
        label_col=cfg.label_col,
    )
