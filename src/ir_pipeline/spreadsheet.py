"""Parse an IR historical-data spreadsheet into canonical KPI series.

Config-driven and layout-tolerant: per sheet we auto-detect the date-header row
(the row in the first dozen with the most quarter-end dates), map each data
column to a `period_end`, then for each configured `SheetKpi` pick the row whose
label-cell contains the configured substring AND carries numeric data, and read
its last N quarters. Headers mix Excel datetimes and "dd/mm/yyyy" strings, so
`_parse_period` normalizes both.

Returns ``{kpi_name: {period_end: value}}`` with `scale` already applied.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import cast

import openpyxl

from ir_pipeline.config import IrConfig

_HEADER_SCAN_ROWS = 14


def _parse_period(cell: object) -> _dt.datetime | None:
    """Normalize a header cell to a quarter-end datetime, or None if not a date."""
    if isinstance(cell, _dt.datetime):
        return cell
    if isinstance(cell, _dt.date):
        return _dt.datetime(cell.year, cell.month, cell.day)
    if isinstance(cell, str):
        s = cell.strip()
        if s.count("/") == 2:  # dd/mm/yyyy
            d, m, y = s.split("/")
            try:
                return _dt.datetime(int(y), int(m), int(d))
            except ValueError:
                return None
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":  # yyyy-mm-dd...
            try:
                return _dt.datetime(int(s[:4]), int(s[5:7]), int(s[8:10]))
            except ValueError:
                return None
    return None


def _header_row(rows: list[tuple[object, ...]]) -> tuple[int, dict[int, _dt.datetime]]:
    """Return (row index, {col_index: period_end}) for the best date-header row."""
    best_idx, best_map = -1, {}
    for i in range(min(_HEADER_SCAN_ROWS, len(rows))):
        col_map = {ci: p for ci, c in enumerate(rows[i]) if (p := _parse_period(c)) is not None}
        if len(col_map) > len(best_map):
            best_idx, best_map = i, col_map
    return best_idx, best_map


def _find_data_row(
    rows: list[tuple[object, ...]],
    label_col: int,
    label_substr: str,
    date_cols: list[int],
) -> int | None:
    """Row index whose label cell matches and that carries the most numeric data."""
    want = label_substr.lower()
    best_idx, best_count = None, 0
    for ri, r in enumerate(rows):
        if label_col >= len(r):
            continue
        lab = r[label_col]
        if not (isinstance(lab, str) and want in lab.lower()):
            continue
        n = sum(1 for c in date_cols if c < len(r) and isinstance(r[c], (int, float)))
        if n > best_count:
            best_idx, best_count = ri, n
    return best_idx


def parse_spreadsheet(
    path: Path, config: IrConfig, max_quarters: int = 8
) -> dict[str, dict[_dt.datetime, float]]:
    """Parse `path` per `config`; return {kpi_name: {period_end: value}}.

    Only the most-recent `max_quarters` per KPI are returned. KPIs whose sheet
    or row can't be located are skipped (not an error — sheets evolve).
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    # Cache each sheet's rows + header map once.
    sheet_cache: dict[str, tuple[list[tuple[object, ...]], dict[int, _dt.datetime]]] = {}
    out: dict[str, dict[_dt.datetime, float]] = {}

    for spec in config.spreadsheet_kpis:
        if spec.sheet not in wb.sheetnames:
            continue
        if spec.sheet not in sheet_cache:
            rows = cast(
                "list[tuple[object, ...]]",
                [tuple(r) for r in wb[spec.sheet].iter_rows(values_only=True)],
            )
            _, col_map = _header_row(rows)
            sheet_cache[spec.sheet] = (rows, col_map)
        rows, col_map = sheet_cache[spec.sheet]
        if not col_map:
            continue
        date_cols = sorted(col_map)
        ri = _find_data_row(rows, config.label_col, spec.row_label, date_cols)
        if ri is None:
            continue
        row = rows[ri]
        series: dict[_dt.datetime, float] = {}
        for ci in date_cols:
            val = row[ci] if ci < len(row) else None
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                series[col_map[ci]] = float(val) * spec.scale
        if series:
            recent = sorted(series)[-max_quarters:]
            out[spec.kpi_name] = {p: series[p] for p in recent}
    wb.close()
    return out
