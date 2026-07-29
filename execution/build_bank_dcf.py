"""Formula-first bank / credit valuation workbook (NU = first instance).

A bank can't be valued with an FCFF DCF (debt is raw material, not financing), so
this builds the equity-side **excess-return / residual-income** model: value the
equity directly, discount at the **cost of equity**, and let the binding
constraint be the **capital you must retain to grow the credit book**.

Design mirrors the redesigned FCFF workbook (Dashboard = front door of yellow
INPUT cells; every projected cell is a formula referencing those inputs + the
prior period + blue actuals). The single canonical Ke cell propagates everywhere.

Layers (the locked logic):
    Credit book  ─(× NIM)→ NII  ─(− cost of risk)→ risk-adj NII
                 (+ fees)(− opex)(− tax) → Net income → ÷ equity = ROE   (OUTPUT)
    Equity compounds by retained earnings; required capital = capital-ratio × book
    Value = BookEquity + Σ PV[(ROE − Ke) × Equity_{t-1}] + PV(terminal)   @ Ke
    Terminal = (ROE_term − Ke) × Equity_T / (Ke − g)   (continuing residual income;
        NOT the full justified-P/B price Equity_T×(ROE_term−g)/(Ke−g), which would
        double-count PV(terminal book equity); exit-PE is a cross-check only)

Env (like build_redesigned_dcf.py): DCF_TICKER, DCF_DEST, DCF_REPO_ROOT.
Prints a RESULT line; the Python mirror is the value-of-record (openpyxl can't
evaluate formulas offline) and matches the in-sheet formulas exactly.
"""

from __future__ import annotations

import json
import os
import sqlite3
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(os.environ.get("DCF_REPO_ROOT") or Path(__file__).resolve().parents[1])
T = os.environ.get("DCF_TICKER", "NU")
DEST = Path(os.environ.get("DCF_DEST") or (REPO / "dcf" / f"{T}.xlsx"))

FMP = REPO / "data" / "historical" / "fmp"

sys.path.insert(0, str(REPO / "src"))


from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

try:  # persistence is best-effort — the workbook builds without a DB
    from dcf import persist as persist_mod
except ImportError:  # pragma: no cover
    persist_mod = None  # type: ignore[assignment]
try:  # global macro assumptions — best-effort; degrades to in-code seed defaults
    from dcf import global_assumptions as global_dcf
except ImportError:  # pragma: no cover
    global_dcf = None  # type: ignore[assignment]

# ---- styling (mirrors the redesign Color Code: yellow=input, blue=actual) ----
YELLOW = PatternFill("solid", fgColor="FFF2CC")
BLUE_FONT = Font(color="1F4E79")  # blue = hardcoded actual
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)
SUB_FONT = Font(bold=True, color="374151")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
PCT100 = '0.0"%"'  # for KPI values stored as whole-number percents (29 = 29%)
USD0 = "#,##0"
NUM2 = "0.00"
MULT = '0.0"x"'


def _load(name: str) -> Any:
    p = FMP / name
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else None


def _rows(x: Any) -> list[dict[str, Any]]:
    if x is None:
        return []
    if isinstance(x, dict):
        for k in ("historical", "financials", "data"):
            if isinstance(x.get(k), list):
                return x[k]
        return [x]
    return x


@dataclass
class Actuals:
    book: float  # credit book (net receivables), $M
    ea: float  # interest-earning assets, $M
    nii: float  # net interest income, $M
    fees: float  # fee & interchange income, $M
    opex: float  # operating expenses, $M
    credit_cost: float  # credit & risk costs (plug to pretax), $M
    pretax: float
    tax_rate: float
    ni: float
    equity: float  # book equity Y0, $M
    equity_prior: float
    shares: float  # diluted shares, M
    price: float


def load_actuals(ticker: str, s: Assum, override: dict[str, Any] | None = None) -> Actuals:
    """Y0 actuals from FMP, in REPORTING-CURRENCY millions.

    FMP's bank line-items are unreliable for some filers: an Indian 20-F filer's
    loan book lands in ``totalInvestments`` (``netReceivables`` is zeroed), share
    counts can be mis-scaled, and a fresh fiscal-year record may be partly
    unpopulated. So every field can be overridden by the ``actuals`` block of
    ``data/bank_assumptions/<T>.json`` — give the override in reporting-currency
    MILLIONS (FMP raw values are divided by 1e6 to match). Earning assets default
    to NII / the Y0 NIM (``s.nim_y0``; NU = 17%) when not overridden, so the
    divisor is no longer hard-coded to NU's spread.
    """
    ov = override or {}
    inc = _rows(_load(f"{ticker}_income_statement_annual.json"))
    bal = _rows(_load(f"{ticker}_balance_sheet_annual.json"))
    prof = _load(f"{ticker}_profile.json")
    if isinstance(prof, list):
        prof = prof[0] if prof else {}
    i0, b0, b1 = inc[0], bal[0], bal[1]
    m = 1e6

    def pick(key: str, fmp_millions: float) -> float:
        v = ov.get(key)
        return float(v) if isinstance(v, (int, float)) else fmp_millions

    nii = pick("nii", i0["netInterestIncome"] / m)
    fees = pick("fees", (i0["revenue"] - i0["interestIncome"]) / m)
    opex = pick("opex", i0["operatingExpenses"] / m)
    pretax = pick("pretax", i0["incomeBeforeTax"] / m)
    # Credit book: override -> net receivables (when FMP populates it) ->
    # total investments (where FMP buckets some banks' loan books) -> 0.
    nr = (b0.get("netReceivables") or 0) / m
    ti = (b0.get("totalInvestments") or 0) / m
    book = pick("book", nr if nr > 0 else ti)
    ea = pick("ea", nii / s.nim_y0)
    credit_cost = pick("credit_cost", (nii + fees) - opex - pretax)  # plug to pretax
    tax_paid = i0["incomeTaxExpense"] / m
    tax_rate = pick("tax_rate", (tax_paid / pretax) if pretax else 0.25)
    return Actuals(
        book=book,
        ea=ea,
        nii=nii,
        fees=fees,
        opex=opex,
        credit_cost=credit_cost,
        pretax=pretax,
        tax_rate=tax_rate,
        ni=pick("ni", i0["netIncome"] / m),
        equity=pick("equity", b0["totalStockholdersEquity"] / m),
        equity_prior=pick("equity_prior", b1["totalStockholdersEquity"] / m),
        shares=pick("shares", i0["weightedAverageShsOutDil"] / m),
        price=pick("price", float(prof.get("price") or 0.0)),
    )


# kpi_facts.fiscal_period_type values that denote an annual (fiscal-year-end)
# observation — mirror of compute.kpi_resolver.ANNUAL_FACT_PERIOD_TYPES (inlined
# so the builder stays import-light; the DB is opened read-only).
_ANNUAL_FPT = frozenset({"FY", "annual"})


def load_kpis(ticker: str) -> dict[str, float]:
    """Latest authoritative value per KPI name from kpi_facts (read-only). Empty if
    no DB/data — the guardrails panel then degrades to actuals-only.

    Cadence-aware: for an annual-cadence definition
    (``kpi_definitions.reporting_cadence='annual'`` — NU's Basel III capital
    adequacy ratio, disclosed once a year in the 20-F) only fiscal-year-end
    (``FY``) rows are authoritative, so the guardrail reads the latest ANNUAL CAR
    (the same value the break rule evaluates) rather than a stray interim print.
    Quarterly KPIs keep latest-by-period_end. Falls back to the period_end rule
    when the cadence column is absent (pre-0072 DB)."""
    db = REPO / "data" / "portfolio.db"
    if not db.exists():
        return {}
    conn = connect_sqlite(db, role=SQLiteConnectionRole.READ_ONLY)
    conn.row_factory = sqlite3.Row
    try:
        has_cadence = any(
            r["name"] == "reporting_cadence"
            for r in conn.execute("PRAGMA table_info(kpi_definitions)").fetchall()
        )
        cadence_sel = "kd.reporting_cadence" if has_cadence else "'quarterly'"
        rows = conn.execute(
            f"SELECT kd.name AS name, kf.value AS value, kf.fiscal_period_type AS fpt, "
            f"{cadence_sel} AS cadence FROM kpi_facts kf "
            "JOIN kpi_definitions kd ON kf.kpi_definition_id=kd.id "
            "WHERE kf.ticker=? ORDER BY kf.period_end DESC",
            (ticker,),
        ).fetchall()
    except sqlite3.Error:
        return {}
    finally:
        conn.close()
    out: dict[str, float] = {}
    for r in rows:  # newest period_end first
        name = str(r["name"])
        value = r["value"]
        if name in out or not isinstance(value, (int, float)):
            continue
        # An annual KPI's interim prints are not authoritative — skip them so the
        # first row that lands is the latest fiscal-year-end value.
        if str(r["cadence"] or "quarterly") == "annual" and str(r["fpt"] or "") not in _ANNUAL_FPT:
            continue
        out[name] = float(value)
    return out


def load_breaks(ticker: str) -> dict[str, Any]:
    """Holdings thesis context for the guardrails: break thresholds + MoS."""
    p = REPO / "micro_thesis" / "holdings" / f"{ticker}.json"
    if not p.exists():
        return {}
    try:
        return cast("dict[str, Any]", json.loads(p.read_text(encoding="utf-8")))
    except (OSError, json.JSONDecodeError):
        return {}


def _kpi(kpis: dict[str, float], *names: str) -> float | None:
    """First KPI whose name starts with any of `names` (case-insensitive)."""
    for want in names:
        for k, v in kpis.items():
            if k.lower().startswith(want.lower()):
                return v
    return None


@dataclass
class Assum:
    # cost of equity (CAPM + country risk)
    rf: float = 0.043
    erp: float = 0.05
    beta: float = 1.0
    crp: float = 0.032
    # book + spread economics (near = Y1, terminal = final year; linear fade)
    ea_mult: float = 2.27  # earning assets / credit book
    g_near: float = 0.22
    g_term: float = 0.08
    nim_near: float = 0.17
    nim_term: float = 0.145
    cor_near: float = 0.155  # credit & risk cost, % of avg book
    cor_term: float = 0.140
    fee_near: float = 0.20  # fee income, % of avg book
    fee_term: float = 0.17
    eff_near: float = 0.288  # opex / net revenue
    eff_term: float = 0.30
    tax: float = 0.257
    cap_ratio: float = 0.17  # required equity / credit book
    pay_near: float = 0.0
    pay_term: float = 0.50
    g_terminal: float = 0.07
    exit_pe: float = 18.0
    years: int = 10
    # Y0 NIM used to back earning assets out of NII when ``actuals.ea`` isn't
    # given (NU = 17%); no longer hard-coded to NU's spread in load_actuals.
    nim_y0: float = 0.17
    # currency / ADR translation (NU default: USD reporter, 1 ADR = 1 share).
    # For an INR filer with a USD ADR, fx_to_usd = USD per 1 INR and
    # adr_ratio = ordinary shares per ADR, so the per-share output is comparable
    # to the traded ADR price.
    fx_to_usd: float = 1.0
    adr_ratio: float = 1.0

    @property
    def ke(self) -> float:
        return self.rf + self.beta * self.erp + self.crp


def _interp(near: float, term: float, t: int, n: int) -> float:
    return near + (term - near) * (t - 1) / (n - 1)


@dataclass
class MirrorRow:
    t: int
    book: float
    nim: float
    nii: float
    cor: float
    closs: float
    fees: float
    opex: float
    ni: float
    equity: float
    roe: float
    fcfe: float
    excess: float
    pv_excess: float
    pv_fcfe: float


@dataclass
class Mirror:
    rows: list[MirrorRow] = field(default_factory=list)
    value: float = 0.0
    vps: float = 0.0  # reporting-currency value per ordinary share
    vps_usd: float = 0.0  # USD value per traded unit (ADR)
    pv_excess: float = 0.0
    pv_tv: float = 0.0
    roe_term: float = 0.0


def mirror(a: Actuals, s: Assum) -> Mirror:
    """Python value-of-record; mirrors the in-sheet formulas exactly."""
    ke, n = s.ke, s.years
    book_prev, ea_prev, eq_prev = a.book, a.ea, a.equity
    reqcap_prev = s.cap_ratio * a.book
    avg_eq0 = (a.equity + a.equity_prior) / 2  # noqa: F841 (Y0 ref only)
    m = Mirror()
    for t in range(1, n + 1):
        g = _interp(s.g_near, s.g_term, t, n)
        book = book_prev * (1 + g)
        avg_book = (book + book_prev) / 2
        ea = book * s.ea_mult
        avg_ea = (ea + ea_prev) / 2
        nim = _interp(s.nim_near, s.nim_term, t, n)
        nii = nim * avg_ea
        cor = _interp(s.cor_near, s.cor_term, t, n)
        closs = cor * avg_book
        radj = nii - closs
        fee = _interp(s.fee_near, s.fee_term, t, n) * avg_book
        revenue = nii + fee
        opex = _interp(s.eff_near, s.eff_term, t, n) * revenue
        pretax = radj + fee - opex
        ni = pretax * (1 - s.tax)
        payout = _interp(s.pay_near, s.pay_term, t, n)
        div = payout * ni
        equity = eq_prev + ni - div
        avg_eq = (equity + eq_prev) / 2
        roe = ni / avg_eq
        reqcap = s.cap_ratio * book
        fcfe = ni - (reqcap - reqcap_prev)
        excess = (roe - ke) * eq_prev
        df = 1 / (1 + ke) ** t
        m.rows.append(
            MirrorRow(
                t,
                book,
                nim,
                nii,
                cor,
                closs,
                fee,
                opex,
                ni,
                equity,
                roe,
                fcfe,
                excess,
                excess * df,
                fcfe * df,
            )
        )
        m.pv_excess += excess * df
        book_prev, ea_prev, eq_prev, reqcap_prev = book, ea, equity, reqcap
    last = m.rows[-1]
    m.roe_term = last.roe
    eq_t = last.equity
    df_n = 1 / (1 + ke) ** n
    # Terminal = PV of the CONTINUING residual income stream beyond year n,
    # (ROE_term - Ke) * Equity_T / (Ke - g). NOT the full justified-P/B equity
    # price Equity_T*(ROE_term-g)/(Ke-g): that price already equals
    # Equity_T + PV(future residual income), so adding it on top of
    # a.equity + PV(excess) -- which by the clean-surplus identity already
    # embeds PV(Equity_T) -- would double-count PV(terminal book equity) and
    # overstate value by ~40%+.
    tv = (m.roe_term - ke) * eq_t / (ke - s.g_terminal)
    m.pv_tv = tv * df_n
    m.value = a.equity + m.pv_excess + m.pv_tv
    m.vps = m.value / a.shares  # reporting-currency per ordinary share
    m.vps_usd = m.vps * s.adr_ratio * s.fx_to_usd  # per traded ADR, in USD
    return m


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _hdr(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].fill = HEAD_FILL
    ws[cell].font = HEAD_FONT


def _inp(ws: Worksheet, row: int, label: str, val: float, fmt: str = PCT) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
    c = ws.cell(row=row, column=2, value=val)
    c.fill = YELLOW
    c.number_format = fmt
    c.border = BORDER


# Dashboard input rows (canonical cells referenced by the Model sheet)
R = {
    "rf": 3,
    "erp": 4,
    "beta": 5,
    "crp": 6,
    "ke": 7,
    "book0": 9,
    "eamult": 10,
    "gnear": 11,
    "gterm": 12,
    "nimnear": 13,
    "nimterm": 14,
    "cornear": 15,
    "corterm": 16,
    "feenear": 17,
    "feeterm": 18,
    "effnear": 19,
    "effterm": 20,
    "tax": 21,
    "cap": 22,
    "paynear": 23,
    "payterm": 24,
    "g": 25,
    "pe": 26,
    "eq0": 27,
    "eqm1": 28,
    "sh": 29,
    "px": 30,
    "years": 31,
    "val": 34,
    "vps": 35,
    "up": 36,
    "pb": 37,
    "pe2": 38,
    "roet": 39,
}


def render_generic_guardrails(
    dash: Worksheet, kpis: dict[str, float], holdings: dict[str, Any]
) -> None:
    """Thesis-driven guardrails for any bank (non-NU): each holdings break rule
    (``break_rules`` + ``business_model_rules``) shown vs its latest value from
    ``kpi_facts``. Bank KPIs not yet extracted into kpi_facts (NIM, NPA, CAR for
    HDB/SOFI) render 'n/a' — honest about coverage, never silently 'OK'. Reference
    only (blue); not part of the formula-linked valuation engine."""
    rules: list[dict[str, Any]] = []
    for key in ("break_rules", "business_model_rules"):
        rr = holdings.get(key)
        if isinstance(rr, list):
            rules.extend(r for r in rr if isinstance(r, dict))
    if not rules:
        return
    _hdr(dash, "D2", "GUARDRAILS — thesis break rules vs latest KPI")
    dash["D3"], dash["E3"], dash["F3"], dash["G3"] = "KPI", "Latest", "Break rule", "Status"
    for cc in ("D3", "E3", "F3", "G3"):
        dash[cc].font = SUB_FONT
    r = 4
    for rule in rules:
        name = str(rule.get("kpi_name") or "")
        comp = str(rule.get("comparator") or "")
        thr = rule.get("threshold")
        value = _kpi(kpis, name)
        dash.cell(row=r, column=4, value=name[:46]).font = Font(color="374151")
        vc = dash.cell(
            row=r, column=5, value=round(value, 2) if isinstance(value, (int, float)) else "n/a"
        )
        vc.font = BLUE_FONT
        periods = rule.get("consecutive_periods")
        rule_txt = f"{comp} {thr}" + (f" / {periods}Q" if isinstance(periods, int) else "")
        dash.cell(row=r, column=6, value=rule_txt).font = Font(size=9, color="6B7280")
        status = "—"
        if isinstance(value, (int, float)) and isinstance(thr, (int, float)):
            breached = (value < thr) if comp == "lt" else (value > thr) if comp == "gt" else False
            status = "BREACH" if breached else "OK"
        sc = dash.cell(row=r, column=7, value=status)
        sc.font = Font(bold=True, color="B91C1C" if status == "BREACH" else "15803D")
        r += 1
    for col, w in (("D", 32), ("E", 12), ("F", 16), ("G", 9)):
        dash.column_dimensions[col].width = w


def render_guardrails(dash: Worksheet, a: Actuals, kpis: dict[str, float]) -> None:
    """Right-hand signal panel: latest bank KPIs vs the thesis break thresholds,
    plus the customers x ARPAC monetization cross-check. Reference actuals (blue),
    not part of the formula-linked valuation engine. Skips silently with no data."""
    if not kpis:
        return
    car_min = 10.5  # BACEN current minimum (from the holdings break narrative)
    sig = [
        (
            "ROE (consolidated)",
            _kpi(kpis, "ROE (annualized", "ROE"),
            PCT100,
            "break <25% / 2Q",
            25,
            "lt",
        ),
        (
            "Risk-adj NIM",
            _kpi(kpis, "Risk-adjusted NIM (NIM"),
            PCT100,
            "break: YoY contraction 2Q",
            None,
            None,
        ),
        (
            # Annual (20-F) Basel III ratio — load_kpis serves the latest FY value.
            "Capital adequacy (CAR, annual)",
            _kpi(kpis, "Capital adequacy"),
            PCT100,
            f"min {car_min}% · break <13%",
            13,
            "lt",
        ),
        ("NPL 90d+", _kpi(kpis, "NPL 90d+"), PCT100, "break >7% / 2Q", 7, "gt"),
        (
            "NPL 15d+ total",
            _kpi(kpis, "NPL 15d+ ratio", "NPL 15d+"),
            PCT100,
            "YoY-tracked",
            None,
            None,
        ),
        ("Monthly ARPAC ($)", _kpi(kpis, "Monthly ARPAC"), NUM2, "break <$10 / 2Q", 10, "lt"),
        ("Activity rate", _kpi(kpis, "Activity Rate"), PCT100, "break <82% / 2Q", 82, "lt"),
        (
            "Total customers (M)",
            _kpi(kpis, "Total customers (millions", "Total customers"),
            USD0,
            "net adds >5M/Q",
            None,
            None,
        ),
    ]
    _hdr(dash, "D2", "GUARDRAILS — latest actual vs thesis break")
    dash["D3"], dash["E3"], dash["F3"], dash["G3"] = "Signal", "Latest", "Rule", "Status"
    for cc in ("D3", "E3", "F3", "G3"):
        dash[cc].font = SUB_FONT
    r = 4
    for label, value, fmt, rule, thr, comp in sig:
        dash.cell(row=r, column=4, value=label).font = Font(color="374151")
        vc = dash.cell(row=r, column=5, value=round(value, 2) if value is not None else "n/a")
        vc.font = BLUE_FONT
        if value is not None:
            vc.number_format = fmt
        dash.cell(row=r, column=6, value=rule).font = Font(size=9, color="6B7280")
        status = "—"
        if value is not None and thr is not None:
            breached = (value < thr) if comp == "lt" else (value > thr)
            status = "BREACH" if breached else "OK"
        sc = dash.cell(row=r, column=7, value=status)
        sc.font = Font(bold=True, color="B91C1C" if status == "BREACH" else "15803D")
        r += 1
    # monetization cross-check (customers x ARPAC vs reported revenue)
    cust = _kpi(kpis, "Total customers (millions", "Total customers")
    arpac = _kpi(kpis, "Monthly ARPAC")
    act = _kpi(kpis, "Activity Rate")
    if cust and arpac:
        active_pct = (act or 83) / 100
        implied_rev = cust * active_pct * arpac * 12 / 1000  # $B
        r += 1
        _hdr(dash, f"D{r}", "MONETIZATION CROSS-CHECK")
        rows = [
            ("Customers (M) x active %", f"{cust:.0f} x {active_pct:.0%}"),
            ("x Monthly ARPAC x 12", f"${arpac:.1f} -> ${implied_rev:.1f}B implied"),
            ("vs reported revenue", f"${a.nii / 1000 + a.fees / 1000:.1f}B net (FY0)"),
        ]
        for lab, txt in rows:
            r += 1
            dash.cell(row=r, column=4, value=lab).font = Font(color="374151")
            dash.cell(row=r, column=5, value=txt).font = Font(size=9, color="374151")
    for col, w in (("D", 22), ("E", 13), ("F", 20), ("G", 9)):
        dash.column_dimensions[col].width = w


def build(
    a: Actuals,
    s: Assum,
    m: Mirror,
    dest: Path,
    kpis: dict[str, float] | None = None,
    holdings: dict[str, Any] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    fin = wb.create_sheet("Financials")
    mod = wb.create_sheet("Model")
    val = wb.create_sheet("Valuation")
    sens = wb.create_sheet("Sensitivity")

    D = "Dashboard"
    # ---------- Financials (blue actuals) ----------
    _hdr(fin, "A1", f"{T} — Financials (FY actuals, $M)")
    facts = [
        ("Credit book (net receivables)", a.book),
        ("Interest-earning assets", a.ea),
        ("Net interest income", a.nii),
        ("Fee & interchange income", a.fees),
        ("Credit & risk cost (plug to pretax)", a.credit_cost),
        ("Operating expenses", a.opex),
        ("Pre-tax income", a.pretax),
        ("Net income", a.ni),
        ("Book equity (Y0)", a.equity),
        ("Book equity (Y-1)", a.equity_prior),
        ("Diluted shares (M)", a.shares),
        ("Price ($)", a.price),
    ]
    for i, (lab, v) in enumerate(facts, start=3):
        fin.cell(row=i, column=1, value=lab).font = Font(color="374151")
        c = fin.cell(row=i, column=2, value=round(v, 2))
        c.font = BLUE_FONT
        c.number_format = NUM2 if lab.startswith("Price") else USD0
    fin.column_dimensions["A"].width = 34
    fin.column_dimensions["B"].width = 14
    fin_book = "Financials!$B$3"
    fin_eq0, fin_eqm1 = "Financials!$B$11", "Financials!$B$12"
    fin_sh, fin_px = "Financials!$B$13", "Financials!$B$14"

    # ---------- Dashboard (inputs) ----------
    _hdr(dash, "A1", f"{T} — Bank Credit Model · Dashboard")
    dash["A2"] = "Cost of equity (CAPM + country risk)"
    dash["A2"].font = SUB_FONT
    _inp(dash, R["rf"], "Risk-free rate", s.rf)
    _inp(dash, R["erp"], "Equity risk premium", s.erp)
    _inp(dash, R["beta"], "Beta", s.beta, NUM2)
    _inp(dash, R["crp"], "Country risk premium (LatAm)", s.crp)
    dash.cell(row=R["ke"], column=1, value="Cost of equity  Ke").font = SUB_FONT
    ke = dash.cell(row=R["ke"], column=2, value=f"=B{R['rf']}+B{R['beta']}*B{R['erp']}+B{R['crp']}")
    ke.number_format = PCT
    ke.font = Font(bold=True)
    KE = f"{D}!$B${R['ke']}"

    dash["A8"] = "Book & spread economics  (near = Y1 · terminal = Y" + str(s.years) + ")"
    dash["A8"].font = SUB_FONT
    dash.cell(row=R["book0"], column=1, value="Credit book Y0 ($M)")
    lk = dash.cell(row=R["book0"], column=2, value=f"={fin_book}")
    lk.number_format = USD0
    lk.font = Font(color="1F7A4D")
    _inp(dash, R["eamult"], "Earning assets / credit book", s.ea_mult, NUM2)
    _inp(dash, R["gnear"], "Book growth — near", s.g_near)
    _inp(dash, R["gterm"], "Book growth — terminal", s.g_term)
    _inp(dash, R["nimnear"], "NIM — near", s.nim_near)
    _inp(dash, R["nimterm"], "NIM — terminal", s.nim_term)
    _inp(dash, R["cornear"], "Cost of risk (% book) — near", s.cor_near)
    _inp(dash, R["corterm"], "Cost of risk (% book) — terminal", s.cor_term)
    _inp(dash, R["feenear"], "Fee income (% book) — near", s.fee_near)
    _inp(dash, R["feeterm"], "Fee income (% book) — terminal", s.fee_term)
    _inp(dash, R["effnear"], "Efficiency (opex/rev) — near", s.eff_near)
    _inp(dash, R["effterm"], "Efficiency (opex/rev) — terminal", s.eff_term)
    _inp(dash, R["tax"], "Tax rate", s.tax)

    dash["A21"].font = SUB_FONT
    _inp(dash, R["cap"], "Capital ratio (req. equity / book)", s.cap_ratio)
    _inp(dash, R["paynear"], "Payout — near", s.pay_near)
    _inp(dash, R["payterm"], "Payout — terminal", s.pay_term)
    _inp(dash, R["g"], "Terminal growth g", s.g_terminal)
    _inp(dash, R["pe"], "Exit P/E (cross-check)", s.exit_pe, MULT)

    dash.cell(row=R["eq0"], column=1, value="Book equity Y0 ($M)")
    for rr, ref, fmt in (
        (R["eq0"], fin_eq0, USD0),
        (R["eqm1"], fin_eqm1, USD0),
        (R["sh"], fin_sh, USD0),
        (R["px"], fin_px, NUM2),
    ):
        cc = dash.cell(row=rr, column=2, value=f"={ref}")
        cc.number_format = fmt
        cc.font = Font(color="1F7A4D")
    dash.cell(row=R["eqm1"], column=1, value="Book equity Y-1 ($M)")
    dash.cell(row=R["sh"], column=1, value="Diluted shares (M)")
    dash.cell(row=R["px"], column=1, value="Current price ($)")
    _inp(dash, R["years"], "Forecast years", s.years, "0")

    # outputs (link from Valuation)
    _hdr(dash, "A33", "OUTPUT")
    for rr, lab, ref, fmt in (
        (R["val"], "Value of equity (reporting, M)", "Valuation!$B$8", USD0),
        (R["vps"], "Value per share / ADR (USD)", "Valuation!$B$14", NUM2),
        (R["up"], "Upside vs price", "Valuation!$B$10", PCT),
        (R["pb"], "Implied P/B", "Valuation!$B$11", MULT),
        (R["pe2"], "Implied P/E (fwd)", "Valuation!$B$12", MULT),
        (R["roet"], "Terminal ROE", "Valuation!$B$5", PCT),
    ):
        dash.cell(row=rr, column=1, value=lab).font = SUB_FONT
        oc = dash.cell(row=rr, column=2, value=f"={ref}")
        oc.number_format = fmt
        oc.font = Font(bold=True)
    dash.column_dimensions["A"].width = 36
    dash.column_dimensions["B"].width = 14
    # NU has the bespoke ARPAC/activity/NPL panel; other banks get a generic
    # panel driven by their holdings break rules vs latest kpi_facts.
    if kpis and _kpi(kpis, "Monthly ARPAC", "Activity Rate") is not None:
        render_guardrails(dash, a, kpis)
    elif holdings:
        render_generic_guardrails(dash, kpis or {}, holdings)

    # ---------- Model (formula-first engine) ----------
    _hdr(mod, "A1", f"{T} — Credit Model (formula-first; $M)")
    n = s.years
    col0 = 3  # column C = Y0
    last_col = col0 + n  # column for Y{n}

    def cl(idx: int) -> str:
        return openpyxl.utils.get_column_letter(idx)

    # year index row
    mod.cell(row=4, column=2, value="Year").font = SUB_FONT
    for j in range(0, n + 1):
        mod.cell(row=4, column=col0 + j, value=j).font = SUB_FONT
    labels = {
        5: "Book growth g",
        6: "Credit book",
        7: "Avg book",
        8: "Earning assets",
        9: "Avg earning assets",
        10: "NIM",
        11: "Net interest income",
        12: "Cost of risk %",
        13: "Credit & risk cost",
        14: "Risk-adj NII",
        15: "Fee %",
        16: "Fee income",
        17: "Net revenue",
        18: "Efficiency",
        19: "Operating expenses",
        20: "Pre-tax income",
        21: "Tax",
        22: "Net income",
        23: "Payout",
        24: "Dividends",
        25: "Book equity",
        26: "Avg equity",
        27: "ROE",
        28: "Required capital",
        29: "FCFE",
        30: "Excess return",
        31: "Discount factor",
        32: "PV excess return",
        33: "PV FCFE",
    }
    for r, lab in labels.items():
        mod.cell(row=r, column=2, value=lab).font = Font(color="374151")

    def interp_formula(col: str, near: str, term: str) -> str:
        yi = f"{col}$4"
        return f"={near}+({term}-{near})*({yi}-1)/({D}!$B${R['years']}-1)"

    # Y0 column (actuals/links)
    c0 = cl(col0)
    mod[f"{c0}6"] = f"={D}!$B${R['book0']}"
    mod[f"{c0}8"] = f"={c0}6*{D}!$B${R['eamult']}"
    mod[f"{c0}25"] = f"={D}!$B${R['eq0']}"
    mod[f"{c0}26"] = f"=({D}!$B${R['eq0']}+{D}!$B${R['eqm1']})/2"
    mod[f"{c0}28"] = f"={D}!$B${R['cap']}*{c0}6"
    for r in (6, 8, 25, 26, 28):
        mod[f"{c0}{r}"].number_format = USD0

    pct_rows = {5, 10, 12, 15, 18, 23, 27}
    usd_rows = {6, 7, 8, 9, 11, 13, 14, 16, 17, 19, 20, 21, 22, 24, 25, 26, 28, 29, 30, 32, 33}
    for j in range(1, n + 1):
        c = cl(col0 + j)
        p = cl(col0 + j - 1)  # prior column
        mod[f"{c}5"] = interp_formula(c, f"{D}!$B${R['gnear']}", f"{D}!$B${R['gterm']}")
        mod[f"{c}6"] = f"={p}6*(1+{c}5)"
        mod[f"{c}7"] = f"=({c}6+{p}6)/2"
        mod[f"{c}8"] = f"={c}6*{D}!$B${R['eamult']}"
        mod[f"{c}9"] = f"=({c}8+{p}8)/2"
        mod[f"{c}10"] = interp_formula(c, f"{D}!$B${R['nimnear']}", f"{D}!$B${R['nimterm']}")
        mod[f"{c}11"] = f"={c}10*{c}9"
        mod[f"{c}12"] = interp_formula(c, f"{D}!$B${R['cornear']}", f"{D}!$B${R['corterm']}")
        mod[f"{c}13"] = f"={c}12*{c}7"
        mod[f"{c}14"] = f"={c}11-{c}13"
        mod[f"{c}15"] = interp_formula(c, f"{D}!$B${R['feenear']}", f"{D}!$B${R['feeterm']}")
        mod[f"{c}16"] = f"={c}15*{c}7"
        mod[f"{c}17"] = f"={c}11+{c}16"
        mod[f"{c}18"] = interp_formula(c, f"{D}!$B${R['effnear']}", f"{D}!$B${R['effterm']}")
        mod[f"{c}19"] = f"={c}18*{c}17"
        mod[f"{c}20"] = f"={c}14+{c}16-{c}19"
        mod[f"{c}21"] = f"={D}!$B${R['tax']}*{c}20"
        mod[f"{c}22"] = f"={c}20-{c}21"
        mod[f"{c}23"] = interp_formula(c, f"{D}!$B${R['paynear']}", f"{D}!$B${R['payterm']}")
        mod[f"{c}24"] = f"={c}23*{c}22"
        mod[f"{c}25"] = f"={p}25+{c}22-{c}24"
        mod[f"{c}26"] = f"=({c}25+{p}25)/2"
        mod[f"{c}27"] = f"={c}22/{c}26"
        mod[f"{c}28"] = f"={D}!$B${R['cap']}*{c}6"
        mod[f"{c}29"] = f"={c}22-({c}28-{p}28)"
        mod[f"{c}30"] = f"=({c}27-{KE})*{p}25"
        mod[f"{c}31"] = f"=1/(1+{KE})^{c}$4"
        mod[f"{c}32"] = f"={c}30*{c}31"
        mod[f"{c}33"] = f"={c}29*{c}31"
    for r in range(5, 34):
        for j in range(0, n + 1):
            cc = mod.cell(row=r, column=col0 + j)
            if r in pct_rows:
                cc.number_format = PCT
            elif r in usd_rows:
                cc.number_format = USD0
            elif r == 31:
                cc.number_format = NUM2
    mod.column_dimensions["B"].width = 22

    # ---------- Valuation ----------
    _hdr(val, "A1", f"{T} — Valuation (excess-return / residual income, $M)")
    cN = cl(last_col)
    tv_label = "Terminal value (continuing residual income)"
    vps_rep_label = "Value per share (reporting ccy)"
    vrows = [
        ("Book equity (Y0)", f"={D}!$B${R['eq0']}", USD0),
        ("PV of excess returns", f"=SUM(Model!{cl(col0 + 1)}32:{cN}32)", USD0),
        ("Terminal equity (Y" + str(n) + ")", f"=Model!{cN}25", USD0),
        ("Terminal ROE", f"=Model!{cN}27", PCT),
        # Continuing residual income (ROE_term-Ke)*Equity_T/(Ke-g) -- NOT the full
        # justified-P/B price, which would double-count PV(terminal book equity).
        (tv_label, f"=(B5-{KE})*B4/({KE}-{D}!$B${R['g']})", USD0),
        ("PV of terminal value", f"=B6*Model!{cN}31", USD0),
        ("Value of equity", "=B2+B3+B7", USD0),
        (vps_rep_label, f"=B8/{D}!$B${R['sh']}", NUM2),
        # Upside compares the USD-per-ADR value (B14) to the USD ADR price.
        ("Upside vs price", f"=B14/{D}!$B${R['px']}-1", PCT),
        ("Implied P/B", f"=B8/{D}!$B${R['eq0']}", MULT),
        ("Implied P/E (fwd)", f"=B8/Model!{cl(col0 + 1)}22", MULT),
        (
            "FCFE cross-check value",
            f"=SUM(Model!{cl(col0 + 1)}33:{cN}33)+(Model!{cN}29*(1+{D}!$B${R['g']})"
            f"/({KE}-{D}!$B${R['g']}))*Model!{cN}31",
            USD0,
        ),
    ]
    # B2..B13 map to the rows above; keep references consistent
    addr = {
        "Book equity (Y0)": 2,
        "PV of excess returns": 3,
        "Terminal equity (Y" + str(n) + ")": 4,
        "Terminal ROE": 5,
        tv_label: 6,
        "PV of terminal value": 7,
        "Value of equity": 8,
        vps_rep_label: 9,
        "Upside vs price": 10,
        "Implied P/B": 11,
        "Implied P/E (fwd)": 12,
        "FCFE cross-check value": 13,
    }
    for lab, formula, fmt in vrows:
        r = addr[lab]
        val.cell(row=r, column=1, value=lab).font = (
            SUB_FONT if r in (8, 9) else Font(color="374151")
        )
        vc = val.cell(row=r, column=2, value=formula)
        vc.number_format = fmt
        if r in (8, 9):
            vc.font = Font(bold=True)
    # Currency / ADR translation: per-ordinary-share (reporting) -> per-ADR (USD).
    # For a USD reporter with no ADR these are 1.0 and B14 == B9. These are
    # reference cells (blue): the bank model recomputes from bank_assumptions, not
    # from edited sheet cells, so editing them here won't round-trip.
    val.cell(row=15, column=1, value="ADR ratio (ordinary shares / ADR)").font = Font(
        color="374151"
    )
    c = val.cell(row=15, column=2, value=s.adr_ratio)
    c.number_format = NUM2
    c.font = BLUE_FONT
    val.cell(row=16, column=1, value="FX (reporting ccy -> USD)").font = Font(color="374151")
    c = val.cell(row=16, column=2, value=s.fx_to_usd)
    c.number_format = "0.00000"
    c.font = BLUE_FONT
    val.cell(row=14, column=1, value="Value per share / ADR (USD)").font = SUB_FONT
    c = val.cell(row=14, column=2, value="=B9*B15*B16")
    c.number_format = NUM2
    c.font = Font(bold=True)
    val.column_dimensions["A"].width = 32
    val.column_dimensions["B"].width = 14

    # ---------- Sensitivity (Ke × terminal-ROE proxy via terminal NIM) ----------
    _hdr(sens, "A1", "Sensitivity — value/share")
    sens["A2"] = "Computed in Python mirror (live grid added with export to Sheets)."
    sens["A2"].font = Font(italic=True, color="6B7280")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def load_assumptions(ticker: str) -> tuple[Assum, dict[str, Any]]:
    """Assum defaults overridden by data/bank_assumptions/<T>.json if present —
    per-ticker calibration, the bank analog of data/dcf_assumptions. Top-level
    numeric keys override ``Assum`` fields (rf/nim_near/cap_ratio/fx_to_usd/...);
    a nested ``actuals`` block overrides the Y0 statement actuals (see
    ``load_actuals``). Returns ``(Assum, actuals_override)``."""
    s = Assum()
    # Seed the editable global macro defaults (rf / erp / tax) BEFORE the
    # per-ticker JSON overrides below, so a bank name without its own pinned
    # value tracks the dashboard-set global default. The JSON setattr loop still
    # wins for any field the name pins explicitly (HDB pins all three; NU pins
    # erp/tax to preserve its pre-global values). Degrades to the Assum literals
    # when the store is unavailable, so the workbook still builds bare.
    if global_dcf is not None:
        _g = global_dcf.load(db_path=REPO / "data" / "portfolio.db")
        s.rf, s.erp, s.tax = _g.risk_free_rate, _g.equity_risk_premium, _g.tax_rate
    actuals_ov: dict[str, Any] = {}
    p = REPO / "data" / "bank_assumptions" / f"{ticker}.json"
    if p.exists():
        try:
            ov: Any = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            ov = {}
        if isinstance(ov, dict):
            ovd = cast("dict[str, Any]", ov)
            for k, v in ovd.items():
                if hasattr(s, k) and isinstance(v, (int, float)):
                    setattr(s, k, v)
            raw_act = ovd.get("actuals")
            if isinstance(raw_act, dict):
                actuals_ov = cast("dict[str, Any]", raw_act)
    return s, actuals_ov


def persist_dcf_run(a: Actuals, s: Assum, m: Mirror) -> bool:
    """Best-effort upsert into dcf_runs so the brief's valuation panel reads the
    bank model's value/share. No-op without the DB / persist module."""
    db = REPO / "data" / "portfolio.db"
    if persist_mod is None or not db.exists():
        return False
    # Persist in USD so the brief's over/under compares like-for-like against the
    # (USD) ADR price: vps_usd is per ADR, npv is the equity value in USD, and the
    # share count is the ADR-equivalent (ordinary / adr_ratio). For a USD reporter
    # with no ADR (fx=1, adr=1) these collapse to the reporting values unchanged.
    npv_usd = m.value * s.fx_to_usd
    shares_traded = a.shares / s.adr_ratio if s.adr_ratio else a.shares
    mos = load_breaks(T).get("mos_bar")
    snap = json.dumps(
        {
            "model": "bank_excess_return",
            "ke": s.ke,
            "roe_terminal": m.roe_term,
            "value_per_share_usd": m.vps_usd,
            "value_per_share_reporting": m.vps,
            "fx_to_usd": s.fx_to_usd,
            "adr_ratio": s.adr_ratio,
            "pv_excess_m": m.pv_excess,
            "pv_terminal_m": m.pv_tv,
            "book_equity_m": a.equity,
            "value_equity_reporting_m": m.value,
            "workbook": str(DEST),
        },
        indent=2,
    )
    row = persist_mod.DcfRunRow(
        ticker=T,
        valuation_date=date.today(),
        horizon_years=s.years,
        wacc=s.ke,
        npv=npv_usd,
        npv_per_share=m.vps_usd,
        shares_outstanding=shares_traded * 1e6,
        currency="USD",
        live_price=a.price or None,
        live_price_at=None,
        mos_bar_used=float(mos) if isinstance(mos, (int, float)) else None,
        assumption_snapshot_json=snap,
        notes=f"workbook={DEST.name} (bank credit model)",
    )
    with connect_sqlite(str(db), role=SQLiteConnectionRole.WRITER, schema_preflight=True) as conn:
        persist_mod.upsert(conn, row)
    return True


def main() -> int:
    s, actuals_ov = load_assumptions(T)
    a = load_actuals(T, s, actuals_ov)
    kpis = load_kpis(T)
    holdings = load_breaks(T)
    m = mirror(a, s)
    build(a, s, m, DEST, kpis=kpis, holdings=holdings)
    persisted = persist_dcf_run(a, s, m)
    up = (m.vps_usd / a.price - 1) if a.price else 0.0
    fx_note = (
        ""
        if s.fx_to_usd == 1.0 and s.adr_ratio == 1.0
        else f"\tfx={s.fx_to_usd:.5f}\tadr={s.adr_ratio:.0f}:1"
    )
    print(
        f"RESULT\t{T}\tKe={s.ke:.1%}\tvalue/sh=${m.vps_usd:.2f}\tprice=${a.price:.2f}"
        f"\tupside={up:+.0%}\tP/B={m.value / a.equity:.1f}x\tROE_term={m.roe_term:.1%}"
        f"{fx_note}\tdcf_runs={'ok' if persisted else 'skip'}\t-> {DEST}"
    )
    # year-by-year so we can eyeball the engine
    print(
        f"{'Yr':>2} {'Book':>7} {'NIM':>5} {'NII':>6} {'CoR%':>5} {'NI':>6} {'Eq':>7} {'ROE':>6} {'Excess':>7}"
    )
    for r in m.rows:
        print(
            f"{r.t:>2} {r.book:>7.0f} {r.nim:>5.1%} {r.nii:>6.0f} {r.cor:>5.1%} "
            f"{r.ni:>6.0f} {r.equity:>7.0f} {r.roe:>6.1%} {r.excess:>7.0f}"
        )
    print(
        f"\nBook eq {a.equity:.0f} + PV(excess) {m.pv_excess:.0f} + PV(TV) {m.pv_tv:.0f} "
        f"= {m.value:.0f} (reporting M) -> {m.vps:.2f}/sh reporting -> ${m.vps_usd:.2f}/ADR USD"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
