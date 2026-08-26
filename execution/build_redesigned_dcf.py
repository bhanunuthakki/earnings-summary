"""Build a Damodaran/MBI-style, formula-driven DCF workbook for one ticker.

Nine sheets: Cover, Dashboard (the control surface), Color Code, WACC, Model,
Financials, Consensus, Valuation, Monte Carlo. Everything derived is an
Excel/Sheets formula; only blue=hardcoded actuals and yellow=assumptions are
literals (green=cross-sheet link, orange=moved-off-consensus). Default terminal =
exit multiple with perpetuity as a cross-check; terminal method + exit basis are
data-validation dropdowns; the Monte Carlo is a live in-sheet simulation. Per-name
Opus assumptions (data/dcf_assumptions/<T>.json["redesign"], from
dcf_opus_assumptions.py) override the consensus-anchored defaults; names flagged
dcf_applicable=false (banks/insurers/asset-managers) are skipped.

Env-driven so a driver can fan out over tickers:
  DCF_TICKER     ticker to build (default AMZN)
  DCF_DEST       output .xlsx path (default dcf/<T>_redesign.xlsx)
  DCF_NAME       display-name override (default from the FMP profile)
  DCF_REPO_ROOT  repo root holding data/ (default: this file's repo root)
"""

from __future__ import annotations

import json
import os
import sqlite3
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# The scenario/sensitivity engine is shared with the reader/refresher so the
# builder-written static cells and the refresh-rewritten ones come from ONE
# implementation. Code lives next to this script regardless of DCF_REPO_ROOT
# (which points at the DATA repo).
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from compute.segment_cache import apply_overrides
from dcf import analyst_segments as analyst_seg_mod
from dcf import (
    assumptions_doc,
    country_risk,
    equity_bridge,
    fade_calibration,
    primary_fact_overlay,
    segment_coverage,
)
from dcf import global_assumptions as global_dcf
from dcf import redesign as redesign_mod
from sqlite_runtime import SQLiteConnectionRole, connect_sqlite

REPO = Path(os.environ.get("DCF_REPO_ROOT") or Path(__file__).resolve().parents[1])
FMP = REPO / "data" / "historical" / "fmp"

# Global macro DCF assumptions — the editable single default for the inputs that
# should be the same across every model (risk-free / ERP / tax). Read once here;
# a per-ticker `_opus` block override still wins at every use site below. Degrades
# to the in-code seed (= the historical literals) when the DB/table is absent, so
# the build is identical to pre-global-assumptions behaviour on a bare checkout.
_g = global_dcf.load(db_path=REPO / "data" / "portfolio.db")

T = os.environ.get("DCF_TICKER", "AMZN")
_pfd = (
    json.loads((FMP / f"{T}_profile.json").read_text(encoding="utf-8"))
    if (FMP / f"{T}_profile.json").exists()
    else [{}]
)
_pfd = (
    (_pfd[0] if isinstance(_pfd, list) and _pfd else _pfd) if isinstance(_pfd, (list, dict)) else {}
)
NAME = (
    os.environ.get("DCF_NAME") or (_pfd.get("companyName") if isinstance(_pfd, dict) else None) or T
)
DEST = Path(os.environ.get("DCF_DEST", str(REPO / "dcf" / f"{T}_redesign.xlsx")))
QUARTERS = 28  # ~7y of quarterly history
N_ACTUAL_FY = 5  # actual FY columns on the Model
N_FC = 10  # forecast years

# ----------------------------------------------------------------------------- styles
BLUE = Font(color="0000CC")
BLUEB = Font(color="0000CC", bold=True)
GREEN = Font(color="008000")
GREENB = Font(color="008000", bold=True)
BLK = Font()
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=15)
BIG = Font(bold=True, size=16, color="1F4E78")
SUB = Font(italic=True, color="666666")
SEC = Font(bold=True, color="FFFFFF")
SECFILL = PatternFill("solid", fgColor="1F4E78")
YEL = PatternFill("solid", fgColor="FFFF00")
ACTH = PatternFill("solid", fgColor="DDEBF7")
FCH = PatternFill("solid", fgColor="FFF2CC")
RIGHT = Alignment(horizontal="right")
WRAP = Alignment(wrap_text=True, vertical="top")
USD, PCT, MULT, PXS, NUM3 = "#,##0", "0.0%", '0.0"x"', '"$"#,##0.00', "0.000"


def put(ws, r, c, v, *, fmt=None, kind="f", bold=False):
    """kind: act=blue hardcode, in=yellow input, f=formula (auto-green if cross-sheet)."""
    cell = ws.cell(r, c, v)
    if fmt:
        cell.number_format = fmt
    if kind == "act":
        cell.font = BLUEB if bold else BLUE
    elif kind == "in":
        cell.font = BLUE
        cell.fill = YEL
    else:
        link = isinstance(v, str) and v.startswith("=") and "!" in v
        if link:
            cell.font = GREENB if bold else GREEN
        else:
            cell.font = BOLD if bold else BLK
    return cell


def band(ws, r, text, ncol):
    for c in range(1, ncol + 1):
        ws.cell(r, c).fill = SECFILL
    ws.cell(r, 1, "  " + text).font = SEC


def ie(expr):
    """Wrap a ratio/growth formula so a blank/zero denominator shows blank,
    not #DIV/0! (Google Sheets is stricter than the offline engine)."""
    return f'=IFERROR({expr},"")'


# ----------------------------------------------------------------------------- data
def load(stmt):
    p = FMP / f"{T}_{stmt}_quarterly.json"
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []


inc, bal, cf = load("income_statement"), load("balance_sheet"), load("cash_flow")


def _apply_primary_fact_overlay(
    statement: primary_fact_overlay.Statement, records: object
) -> tuple[object, dict[str, object]]:
    """Overlay exact primary facts without making the FMP cache a write target."""
    if not isinstance(records, list) or not all(isinstance(record, dict) for record in records):
        detail = {
            "status": "degraded",
            "degraded_reason": "FMP statement payload is not a list of rows",
            "applied": [],
            "conflicts": [],
            "rejected": [],
        }
        _emit_primary_fact_overlay(statement, detail)
        return records, detail
    db_path = REPO / "data" / "portfolio.db"
    if not db_path.exists():
        detail = {
            "status": "degraded",
            "degraded_reason": "portfolio database unavailable",
            "applied": [],
            "conflicts": [],
            "rejected": [],
        }
        _emit_primary_fact_overlay(statement, detail)
        return records, detail
    try:
        with connect_sqlite(str(db_path), role=SQLiteConnectionRole.READ_ONLY) as conn:
            result = primary_fact_overlay.overlay_quarterly_records(
                conn, ticker=T, statement=statement, records=records
            )
    except (OSError, sqlite3.Error) as error:
        detail = {
            "status": "degraded",
            "degraded_reason": f"primary fact overlay unavailable: {error}",
            "applied": [],
            "conflicts": [],
            "rejected": [],
        }
        _emit_primary_fact_overlay(statement, detail)
        return records, detail
    detail = result.to_provenance_dict()
    _emit_primary_fact_overlay(statement, detail)
    return result.records, detail


def _emit_primary_fact_overlay(
    statement: primary_fact_overlay.Statement, detail: dict[str, object]
) -> None:
    print(
        json.dumps(
            {"event": "dcf_primary_fact_overlay", "ticker": T, "statement": statement, **detail},
            sort_keys=True,
        ),
        file=sys.stderr,
    )


inc, PRIMARY_FACT_OVERLAY_INCOME = _apply_primary_fact_overlay("income", inc)
bal, PRIMARY_FACT_OVERLAY_BALANCE = _apply_primary_fact_overlay("balance", bal)
cf, PRIMARY_FACT_OVERLAY_CASH_FLOW = _apply_primary_fact_overlay("cash_flow", cf)
PRIMARY_FACT_OVERLAY = {
    "income": PRIMARY_FACT_OVERLAY_INCOME,
    "balance": PRIMARY_FACT_OVERLAY_BALANCE,
    "cash_flow": PRIMARY_FACT_OVERLAY_CASH_FLOW,
}


def _loadjson(name):
    p = FMP / name
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []


# Apply company-doc overrides (e.g. GOOG Q4'25 8-K product segmentation) so a
# re-fetched contaminated FMP record can't reach the DCF model. Best-effort: no
# override / no DB -> raw FMP data, exactly as before.
_OVR_DB = str(REPO / "data" / "portfolio.db")
prod_seg = apply_overrides(
    _loadjson(f"{T}_product_segments_quarterly.json"),
    ticker=T,
    dim_type="product",
    db_path=_OVR_DB,
)
geo_seg = apply_overrides(
    _loadjson(f"{T}_geo_segments_quarterly.json"),
    ticker=T,
    dim_type="geography",
    db_path=_OVR_DB,
)
est = _loadjson(f"{T}_analyst_estimates_annual.json")
prof = _loadjson(f"{T}_profile.json")
prof = prof[0] if isinstance(prof, list) else prof


def idx(records, segmode=False):
    out = {}
    for r in records:
        per = r.get("period")
        try:
            yr = int(r.get("fiscalYear"))
        except (TypeError, ValueError):
            continue
        if isinstance(per, str) and per.startswith("Q"):
            out[(yr, per)] = r.get("data") if segmode else r
    return out


def detect_fy_periods(
    records_i: dict[tuple[int, str], object],
    default: tuple[str, ...] = ("Q1", "Q2", "Q3", "Q4"),
) -> tuple[str, ...]:
    """Canonical period labels that make up ONE fiscal year for this issuer.

    Generalises the "four quarters per year" assumption to any consistent cadence
    — e.g. a semi-annual filer (BHP) that reports only H1/H2 as Q2/Q4, which sum
    to the fiscal year exactly as four quarters do. The cadence is the LARGEST
    period-set that recurs across >=2 fiscal years, so a single partial year (the
    current in-progress year, or an IPO mid-ramp) is never mistaken for it; with
    too little history to establish one, fall back to quarterly so short-history
    names self-skip below instead of building on a one-off period set.
    """
    by_fy: dict[int, set[str]] = defaultdict(set)
    for y, p in records_i:
        by_fy[y].add(p)
    counts: dict[frozenset[str], int] = defaultdict(int)
    for s in by_fy.values():
        if s:
            counts[frozenset(s)] += 1
    recurring = [s for s, n in counts.items() if n >= 2]
    return tuple(sorted(max(recurring, key=len))) if recurring else default


inc_i, bal_i, cf_i = idx(inc), idx(bal), idx(cf)
pseg_i, gseg_i = idx(prod_seg, True), idx(geo_seg, True)
PERIODS = detect_fy_periods(inc_i)  # ("Q1".."Q4") quarterly · ("Q2","Q4") semi-annual (BHP)
NPERIODS = len(PERIODS)
keys = sorted(inc_i, reverse=True)[:QUARTERS]
keys.reverse()  # oldest -> newest
qlabels = [f"{p} {y}" for (y, p) in keys]
NQ = len(keys)

fy_cols = defaultdict(list)
for pos, (y, p) in enumerate(keys):
    fy_cols[y].append(2 + pos)
_full = sorted(y for y, cs in fy_cols.items() if len(cs) == NPERIODS)
if not _full:
    # Too little FMP history to anchor a forecast — e.g. a name that IPO'd in the
    # last few quarters, for which FMP returns no/partial quarterly statements
    # (FRVO). Emit a clean SKIP like the dcf_applicable=false path below instead
    # of indexing into an empty list.
    _reason = "no quarterly FMP history" if not keys else "no complete fiscal year yet"
    print(f"SKIP\t{T}\t{_reason}\t(insufficient history for a DCF)")
    raise SystemExit(0)
full_fys = [_full[-1]]  # consecutive run ending at the latest full FY (no gap years)
for _y in reversed(_full[:-1]):
    if _y == full_fys[0] - 1:
        full_fys.insert(0, _y)
    else:
        break
full_fys = full_fys[-N_ACTUAL_FY:]
FC_YEARS = list(range(full_fys[-1] + 1, full_fys[-1] + 1 + N_FC))

# current reporting structure = segments present in the latest quarter (drops
# AMZN's pre-2019 legacy segment names that only populate old columns).
_latest_g = gseg_i.get(keys[-1]) or {}
GEO = sorted(
    (s for s, v in _latest_g.items() if isinstance(v, (int, float))), key=lambda s: -_latest_g[s]
)
# Product segments + the whole-company fallback guard (src/dcf/segment_coverage.py).
# The modeled set is the segments present in the LATEST quarter; we fall back to ONE
# whole-company revenue line (income-statement total, anchored to consensus below) when:
#   * <2 segments are reported (no per-segment model to build), OR
#   * the reported segments carry no base-year revenue — FMP sometimes drops segment
#     disclosure for a stretch of years (e.g. LITE leaves FY2025 at zero), OR
#   * the segments cover materially LESS than the base-year income total. This last is
#     the partial-contamination case the bare len/zero checks missed: when FMP's latest
#     quarter drops a big segment (VEEV FY2026 Q4 dropped both R&D segments -> ~45%
#     coverage), `base_revenue_by_segment` silently undercounts and the entire DCF is
#     built on a fraction of the company. Whole-company is always safe on the revenue
#     level (it rebuilds from the COMPLETE income statement at lines below), so we
#     downgrade rather than ship a fair value that's off by the missing-revenue fraction.
_base_fy = full_fys[-1]
# DCF_COVERAGE_FLOOR lets the audit / a before-after check tune the floor; production
# uses the module default (segment_coverage.COVERAGE_FLOOR).
try:
    _cov_floor = float(os.environ.get("DCF_COVERAGE_FLOOR", segment_coverage.COVERAGE_FLOOR))
except ValueError:
    _cov_floor = segment_coverage.COVERAGE_FLOOR
_cov = segment_coverage.resolve_product_segments(
    pseg_i, inc_i, PERIODS, _base_fy, keys[-1], floor=_cov_floor
)
PROD = _cov.prod
SINGLE_SEG = _cov.single_seg
if SINGLE_SEG and _cov.reason and _cov.reason.startswith("coverage"):
    # Loud diagnostic: a contaminated name must NEVER be silently downgraded.
    print(
        f"COVERAGE\t{T}\t{_cov.reason} -> whole-company "
        f"(seg-sum {_cov.seg_base_total / 1e6:,.0f}M vs income {_cov.income_base_total / 1e6:,.0f}M)",
        file=sys.stderr,
    )

# --- Analyst-defined segments (outruns FMP's reported detail) ------------------
# When data/dcf_assumptions/<T>.json carries a valid redesign.analyst_segments
# block, the analyst's split REPLACES the FMP-resolved product segments: the
# builder defines those segments, splits base-year (income-statement) revenue by
# base_pct, and drives per-segment growth from the block. An invalid/absent block
# falls back to the FMP set, logging the reason loudly (never silently half-applied).
_cache_for_seg = REPO / "data" / "dcf_assumptions" / f"{T}.json"
_analyst_raw = None
if _cache_for_seg.exists():
    _cache_json = json.loads(_cache_for_seg.read_text(encoding="utf-8"))
    if isinstance(_cache_json, dict):
        _redesign_block = _cache_json.get("redesign")
        if isinstance(_redesign_block, dict):
            _analyst_raw = _redesign_block.get("analyst_segments")
ANALYST_SEGS = analyst_seg_mod.parse_analyst_segments(_analyst_raw)
if _analyst_raw is not None and not ANALYST_SEGS.valid:
    print(
        f"ANALYST_SEGMENTS\t{T}\tinvalid ({ANALYST_SEGS.reason}) -> FMP segments",
        file=sys.stderr,
    )
if ANALYST_SEGS.valid:
    PROD = ANALYST_SEGS.names
    SINGLE_SEG = False
    print(
        f"ANALYST_SEGMENTS\t{T}\t{len(PROD)} analyst segments override FMP: {', '.join(PROD)}",
        file=sys.stderr,
    )


def m(v):
    return v / 1e6 if isinstance(v, (int, float)) else None


def fy_sum_raw(records_i, field, y):
    tot = 0.0
    for p in PERIODS:
        r = records_i.get((y, p))
        v = (r or {}).get(field)
        if isinstance(v, (int, float)):
            tot += v / 1e6
    return tot


ly = full_fys[-1]
rev_ly = fy_sum_raw(inc_i, "revenue", ly)
if rev_ly <= 0:
    # Base-year revenue missing/zero (broken or empty FMP income data): every ratio
    # below divides by it. SKIP cleanly rather than crash through the ratio block.
    print(f"SKIP\t{T}\tbase-year revenue is zero\t(insufficient data for a DCF)")
    raise SystemExit(0)
ratios_ly = {
    "cogs": fy_sum_raw(inc_i, "costOfRevenue", ly) / rev_ly,
    "rnd": fy_sum_raw(inc_i, "researchAndDevelopmentExpenses", ly) / rev_ly,
    "sga": fy_sum_raw(inc_i, "sellingGeneralAndAdministrativeExpenses", ly) / rev_ly,
    "da": fy_sum_raw(cf_i, "depreciationAndAmortization", ly) / rev_ly,
    "sbc": fy_sum_raw(cf_i, "stockBasedCompensation", ly) / rev_ly,
}
oi_ly = fy_sum_raw(inc_i, "operatingIncome", ly)
other_ly = (
    rev_ly
    - fy_sum_raw(inc_i, "costOfRevenue", ly)
    - fy_sum_raw(inc_i, "researchAndDevelopmentExpenses", ly)
    - fy_sum_raw(inc_i, "sellingGeneralAndAdministrativeExpenses", ly)
    - oi_ly
) / rev_ly


def fade(a, b, n=N_FC):
    return [a + (b - a) * i / (n - 1) for i in range(n)]


# product-segment annual actuals (for growth defaults)
seg_ann = defaultdict(lambda: defaultdict(float))
if ANALYST_SEGS.valid:
    # Analyst split: each year's COMPLETE income-statement revenue apportioned by
    # base_pct. Like whole-company, this rebuilds from the full income statement
    # (never a fraction), and momentum/growth are driven by the block below.
    for y in [full_fys[0] - 1, *full_fys]:
        _rev_y = fy_sum_raw(inc_i, "revenue", y)
        for _name, _pct in ((s.name, s.base_pct) for s in ANALYST_SEGS.segments):
            seg_ann[y][_name] = _pct * _rev_y
elif SINGLE_SEG:
    for y in [full_fys[0] - 1, *full_fys]:
        seg_ann[y]["Total company"] = fy_sum_raw(inc_i, "revenue", y)
else:
    for (y, p), d in pseg_i.items():
        for k, v in (d or {}).items():
            if isinstance(v, (int, float)):
                seg_ann[y][k] += v / 1e6
TAX, EXIT_MULT, TG = _g.tax_rate, 12.0, 0.045

# Default assumptions START AT CONSENSUS (user edits from there). Anchor revenue
# to consensus revenueAvg and the margin path to consensus net income (FMP's ebit
# estimate is unreliable for AMZN: it reports NI > EBIT). Beyond the consensus
# horizon, fade growth to a terminal rate and hold the margin.
est_by_year = {}
for e in est:
    try:
        est_by_year[int(str(e.get("date"))[:4])] = e
    except (TypeError, ValueError):
        pass
cons_rev = {
    y: m(est_by_year[y].get("revenueAvg"))
    for y in FC_YEARS
    if (est_by_year.get(y) or {}).get("revenueAvg")
}
cons_ni = {
    y: m(est_by_year[y].get("netIncomeAvg"))
    for y in FC_YEARS
    if (est_by_year.get(y) or {}).get("netIncomeAvg")
}

# Secondary anchor (estimates-widening): FMP Starter truncates analyst-estimates
# to 10 rows, so the consensus horizon often ends after 2-4 forward years.
# OPTIONALLY extend later years from the persisted yfinance growth consensus
# (data/historical/yfinance/<T>_yf_estimates.json, written by
# execution/fetch_yf_estimates.py — never a live call from this build).
# Provenance discipline: FMP years pass through verbatim, extension years are
# tagged source=yfinance in cons_src and logged as a JSON event — never
# silently blended. Disable with DCF_YF_EXTEND=0.
cons_src = {y: "fmp" for y in set(cons_rev) | set(cons_ni)}
_yf_est_path = REPO / "data" / "historical" / "yfinance" / f"{T}_yf_estimates.json"
if os.environ.get("DCF_YF_EXTEND", "1") != "0" and _yf_est_path.exists():
    from dcf import consensus_extension as _consx

    _yf_growth = _consx.load_yf_growth(_yf_est_path)
    if _yf_growth is not None:
        cons_rev, cons_ni, cons_src = _consx.extend_consensus(
            cons_rev, cons_ni, list(FC_YEARS), _yf_growth
        )
        print(
            json.dumps(_consx.extension_event(T, cons_src, _yf_growth)),
            file=sys.stderr,
        )
cons_years = sorted(cons_rev)
ncons = max(2, len(cons_years))


def _term(s):
    return 0.07 if s in ("Amazon Web Services", "Advertising Services") else 0.05


momentum = {}
for s in PROD:
    g0 = (seg_ann[ly][s] / seg_ann[ly - 1][s] - 1) if seg_ann[ly - 1].get(s) else 0.06
    momentum[s] = fade(max(min(g0, 0.30), -0.05), _term(s))

# revenue: per-segment momentum + a uniform delta each consensus year so the
# segment total lands on consensus; fade to terminal beyond the consensus horizon.
seg_growth = {s: [0.0] * N_FC for s in PROD}
seg_prev = {s: seg_ann[ly][s] for s in PROD}
for j, y in enumerate(FC_YEARS):
    if y in cons_rev:
        mg = {s: momentum[s][j] for s in PROD}
        denom = sum(seg_prev.values())
        delta = (
            (cons_rev[y] - sum(seg_prev[s] * (1 + mg[s]) for s in PROD)) / denom if denom else 0.0
        )
        for s in PROD:
            gg = mg[s] + delta
            seg_growth[s][j] = gg
            seg_prev[s] *= 1 + gg
    else:
        steps = max(1, (N_FC - 1) - (ncons - 1))
        for s in PROD:
            g_start = seg_growth[s][ncons - 1] if ncons else momentum[s][j]
            gg = g_start + (_term(s) - g_start) * ((j - (ncons - 1)) / steps if ncons else 1.0)
            seg_growth[s][j] = gg
            seg_prev[s] *= 1 + gg

# margin: COGS/R&D/SG&A held at last-actual %; Other-opex is the plug set so
# NOPAT (= EBIT x (1-tax)) lands on consensus net income each consensus year.
cogs_pct = [ratios_ly["cogs"]] * N_FC
rnd_pct = [ratios_ly["rnd"]] * N_FC
sga_pct = [ratios_ly["sga"]] * N_FC
other_pct = [0.0] * N_FC
oim_list = []
target_oim = oi_ly / rev_ly
for j, y in enumerate(FC_YEARS):
    if y in cons_ni and y in cons_rev:
        target_oim = (cons_ni[y] / (1 - TAX)) / cons_rev[y]
    oim_list.append(target_oim)
    other_pct[j] = max(
        1 - ratios_ly["cogs"] - ratios_ly["rnd"] - ratios_ly["sga"] - target_oim, 0.02
    )
da_pct = fade(ratios_ly["da"], ratios_ly["da"])
sbc_pct = fade(ratios_ly["sbc"], ratios_ly["sbc"] * 0.6)
# Dashboard SBC % inputs (near-term + terminal), defaulting from the actuals-based
# fade above. Charged as an explicit after-tax expense by the engine (op margins are
# NON-GAAP / SBC-excluded), and the terminal exit-multiple EBITDA is burdened by SBC.
# A per-name _opus override (near/terminal, or a floor for an FMP data gap) is applied
# in the _opus block below.
SBC_NEAR = sbc_pct[0]
SBC_TERM = sbc_pct[-1]
# Capex anchored to Amazon's announced ~$200B 2026 capex (Jassy, Q4'25 call —
# vs Street's ~$147B), then the capex/D&A ratio converges to ~1.05x as the AI
# build-out matures (Damodaran reinvestment normalization).
# 2026 capex: AMZN guided ~$200B; everyone else defaults to last-year capex +10%.
_capex_ly_M = abs(fy_sum_raw(cf_i, "capitalExpenditure", ly)) or (rev_ly * 0.05)
CAPEX_2026_M = 200_000.0 if T == "AMZN" else _capex_ly_M * 1.10
_da_2026 = (cons_rev.get(FC_YEARS[0], rev_ly) * ratios_ly["da"]) or 1.0
cda0 = CAPEX_2026_M / _da_2026
capex_da = fade(cda0, 1.05)
nwc_pct = [0.005] * N_FC

cache = REPO / "data" / "dcf_assumptions" / f"{T}.json"
narr = json.loads(cache.read_text(encoding="utf-8")).get("narrative", "") if cache.exists() else ""
try:
    con = connect_sqlite(str(REPO / "data" / "portfolio.db"), role=SQLiteConnectionRole.READ_ONLY)
    row = con.execute("SELECT live_price FROM dcf_runs WHERE ticker=?", (T,)).fetchone()
    price = (row[0] if row and row[0] else None) or prof.get("price") or 255.0
    con.close()
except Exception:
    price = prof.get("price") or 255.0

# Opus/user per-name assumption block, read once. The dcf_applicable skip + the
# segment/margin/terminal overrides apply further below (after PROD is known); the
# WACC drivers are resolved here because the Monte Carlo base WACC needs them too.
_opus = (
    json.loads(cache.read_text(encoding="utf-8")).get("redesign") if cache.exists() else None
) or {}
_debt_scope_raw = _opus.get("dcf_debt_scope", "interest_bearing_debt_only")
if _debt_scope_raw not in {"interest_bearing_debt_only", "debt_and_lease_obligations"}:
    raise RuntimeError(f"invalid redesign.dcf_debt_scope for {T}: {_debt_scope_raw!r}")
DCF_DEBT_SCOPE = cast("equity_bridge.DebtScope", _debt_scope_raw)
# The immutable Opus baseline (seeded once for passes predating provenance
# tracking) — feeds the Cover "Assumptions by" line and the Assumptions sheet.
_baseline = assumptions_doc.ensure_opus_baseline(cache)
# WACC drivers: a block override wins over the FMP profile beta + textbook rf/ERP/Kd.
# Resolved once and fed to BOTH the Monte Carlo base WACC and the Dashboard yellow
# cells, so an edited beta/ERP survives a from-scratch rebuild (the round-trip is
# closed by refresh_dcf.sync_assumptions_json writing these back to the block).
_beta_override = _opus.get("beta")
BETA = float(_beta_override) if _beta_override is not None else (prof.get("beta") or 1.3)
RF = float(_opus.get("risk_free_rate", _g.risk_free_rate))
ERP = float(_opus.get("equity_risk_premium", _g.equity_risk_premium))
KD = float(_opus.get("cost_of_debt", 0.045))
# Country risk premium: Damodaran's country premiums weighted by where this name
# earns revenue (FMP geo segments). A US/mature name resolves to 0.0 and is left
# exactly where it was; a LatAm/EM name carries the sovereign-risk layer the US
# ERP omits. A per-name override in the block wins (so a hand-set value survives a
# rebuild); else the systematic revenue-weighted value.
_preserved_crp = os.environ.get("DCF_COUNTRY_RISK_OVERRIDE")
if _preserved_crp is not None:
    CRP = float(_preserved_crp)
    _country_risk_context = {
        "event": "dcf_country_risk_context",
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": T,
        "premium": CRP,
        "authority": "preserved_dashboard_override",
        "source_record": None,
    }
elif "country_risk_premium" in _opus:
    CRP = float(_opus["country_risk_premium"])
    _country_risk_context = {
        "event": "dcf_country_risk_context",
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": T,
        "premium": CRP,
        "authority": "owner_override",
        "source_record": None,
    }
else:
    _country_risk_observation = country_risk.country_risk_observation(REPO, T)
    CRP = _country_risk_observation.premium
    _country_risk_context = {
        "event": "dcf_country_risk_context",
        "schema_version": "dcf_country_risk_context.v1",
        "ticker": T,
        "premium": CRP,
        "authority": (
            "systematic_geo"
            if _country_risk_observation.source_record is not None
            else "systematic_default_zero"
        ),
        "source_record": _country_risk_observation.source_record,
    }
print(json.dumps(_country_risk_context, sort_keys=True), file=sys.stderr)

# ----------------------------------------------------------------------------- Monte Carlo
# A reduced-form model (single revenue CAGR, linear margin ramp) calibrated so the
# base case reproduces the full workbook value, then perturbed over Opus-set driver
# distributions. Static snapshot — recomputed each build.
import numpy as np  # noqa: E402

latest = keys[-1]
_bal_latest = bal_i[latest]
# Debt is a governed valuation input, not a normalized-data convenience.  The
# builder may publish only when cash and the approved debt scope resolve from
# exact primary aggregates and every signed lease component. Missing evidence
# is a HOLD; normalized cash/component sums and LT+ST debt are never publishing
# fallbacks because they cannot prove the primary-source perimeter.
_bridge_period_end_raw = _bal_latest.get("date")
_bridge_currency_raw = _bal_latest.get("reportedCurrency")
_bridge_period_end = _bridge_period_end_raw if isinstance(_bridge_period_end_raw, str) else None
_bridge_currency = _bridge_currency_raw if isinstance(_bridge_currency_raw, str) else None
_cash_resolution = (
    equity_bridge.resolve_primary_reported_aggregate(
        _bal_latest,
        aggregate_field="cashAndShortTermInvestments",
        overlay={"statements": PRIMARY_FACT_OVERLAY},
        period_end=_bridge_period_end,
        fiscal_period_type=latest[1],
        currency=_bridge_currency,
    )
    if _bridge_period_end is not None and _bridge_currency is not None
    else None
)
_verified_debt_resolution = (
    equity_bridge.resolve_primary_debt_scope(
        _bal_latest,
        scope=DCF_DEBT_SCOPE,
        overlay={"statements": PRIMARY_FACT_OVERLAY},
        period_end=_bridge_period_end,
        fiscal_period_type=latest[1],
        currency=_bridge_currency,
    )
    if _bridge_period_end is not None and _bridge_currency is not None
    else None
)
if _cash_resolution is None or _verified_debt_resolution is None:
    print(
        json.dumps(
            {
                "event": "dcf_equity_bridge_unavailable",
                "ticker": T,
                "period": f"{latest[1]} {latest[0]}",
                "missing": [
                    name
                    for name, resolution in (
                        ("cash_and_short_term_investments", _cash_resolution),
                        (f"verified_{DCF_DEBT_SCOPE}", _verified_debt_resolution),
                    )
                    if resolution is None
                ],
            },
            sort_keys=True,
        ),
        file=sys.stderr,
    )
    raise RuntimeError("latest balance sheet lacks a verified DCF equity bridge")
cash_now = _cash_resolution.value / 1e6
debt_now = _verified_debt_resolution.value / 1e6
shares_now = m(inc_i[latest].get("weightedAverageShsOutDil"))
if shares_now is None or shares_now <= 0:
    print(
        json.dumps(
            {
                "event": "dcf_equity_bridge_unavailable",
                "ticker": T,
                "period": f"{latest[1]} {latest[0]}",
                "missing": ["positive_diluted_shares"],
            },
            sort_keys=True,
        ),
        file=sys.stderr,
    )
    raise RuntimeError("latest income statement lacks positive diluted shares")
_debt_component_lineage = [
    {**dict(lineage), "operation_sign": sign}
    for lineage, (_field, sign) in zip(
        _verified_debt_resolution.component_lineage,
        _verified_debt_resolution.operations,
        strict=True,
    )
]
print(
    json.dumps(
        {
            "event": "dcf_equity_bridge_context",
            "schema_version": "dcf_equity_bridge_context.v2",
            "ticker": T,
            "period_end": _bridge_period_end,
            "fiscal_period_type": latest[1],
            "reporting_currency": _bridge_currency,
            "cash_m": cash_now,
            "total_debt_m": debt_now,
            "diluted_shares_m": shares_now,
            "cash_basis": _cash_resolution.basis,
            "total_debt_basis": _verified_debt_resolution.debt_basis,
            "debt_scope": _verified_debt_resolution.scope,
            "debt_calculation": _verified_debt_resolution.calculation,
            "debt_operations": [
                {"field": field, "sign": sign}
                for field, sign in _verified_debt_resolution.operations
            ],
            "debt_component_lineage": _debt_component_lineage,
        },
        sort_keys=True,
    ),
    file=sys.stderr,
)
beta = BETA
ke = RF + beta * ERP + CRP
mktcap = price * shares_now
we = mktcap / (mktcap + debt_now) if (mktcap + debt_now) else 1.0
wacc0 = we * ke + (1 - we) * KD * (1 - TAX)


# Dashboard control defaults: per-segment growth collapses to 2 points (near-term
# + terminal); margin to 2 points. The workbook interpolates these, so the Python
# mirror interpolates identically — keeping full_value == the in-sheet value.
g1_def = {s: sum(seg_growth[s][:ncons]) / ncons for s in PROD}  # avg over consensus window
gT_def = {s: seg_growth[s][-1] for s in PROD}
margin_near_def, margin_term_def = oim_list[0], oim_list[-1]

# Analyst-defined segments drive their own near/terminal growth directly (that is
# the whole point of the override — a split FMP does not report, growing on the
# analyst's own path rather than a consensus-reconciled FMP momentum). The engine
# reads these two Dashboard points per segment and fades between them; the margins,
# terminal method and the rest still come from the _opus block below.
if ANALYST_SEGS.valid:
    g1_def = ANALYST_SEGS.near_growth()
    gT_def = ANALYST_SEGS.terminal_growth()

# --- Opus per-name override (if the Opus assumption pass has run for this name) ---
OPUS_BASIS, OPUS_METHOD = "EV/EBITDA", "Exit multiple"
CURRENCY = (inc[0].get("reportedCurrency") if inc else None) or "USD"
_FX_TO_USD = {
    "USD": 1.0,
    "DKK": 0.145,
    "EUR": 1.08,
    "GBP": 1.27,
    "CAD": 0.73,
    "BRL": 0.18,
    "ILS": 0.27,
    "INR": 0.012,
    "JPY": 0.0067,
    "CHF": 1.12,
    "SEK": 0.095,
    "TWD": 0.031,
}
# An unknown reported currency must fail the build, not default to 1.0 — that
# default is how TSM persisted a TWD fair value stamped USD ("97% undervalued")
# into every downstream surface. Failing here keeps the workbook's ×FX formula
# (the single FX source of truth read back by dcf.redesign._read_fx) honest.
if CURRENCY not in _FX_TO_USD:
    print(
        f"FAIL\t{T}\tunknown reported currency {CURRENCY!r} — add its USD rate to "
        "_FX_TO_USD in execution/build_redesigned_dcf.py before building this name",
        file=sys.stderr,
    )
    raise SystemExit(1)
FX = _FX_TO_USD[CURRENCY]
if _opus.get("dcf_applicable") is False:
    print(f"SKIP\t{T}\t{_opus.get('business_model')}\t(FCFF DCF not the right tool)")
    raise SystemExit(0)
if _opus.get("segments"):
    _sg = _opus["segments"]
    # analyst_segments (parsed above) already pinned g1_def/gT_def to the analyst's
    # own per-segment growth; the FMP-named _opus["segments"] growth override does
    # not apply to those analyst names, so only re-growth when NOT analyst-driven.
    if not ANALYST_SEGS.valid:
        g1_def = {s: _sg.get(s, {}).get("near_term_growth", g1_def[s]) for s in PROD}
        gT_def = {s: _sg.get(s, {}).get("terminal_growth", gT_def[s]) for s in PROD}
    margin_near_def = _opus.get("near_term_op_margin", margin_near_def)
    margin_term_def = _opus.get("terminal_op_margin", margin_term_def)
    # SBC %: an explicit near/terminal override wins; a floor guards an FMP data gap
    # (some names — WIX — report 0 SBC for recent quarters, understating the fade).
    SBC_NEAR = float(_opus.get("sbc_pct_near", SBC_NEAR))
    SBC_TERM = float(_opus.get("sbc_pct_terminal", SBC_TERM))
    _sbc_floor = _opus.get("sbc_pct_floor")
    if _sbc_floor is not None:
        SBC_NEAR = max(SBC_NEAR, float(_sbc_floor))
        SBC_TERM = max(SBC_TERM, float(_sbc_floor) * 0.6)
    TAX = _opus.get("tax_rate", _g.tax_rate)
    EXIT_MULT = float(_opus.get("exit_multiple", EXIT_MULT))
    TG = _opus.get("terminal_growth_g", TG)
    OPUS_BASIS = _opus.get("exit_basis", OPUS_BASIS)
    # Keep exit-multiple as the default method (user preference); Opus's perpetuity
    # pick stays available as the cross-check, not the headline.
    narr = _opus.get("narrative") or narr
    if _opus.get("capex_pct_revenue_2026"):
        CAPEX_2026_M = _opus["capex_pct_revenue_2026"] * (
            cons_rev.get(FC_YEARS[0], rev_ly) or rev_ly
        )
    cda0 = CAPEX_2026_M / _da_2026
    capex_da = fade(cda0, float(_opus.get("terminal_capex_da", 1.05)))

# Scalar terminal assumptions (exit multiple, operating margins, SBC %) are
# independent of the segment set, so a bare block override applies even without a
# `segments` sub-block (the guard above requires one). This lets a name carry only a
# terminal re-basing (#837 exit multiple) or a heavy-SBC margin/SBC normalization
# (#838) without fabricating a segments block. A name whose block already set these
# inside the guard is unchanged (same value re-read here); a name with no block keeps
# the code/consensus default.
if isinstance(_opus.get("exit_multiple"), (int, float)):
    EXIT_MULT = float(_opus["exit_multiple"])
if isinstance(_opus.get("near_term_op_margin"), (int, float)):
    margin_near_def = float(_opus["near_term_op_margin"])
if isinstance(_opus.get("terminal_op_margin"), (int, float)):
    margin_term_def = float(_opus["terminal_op_margin"])
# #838 mature SBC normalization: explicit near/terminal SBC % (a mature large-software
# floor, more disciplined than the actuals 0.6x fade that leaves a hyper-grower's SBC
# at ~terminal-margin levels). Applies with or without a segments block.
if isinstance(_opus.get("sbc_pct_near"), (int, float)):
    SBC_NEAR = float(_opus["sbc_pct_near"])
if isinstance(_opus.get("sbc_pct_terminal"), (int, float)):
    SBC_TERM = float(_opus["sbc_pct_terminal"])
if isinstance(_opus.get("sbc_pct_floor"), (int, float)):
    SBC_NEAR = max(SBC_NEAR, float(_opus["sbc_pct_floor"]))
    SBC_TERM = max(SBC_TERM, float(_opus["sbc_pct_floor"]) * 0.6)


# Per-name growth-fade curvature: the convex shape whose revenue path best fits
# Street consensus, weighting near-term years most (dcf.fade_calibration). Runs for
# every name (Opus-override or consensus-default g1/gT); an explicit block override
# wins, and a name with < 2 consensus years keeps the default 2.0.
#
# #840 policy: the consensus fit can pick a linear (p≈1.0) fade for a hyper-grower,
# which keeps growth hot for years and inflates terminal revenue / the pre-SBC base.
# ``calibrate_curvature_with_floor`` floors the effective curvature at 2.0 (convex)
# whenever the revenue-weighted (near−terminal) growth spread exceeds ~8pts — so fast
# decelerators fade front-loaded, while steady names keep their consensus fit.
_cons_by_offset = {y - FC_YEARS[0]: cons_rev[y] for y in cons_rev}
CURV = float(
    _opus.get("growth_fade_curvature")
    or fade_calibration.calibrate_curvature_with_floor(
        {s: seg_ann[ly][s] for s in PROD}, g1_def, gT_def, _cons_by_offset, N_FC
    )
)


def _scen_deltas(raw, seed):
    """Bull/Bear scenario offsets: the block's `scenario_bull`/`scenario_bear`
    override (mirrored back by refresh_dcf.sync_assumptions_json, so user edits
    survive a from-scratch rebuild) over the documented seed defaults."""
    if not isinstance(raw, dict):
        return seed
    return redesign_mod.ScenarioDeltas(
        growth_near=float(raw.get("growth_near", seed.growth_near)),
        growth_term=float(raw.get("growth_term", seed.growth_term)),
        margin_near=float(raw.get("margin_near", seed.margin_near)),
        margin_term=float(raw.get("margin_term", seed.margin_term)),
        exit_multiple=float(raw.get("exit_multiple", seed.exit_multiple)),
        terminal_g=float(raw.get("terminal_g", seed.terminal_g)),
    )


# Thesis-calibrated bear override (Monthly Red Team Phase 1 guard 3 + PR8): the
# holdings JSON's ``bear_deltas`` (when the analyst has named one) seeds the Bear
# column whenever the mirrored ``scenario_bear`` block is absent OR still equals
# the untouched generic BEAR_SEED — a never-edited mirror is the labeled fallback
# ``sync_assumptions_json`` wrote back, not an owner edit. A seed-differing
# mirror (a real owner edit) still wins unconditionally
# (``redesign.resolve_mirrored_bear``).
_holdings_path = REPO / "micro_thesis" / "holdings" / f"{T.upper()}.json"
_holdings_raw: dict | None = None
if _holdings_path.exists():
    try:
        _hd = json.loads(_holdings_path.read_text(encoding="utf-8"))
        _holdings_raw = _hd if isinstance(_hd, dict) else None
    except (OSError, ValueError):
        _holdings_raw = None

BULL_D = _scen_deltas(_opus.get("scenario_bull"), redesign_mod.BULL_SEED)
BEAR_D = redesign_mod.resolve_mirrored_bear(
    _scen_deltas(_opus.get("scenario_bear"), redesign_mod.BEAR_SEED), _holdings_raw
)


def _weights(raw, default):
    """Scenario probability weights (Bull/Base/Bear): the block's `scenario_prior`
    override (LLM-set, mirrored back by refresh_dcf.sync_assumptions_json so owner
    edits survive a from-scratch rebuild) over the symmetric default, normalized to
    sum 1."""
    d_bull, d_base, d_bear = default
    if not isinstance(raw, dict):
        return d_bull, d_base, d_bear
    b = float(raw.get("bull_weight", d_bull))
    m = float(raw.get("base_weight", d_base))
    r = float(raw.get("bear_weight", d_bear))
    s = b + m + r
    if s <= 0 or b < 0 or m < 0 or r < 0:
        return d_bull, d_base, d_bear
    return b / s, m / s, r / s


WEIGHTS = _weights(
    _opus.get("scenario_prior"),
    (
        redesign_mod.DEFAULT_SCENARIO_WEIGHTS["bull"],
        redesign_mod.DEFAULT_SCENARIO_WEIGHTS["base"],
        redesign_mod.DEFAULT_SCENARIO_WEIGHTS["bear"],
    ),
)


def _seg_g(s, j):  # convex near->terminal fade (curvature = consensus-fit CURV)
    frac = ((N_FC - 1 - j) / (N_FC - 1)) ** CURV
    return gT_def[s] + (g1_def[s] - gT_def[s]) * frac


def _oim(j):  # ramp to terminal by the end of the consensus window, then hold
    return margin_near_def + (margin_term_def - margin_near_def) * min(1.0, j / (ncons - 1))


def _sbc_pct(j):  # linear near->terminal fade of the SBC % of revenue
    return SBC_NEAR + (SBC_TERM - SBC_NEAR) * j / (N_FC - 1)


def _project():
    seg = {s: seg_ann[ly][s] for s in PROD}
    prev = sum(seg.values())
    rev, oi, da, vf = [], [], [], []
    for j in range(N_FC):
        for s in PROD:
            seg[s] *= 1 + _seg_g(s, j)
        rr = sum(seg.values())
        ebit = rr * _oim(j)
        nop = ebit * (1 - TAX)
        d, cx, nw = rr * da_pct[j], rr * da_pct[j] * capex_da[j], (rr - prev) * nwc_pct[j]
        # Explicit after-tax SBC charge (op margin is NON-GAAP / SBC-excluded).
        sbc_at = _sbc_pct(j) * rr * (1 - TAX)
        rev.append(rr)
        oi.append(ebit)
        da.append(d)
        vf.append(nop + d - cx - nw - sbc_at)
        prev = rr
    return rev, oi, da, vf


rev_p, oi_p, da_p, vf_p = _project()
# mirror the workbook's SELECTED terminal (Opus may pick perpetuity), then FX -> USD
# EV/EBITDA is BURDENED by terminal-year SBC so the exit multiple applies to real
# (SBC-charged) EBITDA — consistent with charging SBC as an operating expense.
_sbc_term_M = _sbc_pct(N_FC - 1) * rev_p[-1]
_tmetric = {
    "EV/EBITDA": oi_p[-1] - _sbc_term_M + da_p[-1],
    "EV/Sales": rev_p[-1],
    "EV/EBIT": oi_p[-1],
    "EV/FCF": vf_p[-1],
}.get(OPUS_BASIS, oi_p[-1] + da_p[-1])
_tv_exit = _tmetric * EXIT_MULT
_tv_perp = vf_p[-1] * (1 + TG) / (wacc0 - TG) if wacc0 > TG else _tv_exit
_tv = _tv_perp if OPUS_METHOD == "Perpetuity" else _tv_exit
full_value = (
    (
        sum(vf_p[t] / (1 + wacc0) ** (t + 1) for t in range(N_FC))
        + _tv / (1 + wacc0) ** N_FC
        + cash_now
        - debt_now
    )
    / shares_now
    * FX
)

ann_rev = [fy_sum_raw(inc_i, "revenue", y) for y in full_fys]
ann_g = [ann_rev[i] / ann_rev[i - 1] - 1 for i in range(1, len(ann_rev)) if ann_rev[i - 1]]
ann_oim = []
for _y in full_fys:
    _rev = fy_sum_raw(inc_i, "revenue", _y)
    if _rev:
        ann_oim.append(fy_sum_raw(inc_i, "operatingIncome", _y) / _rev)
cagr0 = (rev_p[-1] / rev_ly) ** (1 / N_FC) - 1
mT0 = oi_p[-1] / rev_p[-1] if rev_p[-1] else margin_term_def
SIG = {
    "cagr": float(min(max(np.std(ann_g) * 0.5, 0.012), 0.03)),
    "margin": float(min(max(np.std(ann_oim), 0.015), 0.03)),
    "wacc": 0.005,
    "exit": 1.5,
}


def _reduced(cagrs, mTs, waccs, exits):
    t = np.arange(1, N_FC + 1)
    R = rev_ly * (1 + cagrs[:, None]) ** t[None, :]
    mm = ann_oim[-1] + (mTs[:, None] - ann_oim[-1]) * t[None, :] / N_FC
    ebit = R * mm
    nop = ebit * (1 - TAX)
    d = R * ratios_ly["da"]
    cda = cda0 + (1.05 - cda0) * t / N_FC
    nw = (R - rev_ly * (1 + cagrs[:, None]) ** (t[None, :] - 1)) * 0.005
    vf = nop + d - d * cda[None, :] - nw
    pv = (vf / (1 + waccs[:, None]) ** t[None, :]).sum(axis=1)
    return (
        pv + (ebit[:, -1] + d[:, -1]) * exits / (1 + waccs) ** N_FC + cash_now - debt_now
    ) / shares_now


base_red = float(
    _reduced(np.array([cagr0]), np.array([mT0]), np.array([wacc0]), np.array([EXIT_MULT]))[0]
)
# Thin-data names (recent IPOs like CGEH/FIGR) can drive a NaN through the model:
# a sparse/negative projected revenue base under a fractional-power CAGR, an
# empty-history ratio, etc. An all-NaN mc_vals (which crashes np.histogram below)
# implies base_red itself is NaN — and a non-finite base case means the headline
# value AND the Monte Carlo calibration constant are garbage, so the whole workbook
# is meaningless. SKIP cleanly like the insufficient-history paths above rather
# than emit a "nan" RESULT.
if not (np.isfinite(base_red) and np.isfinite(full_value)):
    print(f"SKIP\t{T}\tnon-finite base valuation\t(insufficient data for a reliable DCF)")
    raise SystemExit(0)
kcal = full_value / base_red if base_red else 1.0
rng = np.random.default_rng(42)
NMC = 10000
mc_vals = (
    _reduced(
        rng.normal(cagr0, SIG["cagr"], NMC),
        rng.normal(mT0, SIG["margin"], NMC),
        np.clip(rng.normal(wacc0, SIG["wacc"], NMC), 0.05, 0.16),
        np.clip(rng.normal(EXIT_MULT, SIG["exit"], NMC), 5, 24),
    )
    * kcal
)
PCTS = [5, 10, 25, 50, 75, 90, 95]
# Drop any non-finite draws before aggregating: a wide driver distribution can push
# an individual trial out of the model's domain (NaN/inf), and np.histogram raises
# "autodetected range of [nan, nan] is not finite" on an all-NaN array (and
# .min()/.max()/percentile raise on an empty one).
mc_vals = mc_vals[np.isfinite(mc_vals)]
if mc_vals.size:
    _hc, _he = np.histogram(mc_vals, bins=18)
    mc_res = {
        "mean": float(mc_vals.mean()),
        "median": float(np.median(mc_vals)),
        "std": float(mc_vals.std()),
        "min": float(mc_vals.min()),
        "max": float(mc_vals.max()),
        "pcts": {p: float(np.percentile(mc_vals, p)) for p in PCTS},
        "p_under": float((mc_vals > price).mean()),
        "p_up20": float((mc_vals > price * 1.2).mean()),
    }
    mc_hist = [(float(_he[i]), float(_he[i + 1]), int(_hc[i])) for i in range(len(_hc))]
else:
    # No finite draws survived (degenerate inputs). The in-sheet Monte Carlo is
    # live-formula based and still works; only this unused Python snapshot is
    # skipped, so the workbook still builds.
    mc_res, mc_hist = {}, []

# ----------------------------------------------------------------------------- Dashboard
# Single control surface: the ~handful of cells that move the answer live here at
# fixed addresses; Model/WACC/Valuation/MC reference them. Per-segment growth is
# collapsed to 2 inputs (near-term + terminal), interpolated by formula.
SEG_ROW0 = 20  # first segment row on the Dashboard
DB = {
    "g1": lambda i: f"Dashboard!$B${SEG_ROW0 + i}",
    "gT": lambda i: f"Dashboard!$C${SEG_ROW0 + i}",
    "ref": lambda i: f"Dashboard!$D${SEG_ROW0 + i}",
    "margin_near": "Dashboard!$B$29",
    "margin_term": "Dashboard!$B$30",
    "tax": "Dashboard!$B$31",
    "capex26": "Dashboard!$B$34",
    "capex_term_da": "Dashboard!$B$35",
    "rf": "Dashboard!$B$38",
    "erp": "Dashboard!$B$39",
    "beta": "Dashboard!$B$40",
    "kd": "Dashboard!$B$41",
    "crp": "Dashboard!$B$47",
    "curv": "Dashboard!$B$49",
    "sbc_near": f"Dashboard!$B${redesign_mod._DB_SBC_NEAR}",
    "sbc_term": f"Dashboard!$B${redesign_mod._DB_SBC_TERM}",
    "method": "Dashboard!$B$43",
    "basis": "Dashboard!$B$44",
    "mult": "Dashboard!$B$45",
    "tg": "Dashboard!$B$46",
    "price": "Dashboard!$B$48",
}

# ----------------------------------------------------------------------------- workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ===== Financials (rich quarterly history + segments + ratios) =====
fs = wb.create_sheet("Financials")
fs.sheet_view.showGridLines = False
fs.column_dimensions["A"].width = 34
for i in range(NQ):
    fs.column_dimensions[get_column_letter(2 + i)].width = 10
put(fs, 1, 1, "Line Item ($M)", bold=True)
for i, lab in enumerate(qlabels):
    put(fs, 1, 2 + i, lab, bold=True).alignment = RIGHT
fs.freeze_panes = "B2"

frow = 2
fin_row = {}
pseg_row = {}


def col(i):
    return get_column_letter(2 + i)


def write_qrow(label, getter, *, kind="act", fmt=USD):
    global frow
    put(fs, frow, 1, label)
    for i, key in enumerate(keys):
        v = getter(i, key)
        if v is not None:
            put(fs, frow, 2 + i, v, fmt=fmt, kind=kind)
    frow += 1
    return frow - 1


def write_yoy(target_row):
    global frow
    put(fs, frow, 1, "    % YoY")
    for i in range(NQ):
        if i >= NPERIODS:  # one fiscal year back = NPERIODS columns (4 quarterly, 2 semi-annual)
            put(
                fs,
                frow,
                2 + i,
                ie(f"{col(i)}{target_row}/{col(i - NPERIODS)}{target_row}-1"),
                fmt=PCT,
            )
    frow += 1


def write_pct(target_row, base_row, label="    % of revenue"):
    global frow
    put(fs, frow, 1, label)
    for i in range(NQ):
        put(fs, frow, 2 + i, ie(f"{col(i)}{target_row}/{col(i)}{base_row}"), fmt=PCT)
    frow += 1


band(fs, frow, "INCOME STATEMENT", NQ + 1)
frow += 1
rev_r = write_qrow("Revenue", lambda i, k: m(inc_i.get(k, {}).get("revenue")))
fin_row["Revenue"] = rev_r
write_yoy(rev_r)
for lab, fld in [
    ("Cost of Revenue", "costOfRevenue"),
    ("R&D Expense", "researchAndDevelopmentExpenses"),
    ("SG&A Expense", "sellingGeneralAndAdministrativeExpenses"),
]:
    rr = write_qrow(lab, lambda i, k, f=fld: m(inc_i.get(k, {}).get(f)))
    fin_row[lab] = rr
    write_pct(rr, rev_r)
gp_r = write_qrow("Gross Profit", lambda i, k: m(inc_i.get(k, {}).get("grossProfit")))
write_pct(gp_r, rev_r, "    gross margin")
oi_r = write_qrow("Operating Income", lambda i, k: m(inc_i.get(k, {}).get("operatingIncome")))
fin_row["Operating Income"] = oi_r
write_pct(oi_r, rev_r, "    operating margin")
ni_r = write_qrow("Net Income", lambda i, k: m(inc_i.get(k, {}).get("netIncome")))
fin_row["Net Income"] = ni_r
write_pct(ni_r, rev_r, "    net margin")
fin_row["Diluted Shares (M)"] = write_qrow(
    "Diluted Shares (M)", lambda i, k: m(inc_i.get(k, {}).get("weightedAverageShsOutDil"))
)

if ANALYST_SEGS.valid:
    # Analyst-defined segments: FMP has no per-quarter data for these names, so each
    # segment row is base_pct x that quarter's actual revenue — written as HARDCODED
    # actuals (blue), NOT a formula, so the offline reader (redesign.py._find_row/
    # _fy_sum, which reads cell VALUES) picks up base-year revenue. The base-FY sum
    # = base_pct x FY revenue — exactly the split the analyst set.
    band(fs, frow, "REVENUE BY SEGMENT — ANALYST-DEFINED (% YoY, % of revenue)", NQ + 1)
    frow += 1
    _pct_by_name = {s.name: s.base_pct for s in ANALYST_SEGS.segments}
    for s in PROD:
        pct = _pct_by_name[s]
        rr = write_qrow(
            s, lambda i, k, p=pct: rv * p if (rv := m(inc_i.get(k, {}).get("revenue"))) else None
        )
        pseg_row[s] = rr
        write_yoy(rr)
        write_pct(rr, rev_r)
elif not SINGLE_SEG:
    band(fs, frow, "REVENUE BY SEGMENT — PRODUCT (% YoY, % of revenue)", NQ + 1)
    frow += 1
    for s in PROD:
        rr = write_qrow(s, lambda i, k, seg=s: m((pseg_i.get(k) or {}).get(seg)))
        pseg_row[s] = rr
        write_yoy(rr)
        write_pct(rr, rev_r)

if GEO:
    band(fs, frow, "REVENUE BY SEGMENT — GEOGRAPHY (% YoY, % of revenue)", NQ + 1)
    frow += 1
    for s in GEO:
        rr = write_qrow(s, lambda i, k, seg=s: m((gseg_i.get(k) or {}).get(seg)))
        write_yoy(rr)
        write_pct(rr, rev_r)

band(fs, frow, "BALANCE SHEET", NQ + 1)
frow += 1


def _bs_getter(field: str):
    """Keep workbook cash/debt rows identical to the bridge used by the builder."""
    aggregate_components = {
        "cashAndShortTermInvestments": ("cashAndCashEquivalents", "shortTermInvestments"),
        "totalDebt": ("longTermDebt", "shortTermDebt"),
    }
    component_fields = aggregate_components.get(field)
    if component_fields is None:
        return lambda i, k, f=field: m(bal_i.get(k, {}).get(f))

    def _aggregate(i, k):
        b = bal_i.get(k, {})
        resolved = equity_bridge.resolve_complete_aggregate(
            b,
            aggregate_field=field,
            component_fields=component_fields,
        )
        return m(resolved.value) if resolved is not None else None

    return _aggregate


for lab, fld in [
    ("Cash & ST Investments", "cashAndShortTermInvestments"),
    ("Total Current Assets", "totalCurrentAssets"),
    ("PP&E (net)", "propertyPlantEquipmentNet"),
    ("Total Assets", "totalAssets"),
    ("Total Current Liabilities", "totalCurrentLiabilities"),
    # TOTAL debt (long- + short-term), not longTermDebt: convertible notes and
    # near-maturity paper sit in shortTermDebt (LITE FQ3'26: $3,251M of $3,314M),
    # so longTermDebt-only understates net debt and overstates equity value.
    ("Total Debt", "totalDebt"),
    ("Total Equity", "totalStockholdersEquity"),
]:
    fin_row[lab] = write_qrow(lab, _bs_getter(fld))

band(fs, frow, "CASH FLOW", NQ + 1)
frow += 1
for lab, fld in [
    ("D&A", "depreciationAndAmortization"),
    ("Stock-based Compensation", "stockBasedCompensation"),
    ("Change in Working Capital", "changeInWorkingCapital"),
    ("Operating Cash Flow", "operatingCashFlow"),
    ("Capex", "capitalExpenditure"),
    ("Free Cash Flow", "freeCashFlow"),
]:
    fin_row[lab] = write_qrow(lab, lambda i, k, f=fld: m(cf_i.get(k, {}).get(f)))
put(
    fs, frow + 1, 1, "Blue = hardcoded actuals from FMP filings. Same-sheet ratios are formulas."
).font = SUB
LAST = col(NQ - 1)  # latest quarter column


def fref(label, y):
    cs = sorted(fy_cols[y])
    r = fin_row[label]
    return f"SUM(Financials!{get_column_letter(cs[0])}{r}:{get_column_letter(cs[-1])}{r})"


def pseg_fy(seg, y):
    if SINGLE_SEG:
        return fref("Revenue", y)  # single revenue line = total company
    cs = sorted(fy_cols[y])
    r = pseg_row[seg]
    return f"SUM(Financials!{get_column_letter(cs[0])}{r}:{get_column_letter(cs[-1])}{r})"


# ===== WACC calculator =====
wc = wb.create_sheet("WACC")
wc.sheet_view.showGridLines = False
wc.column_dimensions["A"].width = 34
wc.column_dimensions["B"].width = 14
put(wc, 1, 1, "WACC Calculator", bold=True).font = TITLE
# Cost of equity = rf + beta*erp + country risk premium (Damodaran additive
# form). The CRP row sits between beta and the CAPM line, so cost of equity is
# B2+B4*B3+B5; every downstream row index shifts by one vs the pre-CRP layout.
wrows = [
    ("Risk-free rate (10Y)", f"={DB['rf']}", PCT, "f"),
    ("Equity risk premium", f"={DB['erp']}", PCT, "f"),
    ("Beta (levered)", f"={DB['beta']}", NUM3, "f"),
    ("Country risk premium", f"={DB['crp']}", PCT, "f"),
    ("Cost of equity (CAPM)", "=B2+B4*B3+B5", PCT, "f"),
    ("Pre-tax cost of debt", f"={DB['kd']}", PCT, "f"),
    ("Tax rate", f"={DB['tax']}", PCT, "f"),
    ("After-tax cost of debt", "=B7*(1-B8)", PCT, "f"),
    (
        "Market cap (E)",
        f"={DB['price']}*Financials!{LAST}{fin_row['Diluted Shares (M)']}",
        USD,
        "f",
    ),
    ("Total debt (D)", f"=Financials!{LAST}{fin_row['Total Debt']}", USD, "f"),
    ("Equity weight", "=B10/(B10+B11)", PCT, "f"),
    ("Debt weight", "=B11/(B10+B11)", PCT, "f"),
]
for i, (lab, v, fmt, kind) in enumerate(wrows):
    put(wc, 2 + i, 1, lab)
    put(wc, 2 + i, 2, v, fmt=fmt, kind=kind)
WACC_ROW = 2 + len(wrows) + 1
put(wc, WACC_ROW, 1, "WACC", bold=True)
put(wc, WACC_ROW, 2, "=B12*B6+B13*B9", fmt=PCT, bold=True)
put(
    wc,
    WACC_ROW + 2,
    1,
    "Yellow = edit. CAPM cost of equity (incl. country risk premium); "
    "market-value weights. Feeds Valuation!WACC.",
).font = SUB

# ===== Model =====
md = wb.create_sheet("Model")
md.sheet_view.showGridLines = False
md.column_dimensions["A"].width = 34
ALL = full_fys + FC_YEARS
yc = {y: 2 + i for i, y in enumerate(ALL)}
fcj = {y: i for i, y in enumerate(FC_YEARS)}


def mc(y):
    return get_column_letter(yc[y])


put(md, 1, 1, "$M unless noted", bold=True)
for y in ALL:
    c = put(md, 1, yc[y], f"FY{y}{'A' if y in full_fys else 'E'}", bold=True)
    c.alignment = RIGHT
    c.fill = ACTH if y in full_fys else FCH
    md.column_dimensions[mc(y)].width = 10
md.freeze_panes = "B2"

r = 3
band(
    md,
    r,
    "REVENUE BY SEGMENT (yellow = growth assumption; actuals link to Financials)",
    len(ALL) + 1,
)
r += 1
seg_rev_row = {}
for i, s in enumerate(PROD):
    seg_rev_row[s] = r
    put(md, r, 1, s)
    for y in full_fys:
        put(md, r, yc[y], f"={pseg_fy(s, y)}", fmt=USD)  # green link
    for y in FC_YEARS:
        put(md, r, yc[y], f"={mc(y - 1)}{r}*(1+{mc(y)}{r + 1})", fmt=USD)
    r += 1
    put(md, r, 1, "    growth % (convex fade; curvature from Dashboard)")
    for y in full_fys[1:]:
        put(md, r, yc[y], ie(f"{mc(y)}{r - 1}/{mc(y - 1)}{r - 1}-1"), fmt=PCT)
    for y in FC_YEARS:  # convex near->terminal fade, curvature = Dashboard cell (green link)
        put(
            md,
            r,
            yc[y],
            f"={DB['gT'](i)}+({DB['g1'](i)}-{DB['gT'](i)})"
            f"*(({N_FC - 1}-{fcj[y]})/{N_FC - 1})^{DB['curv']}",
            fmt=PCT,
        )
    r += 1
rev_row = r
put(md, r, 1, "Total revenue", bold=True)
for y in ALL:
    cells = "+".join(f"{mc(y)}{seg_rev_row[s]}" for s in PROD)
    put(md, r, yc[y], f"={cells}", fmt=USD, bold=True)
r += 1
put(md, r, 1, "    growth %")
for y in ALL[1:]:
    put(md, r, yc[y], ie(f"{mc(y)}{rev_row}/{mc(y - 1)}{rev_row}-1"), fmt=PCT)
r += 2

band(md, r, "COST STRUCTURE (yellow = % of revenue)", len(ALL) + 1)
r += 1
cost_rows = {}
for lab, fin_lab, ser in [
    ("Cost of revenue", "Cost of Revenue", cogs_pct),
    ("R&D / Technology", "R&D Expense", rnd_pct),
    ("SG&A", "SG&A Expense", sga_pct),
]:
    cost_rows[lab] = r
    put(md, r, 1, lab)
    for y in full_fys:
        put(md, r, yc[y], f"={fref(fin_lab, y)}", fmt=USD)
    for y in FC_YEARS:
        put(md, r, yc[y], f"={mc(y)}{rev_row}*{mc(y)}{r + 1}", fmt=USD)
    r += 1
    put(md, r, 1, "    % of revenue")
    for y in full_fys:
        put(md, r, yc[y], ie(f"{mc(y)}{r - 1}/{mc(y)}{rev_row}"), fmt=PCT)
    for y in FC_YEARS:
        put(md, r, yc[y], ser[fcj[y]], fmt=PCT, kind="in")
    r += 1
other_row = r
put(md, r, 1, "Other opex (fulfilment, etc.)")
for y in full_fys:
    parts = "+".join(f"{mc(y)}{cost_rows[c]}" for c in cost_rows)
    put(md, r, yc[y], f"={mc(y)}{rev_row}-({parts})-{fref('Operating Income', y)}", fmt=USD)
for y in FC_YEARS:  # plug so operating margin = Dashboard near->terminal ramp (green link)
    j = fcj[y]
    parts = "+".join(f"{mc(y)}{cost_rows[c]}" for c in cost_rows)
    ramp = f"({DB['margin_near']}+({DB['margin_term']}-{DB['margin_near']})*MIN(1,{j}/{ncons - 1}))"
    put(md, r, yc[y], f"={mc(y)}{rev_row}*(1-{ramp})-({parts})", fmt=USD)
r += 1
put(md, r, 1, "    % of revenue")
for y in ALL:
    put(md, r, yc[y], ie(f"{mc(y)}{other_row}/{mc(y)}{rev_row}"), fmt=PCT)
r += 2
oi_row = r
put(md, r, 1, "Operating income (EBIT)", bold=True)
for y in ALL:
    parts = "+".join(f"{mc(y)}{cost_rows[c]}" for c in cost_rows)
    put(md, r, yc[y], f"={mc(y)}{rev_row}-({parts})-{mc(y)}{other_row}", fmt=USD, bold=True)
r += 1
put(md, r, 1, "    operating margin")
for y in ALL:
    put(md, r, yc[y], ie(f"{mc(y)}{oi_row}/{mc(y)}{rev_row}"), fmt=PCT)
r += 2

band(md, r, "FCF BRIDGE (NOPAT + D&A + SBC - dNWC - Capex)", len(ALL) + 1)
r += 1
nopat_row = r
put(md, r, 1, "NOPAT = EBIT x (1 - tax)")
for y in ALL:
    put(md, r, yc[y], f"={mc(y)}{oi_row}*(1-Valuation!$B$5)", fmt=USD)
r += 1
da_row = r
put(md, r, 1, "+ D&A")
for y in full_fys:
    put(md, r, yc[y], f"={fref('D&A', y)}", fmt=USD)
for y in FC_YEARS:
    put(md, r, yc[y], f"={mc(y)}{rev_row}*{da_pct[fcj[y]]:.4f}", fmt=USD)
r += 1
sbc_row = r
put(md, r, 1, "+ SBC (non-cash)")
for y in full_fys:
    put(md, r, yc[y], f"={fref('Stock-based Compensation', y)}", fmt=USD)
for y in FC_YEARS:  # SBC% = Dashboard near->terminal linear fade (green link)
    j = fcj[y]
    sbc_ramp = f"({DB['sbc_near']}+({DB['sbc_term']}-{DB['sbc_near']})*{j}/{N_FC - 1})"
    put(md, r, yc[y], f"={mc(y)}{rev_row}*{sbc_ramp}", fmt=USD)
r += 1
nwc_row = r
put(md, r, 1, "- dNWC")
for y in FC_YEARS:
    put(md, r, yc[y], f"=({mc(y)}{rev_row}-{mc(y - 1)}{rev_row})*{nwc_pct[fcj[y]]:.4f}", fmt=USD)
r += 1
capex_row = r
put(md, r, 1, "- Capex")
for y in full_fys:
    put(md, r, yc[y], f"=-{fref('Capex', y)}", fmt=USD)
for y in FC_YEARS:
    put(md, r, yc[y], f"={mc(y)}{da_row}*{mc(y)}{r + 1}", fmt=USD)
r += 1
put(md, r, 1, "    Capex / D&A  (from Dashboard: $cap'26 -> terminal)")
for y in full_fys:
    put(md, r, yc[y], ie(f"{mc(y)}{capex_row}/{mc(y)}{da_row}"), fmt=MULT)
_da26 = f"{mc(FC_YEARS[0])}{da_row}"  # D&A in the first forecast year
for y in FC_YEARS:  # 2026 ratio = $200B / D&A; fades to the Dashboard terminal ratio
    j = fcj[y]
    r0 = f"({DB['capex26']}/{_da26})"
    put(md, r, yc[y], f"={r0}+({DB['capex_term_da']}-{r0})*{j}/{N_FC - 1}", fmt=MULT)
r += 2
fcff_row = r
put(md, r, 1, "FCFF (firm)", bold=True)
for y in ALL:
    put(
        md,
        r,
        yc[y],
        f"={mc(y)}{nopat_row}+{mc(y)}{da_row}+{mc(y)}{sbc_row}-{mc(y)}{nwc_row}-{mc(y)}{capex_row}",
        fmt=USD,
        bold=True,
    )
r += 1
valfcf_row = r
put(md, r, 1, "Valuation FCF (NOPAT + D&A - dNWC - Capex - after-tax SBC)", bold=True)
for y in ALL:
    # Charge SBC as a real after-tax expense: back out the FCFF add-back (- SBC),
    # then charge the after-tax cash cost (- SBC*(1-tax)). Net vs FCFF: -SBC*(2-tax).
    # The op margin is NON-GAAP (SBC-excluded), so this restores GAAP-equivalent FCF.
    put(
        md,
        r,
        yc[y],
        f"={mc(y)}{fcff_row}-{mc(y)}{sbc_row}-{mc(y)}{sbc_row}*(1-Valuation!$B$5)",
        fmt=USD,
        bold=True,
    )
r += 2

# --- Returns on capital & efficiency (level + incremental) ---
band(
    md,
    r,
    "RETURNS ON CAPITAL & EFFICIENCY  (is new revenue/capital earning its keep?)",
    len(ALL) + 1,
)
r += 1
EQ_F, DB_F, CA_F = (
    fin_row["Total Equity"],
    fin_row["Total Debt"],
    fin_row["Cash & ST Investments"],
)


def q4(y):
    return get_column_letter(sorted(fy_cols[y])[-1])


ic_row = r
put(md, r, 1, "Invested capital (debt + equity - cash)")
for y in full_fys:
    put(
        md,
        r,
        yc[y],
        f"=Financials!{q4(y)}{EQ_F}+Financials!{q4(y)}{DB_F}-Financials!{q4(y)}{CA_F}",
        fmt=USD,
    )
for y in FC_YEARS:
    put(
        md,
        r,
        yc[y],
        f"={mc(y - 1)}{ic_row}+{mc(y)}{capex_row}-{mc(y)}{da_row}+{mc(y)}{nwc_row}",
        fmt=USD,
    )
r += 1
eq_row = r
put(md, r, 1, "Book equity (year-end)")
for y in full_fys:
    put(md, r, yc[y], f"=Financials!{q4(y)}{EQ_F}", fmt=USD)
for y in FC_YEARS:
    put(md, r, yc[y], f"={mc(y - 1)}{eq_row}+{mc(y)}{nopat_row}", fmt=USD)
r += 1
roic_row = r
put(md, r, 1, "  ROIC (NOPAT / beg. invested capital)", bold=True)
for y in ALL:
    if (y - 1) in yc:
        put(md, r, yc[y], ie(f"{mc(y)}{nopat_row}/{mc(y - 1)}{ic_row}"), fmt=PCT)
r += 1
put(md, r, 1, "  ROE (earnings / beg. equity)", bold=True)
for y in ALL:
    if (y - 1) in yc:
        earn = fref("Net Income", y) if y in full_fys else f"{mc(y)}{nopat_row}"
        put(md, r, yc[y], ie(f"{earn}/{mc(y - 1)}{eq_row}"), fmt=PCT)
r += 1
put(md, r, 1, "  ROIC - WACC spread (value created)")
for y in ALL:
    if (y - 1) in yc:
        put(md, r, yc[y], ie(f"{mc(y)}{roic_row}-WACC!$B${WACC_ROW}"), fmt=PCT)
r += 1
put(md, r, 1, "  Incremental op margin (ΔEBIT / ΔRev)")
for y in ALL:
    if (y - 1) in yc:
        put(
            md,
            r,
            yc[y],
            ie(f"({mc(y)}{oi_row}-{mc(y - 1)}{oi_row})/({mc(y)}{rev_row}-{mc(y - 1)}{rev_row})"),
            fmt=PCT,
        )
r += 1
put(md, r, 1, "  ROIIC (ΔNOPAT / ΔInvested capital)", bold=True)
for y in ALL:
    if (y - 1) in yc:
        put(
            md,
            r,
            yc[y],
            ie(
                f"({mc(y)}{nopat_row}-{mc(y - 1)}{nopat_row})/({mc(y)}{ic_row}-{mc(y - 1)}{ic_row})"
            ),
            fmt=PCT,
        )
r += 1
put(md, r, 1, "  Sales-to-capital (ΔRev / ΔInvested capital)")
for y in ALL:
    if (y - 1) in yc:
        put(
            md,
            r,
            yc[y],
            ie(f"({mc(y)}{rev_row}-{mc(y - 1)}{rev_row})/({mc(y)}{ic_row}-{mc(y - 1)}{ic_row})"),
            fmt='0.00"x"',
        )
r += 1
put(md, r, 1, "  Reinvestment rate (ΔInvested capital / NOPAT)")
for y in ALL:
    if (y - 1) in yc:
        put(md, r, yc[y], ie(f"({mc(y)}{ic_row}-{mc(y - 1)}{ic_row})/{mc(y)}{nopat_row}"), fmt=PCT)
put(md, r + 2, 1, "Blue/green = actual  ·  Yellow = assumption  ·  Black = formula").font = SUB

# ===== Consensus =====
cs = wb.create_sheet("Consensus")
cs.sheet_view.showGridLines = False
cs.column_dimensions["A"].width = 34
est_by_year = {}
for e in est:
    try:
        est_by_year[int(str(e.get("date"))[:4])] = e
    except (TypeError, ValueError):
        pass
CYEARS = [y for y in FC_YEARS if y in est_by_year][:6]
put(cs, 1, 1, "Full consensus check — model vs Street (FMP estimates, $M)", bold=True).font = TITLE
put(cs, 2, 1, "Fiscal year", bold=True)
for j, y in enumerate(CYEARS):
    put(cs, 2, 2 + j, y, bold=True).alignment = RIGHT
    cs.column_dimensions[get_column_letter(2 + j)].width = 12

DIFMT = "+0.0%;(0.0%)"
shares_ref = f"Financials!{LAST}{fin_row['Diluted Shares (M)']}"
metrics = [
    ("Revenue", "revenueAvg", lambda y: f"=Model!{mc(y)}{rev_row}", USD),
    ("EBITDA*", "ebitdaAvg", lambda y: f"=Model!{mc(y)}{oi_row}+Model!{mc(y)}{da_row}", USD),
    ("EBIT*", "ebitAvg", lambda y: f"=Model!{mc(y)}{oi_row}", USD),
    ("SG&A", "sgaExpenseAvg", lambda y: f"=Model!{mc(y)}{cost_rows['SG&A']}", USD),
    ("Net income (NOPAT)", "netIncomeAvg", lambda y: f"=Model!{mc(y)}{nopat_row}", USD),
    ("EPS", "epsAvg", lambda y: f"=Model!{mc(y)}{nopat_row}/{shares_ref}", PXS),
]
band(cs, 3, "EVERY AVAILABLE FMP CONSENSUS FIELD (consensus / model / % diff)", len(CYEARS) + 1)
r = 4
for lab, fld, mexpr, fmt in metrics:
    put(cs, r, 1, f"{lab} — consensus", bold=True)
    for j, y in enumerate(CYEARS):
        v = m(est_by_year[y].get(fld)) if fmt != PXS else est_by_year[y].get(fld)
        if isinstance(v, (int, float)):
            put(cs, r, 2 + j, v, fmt=fmt, kind="act")
    crow = r
    r += 1
    put(cs, r, 1, "  model")
    for j, y in enumerate(CYEARS):
        put(cs, r, 2 + j, mexpr(y), fmt=fmt)
    r += 1
    put(cs, r, 1, "  vs consensus")
    for j in range(len(CYEARS)):
        cl = get_column_letter(2 + j)
        put(cs, r, 2 + j, ie(f"{cl}{r - 1}/{cl}{crow}-1"), fmt=DIFMT)
    r += 2
put(
    cs,
    r,
    1,
    "Defaults anchored to consensus revenue + net income (≈0% diff). Other rows show where the model sits vs Street.",
).font = SUB
put(
    cs,
    r + 1,
    1,
    "*FMP's EBITDA/EBIT estimates look unreliable for AMZN (it reports NI > EBIT) — don't over-read those two.",
).font = SUB

# ===== Valuation =====
vs = wb.create_sheet("Valuation")
vs.sheet_view.showGridLines = False
vs.column_dimensions["A"].width = 36
vs.column_dimensions["B"].width = 14
for i in range(N_FC):
    vs.column_dimensions[get_column_letter(3 + i)].width = 9
put(vs, 1, 1, f"{NAME} — DCF Valuation", bold=True).font = TITLE
vin = [  # all controlled from the Dashboard (green links); edit there, not here
    ("Current price", f"={DB['price']}", PXS, "f"),
    ("WACC", f"=WACC!B{WACC_ROW}", PCT, "f"),
    ("Terminal growth (g)", f"={DB['tg']}", PCT, "f"),
    ("Tax rate", f"={DB['tax']}", PCT, "f"),
    ("Terminal method", f"={DB['method']}", None, "f"),
    ("Exit basis", f"={DB['basis']}", None, "f"),
    ("Exit multiple", f"={DB['mult']}", MULT, "f"),
]
for i, (lab, v, fmt, kind) in enumerate(vin):
    put(vs, 2 + i, 1, lab)
    put(vs, 2 + i, 2, v, fmt=fmt, kind=kind)
B = {
    "price": "$B$2",
    "wacc": "$B$3",
    "g": "$B$4",
    "tax": "$B$5",
    "method": "$B$6",
    "basis": "$B$7",
    "mult": "$B$8",
}

w0 = 11
put(vs, w0, 1, "Forecast year", bold=True)
for j, y in enumerate(FC_YEARS):
    put(vs, w0, 3 + j, y, bold=True).alignment = RIGHT


def vc(j):
    return get_column_letter(3 + j)


put(vs, w0 + 1, 1, "FCFF ($M)")
for j, y in enumerate(FC_YEARS):
    put(vs, w0 + 1, 3 + j, f"=Model!{mc(y)}{valfcf_row}", fmt=USD)
put(vs, w0 + 2, 1, "PV factor")
for j in range(N_FC):
    put(vs, w0 + 2, 3 + j, f"=1/(1+{B['wacc']})^{j + 1}", fmt=NUM3)
put(vs, w0 + 3, 1, "PV of FCFF ($M)")
for j in range(N_FC):
    put(vs, w0 + 3, 3 + j, f"={vc(j)}{w0 + 1}*{vc(j)}{w0 + 2}", fmt=USD)
sumpv = w0 + 5
put(vs, sumpv, 1, "Sum PV of FCFF", bold=True)
put(vs, sumpv, 2, f"=SUM({vc(0)}{w0 + 3}:{vc(N_FC - 1)}{w0 + 3})", fmt=USD, bold=True)

ty = mc(FC_YEARS[-1])
tr = sumpv + 2
band(vs, tr, "TERMINAL VALUE (exit multiple = default; perpetuity = cross-check)", 3)
put(vs, tr + 1, 1, "Terminal-year FCFF")
put(vs, tr + 1, 2, f"={vc(N_FC - 1)}{w0 + 1}", fmt=USD)
put(vs, tr + 2, 1, "Terminal revenue")
put(vs, tr + 2, 2, f"=Model!{ty}{rev_row}", fmt=USD)
put(vs, tr + 3, 1, "Terminal EBIT")
put(vs, tr + 3, 2, f"=Model!{ty}{oi_row}", fmt=USD)
put(vs, tr + 4, 1, "Terminal EBITDA (SBC-burdened)")
# Burden EBITDA by terminal SBC so the exit multiple applies to REAL, SBC-charged
# EBITDA — consistent with charging SBC as an operating expense in the FCF stream.
put(vs, tr + 4, 2, f"=Model!{ty}{oi_row}+Model!{ty}{da_row}-Model!{ty}{sbc_row}", fmt=USD)
put(vs, tr + 5, 1, "Terminal metric (per exit basis)")
put(
    vs,
    tr + 5,
    2,
    f'=IF({B["basis"]}="EV/EBITDA",B{tr + 4},IF({B["basis"]}="EV/Sales",B{tr + 2},IF({B["basis"]}="EV/EBIT",B{tr + 3},B{tr + 1})))',
    fmt=USD,
)
put(vs, tr + 6, 1, "TV (exit) = metric x multiple")
put(vs, tr + 6, 2, f"=B{tr + 5}*{B['mult']}", fmt=USD)
put(vs, tr + 7, 1, "PV of TV — exit multiple")
put(vs, tr + 7, 2, f"=B{tr + 6}/(1+{B['wacc']})^{N_FC}", fmt=USD)
put(vs, tr + 8, 1, "TV (perpetuity) = FCFF(1+g)/(WACC-g)")
put(vs, tr + 8, 2, f"=B{tr + 1}*(1+{B['g']})/({B['wacc']}-{B['g']})", fmt=USD)
put(vs, tr + 9, 1, "PV of TV — perpetuity")
put(vs, tr + 9, 2, f"=B{tr + 8}/(1+{B['wacc']})^{N_FC}", fmt=USD)
put(vs, tr + 10, 1, "PV of TV — SELECTED", bold=True)
put(vs, tr + 10, 2, f'=IF({B["method"]}="Perpetuity",B{tr + 9},B{tr + 7})', fmt=USD, bold=True)

br = tr + 12
band(vs, br, "EQUITY BRIDGE -> VALUE / SHARE", 3)
put(vs, br + 1, 1, "Operating value (enterprise value)", bold=True)
put(vs, br + 1, 2, f"=B{sumpv}+B{tr + 10}", fmt=USD, bold=True)
put(vs, br + 2, 1, "+ Cash & ST investments")
put(vs, br + 2, 2, f"=Financials!{LAST}{fin_row['Cash & ST Investments']}", fmt=USD)
put(vs, br + 3, 1, "- Long-term debt")
put(vs, br + 3, 2, f"=-Financials!{LAST}{fin_row['Total Debt']}", fmt=USD)
put(vs, br + 4, 1, "Equity value", bold=True)
put(vs, br + 4, 2, f"=B{br + 1}+B{br + 2}+B{br + 3}", fmt=USD, bold=True)
put(vs, br + 5, 1, "Diluted shares (M)")
put(vs, br + 5, 2, f"=Financials!{LAST}{fin_row['Diluted Shares (M)']}", fmt=USD)
vps_row = br + 6
put(vs, vps_row, 1, "VALUE / SHARE", bold=True)
put(
    vs, vps_row, 2, f"=B{br + 4}/B{br + 5}*{FX}", fmt=PXS
).font = BIG  # *FX -> USD (1.0 for USD names)

ck = vps_row + 2
band(vs, ck, "CROSS-CHECKS & RETURN", 3)
put(vs, ck + 1, 1, "Terminal weight (% of EV)")
put(vs, ck + 1, 2, f"=B{tr + 10}/B{br + 1}", fmt=PCT)
put(vs, ck + 2, 1, "Implied perpetuity g (from exit multiple)")
put(vs, ck + 2, 2, f"=(B{tr + 6}*{B['wacc']}-B{tr + 1})/(B{tr + 6}+B{tr + 1})", fmt=PCT)
put(vs, ck + 3, 1, "Implied exit EV/EBITDA (from perpetuity)")
put(vs, ck + 3, 2, f"=B{tr + 8}/B{tr + 4}", fmt=MULT)
put(vs, ck + 4, 1, "Upside / (downside) to fair value")
put(vs, ck + 4, 2, f"=B{vps_row}/{B['price']}-1", fmt="+0%;(0%)")
irr = ck + 6
put(vs, irr, 1, "Owner cashflow / share")
put(vs, irr, 2, f"=-{B['price']}", fmt=PXS)
for j, y in enumerate(FC_YEARS):
    extra = f"+B{vps_row}" if j == N_FC - 1 else ""
    put(vs, irr, 3 + j, f"={vc(j)}{w0 + 1}/B{br + 5}{extra}", fmt="0.00")
put(vs, irr + 1, 1, "IRR (buy at price, hold 10y)", bold=True)
put(vs, irr + 1, 2, f"=IRR(B{irr}:{vc(N_FC - 1)}{irr})", fmt=PCT, bold=True)

# ===== Monte Carlo (LIVE — recalculates in-sheet on every edit) =====
from openpyxl.chart import BarChart, Reference  # noqa: E402

NTRIAL = 1000
HARR = "{1;2;3;4;5;6;7;8;9;10}"  # year vector for the in-cell SUMPRODUCT
mcs = wb.create_sheet("Monte Carlo")
mcs.sheet_view.showGridLines = False
mcs.column_dimensions["A"].width = 30
for ltr in "BCD":
    mcs.column_dimensions[ltr].width = 13
put(mcs, 1, 1, "Monte Carlo simulation (live)", bold=True).font = TITLE
put(
    mcs,
    2,
    1,
    f"{NTRIAL:,} trials, recomputed in-sheet on every edit. Edit the yellow driver distributions to re-roll. Engine grid is to the right (cols H+).",
).font = SUB

band(mcs, 4, "DRIVER DISTRIBUTIONS (yellow = edit; Opus-set defaults)", 4)
for j, h in enumerate(["Driver", "Dist", "Mean", "Std dev"]):
    put(mcs, 5, 1 + j, h, bold=True)
_cagr_link = f"=(Model!{mc(FC_YEARS[-1])}{rev_row}/Model!{mc(full_fys[-1])}{rev_row})^(1/{N_FC})-1"
for i, (lab, mu, sd, fmt) in enumerate(
    [
        ("Revenue 10y CAGR", _cagr_link, SIG["cagr"], PCT),
        ("Terminal operating margin", f"={DB['margin_term']}", SIG["margin"], PCT),
        ("WACC", f"=WACC!B{WACC_ROW}", SIG["wacc"], PCT),
        ("Exit EV/EBITDA multiple", f"={DB['mult']}", SIG["exit"], MULT),
    ]
):
    put(mcs, 6 + i, 1, lab)
    put(mcs, 6 + i, 2, "Normal")
    put(mcs, 6 + i, 3, mu, fmt=fmt, kind="f")  # mean links to model/Dashboard (auto-tracks)
    put(mcs, 6 + i, 4, sd, fmt=fmt, kind="in")  # std stays editable

VR = "$O$2:$O$" + str(NTRIAL + 1)
PR = "Valuation!$B$2"
band(mcs, 11, "SIMULATION OUTPUT — value / share (live)", 4)
for i, (lab, f) in enumerate(
    [
        ("Mean", f"=AVERAGE({VR})"),
        ("Median (P50)", f"=MEDIAN({VR})"),
        ("Std dev", f"=STDEV({VR})"),
        ("Min", f"=MIN({VR})"),
        ("Max", f"=MAX({VR})"),
        ("Current price", f"={PR}"),
        ("Base case (full model)", f"=Valuation!B{vps_row}"),
    ]
):
    put(mcs, 12 + i, 1, lab)
    put(mcs, 12 + i, 2, f, fmt=PXS, bold=(lab == "Median (P50)"))
MINC, MAXC = "$B$15", "$B$16"
MC_PUNDER_ROW = 20
put(mcs, 20, 1, "P(undervalued at current price)", bold=True)
put(mcs, 20, 2, f'=COUNTIF({VR},">"&{PR})/{NTRIAL}', fmt=PCT, bold=True)
put(mcs, 21, 1, "P(>20% upside)")
put(mcs, 21, 2, f'=COUNTIF({VR},">"&{PR}*1.2)/{NTRIAL}', fmt=PCT)
band(mcs, 23, "PERCENTILES (live)", 4)
for i, p in enumerate(PCTS):
    put(mcs, 24 + i, 1, f"P{p}")
    put(mcs, 24 + i, 2, f"=PERCENTILE({VR},{p / 100})", fmt=PXS)
NB, hb = 15, 32
band(mcs, hb, "VALUE DISTRIBUTION (live histogram)", 4)
put(mcs, hb + 1, 1, "Bin midpoint ($/sh)", bold=True)
put(mcs, hb + 1, 2, "Count", bold=True)
for i in range(NB):
    lo = f"({MINC}+{i}*({MAXC}-{MINC})/{NB})"
    hi = f"({MINC}+{i + 1}*({MAXC}-{MINC})/{NB})"
    put(mcs, hb + 2 + i, 1, f"={MINC}+{i + 0.5}*({MAXC}-{MINC})/{NB}", fmt='"$"#,##0')
    put(
        mcs,
        hb + 2 + i,
        2,
        f'=COUNTIFS({VR},">="&{lo},{VR},"<"&{hi})' if i < NB - 1 else f'=COUNTIF({VR},">="&{lo})',
    )
chart = BarChart()
chart.type, chart.legend = "col", None
chart.title = f"Value / share — {NTRIAL:,} trials"
chart.height, chart.width = 8.5, 15
chart.add_data(
    Reference(mcs, min_col=2, min_row=hb + 2, max_row=hb + 1 + NB), titles_from_data=False
)
chart.set_categories(Reference(mcs, min_col=1, min_row=hb + 2, max_row=hb + 1 + NB))
mcs.add_chart(chart, "A50")

# --- engine: constants (H:I) + trial grid (K:O), referenced by the dashboard ---
mcs.column_dimensions["G"].width = 3
mcs.column_dimensions["H"].width = 20
put(mcs, 1, 8, "SIMULATION ENGINE — leave as-is", bold=True)
for i, (lab, v) in enumerate(
    [
        ("R0 (base revenue)", f"=Model!{mc(full_fys[-1])}{rev_row}"),
        ("m0 (base op margin)", f"=Model!{mc(full_fys[-1])}{oi_row + 1}"),
        ("da (D&A / revenue)", ratios_ly["da"]),
        ("cda0 (2026 capex/D&A)", cda0),
        ("nwc (% incr revenue)", 0.005),
        ("tax", "=Valuation!$B$5"),
        ("k (calibration)", kcal),
        ("cash", f"=Financials!{LAST}{fin_row['Cash & ST Investments']}"),
        ("debt", f"=Financials!{LAST}{fin_row['Total Debt']}"),
        ("shares", f"=Financials!{LAST}{fin_row['Diluted Shares (M)']}"),
    ]
):
    put(mcs, 2 + i, 8, lab)
    put(mcs, 2 + i, 9, v, fmt=(NUM3 if isinstance(v, float) else None))
for j, h in enumerate(["cagr", "term margin", "wacc", "exit", "value/share"]):
    put(mcs, 1, 11 + j, h, bold=True)


def val_formula(rs):
    return (
        "=$I$8*(SUMPRODUCT(($I$2*(1+K"
        + rs
        + ")^"
        + HARR
        + ")*(($I$3+(L"
        + rs
        + "-$I$3)*"
        + HARR
        + "/10)*(1-$I$7)+$I$4-$I$4*($I$5+(1.05-$I$5)*"
        + HARR
        + "/10)-(K"
        + rs
        + "/(1+K"
        + rs
        + "))*$I$6)/(1+M"
        + rs
        + ")^"
        + HARR
        + ")+$I$2*(1+K"
        + rs
        + ")^10*(L"
        + rs
        + "+$I$4)*N"
        + rs
        + "/(1+M"
        + rs
        + ")^10+$I$9-$I$10)/$I$11"
    )


for rr in range(2, NTRIAL + 2):
    rs = str(rr)
    put(mcs, rr, 11, "=MAX(-0.05,NORMINV(RAND(),$C$6,$D$6))", fmt=PCT)
    put(mcs, rr, 12, "=MAX(0.02,NORMINV(RAND(),$C$7,$D$7))", fmt=PCT)
    put(mcs, rr, 13, "=MIN(0.16,MAX(0.05,NORMINV(RAND(),$C$8,$D$8)))", fmt=PCT)
    put(mcs, rr, 14, "=MIN(24,MAX(5,NORMINV(RAND(),$C$9,$D$9)))", fmt=MULT)
    put(mcs, rr, 15, val_formula(rs), fmt=PXS)
# hide the simulation engine (constants H:I + 1,000-row trial grid K:O)
for ltr in "HIJKLMNO":
    mcs.column_dimensions[ltr].hidden = True

# ===== Dashboard (the front door — all primary controls + sanity checks) =====
from openpyxl.formatting.rule import FormulaRule  # noqa: E402

dsh = wb.create_sheet("Dashboard")
dsh.sheet_view.showGridLines = False
for ltr, w in [("A", 30), ("B", 13), ("C", 13), ("D", 14), ("E", 46)]:
    dsh.column_dimensions[ltr].width = w
put(dsh, 1, 1, f"{NAME} ({T}) — Control Dashboard", bold=True).font = TITLE
put(
    dsh,
    2,
    1,
    "Edit the yellow cells; everything recomputes. Orange = a driver you've moved off consensus.",
).font = SUB

band(dsh, 4, "VERDICT", 5)
put(dsh, 5, 1, "Fair value / share", bold=True)
put(dsh, 5, 2, f"=Valuation!B{vps_row}", fmt=PXS).font = BIG
put(dsh, 6, 1, "Current price")
put(dsh, 6, 2, f"={DB['price']}", fmt=PXS)
put(dsh, 7, 1, "Upside / (downside)")
put(dsh, 7, 2, f"=Valuation!B{ck + 4}", fmt="+0%;(0%)")
put(dsh, 8, 1, "Verdict", bold=True)
put(dsh, 8, 2, f'=IF(Valuation!B{vps_row}>{DB["price"]},"Undervalued","Overvalued")', bold=True)
put(dsh, 9, 1, "Monte Carlo P(undervalued)")
put(dsh, 9, 2, f"='Monte Carlo'!B{MC_PUNDER_ROW}", fmt=PCT)

band(dsh, 11, "GUARDRAIL SIGNALS (are my assumptions sane?)", 5)
for j, h in enumerate(["Signal", "Value", "Flag", "Healthy"]):
    put(dsh, 12, 1 + j, h, bold=True)
roic26 = f"Model!{mc(FC_YEARS[0])}{roic_row}"
# Beginning invested capital for the 2026 ROIC (the prior FY's IC cell). When it
# is <= 0 the ROIC ratio (NOPAT / IC) is mathematically meaningless, not a value
# signal — negative-book-equity cannon-ballers (BKNG, AZO, DPZ, MCD, HD) run
# invested capital negative after cumulative buybacks, which flips ROIC hugely
# negative and would false-flag "! destroys" on some of the market's highest
# cash-return businesses. The guardrail flag renders "n/a" for those names
# instead; the underlying ROIC row is left untouched (this is interpretation,
# not model math).
ic_beg26 = f"Model!{mc(FC_YEARS[0] - 1)}{ic_row}"
tw, ig, pu = f"Valuation!B{ck + 1}", f"Valuation!B{ck + 2}", f"'Monte Carlo'!B{MC_PUNDER_ROW}"
guards = [
    ("Terminal weight (% of EV)", f"={tw}", PCT, f'=IF({tw}>0.9,"! high","ok")', "< 90%"),
    (
        "Implied g vs exit multiple",
        f"={ig}",
        PCT,
        f'=IF({ig}>{DB["tg"]}+0.02,"! rich","ok")',
        "near term g",
    ),
    (
        "ROIC - WACC (2026)",
        f'=IF({ic_beg26}<=0,"n/a (neg. capital)",{roic26}-WACC!B{WACC_ROW})',
        PCT,
        f'=IF({ic_beg26}<=0,"n/a",IF({roic26}-WACC!B{WACC_ROW}<0,"! destroys","creates"))',
        "> 0",
    ),
    (
        "Revenue vs consensus (2026)",
        "=Consensus!B6",
        PCT,
        '=IF(ABS(Consensus!B6)>0.05,"! off-Street","on")',
        "+/- 5%",
    ),
    ("MC P(undervalued)", f"={pu}", PCT, f'=IF({pu}<0.5,"! coin-flip","ok")', "context"),
]
for i, (lab, vf, fmt, flag, hr) in enumerate(guards):
    put(dsh, 13 + i, 1, lab)
    put(dsh, 13 + i, 2, vf, fmt=fmt)
    put(dsh, 13 + i, 3, flag, bold=True)
    put(dsh, 13 + i, 4, hr).font = SUB

band(dsh, 18, "REVENUE DRIVERS — near-term & terminal growth per segment", 5)
for j, h in enumerate(["Segment", "Near-term", "Terminal", "Consensus", "Thesis claim"]):
    put(dsh, 19, 1 + j, h, bold=True)
seg_notes = {
    "Amazon Web Services": "AI/cloud reaccel + capacity adds",
    "Advertising Services": "ad load + retail media",
    "Online Stores": "share gains, GMV growth",
    "Third-Party Seller Services": "3P mix + fee/ad take",
    "Subscription Services": "Prime price + member adds",
    "Physical Stores": "grocery/format expansion",
    "Other Services": "misc / emerging",
}
for i, s in enumerate(PROD):
    rr = SEG_ROW0 + i
    put(dsh, rr, 1, s)
    put(dsh, rr, 2, g1_def[s], fmt=PCT, kind="in")
    put(dsh, rr, 3, gT_def[s], fmt=PCT, kind="in")
    put(dsh, rr, 4, g1_def[s], fmt=PCT, kind="act")
    put(dsh, rr, 5, seg_notes.get(s, "")).font = SUB

band(dsh, 28, "PROFITABILITY & TAX", 5)
put(dsh, 29, 1, "Near-term operating margin")
put(dsh, 29, 2, margin_near_def, fmt=PCT, kind="in")
put(dsh, 29, 4, margin_near_def, fmt=PCT, kind="act")
put(dsh, 29, 5, "consensus-implied; ramps to terminal").font = SUB
put(dsh, 30, 1, "Terminal operating margin")
put(dsh, 30, 2, margin_term_def, fmt=PCT, kind="in")
put(dsh, 30, 5, "AI/ads mix + operating leverage").font = SUB
put(dsh, 31, 1, "Tax rate")
put(dsh, 31, 2, TAX, fmt=PCT, kind="in")

band(dsh, 33, "CAPITAL INTENSITY", 5)
put(dsh, 34, 1, "2026 capex ($M)")
put(dsh, 34, 2, CAPEX_2026_M, fmt=USD, kind="in")
put(dsh, 34, 4, CAPEX_2026_M, fmt=USD, kind="act")
put(
    dsh,
    34,
    5,
    "Amazon guided ~$200B (AI/datacenters)"
    if T == "AMZN"
    else "~last-year capex +10%; set to guidance",
).font = SUB
put(dsh, 35, 1, "Terminal capex / D&A")
put(dsh, 35, 2, 1.05, fmt=MULT, kind="in")
put(dsh, 35, 5, "converges to ~1.0x as build-out matures").font = SUB

band(dsh, 37, "DISCOUNT RATE & TERMINAL", 5)
put(dsh, 38, 1, "Risk-free rate (10Y)")
put(dsh, 38, 2, RF, fmt=PCT, kind="in")
put(dsh, 39, 1, "Equity risk premium")
put(dsh, 39, 2, ERP, fmt=PCT, kind="in")
put(dsh, 40, 1, "Beta (levered)")
put(dsh, 40, 2, round(BETA, 3), fmt=NUM3, kind="in")
put(dsh, 40, 5, "FMP raw beta; Damodaran bottom-up ~1.2").font = SUB
put(dsh, 41, 1, "Pre-tax cost of debt")
put(dsh, 41, 2, KD, fmt=PCT, kind="in")
put(dsh, 42, 1, "WACC (computed)", bold=True)
put(dsh, 42, 2, f"=WACC!B{WACC_ROW}", fmt=PCT, bold=True)
put(dsh, 43, 1, "Terminal method")
put(dsh, 43, 2, OPUS_METHOD, kind="in")
put(dsh, 44, 1, "Exit basis")
put(dsh, 44, 2, OPUS_BASIS, kind="in")
put(dsh, 45, 1, "Exit multiple")
put(dsh, 45, 2, EXIT_MULT, fmt=MULT, kind="in")
put(dsh, 46, 1, "Terminal growth (g)")
put(dsh, 46, 2, TG, fmt=PCT, kind="in")
put(dsh, 47, 1, "Country risk premium")
put(dsh, 47, 2, CRP, fmt=PCT, kind="in")
put(
    dsh,
    47,
    5,
    f"Damodaran country premiums, revenue-weighted ({country_risk.COUNTRY_CRP_AS_OF}); 0 for US/mature",
).font = SUB
put(dsh, 48, 1, "Current price")
put(dsh, 48, 2, price, fmt=PXS, kind="in")
put(dsh, 49, 1, "Growth fade curvature")
put(dsh, 49, 2, CURV, fmt=NUM3, kind="in")
put(
    dsh,
    49,
    5,
    "convexity of the near->terminal growth fade, fit to consensus (1.0=linear); "
    "higher = faster early deceleration",
).font = SUB

# --- SCENARIOS: Bull/Bear as user-editable DELTAS vs the Base yellow cells ---
# Rows/cols are the redesign-module contract (SCEN_ROW_*); the fair-value row is
# Python-computed (static) — written below via write_scenario_fair_values and
# rewritten by every refresh from the then-current inputs.
band(dsh, 50, "SCENARIOS — Bull / Bear offsets vs Base (edit the yellow Δs)", 5)
for j, h in enumerate(["Scenario lever", "Base", "Bull Δ", "Bear Δ"]):
    put(dsh, 51, 1 + j, h, bold=True)
R = redesign_mod
for row, lab, base_ref, fmt, bull_v, bear_v in [
    (
        R.SCEN_ROW_GROWTH_NEAR,
        "Segment growth, near-term",
        "per segment",
        PCT,
        BULL_D.growth_near,
        BEAR_D.growth_near,
    ),
    (
        R.SCEN_ROW_GROWTH_TERM,
        "Segment growth, terminal",
        "per segment",
        PCT,
        BULL_D.growth_term,
        BEAR_D.growth_term,
    ),
    (
        R.SCEN_ROW_MARGIN_NEAR,
        "Operating margin, near-term",
        "=B29",
        PCT,
        BULL_D.margin_near,
        BEAR_D.margin_near,
    ),
    (
        R.SCEN_ROW_MARGIN_TERM,
        "Operating margin, terminal",
        "=B30",
        PCT,
        BULL_D.margin_term,
        BEAR_D.margin_term,
    ),
    (
        R.SCEN_ROW_EXIT_MULT,
        "Exit multiple",
        "=B45",
        MULT,
        BULL_D.exit_multiple,
        BEAR_D.exit_multiple,
    ),
    (R.SCEN_ROW_TG, "Terminal growth (g)", "=B46", PCT, BULL_D.terminal_g, BEAR_D.terminal_g),
]:
    put(dsh, row, 1, lab)
    delta_fmt = f"+{fmt};-{fmt}"  # signed: a Δ of +0.02 renders "+2.0%"
    if base_ref.startswith("="):
        put(dsh, row, 2, base_ref, fmt=fmt)
    else:
        put(dsh, row, 2, base_ref).font = SUB
    put(dsh, row, 3, bull_v, fmt=delta_fmt, kind="in")
    put(dsh, row, 4, bear_v, fmt=delta_fmt, kind="in")
put(dsh, R.SCEN_FV_ROW, 1, "Fair value / share — Base · Bull · Bear", bold=True)
put(dsh, R.SCEN_FV_ROW + 1, 1, "Upside vs current price")
for j, cl in enumerate("BCD"):
    put(dsh, R.SCEN_FV_ROW + 1, 2 + j, ie(f"{cl}{R.SCEN_FV_ROW}/$B$48-1"), fmt="+0%;(0%)")
put(
    dsh,
    R.SCEN_FV_ROW + 2,
    1,
    "Bull/Bear = Base shifted by the Δ columns (uniform across segments). Row 58 + the "
    "Sensitivity sheet are Python-computed — run refresh_dcf after editing.",
).font = SUB

# --- SCENARIO WEIGHTS: probability mass on Base/Bull/Bear (yellow, owner-editable) ---
# Aligned under the row-51 Base/Bull/Bear column headers; seeded from the LLM
# scenario_prior (or the symmetric default). dcf.scenario_reward consumes these;
# refresh_dcf captures + re-injects them so an owner edit always wins.
_w_bull, _w_base, _w_bear = WEIGHTS
put(
    dsh,
    R.SCEN_WEIGHTS_LABEL_ROW,
    1,
    "Scenario probability weights (edit — must sum to 1.0; LLM-set, owner overrides):",
).font = SUB
put(dsh, R.SCEN_WEIGHTS_ROW, 1, "Probability weight")
put(dsh, R.SCEN_WEIGHTS_ROW, R.SCEN_COL_WEIGHT_BASE, _w_base, fmt="0%", kind="in")
put(dsh, R.SCEN_WEIGHTS_ROW, R.SCEN_COL_BULL, _w_bull, fmt="0%", kind="in")
put(dsh, R.SCEN_WEIGHTS_ROW, R.SCEN_COL_BEAR, _w_bear, fmt="0%", kind="in")

# --- STOCK-BASED COMPENSATION: explicit after-tax charge (yellow, owner-editable) ---
# Sits BELOW the SCENARIOS block (rows 64-66) so no existing Dashboard address
# shifts. Operating margins stay NON-GAAP (SBC-excluded, comparable across names);
# the engine charges SBC*(1-tax) as an explicit expense and burdens the terminal
# exit-multiple EBITDA by SBC. Defaults from the actuals-based sbc_pct fade.
band(dsh, redesign_mod._DB_SBC_BAND, "STOCK-BASED COMPENSATION (charged after-tax)", 5)
put(dsh, redesign_mod._DB_SBC_NEAR, 1, "SBC % of revenue — near-term")
put(dsh, redesign_mod._DB_SBC_NEAR, 2, SBC_NEAR, fmt=PCT, kind="in")
put(
    dsh,
    redesign_mod._DB_SBC_NEAR,
    5,
    "charged as an after-tax expense; op margin is non-GAAP (SBC-excluded)",
).font = SUB
put(dsh, redesign_mod._DB_SBC_TERM, 1, "SBC % of revenue — terminal")
put(dsh, redesign_mod._DB_SBC_TERM, 2, SBC_TERM, fmt=PCT, kind="in")
put(
    dsh,
    redesign_mod._DB_SBC_TERM,
    5,
    "fades near->terminal; also burdens the exit-multiple EBITDA",
).font = SUB

ddm = DataValidation(type="list", formula1='"Perpetuity,Exit multiple"', allow_blank=False)
ddb = DataValidation(type="list", formula1='"EV/EBITDA,EV/Sales,EV/EBIT,EV/FCF"', allow_blank=False)
dsh.add_data_validation(ddm)
dsh.add_data_validation(ddb)
ddm.add("B43")
ddb.add("B44")
ORANGE = PatternFill("solid", fgColor="FFD9A0")
dsh.conditional_formatting.add(
    f"B{SEG_ROW0}:B{SEG_ROW0 + len(PROD) - 1}",
    FormulaRule(formula=[f"$B{SEG_ROW0}<>$D{SEG_ROW0}"], fill=ORANGE),
)
dsh.conditional_formatting.add("B29", FormulaRule(formula=["$B29<>$D29"], fill=ORANGE))
dsh.conditional_formatting.add("B34", FormulaRule(formula=["$B34<>$D34"], fill=ORANGE))

# ===== Color Code =====
cc = wb.create_sheet("Color Code")
cc.sheet_view.showGridLines = False
cc.column_dimensions["A"].width = 16
cc.column_dimensions["B"].width = 74
put(cc, 1, 1, "Color", bold=True)
put(cc, 1, 2, "Meaning", bold=True)
put(cc, 2, 1, "Black")
put(cc, 2, 2, "Formula-driven (same-sheet)")
put(cc, 3, 1, "Blue").font = BLUE
put(cc, 3, 2, "Hard-coded actuals (from FMP filings)")
yc2 = put(cc, 4, 1, "Yellow")
yc2.fill = YEL
yc2.font = BLUE
put(cc, 4, 2, "An assumption you can change — edit these to drive the model")
put(cc, 5, 1, "Green").font = GREEN
put(cc, 5, 2, "A formula that links to another sheet")
oc = put(cc, 6, 1, "Orange")
oc.fill = PatternFill("solid", fgColor="FFD9A0")
put(cc, 6, 2, "(Dashboard) an assumption you've moved off consensus — your active bet")
put(cc, 7, 1, "How to use", bold=True)
put(
    cc,
    7,
    2,
    "Edit yellow cells (Model assumptions, WACC drivers, Valuation toggles). Everything else recomputes.",
)

# ===== Cover =====
cv = wb.create_sheet("Cover")
cv.sheet_view.showGridLines = False
cv.column_dimensions["A"].width = 28
cv.column_dimensions["B"].width = 48
put(cv, 1, 1, f"{NAME} ({T})").font = TITLE
put(cv, 2, 1, "DCF valuation — exit-multiple terminal · defaults anchored to consensus").font = SUB
_assumptions_by = (
    "builder defaults (consensus-anchored; no Opus pass)"
    if _baseline is None
    else f"Opus 4.8 — values as of {_baseline.as_of}"
    + (" (provenance seeded)" if _baseline.seeded else "")
)
for i, (k, v) in enumerate(
    [
        ("Last updated", date.today().isoformat()),
        ("Data pulled (FMP)", qlabels[-1]),
        ("Assumptions by", _assumptions_by),
    ]
):
    put(cv, 4 + i, 1, k, bold=True)
    put(cv, 4 + i, 2, v)
band(cv, 8, "VALUATION SUMMARY", 2)
put(cv, 10, 1, "Fair value / share", bold=True)
put(cv, 10, 2, f"=Valuation!B{vps_row}", fmt=PXS).font = BIG
put(cv, 11, 1, "Current price")
put(cv, 11, 2, "=Valuation!$B$2", fmt=PXS)
put(cv, 12, 1, "Upside / (downside)")
put(cv, 12, 2, f"=Valuation!B{ck + 4}", fmt="+0%;(0%)")
put(cv, 13, 1, "Verdict", bold=True)
put(cv, 13, 2, f'=IF(Valuation!B{vps_row}>Valuation!$B$2,"Undervalued","Overvalued")', bold=True)
put(cv, 14, 1, "Monte Carlo P(undervalued)")
put(cv, 14, 2, f"='Monte Carlo'!B{MC_PUNDER_ROW}", fmt=PCT)
band(cv, 15, "KEY ASSUMPTIONS", 2)
for i, (k, v, fmt) in enumerate(
    [
        ("WACC (from WACC tab)", f"=WACC!B{WACC_ROW}", PCT),
        ("Terminal method", "=Valuation!$B$6", None),
        ("Exit basis / multiple", "=Valuation!$B$7", None),
        ("Exit multiple", "=Valuation!$B$8", MULT),
        ("Terminal growth (g)", "=Valuation!$B$4", PCT),
        ("Terminal weight", f"=Valuation!B{ck + 1}", PCT),
        ("Forecast horizon", f"{N_FC} years", None),
    ]
):
    put(cv, 17 + i, 1, k)
    put(cv, 17 + i, 2, v, fmt=fmt)
band(cv, 25, "THE STORY (Opus)", 2)
put(cv, 27, 1, narr or "(Opus narrative)").alignment = WRAP
cv.merge_cells("A27:B43")

# ===== Scenarios + Sensitivity (Python-computed static cells) =====
# One engine (src/dcf/redesign.py) serves both the builder and the refresher, so
# the fresh-build cells and the post-edit refresh rewrites can never drift. The
# WACC here mirrors the reader's derivation exactly (CAPM + market-value weights
# with the FINAL tax rate — wacc0 above predates a possible Opus tax override).
_inp = redesign_mod.RedesignInputs(
    segments=tuple(PROD),
    base_revenue_by_segment={s: seg_ann[ly][s] for s in PROD},
    near_growth_by_segment=dict(g1_def),
    terminal_growth_by_segment=dict(gT_def),
    near_op_margin=margin_near_def,
    terminal_op_margin=margin_term_def,
    tax_rate=TAX,
    capex_2026_m=CAPEX_2026_M,
    terminal_capex_da=capex_da[-1],
    da_ratio=ratios_ly["da"],
    # The reader derives this from the Consensus sheet's year headers (≤6
    # columns), so mirror that here — keeps these static cells byte-identical
    # to what the first refresh would rewrite.
    consensus_years=max(2, min(N_FC, len(CYEARS))),
    wacc=we * ke + (1 - we) * KD * (1 - TAX),
    beta=BETA,
    risk_free_rate=RF,
    equity_risk_premium=ERP,
    cost_of_debt=KD,
    country_risk_premium=CRP,
    growth_fade_curvature=CURV,
    near_sbc_pct=SBC_NEAR,
    terminal_sbc_pct=SBC_TERM,
    terminal_method=OPUS_METHOD,
    terminal_basis=OPUS_BASIS,
    exit_multiple=EXIT_MULT,
    terminal_growth_g=TG,
    current_price=price,
    cash_m=cash_now,
    total_debt_m=debt_now,
    diluted_shares_m=shares_now,
    fx_to_usd=FX,
    bull_deltas=BULL_D,
    bear_deltas=BEAR_D,
)
_sv = redesign_mod.scenario_values(_inp)
redesign_mod.write_scenario_fair_values(wb, _sv)
redesign_mod.write_sensitivity_sheet(wb, redesign_mod.sensitivity_grid(_inp))

order = [
    "Cover",
    "Dashboard",
    "Color Code",
    "WACC",
    "Model",
    "Financials",
    "Consensus",
    "Valuation",
    "Sensitivity",
    "Monte Carlo",
]
for pos, name in enumerate(order):
    wb.move_sheet(name, -(wb.sheetnames.index(name) - pos))

# Assumption provenance LAST (it self-inserts directly after the Dashboard, so
# the explicit reorder above must already have run): the Assumptions sheet +
# yellow-cell comments, classified against the Opus baseline. No ledger update
# here — a from-scratch build's inputs come from the same JSON the ledger
# compares against, so divergence is already recorded by the refresher.
assumptions_doc.write_provenance_into(wb, _inp, cache, ticker=T, update_ledger=False)

DEST.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(DEST))
_up = (full_value / price - 1) if price else 0.0
_seg = "single" if SINGLE_SEG else str(len(PROD))
# Surface the base-FY segment coverage on every build so a contaminated name is
# never silently downgraded (the `cov=` field; see the COVERAGE stderr line above
# for the loud warning when the floor forced whole-company).
_cov_str = f"{_cov.coverage:.2f}" if _cov.coverage is not None else "n/a"
print(
    f"RESULT\t{T}\t{full_value:.2f}\t{price:.2f}\t{_up:+.3f}\t{_seg}\tcov={_cov_str}\tcurv={CURV:.2f}\t{DEST}"
)
_b = f"{_sv.bull:.2f}" if _sv.bull is not None else "n/a"
_r = f"{_sv.bear:.2f}" if _sv.bear is not None else "n/a"
print(f"SCENARIOS\t{T}\tbase={_sv.base:.2f}\tbull={_b}\tbear={_r}")
print("consensus years:", CYEARS)
