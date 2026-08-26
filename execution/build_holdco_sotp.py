"""Formula-first sum-of-the-parts / NAV model for a capital-allocator holdco
(BN = Brookfield Corporation, first instance).

A holdco that consolidates non-recourse subsidiary debt + large minority interest
can't be valued on consolidated earnings or an FCFF DCF — the value is the NAV of
the parts BN's shareholders actually own. This builds the SOTP, corrected per an
independent Opus review:

  ① Asset management  = BAM LTM FRE × FRE multiple × BN ownership %
                        (calibrate the multiple to BAM's own listed valuation —
                        the stake is marketable, so the mark shouldn't stray far
                        from 0.74 × BAM's market cap)
  ② Carried interest  = BN-RETAINED ONLY, disclosure-shaped:
                        accrued × legacy margin (mgmt: ~65%) × (1 − realization
                        haircut) + BN-net annual generated carry × a capitalization
                        multiple, all after tax. (NEVER × the 74% BAM ratio — BAM's
                        2/3 of post-2022 carry belongs to ①'s multiple.)
  ③ Insurance (BWS)   = BWS distributable earnings × a DE multiple   (NOT a bank
                        ROE-excess-return model — a spread insurer's value isn't ROE-on-equity)
  ④ Invested capital  = listed affiliates at market × ownership + private/RE × (1 − haircut)
  ⑤ − Corporate       = recourse debt + preferred + PV(overhead).  Does NOT subtract
                        the ~$250B non-recourse asset-level debt (it's inside ④/①/③).

  SOTP equity = ① + ② + ③ + ④ − ⑤ ;  ÷ diluted shares (incl. BNT exchangeables).

The holdco discount is an OUTPUT (price-to-NAV gap = the thesis), not an input — only
PV(corporate overhead) is deducted. Scenarios follow the S6 bear/base/bull convention
(base feeds dcf_runs; bull/bear ride in assumption_snapshot_json["scenarios"]), and a
reverse-solve (what the market implies for carry + private real estate at the current
price) is the most useful artifact. DE-capitalization is a sanity BAND, not an
independent cross-check (DE already sums the same four buckets).

Editability: every mark loads from data/dcf_assumptions/<T>.json["sotp"]["marks"]
(each entry {"value": x, "note": "written justification"} — the note is the
provenance surfaced by the assumptions tooling), then any yellow Dashboard cell
edited in an existing v2 workbook overrides the JSON (the redesign capture-inject
convention; price always refreshes live), and the effective values sync back to
the JSON so it stays the from-scratch source of truth.

Env (like build_redesigned_dcf.py): DCF_TICKER, DCF_DEST, DCF_REPO_ROOT. Values in $B.
A Python value-of-record mirrors the in-sheet formulas exactly (openpyxl can't
evaluate offline) — verified against the `formulas` lib in tests.
"""

from __future__ import annotations

import dataclasses
import json
import os
import sys
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

CODE_ROOT = Path(__file__).resolve().parents[1]
REPO = Path(os.environ.get("DCF_REPO_ROOT") or CODE_ROOT)
T = os.environ.get("DCF_TICKER", "BN")
DEST = Path(os.environ.get("DCF_DEST") or (REPO / "dcf" / f"{T}.xlsx"))

sys.path.insert(0, str(CODE_ROOT / "src"))


from dcf import reverse_valuation as reverse_valuation_mod  # noqa: E402
from dcf.provenance import (  # noqa: E402
    build_file_provenance,
    build_file_source_record,
    schema_supports_provenance,
)
from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

try:  # persistence is best-effort — the workbook builds without a DB
    from dcf import persist as persist_mod
except ImportError:  # pragma: no cover
    persist_mod = None  # type: ignore[assignment]
try:  # global macro assumptions -- best-effort; recorded for transparency only
    from dcf import global_assumptions as global_dcf
except ImportError:  # pragma: no cover
    global_dcf = None  # type: ignore[assignment]


def _global_assumptions_note() -> dict[str, object]:
    """The editable global macro inputs (risk-free / ERP / tax) in effect at
    build time, recorded in the snapshot for transparency. This is a NAV / SOTP
    model — it does not discount a cash-flow stream — so the globals do NOT
    affect the valuation: ``ke`` here is metadata only and carry tax is a
    separate realization rate. The dashboard surfaces this as 'not used'."""
    if global_dcf is None:  # pragma: no cover
        return {"applies_to_valuation": False, "note": "global module unavailable"}
    g = global_dcf.load(db_path=REPO / "data" / "portfolio.db")
    return {
        "risk_free_rate": g.risk_free_rate,
        "equity_risk_premium": g.equity_risk_premium,
        "tax_rate": g.tax_rate,
        "applies_to_valuation": False,
        "note": "NAV/SOTP model -- globals recorded for transparency, not used in the valuation",
    }


YELLOW = PatternFill("solid", fgColor="FFF2CC")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)
SUB_FONT = Font(bold=True, color="374151")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USDB = '#,##0.0,"B"'  # value already in $B
USD2 = "0.00"
PCT = "0.0%"
MULT = '0.0"x"'


@dataclass
class Sotp:
    """All inputs in $B unless noted. Defaults = the 2026-06 calibration against
    Q1'26 disclosures (BN Q1'26 release, BAM Q1'26 8-K, Sep-2025 Investor Day);
    the written justification per mark lives in
    data/dcf_assumptions/BN.json["sotp"]["marks"], which overrides these."""

    # ① asset management (fee business)
    bam_fre: float = 3.1  # BAM LTM fee-related earnings (Q1'26 8-K)
    fre_mult: float = 24.0  # ≈ BAM's own trading multiple (~$75B / $3.1B LTM FRE)
    bn_own: float = 0.74  # 70% direct + ~4% via BWS (Q1'26 release)
    # ② carried interest — BN-retained only, disclosure-shaped
    carry_accrued_gross: float = 11.8  # accumulated unrealized (gross), Q1'26 disclosed
    carry_legacy_margin: float = 0.65  # mgmt: legacy funds ~65% combined margin
    carry_haircut: float = 0.30  # realization/timing haircut on accrued
    carry_future_net_annual: float = 1.5  # BN-NET generated carry/yr (mgmt: $0.5B'25→$3B'30)
    carry_future_mult: float = 5.0  # risk-haircut capitalization (mgmt plan uses 10x)
    carry_tax: float = 0.18  # cash tax on realized carry
    # ③ insurance (Brookfield Wealth Solutions) — DE multiple, NOT a bank model
    bws_de: float = 1.72  # BWS annualized DE (Q1'26: $430M x 4)
    bws_mult: float = 13.0  # mgmt plan 15x; annuity comps ~8-12x earnings
    # ④ invested capital
    ic_listed: float = 23.5  # BEP 45% + BIP ~27% + BBU 69% at market (June 2026)
    ic_private: float = 38.0  # direct fund investments $12B + real estate $26B (IFRS)
    ic_re_haircut: float = 0.25  # haircut on private/RE (office-heavy RE is the contested line)
    # ⑤ corporate (subtract)
    corp_recourse_debt: float = 14.3  # recourse corporate borrowings (FY25 letter)
    corp_preferred: float = 4.1
    corp_overhead_pv: float = 6.0  # PV of corporate G&A = the real "holdco cost"
    # discount / market
    ke: float = (
        0.10  # blended cost of equity (β-1.85 CAPM ~13-14% is too punitive on the fee annuity)
    )
    shares_m: float = 2371.0  # diluted, incl. exchangeables (Q1'26 LTM DE $5.5B / $2.32)
    price: float = 44.61
    plan_value: float = 68.0  # management's plan value/share (Sep-2025 Investor Day)


def _am(s: Sotp) -> float:
    return s.bam_fre * s.fre_mult * s.bn_own


def _carry(s: Sotp) -> float:
    """Accrued at the legacy margin less a realization haircut, plus BN-net
    future generated carry capitalized — both after tax. Mirrors the disclosure
    shape: legacy funds are 100% BN at ~65% margin; post-2022 carry reaches BN
    as a net 33% royalty, so the future term takes BN-NET dollars directly."""
    accrued = s.carry_accrued_gross * s.carry_legacy_margin * (1 - s.carry_haircut)
    future = s.carry_future_net_annual * s.carry_future_mult
    return (accrued + future) * (1 - s.carry_tax)


def _bws(s: Sotp) -> float:
    return s.bws_de * s.bws_mult


def _ic(s: Sotp) -> float:
    return s.ic_listed + s.ic_private * (1 - s.ic_re_haircut)


def _corp(s: Sotp) -> float:
    return s.corp_recourse_debt + s.corp_preferred + s.corp_overhead_pv


def value(s: Sotp) -> tuple[float, float]:
    """(SOTP equity $B, value per share $)."""
    eq = _am(s) + _carry(s) + _bws(s) + _ic(s) - _corp(s)
    return eq, eq * 1000.0 / s.shares_m


@dataclass(frozen=True)
class SyncResult:
    """The BN workbook-to-owner-JSON sync outcome, kept durable on the run."""

    status: str  # synced | failed: <detail> | not_applicable
    synced_at: datetime | None = None


def reverse_valuation(s: Sotp, eq: float, vps: float) -> dict[str, object] | None:
    """Persist BN's existing exact market-implied carry/private-RE residual."""
    if s.price <= 0 or vps <= 0:
        return None
    floor = _am(s) + _bws(s) + s.ic_listed - _corp(s)
    implied_carry_re = s.price * s.shares_m / 1000.0 - floor
    modeled_carry_re = _carry(s) + s.ic_private * (1 - s.ic_re_haircut)
    residual = reverse_valuation_mod.residual_lever(
        lever_id="implied_carry_private_re_value",
        label="Market-implied carry + private / real-estate value",
        unit="usd_b",
        base_value=modeled_carry_re,
        implied_value=implied_carry_re,
        note="Market equity less asset management, insurance, listed affiliates, and corporate costs.",
    )
    return reverse_valuation_mod.ReverseValuation(
        archetype="holdco_sotp",
        price=s.price,
        base_value_per_share_usd=vps,
        valuation_scope="equity",
        levers=(residual,),
    ).to_snapshot_dict()


def _profile_price_metadata(ticker: str) -> tuple[datetime | None, str | None]:
    """The existing cached-profile price policy, with auditable observation time."""
    profile = REPO / "data" / "historical" / "fmp" / f"{ticker}_profile.json"
    if not profile.is_file():
        return None, "assumption_seed"
    try:
        raw: object = json.loads(profile.read_text(encoding="utf-8"))
        item = raw[0] if isinstance(raw, list) and raw else raw
        if not isinstance(item, dict) or not isinstance(item.get("price"), (int, float)):
            return None, "fmp_profile_unusable"
    except (OSError, ValueError, TypeError):
        return None, "fmp_profile_unusable"
    return datetime.fromtimestamp(profile.stat().st_mtime, tz=UTC), "fmp_profile"


def persist_dcf_run(
    eq: float,
    vps: float,
    price: float,
    ke: float,
    snapshot: dict[str, object],
    sync_result: SyncResult | None = None,
    source_records: tuple[dict[str, object], ...] = (),
) -> bool:
    """Best-effort upsert into dcf_runs so the brief's valuation panel reads the
    SOTP value/share. Shape-agnostic (BN or BRK). No-op without the DB / persist module."""
    db = REPO / "data" / "portfolio.db"
    if persist_mod is None or not db.exists() or not vps:
        return False
    holdings = REPO / "micro_thesis" / "holdings" / f"{T}.json"
    mos: object = None
    if holdings.exists():
        try:
            mos = json.loads(holdings.read_text(encoding="utf-8")).get("mos_bar")
        except (OSError, json.JSONDecodeError):
            mos = None
    live_price_at, live_price_source = _profile_price_metadata(T)
    snapshot_payload = {**snapshot, "workbook": str(DEST)}
    provenance = build_file_provenance(
        ticker=T,
        repo_root=REPO,
        workbook_path=DEST,
        engine_version="holdco_sotp_v1",
        effective_inputs=(
            snapshot.get("marks", {}) if isinstance(snapshot.get("marks"), dict) else {}
        ),
        assumption_snapshot=snapshot_payload,
        live_price=price or None,
        live_price_at=live_price_at,
        live_price_source=live_price_source,
        source_files=(
            (REPO / "data" / "dcf_assumptions" / f"{T}.json", "owner_assumptions"),
            (REPO / "data" / "historical" / "fmp" / f"{T}_profile.json", "company_profile"),
            (REPO / "micro_thesis" / "holdings" / f"{T}.json", "holding_policy"),
        ),
        source_records=source_records,
    )
    row = persist_mod.DcfRunRow(
        ticker=T,
        valuation_date=date.today(),
        horizon_years=0,
        wacc=ke,
        npv=eq * 1000.0,
        npv_per_share=vps,
        shares_outstanding=eq * 1e9 / vps,  # back out shares from equity/value-per-share
        currency="USD",
        live_price=price or None,
        live_price_at=live_price_at,
        mos_bar_used=float(mos) if isinstance(mos, (int, float)) else None,
        assumption_snapshot_json=json.dumps(snapshot_payload, indent=2),
        notes=f"workbook={DEST.name} ({snapshot.get('model', 'holdco SOTP')})",
        assumptions_sync_status=None,
        assumptions_synced_at=None,
        provenance=provenance,
    )
    with connect_sqlite(str(db), role=SQLiteConnectionRole.WRITER, schema_preflight=True) as conn:
        columns = {str(item[1]) for item in conn.execute("PRAGMA table_info(dcf_runs)")}
        if {
            "assumptions_sync_status",
            "assumptions_synced_at",
        }.issubset(columns) and sync_result is not None:
            row = dataclasses.replace(
                row,
                assumptions_sync_status=(
                    sync_result.status if sync_result.status != "not_applicable" else None
                ),
                assumptions_synced_at=sync_result.synced_at,
            )
        if not schema_supports_provenance(conn):
            sys.stderr.write(
                json.dumps(
                    {
                        "event": "dcf_provenance_not_persisted",
                        "ticker": T,
                        "reason": "database schema lacks provenance columns",
                    }
                )
                + "\n"
            )
            row = dataclasses.replace(row, provenance=None)
        persist_mod.upsert(conn, row)
    return True


def _persist_then_sync_bn(
    s: Sotp,
    eq: float,
    vps: float,
    snapshot: dict[str, object],
    source_records: tuple[dict[str, object], ...] = (),
) -> tuple[bool, SyncResult]:
    """Persist effective marks before mutating their owner-authority JSON."""
    pending = SyncResult("pending")
    snapshot["assumption_provenance"] = {
        "authority": f"data/dcf_assumptions/{T}.json",
        "workbook_capture": "supported",
        "sync_status": pending.status,
    }
    persisted = persist_dcf_run(eq, vps, s.price, s.ke, snapshot, pending, source_records)
    if not persisted:
        return False, SyncResult("not_attempted: DCF run was not persisted")

    sync_result = _sync_sotp_json_result(T, s)
    assumption_provenance = snapshot.get("assumption_provenance")
    if isinstance(assumption_provenance, dict):
        assumption_provenance["sync_status"] = sync_result.status
    recorded = persist_dcf_run(eq, vps, s.price, s.ke, snapshot, sync_result, source_records)
    return recorded, sync_result


def _run_bn() -> int:
    prior_workbook_input = (
        build_file_source_record(
            DEST,
            role="owner_workbook_inputs",
            repo_root=REPO,
        )
        if _capture_bn_inputs(DEST)
        else None
    )
    s, notes = _load(T)
    eq, vps = value(s)
    # scenarios — S6 convention: base = the calibrated marks (feeds dcf_runs);
    # bull = no haircuts (full accrued carry + full IFRS private/RE marks);
    # bear = zero credit for carry AND private/RE.
    bull = _scn(s, carry_haircut=0.0, ic_re_haircut=0.0)
    bear = _scn(s, carry_zero=True, ic_private=0.0)
    build(s, DEST, notes=notes, scenarios=(bear, vps, bull))
    snap: dict[str, object] = {
        "model": "holdco_sotp",
        "ke": s.ke,
        "sotp_equity_b": eq,
        "value_per_share_usd": vps,
        "plan_value": s.plan_value,
        "asset_mgmt_b": _am(s),
        "carry_b": _carry(s),
        "insurance_b": _bws(s),
        "invested_capital_b": _ic(s),
        "corporate_b": -_corp(s),
        "marks": {field: getattr(s, field) for field, _row, _label, _fmt in _SOTP_SPEC},
        "global_assumptions": _global_assumptions_note(),
        "scenarios": {
            "base": {"fair_value_per_share_usd": vps},
            "bull": {"fair_value_per_share_usd": bull},
            "bear": {"fair_value_per_share_usd": bear},
        },
    }
    reverse = reverse_valuation(s, eq, vps)
    if reverse is not None:
        snap["reverse_valuation"] = reverse
    persisted, sync_result = _persist_then_sync_bn(
        s,
        eq,
        vps,
        snap,
        (prior_workbook_input,) if prior_workbook_input is not None else (),
    )
    # reverse-solve: what the market implies for carry + private RE at the price
    implied_eq = s.price * s.shares_m / 1000.0
    floor = _am(s) + _bws(s) + s.ic_listed - _corp(s)  # AM + BWS + listed only − corp
    implied_carry_re = implied_eq - floor
    model_carry_re = _carry(s) + s.ic_private * (1 - s.ic_re_haircut)
    print(
        f"RESULT\t{T}\tSOTP/sh=${vps:.2f}\tprice=${s.price:.2f}\tupside={vps / s.price - 1:+.0%}"
        f"\tvs plan ${s.plan_value:.0f}={vps / s.plan_value - 1:+.0%}"
        f"\tdcf_runs={'ok' if persisted else 'skip'}\tjson_sync={sync_result.status}"
        f"\t-> {DEST}"
    )
    print(f"  (1) Asset mgmt (FRE x{s.fre_mult:.0f} x {s.bn_own:.0%})... ${_am(s):6.1f}B")
    print(f"  (2) Carry (BN-retained, net margin/haircut/tax) ${_carry(s):6.1f}B")
    print(f"  (3) Insurance BWS (DE x{s.bws_mult:.0f})............ ${_bws(s):6.1f}B")
    print(f"  (4) Invested capital (listed+private-haircut). ${_ic(s):6.1f}B")
    print(f"  (5) - Corporate (recourse debt+pref+overhead). ${-_corp(s):6.1f}B")
    print(f"  = SOTP equity ................................ ${eq:6.1f}B  -> ${vps:.2f}/sh")
    print(f"\n  Scenarios: bear ${bear:.2f} (carry+RE=0) | base ${vps:.2f} | bull ${bull:.2f}")
    print(
        f"  REVERSE-SOLVE: at ${s.price:.2f}, market implies ${implied_carry_re:.1f}B for "
        f"carry+private-RE vs ${model_carry_re:.1f}B modeled -- the thesis is in that gap."
    )
    return 0


def _scn(s: Sotp, **over: object) -> float:
    import copy

    s2 = copy.copy(s)
    carry_zero = bool(over.pop("carry_zero", False))
    for k, v in over.items():
        setattr(s2, k, v)
    if carry_zero:
        s2.carry_accrued_gross = 0.0
        s2.carry_future_net_annual = 0.0
    return value(s2)[1]


def _sotp_json_path(ticker: str) -> Path:
    return REPO / "data" / "dcf_assumptions" / f"{ticker.upper()}.json"


def _json_marks(ticker: str) -> tuple[dict[str, float], dict[str, str]]:
    """(values, notes) from data/dcf_assumptions/<T>.json["sotp"]["marks"].

    Each mark is {"value": x, "note": "justification"} (or a bare number);
    unknown keys are ignored so the JSON can carry marks for future fields."""
    path = _sotp_json_path(ticker)
    if not path.exists():
        return {}, {}
    try:
        raw: object = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}, {}
    if not isinstance(raw, dict):
        return {}, {}
    sotp_obj = cast("dict[str, object]", raw).get("sotp")
    if not isinstance(sotp_obj, dict):
        return {}, {}
    marks_obj = cast("dict[str, object]", sotp_obj).get("marks")
    if not isinstance(marks_obj, dict):
        return {}, {}
    values: dict[str, float] = {}
    notes: dict[str, str] = {}
    for key, entry in cast("dict[str, object]", marks_obj).items():
        if isinstance(entry, (int, float)) and not isinstance(entry, bool):
            values[key] = float(entry)
            continue
        if isinstance(entry, dict):
            ed = cast("dict[str, object]", entry)
            v = ed.get("value")
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values[key] = float(v)
            n = ed.get("note")
            if isinstance(n, str) and n:
                notes[key] = n
    return values, notes


_SOTP_MARKER = "sotp-inputs-v2"


def _capture_bn_inputs(path: Path) -> dict[str, float]:
    """Yellow-cell values from an existing BN workbook, keyed by Sotp field.

    Only workbooks carrying the v2 marker (Dashboard!D1) are captured — v1
    workbooks hold the retired pre-calibration seeds and must not override the
    calibrated JSON marks on their first rebuild. Price is never captured (it
    always refreshes live). Label-keyed so row drift can't mis-assign values."""
    if not path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except (OSError, KeyError, InvalidFileException):
        return {}
    if "Dashboard" not in wb.sheetnames:
        return {}
    dsh = wb["Dashboard"]
    if dsh["D1"].value != _SOTP_MARKER:
        return {}
    by_label = {label: field for field, _row, label, _fmt in _SOTP_SPEC}
    out: dict[str, float] = {}
    for label_cell, value_cell in dsh.iter_rows(min_row=2, max_row=40, min_col=1, max_col=2):
        field = by_label.get(str(label_cell.value))
        v = value_cell.value
        if field and field != "price" and isinstance(v, (int, float)) and not isinstance(v, bool):
            out[field] = float(v)
    return out


def _sync_sotp_json_result(ticker: str, s: Sotp) -> SyncResult:
    """Mirror the effective marks back into the assumptions JSON (the redesign
    `sync_assumptions_json` convention): numeric values only — the per-mark
    justification notes are never touched. Price is live, not a mark, so it is
    skipped. The detailed outcome is persisted with the DCF run."""
    path = _sotp_json_path(ticker)
    if not path.exists():
        return SyncResult("not_applicable")
    try:
        raw: object = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return SyncResult("failed: unreadable assumptions JSON")
    if not isinstance(raw, dict):
        return SyncResult("failed: assumptions JSON is not an object")
    data = cast("dict[str, object]", raw)
    sotp_obj = data.setdefault("sotp", {})
    if not isinstance(sotp_obj, dict):
        return SyncResult("failed: sotp key is not an object")
    sotp = cast("dict[str, object]", sotp_obj)
    marks_obj = sotp.setdefault("marks", {})
    if not isinstance(marks_obj, dict):
        return SyncResult("failed: marks key is not an object")
    marks = cast("dict[str, object]", marks_obj)
    for field, _row, _label, _fmt in _SOTP_SPEC:
        if field == "price":
            continue
        entry = marks.get(field)
        if isinstance(entry, dict):
            cast("dict[str, object]", entry)["value"] = getattr(s, field)
        else:
            marks[field] = {"value": getattr(s, field)}
    sotp["last_synced"] = date.today().isoformat()
    try:
        path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    except OSError as exc:
        return SyncResult(f"failed: write failed: {exc}")
    return SyncResult("synced", datetime.now(UTC).replace(tzinfo=None))


def _sync_sotp_json(ticker: str, s: Sotp) -> bool:
    """Compatibility wrapper for existing direct callers and tests."""
    return _sync_sotp_json_result(ticker, s).status == "synced"


def _load(ticker: str) -> tuple[Sotp, dict[str, str]]:
    """Effective marks + their justification notes. Precedence (the redesign
    convention): calibrated dataclass defaults < JSON marks < yellow cells the
    user edited in an existing v2 workbook; price always refreshes from FMP."""
    s = Sotp()
    values, notes = _json_marks(ticker)
    fields = {f.name for f in dataclasses.fields(Sotp)}
    for key, v in values.items():
        if key in fields:
            setattr(s, key, v)
    for key, v in _capture_bn_inputs(DEST).items():
        setattr(s, key, v)
    prof = REPO / "data" / "historical" / "fmp" / f"{ticker}_profile.json"
    if prof.exists():
        try:
            d = json.loads(prof.read_text(encoding="utf-8"))
            if isinstance(d, list):
                d = d[0] if d else {}
            if isinstance(d, dict) and d.get("price"):
                s.price = float(d["price"])
        except (OSError, json.JSONDecodeError, ValueError, KeyError):
            pass
    return s, notes


# --------------------------------------------------------------------------- #
def _hdr(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].fill = HEAD_FILL
    ws[cell].font = HEAD_FONT


def _inp(ws: Worksheet, row: int, label: str, val: float, fmt: str = USDB, note: str = "") -> None:
    ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
    c = ws.cell(row=row, column=2, value=val)
    c.fill = YELLOW
    c.number_format = fmt
    c.border = BORDER
    if note:
        n = ws.cell(row=row, column=3, value=note)
        n.font = Font(color="9CA3AF", italic=True, size=9)


# Dashboard input spec — the single source of truth for build (rows/labels),
# capture (label→field), and the JSON sync (field list). (field, row, label, fmt);
# the SOTP sheet formulas reference these rows, so keep them aligned.
_SOTP_SPEC: list[tuple[str, int, str, str]] = [
    ("bam_fre", 3, "BAM total FRE (LTM)", USDB),
    ("fre_mult", 4, "FRE multiple", MULT),
    ("bn_own", 5, "BN ownership of BAM", PCT),
    ("carry_accrued_gross", 7, "Accrued carry (gross, legacy=100% BN)", USDB),
    ("carry_legacy_margin", 8, "Legacy carry margin (after costs)", PCT),
    ("carry_haircut", 9, "Realization/timing haircut", PCT),
    ("carry_future_net_annual", 10, "Future carry / yr (BN-net)", USDB),
    ("carry_future_mult", 11, "Future-carry capitalization", MULT),
    ("carry_tax", 12, "Cash tax on carry", PCT),
    ("bws_de", 14, "BWS distributable earnings (annualized)", USDB),
    ("bws_mult", 15, "BWS DE multiple", MULT),
    ("ic_listed", 17, "Listed affiliates @ market × own", USDB),
    ("ic_private", 18, "Private + real estate (IFRS)", USDB),
    ("ic_re_haircut", 19, "Private/RE haircut", PCT),
    ("corp_recourse_debt", 21, "Recourse corporate debt", USDB),
    ("corp_preferred", 22, "Preferred equity", USDB),
    ("corp_overhead_pv", 23, "PV corporate overhead", USDB),
    ("ke", 25, "Blended cost of equity Ke", PCT),
    ("shares_m", 26, "Diluted shares (M, incl. exchangeables)", "#,##0"),
    ("price", 27, "Current price ($)", USD2),
    ("plan_value", 28, "Management plan value ($)", USD2),
]
_SOTP_ROW = {field: row for field, row, _label, _fmt in _SOTP_SPEC}
# Section headers: row → title.
_SOTP_SECTIONS: list[tuple[int, str]] = [
    (2, "① Asset management (fee business)"),
    (6, "② Carried interest — BN-RETAINED ONLY (never ×74%)"),
    (13, "③ Insurance (BWS) — DE multiple"),
    (16, "④ Invested capital"),
    (20, "⑤ Corporate (subtract)"),
    (24, "Discount / market"),
]


def build(
    s: Sotp,
    dest: Path,
    notes: dict[str, str] | None = None,
    scenarios: tuple[float, float, float] | None = None,
) -> None:
    """Write the workbook. ``notes`` = per-mark justification (column C, from the
    assumptions JSON). ``scenarios`` = (bear, base, bull) value/share — Python
    statics rewritten on every refresh (the S6 sensitivity-grid convention)."""
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    sotp = wb.create_sheet("SOTP")
    scen = wb.create_sheet("Scenarios")
    D = "Dashboard"
    notes = notes or {}

    _hdr(dash, "A1", f"{T} — Holdco Sum-of-the-Parts · Dashboard ($B)")
    # capture marker: only v2 workbooks are edit-preserved across rebuilds
    dash["D1"] = _SOTP_MARKER
    dash["D1"].font = Font(color="D1D5DB", size=8)
    for row, title in _SOTP_SECTIONS:
        dash.cell(row=row, column=1, value=title).font = SUB_FONT
    for field, row, label, fmt in _SOTP_SPEC:
        _inp(dash, row, label, getattr(s, field), fmt, note=notes.get(field, ""))
    dash.column_dimensions["A"].width = 40
    dash.column_dimensions["B"].width = 13
    dash.column_dimensions["C"].width = 80

    # ---- SOTP build (formula-first off the Dashboard inputs) ----
    _hdr(sotp, "A1", f"{T} — Sum-of-the-Parts build ($B)")

    def b(field: str) -> str:
        return f"{D}!$B${_SOTP_ROW[field]}"

    rows = [
        ("① Asset management", f"={b('bam_fre')}*{b('fre_mult')}*{b('bn_own')}"),
        (
            "② Carried interest (BN-retained)",
            f"=({b('carry_accrued_gross')}*{b('carry_legacy_margin')}*(1-{b('carry_haircut')})"
            f"+{b('carry_future_net_annual')}*{b('carry_future_mult')})*(1-{b('carry_tax')})",
        ),
        ("③ Insurance (BWS)", f"={b('bws_de')}*{b('bws_mult')}"),
        (
            "④ Invested capital",
            f"={b('ic_listed')}+{b('ic_private')}*(1-{b('ic_re_haircut')})",
        ),
        (
            "⑤ − Corporate",
            f"=-({b('corp_recourse_debt')}+{b('corp_preferred')}+{b('corp_overhead_pv')})",
        ),
        ("= SOTP equity value", "=SUM(B3:B7)"),
        ("÷ shares → value / share ($)", f"=B8*1000/{b('shares_m')}"),
        ("Upside vs price", f"=B9/{b('price')}-1"),
        ("Discount to plan value", f"=B9/{b('plan_value')}-1"),
        (
            "Market-implied $ for carry+private-RE",
            f"={b('price')}*{b('shares_m')}/1000-({b('bam_fre')}*{b('fre_mult')}*{b('bn_own')}"
            f"+{b('bws_de')}*{b('bws_mult')}+{b('ic_listed')}"
            f"-({b('corp_recourse_debt')}+{b('corp_preferred')}+{b('corp_overhead_pv')}))",
        ),
        ("Modeled $ for carry+private-RE", f"=B4+{b('ic_private')}*(1-{b('ic_re_haircut')})"),
    ]
    rr = 3
    for label, formula in rows:
        sotp.cell(row=rr, column=1, value=label).font = (
            SUB_FONT if "SOTP equity" in label or "value / share" in label else Font(color="374151")
        )
        c = sotp.cell(row=rr, column=2, value=formula)
        c.number_format = (
            USD2
            if ("share" in label or "$)" in label)
            else (PCT if "vs price" in label or "Discount" in label else USDB)
        )
        if "value / share" in label:
            c.font = Font(bold=True)
        rr += 1
    sotp.column_dimensions["A"].width = 40
    sotp.column_dimensions["B"].width = 13

    # ---- Scenarios — S6 bear/base/bull convention. The values are Python
    #      statics rewritten on every refresh (openpyxl can't evaluate offline);
    #      base is what feeds dcf_runs, bull/bear ride in the snapshot JSON. ----
    _hdr(scen, "A1", "Scenarios (bear · base · bull)")
    if scenarios is not None:
        bear, base_v, bull = scenarios
        scen["A3"] = "Scenario"
        scen["B3"] = "Value / share ($)"
        scen["C3"] = "vs price"
        for cell in ("A3", "B3", "C3"):
            scen[cell].font = SUB_FONT
        for i, (name, v) in enumerate(
            [
                ("Bear (carry=0, private/RE=0)", bear),
                ("Base (calibrated marks)", base_v),
                ("Bull (no haircuts)", bull),
            ],
            start=4,
        ):
            scen.cell(row=i, column=1, value=name).font = Font(color="374151")
            vc = scen.cell(row=i, column=2, value=v)
            vc.number_format = USD2
            if s.price:
                pc = scen.cell(row=i, column=3, value=v / s.price - 1)
                pc.number_format = PCT
    scen_notes = [
        "Bear: zero credit for the carried-interest stack AND the private/RE marks",
        "      (AM + BWS + listed affiliates − corporate only).",
        "Base: calibrated marks — accrued carry at the legacy margin less a",
        "      realization haircut; private/RE at IFRS less the haircut.",
        "Bull: no haircuts — full accrued carry and full IFRS private/RE marks.",
        "",
        "2026-06 calibration: the reverse-solve on the SOTP sheet shows the market",
        "discounting the carry + private/RE buckets vs the base marks — it is NOT",
        "paying zero for them (the pre-calibration framing; that was an artifact of",
        "a mis-seeded listed-affiliates mark). Management's plan value additionally",
        "capitalizes target carry at 10x and BWS DE at 15x.",
        "",
        "Holdco discount is an OUTPUT (price-to-NAV gap), not an input — only PV of",
        "corporate overhead is deducted. DE-capitalization is a sanity band, not an",
        "independent cross-check (DE already sums the same four buckets).",
    ]
    for i, n in enumerate(scen_notes, start=9):
        scen.cell(row=i, column=1, value=n).font = Font(color="374151")
    scen.column_dimensions["A"].width = 78
    scen.column_dimensions["B"].width = 16
    scen.column_dimensions["C"].width = 10

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


# =========================================================================== #
# Berkshire shape — the two-column SOTP (investments + operating earnings).
# A capital-allocator holdco like BN, but a totally different decomposition:
# its value is the marked investment portfolio + the capitalized operating
# businesses, with the insurance float as free leverage embedded in the
# portfolio. Buffett's own framing. Kept separate from the BN path so the live
# BN model is untouched.
# =========================================================================== #
@dataclass
class BrkSotp:
    """Berkshire two-column SOTP ($B). Calibrate vs the 10-K."""

    eq_bonds: float = 336.4  # long-term investments (marketable equities + fixed income) @ market
    cash_tbills: float = 373.3  # cash + short-term Treasuries
    op_earn: float = 31.0  # after-tax operating earnings EX insurance investment income (already in the portfolio)
    op_mult: float = (
        13.0  # capitalization of the controlled operating businesses (BNSF/BHE/MSR + underwriting)
    )
    dtl: float = 87.0  # deferred tax on unrealized investment gains
    dtl_hair: float = 0.50  # fraction of the DTL to deduct (PV / probability of the future tax)
    corp: float = 0.0  # net corporate adjustments (parent debt is small; sub debt is non-recourse)
    shares_m: float = 2160.0  # B-equivalent shares
    price: float = 488.30


def _brk_value(s: BrkSotp) -> tuple[float, float]:
    investments = s.eq_bonds + s.cash_tbills
    eq = investments + s.op_earn * s.op_mult - s.dtl * s.dtl_hair - s.corp
    return eq, eq * 1000.0 / s.shares_m


def _brk_load() -> BrkSotp:
    """Seed the marked investments, deferred tax, shares and price from FMP so the
    big balance-sheet items auto-update; operating earnings stays a calibrated default."""
    s = BrkSotp()
    fmp = REPO / "data" / "historical" / "fmp"

    def _rows(name):
        p = fmp / name
        if not p.exists():
            return []
        raw = json.loads(p.read_text(encoding="utf-8"))
        return raw.get("historical", [raw]) if isinstance(raw, dict) else raw

    try:
        bal = _rows(f"{T}_balance_sheet_annual.json")
        if bal:
            b = bal[0]
            s.cash_tbills = float(b.get("cashAndShortTermInvestments") or s.cash_tbills * 1e9) / 1e9
            s.eq_bonds = float(b.get("longTermInvestments") or s.eq_bonds * 1e9) / 1e9
            s.dtl = float(b.get("deferredTaxLiabilitiesNonCurrent") or s.dtl * 1e9) / 1e9
        inc = _rows(f"{T}_income_statement_annual.json")
        if inc:
            s.shares_m = float(inc[0].get("weightedAverageShsOutDil") or s.shares_m * 1e6) / 1e6
        prof = json.loads((fmp / f"{T}_profile.json").read_text(encoding="utf-8"))
        if isinstance(prof, list):
            prof = prof[0] if prof else {}
        if isinstance(prof, dict) and prof.get("price"):
            s.price = float(prof["price"])
    except (OSError, json.JSONDecodeError, KeyError, ValueError, TypeError):
        pass
    return s


_RB = {
    "eq_bonds": 3,
    "cash_tbills": 4,
    "op_earn": 6,
    "op_mult": 7,
    "dtl": 9,
    "dtl_hair": 10,
    "corp": 12,
    "shares": 14,
    "price": 15,
}


def _brk_build(s: BrkSotp, dest: Path) -> None:
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    sotp = wb.create_sheet("SOTP")
    notes_ws = wb.create_sheet("Method")
    D = "Dashboard"

    _hdr(dash, "A1", f"{T} — Berkshire two-column SOTP · Dashboard ($B)")
    dash["A2"] = "① Investment portfolio (marked to market)"
    dash["A2"].font = SUB_FONT
    _inp(dash, _RB["eq_bonds"], "Equities + fixed income @ market", s.eq_bonds)
    _inp(dash, _RB["cash_tbills"], "Cash + Treasury bills", s.cash_tbills)
    dash["A5"] = "② Operating businesses (ex insurance investment income)"
    dash["A5"].font = SUB_FONT
    _inp(dash, _RB["op_earn"], "After-tax operating earnings (ex inv. income)", s.op_earn)
    _inp(dash, _RB["op_mult"], "Operating multiple", s.op_mult, MULT)
    dash["A8"] = "③ Deferred tax + corporate (subtract)"
    dash["A8"].font = SUB_FONT
    _inp(dash, _RB["dtl"], "Deferred tax on unrealized gains", s.dtl)
    _inp(dash, _RB["dtl_hair"], "DTL haircut (PV/probability)", s.dtl_hair, PCT)
    _inp(dash, _RB["corp"], "Net corporate / other", s.corp)
    dash["A13"] = "Market"
    dash["A13"].font = SUB_FONT
    _inp(dash, _RB["shares"], "Shares (M, B-equivalent)", s.shares_m, "#,##0")
    _inp(dash, _RB["price"], "Current price ($)", s.price, USD2)
    dash.column_dimensions["A"].width = 44
    dash.column_dimensions["B"].width = 13

    _hdr(sotp, "A1", f"{T} — Two-column SOTP build ($B)")

    def b(r: int) -> str:
        return f"{D}!$B${r}"

    rows = [
        ("① Investments @ market", f"={b(_RB['eq_bonds'])}+{b(_RB['cash_tbills'])}"),
        ("② Operating businesses", f"={b(_RB['op_earn'])}*{b(_RB['op_mult'])}"),
        ("③ − Deferred tax on gains", f"=-{b(_RB['dtl'])}*{b(_RB['dtl_hair'])}"),
        ("④ − Corporate / other", f"=-{b(_RB['corp'])}"),
        ("= SOTP equity value", "=SUM(B3:B6)"),
        ("÷ shares → value / share ($)", f"=B7*1000/{b(_RB['shares'])}"),
        ("Upside vs price", f"=B8/{b(_RB['price'])}-1"),
        (
            "col-1: investments / share ($)",
            f"=({b(_RB['eq_bonds'])}+{b(_RB['cash_tbills'])})*1000/{b(_RB['shares'])}",
        ),
        (
            "col-2: operating value / share ($)",
            f"={b(_RB['op_earn'])}*{b(_RB['op_mult'])}*1000/{b(_RB['shares'])}",
        ),
    ]
    for i, (label, formula) in enumerate(rows, start=3):
        headline = label.startswith("= SOTP") or label.startswith("÷ shares")
        sotp.cell(row=i, column=1, value=label).font = (
            Font(bold=True) if headline else Font(color="374151")
        )
        c = sotp.cell(row=i, column=2, value=formula)
        c.number_format = USD2 if "share" in label else (PCT if "vs price" in label else USDB)
        if headline:
            c.font = Font(bold=True)
    sotp.column_dimensions["A"].width = 38
    sotp.column_dimensions["B"].width = 13

    _hdr(notes_ws, "A1", "Method — Buffett's two columns")
    for i, n in enumerate(
        [
            "① Investments are MARKED to market (public) — this captures the insurance",
            "   float's value: the float funds the portfolio at ~zero cost (free leverage).",
            "② Operating earnings EXCLUDE insurance investment income — that income is",
            "   already counted by marking the portfolio in ①. Capitalizing it too would",
            "   double-count. Capitalize only the controlled-business earnings (BNSF, BHE,",
            "   MSR, Pilot + insurance UNDERWRITING).",
            "③ Deferred tax on unrealized gains is a real but deferred/contingent claim —",
            "   haircut it (Berkshire rarely liquidates), don't take it at full face.",
            "No holdco discount input: Berkshire trades ~AT its two-column value, so the",
            "price-to-SOTP gap is the output. Swing variables: the operating multiple and",
            "the DTL haircut.",
        ],
        start=3,
    ):
        notes_ws.cell(row=i, column=1, value=n).font = Font(color="374151")
    notes_ws.column_dimensions["A"].width = 80

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def _run_brk() -> int:
    s = _brk_load()
    eq, vps = _brk_value(s)
    _brk_build(s, DEST)
    investments = s.eq_bonds + s.cash_tbills
    op_value = s.op_earn * s.op_mult
    dtl_ded = s.dtl * s.dtl_hair
    snap: dict[str, object] = {
        "model": "holdco_sotp_brk",
        "investments_b": investments,
        "operating_value_b": op_value,
        "dtl_deduction_b": dtl_ded,
        "corporate_b": -s.corp,
        "sotp_equity_b": eq,
        "value_per_share_usd": vps,
        "global_assumptions": _global_assumptions_note(),
    }
    persisted = persist_dcf_run(eq, vps, s.price, 1.0 / s.op_mult, snap)
    print(
        f"RESULT\t{T}\tSOTP/sh=${vps:.2f}\tprice=${s.price:.2f}\tupside={vps / s.price - 1:+.0%}"
        f"\tdcf_runs={'ok' if persisted else 'skip'}\t-> {DEST}"
    )
    print(
        f"  (1) Investments @ market (eq/bonds {s.eq_bonds:.0f} + cash/T-bills {s.cash_tbills:.0f}) ${investments:7.1f}B"
    )
    print(
        f"  (2) Operating businesses (op-earn {s.op_earn:.0f} x {s.op_mult:.0f}x)........ ${op_value:7.1f}B"
    )
    print(
        f"  (3) - Deferred tax on gains ({s.dtl:.0f} x {s.dtl_hair:.0%})............ ${-dtl_ded:7.1f}B"
    )
    print(f"  (4) - Corporate / other................................. ${-s.corp:7.1f}B")
    print(
        f"  = SOTP equity ......................................... ${eq:7.1f}B  -> ${vps:.2f}/sh"
    )
    print(
        f"\n  Two columns: investments ${investments * 1000 / s.shares_m:.0f}/sh + operating "
        f"${op_value * 1000 / s.shares_m:.0f}/sh (less DTL) -- Berkshire trades ~AT intrinsic (no holdco discount)."
    )
    return 0


_BRK_TICKERS = {"BRK-B", "BRK-A", "BRK.B", "BRK"}


def main() -> int:
    return _run_brk() if T.upper() in _BRK_TICKERS else _run_bn()


if __name__ == "__main__":
    raise SystemExit(main())
