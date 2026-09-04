"""Run a deterministic, provider-free DCF benchmark in disposable storage."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import resource
import subprocess
import sys
import tempfile
import time
from collections.abc import Callable
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
from quality.performance import CausalRunEnvelope  # noqa: E402


def _rss() -> int:
    value = int(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
    return value if platform.system() == "Darwin" else value * 1024


def _rss_total() -> int:
    """Return an aggregate upper bound from process and child high-water RSS."""
    own = _rss()
    children = int(resource.getrusage(resource.RUSAGE_CHILDREN).ru_maxrss)
    if platform.system() != "Darwin":
        children *= 1024
    return own + max(0, children)


def _event(name: str, **fields: object) -> None:
    print(json.dumps({"event": name, **fields}, sort_keys=True), file=sys.stderr, flush=True)


def _stage(name: str, action: Callable[[], None]) -> tuple[float, int]:
    started = time.perf_counter()
    action()
    elapsed = max(0.000001, time.perf_counter() - started)
    peak = _rss_total()
    _event(
        "dcf_stage_finished",
        stage=name,
        elapsed_seconds=elapsed,
        peak_rss_bytes=peak,
        rss_semantics="process_plus_children_high_water_upper_bound",
    )
    return elapsed, peak


def workbook_semantic_hash(path: Path) -> tuple[str, str, tuple[str, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    formulas: list[str] = []
    values: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = repr(cell.value)
                values.append(f"{sheet.title}!{cell.coordinate}={value}")
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    # Formula/value parity alone misses workbook-level behavior. Keep this
    # metadata explicit and deterministic so defined names, calculation
    # policy, and hidden sheets cannot drift while cell hashes still match.
    defined_names: list[dict[str, object]] = []
    collection = getattr(workbook, "defined_names", {})
    candidates = collection.values() if hasattr(collection, "values") else ()
    for defined_name in candidates:
        attrs = (
            "name",
            "localSheetId",
            "hidden",
            "function",
            "vbProcedure",
            "xlm",
            "functionGroupId",
            "shortcutKey",
            "publishToServer",
            "workbookParameter",
            "attr_text",
        )
        defined_names.append(
            {attribute: repr(getattr(defined_name, attribute, None)) for attribute in attrs}
        )
    calculation = getattr(workbook, "calculation", None)
    calculation_attrs = getattr(calculation, "__attrs__", ())
    calc_settings = {
        attribute: repr(getattr(calculation, attribute, None)) for attribute in calculation_attrs
    }
    metadata = {
        "defined_names": sorted(defined_names, key=lambda item: repr(item)),
        "calculation": calc_settings,
        "sheet_states": [(sheet.title, sheet.sheet_state) for sheet in workbook.worksheets],
        "active_sheet": getattr(workbook.active, "title", None),
    }
    values.append(f"__workbook_metadata__={json.dumps(metadata, sort_keys=True)}")
    return (
        hashlib.sha256("\n".join(values).encode()).hexdigest(),
        hashlib.sha256("\n".join(formulas).encode()).hexdigest(),
        tuple(workbook.sheetnames),
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=ROOT)
    args = parser.parse_args(argv)
    started = time.perf_counter()
    _event("dcf_workload_started")
    try:
        with tempfile.TemporaryDirectory(prefix="bha115-dcf-") as temp_name:
            fixture_root = Path(temp_name) / "repo"
            fixture_root.mkdir()
            from quality.dcf_benchmark_fixture import write_fixture

            fixture_evidence = write_fixture(
                fixture_root,
                "TESTCO",
                migration_root=ROOT,
            )
            workbook_path = Path(temp_name) / "benchmark.xlsx"
            state: dict[str, str] = {}
            stage_timings: dict[str, float] = {}
            stage_rss: dict[str, int] = {}

            def build() -> None:
                env = dict(
                    os.environ,
                    DCF_TICKER="TESTCO",
                    DCF_REPO_ROOT=str(fixture_root),
                    DCF_DEST=str(workbook_path),
                )
                env.pop("EARNINGS_SUMMARY_DB_PATH", None)
                subprocess.run(
                    [sys.executable, str(ROOT / "execution" / "build_redesigned_dcf.py")],
                    cwd=ROOT,
                    env=env,
                    capture_output=True,
                    check=True,
                    text=True,
                )

            stage_timings["builder"], stage_rss["builder"] = _stage("builder", build)
            semantic_hash, formula_hash, sheets = workbook_semantic_hash(workbook_path)
            state["semantic_sha256"] = semantic_hash
            state["formula_sha256"] = formula_hash
            state["artifact_sha256"] = hashlib.sha256(workbook_path.read_bytes()).hexdigest()

            def formulas() -> None:
                actual, formulas, actual_sheets = workbook_semantic_hash(workbook_path)
                if (actual, formulas, actual_sheets) != (
                    state["semantic_sha256"],
                    state["formula_sha256"],
                    sheets,
                ):
                    raise RuntimeError("DCF workbook semantic identity changed during audit")

            stage_timings["formula-audit"], stage_rss["formula-audit"] = _stage(
                "formula-audit", formulas
            )

            def receipt() -> None:
                state["receipt_sha256"] = hashlib.sha256(
                    json.dumps(state, sort_keys=True).encode()
                ).hexdigest()

            stage_timings["receipt"], stage_rss["receipt"] = _stage("receipt", receipt)
            second_path = Path(temp_name) / "benchmark-second.xlsx"
            workbook_path.replace(second_path)
            stage_timings["parity-build"], stage_rss["parity-build"] = _stage("parity-build", build)
            second_semantic, second_formula, second_sheets = workbook_semantic_hash(workbook_path)
            semantic_parity = (second_semantic, second_formula, second_sheets) == (
                state["semantic_sha256"],
                state["formula_sha256"],
                sheets,
            )
            if not semantic_parity:
                raise RuntimeError("equivalent DCF builds failed semantic parity")
            second_artifact_sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
            byte_parity = second_artifact_sha256 == state["artifact_sha256"]
            revision = subprocess.check_output(
                ["git", "-C", str(args.repo_root.resolve()), "rev-parse", "HEAD"], text=True
            ).strip()
            envelope = CausalRunEnvelope(
                sql_statements=0,
                rows=len(sheets),
                elapsed_seconds=max(0.000001, time.perf_counter() - started),
                peak_rss_bytes=_rss_total(),
                alembic_revision=fixture_evidence.alembic_revision,
                alembic_invocations=fixture_evidence.alembic_invocations,
                migration_elapsed_seconds=fixture_evidence.migration_elapsed_seconds,
                schema_object_count=fixture_evidence.schema_object_count,
                query_plan_sha256=None,
                connection_role="none",
                stage="dcf",
                revision=revision,
                artifact_sha256=state["artifact_sha256"],
                artifact_parity_sha256=second_artifact_sha256,
                artifact_byte_parity=byte_parity,
                semantic_parity=semantic_parity,
                formula_sha256=state["formula_sha256"],
                receipt_sha256=state["receipt_sha256"],
                artifact_sheet_names=sheets,
                stage_timings=stage_timings,
                stage_peak_rss_bytes=stage_rss,
                rss_semantics="process_plus_children_high_water_upper_bound",
            )
    except Exception as exc:
        _event("dcf_workload_failed", error=type(exc).__name__)
        return 1
    print(envelope.model_dump_json())
    _event("dcf_workload_finished", elapsed_seconds=envelope.elapsed_seconds)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
