"""Customer-driven, three-stream platform valuation for Nu Holdings (NU).

The bank excess-return model treats NU as a maturing credit lender (ROE
compressing, loan-book growth fading, fee/float jammed into a thin "fee % of
book"). But NU's own Managerial P&L splits value creation into THREE streams and
~58% of gross profit is now non-credit:

    Credit Income  (lending spread, capital-intensive)
    Float Income   (the deposit franchise — interest on float, capital-light)
    Fee Income     (interchange, insurance, investments, marketplace, NuCel — capital-light)

So value NU the way it actually compounds — a customer-acquisition + monetization
platform:

    Revenue   = active customers x ARPAC x 12          (user growth x cross-sell)
    Gross profit = Revenue x GP margin, split Credit / Float / Fee (mix shifts to fee)
    Net income   = GP - Opex (operating leverage) - tax
    FCFE         = NI - growth in required capital (only the CREDIT book ties up
                   capital; the float/fee franchises are capital-light, so most of
                   NI is distributable -> the capital-light streams are credited)
    Equity value = PV(FCFE) + PV(terminal FCFE)   @ cost of equity

Primary = FCFE (credits the capital-light non-credit growth). Cross-checks:
residual income (full-retention floor — the bank-model lens) and an exit-P/E on
terminal NI. A reverse-solve shows what customer/ARPAC growth the market price
implies. Base case grounded in the Q4'25 deck; every driver is editable.

Env (like build_bank_dcf.py): DCF_TICKER, DCF_DEST, DCF_REPO_ROOT. Values in $M;
customers in millions, ARPAC in $/month. A Python mirror is the value-of-record
and matches the in-sheet formulas exactly.
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import asdict, dataclass, field, replace
from datetime import date
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CODE_ROOT = Path(__file__).resolve().parents[1]
REPO = Path(os.environ.get("DCF_REPO_ROOT") or CODE_ROOT)
T = os.environ.get("DCF_TICKER", "NU")
DEST = Path(os.environ.get("DCF_DEST") or (REPO / "dcf" / f"{T}.xlsx"))

sys.path.insert(0, str(CODE_ROOT / "src"))


from dcf import reverse_valuation as reverse_valuation_mod  # noqa: E402
from dcf.artifact_promotion import (  # noqa: E402
    ArtifactPromotion,
    live_path_from_env,
    promotion_from_env,
)
from dcf.provenance import build_file_provenance, schema_supports_provenance  # noqa: E402
from dcf.specialized_price import (  # noqa: E402
    SpecializedPriceObservation,
    price_seed_source_files,
    resolve_specialized_price,
)
from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

try:  # persistence is best-effort -- the workbook builds without a DB
    from dcf import persist as persist_mod
except ImportError:  # pragma: no cover
    persist_mod = None  # type: ignore[assignment]
try:  # global macro assumptions -- best-effort; degrades to in-code seed defaults
    from dcf import global_assumptions as global_dcf
except ImportError:  # pragma: no cover
    global_dcf = None  # type: ignore[assignment]
try:  # scenario emission (Monthly Red Team PR8) -- best-effort like persistence
    from dcf import redesign as redesign_mod
except ImportError:  # pragma: no cover
    redesign_mod = None  # type: ignore[assignment]

YELLOW = PatternFill("solid", fgColor="FFF2CC")
BLUE_FONT = Font(color="1F4E79")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)
SUB_FONT = Font(bold=True, color="374151")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
USD0 = "#,##0"
NUM1 = "0.0"
NUM2 = "0.00"
MULT = '0.0"x"'


@dataclass
class Assum:
    """Base case grounded in NU's Q4'25 deck. All $M; customers in millions,
    ARPAC in $/month. Rates fade linearly from `_near` (Y1) to `_term` (Y_n)."""

    # engine: customers x ARPAC (the user-growth + cross-sell drivers)
    cust0: float = 135.0  # total customers (M), Q1'26
    custg_near: float = 0.09  # customer growth Y1 (Brazil maturing + MX/CO/US compounding)
    custg_term: float = 0.03
    arpac0: float = 14.9  # ARPAC $/mo, calibrated so Y0 revenue ties to the ~$20B run-rate
    arpacg_near: float = 0.08  # ARPAC expansion Y1 (cross-sell deepening, cohort maturation)
    arpacg_term: float = 0.03
    activity: float = 0.83  # monthly-active / total
    # margins (gross profit = revenue x gp margin; opex = revenue x opex %)
    gpm_near: float = 0.392  # Y0 GP $7.84B / revenue $20B
    gpm_term: float = 0.42  # mix shift to high-margin fee/float + operating leverage
    opex_near: float = 0.131  # Y0 opex $2.61B / revenue $20B
    opex_term: float = 0.10  # operating leverage (efficiency improving)
    tax: float = 0.28  # normalized effective tax (Q4'25 ~31.6%; Q1'26 8.7% was a one-off)
    # gross-profit stream mix (Credit / Float; Fee = 1 - Credit - Float). Shows the
    # three streams; credit share declines, fee share rises (cross-sell).
    cr_near: float = 0.42
    cr_term: float = 0.37
    fl_near: float = 0.29
    fl_term: float = 0.30
    # capital: only the credit book ties up capital (fee/float capital-light)
    cb0: float = 31200.0  # credit book Y0 ($31.2B: cards $20.2B + loans $11.0B)
    cbg_near: float = 0.11
    cbg_term: float = 0.05
    cap_ratio: float = 0.12  # required equity / credit book
    # discounting / terminal
    eq0: float = 11330.0  # book equity Y0 ($11.33B, Q4'25) -- for the RI cross-check
    ke: float = 0.125  # cost of equity (LatAm); explicit unless derive_ke_capm is set
    # Opt-in CAPM: when derive_ke_capm != 0, ke is recomputed from the editable
    # GLOBAL risk-free + ERP as rf + beta*erp + country_risk_premium, so a dashboard
    # change to the macro inputs flows into this model's discount rate. Off by
    # default (ke stays the explicit scalar above -> zero drift). Defaults
    # approximately reproduce 0.125 (0.043 + 1.15*0.045 + 0.0275 ~ 0.123).
    beta: float = 1.15
    country_risk_premium: float = 0.0275  # LatAm (Brazil) blended CRP
    derive_ke_capm: int = 0
    g_term: float = 0.07
    terminal_roe: float = 0.20  # sustainable ROE on RETAINED capital; terminal reinvestment = g/ROE
    exit_pe: float = 16.0  # exit P/E on terminal NI (cross-check)
    years: int = 10
    shares: float = 4907.0  # diluted shares (M)
    price: float = 12.29
    global_assumption_source: dict[str, object] = field(default_factory=dict, repr=False)
    price_seed_source: str = field(default="model_seed", repr=False)
    price_seed_path: str | None = field(default=None, repr=False)


def _interp(near: float, term: float, t: int, n: int) -> float:
    return near if n <= 1 else near + (term - near) * (t - 1) / (n - 1)


@dataclass
class Row:
    t: int
    cust: float
    arpac: float
    revenue: float
    gp: float
    gp_credit: float
    gp_float: float
    gp_fee: float
    ni: float
    cb: float
    reqcap: float
    fcfe: float
    equity: float
    ri: float
    df: float


@dataclass
class Mirror:
    rows: list[Row] = field(default_factory=list)
    pv_fcfe: float = 0.0
    pv_tv: float = 0.0
    value_fcfe: float = 0.0
    vps: float = 0.0
    value_ri: float = 0.0
    vps_ri: float = 0.0
    value_pe: float = 0.0
    vps_pe: float = 0.0
    roe_term: float = 0.0


def mirror(s: Assum) -> Mirror:
    ke, n = s.ke, s.years
    m = Mirror()
    cust_p, arpac_p, cb_p = s.cust0, s.arpac0, s.cb0
    reqcap_p = s.cap_ratio * s.cb0
    eq_p = s.eq0
    for t in range(1, n + 1):
        cust = cust_p * (1 + _interp(s.custg_near, s.custg_term, t, n))
        arpac = arpac_p * (1 + _interp(s.arpacg_near, s.arpacg_term, t, n))
        revenue = cust * s.activity * arpac * 12.0
        gp = revenue * _interp(s.gpm_near, s.gpm_term, t, n)
        cr = _interp(s.cr_near, s.cr_term, t, n)
        fl = _interp(s.fl_near, s.fl_term, t, n)
        gp_credit, gp_float, gp_fee = gp * cr, gp * fl, gp * (1 - cr - fl)
        opex = revenue * _interp(s.opex_near, s.opex_term, t, n)
        ni = (gp - opex) * (1 - s.tax)
        cb = cb_p * (1 + _interp(s.cbg_near, s.cbg_term, t, n))
        reqcap = s.cap_ratio * cb
        fcfe = ni - (reqcap - reqcap_p)
        eq = eq_p + ni  # full-retention path (for the residual-income cross-check)
        ri = ni - ke * eq_p
        df = 1 / (1 + ke) ** t
        m.rows.append(
            Row(
                t,
                cust,
                arpac,
                revenue,
                gp,
                gp_credit,
                gp_float,
                gp_fee,
                ni,
                cb,
                reqcap,
                fcfe,
                eq,
                ri,
                df,
            )
        )
        m.pv_fcfe += fcfe * df
        cust_p, arpac_p, cb_p, reqcap_p, eq_p = cust, arpac, cb, reqcap, eq
    last = m.rows[-1]
    df_n = last.df
    # FCFE terminal: sustainable Gordon. A g%-perpetual-growth business earning
    # ROE must reinvest g/ROE of earnings, so terminal FCFE = NI*(1+g)*(1-g/ROE).
    # (FCFE_n*(1+g) over-distributes -> an implausibly rich terminal multiple.)
    m.roe_term = s.terminal_roe
    ni_n1 = last.ni * (1 + s.g_term)
    tv = ni_n1 * (1 - s.g_term / s.terminal_roe) / (ke - s.g_term)
    m.pv_tv = tv * df_n
    m.value_fcfe = m.pv_fcfe + m.pv_tv
    m.vps = m.value_fcfe / s.shares
    # residual-income cross-check (full retention -> conservative floor / bank-model lens)
    pv_ri = sum(r.ri * r.df for r in m.rows)
    ri_roe = last.ni / last.equity  # end-equity ROE on the full-retention path
    cont_ri = (ri_roe - ke) * last.equity / (ke - s.g_term)
    m.value_ri = s.eq0 + pv_ri + cont_ri * df_n
    m.vps_ri = m.value_ri / s.shares
    # exit-P/E cross-check on terminal NI
    m.value_pe = last.ni * s.exit_pe * df_n + m.pv_fcfe
    m.vps_pe = m.value_pe / s.shares
    return m


def reverse_valuation(s: Assum, m: Mirror) -> dict[str, object] | None:
    """What the market implies for NU's own platform-FCFE mechanics.

    The growth inversion deliberately applies the same shift to customer and
    ARPAC near-term growth that the existing scenario mapping uses.  It does
    not invent a preference between the two growth engines.  Terminal ROE is
    reported separately because it is the model's FCFE terminal reinvestment
    lever, not an FCFF exit multiple.
    """
    if s.price <= 0 or m.vps <= 0:
        return None
    joint_growth = reverse_valuation_mod.solve_lever(
        lever_id="implied_joint_customer_arpac_growth_shift",
        label="Implied joint customer / ARPAC near-growth shift",
        unit="pct",
        base_value=0.0,
        method="monotonic_bisection",
        price_at=lambda shift: (
            mirror(
                replace(
                    s,
                    custg_near=s.custg_near + shift,
                    arpacg_near=s.arpacg_near + shift,
                )
            ).vps
        ),
        target_price=s.price,
        lower_bound=-0.50,
        upper_bound=0.50,
    )
    terminal_roe = reverse_valuation_mod.solve_lever(
        lever_id="implied_terminal_roe",
        label="Implied terminal ROE",
        unit="pct",
        base_value=s.terminal_roe,
        method="monotonic_bisection",
        price_at=lambda roe: mirror(replace(s, terminal_roe=roe)).vps,
        target_price=s.price,
        lower_bound=max(s.g_term + 0.01, 0.01),
        upper_bound=1.00,
    )
    return reverse_valuation_mod.ReverseValuation(
        archetype="platform_dcf",
        price=s.price,
        base_value_per_share_usd=m.vps,
        valuation_scope="equity",
        levers=(joint_growth, terminal_roe),
    ).to_snapshot_dict()


def load_assumptions(ticker: str) -> Assum:
    """Assum defaults overridden by data/bank_assumptions/<T>_platform.json."""
    s = Assum()
    # Seed the editable global tax default before the per-ticker JSON below, so
    # an unpinned platform name tracks the dashboard-set global while a pinned
    # tax still wins (NU pins 0.28 -- a genuinely company-specific Brazilian rate).
    global_loaded = (
        global_dcf.load_with_provenance(db_path=REPO / "data" / "portfolio.db")
        if global_dcf is not None
        else None
    )
    if global_loaded is not None:
        s.global_assumption_source = global_loaded.source_record
        s.tax = global_loaded.assumptions.tax_rate
    else:
        s.global_assumption_source = {
            "role": "global_dcf_assumptions",
            "status": "module_unavailable",
            "observed_at": None,
        }
    p = REPO / "data" / "bank_assumptions" / f"{ticker}_platform.json"
    if p.exists():
        try:
            ov: Any = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            ov = {}
        if isinstance(ov, dict):
            for k, v in cast("dict[str, Any]", ov).items():
                if hasattr(s, k) and isinstance(v, (int, float)):
                    setattr(s, k, v)
                    if k == "price":
                        s.price_seed_source = "owner_assumptions"
                        s.price_seed_path = f"data/bank_assumptions/{ticker}_platform.json"
    # Opt-in: derive ke from the global risk-free/ERP when the name asks for it.
    # Runs after the JSON overrides so beta / CRP / the flag can be tuned per name.
    if global_loaded is not None and s.derive_ke_capm:
        s.ke = (
            global_loaded.assumptions.risk_free_rate
            + s.beta * global_loaded.assumptions.equity_risk_premium
            + s.country_risk_premium
        )
    prof = REPO / "data" / "historical" / "fmp" / f"{ticker}_profile.json"
    if prof.exists():
        try:
            d: Any = json.loads(prof.read_text(encoding="utf-8"))
            if isinstance(d, list):
                d = d[0] if d else {}
            if isinstance(d, dict) and d.get("price"):
                s.price = float(d["price"])
                s.price_seed_source = "fmp_profile"
                s.price_seed_path = f"data/historical/fmp/{ticker}_profile.json"
        except (OSError, json.JSONDecodeError, ValueError, KeyError):
            pass
    return s


# --------------------------------------------------------------------------- #
# Scenario emission (Monthly Red Team PR8).
#
# The redesigned FCFF refresher persists a ``scenarios`` block (bull/base/bear
# fair values per share + the bear leg's provenance) that every risk consumer
# reads (``dcf.scenario_reward.parse_scenario_fair_values``, ``bear_lint``,
# ``portfolio_tail_stress``). This bespoke platform builder historically wrote
# none, so NU had NO modeled downside anywhere. The block below maps the shared
# 6-lever ``ScenarioDeltas`` vocabulary onto THIS model's own levers:
#
#   growth deltas   -> customer growth AND ARPAC growth (near/terminal) — the
#                      two engines of platform revenue
#   margin deltas   -> gross margin (near/terminal) — a credit-cycle NPL spike
#                      lands in NU's managerial P&L as gross-profit compression
#   exit multiple Δ -> sustainable terminal ROE, 1pp per turn. The terminal here
#                      is Gordon NI*(1+g)*(1-g/ROE)/(ke-g); its implied multiple
#                      compresses through ROE, so "-7 turns" reads as ROE -7pp
#   terminal g Δ    -> terminal growth g
#
# The bear deltas come from the holdings JSON's thesis-calibrated ``bear_deltas``
# when present (provenance "thesis"), else the generic BEAR_SEED (provenance
# "seed" — a labeled fallback, flagged by bear_lint on portfolio names). The
# arithmetic is deliberately coarse: an honest thesis-break bear, not precision.
# --------------------------------------------------------------------------- #
def _load_holdings(ticker: str) -> dict[str, object] | None:
    path = REPO / "micro_thesis" / "holdings" / f"{ticker.upper()}.json"
    if not path.exists():
        return None
    try:
        data: Any = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return cast("dict[str, object]", data) if isinstance(data, dict) else None


def scenario_assumptions(s: Assum, deltas: Any) -> Assum:
    """``s`` shifted by one scenario's ``dcf.redesign.ScenarioDeltas`` under the
    documented lever mapping. Guardrails keep the Gordon terminal well-posed
    (g < ke, ROE > g) so a severe bear stays a small positive value rather than
    a sign-flipped artifact."""
    import copy

    s2 = copy.copy(s)
    s2.custg_near += deltas.growth_near
    s2.custg_term += deltas.growth_term
    s2.arpacg_near += deltas.growth_near
    s2.arpacg_term += deltas.growth_term
    s2.gpm_near += deltas.margin_near
    s2.gpm_term += deltas.margin_term
    s2.terminal_roe += deltas.exit_multiple / 100.0
    s2.g_term += deltas.terminal_g
    s2.g_term = min(s2.g_term, s2.ke - 0.01)
    s2.terminal_roe = max(s2.terminal_roe, s2.g_term + 0.01)
    return s2


def scenarios_block(s: Assum, m: Mirror, holdings: dict[str, object] | None) -> dict[str, object]:
    """The ``scenarios`` payload for ``dcf_runs.assumption_snapshot_json`` —
    structurally identical to ``refresh_dcf._redesign_snapshot``'s block, so
    ``parse_scenario_fair_values`` / ``parse_scenario_bear_provenance`` read it
    unchanged. Requires ``redesign_mod`` (caller gates on it)."""
    import dataclasses as _dc

    bull_d = redesign_mod.BULL_SEED
    bear_d = redesign_mod.thesis_bear_seed(holdings)
    provenance = "thesis" if redesign_mod.parse_thesis_bear_deltas(holdings) is not None else "seed"
    bull_vps = mirror(scenario_assumptions(s, bull_d)).vps
    bear_vps = mirror(scenario_assumptions(s, bear_d)).vps
    return {
        "base": {"fair_value_per_share_usd": m.vps},
        "bull": {"fair_value_per_share_usd": bull_vps, "deltas": _dc.asdict(bull_d)},
        "bear": {
            "fair_value_per_share_usd": bear_vps,
            "deltas": _dc.asdict(bear_d),
            "provenance": provenance,
        },
    }


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _hdr(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].fill = HEAD_FILL
    ws[cell].font = HEAD_FONT


def _inp(ws: Worksheet, row: int, label: str, val: float, fmt: str) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
    c = ws.cell(row=row, column=2, value=val)
    c.fill = YELLOW
    c.number_format = fmt
    c.border = BORDER


R = {
    "cust0": 3,
    "custgn": 4,
    "custgt": 5,
    "arpac0": 6,
    "arpacgn": 7,
    "arpacgt": 8,
    "activity": 9,
    "gpmn": 10,
    "gpmt": 11,
    "opexn": 12,
    "opext": 13,
    "tax": 14,
    "crn": 15,
    "crt": 16,
    "fln": 17,
    "flt": 18,
    "cb0": 19,
    "cbgn": 20,
    "cbgt": 21,
    "cap": 22,
    "eq0": 23,
    "ke": 24,
    "g": 25,
    "pe": 26,
    "years": 27,
    "sh": 28,
    "px": 29,
    "troe": 30,
    # outputs
    "val": 33,
    "vps": 34,
    "up": 35,
    "vpsri": 36,
    "vpspe": 37,
    "roet": 38,
}


def build(s: Assum, m: Mirror, dest: Path, holdings: dict[str, object] | None = None) -> None:
    """``holdings`` (Monthly Red Team PR10) is the already-loaded
    ``micro_thesis/holdings/<T>.json`` dict, threaded through so the Scenario
    sheet's Bear/Bull rows derive from the SAME ``scenario_assumptions()`` call
    ``scenarios_block`` uses for the persisted ``dcf_runs`` snapshot — the sheet
    can no longer show a bear the rest of the platform (bear_lint, tail stress,
    red-team evidence packs) disagrees with. ``None`` (no holdings on file / not
    passed) degrades to the generic ``BEAR_SEED`` fallback, same as
    ``scenarios_block``."""
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    mod = wb.create_sheet("Model")
    val = wb.create_sheet("Valuation")
    scn = wb.create_sheet("Scenario")
    D = "Dashboard"

    # ---------- Dashboard ----------
    _hdr(dash, "A1", f"{T} - Customer-Driven Platform DCF | Dashboard")
    dash["A2"] = "Engine: customers x ARPAC -> revenue; Credit/Float/Fee gross profit"
    dash["A2"].font = SUB_FONT
    rows = [
        ("cust0", "Customers Y0 (M)", s.cust0, USD0),
        ("custgn", "Customer growth - near", s.custg_near, PCT),
        ("custgt", "Customer growth - terminal", s.custg_term, PCT),
        ("arpac0", "ARPAC Y0 ($/mo)", s.arpac0, NUM2),
        ("arpacgn", "ARPAC growth - near", s.arpacg_near, PCT),
        ("arpacgt", "ARPAC growth - terminal", s.arpacg_term, PCT),
        ("activity", "Activity rate", s.activity, PCT),
        ("gpmn", "Gross margin - near", s.gpm_near, PCT),
        ("gpmt", "Gross margin - terminal", s.gpm_term, PCT),
        ("opexn", "Opex % rev - near", s.opex_near, PCT),
        ("opext", "Opex % rev - terminal", s.opex_term, PCT),
        ("tax", "Tax rate", s.tax, PCT),
        ("crn", "Credit GP share - near", s.cr_near, PCT),
        ("crt", "Credit GP share - terminal", s.cr_term, PCT),
        ("fln", "Float GP share - near", s.fl_near, PCT),
        ("flt", "Float GP share - terminal", s.fl_term, PCT),
        ("cb0", "Credit book Y0 ($M)", s.cb0, USD0),
        ("cbgn", "Credit book growth - near", s.cbg_near, PCT),
        ("cbgt", "Credit book growth - terminal", s.cbg_term, PCT),
        ("cap", "Capital ratio (req eq / book)", s.cap_ratio, PCT),
        ("eq0", "Book equity Y0 ($M)", s.eq0, USD0),
        ("ke", "Cost of equity Ke", s.ke, PCT),
        ("g", "Terminal growth g", s.g_term, PCT),
        ("pe", "Exit P/E (cross-check)", s.exit_pe, MULT),
        ("years", "Forecast years", s.years, "0"),
        ("sh", "Diluted shares (M)", s.shares, USD0),
        ("px", "Current price ($)", s.price, NUM2),
        ("troe", "Terminal ROE (sustainable)", s.terminal_roe, PCT),
    ]
    for key, lab, v, fmt in rows:
        _inp(dash, R[key], lab, v, fmt)

    _hdr(dash, "A32", "OUTPUT")
    for rr, lab, ref, fmt in (
        (R["val"], "Equity value - FCFE ($M)", "Valuation!$B$8", USD0),
        (R["vps"], "Value per share - FCFE ($)", "Valuation!$B$9", NUM2),
        (R["up"], "Upside vs price", "Valuation!$B$10", PCT),
        (R["vpsri"], "Value/share - residual income", "Valuation!$B$12", NUM2),
        (R["vpspe"], "Value/share - exit P/E", "Valuation!$B$14", NUM2),
        (R["roet"], "Terminal ROE", "Valuation!$B$15", PCT),
    ):
        dash.cell(row=rr, column=1, value=lab).font = SUB_FONT
        oc = dash.cell(row=rr, column=2, value=f"={ref}")
        oc.number_format = fmt
        oc.font = Font(bold=True)
    dash.column_dimensions["A"].width = 34
    dash.column_dimensions["B"].width = 14

    # ---------- Model (formula-first engine) ----------
    _hdr(mod, "A1", f"{T} - Platform engine (formula-first; $M)")
    n = s.years
    col0 = 3  # column C = Y1

    def cl(i: int) -> str:
        return get_column_letter(i)

    labels = {
        4: "Year",
        5: "Customers (M)",
        6: "ARPAC ($/mo)",
        7: "Revenue",
        8: "Gross profit",
        9: "  Credit GP",
        10: "  Float GP",
        11: "  Fee GP",
        12: "Opex",
        13: "Net income",
        14: "Credit book",
        15: "Required capital",
        16: "FCFE",
        17: "Discount factor",
        18: "PV FCFE",
        19: "Book equity (retain)",
        20: "Residual income",
        21: "PV residual income",
    }
    for r, lab in labels.items():
        mod.cell(row=r, column=2, value=lab).font = Font(color="374151") if r != 4 else SUB_FONT

    def dref(key: str) -> str:  # Dashboard absolute cell ref for input `key`
        return f"{D}!$B${R[key]}"

    def interp_f(c: str, near_key: str, term_key: str) -> str:
        near, term = dref(near_key), dref(term_key)
        return f"{near}+({term}-{near})*({c}$4-1)/({dref('years')}-1)"

    for j in range(1, n + 1):
        c = cl(col0 + j - 1)
        p = cl(col0 + j - 2)  # prior column (j==1 references Dashboard Y0 instead)
        mod[f"{c}4"] = j
        if j == 1:
            mod[f"{c}5"] = f"={dref('cust0')}*(1+{interp_f(c, 'custgn', 'custgt')})"
            mod[f"{c}6"] = f"={dref('arpac0')}*(1+{interp_f(c, 'arpacgn', 'arpacgt')})"
            mod[f"{c}14"] = f"={dref('cb0')}*(1+{interp_f(c, 'cbgn', 'cbgt')})"
            mod[f"{c}19"] = f"={dref('eq0')}+{c}13"
            mod[f"{c}16"] = f"={c}13-({c}15-{dref('cap')}*{dref('cb0')})"
            mod[f"{c}20"] = f"={c}13-{dref('ke')}*{dref('eq0')}"
        else:
            mod[f"{c}5"] = f"={p}5*(1+{interp_f(c, 'custgn', 'custgt')})"
            mod[f"{c}6"] = f"={p}6*(1+{interp_f(c, 'arpacgn', 'arpacgt')})"
            mod[f"{c}14"] = f"={p}14*(1+{interp_f(c, 'cbgn', 'cbgt')})"
            mod[f"{c}19"] = f"={p}19+{c}13"
            mod[f"{c}16"] = f"={c}13-({c}15-{p}15)"
            mod[f"{c}20"] = f"={c}13-{dref('ke')}*{p}19"
        mod[f"{c}7"] = f"={c}5*{dref('activity')}*{c}6*12"
        mod[f"{c}8"] = f"={c}7*({interp_f(c, 'gpmn', 'gpmt')})"
        mod[f"{c}9"] = f"={c}8*({interp_f(c, 'crn', 'crt')})"
        mod[f"{c}10"] = f"={c}8*({interp_f(c, 'fln', 'flt')})"
        mod[f"{c}11"] = f"={c}8-{c}9-{c}10"
        mod[f"{c}12"] = f"={c}7*({interp_f(c, 'opexn', 'opext')})"
        mod[f"{c}13"] = f"=({c}8-{c}12)*(1-{dref('tax')})"
        mod[f"{c}15"] = f"={dref('cap')}*{c}14"
        mod[f"{c}17"] = f"=1/(1+{dref('ke')})^{c}4"
        mod[f"{c}18"] = f"={c}16*{c}17"
        mod[f"{c}21"] = f"={c}20*{c}17"
    for r in range(5, 22):
        for j in range(1, n + 1):
            cc = mod.cell(row=r, column=col0 + j - 1)
            cc.number_format = NUM2 if r in (6, 17) else (NUM1 if r == 5 else USD0)
    mod.column_dimensions["B"].width = 20

    # ---------- Valuation ----------
    _hdr(val, "A1", f"{T} - Valuation ($M)")
    cN = cl(col0 + n - 1)
    c1 = cl(col0)
    vrows = [
        ("PV of FCFE (yrs 1-N)", f"=SUM(Model!{c1}18:{cN}18)", USD0, 2),
        (
            "Terminal FCFE (sustainable)",
            f"=Model!{cN}13*(1+{D}!$B${R['g']})*(1-{D}!$B${R['g']}/{D}!$B${R['troe']})/({D}!$B${R['ke']}-{D}!$B${R['g']})",
            USD0,
            3,
        ),
        ("PV of terminal value", f"=B3*Model!{cN}17", USD0, 4),
        ("Terminal NI", f"=Model!{cN}13", USD0, 5),
        ("Terminal customers (M)", f"=Model!{cN}5", NUM1, 6),
        ("Terminal ARPAC ($/mo)", f"=Model!{cN}6", NUM2, 7),
        ("Equity value - FCFE", "=B2+B4", USD0, 8),
        ("Value per share - FCFE ($)", f"=B8/{D}!$B${R['sh']}", NUM2, 9),
        ("Upside vs price", f"=B9/{D}!$B${R['px']}-1", PCT, 10),
        (
            "Residual-income value (floor)",
            f"={D}!$B${R['eq0']}+SUM(Model!{c1}21:{cN}21)+((Model!{cN}13/Model!{cN}19)-{D}!$B${R['ke']})*Model!{cN}19/({D}!$B${R['ke']}-{D}!$B${R['g']})*Model!{cN}17",
            USD0,
            11,
        ),
        ("Value/share - residual income", f"=B11/{D}!$B${R['sh']}", NUM2, 12),
        ("Exit-P/E value", f"=Model!{cN}13*{D}!$B${R['pe']}*Model!{cN}17+B2", USD0, 13),
        ("Value/share - exit P/E", f"=B13/{D}!$B${R['sh']}", NUM2, 14),
        ("Terminal ROE (sustainable)", f"={D}!$B${R['troe']}", PCT, 15),
    ]
    for lab, formula, fmt, rr in vrows:
        val.cell(row=rr, column=1, value=lab).font = (
            SUB_FONT if rr in (8, 9) else Font(color="374151")
        )
        vc = val.cell(row=rr, column=2, value=formula)
        vc.number_format = fmt
        if rr in (8, 9):
            vc.font = Font(bold=True)
    val.column_dimensions["A"].width = 32
    val.column_dimensions["B"].width = 14

    # ---------- Scenario ----------
    _hdr(scn, "A1", "Scenarios & reverse-solve")
    scn["A2"], scn["B2"], scn["C2"], scn["D2"], scn["E2"] = (
        "Scenario",
        "Cust g (near)",
        "ARPAC g (near)",
        "GP margin (term)",
        "Value/share (FCFE)",
    )
    for cc in ("A2", "B2", "C2", "D2", "E2"):
        scn[cc].font = SUB_FONT
    r = 3
    if redesign_mod is not None:
        # Monthly Red Team PR10: Bear/Bull rows come from the SAME
        # scenario_assumptions() call scenarios_block() uses for the persisted
        # dcf_runs snapshot (bear deltas from holdings bear_deltas when present,
        # provenance "thesis"; else the generic BEAR_SEED, provenance "seed") —
        # this sheet and the snapshot every risk consumer reads (bear_lint,
        # tail stress, red-team evidence packs) can no longer disagree.
        bull_d = redesign_mod.BULL_SEED
        bear_d = redesign_mod.thesis_bear_seed(holdings)
        bear_provenance = (
            "thesis" if redesign_mod.parse_thesis_bear_deltas(holdings) is not None else "seed"
        )
        scenario_rows: list[tuple[str, Assum]] = [
            ("Bear", scenario_assumptions(s, bear_d)),
            ("Base", s),
            ("Bull", scenario_assumptions(s, bull_d)),
        ]
        for name, s2 in scenario_rows:
            v = mirror(s2).vps
            scn.cell(row=r, column=1, value=name).font = Font(bold=(name == "Base"), color="374151")
            for col, vv in zip(
                ("B", "C", "D"),
                (s2.custg_near, s2.arpacg_near, s2.gpm_term),
                strict=True,
            ):
                cc = scn[f"{col}{r}"]
                cc.value = vv
                cc.number_format = PCT
            ec = scn.cell(row=r, column=5, value=round(v, 2))
            ec.number_format = NUM2
            ec.font = Font(bold=(name == "Base"))
            r += 1
        prov_label = (
            "bear from holdings bear_deltas (thesis)"
            if bear_provenance == "thesis"
            else "bear from generic BEAR_SEED (seed fallback -- no holdings bear_deltas on file)"
        )
        scn.cell(row=r, column=1, value="Bear provenance").font = Font(italic=True, color="6B7280")
        prov_cell = scn.cell(row=r, column=2, value=prov_label)
        prov_cell.font = Font(italic=True, color="6B7280", size=9)
        scn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1
    else:  # pragma: no cover - import failure only, exercised by no test env
        # dcf.redesign unavailable: degrade LOUDLY (an empty/stale scenario row
        # is exactly the PR10 bug) rather than silently falling back to the old
        # hardcoded Bear/Bull levers.
        scn.cell(
            row=r, column=1, value="Scenarios unavailable (dcf.redesign import failed)"
        ).font = Font(italic=True, color="B91C1C")
        r += 1
    r += 1
    _hdr(scn, f"A{r}", "REVERSE-SOLVE - what the price implies")
    r += 1
    implied_eq = s.price * s.shares
    rs = [
        ("Market equity value ($M)", f"{implied_eq:,.0f}"),
        ("Model FCFE equity value ($M)", f"{m.value_fcfe:,.0f}"),
        ("Model implies vs price", f"{m.vps / s.price - 1:+.0%}"),
        ("FCFE floor (resid. income) / sh", f"${m.vps_ri:.2f}"),
        ("Exit-P/E cross-check / sh", f"${m.vps_pe:.2f}"),
    ]
    for lab, txt in rs:
        scn.cell(row=r, column=1, value=lab).font = Font(color="374151")
        scn.cell(row=r, column=2, value=txt).font = Font(size=10, color="374151")
        r += 1
    scn.column_dimensions["A"].width = 36
    scn.column_dimensions["B"].width = 22

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def persist_dcf_run(
    s: Assum,
    m: Mirror,
    holdings: dict[str, object] | None = None,
    price_observation: SpecializedPriceObservation | None = None,
    *,
    artifact_promotion: ArtifactPromotion | None = None,
) -> bool:
    """``holdings=None`` (the pre-PR10 2-arg call shape every test/caller uses)
    loads ``micro_thesis/holdings/<T>.json`` itself, same as before. ``main()``
    now passes the SAME dict ``build()``'s Scenario sheet used, so a mid-run
    file edit can never make the sheet and the persisted snapshot disagree —
    a ticker with genuinely no holdings JSON still resolves to ``None`` either
    way, so this collapses "not passed" and "no holdings on file" safely."""
    db = REPO / "data" / "portfolio.db"
    if persist_mod is None or not db.exists() or not m.vps:
        return False
    if holdings is None:
        holdings = _load_holdings(T)
    mos: object = holdings.get("mos_bar") if holdings else None
    observed_at = price_observation.observed_at if price_observation is not None else None
    price_source = (
        price_observation.source_name if price_observation is not None else "assumption_seed"
    )
    live_workbook = live_path_from_env(DEST)
    snap_payload: dict[str, object] = {
        "model": "platform_dcf",
        "value_per_share_fcfe": m.vps,
        "value_per_share_residual_income": m.vps_ri,
        "value_per_share_exit_pe": m.vps_pe,
        "equity_value_fcfe_m": m.value_fcfe,
        "terminal_roe": m.roe_term,
        "terminal_customers_m": m.rows[-1].cust,
        "terminal_arpac": m.rows[-1].arpac,
        "ke": s.ke,
        "workbook": str(live_workbook),
        "assumption_provenance": {
            "authority": f"data/bank_assumptions/{T}_platform.json",
            "workbook_capture": "unsupported",
            "sync_status": "not_applicable",
        },
    }
    if redesign_mod is not None:
        snap_payload["scenarios"] = scenarios_block(s, m, holdings)
    reverse = reverse_valuation(s, m)
    if reverse is not None:
        snap_payload["reverse_valuation"] = reverse
    snap = json.dumps(snap_payload, indent=2)
    provenance = build_file_provenance(
        ticker=T,
        repo_root=REPO,
        workbook_path=DEST,
        workbook_locator_path=live_workbook,
        engine_version="nu_platform_fcfe_v1",
        effective_inputs=asdict(s),
        assumption_snapshot=snap_payload,
        live_price=s.price or None,
        live_price_at=observed_at,
        live_price_source=price_source,
        source_files=(
            (REPO / "data" / "bank_assumptions" / f"{T}_platform.json", "owner_assumptions"),
            (REPO / "micro_thesis" / "holdings" / f"{T}.json", "holding_policy"),
            *(
                price_seed_source_files(REPO, price_observation)
                if price_observation is not None
                else ()
            ),
        ),
        source_records=(s.global_assumption_source,),
    )
    row = persist_mod.DcfRunRow(
        ticker=T,
        valuation_date=date.today(),
        horizon_years=s.years,
        wacc=s.ke,
        npv=m.value_fcfe,
        npv_per_share=m.vps,
        shares_outstanding=s.shares * 1e6,
        currency="USD",
        live_price=s.price or None,
        live_price_at=observed_at,
        mos_bar_used=float(mos) if isinstance(mos, (int, float)) else None,
        assumption_snapshot_json=snap,
        notes=f"workbook={live_workbook.name} (customer-driven platform DCF)",
        provenance=provenance,
    )
    with connect_sqlite(str(db), role=SQLiteConnectionRole.WRITER, schema_preflight=True) as conn:
        if not schema_supports_provenance(conn):
            sys.stderr.write(
                json.dumps(
                    {
                        "event": "dcf_provenance_not_persisted",
                        "ticker": T,
                        "reason": "database schema lacks provenance columns",
                    }
                )
                + "\n"
            )
            row = replace(row, provenance=None)
        if artifact_promotion is None:
            return persist_mod.upsert(conn, row)
        return persist_mod.upsert(conn, row, artifact_promotion=artifact_promotion)


def main() -> int:
    s = load_assumptions(T)
    price_observation = resolve_specialized_price(
        REPO,
        T,
        fallback_price=s.price,
        fallback_source_name=s.price_seed_source,
        fallback_source_path=s.price_seed_path,
    )
    s.price = price_observation.price
    m = mirror(s)
    # Loaded once and threaded through both the Scenario sheet (build) and the
    # persisted snapshot (persist_dcf_run) — PR10: one holdings read, one bear,
    # never two that could drift on a mid-run file edit.
    holdings = _load_holdings(T)
    build(s, m, DEST, holdings)
    artifact_promotion = promotion_from_env(DEST)
    if os.environ.get("DCF_PERSIST", "1") != "1":
        persisted = False
    elif artifact_promotion is not None:
        persisted = persist_dcf_run(
            s,
            m,
            holdings,
            price_observation,
            artifact_promotion=artifact_promotion,
        )
    else:
        persisted = persist_dcf_run(s, m, holdings, price_observation)
    up = (m.vps / s.price - 1) if s.price else 0.0
    last = m.rows[-1]
    print(
        f"RESULT\t{T}\tvalue/sh(FCFE)=${m.vps:.2f}\tprice=${s.price:.2f}\tupside={up:+.0%}"
        f"\tRI=${m.vps_ri:.2f}\texitPE=${m.vps_pe:.2f}\tROE_term={m.roe_term:.0%}"
        f"\tdcf_runs={'ok' if persisted else 'skip'}\t-> {DEST}"
    )
    print(f"{'Yr':>2} {'Cust':>6} {'ARPAC':>6} {'Rev':>7} {'GP':>7} {'NI':>7} {'FCFE':>7}")
    for r in m.rows:
        print(
            f"{r.t:>2} {r.cust:>6.0f} {r.arpac:>6.1f} {r.revenue:>7.0f} {r.gp:>7.0f} {r.ni:>7.0f} {r.fcfe:>7.0f}"
        )
    print(
        f"\nTerminal: {last.cust:.0f}M customers x ${last.arpac:.1f} ARPAC -> rev ${last.revenue / 1000:.0f}B, "
        f"NI ${last.ni / 1000:.1f}B, ROE {m.roe_term:.0%}"
    )
    print(
        f"FCFE ${m.vps:.2f} | RI floor ${m.vps_ri:.2f} | exit-PE ${m.vps_pe:.2f}  (vs ${s.price:.2f})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
