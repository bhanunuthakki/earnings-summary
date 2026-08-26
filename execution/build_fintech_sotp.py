"""Hybrid sum-of-the-parts model for SoFi Technologies (SOFI).

A pure bank/credit excess-return model is the WRONG primary lens for SOFI: ~40% of
revenue is fee/platform income (Galileo tech platform, SoFi Money deposits, Invest
AUM, Credit Card, cross-sell) that a NIM-on-a-loan-book model crushes by folding it
into a thin "fee % of book" line. So value SOFI as the sum of its three reported
segments, each on the multiple appropriate to its economics, less a tapering DCF of
the corporate/unallocated overhead:

    (1) Lending             = contribution profit x a consumer-credit multiple
                              (cyclical personal loans, deposit-funded, gain-on-sale)
    (2) Financial Services  = contribution profit x a growth fee/deposit multiple
                              (SoFi Money deposits, Invest/AUM, Credit Card, cross-sell)
    (3) Technology Platform = contribution profit x a BaaS/payments-SaaS multiple
                              (Galileo + Technisys; ~32% margin, ~14% growth)
    (4) - Corporate         = an explicit DCF of the unallocated overhead + SBC,
                              GROWING SLOWER than the segments (operating leverage)
                              so the drag is finite and tapering, not a flat
                              perpetuity capitalization
    (5) - Net corporate debt

    SOTP equity = (1) + (2) + (3) - (4) - (5) ;  / diluted shares.

Segment contribution profit is SOFI's own CODM segment-profit measure (10-K segment
footnote). The most useful artifact is the REVERSE-SOLVE: at the market price, what
segment multiples / corporate operating-leverage must hold -- i.e. how much
ROE/growth inflection is priced in. SOFI's whole equity story is that inflection, so
the SOTP makes the pieces visible rather than pretending one number is precise.

Env (like build_bank_dcf.py): DCF_TICKER, DCF_DEST, DCF_REPO_ROOT. Values in $M;
display in $B. A Python mirror is the value-of-record (openpyxl can't evaluate
formulas offline) and matches the in-sheet formulas exactly.
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import asdict, dataclass, field, replace
from datetime import date
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(os.environ.get("DCF_REPO_ROOT") or Path(__file__).resolve().parents[1])
T = os.environ.get("DCF_TICKER", "SOFI")
DEST = Path(os.environ.get("DCF_DEST") or (REPO / "dcf" / f"{T}.xlsx"))

sys.path.insert(0, str(REPO / "src"))


from dcf.artifact_promotion import ArtifactPromotion, promotion_from_env  # noqa: E402
from dcf.provenance import build_file_provenance, schema_supports_provenance  # noqa: E402
from dcf.specialized_price import (  # noqa: E402
    SpecializedPriceObservation,
    price_seed_source_files,
    resolve_specialized_price,
)
from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

try:  # persistence is best-effort -- the workbook builds without a DB
    from dcf import persist as persist_mod
except ImportError:  # pragma: no cover
    persist_mod = None  # type: ignore[assignment]
try:  # global macro assumptions -- best-effort; degrades to in-code seed defaults
    from dcf import global_assumptions as global_dcf
except ImportError:  # pragma: no cover
    global_dcf = None  # type: ignore[assignment]

YELLOW = PatternFill("solid", fgColor="FFF2CC")
BLUE_FONT = Font(color="1F4E79")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)
SUB_FONT = Font(bold=True, color="374151")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USD0 = "#,##0"
USD2 = "0.00"
PCT = "0.0%"
MULT = '0.0"x"'


@dataclass
class Sotp:
    """All $ in millions unless noted. Segment contribution profit = SOFI FY2025
    10-K segment footnote (CODM segment-profit measure). Multiples are EV/CP and
    are the editable swing -- calibrate to your conviction on each franchise."""

    # (1) / (2) / (3) segment contribution profit (FY2025 actual, $M)
    lending_cp: float = 1016.9
    fs_cp: float = 792.9
    tech_cp: float = 144.4
    # segment EV / contribution-profit multiples (the editable judgment)
    lending_mult: float = 6.5  # consumer lender, cyclical credit, deposit-funded
    fs_mult: float = 13.0  # fast-growing deposit/fee/AUM/cross-sell franchise (88% rev growth)
    tech_mult: float = 14.0  # Galileo BaaS/payments platform (~32% margin, ~14% growth)
    # (4) corporate / unallocated overhead -- tapering DCF (operating leverage):
    # corporate grows much slower than the segments, so its PV is a finite drag
    # rather than a flat perpetuity capitalization.
    corp_cost: float = 1000.0  # year-0 normalized corporate + SBC overhead ($M)
    corp_growth: float = 0.02  # corporate annual growth (<< segment growth -> op leverage)
    corp_years: float = 10.0  # explicit DCF horizon
    corp_terminal_mult: float = 2.0  # terminal corporate residual capitalization (modest)
    ke: float = 0.133  # cost of equity (discounts the corporate stream; high-beta fintech)
    # Opt-in CAPM: when derive_ke_capm != 0, ke is recomputed from the editable
    # GLOBAL risk-free + ERP as rf + beta*erp, so a dashboard change to the macro
    # inputs flows into this model. Off by default (ke stays the explicit scalar
    # above -> zero drift). beta 2.0 reproduces the default exactly: 0.043 + 2.0*0.045.
    # (Tax is not a SOTP input -- this is an EV/contribution-profit multiples model --
    # so the global tax rate intentionally does not apply here.)
    beta: float = 2.0
    derive_ke_capm: int = 0
    # (5) net corporate debt ($M; SOFI cash roughly offsets convertible/warehouse debt)
    net_corp_debt: float = 0.0
    # reference drivers (Segments sheet; not in the valuation math)
    deposits_b: float = 37.5  # total deposits YE2025
    members_m: float = 13.7  # total members YE2025
    invest_aum_b: float = 3.4  # SoFi Invest AUM (modest vs deposits/interchange)
    adj_ebitda: float = 1050.0  # FY2025 adjusted EBITDA (cross-check)
    gaap_pretax: float = 525.9  # FY2025 income before taxes (reconciliation anchor)
    shares_m: float = 1281.0  # diluted shares
    price: float = 15.71
    price_seed_source: str = field(default="model_seed", repr=False)
    price_seed_path: str | None = field(default=None, repr=False)
    global_assumption_source: dict[str, object] = field(default_factory=dict, repr=False)


def seg_vals(s: Sotp) -> tuple[float, float, float]:
    """(lending, financial-services, technology) segment values, $M."""
    return s.lending_cp * s.lending_mult, s.fs_cp * s.fs_mult, s.tech_cp * s.tech_mult


def corp_drag(s: Sotp) -> float:
    """PV of the corporate/unallocated overhead, $M. An explicit DCF: corporate
    grows at ``corp_growth`` (deliberately far below segment growth -> operating
    leverage), discounted at Ke over ``corp_years``, plus a modest terminal
    capitalization of the final-year cost. Replaces a flat perpetuity, so a
    hyper-grower isn't over-penalized for today's (partly growth-investment) opex."""
    ke, n = s.ke, int(s.corp_years)
    pv = sum(s.corp_cost * (1 + s.corp_growth) ** t / (1 + ke) ** t for t in range(1, n + 1))
    corp_n = s.corp_cost * (1 + s.corp_growth) ** n
    terminal = corp_n * s.corp_terminal_mult / (1 + ke) ** n
    return pv + terminal


def value(s: Sotp) -> tuple[float, float]:
    """(SOTP equity $M, value per share $)."""
    lend, fs, tech = seg_vals(s)
    eq = lend + fs + tech - corp_drag(s) - s.net_corp_debt
    return eq, eq / s.shares_m


def _load(ticker: str) -> Sotp:
    """Calibrated defaults; refresh price from FMP profile, then apply any
    per-ticker overrides from data/bank_assumptions/<T>_sotp.json (the SOTP analog
    of the bank model's per-ticker assumptions)."""
    s = Sotp()
    prof = REPO / "data" / "historical" / "fmp" / f"{ticker}_profile.json"
    if prof.exists():
        try:
            d: Any = json.loads(prof.read_text(encoding="utf-8"))
            if isinstance(d, list):
                d = d[0] if d else {}
            if isinstance(d, dict) and d.get("price"):
                s.price = float(d["price"])
                s.price_seed_source = "fmp_profile"
                s.price_seed_path = f"data/historical/fmp/{ticker}_profile.json"
        except (OSError, json.JSONDecodeError, ValueError, KeyError):
            pass
    ov_path = REPO / "data" / "bank_assumptions" / f"{ticker}_sotp.json"
    if ov_path.exists():
        try:
            ov: Any = json.loads(ov_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            ov = {}
        if isinstance(ov, dict):
            for k, v in cast("dict[str, Any]", ov).items():
                if hasattr(s, k) and isinstance(v, (int, float)):
                    setattr(s, k, v)
                    if k == "price":
                        s.price_seed_source = "owner_assumptions"
                        s.price_seed_path = f"data/bank_assumptions/{ticker}_sotp.json"
    # Opt-in: derive ke from the global risk-free/ERP when the name asks for it
    # (after JSON overrides so beta / the flag can be tuned per name).
    global_loaded = (
        global_dcf.load_with_provenance(db_path=REPO / "data" / "portfolio.db")
        if global_dcf is not None
        else None
    )
    if global_loaded is not None:
        s.global_assumption_source = dict(global_loaded.source_record)
        if s.derive_ke_capm:
            s.ke = (
                global_loaded.assumptions.risk_free_rate
                + s.beta * global_loaded.assumptions.equity_risk_premium
            )
    else:
        s.global_assumption_source = {
            "role": "global_dcf_assumptions",
            "status": "module_unavailable",
            "observed_at": None,
        }
    global_fields_effective = bool(s.derive_ke_capm and global_loaded is not None)
    s.global_assumption_source["effective_fields"] = (
        ["risk_free_rate", "equity_risk_premium"] if global_fields_effective else []
    )
    s.global_assumption_source["influences_calculation"] = global_fields_effective
    return s


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _hdr(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].fill = HEAD_FILL
    ws[cell].font = HEAD_FONT


def _inp(ws: Worksheet, row: int, label: str, val: float, fmt: str) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
    c = ws.cell(row=row, column=2, value=val)
    c.fill = YELLOW
    c.number_format = fmt
    c.border = BORDER


# Dashboard input rows (canonical cells the other sheets reference)
R = {
    "lcp": 3,
    "fcp": 4,
    "tcp": 5,
    "lmult": 6,
    "fmult": 7,
    "tmult": 8,
    "ccost": 9,
    "cgrow": 10,
    "cyears": 11,
    "ctermmult": 12,
    "ke": 13,
    "ndebt": 14,
    "sh": 15,
    "px": 16,
    # outputs (linked from Valuation)
    "eq": 19,
    "vps": 20,
    "up": 21,
    "blend": 22,
}


def build(s: Sotp, dest: Path) -> None:
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    seg = wb.create_sheet("Segments")
    corp = wb.create_sheet("Corporate")
    val = wb.create_sheet("Valuation")
    scn = wb.create_sheet("Scenario")

    D = "Dashboard"

    # ---------- Dashboard (inputs) ----------
    _hdr(dash, "A1", f"{T} - Hybrid Sum-of-the-Parts | Dashboard")
    dash["A2"] = "Segment contribution profit ($M, FY2025) x EV/CP multiple"
    dash["A2"].font = SUB_FONT
    _inp(dash, R["lcp"], "Lending - contribution profit ($M)", s.lending_cp, USD0)
    _inp(dash, R["fcp"], "Financial Services - contribution profit ($M)", s.fs_cp, USD0)
    _inp(dash, R["tcp"], "Technology Platform - contribution profit ($M)", s.tech_cp, USD0)
    _inp(dash, R["lmult"], "Lending - EV/CP multiple", s.lending_mult, MULT)
    _inp(dash, R["fmult"], "Financial Services - EV/CP multiple", s.fs_mult, MULT)
    _inp(dash, R["tmult"], "Technology Platform - EV/CP multiple", s.tech_mult, MULT)
    _inp(dash, R["ccost"], "Corporate overhead Y0 ($M/yr)", s.corp_cost, USD0)
    _inp(dash, R["cgrow"], "Corporate growth (vs faster segments)", s.corp_growth, PCT)
    _inp(dash, R["cyears"], "Corporate DCF horizon (yrs)", s.corp_years, "0")
    _inp(dash, R["ctermmult"], "Corporate terminal multiple", s.corp_terminal_mult, MULT)
    _inp(dash, R["ke"], "Cost of equity Ke (discount rate)", s.ke, PCT)
    _inp(dash, R["ndebt"], "Net corporate debt ($M)", s.net_corp_debt, USD0)
    _inp(dash, R["sh"], "Diluted shares (M)", s.shares_m, USD0)
    _inp(dash, R["px"], "Current price ($)", s.price, USD2)

    _hdr(dash, "A18", "OUTPUT")
    for rr, lab, ref, fmt in (
        (R["eq"], "SOTP equity value ($M)", "Valuation!$B$9", USD0),
        (R["vps"], "Value per share ($)", "Valuation!$B$10", USD2),
        (R["up"], "Upside vs price", "Valuation!$B$11", PCT),
        (R["blend"], "Implied blended EV/CP", "Valuation!$B$12", MULT),
    ):
        dash.cell(row=rr, column=1, value=lab).font = SUB_FONT
        oc = dash.cell(row=rr, column=2, value=f"={ref}")
        oc.number_format = fmt
        oc.font = Font(bold=True)
    dash.column_dimensions["A"].width = 42
    dash.column_dimensions["B"].width = 14

    # ---------- Segments (blue actuals + drivers) ----------
    _hdr(seg, "A1", f"{T} - Segments (FY2025 actual, $M)")
    seg["A2"], seg["B2"], seg["C2"], seg["D2"] = (
        "Segment",
        "Net revenue",
        "Contribution profit",
        "CP margin",
    )
    for cc in ("A2", "B2", "C2", "D2"):
        seg[cc].font = SUB_FONT
    seg_rows = [
        ("Lending", 1848.9, s.lending_cp),
        ("Financial Services", 1542.0, s.fs_cp),
        ("Technology Platform (Galileo)", 450.2, s.tech_cp),
    ]
    r = 3
    for name, rev, cp in seg_rows:
        seg.cell(row=r, column=1, value=name).font = Font(color="374151")
        c = seg.cell(row=r, column=2, value=round(rev, 1))
        c.number_format = USD0
        c.font = BLUE_FONT
        c = seg.cell(row=r, column=3, value=round(cp, 1))
        c.number_format = USD0
        c.font = BLUE_FONT
        c = seg.cell(row=r, column=4, value=cp / rev if rev else 0)
        c.number_format = PCT
        c.font = BLUE_FONT
        r += 1
    seg.cell(row=r, column=1, value="Reportable segments total").font = SUB_FONT
    tc = seg.cell(row=r, column=3, value=round(s.lending_cp + s.fs_cp + s.tech_cp, 1))
    tc.number_format = USD0
    tc.font = Font(bold=True)
    r += 2
    _hdr(seg, f"A{r}", "RECONCILIATION + DRIVERS (reference)")
    r += 1
    refs = [
        ("Corporate/Other net revenue ($M)", -227.8),
        ("Adjusted EBITDA ($M, FY2025)", s.adj_ebitda),
        ("GAAP income before taxes ($M)", s.gaap_pretax),
        (
            "Implied corporate/unallocated drag ($M)",
            s.lending_cp + s.fs_cp + s.tech_cp - s.gaap_pretax,
        ),
        ("Total deposits ($B, YE2025)", s.deposits_b),
        ("Total members (M, YE2025)", s.members_m),
        ("SoFi Invest AUM ($B)", s.invest_aum_b),
    ]
    for lab, v in refs:
        seg.cell(row=r, column=1, value=lab).font = Font(color="374151")
        c = seg.cell(row=r, column=2, value=round(v, 1))
        c.number_format = USD0 if abs(v) > 50 else USD2
        c.font = BLUE_FONT
        r += 1
    seg.column_dimensions["A"].width = 40
    for col in ("B", "C", "D"):
        seg.column_dimensions[col].width = 18

    # ---------- Corporate (tapering DCF, formula-first) ----------
    _hdr(corp, "A1", f"{T} - Corporate overhead (tapering DCF, $M)")
    corp["A2"] = "Corporate grows slower than the segments (operating leverage)."
    corp["A2"].font = Font(italic=True, color="6B7280")
    n = int(s.corp_years)
    col0 = 3  # column C = year 1
    corp.cell(row=4, column=2, value="Year").font = SUB_FONT
    for lab_row, lab in ((5, "Corporate cost"), (6, "Discount factor"), (7, "PV corporate")):
        corp.cell(row=lab_row, column=2, value=lab).font = Font(color="374151")
    for j in range(1, n + 1):
        c = get_column_letter(col0 + j - 1)
        corp[f"{c}4"] = j
        corp[f"{c}5"] = f"={D}!$B${R['ccost']}*(1+{D}!$B${R['cgrow']})^{c}4"
        corp[f"{c}6"] = f"=1/(1+{D}!$B${R['ke']})^{c}4"
        corp[f"{c}7"] = f"={c}5*{c}6"
        corp[f"{c}5"].number_format = USD0
        corp[f"{c}6"].number_format = USD2
        corp[f"{c}7"].number_format = USD0
    cN = get_column_letter(col0 + n - 1)
    c1 = get_column_letter(col0)
    corp["B9"] = "PV corporate (yrs 1-N)"
    corp["B10"] = "Terminal corporate PV"
    corp["B11"] = "Total corporate drag"
    corp["C9"] = f"=SUM({c1}7:{cN}7)"
    corp["C10"] = f"={cN}5*{D}!$B${R['ctermmult']}*{cN}6"
    corp["C11"] = "=C9+C10"
    for cc in ("C9", "C10", "C11"):
        corp[cc].number_format = USD0
    corp["B11"].font = SUB_FONT
    corp["C11"].font = Font(bold=True)
    corp.column_dimensions["B"].width = 22

    # ---------- Valuation (formula-first SOTP bridge) ----------
    _hdr(val, "A1", f"{T} - Sum-of-the-Parts ($M)")
    vrows = [
        ("(1) Lending value", f"={D}!$B${R['lcp']}*{D}!$B${R['lmult']}", USD0, 2),
        ("(2) Financial Services value", f"={D}!$B${R['fcp']}*{D}!$B${R['fmult']}", USD0, 3),
        ("(3) Technology Platform value", f"={D}!$B${R['tcp']}*{D}!$B${R['tmult']}", USD0, 4),
        ("Segment value subtotal", "=B2+B3+B4", USD0, 5),
        ("(4) Corporate overhead (tapering DCF)", "=Corporate!C11", USD0, 6),
        ("(5) Net corporate debt", f"={D}!$B${R['ndebt']}", USD0, 7),
        (
            "Total contribution profit",
            f"={D}!$B${R['lcp']}+{D}!$B${R['fcp']}+{D}!$B${R['tcp']}",
            USD0,
            8,
        ),
        ("SOTP equity value (= subtotal - corp - debt)", "=B5-B6-B7", USD0, 9),
        ("Value per share ($)", f"=B9/{D}!$B${R['sh']}", USD2, 10),
        ("Upside vs price", f"=B10/{D}!$B${R['px']}-1", PCT, 11),
        ("Implied blended EV/CP (segments)", "=B5/B8", MULT, 12),
    ]
    for lab, formula, fmt, rr in vrows:
        val.cell(row=rr, column=1, value=lab).font = (
            SUB_FONT if rr in (9, 10) else Font(color="374151")
        )
        vc = val.cell(row=rr, column=2, value=formula)
        vc.number_format = fmt
        if rr in (9, 10):
            vc.font = Font(bold=True)
    val.column_dimensions["A"].width = 40
    val.column_dimensions["B"].width = 14

    # ---------- Scenario + reverse-solve (computed values) ----------
    _hdr(scn, "A1", "Scenarios & reverse-solve")
    scn["A2"], scn["B2"], scn["C2"], scn["D2"], scn["E2"] = (
        "Scenario",
        "Lending x",
        "FS x",
        "Tech x",
        "Value/share",
    )
    for cc in ("A2", "B2", "C2", "D2", "E2"):
        scn[cc].font = SUB_FONT
    scenarios = [
        ("Bear", 5.0, 9.0, 10.0),
        ("Base", s.lending_mult, s.fs_mult, s.tech_mult),
        ("Bull", 9.0, 18.0, 18.0),
    ]
    r = 3
    for name, lm, fm, tm in scenarios:
        import copy

        s2 = copy.copy(s)
        s2.lending_mult, s2.fs_mult, s2.tech_mult = lm, fm, tm
        _, vps = value(s2)
        scn.cell(row=r, column=1, value=name).font = Font(bold=(name == "Base"), color="374151")
        for col, v in zip(("B", "C", "D"), (lm, fm, tm), strict=True):
            cc = scn[f"{col}{r}"]
            cc.value = v
            cc.number_format = MULT
        ec = scn.cell(row=r, column=5, value=round(vps, 2))
        ec.number_format = USD2
        ec.font = Font(bold=(name == "Base"))
        r += 1

    r += 1
    _hdr(scn, f"A{r}", "REVERSE-SOLVE - what the market price implies")
    r += 1
    implied_eq = s.price * s.shares_m
    lend, fs, tech = seg_vals(s)
    seg_subtotal = lend + fs + tech
    total_cp = s.lending_cp + s.fs_cp + s.tech_cp
    implied_seg = implied_eq + corp_drag(s) + s.net_corp_debt
    seg_premium = implied_seg / seg_subtotal - 1
    implied_blend = implied_seg / total_cp
    rs = [
        ("Market equity value ($M)", f"{implied_eq:,.0f}"),
        ("Model segment subtotal ($M)", f"{seg_subtotal:,.0f}"),
        ("Implied segment value at price ($M)", f"{implied_seg:,.0f}"),
        ("-> segment multiples must be higher by", f"{seg_premium:+.0%}"),
        (
            "-> implied blended EV/CP",
            f"{implied_blend:.1f}x  (model {seg_subtotal / total_cp:.1f}x)",
        ),
        ("...or corporate must shrink faster than modeled (operating leverage)", ""),
    ]
    for lab, txt in rs:
        scn.cell(row=r, column=1, value=lab).font = Font(color="374151")
        scn.cell(row=r, column=2, value=txt).font = Font(size=10, color="374151")
        r += 1
    scn.column_dimensions["A"].width = 54
    scn.column_dimensions["B"].width = 26

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def persist_dcf_run(
    s: Sotp,
    eq: float,
    vps: float,
    price_observation: SpecializedPriceObservation | None = None,
    *,
    artifact_promotion: ArtifactPromotion | None = None,
) -> bool:
    """Best-effort upsert into dcf_runs so the brief's valuation panel reads the
    SOTP value/share. No-op without the DB / persist module."""
    db = REPO / "data" / "portfolio.db"
    if persist_mod is None or not db.exists() or not vps:
        return False
    holdings = REPO / "micro_thesis" / "holdings" / f"{T}.json"
    mos: object = None
    if holdings.exists():
        try:
            mos = json.loads(holdings.read_text(encoding="utf-8")).get("mos_bar")
        except (OSError, json.JSONDecodeError):
            mos = None
    lend, fs, tech = seg_vals(s)
    snap_payload: dict[str, object] = {
        "model": "fintech_sotp",
        "value_per_share_usd": vps,
        "sotp_equity_m": eq,
        "lending_value_m": lend,
        "financial_services_value_m": fs,
        "technology_platform_value_m": tech,
        "corporate_drag_m": -corp_drag(s),
        "multiples": {"lending": s.lending_mult, "fs": s.fs_mult, "tech": s.tech_mult},
        "corp_dcf": {
            "y0": s.corp_cost,
            "growth": s.corp_growth,
            "years": s.corp_years,
            "terminal_mult": s.corp_terminal_mult,
            "ke": s.ke,
        },
        "workbook": str(DEST),
    }
    snap = json.dumps(snap_payload, indent=2)
    observed_at = price_observation.observed_at if price_observation is not None else None
    price_source = (
        price_observation.source_name if price_observation is not None else "assumption_seed"
    )
    provenance = build_file_provenance(
        ticker=T,
        repo_root=REPO,
        workbook_path=DEST,
        engine_version="fintech_sotp_v1",
        effective_inputs=asdict(s),
        assumption_snapshot=snap_payload,
        live_price=s.price or None,
        live_price_at=observed_at,
        live_price_source=price_source,
        source_files=(
            (
                REPO / "data" / "bank_assumptions" / f"{T}_sotp.json",
                "owner_assumptions",
            ),
            (REPO / "micro_thesis" / "holdings" / f"{T}.json", "holding_policy"),
            *(
                price_seed_source_files(REPO, price_observation)
                if price_observation is not None
                else ()
            ),
        ),
        source_records=((s.global_assumption_source,) if s.global_assumption_source else ()),
    )
    row = persist_mod.DcfRunRow(
        ticker=T,
        valuation_date=date.today(),
        horizon_years=int(s.corp_years),
        wacc=s.ke,
        npv=eq,
        npv_per_share=vps,
        shares_outstanding=s.shares_m * 1e6,
        currency="USD",
        live_price=s.price or None,
        live_price_at=observed_at,
        mos_bar_used=float(mos) if isinstance(mos, (int, float)) else None,
        assumption_snapshot_json=snap,
        notes=f"workbook={DEST.name} (fintech SOTP)",
        provenance=provenance,
    )
    with connect_sqlite(str(db), role=SQLiteConnectionRole.WRITER, schema_preflight=True) as conn:
        if not schema_supports_provenance(conn):
            row = replace(row, provenance=None)
        if artifact_promotion is None:
            return persist_mod.upsert(conn, row)
        return persist_mod.upsert(conn, row, artifact_promotion=artifact_promotion)


def main() -> int:
    s = _load(T)
    price_observation = resolve_specialized_price(
        REPO,
        T,
        fallback_price=s.price,
        fallback_source_name=s.price_seed_source,
        fallback_source_path=s.price_seed_path,
    )
    s.price = price_observation.price
    eq, vps = value(s)
    build(s, DEST)
    artifact_promotion = promotion_from_env(DEST)
    persisted = (
        persist_dcf_run(
            s,
            eq,
            vps,
            price_observation,
            artifact_promotion=artifact_promotion,
        )
        if artifact_promotion is not None
        else persist_dcf_run(s, eq, vps, price_observation)
    )
    lend, fs, tech = seg_vals(s)
    drag = corp_drag(s)
    up = (vps / s.price - 1) if s.price else 0.0
    print(
        f"RESULT\t{T}\tSOTP/sh=${vps:.2f}\tprice=${s.price:.2f}\tupside={up:+.0%}"
        f"\tdcf_runs={'ok' if persisted else 'skip'}\t-> {DEST}"
    )
    print(
        f"  (1) Lending        CP {s.lending_cp:>7.0f} x {s.lending_mult:.1f}x = ${lend / 1000:6.1f}B"
    )
    print(f"  (2) Financial Svcs CP {s.fs_cp:>7.0f} x {s.fs_mult:.1f}x = ${fs / 1000:6.1f}B")
    print(f"  (3) Tech Platform  CP {s.tech_cp:>7.0f} x {s.tech_mult:.1f}x = ${tech / 1000:6.1f}B")
    print(
        f"  (4) - Corporate DCF (g={s.corp_growth:.0%}, {int(s.corp_years)}y, Ke={s.ke:.1%}) "
        f"= ${-drag / 1000:6.1f}B"
    )
    print(f"  (5) - Net corp debt                     = ${-s.net_corp_debt / 1000:6.1f}B")
    print(f"  = SOTP equity ${eq / 1000:.1f}B  -> ${vps:.2f}/sh  (vs ${s.price:.2f}, {up:+.0%})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
