"""Sum-of-the-parts platform valuation for MercadoLibre (MELI).

A single FCFF DCF mis-frames MELI: ~44% of revenue is Fintech, and inside Fintech
the Mercado Pago credit book is a spread-lending business that consumes regulatory
capital and throws credit losses — an operating-FCFF model treats its receivable
growth as a working-capital drain and prices its earnings on an EBITDA multiple,
both wrong. So value MELI the way it is actually built — three parts, two methods:

    Commerce (3P marketplace + ads + 1P + logistics)   operating FCFF  @ WACC
    Fintech-payments (acquiring, float, fees)           operating FCFF  @ WACC
    Fintech-credit (Mercado Pago credit book)           excess-return / FCFE @ Ke

    Equity value = Operating EV (Commerce + Fintech-payments)
                 + Credit-book equity value (FCFE on the lending franchise)
                 + net non-operating cash
    per share    = Equity value / diluted shares

The credit book is valued on its own cost of equity (lending is riskier than the
capital-light operating franchises) and charged the growth in required capital it
consumes — the capital intensity an FCFF model can't see. The operating side fades
growth on the same CONVEX curve as the redesigned FCFF engine
(``dcf.redesign.GROWTH_FADE_CURVATURE``), so the two models speak the same language.

Base case grounded in MELI's FY2025 10-K segment note (Commerce $16,294M + Fintech
$12,599M = $28,893M total, reconciling exactly), with Fintech split into
payments/services (~$6,741M, capital-light) and credit/financial income (~$5,858M,
the lending book). Segment operating margins are modelling assumptions (MELI
reports profit by geography, not by Commerce/Fintech), anchored so the blend tracks
the consolidated 11.1% → high-teens path. Every driver is editable; a Python mirror
is the value-of-record and matches the in-sheet formulas exactly.

Env (like build_nu_platform_dcf.py): DCF_TICKER, DCF_DEST, DCF_REPO_ROOT. Values in
$M; shares in millions. Discount rates carry the revenue-weighted Damodaran country
risk premium (``dcf.country_risk``), consistent with the FCFF engine.
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(os.environ.get("DCF_REPO_ROOT") or Path(__file__).resolve().parents[1])
T = os.environ.get("DCF_TICKER", "MELI")
DEST = Path(os.environ.get("DCF_DEST") or (REPO / "dcf" / f"{T}.xlsx"))

# Import co-located package code (this script's repo), NOT REPO/src — REPO points
# at the DATA repo (which may be a different checkout, e.g. a worktree's data lives
# in the main repo), so resolving code from it would load a stale/foreign copy.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))


from sqlite_runtime import SQLiteConnectionRole, connect_sqlite  # noqa: E402

try:  # persistence is best-effort -- the workbook builds without a DB
    from dcf import persist as persist_mod
except ImportError:  # pragma: no cover
    persist_mod = None  # type: ignore[assignment]
try:  # global macro assumptions + country risk -- best-effort; in-code defaults else
    from dcf import country_risk
    from dcf import global_assumptions as global_dcf
except ImportError:  # pragma: no cover
    country_risk = None  # type: ignore[assignment]
    global_dcf = None  # type: ignore[assignment]

# Growth fade curvature — kept identical to the redesigned FCFF engine so the two
# models decelerate growth the same way (convex, front-loaded). Imported when the
# package is available; falls back to the same literal.
try:
    from dcf.redesign import GROWTH_FADE_CURVATURE as _CURVATURE
except ImportError:  # pragma: no cover
    _CURVATURE = 2.0
try:  # scenario emission (Monthly Red Team PR8) -- best-effort like persistence
    from dcf import redesign as redesign_mod
except ImportError:  # pragma: no cover
    redesign_mod = None  # type: ignore[assignment]

YELLOW = PatternFill("solid", fgColor="FFF2CC")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True)
SUB_FONT = Font(bold=True, color="374151")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
USD0 = "#,##0"
NUM1 = "0.0"
NUM2 = "0.00"
MULT = '0.0"x"'


@dataclass
class Assum:
    """Base case grounded in MELI's FY2025 10-K segment note. All $M; shares in
    millions. Growth fades CONVEXLY near→terminal; margins/NIMAL ramp linearly."""

    # ---- Commerce (3P marketplace + ads + 1P + logistics) — operating FCFF ----
    comm_rev0: float = 16294.0  # FY2025 Commerce net revenue
    comm_g_near: float = 0.22  # near-term growth (GMV + ads + 1P), fades convexly
    comm_g_term: float = 0.05
    comm_margin_near: float = 0.12  # EBIT margin Y1 (compressed: free-shipping, 1P, AI invest)
    comm_margin_term: float = 0.20  # mature margin (logistics density + ads mix)

    # ---- Fintech-payments (acquiring, float, fees) — operating FCFF, capital-light ----
    fpay_rev0: float = 6741.0  # FY2025 Fintech ex-credit (services $6,678M + product $63M)
    fpay_g_near: float = 0.26
    fpay_g_term: float = 0.06
    fpay_margin_near: float = 0.16
    fpay_margin_term: float = 0.26  # capital-light take-rate + float economics

    # ---- shared operating drivers ----
    da_pct: float = 0.04  # D&A as % of operating revenue
    capex_pct_near: float = 0.055  # logistics build-out heavy early
    capex_pct_term: float = 0.035
    nwc_pct: float = 0.02  # working-capital draw on incremental operating revenue
    tax: float = 0.28
    op_exit_ebitda_mult: float = 12.0  # terminal EV/EBITDA on the operating block
    wacc: float = 0.135  # operating discount rate (set from CAPM+CRP when derive_capm)

    # ---- Fintech-credit (Mercado Pago credit book) — excess-return / FCFE ----
    cb0: float = 13000.0  # credit portfolio Y0 ($M, ~FY2025 end; Q1'26 $14.6B)
    cbg_near: float = 0.34  # book growth (credit card + consumer), fades convexly
    cbg_term: float = 0.08
    nimal_near: float = 0.178  # net interest margin AFTER losses (Q1'26 17.8%)
    nimal_term: float = 0.150  # PM floor; card-mix compresses the spread
    credit_opex_ratio: float = 0.05  # credit-specific opex (origination/servicing) as % of book
    cap_ratio: float = 0.15  # required equity / credit book
    credit_ke: float = 0.16  # cost of equity for the lending franchise (riskier)
    credit_g_term: float = 0.08
    credit_terminal_roe: float = 0.25  # sustainable ROE on the credit book

    # ---- bridge / discounting ----
    net_cash: float = 0.0  # corporate net cash EXCL. credit funding + payment float (editable)
    g_term: float = 0.045  # operating perpetuity-cross-check growth (~risk-free)
    years: int = 10
    shares: float = 50.697  # diluted shares (M)
    price: float = 1684.0

    # ---- opt-in CAPM discount rates from the editable global rf/ERP + country CRP ----
    # When derive_capm != 0, wacc and credit_ke are recomputed from rf + beta*erp + crp
    # so a dashboard macro change flows through. Off by default (explicit scalars win).
    beta_op: float = 1.30
    beta_credit: float = 1.55
    country_risk_premium: float = 0.0  # filled from dcf.country_risk at load when 0
    derive_capm: int = 1


def _interp(near: float, term: float, t: int, n: int) -> float:
    """Linear ramp from ``near`` (year 1) to ``term`` (year n)."""
    return near if n <= 1 else near + (term - near) * (t - 1) / (n - 1)


def _fade(near: float, term: float, t: int, n: int) -> float:
    """Convex growth fade: ``term + (near-term)*((n-t)/(n-1))**curvature``.

    Year 1 = near, year n = term, front-loaded deceleration — identical in shape
    to ``dcf.redesign``'s fade so the SOTP operating block and the FCFF engine
    decelerate the same way.
    """
    if n <= 1:
        return near
    frac = ((n - t) / (n - 1)) ** _CURVATURE
    return term + (near - term) * frac


@dataclass
class Row:
    t: int
    # operating
    comm_rev: float
    fpay_rev: float
    op_rev: float
    op_ebit: float
    op_da: float
    op_capex: float
    op_fcff: float
    # credit
    cb: float
    credit_ni: float
    reqcap: float
    credit_fcfe: float
    # discounting
    df_op: float
    df_cr: float


@dataclass
class Mirror:
    rows: list[Row] = field(default_factory=list)
    pv_op_fcff: float = 0.0
    op_terminal_ev: float = 0.0
    pv_op_terminal: float = 0.0
    operating_ev: float = 0.0
    pv_credit_fcfe: float = 0.0
    credit_terminal: float = 0.0
    pv_credit_terminal: float = 0.0
    credit_equity_value: float = 0.0
    equity_value: float = 0.0
    vps: float = 0.0
    # cross-checks
    op_terminal_revenue: float = 0.0
    credit_terminal_roe: float = 0.0
    terminal_blended_op_margin: float = 0.0


def mirror(s: Assum) -> Mirror:
    n = s.years
    m = Mirror()
    comm_p, fpay_p, op_rev_p = s.comm_rev0, s.fpay_rev0, s.comm_rev0 + s.fpay_rev0
    cb_p = s.cb0
    reqcap_p = s.cap_ratio * s.cb0
    for t in range(1, n + 1):
        # --- operating: Commerce + Fintech-payments -> FCFF ---
        comm_rev = comm_p * (1 + _fade(s.comm_g_near, s.comm_g_term, t, n))
        fpay_rev = fpay_p * (1 + _fade(s.fpay_g_near, s.fpay_g_term, t, n))
        op_rev = comm_rev + fpay_rev
        op_ebit = comm_rev * _interp(
            s.comm_margin_near, s.comm_margin_term, t, n
        ) + fpay_rev * _interp(s.fpay_margin_near, s.fpay_margin_term, t, n)
        op_da = op_rev * s.da_pct
        op_capex = op_rev * _interp(s.capex_pct_near, s.capex_pct_term, t, n)
        op_dnwc = (op_rev - op_rev_p) * s.nwc_pct
        op_fcff = op_ebit * (1 - s.tax) + op_da - op_capex - op_dnwc
        df_op = 1 / (1 + s.wacc) ** t

        # --- credit book: NIMAL spread, capital-charged -> FCFE ---
        cb = cb_p * (1 + _fade(s.cbg_near, s.cbg_term, t, n))
        avg_book = (cb + cb_p) / 2.0
        nimal = _interp(s.nimal_near, s.nimal_term, t, n)
        credit_pretax = avg_book * (nimal - s.credit_opex_ratio)
        credit_ni = credit_pretax * (1 - s.tax)
        reqcap = s.cap_ratio * cb
        credit_fcfe = credit_ni - (reqcap - reqcap_p)
        df_cr = 1 / (1 + s.credit_ke) ** t

        m.rows.append(
            Row(
                t,
                comm_rev,
                fpay_rev,
                op_rev,
                op_ebit,
                op_da,
                op_capex,
                op_fcff,
                cb,
                credit_ni,
                reqcap,
                credit_fcfe,
                df_op,
                df_cr,
            )
        )
        m.pv_op_fcff += op_fcff * df_op
        m.pv_credit_fcfe += credit_fcfe * df_cr
        comm_p, fpay_p, op_rev_p, cb_p, reqcap_p = comm_rev, fpay_rev, op_rev, cb, reqcap

    last = m.rows[-1]
    # operating terminal: EV/EBITDA exit on terminal operating EBITDA
    m.op_terminal_ev = (last.op_ebit + last.op_da) * s.op_exit_ebitda_mult
    m.pv_op_terminal = m.op_terminal_ev * last.df_op
    m.operating_ev = m.pv_op_fcff + m.pv_op_terminal

    # credit terminal: sustainable Gordon on credit NI (reinvest g/ROE)
    ni_n1 = last.credit_ni * (1 + s.credit_g_term)
    m.credit_terminal = (
        ni_n1 * (1 - s.credit_g_term / s.credit_terminal_roe) / (s.credit_ke - s.credit_g_term)
    )
    m.pv_credit_terminal = m.credit_terminal * last.df_cr
    m.credit_equity_value = m.pv_credit_fcfe + m.pv_credit_terminal

    m.equity_value = m.operating_ev + s.net_cash + m.credit_equity_value
    m.vps = m.equity_value / s.shares if s.shares else 0.0
    m.op_terminal_revenue = last.op_rev
    m.credit_terminal_roe = (last.credit_ni / last.reqcap) if last.reqcap else 0.0
    m.terminal_blended_op_margin = (last.op_ebit / last.op_rev) if last.op_rev else 0.0
    return m


def load_assumptions(ticker: str) -> Assum:
    """Assum defaults overridden by data/bank_assumptions/<T>_sotp.json, then the
    editable global tax + (opt-in) CAPM-derived discount rates with the
    revenue-weighted Damodaran country risk premium."""
    s = Assum()
    db = REPO / "data" / "portfolio.db"
    if global_dcf is not None:
        s.tax = global_dcf.load(db_path=db).tax_rate
    # Systematic country risk premium (revenue-weighted), filled when unset.
    if country_risk is not None and not s.country_risk_premium:
        s.country_risk_premium = country_risk.country_risk_premium(REPO, ticker)
    p = REPO / "data" / "bank_assumptions" / f"{ticker}_sotp.json"
    if p.exists():
        try:
            ov: Any = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            ov = {}
        if isinstance(ov, dict):
            for k, v in cast("dict[str, Any]", ov).items():
                if hasattr(s, k) and isinstance(v, (int, float)):
                    setattr(s, k, v)
    # Opt-in: derive both discount rates from the editable global rf/ERP + CRP.
    if global_dcf is not None and s.derive_capm:
        s.wacc = global_dcf.capm_ke(
            s.beta_op, country_risk_premium=s.country_risk_premium, db_path=db
        )
        s.credit_ke = global_dcf.capm_ke(
            s.beta_credit, country_risk_premium=s.country_risk_premium, db_path=db
        )
    prof = REPO / "data" / "historical" / "fmp" / f"{ticker}_profile.json"
    if prof.exists():
        try:
            d: Any = json.loads(prof.read_text(encoding="utf-8"))
            if isinstance(d, list):
                d = d[0] if d else {}
            if isinstance(d, dict) and d.get("price"):
                s.price = float(d["price"])
        except (OSError, json.JSONDecodeError, ValueError, KeyError):
            pass
    return s


# --------------------------------------------------------------------------- #
# Scenario emission (Monthly Red Team PR8).
#
# Mirrors the redesigned refresher's persisted ``scenarios`` block (bull/base/
# bear fair values + bear provenance) so MELI stops being invisible to every
# scenario consumer (``dcf.scenario_reward``, ``bear_lint``, tail stress). The
# shared 6-lever ``ScenarioDeltas`` vocabulary maps onto THIS SOTP's levers:
#
#   growth deltas   -> Commerce, Fintech-payments AND credit-book growth
#                      (near/terminal) — a demand shock hits all three engines
#   margin deltas   -> Commerce + Fintech-payments EBIT margins AND the credit
#                      book's NIMAL (near/terminal) — a price war compresses the
#                      operating segments while a credit cycle compresses NIMAL
#   exit multiple Δ -> the operating exit EV/EBITDA (turns, floored at 1x). The
#                      credit leg's Gordon terminal compresses through NIMAL/g,
#                      not through this multiple
#   terminal g Δ    -> credit terminal growth g (and the operating perpetuity
#                      cross-check g, cosmetic)
#
# Bear deltas: holdings ``bear_deltas`` when present (provenance "thesis"), else
# the generic BEAR_SEED (provenance "seed" — a labeled fallback). Legible over
# precise, per the red-team directive.
# --------------------------------------------------------------------------- #
def _load_holdings(ticker: str) -> dict[str, object] | None:
    path = REPO / "micro_thesis" / "holdings" / f"{ticker.upper()}.json"
    if not path.exists():
        return None
    try:
        data: Any = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return cast("dict[str, object]", data) if isinstance(data, dict) else None


def scenario_assumptions(s: Assum, deltas: Any) -> Assum:
    """``s`` shifted by one scenario's ``dcf.redesign.ScenarioDeltas`` under the
    documented lever mapping. Guardrails keep the credit Gordon terminal
    well-posed (g < ke, ROE > g) and the exit multiple positive."""
    import copy

    s2 = copy.copy(s)
    for attr in ("comm_g_near", "fpay_g_near", "cbg_near"):
        setattr(s2, attr, getattr(s2, attr) + deltas.growth_near)
    for attr in ("comm_g_term", "fpay_g_term", "cbg_term"):
        setattr(s2, attr, getattr(s2, attr) + deltas.growth_term)
    for attr in ("comm_margin_near", "fpay_margin_near", "nimal_near"):
        setattr(s2, attr, getattr(s2, attr) + deltas.margin_near)
    for attr in ("comm_margin_term", "fpay_margin_term", "nimal_term"):
        setattr(s2, attr, getattr(s2, attr) + deltas.margin_term)
    s2.op_exit_ebitda_mult = max(1.0, s2.op_exit_ebitda_mult + deltas.exit_multiple)
    s2.credit_g_term += deltas.terminal_g
    s2.g_term += deltas.terminal_g
    s2.credit_g_term = min(s2.credit_g_term, s2.credit_ke - 0.01)
    s2.credit_terminal_roe = max(s2.credit_terminal_roe, s2.credit_g_term + 0.01)
    return s2


def scenarios_block(s: Assum, m: Mirror, holdings: dict[str, object] | None) -> dict[str, object]:
    """The ``scenarios`` payload for ``dcf_runs.assumption_snapshot_json`` —
    structurally identical to ``refresh_dcf._redesign_snapshot``'s block, so
    ``parse_scenario_fair_values`` / ``parse_scenario_bear_provenance`` read it
    unchanged. Requires ``redesign_mod`` (caller gates on it)."""
    import dataclasses as _dc

    bull_d = redesign_mod.BULL_SEED
    bear_d = redesign_mod.thesis_bear_seed(holdings)
    provenance = "thesis" if redesign_mod.parse_thesis_bear_deltas(holdings) is not None else "seed"
    bull_vps = mirror(scenario_assumptions(s, bull_d)).vps
    bear_vps = mirror(scenario_assumptions(s, bear_d)).vps
    return {
        "base": {"fair_value_per_share_usd": m.vps},
        "bull": {"fair_value_per_share_usd": bull_vps, "deltas": _dc.asdict(bull_d)},
        "bear": {
            "fair_value_per_share_usd": bear_vps,
            "deltas": _dc.asdict(bear_d),
            "provenance": provenance,
        },
    }


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _hdr(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].fill = HEAD_FILL
    ws[cell].font = HEAD_FONT


def _inp(ws: Worksheet, row: int, label: str, val: float, fmt: str) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
    c = ws.cell(row=row, column=2, value=val)
    c.fill = YELLOW
    c.number_format = fmt
    c.border = BORDER


# Dashboard input rows (column B = the editable yellow cell).
R = {
    "comm_rev0": 3,
    "comm_gn": 4,
    "comm_gt": 5,
    "comm_mn": 6,
    "comm_mt": 7,
    "fpay_rev0": 8,
    "fpay_gn": 9,
    "fpay_gt": 10,
    "fpay_mn": 11,
    "fpay_mt": 12,
    "da": 13,
    "capxn": 14,
    "capxt": 15,
    "nwc": 16,
    "tax": 17,
    "opmult": 18,
    "wacc": 19,
    "cb0": 20,
    "cbgn": 21,
    "cbgt": 22,
    "niman": 23,
    "nimat": 24,
    "cropex": 25,
    "cap": 26,
    "cke": 27,
    "cgt": 28,
    "croe": 29,
    "netcash": 30,
    "years": 31,
    "sh": 32,
    "px": 33,
    "curv": 34,
    # outputs
    "opev": 37,
    "crev": 38,
    "eqv": 39,
    "vps": 40,
    "up": 41,
    "termrev": 42,
    "croet": 43,
}


def build(s: Assum, m: Mirror, dest: Path, holdings: dict[str, object] | None = None) -> None:
    """``holdings`` (Monthly Red Team PR10) is the already-loaded
    ``micro_thesis/holdings/<T>.json`` dict, threaded through so the Scenario
    sheet's Bear/Bull rows derive from the SAME ``scenario_assumptions()`` call
    ``scenarios_block`` uses for the persisted ``dcf_runs`` snapshot — the sheet
    can no longer show a bear the rest of the platform (bear_lint, tail stress,
    red-team evidence packs) disagrees with. ``None`` (no holdings on file / not
    passed) degrades to the generic ``BEAR_SEED`` fallback, same as
    ``scenarios_block``."""
    wb = openpyxl.Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    mod = wb.create_sheet("Model")
    val = wb.create_sheet("Valuation")
    scn = wb.create_sheet("Scenario")
    D = "Dashboard"

    # ---------- Dashboard ----------
    _hdr(dash, "A1", f"{T} - Sum-of-the-Parts Platform DCF | Dashboard")
    dash["A2"] = (
        "Commerce + Fintech-payments (operating FCFF @ WACC) + credit book (excess-return @ Ke)"
    )
    dash["A2"].font = SUB_FONT
    rows = [
        ("comm_rev0", "Commerce revenue Y0 ($M)", s.comm_rev0, USD0),
        ("comm_gn", "Commerce growth - near", s.comm_g_near, PCT),
        ("comm_gt", "Commerce growth - terminal", s.comm_g_term, PCT),
        ("comm_mn", "Commerce EBIT margin - near", s.comm_margin_near, PCT),
        ("comm_mt", "Commerce EBIT margin - terminal", s.comm_margin_term, PCT),
        ("fpay_rev0", "Fintech-payments revenue Y0 ($M)", s.fpay_rev0, USD0),
        ("fpay_gn", "Fintech-pay growth - near", s.fpay_g_near, PCT),
        ("fpay_gt", "Fintech-pay growth - terminal", s.fpay_g_term, PCT),
        ("fpay_mn", "Fintech-pay EBIT margin - near", s.fpay_margin_near, PCT),
        ("fpay_mt", "Fintech-pay EBIT margin - terminal", s.fpay_margin_term, PCT),
        ("da", "D&A % of operating revenue", s.da_pct, PCT),
        ("capxn", "Capex % rev - near", s.capex_pct_near, PCT),
        ("capxt", "Capex % rev - terminal", s.capex_pct_term, PCT),
        ("nwc", "Working capital % of incr. rev", s.nwc_pct, PCT),
        ("tax", "Tax rate", s.tax, PCT),
        ("opmult", "Operating exit EV/EBITDA", s.op_exit_ebitda_mult, MULT),
        ("wacc", "WACC (operating discount)", s.wacc, PCT),
        ("cb0", "Credit book Y0 ($M)", s.cb0, USD0),
        ("cbgn", "Credit book growth - near", s.cbg_near, PCT),
        ("cbgt", "Credit book growth - terminal", s.cbg_term, PCT),
        ("niman", "NIMAL - near", s.nimal_near, PCT),
        ("nimat", "NIMAL - terminal", s.nimal_term, PCT),
        ("cropex", "Credit opex % of book", s.credit_opex_ratio, PCT),
        ("cap", "Capital ratio (req eq / book)", s.cap_ratio, PCT),
        ("cke", "Credit cost of equity Ke", s.credit_ke, PCT),
        ("cgt", "Credit terminal growth g", s.credit_g_term, PCT),
        ("croe", "Credit terminal ROE", s.credit_terminal_roe, PCT),
        ("netcash", "Net non-operating cash ($M)", s.net_cash, USD0),
        ("years", "Forecast years", s.years, "0"),
        ("sh", "Diluted shares (M)", s.shares, NUM1),
        ("px", "Current price ($)", s.price, NUM2),
        ("curv", "Growth fade curvature", _CURVATURE, NUM1),
    ]
    for key, lab, v, fmt in rows:
        _inp(dash, R[key], lab, v, fmt)

    _hdr(dash, "A36", "OUTPUT")
    for rr, lab, ref, fmt in (
        (R["opev"], "Operating EV ($M)", "Valuation!$B$5", USD0),
        (R["crev"], "Credit equity value ($M)", "Valuation!$B$9", USD0),
        (R["eqv"], "Equity value ($M)", "Valuation!$B$11", USD0),
        (R["vps"], "Value per share ($)", "Valuation!$B$12", NUM2),
        (R["up"], "Upside vs price", "Valuation!$B$13", PCT),
        (R["termrev"], "Terminal operating revenue ($M)", "Valuation!$B$14", USD0),
        (R["croet"], "Terminal credit ROE", "Valuation!$B$15", PCT),
    ):
        dash.cell(row=rr, column=1, value=lab).font = SUB_FONT
        oc = dash.cell(row=rr, column=2, value=f"={ref}")
        oc.number_format = fmt
        oc.font = Font(bold=True)
    dash.column_dimensions["A"].width = 36
    dash.column_dimensions["B"].width = 14

    # ---------- Model (formula-first engine) ----------
    _hdr(mod, "A1", f"{T} - SOTP engine (formula-first; $M)")
    n = s.years
    col0 = 3  # column C = Y1

    def cl(i: int) -> str:
        return get_column_letter(i)

    labels = {
        3: "Year",
        4: "Commerce revenue",
        5: "Fintech-pay revenue",
        6: "Operating revenue",
        7: "Operating EBIT",
        8: "  D&A",
        9: "  Capex",
        10: "Operating FCFF",
        11: "DF (WACC)",
        12: "PV operating FCFF",
        13: "Credit book",
        14: "Credit NI",
        15: "Required capital",
        16: "Credit FCFE",
        17: "DF (Ke)",
        18: "PV credit FCFE",
    }
    for r, lab in labels.items():
        mod.cell(row=r, column=2, value=lab).font = SUB_FONT if r == 3 else Font(color="374151")

    def dref(key: str) -> str:
        return f"{D}!$B${R[key]}"

    def fade_f(c: str, near_key: str, term_key: str) -> str:
        """Excel convex-fade formula mirroring _fade for the year in row 3 of col c."""
        near, term = dref(near_key), dref(term_key)
        yrs = dref("years")
        return f"{term}+({near}-{term})*(({yrs}-{c}$3)/({yrs}-1))^{dref('curv')}"

    def interp_f(c: str, near_key: str, term_key: str) -> str:
        near, term = dref(near_key), dref(term_key)
        return f"{near}+({term}-{near})*({c}$3-1)/({dref('years')}-1)"

    for j in range(1, n + 1):
        c = cl(col0 + j - 1)
        p = cl(col0 + j - 2)  # prior column (j==1 references Dashboard Y0)
        mod[f"{c}3"] = j
        if j == 1:
            mod[f"{c}4"] = f"={dref('comm_rev0')}*(1+{fade_f(c, 'comm_gn', 'comm_gt')})"
            mod[f"{c}5"] = f"={dref('fpay_rev0')}*(1+{fade_f(c, 'fpay_gn', 'fpay_gt')})"
            mod[f"{c}13"] = f"={dref('cb0')}*(1+{fade_f(c, 'cbgn', 'cbgt')})"
            prev_oprev = f"({dref('comm_rev0')}+{dref('fpay_rev0')})"
            prev_cb = dref("cb0")
            prev_reqcap = f"{dref('cap')}*{dref('cb0')}"
        else:
            mod[f"{c}4"] = f"={p}4*(1+{fade_f(c, 'comm_gn', 'comm_gt')})"
            mod[f"{c}5"] = f"={p}5*(1+{fade_f(c, 'fpay_gn', 'fpay_gt')})"
            mod[f"{c}13"] = f"={p}13*(1+{fade_f(c, 'cbgn', 'cbgt')})"
            prev_oprev = f"{p}6"
            prev_cb = f"{p}13"
            prev_reqcap = f"{p}15"
        mod[f"{c}6"] = f"={c}4+{c}5"
        mod[f"{c}7"] = (
            f"={c}4*({interp_f(c, 'comm_mn', 'comm_mt')})+{c}5*({interp_f(c, 'fpay_mn', 'fpay_mt')})"
        )
        mod[f"{c}8"] = f"={c}6*{dref('da')}"
        mod[f"{c}9"] = f"={c}6*({interp_f(c, 'capxn', 'capxt')})"
        mod[f"{c}10"] = f"={c}7*(1-{dref('tax')})+{c}8-{c}9-({c}6-{prev_oprev})*{dref('nwc')}"
        mod[f"{c}11"] = f"=1/(1+{dref('wacc')})^{c}3"
        mod[f"{c}12"] = f"={c}10*{c}11"
        # credit: NI on avg book at NIMAL net of opex; FCFE charges capital growth
        mod[f"{c}14"] = (
            f"=(({c}13+{prev_cb})/2)*(({interp_f(c, 'niman', 'nimat')})-{dref('cropex')})*(1-{dref('tax')})"
        )
        mod[f"{c}15"] = f"={dref('cap')}*{c}13"
        mod[f"{c}16"] = f"={c}14-({c}15-{prev_reqcap})"
        mod[f"{c}17"] = f"=1/(1+{dref('cke')})^{c}3"
        mod[f"{c}18"] = f"={c}16*{c}17"
    for r in range(4, 19):
        for j in range(1, n + 1):
            cc = mod.cell(row=r, column=col0 + j - 1)
            cc.number_format = NUM2 if r in (11, 17) else USD0
    mod.column_dimensions["B"].width = 20

    # ---------- Valuation (SOTP sum) ----------
    _hdr(val, "A1", f"{T} - Sum-of-the-Parts ($M)")
    cN = cl(col0 + n - 1)
    c1 = cl(col0)
    vrows = [
        ("PV operating FCFF (yrs 1-N)", f"=SUM(Model!{c1}12:{cN}12)", USD0, 2),
        (
            "Operating terminal EV (EV/EBITDA)",
            f"=(Model!{cN}7+Model!{cN}8)*{D}!$B${R['opmult']}",
            USD0,
            3,
        ),
        ("PV operating terminal", f"=B3*Model!{cN}11", USD0, 4),
        ("Operating EV", "=B2+B4", USD0, 5),
        ("PV credit FCFE (yrs 1-N)", f"=SUM(Model!{c1}18:{cN}18)", USD0, 6),
        (
            "Credit terminal value (sustainable)",
            f"=Model!{cN}14*(1+{D}!$B${R['cgt']})*(1-{D}!$B${R['cgt']}/{D}!$B${R['croe']})/({D}!$B${R['cke']}-{D}!$B${R['cgt']})",
            USD0,
            7,
        ),
        ("PV credit terminal", f"=B7*Model!{cN}17", USD0, 8),
        ("Credit equity value", "=B6+B8", USD0, 9),
        ("+ Net non-operating cash", f"={D}!$B${R['netcash']}", USD0, 10),
        ("Equity value", "=B5+B9+B10", USD0, 11),
        ("Value per share ($)", f"=B11/{D}!$B${R['sh']}", NUM2, 12),
        ("Upside vs price", f"=B12/{D}!$B${R['px']}-1", PCT, 13),
        ("Terminal operating revenue", f"=Model!{cN}6", USD0, 14),
        ("Terminal credit ROE", f"=Model!{cN}14/Model!{cN}15", PCT, 15),
    ]
    for lab, formula, fmt, rr in vrows:
        val.cell(row=rr, column=1, value=lab).font = (
            SUB_FONT if rr in (5, 9, 11, 12) else Font(color="374151")
        )
        vc = val.cell(row=rr, column=2, value=formula)
        vc.number_format = fmt
        if rr in (11, 12):
            vc.font = Font(bold=True)
    val.column_dimensions["A"].width = 36
    val.column_dimensions["B"].width = 14

    # ---------- Scenario ----------
    _hdr(scn, "A1", "Scenarios & sum-of-the-parts bridge")
    scn["A2"], scn["B2"], scn["C2"], scn["D2"], scn["E2"] = (
        "Scenario",
        "Commerce g (near)",
        "Comm margin (term)",
        "NIMAL (term)",
        "Value/share",
    )
    for cc in ("A2", "B2", "C2", "D2", "E2"):
        scn[cc].font = SUB_FONT
    r = 3
    if redesign_mod is not None:
        # Monthly Red Team PR10: Bear/Bull rows come from the SAME
        # scenario_assumptions() call scenarios_block() uses for the persisted
        # dcf_runs snapshot (bear deltas from holdings bear_deltas when present,
        # provenance "thesis"; else the generic BEAR_SEED, provenance "seed") —
        # this sheet and the snapshot every risk consumer reads (bear_lint,
        # tail stress, red-team evidence packs) can no longer disagree.
        bull_d = redesign_mod.BULL_SEED
        bear_d = redesign_mod.thesis_bear_seed(holdings)
        bear_provenance = (
            "thesis" if redesign_mod.parse_thesis_bear_deltas(holdings) is not None else "seed"
        )
        scenario_rows: list[tuple[str, Assum]] = [
            ("Bear", scenario_assumptions(s, bear_d)),
            ("Base", s),
            ("Bull", scenario_assumptions(s, bull_d)),
        ]
        for name, s2 in scenario_rows:
            v = mirror(s2).vps
            scn.cell(row=r, column=1, value=name).font = Font(bold=(name == "Base"), color="374151")
            for col, vv in zip(
                ("B", "C", "D"),
                (s2.comm_g_near, s2.comm_margin_term, s2.nimal_term),
                strict=True,
            ):
                cc = scn[f"{col}{r}"]
                cc.value = vv
                cc.number_format = PCT
            ec = scn.cell(row=r, column=5, value=round(v, 2))
            ec.number_format = NUM2
            ec.font = Font(bold=(name == "Base"))
            r += 1
        prov_label = (
            "bear from holdings bear_deltas (thesis)"
            if bear_provenance == "thesis"
            else "bear from generic BEAR_SEED (seed fallback -- no holdings bear_deltas on file)"
        )
        scn.cell(row=r, column=1, value="Bear provenance").font = Font(italic=True, color="6B7280")
        prov_cell = scn.cell(row=r, column=2, value=prov_label)
        prov_cell.font = Font(italic=True, color="6B7280", size=9)
        scn.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1
    else:  # pragma: no cover - import failure only, exercised by no test env
        # dcf.redesign unavailable: degrade LOUDLY (an empty/stale scenario row
        # is exactly the PR10 bug) rather than silently falling back to the old
        # hardcoded Bear/Bull levers.
        scn.cell(
            row=r, column=1, value="Scenarios unavailable (dcf.redesign import failed)"
        ).font = Font(italic=True, color="B91C1C")
        r += 1
    r += 1
    _hdr(scn, f"A{r}", "SUM-OF-THE-PARTS BRIDGE ($M)")
    r += 1
    bridge = [
        ("Operating EV (Commerce + Fintech-pay)", f"{m.operating_ev:,.0f}"),
        ("Credit-book equity value (FCFE @ Ke)", f"{m.credit_equity_value:,.0f}"),
        ("Net non-operating cash", f"{s.net_cash:,.0f}"),
        ("= Equity value", f"{m.equity_value:,.0f}"),
        ("Value per share", f"${m.vps:,.2f}  (vs ${s.price:,.2f}, {m.vps / s.price - 1:+.0%})"),
        ("Terminal credit ROE", f"{m.credit_terminal_roe:.0%}"),
        ("Terminal blended op margin", f"{m.terminal_blended_op_margin:.0%}"),
    ]
    for lab, txt in bridge:
        scn.cell(row=r, column=1, value=lab).font = Font(color="374151")
        scn.cell(row=r, column=2, value=txt).font = Font(size=10, color="374151")
        r += 1
    scn.column_dimensions["A"].width = 40
    scn.column_dimensions["B"].width = 26

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def persist_dcf_run(s: Assum, m: Mirror, holdings: dict[str, object] | None = None) -> bool:
    """``holdings=None`` (the pre-PR10 2-arg call shape every test/caller uses)
    loads ``micro_thesis/holdings/<T>.json`` itself, same as before. ``main()``
    now passes the SAME dict ``build()``'s Scenario sheet used, so a mid-run
    file edit can never make the sheet and the persisted snapshot disagree —
    a ticker with genuinely no holdings JSON still resolves to ``None`` either
    way, so this collapses "not passed" and "no holdings on file" safely."""
    db = REPO / "data" / "portfolio.db"
    if persist_mod is None or not db.exists() or not m.vps:
        return False
    if holdings is None:
        holdings = _load_holdings(T)
    mos: object = holdings.get("mos_bar") if holdings else None
    snap_payload: dict[str, object] = {
        "model": "meli_platform_sotp",
        "value_per_share": m.vps,
        "operating_ev_m": m.operating_ev,
        "credit_equity_value_m": m.credit_equity_value,
        "equity_value_m": m.equity_value,
        "terminal_operating_revenue_m": m.op_terminal_revenue,
        "terminal_credit_roe": m.credit_terminal_roe,
        "terminal_blended_op_margin": m.terminal_blended_op_margin,
        "wacc": s.wacc,
        "credit_ke": s.credit_ke,
        "country_risk_premium": s.country_risk_premium,
        "workbook": str(DEST),
    }
    if redesign_mod is not None:
        snap_payload["scenarios"] = scenarios_block(s, m, holdings)
    snap = json.dumps(snap_payload, indent=2)
    row = persist_mod.DcfRunRow(
        ticker=T,
        valuation_date=date.today(),
        horizon_years=s.years,
        wacc=s.wacc,
        npv=m.equity_value,
        npv_per_share=m.vps,
        shares_outstanding=s.shares * 1e6,
        currency="USD",
        live_price=s.price or None,
        live_price_at=None,
        mos_bar_used=float(mos) if isinstance(mos, (int, float)) else None,
        assumption_snapshot_json=snap,
        notes=f"workbook={DEST.name} (MELI sum-of-the-parts platform DCF)",
    )
    with connect_sqlite(str(db), role=SQLiteConnectionRole.WRITER, schema_preflight=True) as conn:
        persist_mod.upsert(conn, row)
    return True


def main() -> int:
    s = load_assumptions(T)
    m = mirror(s)
    # Loaded once and threaded through both the Scenario sheet (build) and the
    # persisted snapshot (persist_dcf_run) — PR10: one holdings read, one bear,
    # never two that could drift on a mid-run file edit.
    holdings = _load_holdings(T)
    build(s, m, DEST, holdings)
    persisted = (
        persist_dcf_run(s, m, holdings) if os.environ.get("DCF_PERSIST", "1") == "1" else False
    )
    up = (m.vps / s.price - 1) if s.price else 0.0
    print(
        f"RESULT\t{T}\tvalue/sh=${m.vps:,.2f}\tprice=${s.price:,.2f}\tupside={up:+.0%}"
        f"\topEV=${m.operating_ev:,.0f}M\tcreditEq=${m.credit_equity_value:,.0f}M"
        f"\tWACC={s.wacc:.1%}\tcreditKe={s.credit_ke:.1%}\tdcf_runs={'ok' if persisted else 'skip'}\t-> {DEST}"
    )
    print(
        f"{'Yr':>2} {'CommRev':>8} {'PayRev':>7} {'OpEBIT':>7} {'OpFCFF':>7} {'Book':>7} {'CrNI':>6} {'CrFCFE':>7}"
    )
    for r in m.rows:
        print(
            f"{r.t:>2} {r.comm_rev:>8,.0f} {r.fpay_rev:>7,.0f} {r.op_ebit:>7,.0f} {r.op_fcff:>7,.0f} "
            f"{r.cb:>7,.0f} {r.credit_ni:>6,.0f} {r.credit_fcfe:>7,.0f}"
        )
    print(
        f"\nSOTP: Operating EV ${m.operating_ev / 1000:,.1f}B + Credit equity ${m.credit_equity_value / 1000:,.1f}B "
        f"+ net cash ${s.net_cash / 1000:,.1f}B = equity ${m.equity_value / 1000:,.1f}B"
    )
    print(
        f"Value/share ${m.vps:,.2f} vs ${s.price:,.2f} ({up:+.0%}) | terminal op rev "
        f"${m.op_terminal_revenue / 1000:,.0f}B, credit ROE {m.credit_terminal_roe:.0%}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
