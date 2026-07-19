"""Dump the structure of the analyst reference DCFs so we can match their anatomy:
sheet list, row-by-row labels, formula-vs-hardcode, and the input-cell color map.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

EX = Path(r"C:\Users\Bhanu\.gemini\antigravity\scratch\earnings-summary\examples\dcf")
FILES = ["AMZN-Feb-06-2026.xlsx", "META-Feb-02-2026.xlsx", "GOOG-Mar-09-2023.xlsx"]


def argb(color):
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and rgb not in ("00000000",):
        return rgb
    return None


def dump(path, deep):
    wb = openpyxl.load_workbook(str(path), data_only=False)
    print("\n" + "=" * 90)
    print("FILE:", path.name, "| sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n  --- SHEET '{ws.title}'  dims={ws.dimensions} maxrow={ws.max_row} maxcol={ws.max_column}")
        if not deep:
            continue
        fills, fonts = {}, {}
        for row in ws.iter_rows():
            for cell in row:
                f = argb(cell.fill.fgColor) if cell.fill and cell.fill.patternType else None
                ft = argb(cell.font.color) if cell.font else None
                if f:
                    fills[f] = fills.get(f, 0) + 1
                if ft:
                    fonts[ft] = fonts.get(ft, 0) + 1
        print("    fill colors:", fills)
        print("    font colors:", fonts)
        # row-by-row: colA label, then for cols B..H show value/formula + input-color flag
        for r in range(1, min(ws.max_row, 70) + 1):
            a = ws.cell(r, 1).value
            cells = []
            for c in range(2, min(ws.max_column, 9) + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                isf = isinstance(v, str) and v.startswith("=")
                fill = argb(cell.fill.fgColor) if cell.fill and cell.fill.patternType else None
                ftc = argb(cell.font.color) if cell.font else None
                tag = "F" if isf else "v"
                mark = ""
                if fill:
                    mark += f"[bg{fill[-6:]}]"
                if ftc:
                    mark += f"[fc{ftc[-6:]}]"
                sval = str(v)
                if len(sval) > 40:
                    sval = sval[:40] + "..."
                cells.append(f"{get_column_letter(c)}={tag}:{sval}{mark}")
            if a is not None or cells:
                la = str(a) if a is not None else ""
                if len(la) > 34:
                    la = la[:34] + "..."
                print(f"    r{r:>2} | {la:<36} | " + "  ".join(cells))


for i, fn in enumerate(FILES):
    dump(EX / fn, deep=(i == 0))  # deep dump AMZN; just sheet lists for META/GOOG
